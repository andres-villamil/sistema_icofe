from django.shortcuts import render
from .forms import CreateEntitieForm, EditedEntitieForm
from django.contrib.auth.decorators import login_required
from .models import Entidades_oe
from ooee.models import OperacionEstadistica
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
##  report by excel ##
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Font, Fill, Border, Side,  GradientFill, Alignment, Color, PatternFill
from django.views.generic import TemplateView
from itertools import chain
from django.core import serializers

# Create your views here.

@login_required
def createEntitieView(request):
    count_entities = Entidades_oe.objects.count()
    create_form = CreateEntitieForm(request.POST or None)
    #print("validoq llega ",create_form)
    # if this is a POST request we need to process the form data
    if request.method == "POST": 
        create_form = CreateEntitieForm(request.POST)
        if create_form.is_valid():
            # process the data in form.cleaned_data as required
            instancia = create_form.save(commit=False)
            generate_cod = count_entities + 72 #Generar codigo para la entidad
            instancia.codigo = "EN" + str(generate_cod)
            instancia.save()
            # redirect to a new URL:
            #print("save")
            return redirect('entities:all_entities')

    # if a GET (or any other method) we'll create a blank form
    else:
        create_form = CreateEntitieForm()
        #print("no guardo")
    return render(request, 'entities/created_entities.html', {'create_form': create_form, 'count_entities': count_entities })


@login_required
def entities_edit(request, pk):
    count_entities = Entidades_oe.objects.count() 
    entitie =  Entidades_oe.objects.get(pk=pk)
    edit_form = EditedEntitieForm(instance=entitie)
    #print("editar entidad", edit_form)
    args = {}
    if request.method == "POST":
        edit_form = EditedEntitieForm(request.POST, instance=entitie)
        if edit_form.is_valid():
           editEnt_form = edit_form.save(commit=False)
           editEnt_form.save()
           return redirect('entities:publish_entities')
    return render(request, 'entities/edit_entitie.html', {'edit_form': edit_form, 'count_entities': count_entities, 'entitie': entitie})


@login_required
def allEntities(request):
   
    entities = Entidades_oe.objects.all().order_by('nombre')
    count_entities = entities.count()
    entitiesList = list(Entidades_oe.objects.all())
    total_oe = []
    for index, item in enumerate(entitiesList):
        identifier = index + 1
        ooee = OperacionEstadistica.objects.filter(entidad=identifier)
        count_ooee = ooee.count()
        total_oe.append({"entidad_oe": identifier , "numberOe":count_ooee})
       
    return render(request, 'entities/allEntities.html',  {'entities': entities, 'count_entities': count_entities, 'total_oe': total_oe })



def publishedEntities(request):
    entities = Entidades_oe.objects.filter(estado=1).order_by('nombre') #entidades publicadas 
    count_entities = entities.count()
    entitiesList = list(Entidades_oe.objects.all())
    total_oe = []
    for index, item in enumerate(entitiesList):
        identifier = index + 1
        ooee = OperacionEstadistica.objects.filter(entidad=identifier)
        count_ooee = ooee.count()
        total_oe.append({"entidad_oe": identifier , "numberOe":count_ooee}) 
    return render(request, 'entities/published_entities.html',  {'entities': entities, 'count_entities': count_entities, 'total_oe': total_oe})

@login_required
def disableEntities(request):
    entities = Entidades_oe.objects.filter(estado=2).order_by('nombre')  
    count_entities = entities.count() 
    return render(request, 'entities/disable_entities.html',  {'entities': entities, 'count_entities': count_entities})




def entities_detail(request, pk):
    count_entities = Entidades_oe.objects.count()  
    entitie = get_object_or_404(Entidades_oe, pk=pk)
    #print("entidades",entitie.pk)
    return render(request, 'entities/entities_detail.html', {'entitie': entitie, 'count_entities': count_entities})

@login_required
def entitie_delete(request, id):
    count_entities = Entidades_oe.objects.count() 
    context = {}

    obj = get_object_or_404(Entidades_oe, id = id)

    if request.method == "POST":
       obj.delete()
       return redirect('entities:all_entities')
    
    return render(request, "entities/delete_entitie.html",  {'obj': obj,  'count_entities': count_entities}, context ) 


import json
def entidadesPublicadas(request):

    entidades = Entidades_oe.objects.filter(estado=1)

    response = serializers.serialize("json", entidades, fields=('id', 'nombre', 'codigo'))
    return HttpResponse(response, content_type='application/json')


# ----------------------------report entidades by openpyxl---------------------


class report_entities_xls(TemplateView):
    def get(self, request, *args, **kwargs):
        entities = Entidades_oe.objects.all()
        wb = Workbook()
        ws = wb.active

		## size rows

        ws.row_dimensions[1].height = 55
        ws.row_dimensions[2].height = 30

		## size column
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 100
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 55
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 55
        ws.column_dimensions['H'].width = 55
        ws.column_dimensions['I'].width = 45
        ws.column_dimensions['J'].width = 45
        ws.column_dimensions['K'].width = 25
        ws.column_dimensions['L'].width = 55
        ws.column_dimensions['M'].width = 90
        ws.column_dimensions['N'].width = 55
        ws.column_dimensions['O'].width = 25
        ws.column_dimensions['P'].width = 25
        ws.column_dimensions['Q'].width = 55
        ws.column_dimensions['R'].width = 100
        ws.column_dimensions['S'].width = 55
        ws.column_dimensions['T'].width = 25
        ws.column_dimensions['U'].width = 20
        ws.column_dimensions['V'].width = 20
       
        


		##insert image
        #img = openpyxl.drawing.image.Image('media/pictures/logoSEN.png')
        #img.anchor = 'A1'	
        #ws.add_image(img)


		##styles
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
		
        #title_cell = ws['B1']
        title_cell = ws['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de entidades'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')
        title_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
	
        ws.merge_cells('A1:V1')
        ws.merge_cells('A2:G2')
        ws.merge_cells('H2:K2')
        ws.merge_cells('L2:P2')
        ws.merge_cells('Q2:U2')
        

        ## insert heads groups
        codigo_cell = ws['A2']
        codigo_cell.value = 'Entidad Responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['H2']
        codigo_cell.value = 'Director de la Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['L2']
        codigo_cell.value = 'Oficina de información estadística, oficina de planeación o similar'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['Q2']
        codigo_cell.value = 'Delegado para el SEN'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['U2']
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['V2']
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

       
        #end headsgroups

        codigo_cell = ws['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        nombre_cell = ws['B3']
        nombre_cell.value = 'Nombre'
        nombre_cell.font = Font(bold=True)
        nombre_cell.alignment = Alignment(horizontal='center')
        nombre_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        nit_cell = ws['C3'] 
        nit_cell.value = 'Nit'
        nit_cell.font = Font(bold=True)
        nit_cell.alignment = Alignment(horizontal='center')
        nit_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
		
        tipo_entidad_cell = ws['D3'] 
        tipo_entidad_cell.value = 'Tipo de Entidad'
        tipo_entidad_cell.font = Font(bold=True)
        tipo_entidad_cell.alignment = Alignment(horizontal='center')
        tipo_entidad_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

		
        direccion_cell = ws['E3']
        direccion_cell.value = 'Dirección'
        direccion_cell.font = Font(bold=True)
        direccion_cell.alignment = Alignment(horizontal='center')
        direccion_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
		
        telefono_cell = ws['F3'] 
        telefono_cell.value = 'Teléfono'
        telefono_cell.font = Font(bold=True)
        telefono_cell.alignment = Alignment(horizontal='center')
        telefono_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


        url_cell = ws['G3'] 
        url_cell.value = 'Pagina Web'
        url_cell.font = Font(bold=True)
        url_cell.alignment = Alignment(horizontal='center')
        url_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        director_cell = ws['H3'] 
        director_cell.value = 'Nombre'
        director_cell.font = Font(bold=True)
        director_cell.alignment = Alignment(horizontal='center')
        director_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        cargo_director_cell = ws['I3'] 
        cargo_director_cell.value = 'Cargo'
        cargo_director_cell.font = Font(bold=True)
        cargo_director_cell.alignment = Alignment(horizontal='center')
        cargo_director_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        correo_director_cell = ws['J3'] 
        correo_director_cell.value = 'Correo'
        correo_director_cell.font = Font(bold=True)
        correo_director_cell.alignment = Alignment(horizontal='center')
        correo_director_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        tel_director_cell = ws['K3'] 
        tel_director_cell.value = 'Teléfono'
        tel_director_cell.font = Font(bold=True)
        tel_director_cell.alignment = Alignment(horizontal='center')
        tel_director_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        nombre_planeacion_cell = ws['L3'] 
        nombre_planeacion_cell.value = 'Nombre'
        nombre_planeacion_cell.font = Font(bold=True)
        nombre_planeacion_cell.alignment = Alignment(horizontal='center')
        nombre_planeacion_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        cargo_planeacion_cell = ws['M3'] 
        cargo_planeacion_cell.value = 'Cargo'
        cargo_planeacion_cell.font = Font(bold=True)
        cargo_planeacion_cell.alignment = Alignment(horizontal='center')
        cargo_planeacion_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        correo_planeacion_cell = ws['N3'] 
        correo_planeacion_cell.value = 'Correo electrónico'
        correo_planeacion_cell.font = Font(bold=True)
        correo_planeacion_cell.alignment = Alignment(horizontal='center')
        correo_planeacion_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        tel_planeacion_cell = ws['O3'] 
        tel_planeacion_cell.value = 'Teléfono'
        tel_planeacion_cell.font = Font(bold=True)
        tel_planeacion_cell.alignment = Alignment(horizontal='center')
        tel_planeacion_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ext_planeacion_cell = ws['P3'] 
        ext_planeacion_cell.value = 'Extensión'
        ext_planeacion_cell.font = Font(bold=True)
        ext_planeacion_cell.alignment = Alignment(horizontal='center')
        ext_planeacion_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        nombre_contacto_cell = ws['Q3'] 
        nombre_contacto_cell.value = 'Nombre'
        nombre_contacto_cell.font = Font(bold=True)
        nombre_contacto_cell.alignment = Alignment(horizontal='center')
        nombre_contacto_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        cargo_contacto_cell = ws['R3'] 
        cargo_contacto_cell.value = 'Cargo'
        cargo_contacto_cell.font = Font(bold=True)
        cargo_contacto_cell.alignment = Alignment(horizontal='center')
        cargo_contacto_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        correo_contacto_cell = ws['S3'] 
        correo_contacto_cell.value = 'Correo electrónico'
        correo_contacto_cell.font = Font(bold=True)
        correo_contacto_cell.alignment = Alignment(horizontal='center')
        correo_contacto_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        tel_contacto_cell = ws['T3'] 
        tel_contacto_cell.value = 'Teléfono'
        tel_contacto_cell.font = Font(bold=True)
        tel_contacto_cell.alignment = Alignment(horizontal='center')
        tel_contacto_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ext_contacto_cell = ws['U3'] 
        ext_contacto_cell.value = 'Extensión'
        ext_contacto_cell.font = Font(bold=True)
        ext_contacto_cell.alignment = Alignment(horizontal='center')
        ext_contacto_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
		
        orden_territorial_cell = ws['V3'] 
        orden_territorial_cell.value = 'Orden Territorial'
        orden_territorial_cell.font = Font(bold=True)
        orden_territorial_cell.alignment = Alignment(horizontal='center')
        orden_territorial_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        cont = 4
        for entitie in entities:
            ws.cell(row = cont, column = 1).value = entitie.codigo
            ws.cell(row = cont, column = 2).value = entitie.nombre
            ws.cell(row = cont, column = 3).value = entitie.nit
            ws.cell(row = cont, column = 4).value = str(entitie.tipo_entidad)
            ws.cell(row = cont, column = 5).value = entitie.direccion
            ws.cell(row = cont, column = 6).value = entitie.telefono
            ws.cell(row = cont, column = 7).value = entitie.pagina_web
            ws.cell(row = cont, column = 8).value = entitie.nombre_dir
            ws.cell(row = cont, column = 9).value = entitie.cargo_dir
            ws.cell(row = cont, column = 10).value = entitie.correo_dir
            ws.cell(row = cont, column = 11).value = entitie.telefono_dir
            ws.cell(row = cont, column = 12).value = entitie.nombre_pla
            ws.cell(row = cont, column = 13).value = entitie.cargo_pla
            ws.cell(row = cont, column = 14).value = entitie.correo_pla
            ws.cell(row = cont, column = 15).value = entitie.telefono_pla
            ws.cell(row = cont, column = 15).value = entitie.extension_pla
            ws.cell(row = cont, column = 17).value = entitie.nombre_cont
            ws.cell(row = cont, column = 18).value = entitie.cargo_cont
            ws.cell(row = cont, column = 19).value = entitie.correo_cont
            ws.cell(row = cont, column = 20).value = entitie.telefono_cont
            ws.cell(row = cont, column = 21).value = entitie.extension_cont
            if entitie.ord_ter == None:
                ws.cell(row = cont, column = 22).value =  ""
            else: 
                str(entitie.ord_ter)

            ws.cell(row = cont, column = 1).border = thin_border
            ws.cell(row = cont, column = 2).border = thin_border
            ws.cell(row = cont, column = 3).border = thin_border
            ws.cell(row = cont, column = 4).border = thin_border
            ws.cell(row = cont, column = 5).border = thin_border
            ws.cell(row = cont, column = 6).border = thin_border
            ws.cell(row = cont, column = 7).border = thin_border
            ws.cell(row = cont, column = 8).border = thin_border
            ws.cell(row = cont, column = 9).border = thin_border
            ws.cell(row = cont, column = 10).border = thin_border
            ws.cell(row = cont, column = 11).border = thin_border
            ws.cell(row = cont, column = 12).border = thin_border
            ws.cell(row = cont, column = 13).border = thin_border
            ws.cell(row = cont, column = 14).border = thin_border
            ws.cell(row = cont, column = 15).border = thin_border
            ws.cell(row = cont, column = 16).border = thin_border
            ws.cell(row = cont, column = 17).border = thin_border
            ws.cell(row = cont, column = 18).border = thin_border
            ws.cell(row = cont, column = 19).border = thin_border
            ws.cell(row = cont, column = 20).border = thin_border
            ws.cell(row = cont, column = 21).border = thin_border
            ws.cell(row = cont, column = 22).border = thin_border
           
                           
            cont+=1
		
        file_name = "reporte_entidades.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(file_name)
        response['Content-Disposition'] = content
        wb.save(response)
        return response 
