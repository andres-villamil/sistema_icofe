from django.shortcuts import render
from login.models import Role, User, Profile
from rraa.models import RegistroAdministrativo
from demandas.models import demandaInfor

from .forms import CreateOEForm,  CreateEntidadFaseForm, MB_NormaForm, MB_RequerimientoForm, MB_PrinUsuariosForm, MC_UnidadObservacionForm, \
MC_ObtencionDatoForm, MC_MuestreoProbabilisticoForm, MC_MuestreoNoProbabilisticoForm, MC_TipoMarcoForm, MC_DocsDesarrolloForm, MC_ConceptosEstandarizadosForm, \
MC_ClasificacionesForm, MC_CoberturaGeograficaForm, MC_DesagregacionInformacionForm,  MC_FuenteFinanciacionForm, MC_listaVariableFormset, MD_MedioDatosForm, \
MD_PeriodicidadOeForm, MD_HerramProcesamientoForm,  ME_AnalisisResultadosForm, MF_MediosDifusionForm, MF_FechaPublicacionForm, MF_FrecuenciaDifusionForm, \
MF_ProductosDifundirForm, MF_OtrosProductosForm, MF_ResultadosSimilaresForm, MF_HPSistemaInfoForm,  MC_resultadoEstadisticoFormset, \
EditOEForm, EditMB_NormaForm, EditMB_RequerimientoForm, EditMB_PrinUsuariosForm, EditMC_UnidadObservacionForm, \
EditMC_ObtencionDatoForm, EditMC_MuestreoProbabilisticoForm, EditMC_MuestreoNoProbabilisticoForm, EditMC_TipoMarcoForm, EditMC_DocsDesarrolloForm, EditMC_ConceptosEstandarizadosForm, \
EditMC_ClasificacionesForm, EditMC_CoberturaGeograficaForm, EditMC_DesagregacionInformacionForm,  EditMC_FuenteFinanciacionForm,  EditMD_MedioDatosForm, \
EditMD_PeriodicidadOeForm, EditMD_HerramProcesamientoForm,  EditME_AnalisisResultadosForm, EditMF_MediosDifusionForm, EditMF_FechaPublicacionForm, EditMF_FrecuenciaDifusionForm, \
EditMF_ProductosDifundirForm, EditMF_OtrosProductosForm, EditMF_ResultadosSimilaresForm, EditMF_HPSistemaInfoForm, \
    MB_EntidadFasesFormset, CommentForm, NovedadForm, CriticaForm, EvaluacionCalidadForm, EditEvaluacionCalidadForm


from django.contrib.auth.decorators import login_required


from .models import Entidades_oe, OoeeState, AreaTematica, Tema, FasesProceso, MB_EntidadFases, Norma, MB_Norma, Requerimientos, MB_Requerimientos, \
PrinUsuarios, MB_PrinUsuarios, OperacionEstadistica, OoeeLog, Comment, UnidadObservacion, MC_UnidadObservacion, TipoOperacion, ObtencionDato, \
MC_ObtencionDato, MuestreoProbabilistico, MC_MuestreoProbabilistico, MuestreoNoProbabilistico, MC_MuestreoNoProbabilistico, TipoMarco, MC_TipoMarco, \
DocsDesarrollo, MC_DocsDesarrollo, ConceptosEstandarizados, MC_ConceptosEstandarizados, Clasificaciones, MC_Clasificaciones, CoberturaGeografica , \
MC_CoberturaGeografica, DesagregacionInformacion, MC_DesagregacionInformacion, FuenteFinanciacion,  MC_FuenteFinanciacion, MC_listaVariable, \
MedioDatos, MD_MedioDatos, PeriodicidadOe, MD_PeriodicidadOe, HerramProcesamiento ,MD_HerramProcesamiento,  ME_AnalisisResultados,  AnalisisResultados, \
MediosDifusion, MF_MediosDifusion, FechaPublicacion, MF_FechaPublicacion, FrecuenciaDifusion, MF_FrecuenciaDifusion, ProductosDifundir, MF_ProductosDifundir, \
OtrosProductos, MF_OtrosProductos, MF_ResultadosSimilares, MF_HPSistemaInfo, MC_ResultadoEstadistico, TipoNovedad, NovedadActualizacion, EstadoActualizacion, \
    EstadoCritica, Critica, EstadoEvaluacion, ResultadoEvaluacion, PlanDeMejoramiento, EvaluacionCalidad



from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render, redirect, get_object_or_404
##  report by excel ##
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Font, Fill, Border, Side, GradientFill, Alignment, Color, PatternFill
##  end report by excel ##
from django.contrib import messages 
from django.views.generic import TemplateView
from django.forms import inlineformset_factory
from django.core import serializers
from .filters import TopicsFilter, NameFilterOOEE
from itertools import chain
from django.core.mail import send_mail
from django.utils import timezone

from django.db.models import Q

from django.http import JsonResponse
from itertools import groupby
from django.db.models import Count


####libreria reporte excel
import xlsxwriter
import io
from datetime import datetime
from django.views.generic import View

from django.utils.decorators import method_decorator


# Create your views here.


@login_required
def createOOEEView(request):

    user = request.user
    #print("___user______", user.profile.entidad )
    entidad_creadora = user.profile.entidad.pk


    entities = Entidades_oe.objects.filter(estado_id=1) #entidades publicadas
    count_entities = entities.count()
    count_oe = OperacionEstadistica.objects.count()
    #print("contador de oe", count_oe)
    createoe_form = CreateOEForm(request.POST or None, request.FILES or None)   #form crear oe
    EntidadFase_form = CreateEntidadFaseForm(request.POST or None) # form crear fase
    norma_form = MB_NormaForm(request.POST or None) # form comentarios norma
    requerimiento_form = MB_RequerimientoForm(request.POST or None) # form comentarios requerimiento
    prinUsuarios_form = MB_PrinUsuariosForm(request.POST or None) # form comentarios usuarios principales
    unidadObs_form = MC_UnidadObservacionForm(request.POST or None) # form comentarios unidad de observacion
    obtDato_form = MC_ObtencionDatoForm(request.POST or None) # form comentarios unidad de obtención de datos
    muesProb_form = MC_MuestreoProbabilisticoForm(request.POST or None) # form comentarios muestreo probabibilistico
    muesNoProb_form = MC_MuestreoNoProbabilisticoForm(request.POST or None) # form comentarios muestreo No probabibilistico
    tipoMarco_Form = MC_TipoMarcoForm(request.POST or None) # form comentarios tipo de marco
    docDes_form = MC_DocsDesarrolloForm(request.POST or None) # form comentarios documentos que se elaboran para el desarrollo de la oe
    concepEstan_form = MC_ConceptosEstandarizadosForm(request.POST or None) # form comentarios oe utiliza conceptos estandaraizados
    nomClas_Form = MC_ClasificacionesForm(request.POST or None) # form comentarios oe utiliza nomenclaturas y/o clasificaciones
    cobGeo_form = MC_CoberturaGeograficaForm(request.POST or None) # form comentarios oe cobertura geografica
    desInfo_form = MC_DesagregacionInformacionForm(request.POST or None) # form desagregacion oe
    fuenteFin_form = MC_FuenteFinanciacionForm(request.POST or None) # form fuentes de financiación
    
    #Modulo D *****  
    medDat_form = MD_MedioDatosForm(request.POST or None) 
    periodOe_form = MD_PeriodicidadOeForm(request.POST  or None) 
    herrProc_form = MD_HerramProcesamientoForm(request.POST  or None)

    #Modulo E ***** 
    anaResul_form = ME_AnalisisResultadosForm(request.POST or None)

    #Modulo F
    me_dif_form = MF_MediosDifusionForm(request.POST  or None)
    fecPubli_form = MF_FechaPublicacionForm(request.POST  or None)
    freDif_form = MF_FrecuenciaDifusionForm(request.POST or None)
    prod_dif_form = MF_ProductosDifundirForm(request.POST  or None)
    otroProd_form = MF_OtrosProductosForm(request.POST  or None)
    resultSimi_form = MF_ResultadosSimilaresForm(request.POST  or None)
    hpsistInfo_form = MF_HPSistemaInfoForm(request.POST  or None)

    #novedad
    novedad_form = NovedadForm()
    #novedades = post_oe.novedadactualizacions.filter(active=True)
    new_novedad = None
    

    registered = False
    
    # if this is a POST request we need to process the form data
    if request.method == 'GET':     
        formListset = MC_listaVariableFormset(queryset=MC_listaVariable.objects.none(), prefix='listas')
        formResultset = MC_resultadoEstadisticoFormset(queryset=MC_ResultadoEstadistico.objects.none(), prefix='resultados')
        formEntiFasset  = MB_EntidadFasesFormset(queryset=MB_EntidadFases.objects.none(), prefix='entidadfase')

    elif request.method == "POST": 
        
        createoe_form = CreateOEForm(request.POST, request.FILES)
        formEntiFasset  = MB_EntidadFasesFormset(request.POST, prefix='entidadfase')   # pregunta 3 DE MODULO B entidad y fases que intervienen
        norma_form = MB_NormaForm(request.POST)
        requerimiento_form = MB_RequerimientoForm(request.POST)
        prinUsuarios_form = MB_PrinUsuariosForm(request.POST)
        unidadObs_form = MC_UnidadObservacionForm(request.POST)
        obtDato_form = MC_ObtencionDatoForm(request.POST)
        muesProb_form =MC_MuestreoProbabilisticoForm(request.POST)
        muesProb_form =MC_MuestreoProbabilisticoForm(request.POST)
        muesNoProb_form = MC_MuestreoNoProbabilisticoForm(request.POST)
        tipoMarco_Form = MC_TipoMarcoForm(request.POST)
        docDes_form = MC_DocsDesarrolloForm(request.POST)
        concepEstan_form = MC_ConceptosEstandarizadosForm(request.POST)
        nomClas_Form = MC_ClasificacionesForm(request.POST)
        cobGeo_form = MC_CoberturaGeograficaForm(request.POST)
        desInfo_form = MC_DesagregacionInformacionForm(request.POST)
        fuenteFin_form = MC_FuenteFinanciacionForm(request.POST)
        formListset = MC_listaVariableFormset(request.POST, prefix='listas') #PREGUNTA 14 MODULO C***
        formResultset = MC_resultadoEstadisticoFormset(request.POST, prefix='resultados') #PREGUNTA 15 MODULO C***
          
        ##Modulo D
        medDat_form = MD_MedioDatosForm(request.POST) 
        periodOe_form = MD_PeriodicidadOeForm(request.POST) 
        herrProc_form = MD_HerramProcesamientoForm(request.POST)

        #Modulo E ***** 
        anaResul_form = ME_AnalisisResultadosForm(request.POST)

        #Modulo F
        me_dif_form = MF_MediosDifusionForm(request.POST)
        fecPubli_form = MF_FechaPublicacionForm(request.POST)
        freDif_form = MF_FrecuenciaDifusionForm(request.POST)
        prod_dif_form = MF_ProductosDifundirForm(request.POST)
        otroProd_form = MF_OtrosProductosForm(request.POST)
        resultSimi_form = MF_ResultadosSimilaresForm(request.POST)
        hpsistInfo_form = MF_HPSistemaInfoForm(request.POST)
 
        ## Novedades para cuando es creada
        novedad_form = NovedadForm(data=request.POST)

        
        if createoe_form.is_valid() and formEntiFasset.is_valid() and  norma_form.is_valid() and requerimiento_form.is_valid() and prinUsuarios_form.is_valid() and unidadObs_form.is_valid() and \
            obtDato_form.is_valid() and muesProb_form.is_valid() and muesNoProb_form.is_valid() and  tipoMarco_Form.is_valid() and \
                docDes_form.is_valid() and  concepEstan_form.is_valid() and nomClas_Form.is_valid() and desInfo_form.is_valid() and \
                    fuenteFin_form.is_valid() and formListset.is_valid() and formResultset.is_valid() and medDat_form.is_valid() and periodOe_form.is_valid() and herrProc_form.is_valid() and \
                        anaResul_form.is_valid() and  me_dif_form.is_valid() and fecPubli_form.is_valid() and freDif_form.is_valid() and \
                            prod_dif_form.is_valid() and  otroProd_form.is_valid() and resultSimi_form.is_valid() and hpsistInfo_form.is_valid() and novedad_form.is_valid():
             
            instancia = createoe_form.save(commit=False)

            generate_cod = count_oe + 134  ## generar nuevas ooee codigo a partir de 600
            instancia.codigo_oe = "OE" + str(generate_cod)
            instancia.creado_por = user

            if 'variable_file' in request.FILES:
                instancia.variable_file = request.FILES['variable_file']

            if 'anexos' in request.FILES:
                instancia.anexos = request.FILES['anexos']
            
            instancia.save()
            obj_id = instancia.id  ##obtener id del objeto que estoy creando

            for formLista in formListset:
                # so that `ooee` instance can be attached.
                formListaVariable = formLista.save(commit=False)
                formListaVariable.ooee_id = obj_id
                formListaVariable.instancia = instancia
                if formListaVariable.lista_var == "":
                        formListaVariable = formLista.save(commit=False)
                else: 
                    formListaVariable.save()

            #print("id oe",obj_id)

            for formResult in formResultset:
                # so that `ooee` instance can be attached.
                formResultEstadis = formResult.save(commit=False)
                formResultEstadis.ooee_id = obj_id
                formResultEstadis.instancia = instancia
                if formResultEstadis.resultEstad == "":
                    formResultEstadis = formResult.save(commit=False)
                else:
                    formResultEstadis.save()

            for formEntiFase in formEntiFasset:

                formEntidadFase = formEntiFase.save(commit=False)
                formEntidadFase.ooee_id = obj_id
                formEntidadFase.instancia = instancia
                if formEntidadFase.nombre_entifas == "":
                    formEntidadFase = formEntiFase.save(commit=False)
                else:
                    # Save the entitie to the database
                    formEntidadFase.save() 
                    formEntiFase.save_m2m()  
                 
        ## form norma
            addNorma = norma_form.save(commit=False)
            addNorma.ooee_id = obj_id
            addNorma.save()
        ## form requerimientos
            addRequerimiento = requerimiento_form.save(commit=False)
            addRequerimiento.ooee_id = obj_id
            addRequerimiento.save()
        ## form Principales usuarios
            addPriUser = prinUsuarios_form.save(commit=False)
            addPriUser.ooee_id = obj_id
            addPriUser.save()
        ## form unidad de observación
            addUnidad = unidadObs_form.save(commit=False)
            addUnidad.ooee_id = obj_id
            addUnidad.save()
        ## Obtención de datos
            addObtDato = obtDato_form.save(commit=False)
            addObtDato.ooee_id = obj_id
            addObtDato.save()
        ## Muestreo Probabilistico
            addMuesProb = muesProb_form.save(commit=False)
            addMuesProb.ooee_id = obj_id
            addMuesProb.save()
        ## Muestreo no probabilistico
            addMuesNoProb =  muesNoProb_form.save(commit=False)
            addMuesNoProb.ooee_id = obj_id
            addMuesNoProb.save()
        ## Tipo de marco
            addTipoMarco = tipoMarco_Form.save(commit=False)
            addTipoMarco.ooee_id = obj_id
            addTipoMarco.save()
        ## Documentos de desarrollo
            addDocDes = docDes_form.save(commit=False)
            addDocDes.ooee_id = obj_id
            addDocDes.save()
        ## conceptos estandarizados
            addConcepEst = concepEstan_form.save(commit=False)
            addConcepEst.ooee_id = obj_id
            addConcepEst.save()
        ## Clasisificaciones y nomenclatura
            addnomClas = nomClas_Form.save(commit=False)
            addnomClas.ooee_id = obj_id
            addnomClas.save()
        ## Cobertura Geografica
            addcobGeo = cobGeo_form.save(commit=False)
            addcobGeo.ooee_id = obj_id
            addcobGeo.save()
        ## Desagregación de la información
            adddesInfo = desInfo_form.save(commit=False)
            adddesInfo.ooee_id = obj_id
            adddesInfo.save()
        ## Fuente financiación
            addfuenteFin = fuenteFin_form.save(commit=False)
            addfuenteFin.ooee_id = obj_id
            addfuenteFin.save()
    

        ### Modulo D  ###

        ##Medio Obtención de datos
            addmedDat = medDat_form.save(commit=False)
            addmedDat.ooee_id = obj_id
            addmedDat.save()
        ## periodicidad de recolección
            addperiodOe = periodOe_form.save(commit=False)
            addperiodOe.ooee_id = obj_id
            addperiodOe.save()
        ## herramientas son utilizadas en el procesamiento de los datos
            addherrProc= herrProc_form.save(commit=False)
            addherrProc.ooee_id = obj_id
            addherrProc.save()
        
        ### Modulo E  ###

        ##tipo de análisis
            addanaResul = anaResul_form.save(commit=False)
            addanaResul.ooee_id = obj_id
            addanaResul.save()

        ### Modulo F  ###

        ##A través de qué medio(s) difunde los resultados estadísticos
            addme_diF =  me_dif_form.save(commit=False)
            addme_diF.ooee_id = obj_id
            addme_diF.save()
        ## fecha de publicación de resultados estadísticos
            addfecPubli = fecPubli_form.save(commit=False)
            addfecPubli.ooee_id = obj_id
            addfecPubli.save()
        ## Cuál es la frecuencia de difusión de los resultados estadísticos
            addfreDif = freDif_form.save(commit=False)
            addfreDif.ooee_id = obj_id
            addfreDif.save()
        ## Cuáles productos utiliza para difundir los resultados estadísticos
            addprod_dif = prod_dif_form.save(commit=False)
            addprod_dif.ooee_id = obj_id
            addprod_dif.save()
        ##  otros productos estadísticos de la OE están disponibles para consulta de los usuarios
            addotroProd = otroProd_form.save(commit=False)
            addotroProd.ooee_id = obj_id
            addotroProd.save()
        ## Conoce si otra entidad produce resultados similares a los de la operación estadística
            addresultSimi = resultSimi_form.save(commit=False)
            addresultSimi.ooee_id = obj_id
            addresultSimi.save()
        ## La operación estadística hace parte de algún sistema de información
            addhpsistInfo = hpsistInfo_form.save(commit=False)
            addhpsistInfo.ooee_id = obj_id
            addhpsistInfo.save()

        ##  novedades
            new_novedad = novedad_form.save(commit=False)
            new_novedad.post_oe_id = obj_id
            new_novedad.name_nov = request.user
            new_novedad.est_actualiz_id = 4
            new_novedad.descrip_novedad = "Se crea operación estadística."
            new_novedad.save()


        ## form 
            createoe_form.save_m2m() ## metodo para guardar relaciones manyTomany
            registered = True
            # redirect to a new URL:
            #print("save")
            ##return redirect('ooee:all_ooee')

        else:    
            #print("no guardo___________________________")
            messages.error(request, "Las respuestas NO se han guardado verifique si tiene errores")
    
        
    return render(request, 'ooee/created_ooee.html', {'createoe_form': createoe_form, 'formEntiFasset': formEntiFasset, 'norma_form': norma_form,
    'requerimiento_form': requerimiento_form, 'prinUsuarios_form': prinUsuarios_form, 'unidadObs_form': unidadObs_form, 'obtDato_form': obtDato_form,
    'muesProb_form':  muesProb_form, 'muesNoProb_form': muesNoProb_form, 'tipoMarco_Form': tipoMarco_Form, 'docDes_form': docDes_form, 
    'concepEstan_form': concepEstan_form, 'nomClas_Form': nomClas_Form,  'cobGeo_form':  cobGeo_form,  'desInfo_form' : desInfo_form, 
    'fuenteFin_form' : fuenteFin_form, 'formListset': formListset, 'formResultset': formResultset, 'medDat_form': medDat_form, 'periodOe_form': periodOe_form, 'herrProc_form': herrProc_form,
    'anaResul_form': anaResul_form, 'me_dif_form': me_dif_form, 'fecPubli_form': fecPubli_form, 'freDif_form': freDif_form, 'prod_dif_form': prod_dif_form,
    'otroProd_form' : otroProd_form, 'resultSimi_form': resultSimi_form, 'hpsistInfo_form': hpsistInfo_form, 'count_entities': count_entities, 'registered': registered,
    'new_novedad': new_novedad, 'novedad_form': novedad_form, 'entidad_creadora': entidad_creadora })


@login_required
def nuevas_operaciones(request):

    entities = Entidades_oe.objects.filter(estado=1)
    count_entities = entities.count()
    user = request.user
    entidadByUser = user.profile.entidad.id

    if user.is_authenticated == True and str(user) == "sdazag": # 1

        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (tema=2) | Q (tema=7) | Q (tema=19) | Q (tema=28))
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })

    elif user.is_authenticated == True and str(user) == "dvlizarazog": # 4 
        
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (tema=9) | Q (tema=29) | Q (tema=30))
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })

    elif user.is_authenticated == True and str(user) == "ancardenasc": # 5 frherreran freddy es administrador se cambia por gavargasr
        
        filtros_tema_oe = OperacionEstadistica.objects.filter(tema=10)
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })

    elif user.is_authenticated == True and str(user) == "pczambranog": # 6
        
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (tema=11) | Q (tema=12) | Q (tema=25))
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })

    elif user.is_authenticated == True and str(user) == "mjpradag": # 7  mjpradag  (mppulidor hizo cesión de contrato a mjpradag)
    
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (tema=13) | Q (tema=24) | Q (tema=26) | Q (tema=27))
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })


    elif user.is_authenticated == True and str(user) == "eeguayazans": # 8 eeguayazans Edgar Eduardo Guayazan Sierra
        
        filtros_tema_oe  = OperacionEstadistica.objects.filter(Q (tema=4) | Q (tema=8) | Q (tema=20))
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })

    elif user.is_authenticated == True and str(user) == "mlbarretob": # 9

        filtros_tema_oe  = OperacionEstadistica.objects.filter(Q (tema=21) | Q (tema=22) | Q (tema=23))
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })

    ## start coordinadores dimpe, cuentas y censos
    
    ### Usuarios Dirección de Censos y Demografía ###
    elif user.is_authenticated == True and str(user) == "emvallec": #emvallec

        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE16") | Q (creado_por__username="emvallec")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })
        
    elif user.is_authenticated == True and str(user) == "hcoteo":  #hcoteo 

        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE39") | Q(codigo_oe="OE53") | Q (creado_por__username="hcoteo")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })

    elif user.is_authenticated == True and str(user) == "mfospinab":  #mfospinab

        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE168") | Q(codigo_oe="OE169") | Q (creado_por__username="mfospinab")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })

    ### end Usuarios Dirección de Censos y Demografía ####

    ### Usuarios Dimpe
    elif user.is_authenticated == True and str(user) == "alsoto":  #alsoto

        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE268") | Q(codigo_oe="OE269") | Q(codigo_oe="OE34") | Q(codigo_oe="OE427")
        | Q(codigo_oe="OE491") | Q(codigo_oe="OE495") | Q (codigo_oe="OE496") | Q (codigo_oe="OE497") | Q (codigo_oe="OE498") | Q (codigo_oe="OE499")
        | Q (codigo_oe="OE500") | Q (codigo_oe="OE501") | Q (codigo_oe="OE502") | Q (codigo_oe="OE503") | Q (codigo_oe="OE504") | Q (codigo_oe="OE505") | Q (codigo_oe="OE506")
        | Q (codigo_oe="OE507") | Q (codigo_oe="OE508") | Q (codigo_oe="OE509") | Q (codigo_oe="OE51") | Q (codigo_oe="OE59") | Q (codigo_oe="OE71") | Q (codigo_oe="OE600")
        | Q (creado_por__username="alsoto")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })

    
    elif user.is_authenticated == True and str(user) == "cequinonesl":  #cequinonesl

        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE13") | Q(codigo_oe="OE221") | Q(codigo_oe="OE222") | Q(codigo_oe="OE360")
        | Q(codigo_oe="OE37") | Q(codigo_oe="OE451") | Q (codigo_oe="OE494") | Q (codigo_oe="OE9") | Q (creado_por__username="cequinonesl")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })

    
    elif user.is_authenticated == True and str(user) == "cequinonesl_2":  #cequinonesl_2 *** se crea porque el usuario no presta credenciales dane

        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE13") | Q(codigo_oe="OE221") | Q(codigo_oe="OE222") | Q(codigo_oe="OE360")
        | Q(codigo_oe="OE37") | Q(codigo_oe="OE451") | Q (codigo_oe="OE9") | Q (creado_por__username="cequinonesl_2")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })
        

    elif user.is_authenticated == True and str(user) == "aramosh": #dcpenab no continua trabajando en dane y se reemplaza por aramosh
        
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE129") | Q(codigo_oe="OE20") | Q(codigo_oe="OE297") | Q(codigo_oe="OE35")
        | Q(codigo_oe="OE36") | Q(codigo_oe="OE490") | Q(codigo_oe="OE449") | Q (codigo_oe="OE50") | Q (codigo_oe="OE556") | Q (codigo_oe="OE620") | Q (creado_por__username="aramosh")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })

        
    elif user.is_authenticated == True and str(user) == "eagarzona": #eagarzona
        
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE10") | Q(codigo_oe="OE104") | Q(codigo_oe="OE12") | Q(codigo_oe="OE239")
        | Q(codigo_oe="OE240") | Q(codigo_oe="OE533") | Q (creado_por__username="eagarzona")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })

    elif user.is_authenticated == True and str(user) == "gaportillac":  #gaportillac
        
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE208") | Q(codigo_oe="OE450") | Q (creado_por__username="gaportillac")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })


    elif user.is_authenticated == True and str(user) == "ircastillop": #ircastillop
        
        filtros_tema_oe =  OperacionEstadistica.objects.filter(Q (codigo_oe="OE101") | Q(codigo_oe="OE103") | Q(codigo_oe="OE162") | Q(codigo_oe="OE283")
        | Q(codigo_oe="OE493") | Q (creado_por__username="ircastillop")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })


    elif user.is_authenticated == True and str(user) == "jjquinchuac":  #jjquinchuac
        
        filtros_tema_oe =  OperacionEstadistica.objects.filter(Q (codigo_oe="OE284") | Q (creado_por__username="jjquinchuac")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })
        

    elif user.is_authenticated == True and str(user) == "laguion": #laguion
         
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE14") | Q(codigo_oe="OE15") | Q(codigo_oe="OE21") | Q(codigo_oe="OE223")
        | Q(codigo_oe="OE358") | Q (codigo_oe="OE359") | Q (codigo_oe="OE361") | Q (codigo_oe="OE487") | Q(codigo_oe="OE557") 
        | Q (creado_por__username="laguion") ).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })


    elif user.is_authenticated == True and str(user) == "lahernandezv": #lahernandezv
         
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE163") | Q(codigo_oe="OE164") | Q(codigo_oe="OE165") | Q(codigo_oe="OE166")
        | Q(codigo_oe="OE305") | Q (codigo_oe="OE362") | Q (codigo_oe="OE363") | Q (codigo_oe="OE364") | Q(codigo_oe="OE38") | Q(codigo_oe="OE52") 
        | Q (creado_por__username="lahernandezv")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })

    elif user.is_authenticated == True and str(user) == "mhsanchezf": #mhsanchezf
         
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE11") | Q(codigo_oe="OE130") | Q(codigo_oe="OE161") | Q(codigo_oe="OE167")
        | Q(codigo_oe="OE33") | Q (codigo_oe="OE365") | Q (codigo_oe="OE452") | Q (codigo_oe="OE492") | Q (creado_por__username="mhsanchezf")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })

        
    elif user.is_authenticated == True and str(user) == "opandradem": #opandradem
         
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE206") | Q(codigo_oe="OE207") | Q(codigo_oe="OE473") | Q (creado_por__username="opandradem")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })


    elif user.is_authenticated == True and str(user) == "ambernalc": #ambernalc
         
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE34") | Q(codigo_oe="OE51") | Q(codigo_oe="OE59")
        | Q(codigo_oe="OE71") | Q(codigo_oe="OE268") | Q(codigo_oe="OE269") | Q(codigo_oe="OE427") | Q (codigo_oe="OE600")  | Q (creado_por__username="ambernalc")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })

    ### End Usuarios Dimpe

    ### Usuarios Cuentas nacionales

    elif user.is_authenticated == True and str(user) == "apcasasv": # apcasasv
         
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE226") | Q(codigo_oe="OE72") | Q (codigo_oe="OE566") | Q (creado_por__username="apcasasv")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })

        
    elif user.is_authenticated == True and str(user) == "dacobaledam":  #dacobaledam

        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE558") | Q(codigo_oe="OE559") | Q(codigo_oe="OE560") | Q(codigo_oe="OE561")
        | Q(codigo_oe="OE562") | Q (codigo_oe="OE563") | Q (codigo_oe="OE564") | Q (creado_por__username="dacobaledam")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })
              
    elif user.is_authenticated == True and str(user) == "dcubidesl": #dcubidesl
         
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE131") | Q(codigo_oe="OE209") | Q(codigo_oe="OE22") | Q(codigo_oe="OE520")
        | Q(codigo_oe="OE535") | Q (codigo_oe="OE570") | Q (codigo_oe="OE571") | Q (codigo_oe="OE572") | Q (codigo_oe="OE573") | Q (codigo_oe="OE574")
        | Q (codigo_oe="OE575") | Q (creado_por__username="dcubidesl")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })

    elif user.is_authenticated == True and str(user) == "eforerom": #eforerom
         
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE370") | Q(codigo_oe="OE567") | Q(codigo_oe="OE367") | Q (creado_por__username="eforerom")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })

    elif user.is_authenticated == True and str(user) == "kjdavilaa": #kjdavilaa
         
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE370") | Q (creado_por__username="kjdavilaa")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })
        
    elif user.is_authenticated == True and str(user) == "jaherrerab":  #jaherrerab
         
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE271") | Q(codigo_oe="OE569")| Q (creado_por__username="jaherrerab")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })


    elif user.is_authenticated == True and str(user) == "jepalaciosm": #jepalaciosm
         
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE272") | Q(codigo_oe="OE273") | Q(codigo_oe="OE274")
        | Q (creado_por__username="jepalaciosm")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })

    
    elif user.is_authenticated == True and str(user) == "jpcardozot":  #jpcardozot
        
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE170") | Q(codigo_oe="OE270") | Q(codigo_oe="OE369") | Q(codigo_oe="OE568") | Q(codigo_oe="OE370")
        | Q (creado_por__username="jpcardozot")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })


    elif user.is_authenticated == True and str(user) == "nchudtr":  #nchudtr
        
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (codigo_oe="OE272") | Q(codigo_oe="OE273") | Q(codigo_oe="OE274")
        | Q (creado_por__username="nchudtr")).order_by('nombre_oe')
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })

    ## end cuentas nacionales

    ## end coordinadores dimpe, cuentas y censos


    elif user.is_authenticated == True and user.profile.role.id == 1 or user.profile.role.id == 5 or user.profile.role.id == 6 or user.profile.role.id == 7: 

        filtros_tema_oe  = OperacionEstadistica.objects.all()
        novedad = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        for nov in unique_oe:
                
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if  str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })
    
    else:
        
        filtros_tema_oe  = OperacionEstadistica.objects.filter(entidad=entidadByUser)
        novedad = []
        lista_oe = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
        
        for nov in unique_oe:     
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            if str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"oe_identify": int(oe_nov), "last_novedad": str(novedad_last.est_actualiz), "nombre_oe": str(novedad_last.post_oe) })


    return render(request, 'ooee/nuevas_ooee.html', { 'novedad': novedad, 'filtros_tema_oe': filtros_tema_oe, 'count_entities': count_entities })



############ vista para mostrar las operaciones según el rol y la entidad
@login_required
def allOOEE(request):
    
    count_entities = Entidades_oe.objects.filter(estado_id=1).count()
    user = request.user
    
    if user.is_authenticated == True and str(user) == "sdazag": # 1 sdazag
        oes = OperacionEstadistica.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28)).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)
    
        """ elif user.is_authenticated == True and str(user) == "aobandor": # 3 se cambia usuaaria a rol administrador

        oes = OperacionEstadistica.objects.filter(Q (tema=5) | Q (tema=14) | Q (tema=15) | Q (tema=16) | Q (tema=17) | Q (tema=18)).order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes) """

    elif user.is_authenticated == True and str(user) == "dvlizarazog": # 4
        
        oes = OperacionEstadistica.objects.filter(Q (tema_id=9) | Q (tema_id=29) | Q (tema_id=30)).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "ancardenasc": # 5frherreran  se cambia por gavargasr
        
        oes = OperacionEstadistica.objects.filter(tema_id=10).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "lhsanchezz": # 6
        oes = OperacionEstadistica.objects.filter(Q (tema_id=3) | Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "pczambranog": # 6
        oes = OperacionEstadistica.objects.filter(Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)
        
    elif user.is_authenticated == True and str(user) == "mjpradag": # 7  mppulidor se cambia por mjpradag

        oes = OperacionEstadistica.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27)).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)
       
    elif user.is_authenticated == True and str(user) == "eeguayazans": # 8 eeguayazans Edgar Eduardo Guayazan Sierra

        oes = OperacionEstadistica.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20)).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)
        
    elif user.is_authenticated == True and str(user) == "mlbarretob": # 9

        oes = OperacionEstadistica.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23)).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    ### Usuarios Dirección de Censos y Demografía ###
    elif user.is_authenticated == True and str(user) == "emvallec": #emvallec

        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE16") | Q (creado_por__username="emvallec")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "hcoteo":  #hcoteo 
        
        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE39") | Q(codigo_oe="OE53") | Q (creado_por__username="hcoteo")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "mfospinab":  #mfospinab

        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE168") | Q(codigo_oe="OE169") | Q (creado_por__username="mfospinab")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)
    ### end Usuarios Dirección de Censos y Demografía ####

    ### Usuarios Dimpe

    elif user.is_authenticated == True and str(user) == "ambernalc": # 2 ambernalc

        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE34") | Q(codigo_oe="OE51") | Q(codigo_oe="OE59")
        | Q(codigo_oe="OE71") | Q(codigo_oe="OE268") | Q(codigo_oe="OE269") | Q(codigo_oe="OE427") | Q (codigo_oe="OE600") | Q (creado_por__username="ambernalc")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)


    elif user.is_authenticated == True and str(user) == "alsoto":  #alsoto

        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE268") | Q(codigo_oe="OE269") | Q(codigo_oe="OE34") | Q(codigo_oe="OE427")
        | Q(codigo_oe="OE491") | Q(codigo_oe="OE495") | Q (codigo_oe="OE496") | Q (codigo_oe="OE497") | Q (codigo_oe="OE498") | Q (codigo_oe="OE499")
        | Q (codigo_oe="OE500") | Q (codigo_oe="OE501") | Q (codigo_oe="OE502") | Q (codigo_oe="OE503") | Q (codigo_oe="OE504") | Q (codigo_oe="OE505") | Q (codigo_oe="OE506")
        | Q (codigo_oe="OE507") | Q (codigo_oe="OE508") | Q (codigo_oe="OE509") | Q (codigo_oe="OE51") | Q (codigo_oe="OE59") | Q (codigo_oe="OE71") |  Q (codigo_oe="OE600")
        | Q (creado_por__username="alsoto")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "cequinonesl":  #cequinonesl

        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE13") | Q(codigo_oe="OE221") | Q(codigo_oe="OE222") | Q(codigo_oe="OE360")
        | Q(codigo_oe="OE37") | Q(codigo_oe="OE451") | Q (codigo_oe="OE494") | Q (codigo_oe="OE9") | Q (creado_por__username="cequinonesl")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "cequinonesl_2":  #cequinonesl_2 *** se crea porque el usuario no presta credenciales dane

        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE13") | Q(codigo_oe="OE221") | Q(codigo_oe="OE222") | Q(codigo_oe="OE360")
        | Q(codigo_oe="OE37") | Q(codigo_oe="OE451") | Q (codigo_oe="OE494") | Q (codigo_oe="OE9") | Q (creado_por__username="cequinonesl_2")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)       

    elif user.is_authenticated == True and str(user) == "aramosh": #dcpenab no continua trabajando en dane y se reemplaza por aramosh
        
        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE129") | Q(codigo_oe="OE20") | Q(codigo_oe="OE297") | Q(codigo_oe="OE35")
        | Q(codigo_oe="OE36") | Q(codigo_oe="OE449") | Q(codigo_oe="OE490") | Q (codigo_oe="OE50") | Q (codigo_oe="OE556") | Q (codigo_oe="OE620") | Q (creado_por__username="aramosh")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "eagarzona": #eagarzona
        
        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE10") | Q(codigo_oe="OE104") | Q(codigo_oe="OE12") | Q(codigo_oe="OE239")
        | Q(codigo_oe="OE240") | Q(codigo_oe="OE533") | Q (creado_por__username="eagarzona")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "gaportillac":  #gaportillac
        
        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE208") | Q(codigo_oe="OE450") | Q (creado_por__username="gaportillac")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "ircastillop": #ircastillop
        
        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE101") | Q(codigo_oe="OE103") | Q(codigo_oe="OE162") | Q(codigo_oe="OE283")
        | Q(codigo_oe="OE493") | Q (creado_por__username="ircastillop")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "jjquinchuac":  #jjquinchuac
        
        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE284") | Q (creado_por__username="jjquinchuac")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    
    elif user.is_authenticated == True and str(user) == "laguion": #laguion
         
        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE14") | Q(codigo_oe="OE15") | Q(codigo_oe="OE21") | Q(codigo_oe="OE223")
        | Q(codigo_oe="OE358") | Q (codigo_oe="OE359") | Q (codigo_oe="OE361") | Q (codigo_oe="OE487") | Q(codigo_oe="OE557") 
        | Q (creado_por__username="laguion") ).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "lahernandezv": #lahernandezv
         
        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE163") | Q(codigo_oe="OE164") | Q(codigo_oe="OE165") | Q(codigo_oe="OE166")
        | Q(codigo_oe="OE305") | Q (codigo_oe="OE362") | Q (codigo_oe="OE363") | Q (codigo_oe="OE364") | Q(codigo_oe="OE38") | Q(codigo_oe="OE52") 
        | Q (creado_por__username="lahernandezv")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "mhsanchezf": #mhsanchezf
         
        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE11") | Q(codigo_oe="OE130") | Q(codigo_oe="OE161") | Q(codigo_oe="OE167")
        | Q(codigo_oe="OE33") | Q (codigo_oe="OE365") | Q (codigo_oe="OE452") | Q (codigo_oe="OE492") | Q (creado_por__username="mhsanchezf")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "opandradem": #opandradem
         
        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE206") | Q(codigo_oe="OE207") | Q(codigo_oe="OE473") | Q (creado_por__username="opandradem")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    ### End Usuarios Dimpe

    ### Usuarios Cuentas nacionales

    elif user.is_authenticated == True and str(user) == "apcasasv": # apcasasv
         
        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE226") | Q(codigo_oe="OE72") | Q (codigo_oe="OE566") | Q (creado_por__username="apcasasv")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "dacobaledam":  #dacobaledam
         
        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE558") | Q(codigo_oe="OE559") | Q(codigo_oe="OE560") | Q(codigo_oe="OE561")
        | Q(codigo_oe="OE562") | Q (codigo_oe="OE563") | Q (codigo_oe="OE564") | Q (creado_por__username="dacobaledam")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)


    elif user.is_authenticated == True and str(user) == "dcubidesl": #dcubidesl
         
        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE131") | Q(codigo_oe="OE209") | Q(codigo_oe="OE22") | Q(codigo_oe="OE520")
        | Q(codigo_oe="OE535") | Q (codigo_oe="OE570") | Q (codigo_oe="OE571") | Q (codigo_oe="OE572") | Q (codigo_oe="OE573") | Q (codigo_oe="OE574")
        | Q (codigo_oe="OE575") | Q (creado_por__username="dcubidesl")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "eforerom": #eforerom
         
        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE370") | Q(codigo_oe="OE567") | Q(codigo_oe="OE367") | Q (creado_por__username="eforerom")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "kjdavilaa": #kjdavilaa
         
        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE370") | Q (creado_por__username="kjdavilaa")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "jaherrerab":  #jaherrerab
         
        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE271") | Q(codigo_oe="OE569")| Q (creado_por__username="jaherrerab")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "jepalaciosm": #jepalaciosm
         
        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE272") | Q(codigo_oe="OE273") | Q(codigo_oe="OE274")
        | Q (creado_por__username="jepalaciosm")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "jpcardozot":  #jpcardozot
        
        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE170") | Q(codigo_oe="OE270") | Q(codigo_oe="OE369") | Q(codigo_oe="OE568") | Q(codigo_oe="OE370")
        | Q (creado_por__username="jpcardozot")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    elif user.is_authenticated == True and str(user) == "nchudtr":  #nchudtr
        
        oes = OperacionEstadistica.objects.filter(Q (codigo_oe="OE272") | Q(codigo_oe="OE273") | Q(codigo_oe="OE274")
        | Q (creado_por__username="nchudtr")).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)
    ## end cuentas nacionales
    
    elif user.is_authenticated == True and user.profile.role.id == 1 or user.profile.role.id == 5 or user.profile.role.id == 6: ##administrador o rol revisor o rol evaluador de calidad
            
        oes = OperacionEstadistica.objects.all().only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        novedad = []
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)   

    else:
        oes = OperacionEstadistica.objects.filter(entidad_id=user.profile.entidad.id).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').exclude(nombre_est_id=6).order_by('nombre_oe')
        oe_filter = TopicsFilter(request.GET, queryset=oes)
        nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)
        
    page = request.GET.get('page', 1)
    paginator = Paginator(oes, 10)
    results = oes.count()

    try:
        operacion = paginator.page(page)
    except PageNotAnInteger:
        operacion = paginator.page(1)
    except EmptyPage:
        operacion = paginator.page(paginator.num_pages)

    index = operacion.number - 1
    max_index = len(paginator.page_range)
    start_index = index - 3 if index >= 3 else 0
    end_index = index + 3 if index <= max_index - 3 else max_index
    page_start = 1
    page_end = max_index
    page_range = paginator.page_range[start_index:end_index]
    page_number= int(page)
        
    
    return render(request, 'ooee/all_ooee.html',  {'oes': oes, 'filter': oe_filter, 'count_entities': count_entities,
    'operacion': operacion, 'page_range': page_range, 'page_end': page_end, 'page_start': page_start, 'nameoeFilter': nameoeFilter, 'results': results})


## filtros por nombre para que funcione con paginación django all_ooee
class FilterNameAdminOE(TemplateView):
    def get(self, request, *args, **kwargs):
        
        nombre = request.GET.get('nombre_oe') 
        user = request.user

        ### usuarios planificación

        if user.is_authenticated == True and str(user) == "sdazag": # 1 sdazag
            operacion = OperacionEstadistica.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28)).filter(Q (nombre_oe__icontains=nombre)).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "dvlizarazog": # 4
        
            operacion = OperacionEstadistica.objects.filter(Q (tema_id=9) | Q (tema_id=29) | Q (tema_id=30)).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "ancardenasc": # 5frherreran se cambia por gavargasr
        
            operacion = OperacionEstadistica.objects.filter(tema_id=10).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "lhsanchezz": # 6 no continua con contrato
        
            operacion = OperacionEstadistica.objects.filter(Q (tema_id=3) | Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "pczambranog": # 6
        
            operacion = OperacionEstadistica.objects.filter(Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)
        
        elif user.is_authenticated == True and str(user) == "mjpradag": # 7 mppulidor se cambia por mjpradag 

            operacion = OperacionEstadistica.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27)).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)
       
        elif user.is_authenticated == True and str(user) == "eeguayazans": # 8 eeguayazans Edgar Eduardo Guayazan Sierra

            operacion = OperacionEstadistica.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20)).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)
        
        elif user.is_authenticated == True and str(user) == "mlbarretob": # 9

            operacion = OperacionEstadistica.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23)).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)
        
    ### end usuarios planificación

    ### Usuarios Dirección de Censos y Demografía ###
        elif user.is_authenticated == True and str(user) == "emvallec": #emvallec

            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE16") | Q (creado_por__username="emvallec")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "hcoteo":  #hcoteo

            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE39") | Q(codigo_oe="OE53") | Q (creado_por__username="hcoteo")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "mfospinab":  #mfospinab

            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE168") | Q(codigo_oe="OE169") | Q (creado_por__username="mfospinab")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)
    ### end Usuarios Dirección de Censos y Demografía ####

    ### Usuarios Dimpe

        elif user.is_authenticated == True and str(user) == "ambernalc": # 2 ambernalc

            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE34") | Q(codigo_oe="OE51") | Q(codigo_oe="OE59")
            | Q(codigo_oe="OE71") | Q(codigo_oe="OE268") | Q(codigo_oe="OE269") | Q(codigo_oe="OE427") | Q (codigo_oe="OE600") 
            | Q (creado_por__username="ambernalc")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')   
        
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "alsoto":  #alsoto

            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE268") | Q(codigo_oe="OE269") | Q(codigo_oe="OE34") | Q(codigo_oe="OE427")
            | Q(codigo_oe="OE491") | Q(codigo_oe="OE495") | Q (codigo_oe="OE496") | Q (codigo_oe="OE497") | Q (codigo_oe="OE498") | Q (codigo_oe="OE499")
            | Q (codigo_oe="OE500") | Q (codigo_oe="OE501") | Q (codigo_oe="OE502") | Q (codigo_oe="OE503") | Q (codigo_oe="OE504") | Q (codigo_oe="OE505") | Q (codigo_oe="OE506")
            | Q (codigo_oe="OE507") | Q (codigo_oe="OE508") | Q (codigo_oe="OE509") | Q (codigo_oe="OE51") | Q (codigo_oe="OE59") | Q (codigo_oe="OE71") |  Q (codigo_oe="OE600")
            | Q (creado_por__username="alsoto")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "cequinonesl":  #cequinonesl

            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE13") | Q(codigo_oe="OE221") | Q(codigo_oe="OE222") | Q(codigo_oe="OE360")
            | Q(codigo_oe="OE37") | Q(codigo_oe="OE451") | Q (codigo_oe="OE494") | Q (codigo_oe="OE9") | Q (creado_por__username="cequinonesl")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "cequinonesl_2":  #cequinonesl_2 *** se crea porq usuario no presta credenciales dane

            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE13") | Q(codigo_oe="OE221") | Q(codigo_oe="OE222") | Q(codigo_oe="OE360")
            | Q(codigo_oe="OE37") | Q(codigo_oe="OE451") | Q (codigo_oe="OE494") | Q (codigo_oe="OE9") | Q (creado_por__username="cequinonesl_2")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "aramosh": #dcpenab no continua trabajando en dane y se reemplaza por aramosh
        
            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE129") | Q(codigo_oe="OE20") | Q(codigo_oe="OE297") | Q(codigo_oe="OE35")
            | Q(codigo_oe="OE36") | Q(codigo_oe="OE449") | Q(codigo_oe="OE490") | Q (codigo_oe="OE50") | Q (codigo_oe="OE556") | Q (codigo_oe="OE620")
            | Q (creado_por__username="aramosh")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "eagarzona": #eagarzona
        
            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE10") | Q(codigo_oe="OE104") | Q(codigo_oe="OE12") | Q(codigo_oe="OE239")
            | Q(codigo_oe="OE240") | Q(codigo_oe="OE533") | Q (creado_por__username="eagarzona")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "gaportillac":  #gaportillac
        
            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE208") | Q(codigo_oe="OE450") | Q (creado_por__username="gaportillac")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "ircastillop": #ircastillop
        
            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE101") | Q(codigo_oe="OE103") | Q(codigo_oe="OE162") | Q(codigo_oe="OE283")
            | Q(codigo_oe="OE493") | Q (creado_por__username="ircastillop")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "jjquinchuac":  #jjquinchuac
        
            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE284") | Q (creado_por__username="jjquinchuac")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "laguion": #laguion
         
            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE14") | Q(codigo_oe="OE15") | Q(codigo_oe="OE21") | Q(codigo_oe="OE223")
            | Q(codigo_oe="OE358") | Q (codigo_oe="OE359") | Q (codigo_oe="OE361") | Q (codigo_oe="OE487") | Q(codigo_oe="OE557") 
            | Q (creado_por__username="laguion")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "lahernandezv": #lahernandezv
         
            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE163") | Q(codigo_oe="OE164") | Q(codigo_oe="OE165") | Q(codigo_oe="OE166")
            | Q(codigo_oe="OE305") | Q (codigo_oe="OE362") | Q (codigo_oe="OE363") | Q (codigo_oe="OE364") | Q(codigo_oe="OE38") | Q(codigo_oe="OE52") 
            | Q (creado_por__username="lahernandezv")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "mhsanchezf": #mhsanchezf
         
            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE11") | Q(codigo_oe="OE130") | Q(codigo_oe="OE161") | Q(codigo_oe="OE167")
            | Q(codigo_oe="OE33") | Q (codigo_oe="OE365") | Q (codigo_oe="OE452") | Q (codigo_oe="OE492") | Q (creado_por__username="mhsanchezf")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "opandradem": #opandradem
         
            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE206") | Q(codigo_oe="OE207") | Q(codigo_oe="OE473") | Q (creado_por__username="opandradem")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

    ### End Usuarios Dimpe
    ### Usuarios Cuentas nacionales

        elif user.is_authenticated == True and str(user) == "apcasasv": # apcasasv
         
            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE226") | Q(codigo_oe="OE72") | Q (codigo_oe="OE566") | Q (creado_por__username="apcasasv")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')  
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "dacobaledam":  #dacobaledam
         
            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE558") | Q(codigo_oe="OE559") | Q(codigo_oe="OE560") | Q(codigo_oe="OE561")
            | Q(codigo_oe="OE562") | Q (codigo_oe="OE563") | Q (codigo_oe="OE564") | Q (creado_por__username="dacobaledam")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "dcubidesl": #dcubidesl
         
            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE131") | Q(codigo_oe="OE209") | Q(codigo_oe="OE22") | Q(codigo_oe="OE520")
            | Q(codigo_oe="OE535") | Q (codigo_oe="OE570") | Q (codigo_oe="OE571") | Q (codigo_oe="OE572") | Q (codigo_oe="OE573") | Q (codigo_oe="OE574")
            | Q (codigo_oe="OE575") | Q (creado_por__username="dcubidesl")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "eforerom": #eforerom
         
            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE370") | Q(codigo_oe="OE567") | Q(codigo_oe="OE367") | Q (creado_por__username="eforerom")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "kjdavilaa": #kjdavilaa
         
            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE370") | Q (creado_por__username="kjdavilaa")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)
        

        elif user.is_authenticated == True and str(user) == "jaherrerab":  #jaherrerab
         
            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE271") | Q(codigo_oe="OE569")| Q (creado_por__username="jaherrerab")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "jepalaciosm": #jepalaciosm
         
            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE272") | Q(codigo_oe="OE273") | Q(codigo_oe="OE274")
            | Q (creado_por__username="jepalaciosm")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "jpcardozot":  #jpcardozot
        
            operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE170") | Q(codigo_oe="OE270") | Q(codigo_oe="OE369") | Q(codigo_oe="OE568") | Q(codigo_oe="OE370")
            | Q (creado_por__username="jpcardozot")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
        
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        elif user.is_authenticated == True and str(user) == "nchudtr":  #nchudtr
        
            operacion = operacion = OperacionEstadistica.objects.filter(Q (codigo_oe="OE272") | Q(codigo_oe="OE273") | Q(codigo_oe="OE274")
            | Q (creado_por__username="nchudtr")).filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
            
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

    ## end cuentas nacionales
    
        elif user.is_authenticated == True and user.profile.role.id == 1 or user.profile.role.id == 5 or user.profile.role.id == 6: ##administrador o rol revisor o evaluador de calidad
            
            operacion = OperacionEstadistica.objects.filter(nombre_oe__icontains=nombre).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)   

        else:
            operacion = OperacionEstadistica.objects.filter(Q (entidad_id=user.profile.entidad.id)).filter(Q (nombre_oe__icontains=nombre)).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').exclude(nombre_est_id=6).order_by('nombre_oe')
            nameoeFilter = NameFilterOOEE(request.GET, queryset=operacion)

        data = serializers.serialize('json', operacion,
                    fields=('nombre_oe', 'entidad', 'area_tematica', 'tema', 'codigo_oe', 'nombre_est', 'estado_oe_tematico', 'validacion_oe_tematico'), use_natural_foreign_keys=True, use_natural_primary_keys=True)
        return HttpResponse(data, content_type='application/json')




### función ajax para vista all_ooee por filtros
class FilterAllOOEEAjaxView(TemplateView):
    def get(self, request, *args, **kwargs):
        
        id_area_tematica = request.GET.get('area_tematica_id') #opc 1
        id_tema = request.GET.get('tema_id') #opc 2
        id_fase = request.GET.get('fase_id') #opc 3
        id_entidad = request.GET.get('entidad_id') #opc 4

        user = request.user
        #print("auth",user.is_authenticated, "entidad", user.profile.entidad.id)

        ########## user sdazag ################
        if user.is_authenticated == True and str(user) == "sdazag": # 1 sdazag
            oes = OperacionEstadistica.objects.filter(Q (tema=2) | Q (tema=7) | Q (tema=19) | Q (tema=28)).order_by('nombre_oe')
            oe_filter = TopicsFilter(request.GET, queryset=oes)

            ## opc 1 es diferente de vacia
            if  id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_entidad == "":
            
                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica)

            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_entidad == "":

                operacion = oe_filter.qs.filter(tema=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_entidad == "":
                
                operacion = oe_filter.qs.filter(nombre_est=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_entidad != "":
                
                operacion = oe_filter.qs.filter(entidad=id_entidad)

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and id_entidad == "":
                
                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema)  
    
            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_entidad == "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema).filter(nombre_est=id_fase)

            ## opc 1, opc 2, opc 3 es y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_entidad != "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema).filter(nombre_est=id_fase).filter(entidad=id_entidad)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and id_entidad == "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(nombre_est=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and id_entidad != "":

                operacion = oe_filter.qs.filter(area_tematica_id=id_area_tematica).filter(entidad_id=id_entidad)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and id_entidad == "":
        
                operacion = oe_filter.qs.filter(tema=id_tema).filter(nombre_est=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and id_entidad != "":
        
                operacion = oe_filter.qs.filter(tema_id=id_tema).filter(entidad_id=id_entidad)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and id_entidad != "":
        
                operacion = oe_filter.qs.filter(nombre_est=id_fase).filter(entidad=id_entidad)

        #######################  user aobandor se cambia usuario a administrador ###########################  
            """ elif user.is_authenticated == True and str(user) == "aobandor": # 3

            oes = OperacionEstadistica.objects.filter(Q (tema=5) | Q (tema=14) | Q (tema=15) | Q (tema=16) | Q (tema=17) | Q (tema=18)).order_by('nombre_oe')
            oe_filter = TopicsFilter(request.GET, queryset=oes) """

            
        ####################### user dvlizarazog #############################
        elif user.is_authenticated == True and str(user) == "dvlizarazog": # 4
            
            oes = OperacionEstadistica.objects.filter(Q (tema=9) | Q (tema=29) | Q (tema=30)).order_by('nombre_oe')
            oe_filter = TopicsFilter(request.GET, queryset=oes)

            ## opc 1 es diferente de vacia
            if  id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_entidad == "":
            
                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica)

            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_entidad == "":

                operacion = oe_filter.qs.filter(tema=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_entidad == "":
                
                operacion = oe_filter.qs.filter(nombre_est=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_entidad != "":
                
                operacion = oe_filter.qs.filter(entidad=id_entidad)

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and id_entidad == "":
                
                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema)  
    
            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_entidad == "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema).filter(nombre_est=id_fase)

            ## opc 1, opc 2, opc 3 es y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_entidad != "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema).filter(nombre_est=id_fase).filter(entidad=id_entidad)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and id_entidad == "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(nombre_est=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and id_entidad != "":

                operacion = oe_filter.qs.filter(area_tematica_id=id_area_tematica).filter(entidad_id=id_entidad)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and id_entidad == "":
        
                operacion = oe_filter.qs.filter(tema=id_tema).filter(nombre_est=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and id_entidad != "":
        
                operacion = oe_filter.qs.filter(tema_id=id_tema).filter(entidad_id=id_entidad)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and id_entidad != "":
        
                operacion = oe_filter.qs.filter(nombre_est=id_fase).filter(entidad=id_entidad)

        #################### user fherreran ##########################
        elif user.is_authenticated == True and str(user) == "ancardenasc": # 5frherreran se cambia por gavargasr
            
            oes = OperacionEstadistica.objects.filter(tema=10).order_by('nombre_oe')
            oe_filter = TopicsFilter(request.GET, queryset=oes)

            ## opc 1 es diferente de vacia
            if  id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_entidad == "":
            
                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica)

            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_entidad == "":

                operacion = oe_filter.qs.filter(tema=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_entidad == "":
                
                operacion = oe_filter.qs.filter(nombre_est=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_entidad != "":
                
                operacion = oe_filter.qs.filter(entidad=id_entidad)

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and id_entidad == "":
                
                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema)  
    
            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_entidad == "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema).filter(nombre_est=id_fase)

            ## opc 1, opc 2, opc 3 es y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_entidad != "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema).filter(nombre_est=id_fase).filter(entidad=id_entidad)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and id_entidad == "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(nombre_est=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and id_entidad != "":

                operacion = oe_filter.qs.filter(area_tematica_id=id_area_tematica).filter(entidad_id=id_entidad)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and id_entidad == "":
        
                operacion = oe_filter.qs.filter(tema=id_tema).filter(nombre_est=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and id_entidad != "":
        
                operacion = oe_filter.qs.filter(tema_id=id_tema).filter(entidad_id=id_entidad)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and id_entidad != "":
        
                operacion = oe_filter.qs.filter(nombre_est=id_fase).filter(entidad=id_entidad)

        ########################### user lhsanchezz se remplaza por  pczambranog ####################  
        elif user.is_authenticated == True and str(user) == "pczambranog": # 6
            
            oes = OperacionEstadistica.objects.filter(Q (tema=11) | Q (tema=12) | Q (tema=25)).order_by('nombre_oe')
            oe_filter = TopicsFilter(request.GET, queryset=oes)

            ## opc 1 es diferente de vacia
            if  id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_entidad == "":
            
                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica)

            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_entidad == "":

                operacion = oe_filter.qs.filter(tema=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_entidad == "":
                
                operacion = oe_filter.qs.filter(nombre_est=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_entidad != "":
                
                operacion = oe_filter.qs.filter(entidad=id_entidad)

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and id_entidad == "":
                
                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema)  
    
            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_entidad == "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema).filter(nombre_est=id_fase)

            ## opc 1, opc 2, opc 3 es y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_entidad != "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema).filter(nombre_est=id_fase).filter(entidad=id_entidad)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and id_entidad == "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(nombre_est=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and id_entidad != "":

                operacion = oe_filter.qs.filter(area_tematica_id=id_area_tematica).filter(entidad_id=id_entidad)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and id_entidad == "":
        
                operacion = oe_filter.qs.filter(tema=id_tema).filter(nombre_est=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and id_entidad != "":
        
                operacion = oe_filter.qs.filter(tema_id=id_tema).filter(entidad_id=id_entidad)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and id_entidad != "":
        
                operacion = oe_filter.qs.filter(nombre_est=id_fase).filter(entidad=id_entidad)

        ####################### user mjpradag ###########################3
        elif user.is_authenticated == True and str(user) == "mjpradag": # 7  mppulidor se cambia por mjpradag 

            oes = OperacionEstadistica.objects.filter(Q (tema=13) | Q (tema=24) | Q (tema=26) | Q (tema=27)).order_by('nombre_oe')
            oe_filter = TopicsFilter(request.GET, queryset=oes)

            ## opc 1 es diferente de vacia
            if  id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_entidad == "":
            
                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica)

            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_entidad == "":

                operacion = oe_filter.qs.filter(tema=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_entidad == "":
                
                operacion = oe_filter.qs.filter(nombre_est=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_entidad != "":
                
                operacion = oe_filter.qs.filter(entidad=id_entidad)

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and id_entidad == "":
                
                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema)  
    
            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_entidad == "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema).filter(nombre_est=id_fase)

            ## opc 1, opc 2, opc 3 es y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_entidad != "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema).filter(nombre_est=id_fase).filter(entidad=id_entidad)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and id_entidad == "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(nombre_est=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and id_entidad != "":

                operacion = oe_filter.qs.filter(area_tematica_id=id_area_tematica).filter(entidad_id=id_entidad)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and id_entidad == "":
        
                operacion = oe_filter.qs.filter(tema=id_tema).filter(nombre_est=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and id_entidad != "":
        
                operacion = oe_filter.qs.filter(tema_id=id_tema).filter(entidad_id=id_entidad)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and id_entidad != "":
        
                operacion = oe_filter.qs.filter(nombre_est=id_fase).filter(entidad=id_entidad)

        #################### user eeguayazans ########################
        elif user.is_authenticated == True and str(user) == "eeguayazans": # 8 eeguayazans Edgar Eduardo Guayazan Sierra

            oes = OperacionEstadistica.objects.filter(Q (tema=4) | Q (tema=8) | Q (tema=20)).order_by('nombre_oe')
            oe_filter = TopicsFilter(request.GET, queryset=oes)

            ## opc 1 es diferente de vacia
            if  id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_entidad == "":
            
                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica)

            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_entidad == "":

                operacion = oe_filter.qs.filter(tema=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_entidad == "":
                
                operacion = oe_filter.qs.filter(nombre_est=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_entidad != "":
                
                operacion = oe_filter.qs.filter(entidad=id_entidad)

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and id_entidad == "":
                
                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema)  
    
            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_entidad == "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema).filter(nombre_est=id_fase)

            ## opc 1, opc 2, opc 3 es y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_entidad != "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema).filter(nombre_est=id_fase).filter(entidad=id_entidad)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and id_entidad == "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(nombre_est=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and id_entidad != "":

                operacion = oe_filter.qs.filter(area_tematica_id=id_area_tematica).filter(entidad_id=id_entidad)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and id_entidad == "":
        
                operacion = oe_filter.qs.filter(tema=id_tema).filter(nombre_est=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and id_entidad != "":
        
                operacion = oe_filter.qs.filter(tema_id=id_tema).filter(entidad_id=id_entidad)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and id_entidad != "":
        
                operacion = oe_filter.qs.filter(nombre_est=id_fase).filter(entidad=id_entidad)

        ############################user mlbarretob ######################
        elif user.is_authenticated == True and str(user) == "mlbarretob": # 9

            oes = OperacionEstadistica.objects.filter(Q (tema=21) | Q (tema=22) | Q (tema=23)).order_by('nombre_oe')
            oe_filter = TopicsFilter(request.GET, queryset=oes)

            ## opc 1 es diferente de vacia
            if  id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_entidad == "":
            
                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica)

            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_entidad == "":

                operacion = oe_filter.qs.filter(tema=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_entidad == "":
                
                operacion = oe_filter.qs.filter(nombre_est=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_entidad != "":
                
                operacion = oe_filter.qs.filter(entidad=id_entidad)

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and id_entidad == "":
                
                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema)  
    
            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_entidad == "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema).filter(nombre_est=id_fase)

            ## opc 1, opc 2, opc 3 es y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_entidad != "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema).filter(nombre_est=id_fase).filter(entidad=id_entidad)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and id_entidad == "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(nombre_est=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and id_entidad != "":

                operacion = oe_filter.qs.filter(area_tematica_id=id_area_tematica).filter(entidad_id=id_entidad)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and id_entidad == "":
        
                operacion = oe_filter.qs.filter(tema=id_tema).filter(nombre_est=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and id_entidad != "":
        
                operacion = oe_filter.qs.filter(tema_id=id_tema).filter(entidad_id=id_entidad)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and id_entidad != "":
        
                operacion = oe_filter.qs.filter(nombre_est=id_fase).filter(entidad=id_entidad)

        ############################ Rol administrador y revisor #######################
        elif user.is_authenticated == True and user.profile.role.id == 1 or user.profile.role.id == 5: 
            
            oes = OperacionEstadistica.objects.all().order_by('nombre_oe')
            oe_filter = TopicsFilter(request.GET, queryset=oes)

            ## opc 1 es diferente de vacia
            if  id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_entidad == "":
            
                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica)

            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_entidad == "":

                operacion = oe_filter.qs.filter(tema=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_entidad == "":
                
                operacion = oe_filter.qs.filter(nombre_est=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_entidad != "":
                
                operacion = oe_filter.qs.filter(entidad=id_entidad)

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and id_entidad == "":
                
                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema)  
    
            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_entidad == "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema).filter(nombre_est=id_fase)

            ## opc 1, opc 2, opc 3 es y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_entidad != "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema).filter(nombre_est=id_fase).filter(entidad=id_entidad)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and id_entidad == "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(nombre_est=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and id_entidad != "":

                operacion = oe_filter.qs.filter(area_tematica_id=id_area_tematica).filter(entidad_id=id_entidad)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and id_entidad == "":
        
                operacion = oe_filter.qs.filter(tema=id_tema).filter(nombre_est=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and id_entidad != "":
        
                operacion = oe_filter.qs.filter(tema_id=id_tema).filter(entidad_id=id_entidad)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and id_entidad != "":
        
                operacion = oe_filter.qs.filter(nombre_est=id_fase).filter(entidad=id_entidad)

         ###########################  Rol fuente ###############################
        else:
            oes = OperacionEstadistica.objects.filter(entidad=user.profile.entidad.id).exclude(nombre_est_id=6).order_by('nombre_oe')
            oe_filter = TopicsFilter(request.GET, queryset=oes)

            ## opc 1 es diferente de vacia
            if  id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_entidad == "":
            
                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).exclude(nombre_est_id=6)

            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_entidad == "":

                operacion = oe_filter.qs.filter(tema=id_tema).exclude(nombre_est_id=6)  
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_entidad == "":
                
                operacion = oe_filter.qs.filter(nombre_est=id_fase).exclude(nombre_est_id=6)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_entidad != "":
                
                operacion = oe_filter.qs.filter(entidad=id_entidad).exclude(nombre_est_id=6)

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and id_entidad == "":
                
                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema).exclude(nombre_est_id=6)  
    
            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_entidad == "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema).filter(nombre_est=id_fase).exclude(nombre_est_id=6)

            ## opc 1, opc 2, opc 3 es y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_entidad != "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema).filter(nombre_est=id_fase).filter(entidad=id_entidad).exclude(nombre_est_id=6)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and id_entidad == "":

                operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(nombre_est=id_fase).exclude(nombre_est_id=6)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and id_entidad != "":

                operacion = oe_filter.qs.filter(area_tematica_id=id_area_tematica).filter(entidad_id=id_entidad).exclude(nombre_est_id=6)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and id_entidad == "":
        
                operacion = oe_filter.qs.filter(tema=id_tema).filter(nombre_est=id_fase).exclude(nombre_est_id=6)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and id_entidad != "":
        
                operacion = oe_filter.qs.filter(tema_id=id_tema).filter(entidad_id=id_entidad).exclude(nombre_est_id=6)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and id_entidad != "":
        
                operacion = oe_filter.qs.filter(nombre_est=id_fase).filter(entidad=id_entidad).exclude(nombre_est_id=6)

        data = serializers.serialize('json', operacion,
                    fields=('nombre_oe', 'entidad', 'area_tematica', 'tema', 'codigo_oe', 'nombre_est', 'estado_oe_tematico', 'validacion_oe_tematico'), use_natural_foreign_keys=True, use_natural_primary_keys=True)
        return HttpResponse(data, content_type='application/json')


############ end vista para mostrar las operaciones según el rol y la entidad


from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

## vista de consulta de todos los OOEE EN ESTADO PUBLICADO
def consultationModule(request):
    
    results = OperacionEstadistica.objects.filter(nombre_est_id=5).count()
    count_entities = Entidades_oe.objects.filter(estado_id=1).count()
    oes = OperacionEstadistica.objects.filter(nombre_est_id=5).only('nombre_oe', 'entidad', 'area_tematica', 'tema', 'estado_oe_tematico', 'nombre_est').order_by('nombre_oe')
    count_total_oe = oes.count()
    oe_filter = TopicsFilter(request.GET, queryset=oes)
    nameoeFilter = NameFilterOOEE(request.GET, queryset=oes)

    page = request.GET.get('page', 1)
    paginator = Paginator(oes, 10)

    try:
        operacion = paginator.page(page)
    except PageNotAnInteger:
        operacion = paginator.page(1)
    except EmptyPage:
        operacion = paginator.page(paginator.num_pages)

    index = operacion.number - 1
    max_index = len(paginator.page_range)
    start_index = index - 3 if index >= 3 else 0
    end_index = index + 3 if index <= max_index - 3 else max_index
    page_start = 1
    page_end = max_index
    page_range = paginator.page_range[start_index:end_index]
    page_number= int(page)
    #result_show = show_operaciones*page_number if show_operaciones == 12 else results
    #show_incre = page_number if page_number == 1 else result_show - show_operaciones + 1

    
    return render(request, 'ooee/modulo_consulta_oe.html',  {'oes': oes, 'filter': oe_filter,
    'count_entities': count_entities, 'count_total_oe': count_total_oe, 'operacion': operacion,
    'page_range': page_range, 'page_end': page_end, 'page_start': page_start, 'results': results, 'nameoeFilter': nameoeFilter})


## filtros por nombre para que funcione con paginación django modulo_cosulta_oe
class FilterByNameoe(TemplateView):
    def get(self, request, *args, **kwargs):
        
        nombre = request.GET.get('nombre_oe') #opc 1
        #print("que estoy recibiendo", nombre)
        operacion = OperacionEstadistica.objects.filter(nombre_est=5).filter(nombre_oe__icontains=nombre)
        oe_filter = NameFilterOOEE(request.GET, queryset=operacion)
        data = serializers.serialize('json', operacion,
                    fields=('nombre_oe', 'entidad', 'area_tematica', 'tema', 'codigo_oe', 'nombre_est', 'estado_oe_tematico', 'validacion_oe_tematico'), use_natural_foreign_keys=True, use_natural_primary_keys=True)
        return HttpResponse(data, content_type='application/json')


### modulo de consulta *** con base completa ooee publicados
class SearchAjaxView(TemplateView):
    def get(self, request, *args, **kwargs):
        
        id_entidad = request.GET.get('entidad') #opc 1
        id_area_tematica = request.GET.get('area_tematica') #opc 2
        id_tema = request.GET.get('tema') #opc 3

        ## opc 1 es diferente de vacia
        if id_entidad != "" and id_area_tematica == "" and  id_tema == "":

            oes = OperacionEstadistica.objects.filter(nombre_est=5)
            oe_filter = TopicsFilter(request.GET, queryset=oes)
            operacion = oe_filter.qs.filter(entidad=id_entidad)

        ## opc 1 y opc 2 es diferente de vacia
        elif  id_entidad != "" and id_area_tematica != "" and  id_tema == "":
        
            oes = OperacionEstadistica.objects.filter(nombre_est=5)
            oe_filter = TopicsFilter(request.GET, queryset=oes)
            operacion = oe_filter.qs.filter(entidad=id_entidad).filter(area_tematica=id_area_tematica)
        
        ## opc 1, opc 2 y opc 3 es diferente de vacia
        elif id_entidad != "" and id_area_tematica != "" and  id_tema != "":

            oes = OperacionEstadistica.objects.filter(nombre_est=5)
            oe_filter = TopicsFilter(request.GET, queryset=oes)
            operacion = oe_filter.qs.filter(entidad=id_entidad).filter(area_tematica=id_area_tematica).filter(tema=id_tema)
        
        ## opcion 1 y opcion 3 es diferente de vacia
        elif id_entidad != "" and id_area_tematica == "" and  id_tema != "":
        
            oes = OperacionEstadistica.objects.filter(nombre_est=5)
            oe_filter = TopicsFilter(request.GET, queryset=oes)
            operacion = oe_filter.qs.filter(entidad=id_entidad).filter(tema=id_tema)

        #opcion 2 es diferente de vacia
        elif id_entidad == "" and id_area_tematica != "" and  id_tema == "":

            oes = OperacionEstadistica.objects.filter(nombre_est=5)
            oe_filter = TopicsFilter(request.GET, queryset=oes)
            operacion = oe_filter.qs.filter(area_tematica=id_area_tematica)

        ## opc 2 y opc 3 es diferente de vacia
        elif id_entidad == "" and id_area_tematica != "" and  id_tema != "":

            oes = OperacionEstadistica.objects.filter(nombre_est=5)
            oe_filter = TopicsFilter(request.GET, queryset=oes)
            operacion = oe_filter.qs.filter(area_tematica=id_area_tematica).filter(tema=id_tema)

        #opcion 3 es diferente de vacia
        elif id_entidad == "" and id_area_tematica == "" and  id_tema != "":

            oes = OperacionEstadistica.objects.filter(nombre_est=5)
            oe_filter = TopicsFilter(request.GET, queryset=oes)
            operacion = oe_filter.qs.filter(tema=id_tema)

        else:
            operacion = []
        
        data = serializers.serialize('json', operacion,
                    fields=('nombre_oe', 'entidad', 'area_tematica', 'tema', 'codigo_oe', 'nombre_est', 'estado_oe_tematico', 'validacion_oe_tematico'), use_natural_foreign_keys=True, use_natural_primary_keys=True)
        return HttpResponse(data, content_type='application/json')


###########  vistas de consulta externas 

def homeConsultaOE(request):
 
    count_entities = Entidades_oe.objects.filter(estado_id=1).count() #entidades publicadas 
    
    ## OOEE
    
    count_oe_econo = OperacionEstadistica.objects.filter(nombre_est_id=5).filter(area_tematica_id=1).count() 
    count_oe_socioDe =  OperacionEstadistica.objects.filter(nombre_est_id=5).filter(area_tematica_id=2).count() 
    count_oe_amb = OperacionEstadistica.objects.filter(nombre_est_id=5).filter(area_tematica_id=3).count()
    ## RRAA

    count_ra_econo = RegistroAdministrativo.objects.filter(sist_estado_id=5).filter(area_temat_id=1).count() 
    count_ra_socioDe =  RegistroAdministrativo.objects.filter(sist_estado_id=5).filter(area_temat_id=2).count()
    count_ra_amb = RegistroAdministrativo.objects.filter(sist_estado_id=5).filter(area_temat_id=3).count()

    ## DDI
    count_ddi_econo = demandaInfor.objects.filter(nombre_est_id=5).filter(area_tem_id=1).count() 
    count_ddi_socioDe = demandaInfor.objects.filter(nombre_est_id=5).filter(area_tem_id=2).count()
    count_ddi_amb = demandaInfor.objects.filter(nombre_est_id=5).filter(area_tem_id=3).count()

    #total de OOEE, RRAA y DDI en estado publicados
    count_total_oe = OperacionEstadistica.objects.filter(nombre_est_id=5).count()
    count_total_ra = RegistroAdministrativo.objects.filter(sist_estado_id=5).count()
    count_total_ddi = demandaInfor.objects.filter(nombre_est_id=5).count()
     
    return render(request, 'ooee/consulta_ooee.html',  {'count_oe_econo': count_oe_econo, 'count_oe_socioDe': count_oe_socioDe,
    'count_oe_amb': count_oe_amb, 'count_ra_econo': count_ra_econo, 'count_ra_socioDe': count_ra_socioDe, 'count_ra_amb': count_ra_amb,
    'count_entities': count_entities, 'count_total_oe': count_total_oe, 'count_total_ra': count_total_ra,
    'count_total_ddi': count_total_ddi, 'count_ddi_econo': count_ddi_econo, 'count_ddi_socioDe': count_ddi_socioDe,
    'count_ddi_amb': count_ddi_amb})


############################## modulo para iframe de pagina sen ####################################

def homeConsultaSEN(request):

    count_entities = Entidades_oe.objects.filter(estado_id=1).count() 
    
    ## OOEE
   
    count_oe_econo = OperacionEstadistica.objects.filter(nombre_est_id=5).filter(area_tematica_id=1).count()
    count_oe_socioDe =  OperacionEstadistica.objects.filter(nombre_est_id=5).filter(area_tematica_id=2).count()
    count_oe_amb = OperacionEstadistica.objects.filter(nombre_est_id=5).filter(area_tematica_id=3).count()
    ## RRAA 
    count_ra_econo = RegistroAdministrativo.objects.filter(sist_estado_id=5).filter(area_temat_id=1).count() 
    count_ra_socioDe =  RegistroAdministrativo.objects.filter(sist_estado_id=5).filter(area_temat_id=2).count() 
    count_ra_amb = RegistroAdministrativo.objects.filter(sist_estado_id=5).filter(area_temat_id=3).count()

    ## DDI
   
    count_ddi_econo = demandaInfor.objects.filter(nombre_est_id=5).filter(area_tem_id=1).count()
    count_ddi_socioDe = demandaInfor.objects.filter(nombre_est_id=5).filter(area_tem_id=2).count()
    count_ddi_amb = demandaInfor.objects.filter(nombre_est_id=5).filter(area_tem_id=3).count()
    
    #total de OOEE Y RRAA en estado publicados 
    count_total_oe = OperacionEstadistica.objects.filter(nombre_est_id=5).count()
    count_total_ra = RegistroAdministrativo.objects.filter(sist_estado_id=5).count()
    count_total_ddi = demandaInfor.objects.filter(nombre_est_id=5).count()
     
    return render(request, 'ooee/modulo_consulta_sen.html',  {'count_oe_econo': count_oe_econo, 'count_oe_socioDe': count_oe_socioDe,
    'count_oe_amb': count_oe_amb, 'count_ra_econo': count_ra_econo, 'count_ra_socioDe': count_ra_socioDe, 'count_ra_amb': count_ra_amb,
    'count_entities': count_entities, 'count_total_oe': count_total_oe, 'count_total_ra': count_total_ra,
    'count_total_ddi': count_total_ddi, 'count_ddi_econo': count_ddi_econo, 'count_ddi_socioDe': count_ddi_socioDe,
    'count_ddi_amb': count_ddi_amb})



def detailOOEE(request, pk):
    
    entities = Entidades_oe.objects.filter(estado_id=1) #entidades publicadas
    count_entities = entities.count()
    oe = get_object_or_404(OperacionEstadistica, pk=pk)
    entidadFaseFieldText = MB_EntidadFases.objects.filter(ooee_id=oe.pk)
    normaFieldText = MB_Norma.objects.get(ooee_id=oe.pk)
    requerimientoFieldText = MB_Requerimientos.objects.get(ooee_id=oe.pk)
    priusuariosFieldText = MB_PrinUsuarios.objects.get(ooee_id=oe.pk)
    uni_observacionFieldText = MC_UnidadObservacion.objects.get(ooee_id=oe.pk)
    obtencionDatoFieldText = MC_ObtencionDato.objects.get(ooee_id=oe.pk)
    probabiFieldText = MC_MuestreoProbabilistico.objects.get(ooee_id=oe.pk)
    no_probabiFieldText = MC_MuestreoNoProbabilistico.objects.get(ooee_id=oe.pk)
    tipoMarcoFieldText = MC_TipoMarco.objects.get(ooee_id=oe.pk)
    docsDesFieldText = MC_DocsDesarrollo.objects.get(ooee_id=oe.pk)
    listaConcFieldText = MC_ConceptosEstandarizados.objects.get(ooee_id=oe.pk)
    nombreClaFieldText = MC_Clasificaciones.objects.get(ooee_id=oe.pk)
    cobGeoFieldText = MC_CoberturaGeografica.objects.get(ooee_id=oe.pk)
    opcDesagFieldText = MC_DesagregacionInformacion.objects.get(ooee_id=oe.pk)
    fuentesFieldText = MC_FuenteFinanciacion.objects.get(ooee_id=oe.pk)
    listasVarFieldText = MC_listaVariable.objects.filter(ooee_id=oe.pk) #pregunta 14
    resultEstaFieldText = MC_ResultadoEstadistico.objects.filter(ooee_id=oe.pk) #pregunta 15
    #print("filter",resultEstaFieldText)
    medObtFieldText = MD_MedioDatos.objects.get(ooee_id=oe.pk)
    periodicidadFieldText = MD_PeriodicidadOe.objects.get(ooee_id=oe.pk)
    herraProcFieldText = MD_HerramProcesamiento.objects.get(ooee_id=oe.pk)
    anaResuFieldText = ME_AnalisisResultados.objects.get(ooee_id=oe.pk)
    medioDifFieldText = MF_MediosDifusion.objects.get(ooee_id=oe.pk)
    fechapubliFieldText = MF_FechaPublicacion.objects.get(ooee_id=oe.pk)
    frecueDifFieldText = MF_FrecuenciaDifusion.objects.get(ooee_id=oe.pk)
    producDifFieldText = MF_ProductosDifundir.objects.get(ooee_id=oe.pk)
    otroProdFieldText = MF_OtrosProductos.objects.get(ooee_id=oe.pk)
    resulSimilaresFieldText = MF_ResultadosSimilares.objects.get(ooee_id=oe.pk)
    hpSisteInforFieldText = MF_HPSistemaInfo.objects.get(ooee_id=oe.pk)

    ##estado de la novedad
    if NovedadActualizacion.objects.filter(post_oe=oe.pk).exists():
        estadoNovedad = NovedadActualizacion.objects.filter(post_oe=oe.pk)
        for stateNov in estadoNovedad:
            stateNov.est_actualiz
        if stateNov.est_actualiz != None:
            estadoNovOE = stateNov.est_actualiz
            fechaEstado = stateNov.fecha_actualiz
        else:
            estadoNovOE = ""
            fechaEstado = ""
    else:
        estadoNovOE = ""
        fechaEstado = ""
    

    #print("oe",oe.pk)
    return render(request, 'ooee/ooee_detail.html', {'oe': oe, 'entidadFaseFieldText': entidadFaseFieldText, 'normaFieldText': normaFieldText, 
    'requerimientoFieldText': requerimientoFieldText, 'priusuariosFieldText': priusuariosFieldText, 
    'uni_observacionFieldText': uni_observacionFieldText, 'obtencionDatoFieldText': obtencionDatoFieldText,
    'probabiFieldText': probabiFieldText, 'no_probabiFieldText': no_probabiFieldText, 'tipoMarcoFieldText': tipoMarcoFieldText,
    'docsDesFieldText': docsDesFieldText, 'listaConcFieldText': listaConcFieldText, 'nombreClaFieldText': nombreClaFieldText, 
    'cobGeoFieldText': cobGeoFieldText, 'opcDesagFieldText': opcDesagFieldText, 'fuentesFieldText': fuentesFieldText,
    'listasVarFieldText': listasVarFieldText, 'resultEstaFieldText': resultEstaFieldText, 'medObtFieldText': medObtFieldText, 'periodicidadFieldText': periodicidadFieldText,  'herraProcFieldText': herraProcFieldText,
    'anaResuFieldText': anaResuFieldText, 'medioDifFieldText': medioDifFieldText, 'fechapubliFieldText': fechapubliFieldText, 
    'frecueDifFieldText': frecueDifFieldText, 'producDifFieldText': producDifFieldText, 'otroProdFieldText': otroProdFieldText, 
    'resulSimilaresFieldText': resulSimilaresFieldText, 'hpSisteInforFieldText': hpSisteInforFieldText, 
    'estadoNovOE': estadoNovOE, 'fechaEstado': fechaEstado, 'count_entities': count_entities})

########### end vistas de consultas externas


## correo para notificar a la fuente cuando se cambia el estado a devuelto
def sendEmailStatusChange(subject, recipients, body):    
    send_mail(subject, body, 'sen@dane.gov.co', recipients,  fail_silently = False)  

def emailNotificationStateView(request, oe, listaUsuarios):
    
    subject = 'Cambio de estado en formulario de la operación estadística '+ str(oe).replace('\n', '').replace('\r', '')

    recipients = listaUsuarios
      
    body = 'SISTEMA DE IDENTIFICACIÓN Y CARACTERIZACIÓN DE OFERTA Y DEMANDA ESTADÍSTICA DEL SEN\n\n' \
            'La operación estadística: '+ str(oe) + \
			    '\n\n Ha sido DEVUELTO(A)\n' \
                    '\n Por favor ingrese al aplicativo, revise los comentarios generados por el temático y realice los ajustes correspondientes.'

    return subject, recipients, body

## end correo para notificar a la fuente cuando se cambia el estado a devuelto


def sendEmailNotification(subject, recipients, body):    
    send_mail(subject, body, 'sen@dane.gov.co', recipients,  fail_silently = False)  


## envio de notificaciones al editar OOEE
def createEmailEditOOEE(request, oe, temaRes, fieldsEdited, fieldFormsetooee1, fieldFormsetooee2, fieldFormsetooee3):

    #print("que llega", fieldsEdited)
    subject = 'Cambios en formulario de la operación estadística '+ str(oe).replace('\n', '').replace('\r', '')
   
    if temaRes == 2 or temaRes == 7 or temaRes == 19 or temaRes == 28: 

        responsable_oe = 'sdazag@dane.gov.co' # sdazag
                   
    elif temaRes == 5 or temaRes == 14 or temaRes ==  15 or temaRes ==  16 or temaRes ==  17 or temaRes ==  18: 

        responsable_oe = 'aobandor@dane.gov.co' # aobandor
             
    elif temaRes == 9 or temaRes == 29 or temaRes == 30:
        
        responsable_oe = 'dvlizarazog@dane.gov.co' # dvlizarazog
          
    elif temaRes == 10 or temaRes == 6 or temaRes == 1:
        
        responsable_oe = 'gavargasr@dane.gov.co' # frherreran se cambia por gavargasr
              
    #elif temaRes == 11 or temaRes == 12 or temaRes == 25 or temaRes == 3: 
        
    #    responsable_oe = 'lhsanchezz@dane.gov.co' # lhsanchezz

    elif temaRes == 11 or temaRes == 12 or temaRes == 25: 
        
        responsable_oe = 'pczambranog@dane.gov.co' # pczambranog
           
    elif temaRes == 13 or temaRes == 24 or temaRes == 26 or temaRes == 27: 
        
        responsable_oe = 'mjpradag@dane.gov.co' # mppulidor se cambia por mjpradag


    ### Nota importante:  se pone responsable a ruth mientras ingresan los contratistas lhsanchezz
    elif temaRes == 11 or temaRes == 12 or temaRes == 25 or temaRes == 3: 
        responsable_oe = 'rctrianaa@dane.gov.co' # rctrianaa
    ### Nota importante: se pone responsable a ruth mientras ingresan los contratistas lhsanchezz
        
    elif temaRes == 4 or temaRes == 8 or temaRes == 20: 

        responsable_oe = 'eeguayazans@dane.gov.co' # eeguayazans Edgar Eduardo Guayazan Sierra
        
    elif temaRes == 21 or temaRes == 22 or temaRes == 23:  
        
        responsable_oe = 'mlbarretob@dane.gov.co' # mlbarretob
            
    recipients = [ responsable_oe]
      
    body = 'SISTEMA DE IDENTIFICACIÓN Y CARACTERIZACIÓN DE OFERTA Y DEMANDA ESTADÍSTICA DEL SEN\n\n' \
            'La operación estadística: '+ str(oe) + \
			    '\n\nha sido editada en los siguientes campos:\n' \
                    '\n' + str(fieldsEdited).replace('[','').replace(']','').replace(',', '\n').replace("'", "") + \
                        '\n' + str(fieldFormsetooee1).replace('[','').replace(']','').replace(',', '\n').replace("'", "") + \
                            '\n' + str(fieldFormsetooee2).replace('[','').replace(']','').replace(',', '\n').replace("'", "") + \
                                '\n' + str(fieldFormsetooee3).replace('[','').replace(']','').replace(',', '\n').replace("'", "")
                                                                                      
    return subject, recipients, body


@login_required
def ooee_edit(request, pk):
    entities = Entidades_oe.objects.filter(estado_id=1)
    count_entities = entities.count() 
##___***id para instancia de editar los formularios
    post_oe = get_object_or_404(OperacionEstadistica, pk=pk)
    oe_id =  OperacionEstadistica.objects.get(pk=pk)
    print("______", oe_id.codigo_oe)
    code_oe = oe_id.codigo_oe
    
    user = request.user
    entidad_creadora = user.profile.entidad.pk
    
    roleUser = user.profile.role.id
    #print("el rol es", roleUser)
    #print("estado del formulario", oe_id.nombre_est_id)

    entidad_responsable = oe_id.entidad.pk
    sistema_estado = oe_id.nombre_est.pk
    ent_resp2 = oe_id.entidad_resp2
    ent_resp3 = oe_id.entidad_resp3

    ## traer listado de usuarios responsable de la ooee
    user_email = []
    user_responsable = Profile.objects.filter(entidad=oe_id.entidad.pk)
    #print("responsable",user_responsable)
    for us in list(chain(user_responsable)):
        user_email.append(str(us.user.email))
    #print("array de usuarios",user_email)

    ## traer listado de usuarios responsable de la ooee
    
    totalIndicators = MC_ResultadoEstadistico.objects.filter(ooee_id=oe_id.pk) #pregunta 15
    count_indicators =  totalIndicators.count()

    #Modulo B ***** 
    normaoe_id = MB_Norma.objects.get(ooee_id=oe_id.pk)
    requerimientooe_id = MB_Requerimientos.objects.get(ooee_id=oe_id.pk) 
    prinUsuariosoe_id = MB_PrinUsuarios.objects.get(ooee_id=oe_id.pk)
    #Modulo C *****  
    unidadObsoe_id = MC_UnidadObservacion.objects.get(ooee_id=oe_id.pk)
    obtDatooe_id = MC_ObtencionDato.objects.get(ooee_id=oe_id.pk)
    muesProboe_id = MC_MuestreoProbabilistico.objects.get(ooee_id=oe_id.pk)
    muesNoProboe_id = MC_MuestreoNoProbabilistico.objects.get(ooee_id=oe_id.pk)
    tipoMarcooe_id = MC_TipoMarco.objects.get(ooee_id=oe_id.pk)
    docDesoe_id = MC_DocsDesarrollo.objects.get(ooee_id=oe_id.pk)
    concepEstanoe_id = MC_ConceptosEstandarizados.objects.get(ooee_id=oe_id.pk)
    nomClasoe_id = MC_Clasificaciones.objects.get(ooee_id=oe_id.pk)
    cobGeooe_id = MC_CoberturaGeografica.objects.get(ooee_id=oe_id.pk)
    desInfooe_id = MC_DesagregacionInformacion.objects.get(ooee_id=oe_id.pk)
    fuenteFinoe_id = MC_FuenteFinanciacion.objects.get(ooee_id=oe_id.pk)
   
    #Modulo D *****  
    medDatoe_id = MD_MedioDatos.objects.get(ooee_id=oe_id.pk)
    periodOe_id = MD_PeriodicidadOe.objects.get(ooee_id=oe_id.pk)
    herrProcoe_id = MD_HerramProcesamiento.objects.get(ooee_id=oe_id.pk)
    #Modulo E *****  
    anaResuloe_id = ME_AnalisisResultados.objects.get(ooee_id=oe_id.pk)
    #Modulo F *****  
    me_difoe_id =  MF_MediosDifusion.objects.get(ooee_id=oe_id.pk)
    fecPublioe_id = MF_FechaPublicacion.objects.get(ooee_id=oe_id.pk)
    freDifoe_id = MF_FrecuenciaDifusion.objects.get(ooee_id=oe_id.pk)
    prod_difoe_id = MF_ProductosDifundir.objects.get(ooee_id=oe_id.pk)
    otroProdoe_id = MF_OtrosProductos.objects.get(ooee_id=oe_id.pk)
    resultSimioe_id = MF_ResultadosSimilares.objects.get(ooee_id=oe_id.pk)
    hpsistInfooe_id = MF_HPSistemaInfo.objects.get(ooee_id=oe_id.pk)

    ##___***end id para instancia de editar los formularios
    editoe_form = EditOEForm(instance=oe_id)
    #Modulo B *****
    editnorma_form = EditMB_NormaForm(instance=normaoe_id) # form comentarios norma
    editrequerimiento_form = EditMB_RequerimientoForm(instance=requerimientooe_id) # form comentarios requerimiento
    editprinUsuarios_form = EditMB_PrinUsuariosForm(instance=prinUsuariosoe_id) # form comentarios usuarios principales
    #Modulo C *****  
    editunidadObs_form = EditMC_UnidadObservacionForm(instance=unidadObsoe_id) # form comentarios unidad de observacion
    editobtDato_form = EditMC_ObtencionDatoForm(instance=obtDatooe_id) # form comentarios unidad de obtención de datos
    editmuesProb_form = EditMC_MuestreoProbabilisticoForm(instance=muesProboe_id) # form comentarios muestreo probabibilistico
    editmuesNoProb_form = EditMC_MuestreoNoProbabilisticoForm(instance=muesNoProboe_id) # form comentarios muestreo No probabibilistico
    edittipoMarco_Form = EditMC_TipoMarcoForm(instance=tipoMarcooe_id) # form comentarios tipo de marco
    editdocDes_form = EditMC_DocsDesarrolloForm(instance=docDesoe_id) # form comentarios documentos que se elaboran para el desarrollo de la oe
    editconcepEstan_form = EditMC_ConceptosEstandarizadosForm(instance=concepEstanoe_id) # form comentarios oe utiliza conceptos estandaraizados
    editnomClas_Form = EditMC_ClasificacionesForm(instance=nomClasoe_id) # form comentarios oe utiliza nomenclaturas y/o clasificaciones
    editcobGeo_form = EditMC_CoberturaGeograficaForm(instance=cobGeooe_id) # form comentarios oe cobertura geografica
    editdesInfo_form = EditMC_DesagregacionInformacionForm(instance=desInfooe_id) # form desagregacion oe
    editfuenteFin_form = EditMC_FuenteFinanciacionForm(instance=fuenteFinoe_id) # form fuentes de financiación
    #Modulo D *****  
    editmedDat_form = EditMD_MedioDatosForm(instance=medDatoe_id) 
    editperiodOe_form = EditMD_PeriodicidadOeForm(instance=periodOe_id) 
    editherrProc_form = EditMD_HerramProcesamientoForm(instance=herrProcoe_id)
    #Modulo E ***** 
    editanaResul_form = EditME_AnalisisResultadosForm(instance=anaResuloe_id)
    #Modulo F
    editme_dif_form = EditMF_MediosDifusionForm(instance=me_difoe_id)
    editfecPubli_form = EditMF_FechaPublicacionForm(instance=fecPublioe_id)
    editfreDif_form = EditMF_FrecuenciaDifusionForm(instance=freDifoe_id)
    editprod_dif_form = EditMF_ProductosDifundirForm(instance=prod_difoe_id)
    editotroProd_form = EditMF_OtrosProductosForm(instance=otroProdoe_id)
    editresultSimi_form = EditMF_ResultadosSimilaresForm(instance=resultSimioe_id)
    edithpsistInfo_form = EditMF_HPSistemaInfoForm(instance=hpsistInfooe_id)
    
    ## comentarios
    comment_form = CommentForm()
    comments = post_oe.comments.filter(active=True)
    new_comment = None

    ## Novedades
    novedad_form = NovedadForm()
    novedades = post_oe.novedadactualizacions.filter(active=True)
    new_novedad = None

    ##  mostrar el ultimo estado y fecha de actualización
    if post_oe.novedadactualizacions.filter(active=True).exists():
        for noved in novedades:
            noved.est_actualiz
        if noved.est_actualiz != None:
            estadoNovedadOE = noved.est_actualiz
            fechaNovedadOE = noved.fecha_actualiz
        else:
            estadoNovedadOE = ""
            fechaNovedadOE = ""
    else:
        estadoNovedadOE = ""
        fechaNovedadOE = ""

    ##  end mostrar el ultimo estado y fecha de actualización
    
    ## Critica
    critica_form = CriticaForm()
    criticas = post_oe.criticas.filter(active=True)
    new_critica = None

    args = {}
    
    if request.method == 'GET':
        editformEntiFasset  = MB_EntidadFasesFormset(queryset=MB_EntidadFases.objects.filter(ooee_id=oe_id.pk), prefix='entidadfase')
        editformListset = MC_listaVariableFormset(queryset=MC_listaVariable.objects.filter(ooee_id=oe_id.pk), prefix='listas')
        editformResultset = MC_resultadoEstadisticoFormset(queryset=MC_ResultadoEstadistico.objects.filter(ooee_id=oe_id.pk), prefix='resultados')
    
    if request.method == "POST":
        editoe_form = EditOEForm(request.POST, request.FILES, instance=oe_id)
        #editEntidadFase_form = EditEntidadFaseForm(request.POST, instance=oe_id) 
    #Modulo B *****
        editformEntiFasset  = MB_EntidadFasesFormset(request.POST, prefix='entidadfase')
        editnorma_form = EditMB_NormaForm(request.POST, instance=normaoe_id) 
        editrequerimiento_form = EditMB_RequerimientoForm(request.POST, instance=requerimientooe_id) 
        editprinUsuarios_form = EditMB_PrinUsuariosForm(request.POST, instance=prinUsuariosoe_id) 
    #Modulo C ***** 
        editunidadObs_form = EditMC_UnidadObservacionForm(request.POST, instance=unidadObsoe_id) 
        editobtDato_form = EditMC_ObtencionDatoForm(request.POST, instance=obtDatooe_id) 
        editmuesProb_form = EditMC_MuestreoProbabilisticoForm(request.POST, instance=muesProboe_id) 
        editmuesNoProb_form = EditMC_MuestreoNoProbabilisticoForm(request.POST, instance=muesNoProboe_id)
        edittipoMarco_Form = EditMC_TipoMarcoForm(request.POST, instance=tipoMarcooe_id) 
        editdocDes_form = EditMC_DocsDesarrolloForm(request.POST, instance=docDesoe_id) 
        editconcepEstan_form = EditMC_ConceptosEstandarizadosForm(request.POST, instance=concepEstanoe_id) 
        editnomClas_Form = EditMC_ClasificacionesForm(request.POST,instance=nomClasoe_id) 
        editcobGeo_form = EditMC_CoberturaGeograficaForm(request.POST, instance=cobGeooe_id) 
        editdesInfo_form = EditMC_DesagregacionInformacionForm(request.POST, instance=desInfooe_id) 
        editfuenteFin_form = EditMC_FuenteFinanciacionForm(request.POST, instance=fuenteFinoe_id) 
        #***formsets 
        editformListset = MC_listaVariableFormset(request.POST,  prefix='listas') #PREGUNTA 14 *** 
        editformResultset = MC_resultadoEstadisticoFormset(request.POST, prefix='resultados') #PREGUNTA 15 ***
   
       
    #Modulo D *****  
        editmedDat_form = EditMD_MedioDatosForm(request.POST, instance=medDatoe_id) 
        editperiodOe_form = EditMD_PeriodicidadOeForm(request.POST,instance=periodOe_id)
        editherrProc_form = EditMD_HerramProcesamientoForm(request.POST, instance=herrProcoe_id)
    #Modulo E ***** 
        editanaResul_form = EditME_AnalisisResultadosForm(request.POST, instance=anaResuloe_id)
    #Modulo F *****
        editme_dif_form = EditMF_MediosDifusionForm(request.POST, instance=me_difoe_id)
        editfecPubli_form = EditMF_FechaPublicacionForm(request.POST, instance=fecPublioe_id)
        editfreDif_form = EditMF_FrecuenciaDifusionForm(request.POST, instance=freDifoe_id)
        editprod_dif_form = EditMF_ProductosDifundirForm(request.POST,instance=prod_difoe_id)
        editotroProd_form = EditMF_OtrosProductosForm(request.POST, instance=otroProdoe_id)
        editresultSimi_form = EditMF_ResultadosSimilaresForm(request.POST, instance=resultSimioe_id)
        edithpsistInfo_form = EditMF_HPSistemaInfoForm(request.POST, instance=hpsistInfooe_id)
    ##Comentarios ***    
        comment_form = CommentForm(data=request.POST)

    ##Novedades
        novedad_form = NovedadForm(data=request.POST)

    ##Critica
        critica_form = CriticaForm(data=request.POST)

        #print("no guardo",editoe_form )

        ## detectar cambios
        fieldsEdited = []
        fieldFormsetooee1 = []
        fieldFormsetooee2 = []
        fieldFormsetooee3 = []

        if editoe_form.has_changed() or editformEntiFasset.has_changed() or editformListset.has_changed() or editformResultset.has_changed():
            #print("The following fields changed: %s" % ", ".join(editoe_form.changed_data))
            #print("contador de cambios", editoe_form.changed_data)
            
            for index, item in enumerate(editoe_form.changed_data):

                editoe_form.fields[item].widget.attrs['title']
                questionParamTitle = editoe_form.fields[item].widget.attrs['title']
                fieldsEdited.append(questionParamTitle)# array que almacena lista de campos editados

            if editformEntiFasset.has_changed() == True:
                fieldFormsetooee1 = ["Modulo B Pregunta 3: Indique si hay otra(s) entidad(es) que participa(n) en alguna de las fases del proceso estadístico"]
            if editformEntiFasset.has_changed() == True:
                fieldFormsetooee2 = ["Modulo C Pregunta 14: Liste todas las variables que maneja la operación estadística"]
            if editformEntiFasset.has_changed() == True:
                fieldFormsetooee3 = ["Modulo C Pregunta 15: Relacione todos los resultados estadísticos (indicadores o resultados agregados) que produce la operación estadística"]

        if editoe_form.is_valid() and editformEntiFasset.is_valid() and editnorma_form.is_valid() and editrequerimiento_form.is_valid() and editprinUsuarios_form.is_valid()  and \
                editunidadObs_form.is_valid() and editobtDato_form.is_valid() and editmuesProb_form.is_valid() and editmuesNoProb_form.is_valid() and \
                    edittipoMarco_Form.is_valid() and editdocDes_form.is_valid() and editconcepEstan_form.is_valid() and editnomClas_Form.is_valid() and \
                         editcobGeo_form.is_valid() and editdesInfo_form.is_valid and editfuenteFin_form.is_valid() and editmedDat_form.is_valid() and editperiodOe_form.is_valid() and \
                             editherrProc_form.is_valid() and editanaResul_form.is_valid() and editme_dif_form.is_valid() and editfecPubli_form.is_valid() and \
                                 editfreDif_form.is_valid() and editprod_dif_form.is_valid() and editotroProd_form.is_valid() and editresultSimi_form.is_valid() and edithpsistInfo_form.is_valid() and \
                                      editformListset.is_valid() and editformResultset.is_valid() and comment_form.is_valid() and novedad_form.is_valid() and critica_form.is_valid():            
            
            editOOEE_form = editoe_form.save(commit=False)
            #editOOEE_form.codigo_oe = "OE" + str(oe_id.pk)

            ## validaciones según el rol
            ## Si rol fuente --> edita ooee publicada  ***
            if roleUser == 3 and editOOEE_form.nombre_est_id == 5:
                editOOEE_form.nombre_est_id = 2

            elif roleUser == 3 and editOOEE_form.nombre_est_id == 4:
                editOOEE_form.nombre_est_id = 2

            ## Si rol tematico --> edita ooee publicada no descomentar
            #elif roleUser == 4 and editOOEE_form.nombre_est_id == 5:
            #    editOOEE_form.nombre_est_id = 2

            ## end validaciones según el rol
            
            temaRes = editOOEE_form.tema_id
            

            #envio de correo para notificar cuando el estado cambia a devuelto ***
        
            if editOOEE_form.nombre_est_id == 3 and editOOEE_form.entidad_id != 1:
                subject, recipients, body = emailNotificationStateView(user, post_oe, user_email)
                sendEmailStatusChange(subject, recipients, body)
            #end envio de correo para notificar cuando el estado cambia a devuelto
            

            if 'variable_file' in request.FILES:
                editOOEE_form.variable_file = request.FILES['variable_file']

            if 'anexos' in request.FILES:
                editOOEE_form.anexos = request.FILES['anexos']

            ## send correo Notificaciones para tematico al editar formulario **
            subject, recipients, body = createEmailEditOOEE(user, post_oe, temaRes, fieldsEdited, fieldFormsetooee1, fieldFormsetooee2, fieldFormsetooee3)
            sendEmailNotification(subject, recipients, body)
            ## end send correo Notificaciones


            editOOEE_form.save()
            obj_id = editOOEE_form.id
            
            #print("id del formulario a actualizar",obj_id)
            
        #Modulo B ***** 
            formEntiFase = editformEntiFasset.save(commit=False)
            
            if editformEntiFasset.deleted_forms:
                for obj in editformEntiFasset.deleted_objects:
                    obj.delete() 
            else:
                for formEntiFase in editformEntiFasset:
                # so that `ooee` instance can be attached.
                    formEntidadFase = formEntiFase.save(commit=False)
                    formEntidadFase.ooee_id = obj_id
                    formEntidadFase.editOOEE_form = editOOEE_form
                    if formEntidadFase.nombre_entifas == "":
                        formEntidadFase = formEntiFase.save(commit=False)
                    else:
                        # Save the entitie to the database
                        formEntidadFase.save() 
                        formEntiFase.save_m2m()  

            editNormaoe_form =  editnorma_form.save(commit=False)
            editNormaoe_form.save()

            editReqoe_form =  editrequerimiento_form.save(commit=False)
            editReqoe_form.save()

            editprinUsuaoe_form = editprinUsuarios_form.save(commit=False)
            editprinUsuaoe_form.save()
        #Modulo C ***** 
            edituniObsoe_form  = editunidadObs_form.save(commit=False)
            edituniObsoe_form.save()

            editobtDatooe_form = editobtDato_form.save(commit=False)
            editobtDatooe_form.save()

            editmuesProboe_form = editmuesProb_form.save(commit=False)
            editmuesProboe_form.save()

            editmuesNoProboe_form = editmuesNoProb_form.save(commit=False)
            editmuesNoProboe_form.save()

            edittipoMarcooe_Form = edittipoMarco_Form.save(commit=False)
            edittipoMarcooe_Form.save()

            editdocDesoe_form = editdocDes_form.save(commit=False)
            editdocDesoe_form.save()

            editconcepEstanoe_form = editconcepEstan_form.save(commit=False)
            editconcepEstanoe_form.save()
            
            editnomClasoe_Form = editnomClas_Form.save(commit=False)
            editnomClasoe_Form.save()

            editcobGeooe_form = editcobGeo_form.save(commit=False)
            editcobGeooe_form.save()

            editdesInfooe_form = editdesInfo_form.save(commit=False)
            editdesInfooe_form.save()

            editfuenteFinoe_form = editfuenteFin_form.save(commit=False)
            editfuenteFinoe_form.save()
     
            #*** formsets
          
            formLista = editformListset.save(commit=False)
           
            if editformListset.deleted_forms:
                for obj in editformListset.deleted_objects:
                    obj.delete() 
            else:
                for formLista in editformListset:
                # so that `ooee` instance can be attached.
                    formListaVariable = formLista.save(commit=False)
                    formListaVariable.ooee_id = obj_id
                    formListaVariable.editOOEE_form = editOOEE_form
                    if formListaVariable.lista_var == "":
                        formListaVariable = formLista.save(commit=False)
                    else:
                        # Save listvar to the database
                        formListaVariable.save() 

            formResult = editformResultset.save(commit=False)
            if editformResultset.deleted_forms:
                for obj in editformResultset.deleted_objects:
                    obj.delete()
            else:
                for formResult in editformResultset:
                # so that `ooee` instance can be attached.
                    formResultEstadis = formResult.save(commit=False)
                    formResultEstadis.ooee_id = obj_id
                    formResultEstadis.editOOEE_form = editOOEE_form
                    if formResultEstadis.resultEstad == "":
                        formResultEstadis = formResult.save(commit=False)
                    else:
                        # Save resultados o indicadores agregados to the database
                        formResultEstadis.save() 
          
        #Modulo D ***** 
            editmedDatoe_form = editmedDat_form.save(commit=False)
            editmedDatoe_form.save()

            editperiodoooe_form = editperiodOe_form.save(commit=False)
            editperiodoooe_form.save()

            editherrProcoe_form = editherrProc_form.save(commit=False)
            editherrProcoe_form.save()
        #Modulo E ***** 
            editanaResuloe_form = editanaResul_form.save(commit=False)
            editanaResuloe_form.save()
        #Modulo F *****
            editme_difoe_form  = editme_dif_form.save(commit=False)
            editme_difoe_form.save()

            editfecPublioe_form = editfecPubli_form.save(commit=False)
            editfecPublioe_form.save()

            editfreDifoe_form = editfreDif_form.save(commit=False)
            editfreDifoe_form.save()

            editprod_difoe_form = editprod_dif_form.save(commit=False)
            editprod_difoe_form.save()

            editotroProdoe_form = editotroProd_form.save(commit=False)
            editotroProdoe_form.save()

            editresultSimioe_form = editresultSimi_form.save(commit=False)
            editresultSimioe_form.save()

            edithpsistInfooe_form = edithpsistInfo_form.save(commit=False)
            edithpsistInfooe_form.save()

        ## comentarios
            new_comment = comment_form.save(commit=False)
            new_comment.name = request.user
            new_comment.post_oe = post_oe
            if new_comment.body == "":
                new_comment = comment_form.save(commit=False)
            else:
            # Save the comment to the database
                new_comment.save()

        ##  novedades
            new_novedad = novedad_form.save(commit=False)
            new_novedad.name_nov = request.user
            new_novedad.post_oe = post_oe
            if new_novedad.descrip_novedad == "":
                new_novedad = novedad_form.save(commit=False)
            else:
                new_novedad.save()

        ## Criticas
            new_critica = critica_form.save(commit=False)
            new_critica.name_cri = request.user
            new_critica.post_oe = post_oe
            if new_critica.descrip_critica == "":
                new_critica = critica_form.save(commit=False)
            else:
                new_critica.save()

            editoe_form.save_m2m() ## metodo para guardar relaciones manyTomany

            messages.success(request, 'Las respuestas se han guardado con éxito')
            
            #return redirect('ooee:all_ooee')
            return HttpResponseRedirect(request.path_info)
            
        else:
            
            messages.error(request, "Las respuestas NO se han guardado verifique si tiene errores")
            

    return render(request, 'ooee/edit_ooee.html', {'editoe_form': editoe_form,  'editformEntiFasset':  editformEntiFasset, 'editnorma_form': editnorma_form,
    'editrequerimiento_form': editrequerimiento_form, 'editprinUsuarios_form': editprinUsuarios_form, 'editunidadObs_form': editunidadObs_form, 'editobtDato_form': editobtDato_form,
    'editmuesProb_form':  editmuesProb_form, 'editmuesNoProb_form': editmuesNoProb_form, 'edittipoMarco_Form': edittipoMarco_Form, 'editdocDes_form': editdocDes_form, 
    'editconcepEstan_form': editconcepEstan_form, 'editnomClas_Form': editnomClas_Form,  'editcobGeo_form':  editcobGeo_form,  'editdesInfo_form' : editdesInfo_form, 
    'editfuenteFin_form' : editfuenteFin_form, 'editformListset': editformListset, 'editformResultset': editformResultset, 'editmedDat_form': editmedDat_form, 'editperiodOe_form': editperiodOe_form, 'editherrProc_form': editherrProc_form,
    'editanaResul_form': editanaResul_form, 'editme_dif_form': editme_dif_form, 'editfecPubli_form': editfecPubli_form, 'editfreDif_form': editfreDif_form, 'editprod_dif_form': editprod_dif_form,
    'editotroProd_form' : editotroProd_form, 'editresultSimi_form': editresultSimi_form, 'edithpsistInfo_form': edithpsistInfo_form, 'count_entities': count_entities,
    'comments': comments, 'new_comment': new_comment, 'comment_form': comment_form, 'novedades': novedades , 'new_novedad': new_novedad, 'novedad_form': novedad_form, 'criticas': criticas, 'new_critica': new_critica, 'critica_form': critica_form,
    'count_indicators': count_indicators, 'entidad_responsable': entidad_responsable, 'sistema_estado': sistema_estado, 
    'estadoNovedadOE': estadoNovedadOE, 'fechaNovedadOE': fechaNovedadOE, 'entidad_creadora': entidad_creadora, "code_oe": code_oe})


#### pruebas de servicios
import json
def send_json(request):

    ooee = OperacionEstadistica.objects.filter(nombre_est=5).only('nombre_oe', 'entidad', 'correo_resp', 'nombre_resp')

    response = serializers.serialize("json", ooee, fields=('nombre_oe', 'entidad', 'correo_resp', 'nombre_resp'), use_natural_foreign_keys=True, use_natural_primary_keys=True)
    return HttpResponse(response, content_type='application/json')
    
    #return JsonResponse(serializers.serialize('json', ooee), safe=False)
    #return HttpResponse(serializers.serialize('json', ooee), content_type="application/json")



"""def send_json(request):
    qr = OperacionEstadistica.objects.filter(nombre_est=5).only('nombre_oe', 'entidad')
    qr_json = json.dumps(list(qr), ensure_ascii=False, default=str)
    return JsonResponse(qr_json, safe=False)"""


#### end pruebas de servicios

### Modulo de evaluación de calidad
@login_required
def create_eval(request, pk):
    entities = Entidades_oe.objects.filter(estado_id=1)
    count_entities = entities.count() 
    post_oe = get_object_or_404(OperacionEstadistica, pk=pk)
    #print("craaaaa", post_oe)
    evaluacionCalidad_form = EvaluacionCalidadForm()
    evaluaciones = post_oe.evaluacionescalidad.filter(active=True)
    new_evaluacion = None
    registered = False

    if request.method == "POST":

        evaluacionCalidad_form = EvaluacionCalidadForm(data=request.POST)

        if evaluacionCalidad_form.is_valid():

            new_evaluacion = evaluacionCalidad_form.save(commit=False)
            new_evaluacion.name_evaluador =  request.user
            new_evaluacion.post_oe = post_oe
            if new_evaluacion.year_eva == None:
                new_evaluacion = evaluacionCalidad_form.save(commit=False)
            else:
                new_evaluacion.save()
                registered = True
                
        else:
            
            evaluacionCalidad_form = EvaluacionCalidadForm()
            messages.error(request, "Las respuestas NO se han guardado verifique si tiene errores")

    return render(request, 'eval_ooee/createEval_ooee.html', {'count_entities': count_entities, 'post_oe' : post_oe, 'evaluaciones': evaluaciones,
    'new_evaluacion': new_evaluacion, 'evaluacionCalidad_form': evaluacionCalidad_form, 'registered': registered })



@login_required
def edit_eval(request, pk):
    entities = Entidades_oe.objects.filter(estado_id=1)
    count_entities = entities.count()
    eval_id = EvaluacionCalidad.objects.get(pk=pk)
    #print("id de evaluación", eval_id)
    #post_oe = get_object_or_404(EvaluacionCalidad, pk=pk)
    post_oe = get_object_or_404(OperacionEstadistica, pk=pk)
    #print("id que necesito", post_oe, "pk", pk)
    editEvaluacionCalidad_form = EditEvaluacionCalidadForm(instance=eval_id)
    #editEvaluaciones = post_oe.evaluacionescalidad.filter(active=True)

    #print("dont save", editEvaluacionCalidad_form)
    if request.method == "POST":

        editEvaluacionCalidad_form = EditEvaluacionCalidadForm(request.POST, instance=eval_id)

        if editEvaluacionCalidad_form.has_changed():
            #print("The following fields changed: %s" % ", ".join(editEvaluacionCalidad_form.changed_data))
            #print("contador de cambios", editEvaluacionCalidad_form.changed_data )
            cambios = editEvaluacionCalidad_form.changed_data
            

        if editEvaluacionCalidad_form.is_valid():

            new_evaluacion = editEvaluacionCalidad_form.save(commit=False)
            new_evaluacion.fecha_eval_sis = timezone.now()
            new_evaluacion.save()

            messages.success(request, 'los cambios se han guardado con éxito')
            return HttpResponseRedirect(request.path_info)
            
        else:
            messages.error(request, "Las respuestas NO se han guardado verifique si tiene errores")

    return render(request, 'eval_ooee/editEval_ooee.html', {'count_entities': count_entities,
    'editEvaluacionCalidad_form': editEvaluacionCalidad_form, 'post_oe': post_oe })

########*********** Reporte completo de operaciones estadisticas ##############**************

class reportOOEEfull_xls(TemplateView):
    def get(self, request, *args, **kwargs):
             
        ooees = OperacionEstadistica.objects.all()

        normaTextList = list(MB_Norma.objects.all())
        entidadesFasesTextList = list(MB_EntidadFases.objects.all())
        requerimientoTextList = list(MB_Requerimientos.objects.all())
        usuariosPrinTextList = list(MB_PrinUsuarios.objects.all())
        unidadObserTextList = list(MC_UnidadObservacion.objects.all())
        obtencionDatoTextList = list(MC_ObtencionDato.objects.all())
        muestProbaTextList = list(MC_MuestreoProbabilistico.objects.all())
        muestNoProbaTextList = list(MC_MuestreoNoProbabilistico.objects.all())
        tipoMarcoTextList = list(MC_TipoMarco.objects.all())
        docsDesarrolloTextList = list(MC_DocsDesarrollo.objects.all())
        conceptosEstandaTextList = list(MC_ConceptosEstandarizados.objects.all())
        clasificacionesTextList = list(MC_Clasificaciones.objects.all())
        coberGeogTextList = list(MC_CoberturaGeografica.objects.all())
        desagreInfoTextList = list(MC_DesagregacionInformacion.objects.all())
        fuenteFinanTextList = list(MC_FuenteFinanciacion.objects.all())
            
        listaDeVariablesText = list(MC_listaVariable.objects.all()) # Modulo c pregunta 14
        listaResultEstText = list(MC_ResultadoEstadistico.objects.all()) # Modulo c pregunta 15

        medioDatosTextList = list(MD_MedioDatos.objects.all())
        periodicidadTextList = list(MD_PeriodicidadOe.objects.all())
        herraProcesamiTextList = list(MD_HerramProcesamiento.objects.all())
        analResultTextList = list(ME_AnalisisResultados.objects.all())
        medioDifusTextList = list(MF_MediosDifusion.objects.all())
        fechaPublicTextList = list(MF_FechaPublicacion.objects.all())
        frecuenciaDifusionTextList = list(MF_FrecuenciaDifusion.objects.all())
        productosDifundirTextList = list(MF_ProductosDifundir.objects.all())
        otrosProductosTextList = list(MF_OtrosProductos.objects.all())
        resultadosSimiTextList = list(MF_ResultadosSimilares.objects.all())
        hpSistemaInfoTextList = list(MF_HPSistemaInfo.objects.all())

        listaEvaluacionText = list(EvaluacionCalidad.objects.all()) #Evaluación de calidad

        listaCriticaText = list(Critica.objects.all()) #critica de la oe
        listaNovedadText = list(NovedadActualizacion.objects.all()) #novedad  de la oe

        wb = Workbook()
        ws = wb.active
        ws.title = "Directorio OOEE"
        
        sheet2 = wb.create_sheet('Entidades Fases')
        sheet3 = wb.create_sheet('Variables')
        sheet4 = wb.create_sheet('Resultados')
        sheet5 = wb.create_sheet('Eval Calidad')
        sheet6 = wb.create_sheet('Critica')
        sheet7 = wb.create_sheet('Novedad')

        def set_border(ws, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = ws[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(ws,'A1:IG'+str(ooees.count()+3))

		## size rows

        ws.row_dimensions[1].height = 55
        ws.row_dimensions[2].height = 40
        ws.row_dimensions[3].height = 80
       
		## size column
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['I'].width = 25
        ws.column_dimensions['J'].width = 25
        ws.column_dimensions['K'].width = 25
        ws.column_dimensions['L'].width = 25
        ws.column_dimensions['M'].width = 25
        ws.column_dimensions['N'].width = 25
        ws.column_dimensions['O'].width = 25
        ws.column_dimensions['P'].width = 25
        ws.column_dimensions['Q'].width = 25
        ws.column_dimensions['R'].width = 22
        ws.column_dimensions['S'].width = 25
        ws.column_dimensions['T'].width = 25
        ws.column_dimensions['U'].width = 25
        ws.column_dimensions['V'].width = 22
        ws.column_dimensions['W'].width = 25
        ws.column_dimensions['X'].width = 25
        ws.column_dimensions['Y'].width = 25
        ws.column_dimensions['Z'].width = 25

        ws.column_dimensions['AA'].width = 25
        ws.column_dimensions['AB'].width = 25
        ws.column_dimensions['AC'].width = 25
        ws.column_dimensions['AD'].width = 25
        ws.column_dimensions['AE'].width = 25
        ws.column_dimensions['AF'].width = 25
        ws.column_dimensions['AG'].width = 25
        ws.column_dimensions['AH'].width = 25
        ws.column_dimensions['AI'].width = 25
        ws.column_dimensions['AJ'].width = 25
        ws.column_dimensions['AK'].width = 25
        ws.column_dimensions['AL'].width = 25
        ws.column_dimensions['AM'].width = 25
        ws.column_dimensions['AN'].width = 25
        ws.column_dimensions['AO'].width = 25
        ws.column_dimensions['AP'].width = 25
        ws.column_dimensions['AQ'].width = 25
        ws.column_dimensions['AR'].width = 25
        ws.column_dimensions['AS'].width = 25
        ws.column_dimensions['AT'].width = 25
        ws.column_dimensions['AU'].width = 25
        ws.column_dimensions['AV'].width = 25
        ws.column_dimensions['AW'].width = 25
        ws.column_dimensions['AX'].width = 25
        ws.column_dimensions['AY'].width = 25
        ws.column_dimensions['AZ'].width = 25

        ws.column_dimensions['BA'].width = 25
        ws.column_dimensions['BB'].width = 25
        ws.column_dimensions['BC'].width = 25
        ws.column_dimensions['BD'].width = 25
        ws.column_dimensions['BE'].width = 25
        ws.column_dimensions['BF'].width = 25
        ws.column_dimensions['BG'].width = 25
        ws.column_dimensions['BH'].width = 25
        ws.column_dimensions['BI'].width = 25
        ws.column_dimensions['BJ'].width = 25
        ws.column_dimensions['BK'].width = 25
        ws.column_dimensions['BL'].width = 25
        ws.column_dimensions['BM'].width = 25
        ws.column_dimensions['BN'].width = 25
        ws.column_dimensions['BO'].width = 25
        ws.column_dimensions['BP'].width = 25
        ws.column_dimensions['BQ'].width = 25
        ws.column_dimensions['BR'].width = 25
        ws.column_dimensions['BS'].width = 25
        ws.column_dimensions['BT'].width = 25
        ws.column_dimensions['BU'].width = 25
        ws.column_dimensions['BV'].width = 25
        ws.column_dimensions['BW'].width = 25
        ws.column_dimensions['BX'].width = 25
        ws.column_dimensions['BY'].width = 25
        ws.column_dimensions['BZ'].width = 25

        ws.column_dimensions['CA'].width = 25
        ws.column_dimensions['CB'].width = 25
        ws.column_dimensions['CC'].width = 25
        ws.column_dimensions['CD'].width = 25
        ws.column_dimensions['CE'].width = 25
        ws.column_dimensions['CF'].width = 25
        ws.column_dimensions['CG'].width = 25
        ws.column_dimensions['CH'].width = 25
        ws.column_dimensions['CI'].width = 25
        ws.column_dimensions['CJ'].width = 25
        ws.column_dimensions['CK'].width = 25
        ws.column_dimensions['CL'].width = 25
        ws.column_dimensions['CM'].width = 25
        ws.column_dimensions['CN'].width = 25
        ws.column_dimensions['CO'].width = 25
        ws.column_dimensions['CP'].width = 25
        ws.column_dimensions['CQ'].width = 25
        ws.column_dimensions['CR'].width = 25
        ws.column_dimensions['CS'].width = 25
        ws.column_dimensions['CT'].width = 25
        ws.column_dimensions['CU'].width = 25
        ws.column_dimensions['CV'].width = 25
        ws.column_dimensions['CW'].width = 25
        ws.column_dimensions['CX'].width = 25
        ws.column_dimensions['CY'].width = 25
        ws.column_dimensions['CZ'].width = 25

        ws.column_dimensions['DA'].width = 25
        ws.column_dimensions['DB'].width = 25
        ws.column_dimensions['DC'].width = 25
        ws.column_dimensions['DD'].width = 25
        ws.column_dimensions['DE'].width = 25
        ws.column_dimensions['DF'].width = 25
        ws.column_dimensions['DG'].width = 25
        ws.column_dimensions['DH'].width = 25
        ws.column_dimensions['DI'].width = 25
        ws.column_dimensions['DJ'].width = 25
        ws.column_dimensions['DK'].width = 25
        ws.column_dimensions['DL'].width = 25
        ws.column_dimensions['DM'].width = 25
        ws.column_dimensions['DN'].width = 25
        ws.column_dimensions['DO'].width = 25
        ws.column_dimensions['DP'].width = 25
        ws.column_dimensions['DQ'].width = 25
        ws.column_dimensions['DR'].width = 25
        ws.column_dimensions['DS'].width = 25
        ws.column_dimensions['DT'].width = 25
        ws.column_dimensions['DU'].width = 25
        ws.column_dimensions['DV'].width = 25
        ws.column_dimensions['DW'].width = 25
        ws.column_dimensions['DX'].width = 25
        ws.column_dimensions['DY'].width = 25
        ws.column_dimensions['DZ'].width = 25

        ws.column_dimensions['EA'].width = 25
        ws.column_dimensions['EB'].width = 25
        ws.column_dimensions['EC'].width = 25
        ws.column_dimensions['ED'].width = 25
        ws.column_dimensions['EE'].width = 25
        ws.column_dimensions['EF'].width = 25
        ws.column_dimensions['EG'].width = 25
        ws.column_dimensions['EH'].width = 25
        ws.column_dimensions['EI'].width = 25
        ws.column_dimensions['EJ'].width = 25
        ws.column_dimensions['EK'].width = 25
        ws.column_dimensions['EL'].width = 25
        ws.column_dimensions['EM'].width = 25
        ws.column_dimensions['EN'].width = 25
        ws.column_dimensions['EO'].width = 25
        ws.column_dimensions['EP'].width = 25
        ws.column_dimensions['EQ'].width = 25
        ws.column_dimensions['ER'].width = 25
        ws.column_dimensions['ES'].width = 25
        ws.column_dimensions['ET'].width = 25
        ws.column_dimensions['EU'].width = 25
        ws.column_dimensions['EV'].width = 25
        ws.column_dimensions['EW'].width = 25
        ws.column_dimensions['EX'].width = 25
        ws.column_dimensions['EY'].width = 25
        ws.column_dimensions['EZ'].width = 25

        ws.column_dimensions['FA'].width = 25
        ws.column_dimensions['FB'].width = 25
        ws.column_dimensions['FC'].width = 25
        ws.column_dimensions['FD'].width = 25
        ws.column_dimensions['FE'].width = 25
        ws.column_dimensions['FF'].width = 25
        ws.column_dimensions['FG'].width = 25
        ws.column_dimensions['FH'].width = 25
        ws.column_dimensions['FI'].width = 25
        ws.column_dimensions['FJ'].width = 25
        ws.column_dimensions['FK'].width = 25
        ws.column_dimensions['FL'].width = 25
        ws.column_dimensions['FM'].width = 25
        ws.column_dimensions['FN'].width = 25
        ws.column_dimensions['FO'].width = 25
        ws.column_dimensions['FP'].width = 25
        ws.column_dimensions['FQ'].width = 25
        ws.column_dimensions['FR'].width = 25
        ws.column_dimensions['FS'].width = 25
        ws.column_dimensions['FT'].width = 25
        ws.column_dimensions['FU'].width = 25
        ws.column_dimensions['FV'].width = 25
        ws.column_dimensions['FW'].width = 25
        ws.column_dimensions['FX'].width = 25
        ws.column_dimensions['FY'].width = 25
        ws.column_dimensions['FZ'].width = 25

        ws.column_dimensions['GA'].width = 25
        ws.column_dimensions['GB'].width = 25
        ws.column_dimensions['GC'].width = 25
        ws.column_dimensions['GD'].width = 25
        ws.column_dimensions['GE'].width = 25
        ws.column_dimensions['GF'].width = 25
        ws.column_dimensions['GG'].width = 25
        ws.column_dimensions['GH'].width = 25
        ws.column_dimensions['GI'].width = 25
        ws.column_dimensions['GJ'].width = 25
        ws.column_dimensions['GK'].width = 25
        ws.column_dimensions['GL'].width = 25
        ws.column_dimensions['GM'].width = 25
        ws.column_dimensions['GN'].width = 25
        ws.column_dimensions['GO'].width = 25
        ws.column_dimensions['GP'].width = 25
        ws.column_dimensions['GQ'].width = 25
        ws.column_dimensions['GR'].width = 25
        ws.column_dimensions['GS'].width = 25
        ws.column_dimensions['GT'].width = 25
        ws.column_dimensions['GU'].width = 25
        ws.column_dimensions['GV'].width = 25
        ws.column_dimensions['GW'].width = 25
        ws.column_dimensions['GX'].width = 25
        ws.column_dimensions['GY'].width = 25
        ws.column_dimensions['GZ'].width = 25

        ws.column_dimensions['HA'].width = 25
        ws.column_dimensions['HB'].width = 25
        ws.column_dimensions['HC'].width = 25
        ws.column_dimensions['HD'].width = 25
        ws.column_dimensions['HE'].width = 25
        ws.column_dimensions['HF'].width = 25
        ws.column_dimensions['HG'].width = 25
        ws.column_dimensions['HH'].width = 25
        ws.column_dimensions['HI'].width = 25
        ws.column_dimensions['HJ'].width = 25
        ws.column_dimensions['HK'].width = 25
        ws.column_dimensions['HL'].width = 25
        ws.column_dimensions['HM'].width = 25
        ws.column_dimensions['HN'].width = 25
        ws.column_dimensions['HO'].width = 25
        ws.column_dimensions['HP'].width = 25
        ws.column_dimensions['HQ'].width = 25
        ws.column_dimensions['HR'].width = 25
        ws.column_dimensions['HS'].width = 25
        ws.column_dimensions['HT'].width = 25
        ws.column_dimensions['HU'].width = 25
        ws.column_dimensions['HV'].width = 25
        ws.column_dimensions['HW'].width = 25
        ws.column_dimensions['HX'].width = 25
        ws.column_dimensions['HY'].width = 25
        ws.column_dimensions['HZ'].width = 25
        ws.column_dimensions['IA'].width = 25
        ws.column_dimensions['IB'].width = 25
        ws.column_dimensions['IC'].width = 25
        ws.column_dimensions['ID'].width = 25
        ws.column_dimensions['IE'].width = 25
        ws.column_dimensions['IF'].width = 25
        ws.column_dimensions['IG'].width = 25
        

		##insert image
        #img = openpyxl.drawing.image.Image('media/pictures/logoSEN.png')
        #img.anchor = 'A1'	
        #ws.add_image(img)

		##styles
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
		
        title_cell = ws['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de las Operaciones Estadísticas'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')

          
        ws.merge_cells('A1:IG1')
        ws.merge_cells('A2:C2')
        ws.merge_cells('D2:I2')
        ws.merge_cells('J2:L2')
        ws.merge_cells('M2:U2')
        ws.merge_cells('V2:V3')
        ws.merge_cells('W2:AA2')
        ws.merge_cells('AB2:AI2')
        ws.merge_cells('AJ2:AT2')
        ws.merge_cells('AU2:AU3')
        ws.merge_cells('AV2:BC2')
        ws.merge_cells('BD2:BH2')
        ws.merge_cells('BI2:BP2')
        ws.merge_cells('BQ2:BV2')
        ws.merge_cells('BW2:CB2')
        ws.merge_cells('CC2:CH2')
        ws.merge_cells('CI2:CP2')
        ws.merge_cells('CQ2:DB2')
        ws.merge_cells('DC2:DF2')
        ws.merge_cells('DG2:DV2')
        ws.merge_cells('DW2:EL2')
        ws.merge_cells('EM2:EO2')
        ws.merge_cells('EP2:EW2')
        ws.merge_cells('EX2:EY2')
        ws.merge_cells('EZ2:FD2')
        ws.merge_cells('FE2:FE3')
        ws.merge_cells('FF2:FF3')
        ws.merge_cells('FG2:FP2')
        ws.merge_cells('FQ2:FW2')
        ws.merge_cells('FX2:GF2')
        ws.merge_cells('GG2:GG3')
        ws.merge_cells('GH2:GL2')
        ws.merge_cells('GM2:GQ2')
        ws.merge_cells('GR2:GR3')
        ws.merge_cells('GS2:GT2')
        ws.merge_cells('GU2:GY2')
        ws.merge_cells('GZ2:HF2')
        ws.merge_cells('HG2:HM2')
        ws.merge_cells('HN2:HW2')


        ws.merge_cells('HX2:HZ2')
        ws.merge_cells('IA2:IB2')

        ws.merge_cells('IC2:IC3')
        ws.merge_cells('ID2:ID3')
        ws.merge_cells('IE2:IE3')
        ws.merge_cells('IF2:IF3')
        ws.merge_cells('IG2:IG3')
        
        #pin 1

        
        ## insert heads groups
        codigo_cell = ws['A2']
        codigo_cell.value = 'OPERACIÓN ESTADÍSTICA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['D2']
        codigo_cell.value = 'ENTIDAD O ENTIDADES RESPONSABLES'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['J2']
        codigo_cell.value = 'ÁREA TEMÁTICA / TEMA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['M2']
        codigo_cell.value = ' A. IDENTIFICACIÓN '
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['V2']
        codigo_cell.value = 'Indique si hay otra(s) entidad(es) que participa(n) en alguna de las fases del proceso estadístico  (ver hoja 2)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['W2']
        codigo_cell.value = 'Bajo cuál(es) de las siguiente(s) norma(s), se soporta la producción de información de la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AB2']
        codigo_cell.value = 'La operación estadística satisface requerimientos de información de:'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AJ2']
        codigo_cell.value = 'Señale cuáles son los principales usuarios  de la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AU2']
        codigo_cell.value = '¿Cuál es la población objetivo de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['AV2']
        codigo_cell.value = '¿Cuál es la unidad de observación de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BD2']
        codigo_cell.value = '¿Cuál es el tipo de operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['BI2']
        codigo_cell.value = 'Indique de donde se obtienen los datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['BQ2']
        codigo_cell.value = 'Muestreo probabilístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['BW2']
        codigo_cell.value = 'Muestreo No probabilístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['CC2']
        codigo_cell.value = ' ¿La operación estadística cuenta con un marco estadístico para identificar y ubicar las unidades de observación?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['CI2']
        codigo_cell.value = 'Indique cuáles de los siguientes documentos se elaboran para el desarrollo de la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['CQ2']
        codigo_cell.value = 'Indique si la operación estadística utiliza conceptos estandarizados de'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['DC2']
        codigo_cell.value = '¿La operación estadística utiliza nomenclaturas y/o clasificaciones?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['DG2']
        codigo_cell.value = '¿Cuál es la cobertura geográfica de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['DW2']
        codigo_cell.value = 'Desagregación geográfica '
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['EM2']
        codigo_cell.value = 'Desagregación por zona'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['EP2']
        codigo_cell.value = 'Desagregación por grupos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EX2']
        codigo_cell.value = '¿Cuál es el costo anual de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EZ2']
        codigo_cell.value = '¿Cuál(es) son las fuentes de financiación de la operación estadística ?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        

        codigo_cell = ws['FE2']
        codigo_cell.value = 'Liste todas las variables que maneja la operación estadística (Ver hoja 3)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FF2']
        codigo_cell.value = '¿Cuáles son los resultados agregados o indicadores calculados? (Ver hoja 4)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['FG2']
        codigo_cell.value = '¿Cuál es el medio de obtención de los datos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['FQ2']
        codigo_cell.value = '¿Cuál es la periodicidad de recolección o acopio de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['FX2']
        codigo_cell.value = 'Indique cuáles de las siguientes herramientas son utilizadas en el procesamiento de los datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['GG2']
        codigo_cell.value = 'Haga una  breve descripción de la manera cómo se realiza el procesamiento de los datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['GH2']
        codigo_cell.value = '¿Qué tipo de análisis realiza a los resultados obtenidos en la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['GM2']
        codigo_cell.value = '¿A través de qué medio(s) difunde los resultados estadísticos (Agregados, indicadores), a los usuarios?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['GR2']
        codigo_cell.value = 'Indique la dirección de la página web donde se encuentran los resultados estadísticos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        

        codigo_cell = ws['GS2']
        codigo_cell.value = 'Fechas de disponibilidad de los resultados estadísticos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['GU2']
        codigo_cell.value = '¿Cuál es la próxima fecha de publicación de resultados estadísticos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        

        codigo_cell = ws['GZ2']
        codigo_cell.value = '¿Cuál es la frecuencia de difusión de los resultados estadísticos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        
        codigo_cell = ws['HG2']
        codigo_cell.value = '¿Cuáles productos utiliza para difundir los resultados estadísticos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['HN2']
        codigo_cell.value = '¿Qué otros productos estadísticos de la OE están disponibles para consulta de los usuarios?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['HX2']
        codigo_cell.value = '¿Conoce si otra entidad produce resultados similares a los de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        
        codigo_cell = ws['IC2']
        codigo_cell.value = 'Observaciones'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['ID2']
        codigo_cell.value = 'Anexos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['IE2']
        codigo_cell.value = 'Estado del proceso de la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['IF2']
        codigo_cell.value = 'Estado de la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['IG2']
        codigo_cell.value = 'Validación de la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ## pin 2

        ####### insert heads fields

        codigo_cell = ws['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['B3']
        codigo_cell.value = 'Nombre Operación Estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['C3']
        codigo_cell.value = 'Objetivo de la Operación Estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['D3']
        codigo_cell.value = 'Código de la Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center',  wrap_text=True)
       
        codigo_cell = ws['E3']
        codigo_cell.value = 'Entidad responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['F3']
        codigo_cell.value = 'Código de la Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['G3']
        codigo_cell.value = 'Entidad responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['H3']
        codigo_cell.value = 'Código de la Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['I3']
        codigo_cell.value = 'Entidad responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['J3']
        codigo_cell.value = 'Área Tematica'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = ws['K3']
        codigo_cell.value = 'Tema'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['L3']
        codigo_cell.value = 'Tema Compartido'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['M3']
        codigo_cell.value = 'Nombre de la Dependencia'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['N3']
        codigo_cell.value = 'Nombre del Director'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['O3']
        codigo_cell.value = 'Cargo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['P3']
        codigo_cell.value = 'Correo Electrónico del Director'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = ws['Q3']
        codigo_cell.value = 'Teléfono del director'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['R3']
        codigo_cell.value = 'Nombre del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['S3']
        codigo_cell.value = 'Cargo del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['T3']
        codigo_cell.value = 'Correo Electrónico del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['U3']
        codigo_cell.value = 'Teléfono del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['W3']
        codigo_cell.value = 'a. Constitución Política'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['X3']
        codigo_cell.value = 'b. Ley'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['Y3']
        codigo_cell.value = 'c. Decreto (nacional, departamental, municipal)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['Z3']
        codigo_cell.value = 'd. Otra (Resolución, ordenanza, acuerdo)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center',  wrap_text=True)

        codigo_cell = ws['AA3']
        codigo_cell.value = 'e. Ninguna'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AB3']
        codigo_cell.value = 'a. Objetivos de desarrollo Sostenible (ODS)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['AC3']
        codigo_cell.value = 'b. Organización para la Cooperación y el Desarrollo Económico (OCDE)   '
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['AD3']
        codigo_cell.value = 'c. Otros compromisos internacionales'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['AE3']
        codigo_cell.value = 'd. Plan Nacional de Desarrollo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AF3']
        codigo_cell.value = 'e. Cuentas económicas y macroeconómicas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['AG3']
        codigo_cell.value = 'f. Plan Sectorial, Territorial o CONPES'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['AH3']
        codigo_cell.value = 'g. Otro (s)  ¿cuál(es) ?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AI3']
        codigo_cell.value = 'h. Ninguno'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AJ3']
        codigo_cell.value = 'a. Organismos internacionales'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AK3']
        codigo_cell.value = 'b. Presidencia de la República'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AL3']
        codigo_cell.value = 'c. Ministerios'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AM3']
        codigo_cell.value = 'd. Organismos de Control'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AN3']
        codigo_cell.value = 'e. Otras entidades del orden Nacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AO3']
        codigo_cell.value = 'f. Entidades de orden Territorial'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AP3']
        codigo_cell.value = 'g. Gremios'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AQ3']
        codigo_cell.value = 'h. Entidades privadas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AR3']
        codigo_cell.value = 'i. Dependencias de la misma entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AS3']
        codigo_cell.value = 'j. Academia'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AT3']
        codigo_cell.value = 'k. Público en General'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AV3']
        codigo_cell.value = 'a. Empresa'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        codigo_cell = ws['AW3']
        codigo_cell.value = 'b. Establecimiento'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AX3']
        codigo_cell.value = 'c. Hogar'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AY3']
        codigo_cell.value = 'd. Persona'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AZ3']
        codigo_cell.value = 'e. Unidad productora agropecuaria'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BA3']
        codigo_cell.value = 'f. Predio'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BB3']
        codigo_cell.value = 'g. Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BC3']
        codigo_cell.value = 'Cuál(es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['BD3']
        codigo_cell.value = 'a. Aprovechamiento de registro administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BE3']
        codigo_cell.value = 'b. Estadística derivada'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BF3']
        codigo_cell.value = 'c. Muestreo probabilístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BG3']
        codigo_cell.value = 'd. Muestreo no probabilístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BH3']
        codigo_cell.value = 'e. Censo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['BI3']
        codigo_cell.value = 'Registro Administrativos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BJ3']
        codigo_cell.value = 'Listado de RRAA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BK3']
        codigo_cell.value = 'Cuál(es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BL3']
        codigo_cell.value = 'Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BM3']
        codigo_cell.value = 'Operación Estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BN3']
        codigo_cell.value = 'Listado de OOEE'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BO3']
        codigo_cell.value = 'Cuál(es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BP3']
        codigo_cell.value =  'Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BQ3']
        codigo_cell.value =  'Muestreo aleatorio simple'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BR3']
        codigo_cell.value =  'Muestreo aleatorio sistemático'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BS3']
        codigo_cell.value =  'Muestreo aleatorio estratificado'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BT3']
        codigo_cell.value =  'Muestreo aleatorio por conglomerados'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BU3']
        codigo_cell.value =  'Otro'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BV3']
        codigo_cell.value =  '¿Cual?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['BW3']
        codigo_cell.value =  'Muestreo por cuotas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BX3']
        codigo_cell.value =  'Muestreo intencional o de conveniencia'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BY3']
        codigo_cell.value =  'Bola de nieve'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BZ3']
        codigo_cell.value =  'Muestreo discrecional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CA3']
        codigo_cell.value =  'Otro'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CB3']
        codigo_cell.value =  '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['CC3']
        codigo_cell.value =  'Se cuenta con un marco'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CD3']
        codigo_cell.value =  'Marco de lista'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CE3']
        codigo_cell.value =  'Marco de área'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CF3']
        codigo_cell.value =  'Marco geoestadístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CG3']
        codigo_cell.value =  'Otro(s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CH3']
        codigo_cell.value =  '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['CI3']
        codigo_cell.value =  'Metodología'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CJ3']
        codigo_cell.value =  'Ficha Metodológica (técnica)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CK3']
        codigo_cell.value =  'Hojas de vida de indicadores'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CL3']
        codigo_cell.value =  'Manual operativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CM3']
        codigo_cell.value =  'Diccionario de datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CN3']
        codigo_cell.value =  'Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CO3']
        codigo_cell.value = '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CP3']
        codigo_cell.value = 'Ninguno'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        
        codigo_cell = ws['CQ3']
        codigo_cell.value = 'DANE'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CR3']
        codigo_cell.value = 'Organismos internacionales'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CS3']
        codigo_cell.value = '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CT3']
        codigo_cell.value = 'Otra entidad de orden nacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CU3']
        codigo_cell.value = '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CV3']
        codigo_cell.value = 'Leyes, decretos, etc'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CW3']
        codigo_cell.value = '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CX3']
        codigo_cell.value = 'Creación propia'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CY3']
        codigo_cell.value = 'Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CZ3']
        codigo_cell.value = '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DA3']
        codigo_cell.value = 'Ninguno'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DB3']
        codigo_cell.value = '¿Por qué?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        
        codigo_cell = ws['DC3']
        codigo_cell.value = '¿se utilizan nomenclaturas y/o clasificaciones estandarizadas?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DD3']
        codigo_cell.value = 'Si: ¿Cuáles?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DE3']
        codigo_cell.value = 'Si: Otras'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DF3']
        codigo_cell.value = 'No: ¿Por qué?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['DG3']
        codigo_cell.value = 'a. Nacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DH3']
        codigo_cell.value = 'b. Regional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DI3']
        codigo_cell.value = 'b. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DJ3']
        codigo_cell.value = 'b. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DK3']
        codigo_cell.value = 'c. Departamental'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DL3']
        codigo_cell.value = 'c. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DM3']
        codigo_cell.value = 'c. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DN3']
        codigo_cell.value = 'd. Áreas metropolitanas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DO3']
        codigo_cell.value = 'd. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DP3']
        codigo_cell.value = 'd. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DQ3']
        codigo_cell.value = 'e. Municipal'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DR3']
        codigo_cell.value = 'e. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DS3']
        codigo_cell.value = 'e. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DT3']
        codigo_cell.value = 'f. Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DU3']
        codigo_cell.value = 'f. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DV3']
        codigo_cell.value = 'f. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        
        codigo_cell = ws['DW3']
        codigo_cell.value = 'a. Nacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DX3'] 
        codigo_cell.value = 'b. Regional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DY3']
        codigo_cell.value = 'b. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DZ3']
        codigo_cell.value = 'b. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EA3']
        codigo_cell.value = 'c. Departamental'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EB3']
        codigo_cell.value = 'c. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EC3']
        codigo_cell.value = 'c. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['ED3']
        codigo_cell.value = 'd. Áreas metropolitanas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EE3']
        codigo_cell.value = 'd. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EF3']
        codigo_cell.value = 'd. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EG3']
        codigo_cell.value = 'e. Municipal'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EH3']
        codigo_cell.value = 'e. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EI3']
        codigo_cell.value = 'e. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EJ3']
        codigo_cell.value = 'f. Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EK3']
        codigo_cell.value = 'f. ¿Cuántos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EL3']
        codigo_cell.value = 'f. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EM3']
        codigo_cell.value = 'Total'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EN3']
        codigo_cell.value = 'Urbano'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EO3']
        codigo_cell.value = 'Rural'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['EP3']
        codigo_cell.value = 'Sexo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EQ3']
        codigo_cell.value = 'Edad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['ER3']
        codigo_cell.value = 'Grupo étnico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['ES3']
        codigo_cell.value = 'Discapacidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['ET3']
        codigo_cell.value = 'Estrato'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EU3']
        codigo_cell.value = 'Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EV3']
        codigo_cell.value = '¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EW3']
        codigo_cell.value = 'Ninguno'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['EX3']
        codigo_cell.value = 'a. Costo anuaL'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EY3']
        codigo_cell.value = 'b. No sabe'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['EZ3']
        codigo_cell.value = 'Recursos propios'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FA3']
        codigo_cell.value = 'Aportes de otra entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FB3']
        codigo_cell.value = 'Cooperación internacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FC3']
        codigo_cell.value = 'Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FD3']
        codigo_cell.value = '¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        
        codigo_cell = ws['FG3']
        codigo_cell.value = 'a. Formulario físico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FH3']
        codigo_cell.value = 'b. Formulario electrónico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FI3']
        codigo_cell.value = 'c. Dispositivo Móvil de Captura [DMC]'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FJ3']
        codigo_cell.value = 'd. Sistema de Información'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FK3']
        codigo_cell.value = 'd. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FL3']
        codigo_cell.value = 'e. Percepción remota (imágenes satelitales, fotos, sensores, etc)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FM3']
        codigo_cell.value = 'f. Base de datos de registro administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FN3']
        codigo_cell.value = 'g. Resultados estadísticos de otra operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FO3']
        codigo_cell.value = 'h. Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FP3']
        codigo_cell.value = 'h. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['FQ3']
        codigo_cell.value = 'a. Anual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FR3']
        codigo_cell.value = 'b. Semestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FS3']
        codigo_cell.value = 'c. Trimestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FT3']
        codigo_cell.value = 'd. Mensual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FU3']
        codigo_cell.value = 'e. Diaria'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FV3']
        codigo_cell.value = 'f. Otra'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FW3']
        codigo_cell.value = 'f. ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['FX3']
        codigo_cell.value = 'a. Excel'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FY3']
        codigo_cell.value = 'b. Acces'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FZ3']
        codigo_cell.value = 'c. R'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GA3']
        codigo_cell.value = 'd. SAS'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GB3']
        codigo_cell.value = 'e. SPSS'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GC3']
        codigo_cell.value = 'f. Oracle'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GD3']
        codigo_cell.value = 'g. Stata'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GE3']
        codigo_cell.value = 'h. Otra'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GF3']
        codigo_cell.value = 'h. ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        
        codigo_cell = ws['GH3']
        codigo_cell.value = 'a Consistencia y validación de datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GI3']
        codigo_cell.value = 'b Análisis de tendencias y series de tiempo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GJ3']
        codigo_cell.value = 'c Análisis de contexto de los resultados'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GK3']
        codigo_cell.value = 'd Otro.'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GL3']
        codigo_cell.value = 'd. ¿cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['GM3']
        codigo_cell.value = 'a. Página web'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GN3']
        codigo_cell.value = 'b. Medio físico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GO3']
        codigo_cell.value = 'c. Medio electrónico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GP3']
        codigo_cell.value = 'd. Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GQ3']
        codigo_cell.value = 'd. ¿cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['GS3']
        codigo_cell.value = 'Desde mes y año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GT3']
        codigo_cell.value = 'Hasta mes y año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        
        codigo_cell = ws['GU3']
        codigo_cell.value = 'a. Fecha publicación'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GV3']
        codigo_cell.value = 'a. mes/año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GW3']
        codigo_cell.value = 'b. No sabe'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GX3']
        codigo_cell.value = 'c. No hay'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GY3']
        codigo_cell.value = 'c. ¿por qué?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    
        codigo_cell = ws['GZ3']
        codigo_cell.value = 'a. Anual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HA3']
        codigo_cell.value = 'b. Semestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HB3']
        codigo_cell.value = 'c. Trimestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HC3']
        codigo_cell.value = 'd. Mensual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HD3']
        codigo_cell.value = 'e. Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HE3']
        codigo_cell.value = 'e. ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HF3']
        codigo_cell.value = 'f. No está definido'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        
        codigo_cell = ws['HG3']
        codigo_cell.value = 'a. Cuadros de salida (tablas)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HH3']
        codigo_cell.value = 'b. Boletín estadístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HI3']
        codigo_cell.value = 'c. Anuario'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HJ3']
        codigo_cell.value = 'd. Mapas estadísticos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HK3']
        codigo_cell.value = 'e. Bases de datos interactivas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HL3']
        codigo_cell.value = 'f. Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HM3']
        codigo_cell.value = 'f. ¿cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        
        codigo_cell = ws['HN3']
        codigo_cell.value = 'a. Series históricas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HO3']
        codigo_cell.value = 'a. Desde: mes y año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HP3']
        codigo_cell.value = 'a. Hasta: mes y año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HQ3']
        codigo_cell.value = 'b. Microdatos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HR3']
        codigo_cell.value = 'b. Desde: mes y año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HS3']
        codigo_cell.value = 'b. Hasta: mes y año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HT3']
        codigo_cell.value = 'c. Documentos metodológicos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HU3']
        codigo_cell.value = 'c. URL'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HV3']
        codigo_cell.value = 'd. Calendario de difusión'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HW3']
        codigo_cell.value = 'e. Ninguna'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        
        codigo_cell = ws['HX3']
        codigo_cell.value = '¿Conoce si otra entidad produce resultados similares a los de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HY3']
        codigo_cell.value = 'Si: Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HZ3']
        codigo_cell.value = 'Si: Operación estadística/Indicadores'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        
        codigo_cell = ws['IA3']
        codigo_cell.value = '¿La entidad hace parte de un sistema de información estadística de uso  interinstitucional ?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['IB3']
        codigo_cell.value = 'Si: ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ##pin 3

        cont = 4

        for ooee in ooees:
            
            ws.cell(row = cont, column = 1).value = ooee.codigo_oe
            ws.cell(row = cont, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 1).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 2).value = ooee.nombre_oe
            ws.cell(row = cont, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 2).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 3).value = ooee.objetivo_oe
            ws.cell(row = cont, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 3).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 4).value =  ooee.entidad.codigo
            ws.cell(row = cont, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 4).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 5).value =  str(ooee.entidad)
            ws.cell(row = cont, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 5).font = Font(size = "8", name='Barlow')

            if ooee.entidad_resp2 == None:
                ws.cell(row = cont, column = 6).value = ""
                ws.cell(row = cont, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 6).font = Font(size = "8", name='Barlow')

                ws.cell(row = cont, column = 7).value = ""
                ws.cell(row = cont, column = 7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 7).font = Font(size = "8", name='Barlow')
            else:
                ws.cell(row = cont, column = 6).value =  ooee.entidad_resp2.codigo
                ws.cell(row = cont, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 6).font = Font(size = "8", name='Barlow')

                ws.cell(row = cont, column = 7).value = str(ooee.entidad_resp2)
                ws.cell(row = cont, column = 7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 7).font = Font(size = "8", name='Barlow')

            
            if ooee.entidad_resp3 == None:
                ws.cell(row = cont, column = 8).value = ""
                ws.cell(row = cont, column = 8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 8).font = Font(size = "8", name='Barlow')

                ws.cell(row = cont, column = 9).value = ""
                ws.cell(row = cont, column = 9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 9).font = Font(size = "8", name='Barlow')
            else:
                ws.cell(row = cont, column = 8).value = ooee.entidad_resp3.codigo
                ws.cell(row = cont, column = 8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 8).font = Font(size = "8", name='Barlow')

                ws.cell(row = cont, column = 9).value = str(ooee.entidad_resp3)
                ws.cell(row = cont, column = 9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 9).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 10).value = str(ooee.area_tematica)
            ws.cell(row = cont, column = 10).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 10).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 11).value = str(ooee.tema)
            ws.cell(row = cont, column = 11).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 11).font = Font(size = "8", name='Barlow')
            
            ## TEMA COMPARTIDO

            listaTemaCompartido = ooee.tema_compartido.all()
            temaCompartido_array = []
            for indexTemaCompa, itemTemaCompa in enumerate(listaTemaCompartido):
                temaCompartido_array.append(str(itemTemaCompa))
                ws.cell(row = cont, column = 12).value = str(temaCompartido_array)
                ws.cell(row = cont, column = 12).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 12).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 13).value = ooee.nombre_dep
            ws.cell(row = cont, column = 13).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = cont, column = 13).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 14).value = ooee.nombre_dir
            ws.cell(row = cont, column = 14).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 14).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 15).value = ooee.cargo_dir
            ws.cell(row = cont, column = 15).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 15).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 16).value = ooee.correo_dir
            ws.cell(row = cont, column = 16).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 16).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 17).value = ooee.tel_dir
            ws.cell(row = cont, column = 17).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 17).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 18).value = ooee.nombre_resp
            ws.cell(row = cont, column = 18).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 18).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 19).value = ooee.cargo_resp
            ws.cell(row = cont, column = 19).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 19).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 20).value = ooee.correo_resp
            ws.cell(row = cont, column = 20).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 20).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 21).value = ooee.tel_resp
            ws.cell(row = cont, column = 21).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 21).font = Font(size = "8", name='Barlow')
            
            if str(ooee.fase) == "True":
                ws.cell(row = cont, column = 22).value = "Si"
                ws.cell(row = cont, column = 22).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 22).font = Font(size = "8", name='Barlow')

            elif str(ooee.fase) == "False":
                ws.cell(row = cont, column = 22).value = "No"
                ws.cell(row = cont, column = 22).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 22).font = Font(size = "8", name='Barlow')

            #*****lista de OOEE Y RRAA PREGUNTA 4 DEL modulo c
            listaRRAAobtencion = ooee.rraa_lista.all()
            listaRRAAobt_array = []
            for indexListaRRAAobt, itemListaRRAAobt in enumerate(listaRRAAobtencion):
                listaRRAAobt_array.append(str(itemListaRRAAobt))
                ws.cell(row = cont, column = 62).value = str(listaRRAAobt_array).replace('[','').replace(']','')
                ws.cell(row = cont, column = 62).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 62).font = Font(size = "8", name='Barlow')

            listaOOEEobtencion = ooee.ooee_lista.all()
            listaOOEEobt_array = []
            for indexListaOOEEobt, itemListaOOEEobt in enumerate(listaOOEEobtencion):
                listaOOEEobt_array.append(str(itemListaOOEEobt))
                ws.cell(row = cont, column = 66).value = str(listaOOEEobt_array)
                ws.cell(row = cont, column = 66).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 66).font = Font(size = "8", name='Barlow')

            
            listaNorma = ooee.norma.all()
            listaRequerimiento = ooee.requerimientos.all()
            listaPriUser = ooee.pri_usuarios.all()
            listaUnidadObs = ooee.uni_observacion.all()
            listaTipoOper = ooee.tipo_operacion.all()
            listaObtDato =  ooee.obt_dato.all()
            listaTipoProbabilistico = ooee.tipo_probabilistico.all()
            listaTipoNoProbabilistico = ooee.tipo_no_probabilistico.all()
            listaTipoMarco = ooee.tipo_marco.all()
            listaDocsDesarrollo = ooee.docs_des.all()
            ListaListaConc = ooee.lista_conc.all()
            listaClasificaciones =  ooee.nombre_cla.all()
            listaCobGeog = ooee.cob_geo.all()
            listaDesgeo = ooee.des_geo.all()
            listaDesZona = ooee.des_zona.all()
            listaDesGrupo = ooee.des_grupo.all()
            listaFuentes = ooee.fuentes.all()
            listaMediObtDato = ooee.med_obt.all()
            listaPeriodicidad = ooee.periodicidad.all()
            listaHerramProcesa = ooee.h_proc.all()
            listaAnalisResultados = ooee.a_resul.all()
            listaMedioDifusion = ooee.m_dif.all()
            listaFechaPublicacion = ooee.f_publi.all()
            listaFrecuenciaDifusion = ooee.fre_dif.all()
            listaProductosDifundir = ooee.pro_dif.all()
            listaOtrosProductos = ooee.otro_prod.all()

            for index,item in enumerate(listaNorma):
                indice = index
                if  str(item) == 'Ninguna' and index == indice:
                    #print("item", item)
                    ws.cell(row = cont, column = 27).value = str(item)
                    ws.cell(row = cont, column = 27).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 27).font = Font(size = "8", name='Barlow')
                   

            for indexReq, itemReq in enumerate(listaRequerimiento):
                indiceReq = indexReq
                if indexReq == indiceReq and str(itemReq) == 'Ninguno':
                    ws.cell(row = cont, column = 35).value = str(itemReq)
                    ws.cell(row = cont, column = 35).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 35).font = Font(size = "8", name='Barlow')
                   
            
            for indexPriUs, itemPrius in enumerate(listaPriUser):
                indicePriUser = indexPriUs
                if indexPriUs == indicePriUser and str(itemPrius) == 'Público en General':
                    ws.cell(row = cont, column = 46).value = str(itemPrius)
                    ws.cell(row = cont, column = 46).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 46).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 47).value = ooee.pob_obje
            ws.cell(row = cont, column = 47).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 47).font = Font(size = "8", name='Barlow')

            for indexUnidadObs, itemUnidadObs, in enumerate(listaUnidadObs):   
                if  str(itemUnidadObs) == "Empresa":
                    ws.cell(row = cont, column = 48).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 48).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 48).font = Font(size = "8", name='Barlow')
                if  str(itemUnidadObs) == "Establecimiento":   
                    ws.cell(row = cont, column = 49).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 49).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 49).font = Font(size = "8", name='Barlow')
                if  str(itemUnidadObs) == "Hogar":   
                    ws.cell(row = cont, column = 50).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 50).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 50).font = Font(size = "8", name='Barlow')
                if  str(itemUnidadObs) == "Persona":   
                    ws.cell(row = cont, column = 51).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 51).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 51).font = Font(size = "8", name='Barlow')
                if  str(itemUnidadObs) == "Unidad productora agropecuaria":   
                    ws.cell(row = cont, column = 52).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 52).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 52).font = Font(size = "8", name='Barlow')
                if  str(itemUnidadObs) == "Predio":   
                    ws.cell(row = cont, column = 53).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 53).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 53).font = Font(size = "8", name='Barlow')
                if  str(itemUnidadObs) == "Otra (s)":   
                    ws.cell(row = cont, column = 54).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 54).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 54).font = Font(size = "8", name='Barlow')

            for indexListaOper, itemListaOper in enumerate(listaTipoOper):
                if str(itemListaOper) == "Aprovechamiento de registro administrativo":
                    ws.cell(row = cont, column = 56).value = str(itemListaOper)
                    ws.cell(row = cont, column = 56).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 56).font = Font(size = "8", name='Barlow')
                if str(itemListaOper) == "Estadística derivada":
                    ws.cell(row = cont, column = 57).value = str(itemListaOper)
                    ws.cell(row = cont, column = 57).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 57).font = Font(size = "8", name='Barlow')
                if str(itemListaOper) == "Muestreo probabilístico":
                    ws.cell(row = cont, column = 58).value = str(itemListaOper)
                    ws.cell(row = cont, column = 58).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 58).font = Font(size = "8", name='Barlow')
                if str(itemListaOper) == "Muestreo no probabilístico":
                    ws.cell(row = cont, column = 59).value = str(itemListaOper)
                    ws.cell(row = cont, column = 59).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 59).font = Font(size = "8", name='Barlow')
                if str(itemListaOper) == "Censo":
                    ws.cell(row = cont, column = 60).value = str(itemListaOper)
                    ws.cell(row = cont, column = 60).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 60).font = Font(size = "8", name='Barlow')
            
            for indexObtDato, itemObtDato in enumerate(listaObtDato):
                if str(itemObtDato) == "Registro Administrativos":
                    ws.cell(row = cont, column = 61).value = str(itemObtDato)
                    ws.cell(row = cont, column = 61).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 61).font = Font(size = "8", name='Barlow')
                if str(itemObtDato) == "Operación Estadística":
                    ws.cell(row = cont, column = 65).value = str(itemObtDato)
                    ws.cell(row = cont, column = 65).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 65).font = Font(size = "8", name='Barlow')            

            for indexTipoProba, itemTipoProba in enumerate(listaTipoProbabilistico):
                if str(itemTipoProba) == "Muestreo aleatorio simple":
                    ws.cell(row = cont, column = 69).value = str(itemTipoProba)
                    ws.cell(row = cont, column = 69).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 69).font = Font(size = "8", name='Barlow')
                if str(itemTipoProba) == "Muestreo aleatorio sistemático":
                    ws.cell(row = cont, column = 70).value = str(itemTipoProba)
                    ws.cell(row = cont, column = 70).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 70).font = Font(size = "8", name='Barlow')
                if str(itemTipoProba) == "Muestreo aleatorio estratificado":
                    ws.cell(row = cont, column = 71).value = str(itemTipoProba)
                    ws.cell(row = cont, column = 71).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 71).font = Font(size = "8", name='Barlow')
                if str(itemTipoProba) == "Muestreo aleatorio por conglomerados":
                    ws.cell(row = cont, column = 72).value = str(itemTipoProba)
                    ws.cell(row = cont, column = 72).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 72).font = Font(size = "8", name='Barlow')
                if str(itemTipoProba) == "Otro":
                    ws.cell(row = cont, column = 73).value = str(itemTipoProba)
                    ws.cell(row = cont, column = 73).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 73).font = Font(size = "8", name='Barlow')

            for indexTipoNoProba, itemTipoNoProba in enumerate(listaTipoNoProbabilistico):
                if str(itemTipoNoProba) == "Muestreo por cuotas":
                    ws.cell(row = cont, column = 75).value = str(itemTipoNoProba)
                    ws.cell(row = cont, column = 75).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 75).font = Font(size = "8", name='Barlow')
                if str(itemTipoNoProba) == "Muestreo intencional o de conveniencia":
                    ws.cell(row = cont, column = 76).value = str(itemTipoNoProba)
                    ws.cell(row = cont, column = 76).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 76).font = Font(size = "8", name='Barlow')
                if str(itemTipoNoProba) == "Bola de nieve":
                    ws.cell(row = cont, column = 77).value = str(itemTipoNoProba)
                    ws.cell(row = cont, column = 77).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)                               
                    ws.cell(row = cont, column = 77).font = Font(size = "8", name='Barlow')
                if str(itemTipoNoProba) == "Muestreo discrecional":
                    ws.cell(row = cont, column = 78).value = str(itemTipoNoProba)
                    ws.cell(row = cont, column = 78).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 78).font = Font(size = "8", name='Barlow')
                if str(itemTipoNoProba) == "Otro":
                    ws.cell(row = cont, column = 79).value = str(itemTipoNoProba)
                    ws.cell(row = cont, column = 79).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 79).font = Font(size = "8", name='Barlow')
            
            if str(ooee.marco_estad) == "True":

                ws.cell(row = cont, column = 81).value = "Si"
                ws.cell(row = cont, column = 81).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 81).font = Font(size = "8", name='Barlow')
            
            elif str(ooee.marco_estad) == "False":
                ws.cell(row = cont, column = 81).value = "No"
                ws.cell(row = cont, column = 81).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 81).font = Font(size = "8", name='Barlow')
            
            for indexTipoMarco, itemTipoMarco in enumerate(listaTipoMarco):
                if str(itemTipoMarco) == "Marco de lista":
                    ws.cell(row = cont, column = 82).value = str(itemTipoMarco)
                    ws.cell(row = cont, column = 82).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 82).font = Font(size = "8", name='Barlow')
                if str(itemTipoMarco) == "Marco de área":
                    ws.cell(row = cont, column = 83).value = str(itemTipoMarco)
                    ws.cell(row = cont, column = 83).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 83).font = Font(size = "8", name='Barlow')
                if str(itemTipoMarco) == "Marco geoestadístico":
                    ws.cell(row = cont, column = 84).value = str(itemTipoMarco)
                    ws.cell(row = cont, column = 84).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 84).font = Font(size = "8", name='Barlow')
                if str(itemTipoMarco) == "Otro(s)":
                    ws.cell(row = cont, column = 85).value = str(itemTipoMarco)
                    ws.cell(row = cont, column = 85).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)                        
                    ws.cell(row = cont, column = 85).font = Font(size = "8", name='Barlow')
            
            for indexDocsDesarrollo, itemDocsDesarrollo in enumerate(listaDocsDesarrollo):
                if str(itemDocsDesarrollo) == "Metodología":
                    ws.cell(row = cont, column = 87).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 87).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 87).font = Font(size = "8", name='Barlow')
                if str(itemDocsDesarrollo) == "Ficha Metodológica (técnica)":
                    ws.cell(row = cont, column = 88).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 88).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 88).font = Font(size = "8", name='Barlow')
                if str(itemDocsDesarrollo) == "Hojas de vida de indicadores":
                    ws.cell(row = cont, column = 89).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 89).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 89).font = Font(size = "8", name='Barlow')
                if str(itemDocsDesarrollo) == "Manual operativo":
                    ws.cell(row = cont, column = 90).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 90).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 90).font = Font(size = "8", name='Barlow')
                if str(itemDocsDesarrollo) == "Diccionario de datos":
                    ws.cell(row = cont, column = 91).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 91).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 91).font = Font(size = "8", name='Barlow')
                if str(itemDocsDesarrollo) == "Otro (s)":
                    ws.cell(row = cont, column = 92).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 92).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 92).font = Font(size = "8", name='Barlow')
                if str(itemDocsDesarrollo) == "Ninguno":
                    ws.cell(row = cont, column = 94).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 94).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 94).font = Font(size = "8", name='Barlow')
            
            for indexListaConc, itemListaConc in enumerate(ListaListaConc):
                if str(itemListaConc) == "DANE":
                    ws.cell(row = cont, column = 95).value = str(itemListaConc)
                    ws.cell(row = cont, column = 95).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 95).font = Font(size = "8", name='Barlow')
                if str(itemListaConc) == "Organismos internacionales":
                    ws.cell(row = cont, column = 96).value = str(itemListaConc)
                    ws.cell(row = cont, column = 96).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 96).font = Font(size = "8", name='Barlow')
                if str(itemListaConc) == "Otra entidad de orden nacional":
                    ws.cell(row = cont, column = 98).value = str(itemListaConc)
                    ws.cell(row = cont, column = 98).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 98).font = Font(size = "8", name='Barlow')
                if str(itemListaConc) == "Leyes, decretos, etc.":
                    ws.cell(row = cont, column = 100).value = str(itemListaConc)
                    ws.cell(row = cont, column = 100).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 100).font = Font(size = "8", name='Barlow')
                if str(itemListaConc) == "Creación propia":
                    ws.cell(row = cont, column = 102).value = str(itemListaConc)
                    ws.cell(row = cont, column = 102).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 102).font = Font(size = "8", name='Barlow')
                if str(itemListaConc) == "Otra (s)":
                    ws.cell(row = cont, column = 103).value = str(itemListaConc)
                    ws.cell(row = cont, column = 103).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 103).font = Font(size = "8", name='Barlow')
                if str(itemListaConc) == "Ninguno":
                    ws.cell(row = cont, column = 105).value = str(itemListaConc)
                    ws.cell(row = cont, column = 105).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 105).font = Font(size = "8", name='Barlow')

            if str(ooee.nome_clas) == "True":
                ws.cell(row = cont, column = 107).value = "Si"
                ws.cell(row = cont, column = 107).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 107).font = Font(size = "8", name='Barlow')
            
            elif str(ooee.nome_clas) == "False":
                ws.cell(row = cont, column = 107).value = "No"
                ws.cell(row = cont, column = 107).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 107).font = Font(size = "8", name='Barlow')
            
            clasifi_array = []
            for indexClasif, itemClasif in enumerate(listaClasificaciones):
                clasifi_array.append(str(itemClasif))
                ws.cell(row = cont, column = 108).value = str(clasifi_array)
                ws.cell(row = cont, column = 108).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 108).font = Font(size = "8", name='Barlow')
        
            for indexCobGeo, itemCobGeo in enumerate(listaCobGeog):
                
                if str(itemCobGeo) == "Nacional":
                    ws.cell(row = cont, column = 111).value = str(itemCobGeo)
                    ws.cell(row = cont, column = 111).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 111).font = Font(size = "8", name='Barlow')

                if str(itemCobGeo) == "Regional":
                    ws.cell(row = cont, column = 112).value = str(itemCobGeo)
                    ws.cell(row = cont, column = 112).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 112).font = Font(size = "8", name='Barlow')
                
                if str(itemCobGeo) == "Departamental":
                    ws.cell(row = cont, column = 115).value = str(itemCobGeo)
                    ws.cell(row = cont, column = 115).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 115).font = Font(size = "8", name='Barlow')
                
                if str(itemCobGeo) == "Áreas metropolitanas":
                    ws.cell(row = cont, column = 118).value = str(itemCobGeo)
                    ws.cell(row = cont, column = 118).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 118).font = Font(size = "8", name='Barlow')
                
                if str(itemCobGeo) == "Municipal":
                    ws.cell(row = cont, column = 121).value = str(itemCobGeo)
                    ws.cell(row = cont, column = 121).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 121).font = Font(size = "8", name='Barlow')
                
                if str(itemCobGeo) == "Otro (s)":
                    ws.cell(row = cont, column = 124).value = str(itemCobGeo)
                    ws.cell(row = cont, column = 124).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 124).font = Font(size = "8", name='Barlow')


                for indexDesgeo, itemDesgeo in enumerate(listaDesgeo): 
                    
                    if str(itemDesgeo) == "Nacional":
                        ws.cell(row = cont, column = 127).value = str(itemDesgeo)
                        ws.cell(row = cont, column = 127).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 127).font = Font(size = "8", name='Barlow')
                    
                    if str(itemDesgeo) == "Regional":
                        ws.cell(row = cont, column = 128).value = str(itemDesgeo)
                        ws.cell(row = cont, column = 128).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 128).font = Font(size = "8", name='Barlow')

                    if str(itemDesgeo) == "Departamental":
                        ws.cell(row = cont, column = 131).value = str(itemDesgeo)
                        ws.cell(row = cont, column = 131).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 131).font = Font(size = "8", name='Barlow')
                    
                    if str(itemDesgeo) == "Áreas metropolitanas":
                        ws.cell(row = cont, column = 134).value = str(itemDesgeo)
                        ws.cell(row = cont, column = 134).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 134).font = Font(size = "8", name='Barlow')
                    
                    if str(itemDesgeo) == "Municipal":
                        ws.cell(row = cont, column = 137).value = str(itemDesgeo)
                        ws.cell(row = cont, column = 137).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 137).font = Font(size = "8", name='Barlow')
                    
                    if str(itemDesgeo) == "Otro (s)":
                        ws.cell(row = cont, column = 140).value = str(itemDesgeo)
                        ws.cell(row = cont, column = 140).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 140).font = Font(size = "8", name='Barlow')
                
                
                for indexDesZona, itemDesZona in enumerate(listaDesZona):
                    if str(itemDesZona) == "Total":
                        ws.cell(row = cont, column = 143).value = str(itemDesZona)
                        ws.cell(row = cont, column = 143).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 143).font = Font(size = "8", name='Barlow')

                    if str(itemDesZona) == "Urbano":
                        ws.cell(row = cont, column = 144).value = str(itemDesZona)
                        ws.cell(row = cont, column = 144).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 144).font = Font(size = "8", name='Barlow')
                    
                    if str(itemDesZona) == "Rural":
                        ws.cell(row = cont, column = 145).value = str(itemDesZona)
                        ws.cell(row = cont, column = 145).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 145).font = Font(size = "8", name='Barlow')
                    
                
                for indexDesGrupo, itemDesGrupo in enumerate(listaDesGrupo):
                    
                    if str(itemDesGrupo) == "Sexo":
                        ws.cell(row = cont, column = 146).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 146).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 146).font = Font(size = "8", name='Barlow')

                    if str(itemDesGrupo) == "Edad":
                        ws.cell(row = cont, column = 147).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 147).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 147).font = Font(size = "8", name='Barlow')

                    if str(itemDesGrupo) == "Grupo étnico":
                        ws.cell(row = cont, column = 148).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 148).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 148).font = Font(size = "8", name='Barlow')

                    if str(itemDesGrupo) == "Discapacidad":
                        ws.cell(row = cont, column = 149).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 149).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 149).font = Font(size = "8", name='Barlow')

                    if str(itemDesGrupo) == "Estrato":
                        ws.cell(row = cont, column = 150).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 150).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 150).font = Font(size = "8", name='Barlow')

                    if str(itemDesGrupo) == "Otro, ¿cuál?":
                        ws.cell(row = cont, column = 151).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 151).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 151).font = Font(size = "8", name='Barlow')

                    if str(itemDesGrupo) == "Ninguno":
                        ws.cell(row = cont, column = 153).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 153).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 153).font = Font(size = "8", name='Barlow')

                ws.cell(row = cont, column = 154).value = ooee.ca_anual
                ws.cell(row = cont, column = 154).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 154).font = Font(size = "8", name='Barlow') 

                ws.cell(row = cont, column = 155).value = ooee.cb_anual
                ws.cell(row = cont, column = 155).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 155).font = Font(size = "8", name='Barlow')  
               
                for indexFuentes, itemFuentes in enumerate(listaFuentes):
                    
                    if str(itemFuentes) == "Recursos propios":
                        ws.cell(row = cont, column = 156).value = str(itemFuentes)
                        ws.cell(row = cont, column = 156).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 156).font = Font(size = "8", name='Barlow')

                    if str(itemFuentes) == "Aportes de otra entidad":
                        ws.cell(row = cont, column = 157).value = str(itemFuentes)
                        ws.cell(row = cont, column = 157).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 157).font = Font(size = "8", name='Barlow')

                    if str(itemFuentes) == "Cooperación internacional":
                        ws.cell(row = cont, column = 158).value = str(itemFuentes)
                        ws.cell(row = cont, column = 158).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 158).font = Font(size = "8", name='Barlow')

                    if str(itemFuentes) == "Otro (s)":
                        ws.cell(row = cont, column = 159).value = str(itemFuentes)
                        ws.cell(row = cont, column = 159).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 159).font = Font(size = "8", name='Barlow')
                
                
                for indexMediObtDato, itemMediObtDato in enumerate(listaMediObtDato):

                    if str(itemMediObtDato) == "Formulario físico":
                        ws.cell(row = cont, column = 163).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 163).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 163).font = Font(size = "8", name='Barlow')
                    
                    if str(itemMediObtDato) == "Formulario electrónico":
                        ws.cell(row = cont, column = 164).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 164).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 164).font = Font(size = "8", name='Barlow')

                    if str(itemMediObtDato) == "Dispositivo Móvil de Captura [DMC]":
                        ws.cell(row = cont, column = 165).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 165).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 165).font = Font(size = "8", name='Barlow')

                    if str(itemMediObtDato) == "Sistema de Información":
                        ws.cell(row = cont, column = 166).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 166).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 166).font = Font(size = "8", name='Barlow')

                    if str(itemMediObtDato) == "Percepción remota (imágenes satelitales, fotos, sensores, etc)":
                        ws.cell(row = cont, column = 168).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 168).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 168).font = Font(size = "8", name='Barlow')

                    if str(itemMediObtDato) == "Base de datos de registro administrativo":
                        ws.cell(row = cont, column = 169).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 169).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 169).font = Font(size = "8", name='Barlow')

                    if str(itemMediObtDato) == "Resultados estadísticos de otra operación estadística":
                        ws.cell(row = cont, column = 170).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 170).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 170).font = Font(size = "8", name='Barlow')

                    if str(itemMediObtDato) == "Otro (s)":
                        ws.cell(row = cont, column = 171).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 171).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 171).font = Font(size = "8", name='Barlow')

                for indexPeriodicidad, itemPeriodicidad in enumerate(listaPeriodicidad):
                    if str(itemPeriodicidad) == "Anual":
                        ws.cell(row = cont, column = 173).value = str(itemPeriodicidad)
                        ws.cell(row = cont, column = 173).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 173).font = Font(size = "8", name='Barlow')
                    
                    if str(itemPeriodicidad) == "Semestral":
                        ws.cell(row = cont, column = 174).value = str(itemPeriodicidad)
                        ws.cell(row = cont, column = 174).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 174).font = Font(size = "8", name='Barlow')

                    if str(itemPeriodicidad) == "Trimestral":
                        ws.cell(row = cont, column = 175).value = str(itemPeriodicidad)
                        ws.cell(row = cont, column = 175).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 175).font = Font(size = "8", name='Barlow')

                    if str(itemPeriodicidad) == "Mensual":
                        ws.cell(row = cont, column = 176).value = str(itemPeriodicidad)
                        ws.cell(row = cont, column = 176).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 176).font = Font(size = "8", name='Barlow')
                    
                    if str(itemPeriodicidad) == "Diaria":
                        ws.cell(row = cont, column = 177).value = str(itemPeriodicidad)
                        ws.cell(row = cont, column = 177).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 177).font = Font(size = "8", name='Barlow')
                    
                    if str(itemPeriodicidad) == "Otra":
                        ws.cell(row = cont, column = 178).value = str(itemPeriodicidad)
                        ws.cell(row = cont, column = 178).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 178).font = Font(size = "8", name='Barlow')

                for indexHerramProcesa, itemHerramProcesa in enumerate(listaHerramProcesa):
                    
                    if str(itemHerramProcesa) == "Excel" :
                        ws.cell(row = cont, column = 180).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 180).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 180).font = Font(size = "8", name='Barlow')

                    if str(itemHerramProcesa) == "Access" :
                        ws.cell(row = cont, column = 181).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 181).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 181).font = Font(size = "8", name='Barlow')

                    if str(itemHerramProcesa) == "R" :
                        ws.cell(row = cont, column = 182).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 182).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 182).font = Font(size = "8", name='Barlow')

                    if str(itemHerramProcesa) == "SAS" :
                        ws.cell(row = cont, column = 183).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 183).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 183).font = Font(size = "8", name='Barlow')

                    if str(itemHerramProcesa) == "SPSS" :
                        ws.cell(row = cont, column = 184).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 184).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 184).font = Font(size = "8", name='Barlow')
                    
                    if str(itemHerramProcesa) == "Oracle" :
                        ws.cell(row = cont, column = 185).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 185).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 185).font = Font(size = "8", name='Barlow')

                    if str(itemHerramProcesa) == "Stata" :
                        ws.cell(row = cont, column = 186).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 186).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 186).font = Font(size = "8", name='Barlow')
                    
                    if str(itemHerramProcesa) == "Otra (s)" :
                        ws.cell(row = cont, column = 187).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 187).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 187).font = Font(size = "8", name='Barlow')

                ws.cell(row = cont, column = 189).value = str(ooee.descrip_proces)
                ws.cell(row = cont, column = 189).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 189).font = Font(size = "8", name='Barlow')  
               
                for indexAnalisResultados, itemAnalisResultados in enumerate(listaAnalisResultados):
                    
                    if str(itemAnalisResultados) == "Consistencia y validación de datos":
                        ws.cell(row = cont, column = 190).value = str(itemAnalisResultados)
                        ws.cell(row = cont, column = 190).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 190).font = Font(size = "8", name='Barlow')

                    if str(itemAnalisResultados) == "Análisis de tendencias y series de tiempo":
                        ws.cell(row = cont, column = 191).value = str(itemAnalisResultados)
                        ws.cell(row = cont, column = 191).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 191).font = Font(size = "8", name='Barlow')

                    if str(itemAnalisResultados) == "Análisis de contexto de los resultados":
                        ws.cell(row = cont, column = 192).value = str(itemAnalisResultados)
                        ws.cell(row = cont, column = 192).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 192).font = Font(size = "8", name='Barlow') 

                    if str(itemAnalisResultados) == "Otro":
                        ws.cell(row = cont, column = 193).value = str(itemAnalisResultados)
                        ws.cell(row = cont, column = 193).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 193).font = Font(size = "8", name='Barlow')  

                for indexMedioDifusion, itemMedioDifusion in enumerate(listaMedioDifusion):

                    if str(itemMedioDifusion) == "Página web":
                        ws.cell(row = cont, column = 195).value = str(itemMedioDifusion)
                        ws.cell(row = cont, column = 195).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 195).font = Font(size = "8", name='Barlow') 
                    
                    if str(itemMedioDifusion) == "Medio físico":
                        ws.cell(row = cont, column = 196).value = str(itemMedioDifusion)
                        ws.cell(row = cont, column = 196).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 196).font = Font(size = "8", name='Barlow') 

                    if str(itemMedioDifusion) == "Medio electrónico":
                        ws.cell(row = cont, column = 197).value = str(itemMedioDifusion)
                        ws.cell(row = cont, column = 197).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 197).font = Font(size = "8", name='Barlow') 

                    if str(itemMedioDifusion) == "Otro":
                        ws.cell(row = cont, column = 198).value = str(itemMedioDifusion)
                        ws.cell(row = cont, column = 198).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 198).font = Font(size = "8", name='Barlow')  
                
                
                ws.cell(row = cont, column = 200).value = str(ooee.res_est_url)
                ws.cell(row = cont, column = 200).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 200).font = Font(size = "8", name='Barlow')  

                ws.cell(row = cont, column = 201).value = str(ooee.dispo_desde)
                ws.cell(row = cont, column = 201).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 201).font = Font(size = "8", name='Barlow')  

                ws.cell(row = cont, column = 202).value = str(ooee.dispo_hasta)
                ws.cell(row = cont, column = 202).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 202).font = Font(size = "8", name='Barlow')  
                
            for indexFechaPublicacion, itemFechaPublicacion in enumerate(listaFechaPublicacion):  
                
                if str(itemFechaPublicacion) == "fecha Publicación":
                    ws.cell(row = cont, column = 203).value = str(itemFechaPublicacion)
                    ws.cell(row = cont, column = 203).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 203).font = Font(size = "8", name='Barlow')

                if str(itemFechaPublicacion) == "No sabe":
                    ws.cell(row = cont, column = 205).value = str(itemFechaPublicacion)
                    ws.cell(row = cont, column = 205).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 205).font = Font(size = "8", name='Barlow')

                if str(itemFechaPublicacion) == "No hay":
                    ws.cell(row = cont, column = 206).value = str(itemFechaPublicacion)
                    ws.cell(row = cont, column = 206).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 206).font = Font(size = "8", name='Barlow')

            for indexFrecuenciaDifusion, itemFrecuenciaDifusion in enumerate(listaFrecuenciaDifusion):
                
                if str(itemFrecuenciaDifusion) == "Anual":
                    ws.cell(row = cont, column = 208).value = str(itemFrecuenciaDifusion)
                    ws.cell(row = cont, column = 208).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 208).font = Font(size = "8", name='Barlow')

                if str(itemFrecuenciaDifusion) == "Semestral":
                    ws.cell(row = cont, column = 209).value = str(itemFrecuenciaDifusion)
                    ws.cell(row = cont, column = 209).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 209).font = Font(size = "8", name='Barlow')

                if str(itemFrecuenciaDifusion) == "Trimestral":
                    ws.cell(row = cont, column = 210).value = str(itemFrecuenciaDifusion)
                    ws.cell(row = cont, column = 210).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 210).font = Font(size = "8", name='Barlow')

                if str(itemFrecuenciaDifusion) == "Mensual":
                    ws.cell(row = cont, column = 211).value = str(itemFrecuenciaDifusion)
                    ws.cell(row = cont, column = 211).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 211).font = Font(size = "8", name='Barlow')

                if str(itemFrecuenciaDifusion) == "Otro (s)":
                    ws.cell(row = cont, column = 212).value = str(itemFrecuenciaDifusion)
                    ws.cell(row = cont, column = 212).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 212).font = Font(size = "8", name='Barlow')

                if str(itemFrecuenciaDifusion) == "No está definido":
                    ws.cell(row = cont, column = 213).value = str(itemFrecuenciaDifusion)
                    ws.cell(row = cont, column = 213).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 213).font = Font(size = "8", name='Barlow')

            for indexProductosDifundir, itemProductosDifundir in enumerate(listaProductosDifundir):

                if str(itemProductosDifundir) == "Cuadros de salida (tablas)":
                    ws.cell(row = cont, column = 215).value = str(itemProductosDifundir)
                    ws.cell(row = cont, column = 215).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 215).font = Font(size = "8", name='Barlow')
                
                if str(itemProductosDifundir) == "Boletín estadístico":
                    ws.cell(row = cont, column = 216).value = str(itemProductosDifundir)
                    ws.cell(row = cont, column = 216).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 216).font = Font(size = "8", name='Barlow')

                if str(itemProductosDifundir) == "Anuario":
                    ws.cell(row = cont, column = 217).value = str(itemProductosDifundir)
                    ws.cell(row = cont, column = 217).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 217).font = Font(size = "8", name='Barlow')

                if str(itemProductosDifundir) == "Mapas estadísticos":
                    ws.cell(row = cont, column = 218).value = str(itemProductosDifundir)
                    ws.cell(row = cont, column = 218).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 218).font = Font(size = "8", name='Barlow')

                if str(itemProductosDifundir) == "Bases de datos interactivas":
                    ws.cell(row = cont, column = 219).value = str(itemProductosDifundir)
                    ws.cell(row = cont, column = 219).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 219).font = Font(size = "8", name='Barlow')

                if str(itemProductosDifundir) == "Otro (s)":
                    ws.cell(row = cont, column = 220).value = str(itemProductosDifundir)
                    ws.cell(row = cont, column = 220).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 220).font = Font(size = "8", name='Barlow')

            for indexOtrosProductos, itemOtrosProductos in enumerate(listaOtrosProductos):

                if str(itemOtrosProductos) == "Series históricas":
                    ws.cell(row = cont, column = 222).value = str(itemOtrosProductos)
                    ws.cell(row = cont, column = 222).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 222).font = Font(size = "8", name='Barlow')

                if str(itemOtrosProductos) == "Microdatos":
                    ws.cell(row = cont, column = 225).value = str(itemOtrosProductos)
                    ws.cell(row = cont, column = 225).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 225).font = Font(size = "8", name='Barlow')
                
                if str(itemOtrosProductos) == "Documentos metodológicos":
                    ws.cell(row = cont, column = 228).value = str(itemOtrosProductos)
                    ws.cell(row = cont, column = 228).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 228).font = Font(size = "8", name='Barlow')
                
                if str(itemOtrosProductos) == "Calendario de difusión":
                    ws.cell(row = cont, column = 230).value = str(itemOtrosProductos)
                    ws.cell(row = cont, column = 230).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 230).font = Font(size = "8", name='Barlow')
                    
                if str(itemOtrosProductos) == "Ninguna":
                    ws.cell(row = cont, column = 231).value = str(itemOtrosProductos)
                    ws.cell(row = cont, column = 231).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 231).font = Font(size = "8", name='Barlow')

            if str(ooee.conoce_otra) == "True":
                ws.cell(row = cont, column = 232).value = "Si"
                ws.cell(row = cont, column = 232).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 232).font = Font(size = "8", name='Barlow')
            elif str(ooee.conoce_otra) == "False":
                ws.cell(row = cont, column = 232).value = "No"
                ws.cell(row = cont, column = 232).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 232).font = Font(size = "8", name='Barlow')

            if str(ooee.hp_siste_infor) == "True":
                ws.cell(row = cont, column = 235).value = "Si"
                ws.cell(row = cont, column = 235).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 235).font = Font(size = "8", name='Barlow')
            elif str(ooee.hp_siste_infor) == "False":
                ws.cell(row = cont, column = 235).value = "No"
                ws.cell(row = cont, column = 235).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 235).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 237).value = ooee.observaciones
            ws.cell(row = cont, column = 237).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 237).font = Font(size = "8", name='Barlow') 

            ws.cell(row = cont, column = 238).value = str(ooee.anexos)
            ws.cell(row = cont, column = 238).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 238).font = Font(size = "8", name='Barlow')    
            

            ws.cell(row = cont, column = 239).value = str(ooee.nombre_est)
            ws.cell(row = cont, column = 239).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 239).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 240).value = str(ooee.estado_oe_tematico)
            ws.cell(row = cont, column = 240).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 240).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 241).value = str(ooee.validacion_oe_tematico)
            ws.cell(row = cont, column = 241).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 241).font = Font(size = "8", name='Barlow')
            ## SI desea continuar despues de la linea 241 ------------------------- 


        
            cont+=1
        #print("---------------------->", normaTextList)
        for indexNorma, itemNorma in enumerate(normaTextList):
           
            indexNorma =  itemNorma.ooee_id + 3
            ws.cell(row = indexNorma, column = 23).value = str(itemNorma.cp_d)
            ws.cell(row = indexNorma, column = 23).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 23).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexNorma, column = 24).value = str(itemNorma.ley_d)
            ws.cell(row = indexNorma, column = 24).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 24).font = Font(size = "8", name='Barlow')
                    
            ws.cell(row = indexNorma, column = 25).value = str(itemNorma.decreto_d)
            ws.cell(row = indexNorma, column = 25).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 25).font = Font(size = "8", name='Barlow')
                    
            ws.cell(row = indexNorma, column = 26).value = str(itemNorma.otra_d)
            ws.cell(row = indexNorma, column = 26).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 26).font = Font(size = "8", name='Barlow')


        for indexRequerimiento, itemRequerimiento in enumerate(requerimientoTextList):
 
            indexRequerimiento = itemRequerimiento.ooee_id + 3

            ws.cell(row = indexRequerimiento, column = 28).value = str(itemRequerimiento.ri_ods)
            ws.cell(row = indexRequerimiento, column = 28).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexRequerimiento, column = 28).font = Font(size = "8", name='Barlow')
        
            ws.cell(row = indexRequerimiento, column = 29).value = str(itemRequerimiento.ri_ocde)
            ws.cell(row = indexRequerimiento, column = 29).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexRequerimiento, column = 29).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexRequerimiento, column = 30).value = str(itemRequerimiento.ri_ci)
            ws.cell(row = indexRequerimiento, column = 30).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexRequerimiento, column = 30).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexRequerimiento, column = 31).value = str(itemRequerimiento.ri_pnd)
            ws.cell(row = indexRequerimiento, column = 31).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexRequerimiento, column = 31).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexRequerimiento, column = 32).value = str(itemRequerimiento.ri_cem)
            ws.cell(row = indexRequerimiento, column = 32).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  
            ws.cell(row = indexRequerimiento, column = 32).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexRequerimiento, column = 33).value = str(itemRequerimiento.ri_pstc)
            ws.cell(row = indexRequerimiento, column = 33).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexRequerimiento, column = 33).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexRequerimiento, column = 34).value = str(itemRequerimiento.ri_otro)
            ws.cell(row = indexRequerimiento, column = 34).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexRequerimiento, column = 34).font = Font(size = "8", name='Barlow')
           
        for indexPriUser, itemPriUser in enumerate(usuariosPrinTextList):

            indexPriUser = itemPriUser.ooee_id + 3

            ws.cell(row = indexPriUser, column = 36).value = str(itemPriUser.org_int)
            ws.cell(row = indexPriUser, column = 36).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 36).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexPriUser, column = 37).value = str(itemPriUser.pres_rep)
            ws.cell(row = indexPriUser, column = 37).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 37).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexPriUser, column = 38).value = str(itemPriUser.misnit)
            ws.cell(row = indexPriUser, column = 38).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 38).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexPriUser, column = 39).value = str(itemPriUser.org_cont)
            ws.cell(row = indexPriUser, column = 39).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 39).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexPriUser, column = 40).value = str(itemPriUser.o_ent_o_nac)
            ws.cell(row = indexPriUser, column = 40).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 40).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexPriUser, column = 41).value = str(itemPriUser.ent_o_terr)
            ws.cell(row = indexPriUser, column = 41).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 41).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexPriUser, column = 42).value = str(itemPriUser.gremios)
            ws.cell(row = indexPriUser, column = 42).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 42).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexPriUser, column = 43).value = str(itemPriUser.ent_privadas)
            ws.cell(row = indexPriUser, column = 43).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 43).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexPriUser, column = 44).value = str(itemPriUser.dep_misma_entidad)
            ws.cell(row = indexPriUser, column = 44).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 44).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexPriUser, column = 45).value = str(itemPriUser.academia)
            ws.cell(row = indexPriUser, column = 45).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 45).font = Font(size = "8", name='Barlow')
            
        for indexUniObs, itemUniObs in enumerate(unidadObserTextList):
      
            indexUniObs = itemUniObs.ooee_id + 3
                
            ws.cell(row = indexUniObs, column = 55).value = str(itemUniObs.mc_otra)
            ws.cell(row = indexUniObs, column = 55).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexUniObs, column = 55).font = Font(size = "8", name='Barlow')
            
        for indexObtencionDato, itemObtencionDato in enumerate(obtencionDatoTextList):

            indexObtencionDato = itemObtencionDato.ooee_id + 3
            ## si la opcion marcada es Registro Administrativo
            ws.cell(row = indexObtencionDato, column = 63).value = str(itemObtencionDato.mc_ra_cual)
            ws.cell(row = indexObtencionDato, column = 63).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexObtencionDato, column = 63).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexObtencionDato, column = 64).value = str(itemObtencionDato.mc_ra_entidad)
            ws.cell(row = indexObtencionDato, column = 64).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexObtencionDato, column = 64).font = Font(size = "8", name='Barlow')
                        
            ## si la opcion marcada es operacion estadística
            ws.cell(row = indexObtencionDato, column = 67).value = str(itemObtencionDato.mc_oe_cual)
            ws.cell(row = indexObtencionDato, column = 67).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexObtencionDato, column = 67).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexObtencionDato, column = 68).value = str(itemObtencionDato.mc_oe_entidad)
            ws.cell(row = indexObtencionDato, column = 68).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexObtencionDato, column = 68).font = Font(size = "8", name='Barlow')
                

        for indexMuesProbabilistico, itemMuesProbabilistico  in enumerate(muestProbaTextList):

            indexMuesProbabilistico =  itemMuesProbabilistico.ooee_id + 3

            ws.cell(row = indexMuesProbabilistico, column = 74).value = str(itemMuesProbabilistico.prob_otro)
            ws.cell(row = indexMuesProbabilistico, column = 74).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexMuesProbabilistico, column = 74).font = Font(size = "8", name='Barlow')

        for indexMuesNoProbabilistico, itemMuesNoProbabilistico in enumerate(muestNoProbaTextList):

            indexMuesNoProbabilistico = itemMuesNoProbabilistico.ooee_id + 3
                
            ws.cell(row = indexMuesNoProbabilistico, column = 80).value = str(itemMuesNoProbabilistico.no_prob_otro)
            ws.cell(row = indexMuesNoProbabilistico, column = 80).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexMuesNoProbabilistico, column = 80).font = Font(size = "8", name='Barlow')

        for indexTipMar, itemTipMar in enumerate(tipoMarcoTextList):

            indexTipMar = itemTipMar.ooee_id + 3
                
            ws.cell(row = indexTipMar, column = 86).value = str(itemTipMar.otro_tipo_marco)
            ws.cell(row = indexTipMar, column = 86).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexTipMar, column = 86).font = Font(size = "8", name='Barlow')

        for indexdocsDesa, itemdocsDesa in enumerate(docsDesarrolloTextList):

            indexdocsDesa = itemTipMar.ooee_id + 3
                
            ws.cell(row = indexdocsDesa, column = 93).value = str(itemdocsDesa.otro_docs)
            ws.cell(row = indexdocsDesa, column = 93).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdocsDesa, column = 93).font = Font(size = "8", name='Barlow')
    
        for indexConceptosEstanda, itemConceptosEstanda in enumerate(conceptosEstandaTextList):
            
            indexConceptosEstanda = itemConceptosEstanda.ooee_id + 3

            ws.cell(row = indexConceptosEstanda, column = 97).value = str(itemConceptosEstanda.org_in_cuales)
            ws.cell(row = indexConceptosEstanda, column = 97).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConceptosEstanda, column = 97).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexConceptosEstanda, column = 99).value = str(itemConceptosEstanda.ent_ordnac_cuales)
            ws.cell(row = indexConceptosEstanda, column = 99).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConceptosEstanda, column = 99).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexConceptosEstanda, column = 101).value = str(itemConceptosEstanda.leye_dec_cuales)
            ws.cell(row = indexConceptosEstanda, column = 101).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConceptosEstanda, column = 101).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexConceptosEstanda, column = 104).value = str(itemConceptosEstanda.otra_cual_conp)
            ws.cell(row = indexConceptosEstanda, column = 104).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConceptosEstanda, column = 104).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexConceptosEstanda, column = 106).value = str(itemConceptosEstanda.ningu_pq)
            ws.cell(row = indexConceptosEstanda, column = 106).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConceptosEstanda, column = 106).font = Font(size = "8", name='Barlow')
            
        for indexClasificaciones, itemClasificaciones in enumerate(clasificacionesTextList):

            indexClasificaciones = itemClasificaciones.ooee_id + 3

            ws.cell(row = indexClasificaciones, column = 109).value = str(itemClasificaciones.otra_cual_clas)
            ws.cell(row = indexClasificaciones, column = 109).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexClasificaciones, column = 109).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexClasificaciones, column = 110).value = str(itemClasificaciones.no_pq)
            ws.cell(row = indexClasificaciones, column = 110).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexClasificaciones, column = 110).font = Font(size = "8", name='Barlow')
        
        for indexCoberturaGeog, itemCoberturaGeog in enumerate(coberGeogTextList):

            indexCoberturaGeog = itemCoberturaGeog.ooee_id + 3
                
            ws.cell(row = indexCoberturaGeog, column = 113).value = str(itemCoberturaGeog.tot_regional)
            ws.cell(row = indexCoberturaGeog, column = 113).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 113).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 114).value = str(itemCoberturaGeog.cual_regional)
            ws.cell(row = indexCoberturaGeog, column = 114).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 114).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 116).value = str(itemCoberturaGeog.tot_dep)
            ws.cell(row = indexCoberturaGeog, column = 116).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 116).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 117).value = str(itemCoberturaGeog.cual_dep)
            ws.cell(row = indexCoberturaGeog, column = 117).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 117).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 119).value = str(itemCoberturaGeog.tot_are_metr)
            ws.cell(row = indexCoberturaGeog, column = 119).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 119).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 120).value = str(itemCoberturaGeog.cual_are_metr)
            ws.cell(row = indexCoberturaGeog, column = 120).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 120).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 122).value = str(itemCoberturaGeog.tot_mun)
            ws.cell(row = indexCoberturaGeog, column = 122).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 122).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 123).value = str(itemCoberturaGeog.cual_mun)
            ws.cell(row = indexCoberturaGeog, column = 123).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 123).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 125).value = str(itemCoberturaGeog.tot_otro)
            ws.cell(row = indexCoberturaGeog, column = 125).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 125).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 126).value = str(itemCoberturaGeog.cual_otro)
            ws.cell(row = indexCoberturaGeog, column = 126).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 126).font = Font(size = "8", name='Barlow')

        for indexdesagreInfo, itemdesagreInfo in enumerate(desagreInfoTextList):

            indexdesagreInfo = itemdesagreInfo.ooee_id + 3
            ws.cell(row = indexdesagreInfo, column = 129).value = str(itemdesagreInfo.des_tot_regional)
            ws.cell(row = indexdesagreInfo, column = 129).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 129).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexdesagreInfo, column = 130).value = str(itemdesagreInfo.des_cual_regional)
            ws.cell(row = indexdesagreInfo, column = 130).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 130).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexdesagreInfo, column = 132).value = str(itemdesagreInfo.des_tot_dep)
            ws.cell(row = indexdesagreInfo, column = 132).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 132).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 133).value = str(itemdesagreInfo.des_cual_dep)
            ws.cell(row = indexdesagreInfo, column = 133).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 133).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 135).value = str(itemdesagreInfo.des_tot_are_metr)
            ws.cell(row = indexdesagreInfo, column = 135).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 135).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 136).value = str(itemdesagreInfo.des_cual_are_metr)
            ws.cell(row = indexdesagreInfo, column = 136).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 136).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 138).value = str(itemdesagreInfo.des_tot_mun)
            ws.cell(row = indexdesagreInfo, column = 138).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 138).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 139).value = str(itemdesagreInfo.des_cual_mun)
            ws.cell(row = indexdesagreInfo, column = 139).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 139).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 141).value = str(itemdesagreInfo.des_tot_otro)
            ws.cell(row = indexdesagreInfo, column = 141).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 141).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 142).value = str(itemdesagreInfo.des_cual_otro)
            ws.cell(row = indexdesagreInfo, column = 142).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 142).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 152).value = str(itemdesagreInfo.des_grupo_otro)
            ws.cell(row = indexdesagreInfo, column = 152).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 152).font = Font(size = "8", name='Barlow')
            
        for indexFuenteFinan, itemFuenteFinan in enumerate(fuenteFinanTextList):

            indexFuenteFinan = itemFuenteFinan.ooee_id + 3

            ws.cell(row = indexFuenteFinan, column = 160).value = str(itemFuenteFinan.r_otros)
            ws.cell(row = indexFuenteFinan, column = 160).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexFuenteFinan, column = 160).font = Font(size = "8", name='Barlow')
            
        for indexMedioDatos, itemMedioDatos in enumerate(medioDatosTextList):

            indexMedioDatos = itemMedioDatos.ooee_id + 3

            ws.cell(row = indexMedioDatos, column = 167).value = str(itemMedioDatos.sis_info)
            ws.cell(row = indexMedioDatos, column = 167).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexMedioDatos, column = 167).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexMedioDatos, column = 172).value = str(itemMedioDatos.md_otro)
            ws.cell(row = indexMedioDatos, column = 172).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexMedioDatos, column = 172).font = Font(size = "8", name='Barlow')
            
        for indexPeriodicidadOE, itemPeriodicidadOE in enumerate(periodicidadTextList):

            indexPeriodicidadOE = itemPeriodicidadOE.ooee_id + 3

            ws.cell(row = indexMedioDatos, column = 179).value = str(itemPeriodicidadOE.per_otro)
            ws.cell(row = indexMedioDatos, column = 179).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexMedioDatos, column = 179).font = Font(size = "8", name='Barlow')
            
        for indexHerraProcesami, itemHerraProcesami in enumerate(herraProcesamiTextList):

            indexHerraProcesami = itemHerraProcesami.ooee_id + 3

            ws.cell(row = indexHerraProcesami, column = 188).value = str(itemHerraProcesami.herr_otro)
            ws.cell(row = indexHerraProcesami, column = 188).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexHerraProcesami, column = 188).font = Font(size = "8", name='Barlow')
            
        for indexAnalResul, itemAnalResul in enumerate(analResultTextList):
            
            indexAnalResul = itemAnalResul.ooee_id + 3

            ws.cell(row = indexAnalResul, column = 194).value = str(itemAnalResul.ana_otro)
            ws.cell(row = indexAnalResul, column = 194).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexAnalResul, column = 194).font = Font(size = "8", name='Barlow')
            
        for indexMedioDifus, itemMedioDifus in enumerate(medioDifusTextList):
            
            indexMedioDifus = itemMedioDifus.ooee_id + 3

            ws.cell(row = indexMedioDifus, column = 199).value = str(itemMedioDifus.dif_otro)
            ws.cell(row = indexMedioDifus, column = 199).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexMedioDifus, column = 199).font = Font(size = "8", name='Barlow')
            
        for indexFechaPublic, itemFechaPublic in enumerate(fechaPublicTextList):

            indexFechaPublic = itemFechaPublic.ooee_id + 3

            ws.cell(row = indexFechaPublic, column = 204).value = str(itemFechaPublic.fecha)
            ws.cell(row = indexFechaPublic, column = 204).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexFechaPublic, column = 204).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexFechaPublic, column = 207).value = str(itemFechaPublic.no_hay)
            ws.cell(row = indexFechaPublic, column = 207).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexFechaPublic, column = 207).font = Font(size = "8", name='Barlow')
            
        for indexFrecuencDifus, itemFrecuencDifus in enumerate(frecuenciaDifusionTextList):

            indexFrecuencDifus = itemFrecuencDifus.ooee_id + 3

            ws.cell(row = indexFrecuencDifus, column = 214).value = str(itemFrecuencDifus.no_definido)
            ws.cell(row = indexFrecuencDifus, column = 214).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexFrecuencDifus, column = 214).font = Font(size = "8", name='Barlow')

        for indexProductDifund, itemProductDifund in enumerate(productosDifundirTextList):

            indexProductDifund = itemProductDifund.ooee_id + 3

            ws.cell(row = indexProductDifund, column = 221).value = str(itemProductDifund.difundir_otro)
            ws.cell(row = indexProductDifund, column = 221).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexProductDifund, column = 221).font = Font(size = "8", name='Barlow')
        
        for indexOtrosProduct, itemOtrosProduct in enumerate(otrosProductosTextList):

            indexOtrosProduct = itemOtrosProduct.ooee_id + 3

            ws.cell(row = indexOtrosProduct, column = 223).value = str(itemOtrosProduct.ser_hist_desde)
            ws.cell(row = indexOtrosProduct, column = 223).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexOtrosProduct, column = 223).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexOtrosProduct, column = 224).value = str(itemOtrosProduct.ser_hist_hasta)
            ws.cell(row = indexOtrosProduct, column = 224).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexOtrosProduct, column = 224).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexOtrosProduct, column = 226).value = str(itemOtrosProduct.microdatos_desde)
            ws.cell(row = indexOtrosProduct, column = 226).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexOtrosProduct, column = 226).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexOtrosProduct, column = 227).value = str(itemOtrosProduct.microdatos_hasta)
            ws.cell(row = indexOtrosProduct, column = 227).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexOtrosProduct, column = 227).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexOtrosProduct, column = 229).value = str(itemOtrosProduct.op_url)
            ws.cell(row = indexOtrosProduct, column = 229).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexOtrosProduct, column = 229).font = Font(size = "8", name='Barlow')

        for indexResultSimi, itemResultSimi in enumerate(resultadosSimiTextList):
            
            indexResultSimi = itemResultSimi.ooee_id + 3

            ws.cell(row = indexResultSimi, column = 233).value = str(itemResultSimi.rs_entidad)
            ws.cell(row = indexResultSimi, column = 233).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexResultSimi, column = 233).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexResultSimi, column = 234).value = str(itemResultSimi.rs_oe)
            ws.cell(row = indexResultSimi, column = 234).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexResultSimi, column = 234).font = Font(size = "8", name='Barlow')
            
        for indexhpSistemaInfo, itemhpSistemaInfo in enumerate(hpSistemaInfoTextList):
            
            indexhpSistemaInfo = itemhpSistemaInfo.ooee_id + 3

            ws.cell(row = indexhpSistemaInfo, column = 236).value = str(itemhpSistemaInfo.si_cual)
            ws.cell(row = indexhpSistemaInfo, column = 236).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexhpSistemaInfo, column = 236).font = Font(size = "8", name='Barlow')
            
        for row in range(4, ws.max_row + 1):
            ws.row_dimensions[row].height = 90
            
        #############  Hoja 2 entidadades y fases  ##################

        sheet2.merge_cells('A2:B2')
        sheet2.merge_cells('C2:Z2')
        sheet2.merge_cells('A1:HZ1')

        def set_border(sheet2, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet2[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet2,'A1:HZ'+str(ooees.count()+3))
    
        #row dimensions
        sheet2.row_dimensions[1].height = 55
        sheet2.row_dimensions[2].height = 40

        # column width
        sheet2.column_dimensions['A'].width = 20
        sheet2.column_dimensions['B'].width = 20
        sheet2.column_dimensions['C'].width = 20
        sheet2.column_dimensions['D'].width = 20
        sheet2.column_dimensions['E'].width = 20
        sheet2.column_dimensions['F'].width = 20
        sheet2.column_dimensions['G'].width = 20
        sheet2.column_dimensions['H'].width = 20
        sheet2.column_dimensions['I'].width = 20
        sheet2.column_dimensions['J'].width = 20
        sheet2.column_dimensions['K'].width = 20
        sheet2.column_dimensions['L'].width = 20
        sheet2.column_dimensions['M'].width = 20
        sheet2.column_dimensions['N'].width = 20
        sheet2.column_dimensions['O'].width = 20
        sheet2.column_dimensions['P'].width = 20
        sheet2.column_dimensions['Q'].width = 20
        sheet2.column_dimensions['R'].width = 20
        sheet2.column_dimensions['S'].width = 20
        sheet2.column_dimensions['T'].width = 20
        sheet2.column_dimensions['U'].width = 20
        sheet2.column_dimensions['V'].width = 20
        sheet2.column_dimensions['W'].width = 20
        sheet2.column_dimensions['X'].width = 20
        sheet2.column_dimensions['Y'].width = 20
        sheet2.column_dimensions['Z'].width = 20
    
        title_cell = sheet2['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de las Operaciones Estadísticas'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')
	
        codigo_cell = sheet2['A2']
        codigo_cell.value = 'A. OPERACIÓN ESTADÍSTICA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet2['C2']
        codigo_cell.value = '¿Otra entidad es responsable de una o varias fases de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet2['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = sheet2['B3']
        codigo_cell.value = 'Nombre Operación Estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet2['C3']
        codigo_cell.value = 'Indique si hay otra(s) entidad(es) que participa(n) en alguna de las fases del proceso estadístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        contSheet2 = 4
        for ooee in ooees: 
            sheet2.cell(row = contSheet2, column = 1).value = ooee.codigo_oe
            sheet2.cell(row = contSheet2, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet2.cell(row = contSheet2, column = 1).font = Font(size = "8", name='Barlow')

            sheet2.cell(row = contSheet2, column = 2).value = ooee.nombre_oe
            sheet2.cell(row = contSheet2, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet2.cell(row = contSheet2, column = 2).font = Font(size = "8", name='Barlow')
            
            if str(ooee.fase) == "True":

                sheet2.cell(row = contSheet2, column = 3).value = "Si"
                sheet2.cell(row = contSheet2, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet2.cell(row = contSheet2, column = 3).font = Font(size = "8", name='Barlow')
            
            elif  str(ooee.fase) == "False":
            
                sheet2.cell(row = contSheet2, column = 3).value = "No"
                sheet2.cell(row = contSheet2, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet2.cell(row = contSheet2, column = 3).font = Font(size = "8", name='Barlow')
            
            contSheet2+= 1
        
        entidadoeid = []
        for indexEnFas, itemEnfas in enumerate(entidadesFasesTextList):
            
            indexEnFas = itemEnfas.ooee_id + 3
            entidadoeid.append(itemEnfas.ooee_id)
            incrementoEntitie = entidadoeid.count(itemEnfas.ooee_id) + 3
            count_id = entidadoeid.count(itemEnfas.ooee_id)
            c=0 #inicializamos el contador  
            n=6*count_id 
            for i in range(1,n+1):  
                if i%6 == 0:  
                    i = i - 2
                    c+=1
            
            sheet2.cell(row = 3, column = i ).value = "Nombre de Entidad"
            sheet2.cell(row = 3, column = i ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet2.cell(row = 3, column = i ).font = Font(bold=True) 

            sheet2.cell(row = indexEnFas, column = i ).value = str(itemEnfas.nombre_entifas) 
            sheet2.cell(row = indexEnFas, column = i ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet2.cell(row = indexEnFas, column = i ).font = Font(size = "8", name='Barlow') 
            
            listaFases = list(itemEnfas.fases.all())  ##iterar para traer las fases seleccionadas

            for indexFase, itemFase in enumerate(listaFases):
                indiceFase = indexFase 
            
                if str(itemFase) == "Detección y análisis de requerimientos":
                    countFase=0 #inicializamos el contador  
                    faseSelected=6*count_id 
                    for posFase in range(1,n+1):  
                        if posFase%6 == 0:  
                            posFase = posFase - 1
                            countFase+=1

                    sheet2.cell(row = 3, column = posFase ).value = "a. Detección y análisis de requerimientos"
                    sheet2.cell(row = 3, column = posFase ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = 3, column = posFase ).font = Font(bold=True)

                    sheet2.cell(row = indexEnFas, column = posFase).value = str(itemFase)
                    sheet2.cell(row = indexEnFas, column = posFase).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = indexEnFas, column = posFase).font = Font(size = "8", name='Barlow')
                                
                if str(itemFase) == "Diseño y pruebas":

                    countFase=0 #inicializamos el contador  
                    faseSelected=6*count_id 
                    for posFase in range(1,n+1):  
                        if posFase%6 == 0:  
                            posFase = posFase
                            countFase+=1

                    sheet2.cell(row = 3, column = posFase ).value = "b. Diseño y pruebas"
                    sheet2.cell(row = 3, column = posFase ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = 3, column = posFase ).font = Font(bold=True)

                    sheet2.cell(row = indexEnFas, column = posFase).value = str(itemFase)
                    sheet2.cell(row = indexEnFas, column = posFase).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = indexEnFas, column = posFase).font = Font(size = "8", name='Barlow')

                if str(itemFase) == "Ejecución":

                    countFase=0 #inicializamos el contador  
                    faseSelected=6*count_id 
                    for posFase in range(1,n+1):  
                        if posFase%6 == 0:  
                            posFase = posFase + 1
                            countFase+=1

                    sheet2.cell(row = 3, column = posFase ).value = "c. Ejecución"
                    sheet2.cell(row = 3, column = posFase ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = 3, column = posFase ).font = Font(bold=True)

                    sheet2.cell(row = indexEnFas, column = posFase).value = str(itemFase)
                    sheet2.cell(row = indexEnFas, column = posFase).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = indexEnFas, column = posFase).font = Font(size = "8", name='Barlow')
                
                if str(itemFase) == "Análisis":

                    countFase=0 #inicializamos el contador  
                    faseSelected=6*count_id 
                    for posFase in range(1,n+1):  
                        if posFase%6 == 0:  
                            posFase = posFase + 2
                            countFase+=1

                    sheet2.cell(row = 3, column = posFase ).value = "d. Análisis"
                    sheet2.cell(row = 3, column = posFase ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = 3, column = posFase ).font = Font(bold=True)

                    sheet2.cell(row = indexEnFas, column = posFase).value = str(itemFase)
                    sheet2.cell(row = indexEnFas, column = posFase).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = indexEnFas, column = posFase).font = Font(size = "8", name='Barlow')
                
                if str(itemFase) == "Difusión":

                    countFase=0 #inicializamos el contador  
                    faseSelected=6*count_id 
                    for posFase in range(1,n+1):  
                        if posFase%6 == 0:  
                            posFase = posFase + 3
                            countFase+=1

                    sheet2.cell(row = 3, column = posFase ).value = "e. Difusión"
                    sheet2.cell(row = 3, column = posFase ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = 3, column = posFase ).font = Font(bold=True)
                    
                    sheet2.cell(row = indexEnFas, column = posFase).value = str(itemFase)
                    sheet2.cell(row = indexEnFas, column = posFase).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = indexEnFas, column = posFase).font = Font(size = "8", name='Barlow')
                               
        for row in range(4, sheet2.max_row + 1):
            sheet2.row_dimensions[row].height = 90   
        
                    
        #############  Hoja 3 lista de variables  ##################

        sheet3.merge_cells('A2:B2')
        sheet3.merge_cells('C2:Z3')
        sheet3.merge_cells('A1:HZ1')

        def set_border(sheet3, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet3[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet3,'A1:HZ'+str(ooees.count()+3))
    
        #row dimensions
        sheet3.row_dimensions[1].height = 55
        sheet3.row_dimensions[2].height = 40

        # column width
        sheet3.column_dimensions['A'].width = 20
        sheet3.column_dimensions['B'].width = 20
        
        title_cell = sheet3['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de las Operaciones Estadísticas'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')
	
        codigo_cell = sheet3['A2']
        codigo_cell.value = 'A. OPERACIÓN ESTADÍSTICA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = sheet3['C2']
        codigo_cell.value = 'Liste todas las variables que maneja la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')


        codigo_cell = sheet3['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        
        codigo_cell = sheet3['B3']
        codigo_cell.value = 'Nombre Operación Estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        contsheet3 = 4

        for ooee in ooees: 
            sheet3.cell(row = contsheet3, column = 1).value = ooee.codigo_oe
            sheet3.cell(row = contsheet3, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet3.cell(row = contsheet3, column = 1).font = Font(size = "8", name='Barlow')

            sheet3.cell(row = contsheet3, column = 2).value = ooee.nombre_oe
            sheet3.cell(row = contsheet3, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet3.cell(row = contsheet3, column = 2).font = Font(size = "8", name='Barlow')
            contsheet3+= 1
    
        arrayoeId = []
        for indexVariab, itemVariab in enumerate(listaDeVariablesText):
            indexVariab =  itemVariab.ooee_id + 3
            arrayoeId.append(itemVariab.ooee_id)
            
            incrementoVa = arrayoeId.count(itemVariab.ooee_id) + 2
            sheet3.cell(row = indexVariab, column = incrementoVa ).value = str(itemVariab.lista_var)
            sheet3.cell(row = indexVariab, column = incrementoVa).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet3.cell(row = indexVariab, column = incrementoVa).font = Font(size = "8", name='Barlow')
            
        for row in range(4, sheet3.max_row + 1):  ## definir tamaño de rows
            sheet3.row_dimensions[row].height = 90

        #############  Hoja 4 resultados  ##################

        sheet4.merge_cells('A2:B2')
        sheet4.merge_cells('C2:Z3')
        sheet4.merge_cells('A1:HZ1')

        def set_border(sheet4, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet4[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet4,'A1:HZ'+str(ooees.count()+3))
    
        #row dimensions
        sheet4.row_dimensions[1].height = 55
        sheet4.row_dimensions[2].height = 40

        # column width
        sheet4.column_dimensions['A'].width = 20
        sheet4.column_dimensions['B'].width = 20
    
        title_cell = sheet4['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de las Operaciones Estadísticas'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')
	
        codigo_cell = sheet4['A2']
        codigo_cell.value = 'A. OPERACIÓN ESTADÍSTICA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        resultados_cell = sheet4['C2']
        resultados_cell.value = '¿Cuáles son los resultados agregados o indicadores calculados?'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        codigo_cell = sheet5['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        
        codigo_cell = sheet4['B3']
        codigo_cell.value = 'Nombre Operación Estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        contsheet4 = 4

        for ooee in ooees: 
            sheet4.cell(row = contsheet4, column = 1).value = ooee.codigo_oe
            sheet4.cell(row = contsheet4, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet4.cell(row = contsheet4, column = 1).font = Font(size = "8", name='Barlow') 

            sheet4.cell(row = contsheet4, column = 2).value = ooee.nombre_oe
            sheet4.cell(row = contsheet4, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet4.cell(row = contsheet4, column = 2).font = Font(size = "8", name='Barlow') 
            contsheet4+=1

        listaIdoe = []

        for indexResultEsta, itemResultEsta in enumerate(listaResultEstText):
            indexResultEsta =  itemResultEsta.ooee_id + 3
            listaIdoe.append(itemResultEsta.ooee_id)
            
            incrementoRe = listaIdoe.count(itemResultEsta.ooee_id) + 2
            sheet4.cell(row = indexResultEsta, column = incrementoRe).value = str(itemResultEsta.resultEstad)
            sheet4.cell(row = indexResultEsta, column = incrementoRe).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet4.cell(row = indexResultEsta, column = incrementoRe).font = Font(size = "8", name='Barlow')

        for row in range(4, sheet4.max_row + 1):  ## definir tamaño de rows
            sheet4.row_dimensions[row].height = 90

        ############# END Hoja 4 resultados  ##################

        ############## Hoja 5 Evaluación de calidad #############

        
        sheet5.merge_cells('A1:AZ1')
        sheet5.merge_cells('A2:F2')
        sheet5.merge_cells('G2:AZ2')

        def set_border(sheet5, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet5[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet5,'A1:AZ'+str(ooees.count()+3))
        

        #row dimensions
        sheet5.row_dimensions[1].height = 55
        sheet5.row_dimensions[2].height = 40
        sheet5.row_dimensions[3].height = 40

        # column width
        sheet5.column_dimensions['A'].width = 20
        sheet5.column_dimensions['B'].width = 18
        sheet5.column_dimensions['C'].width = 15
        sheet5.column_dimensions['D'].width = 24
        sheet5.column_dimensions['E'].width = 15
        sheet5.column_dimensions['F'].width = 24
        sheet5.column_dimensions['G'].width = 22
        sheet5.column_dimensions['H'].width = 15
        sheet5.column_dimensions['I'].width = 15
        sheet5.column_dimensions['J'].width = 15
        sheet5.column_dimensions['K'].width = 15
        sheet5.column_dimensions['L'].width = 15
        sheet5.column_dimensions['M'].width = 15
        sheet5.column_dimensions['N'].width = 15
        sheet5.column_dimensions['O'].width = 15
        sheet5.column_dimensions['P'].width = 15
        sheet5.column_dimensions['Q'].width = 15
        sheet5.column_dimensions['R'].width = 22
        sheet5.column_dimensions['S'].width = 15
        sheet5.column_dimensions['T'].width = 15
        sheet5.column_dimensions['U'].width = 15
        sheet5.column_dimensions['V'].width = 15
        sheet5.column_dimensions['W'].width = 15
        sheet5.column_dimensions['X'].width = 15
        sheet5.column_dimensions['Y'].width = 20
        sheet5.column_dimensions['Z'].width = 15

        sheet5.column_dimensions['AA'].width = 15
        sheet5.column_dimensions['AB'].width = 15
        sheet5.column_dimensions['AC'].width = 15
        sheet5.column_dimensions['AD'].width = 15
        sheet5.column_dimensions['AE'].width = 15
        sheet5.column_dimensions['AF'].width = 15
        sheet5.column_dimensions['AG'].width = 15
        sheet5.column_dimensions['AH'].width = 15
        sheet5.column_dimensions['AI'].width = 15
        sheet5.column_dimensions['AJ'].width = 15
        sheet5.column_dimensions['AK'].width = 15
        sheet5.column_dimensions['AL'].width = 15
        sheet5.column_dimensions['AM'].width = 15
        sheet5.column_dimensions['AN'].width = 15
        sheet5.column_dimensions['AO'].width = 15
        sheet5.column_dimensions['AP'].width = 15
        sheet5.column_dimensions['AQ'].width = 15
        sheet5.column_dimensions['AR'].width = 15
        sheet5.column_dimensions['AS'].width = 15
        sheet5.column_dimensions['AT'].width = 15
        sheet5.column_dimensions['AU'].width = 15
        sheet5.column_dimensions['AV'].width = 15
        sheet5.column_dimensions['AW'].width = 15
        sheet5.column_dimensions['AX'].width = 15
        sheet5.column_dimensions['AY'].width = 15
        sheet5.column_dimensions['AZ'].width = 15

        title_cell = sheet5['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de las Operaciones Estadísticas'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet5['A2']
        codigo_cell.value = 'IDENTIFICACIÓN'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet5['G2']
        codigo_cell.value = 'EVALUACIÓN DE CALIDAD'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['A3']
        resultados_cell.value = 'Área Temática'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['B3']
        resultados_cell.value = 'Tema'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['C3']
        resultados_cell.value = 'Código Entidad'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['D3']
        resultados_cell.value = 'Nombre Entidad'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['E3']
        resultados_cell.value = 'Código OOEE'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['F3']
        resultados_cell.value = 'Nombre de la Operación Estadística'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        resultados_cell = sheet5['G3']
        resultados_cell.value = 'Número de Evaluaciones'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        contsheet5 = 4

        for ooee in ooees: 
            sheet5.cell(row = contsheet5, column = 1).value = str(ooee.area_tematica)
            sheet5.cell(row = contsheet5, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 1).font = Font(size = "8", name='Barlow') 
            
            sheet5.cell(row = contsheet5, column = 2).value = str(ooee.tema)
            sheet5.cell(row = contsheet5, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 2).font = Font(size = "8", name='Barlow') 
            
            sheet5.cell(row = contsheet5, column = 3).value = ooee.entidad.codigo
            sheet5.cell(row = contsheet5, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 3).font = Font(size = "8", name='Barlow')

            sheet5.cell(row = contsheet5, column = 4).value = str(ooee.entidad)
            sheet5.cell(row = contsheet5, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 4).font = Font(size = "8", name='Barlow')

            sheet5.cell(row = contsheet5, column = 5).value = ooee.codigo_oe
            sheet5.cell(row = contsheet5, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 5).font = Font(size = "8", name='Barlow')

            sheet5.cell(row = contsheet5, column = 6).value = ooee.nombre_oe
            sheet5.cell(row = contsheet5, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 6).font = Font(size = "8", name='Barlow') 
            contsheet5+=1

        idoeEval = []
        for indexEval, itemEval in enumerate(listaEvaluacionText):
            indexEval =  itemEval.post_oe_id + 3
            idoeEval.append(itemEval.post_oe_id)
            
            ### 1 Estado de la Certificación
            
            count_est_cert = idoeEval.count(itemEval.post_oe_id)
            #print("total evaluaciones",count_est_cert)
            contrField1=0 #inicializamos el contador  
            pos_est_cert = 10*count_est_cert
            for incField1 in range(1,pos_est_cert+1):  
                if incField1%10 == 0:  
                    incField1 = incField1 - 2
                    contrField1+=1

            sheet5.cell(row = indexEval, column = 7).value = count_est_cert
            sheet5.cell(row = indexEval, column = 7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = indexEval, column = 7).font = Font(bold=True) 

            sheet5.cell(row = 3, column = incField1).value = "Estado de la Evaluación"
            sheet5.cell(row = 3, column = incField1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = 3, column = incField1).font = Font(bold=True) 

            sheet5.cell(row = indexEval, column = incField1).value = str(itemEval.est_evaluacion)
            sheet5.cell(row = indexEval, column = incField1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = indexEval, column = incField1).font = Font(size = "8", name='Barlow')
  
            ### 2 Resultado año de vigencia

            count_id_oe = idoeEval.count(itemEval.post_oe_id)
            contadorEval=0 #inicializamos el contador  
            posFields = 10*count_id_oe
            for incEval in range(1,posFields+1):  
                if incEval%10 == 0:  
                    incEval = incEval - 1
                    contadorEval+=1

            sheet5.cell(row = 3, column = incEval).value = "Año"
            sheet5.cell(row = 3, column = incEval ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = 3, column = incEval ).font = Font(bold=True) 

            if itemEval.year_eva == None:

                sheet5.cell(row = indexEval, column = incEval).value = ""
                sheet5.cell(row = indexEval, column = incEval).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet5.cell(row = indexEval, column = incEval).font = Font(size = "8", name='Barlow')
            else:
                sheet5.cell(row = indexEval, column = incEval).value = itemEval.year_eva.strftime('%Y')
                sheet5.cell(row = indexEval, column = incEval).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet5.cell(row = indexEval, column = incEval).font = Font(size = "8", name='Barlow')
            
            ### 3 Observaciones de la evaluación

            contadorField3=0 #inicializamos el contador  
            posFields3 = 10*count_id_oe
            for incField3 in range(1,posFields3+1):  
                if incField3%10 == 0:  
                    incField3 = incField3
                    contadorField3+=1

            sheet5.cell(row = 3, column = incField3).value = "Observaciones"
            sheet5.cell(row = 3, column = incField3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = 3, column = incField3).font = Font(bold=True)

            sheet5.cell(row = indexEval, column = incField3).value = str(itemEval.observ_est)
            sheet5.cell(row = indexEval, column = incField3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = indexEval, column = incField3).font = Font(size = "8", name='Barlow')

            ### 4 Metodologia

            contadorField4=0 #inicializamos el contador  
            posFields4 = 10*count_id_oe
            for incField4 in range(1,posFields4+1):  
                if incField4%10 == 0:
                    incField4 = incField4 + 1
                    contadorField4+=1

            sheet5.cell(row = 3, column = incField4).value = "Metodología"
            sheet5.cell(row = 3, column = incField4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = 3, column = incField4).font = Font(bold=True)

            sheet5.cell(row = indexEval, column = incField4).value = str(itemEval.metodologia)
            sheet5.cell(row = indexEval, column = incField4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = indexEval, column = incField4).font = Font(size = "8", name='Barlow')
            
        ### 5 Resultado de la evaluación según la metodologia

            contadorField5=0 #inicializamos el contador  
            posFields5 = 10*count_id_oe
            for incField5 in range(1,posFields5+1):  
                if incField5%10 == 0:  
                    incField5 = incField5 + 2
                    contadorField5+=1

            sheet5.cell(row = 3, column = incField5).value = "Resultado de la evaluación"
            sheet5.cell(row = 3, column = incField5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = 3, column = incField5).font = Font(bold=True)

            if str(itemEval.metodologia) == "ntcpe 1000":

                sheet5.cell(row = indexEval, column = incField5).value = str(itemEval.res_evaluacion)
                sheet5.cell(row = indexEval, column = incField5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet5.cell(row = indexEval, column = incField5).font = Font(size = "8", name='Barlow')

            elif str(itemEval.metodologia) == "matriz de requisitos":
                sheet5.cell(row = indexEval, column = incField5).value = str(itemEval.res_mzrequi)
                sheet5.cell(row = indexEval, column = incField5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet5.cell(row = indexEval, column = incField5).font = Font(size = "8", name='Barlow')
        
        
        ### 6 Observaciones de resultado de la evaluación

            contadorField6=0 #inicializamos el contador  
            posFields6 = 10*count_id_oe
            for incField6 in range(1,posFields6+1):  
                if incField6%10 == 0:  
                    incField6 = incField6 + 3
                    contadorField6+=1

            sheet5.cell(row = 3, column = incField6).value = "Observaciones"
            sheet5.cell(row = 3, column = incField6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = 3, column = incField6).font = Font(bold=True)

            sheet5.cell(row = indexEval, column = incField6).value = str(itemEval.observ_resul)
            sheet5.cell(row = indexEval, column = incField6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = indexEval, column = incField6).font = Font(size = "8", name='Barlow')

        ### 7 Vigencia

            contadorField7=0 #inicializamos el contador  
            posFields7 = 10*count_id_oe
            for incField7 in range(1,posFields7+1):  
                if incField7%10 == 0:  
                    incField7 = incField7 + 4
                    contadorField7+=1

            sheet5.cell(row = 3, column = incField7).value = "Vigencia"
            sheet5.cell(row = 3, column = incField7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = 3, column = incField7).font = Font(bold=True)

            sheet5.cell(row = indexEval, column = incField7).value = "Desde: " + str(itemEval.vigencia_desde) + " " + "\n\n Hasta: " + str(itemEval.vigencia_hasta)
            sheet5.cell(row = indexEval, column = incField7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = indexEval, column = incField7).font = Font(size = "8", name='Barlow')

         ### 8 Plan de Mejoramiento   
            
            contadorField8=0 #inicializamos el contador  
            posFields8 = 10*count_id_oe
            for incField8 in range(1,posFields8+1):  
                if incField8%10 == 0:  
                    incField8 = incField8 + 5
                    contadorField8+=1

            sheet5.cell(row = 3, column = incField8).value = "Plan de mejoramiento"
            sheet5.cell(row = 3, column = incField8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = 3, column = incField8).font = Font(bold=True)

            sheet5.cell(row = indexEval, column = incField8).value = str(itemEval.pla_mejoramiento)
            sheet5.cell(row = indexEval, column = incField8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = indexEval, column = incField8).font = Font(size = "8", name='Barlow')

        ###  9 Seguimiento anual (Vigilancia):   
            
            contadorField8=0 #inicializamos el contador  
            posFields8 = 10*count_id_oe
            for incField8 in range(1,posFields8+1):  
                if incField8%10 == 0:  
                    incField8 = incField8 + 6
                    contadorField8+=1

            sheet5.cell(row = 3, column = incField8).value = "Seguimiento anual (Vigilancia)"
            sheet5.cell(row = 3, column = incField8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = 3, column = incField8).font = Font(bold=True)

            sheet5.cell(row = indexEval, column = incField8).value = str(itemEval.seg_vig)
            sheet5.cell(row = indexEval, column = incField8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = indexEval, column = incField8).font = Font(size = "8", name='Barlow')

        ###  Observaciones seguimiento anual
            
            contadorField9=0 #inicializamos el contador  
            posFields9 = 10*count_id_oe
            for incField9 in range(1,posFields9+1):  
                if incField9%10 == 0:  
                    incField9 = incField9 + 7
                    contadorField9+=1

            sheet5.cell(row = 3, column = incField9).value = "Observaciones"
            sheet5.cell(row = 3, column = incField9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = 3, column = incField9).font = Font(bold=True)

            sheet5.cell(row = indexEval, column = incField9).value = str(itemEval.obs_seg_anual)
            sheet5.cell(row = indexEval, column = incField9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = indexEval, column = incField9).font = Font(size = "8", name='Barlow')
   
        for row in range(4, sheet5.max_row + 1):  ## definir tamaño de rows
            sheet5.row_dimensions[row].height = 90

        ############## End Evaluación de calidad ############# 

        #### hoja 6 critica #####

        sheet6.merge_cells('A1:F1')
        sheet6.merge_cells('A2:F2')

        def set_border(sheet6, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            rows = sheet6[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet6,'A1:F1')
            

        #row dimensions
        sheet6.row_dimensions[1].height = 55
        sheet6.row_dimensions[2].height = 40
        sheet6.row_dimensions[3].height = 40
        sheet6.row_dimensions[4].height = 40
        sheet6.row_dimensions[5].height = 40
        sheet6.row_dimensions[6].height = 40

        # column width
        sheet6.column_dimensions['A'].width = 20
        sheet6.column_dimensions['B'].width = 20
        sheet6.column_dimensions['C'].width = 20
        sheet6.column_dimensions['D'].width = 20
        sheet6.column_dimensions['E'].width = 20
        sheet6.column_dimensions['F'].width = 20

        title_cell = sheet6['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de las Operaciones Estadísticas'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet6['A2']
        codigo_cell.value = 'Críticas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')


        resultados_cell = sheet6['A3']
        resultados_cell.value = 'Código'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        resultados_cell = sheet6['B3']
        resultados_cell.value = 'Nombre de la Operación Estadística'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        resultados_cell = sheet6['C3']
        resultados_cell.value = 'Estado de la crítica'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        resultados_cell = sheet6['D3']
        resultados_cell.value = 'Observaciones de la crítica'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        resultados_cell = sheet6['E3']
        resultados_cell.value = 'Funcionario que realiza la crítica'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        resultados_cell = sheet6['F3']
        resultados_cell.value = 'Fecha en que se realiza la crítica'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        idoeCritica = []
        contsheet6 = 4
        for indexCritica, itemCritica in enumerate(listaCriticaText):
            indexCritica =  itemCritica.post_oe_id + 3
            idoeCritica.append(itemCritica.post_oe_id)
            for oest in ooees:
                if  str(itemCritica.post_oe) == str(oest.nombre_oe):

                    sheet6.cell(row = contsheet6, column = 1).value = oest.codigo_oe
                    sheet6.cell(row = contsheet6, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet6.cell(row = contsheet6, column = 1).font = Font(bold=True) 

                    sheet6.cell(row = contsheet6, column = 2).value = str(itemCritica.post_oe)
                    sheet6.cell(row = contsheet6, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet6.cell(row = contsheet6, column = 2).font = Font(bold=True) 

                    sheet6.cell(row = contsheet6, column = 3).value = str(itemCritica.estado_crit)
                    sheet6.cell(row = contsheet6, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet6.cell(row = contsheet6, column = 3).font = Font(bold=True)

                    sheet6.cell(row = contsheet6, column = 4).value = str(itemCritica.descrip_critica)
                    sheet6.cell(row = contsheet6, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet6.cell(row = contsheet6, column = 4).font = Font(bold=True)

                    sheet6.cell(row = contsheet6, column = 5).value = str(itemCritica.name_cri)
                    sheet6.cell(row = contsheet6, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet6.cell(row = contsheet6, column = 5).font = Font(bold=True)

                    sheet6.cell(row = contsheet6, column = 6).value = str(itemCritica.fecha_critica)
                    sheet6.cell(row = contsheet6, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet6.cell(row = contsheet6, column = 6).font = Font(bold=True)

                    contsheet6+=1

            for row in range(4, sheet6.max_row + 1):  ## definir tamaño de rows
                sheet6.row_dimensions[row].height = 90

        ############## End critica #############

        ############## Hoja 7 Novedad #############
        sheet7.merge_cells('A1:G1')
        sheet7.merge_cells('A2:G2')

        def set_border(sheet7, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet7[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet7,'A1:G1')
            

        #row dimensions
        sheet7.row_dimensions[1].height = 55
        sheet7.row_dimensions[2].height = 40
        sheet7.row_dimensions[3].height = 40
        sheet7.row_dimensions[4].height = 40
        sheet7.row_dimensions[5].height = 40
        sheet7.row_dimensions[6].height = 40
        sheet7.row_dimensions[7].height = 40

        # column width
        sheet7.column_dimensions['A'].width = 20
        sheet7.column_dimensions['B'].width = 20
        sheet7.column_dimensions['C'].width = 20
        sheet7.column_dimensions['D'].width = 20
        sheet7.column_dimensions['E'].width = 20
        sheet7.column_dimensions['F'].width = 20
        sheet7.column_dimensions['G'].width = 20

        title_cell = sheet7['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de las Operaciones Estadísticas'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet7['A2']
        codigo_cell.value = 'Novedades'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')


        resultados_cell = sheet7['A3']
        resultados_cell.value = 'Código'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        resultados_cell = sheet7['B3']
        resultados_cell.value = 'Nombre de la Operación Estadística'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        resultados_cell = sheet7['C3']
        resultados_cell.value = 'Novedad'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        resultados_cell = sheet7['D3']
        resultados_cell.value = 'Estado de la actualización'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        resultados_cell = sheet7['E3']
        resultados_cell.value = 'Descripción de la novedad'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        resultados_cell = sheet7['F3']
        resultados_cell.value = 'Funcionario que realiza la novedad'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        resultados_cell = sheet7['G3']
        resultados_cell.value = 'Fecha en que se realiza la novedad'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        idoeNovedad = []
        contsheet7 = 4
        for indexNovedad, itemNovedad in enumerate(listaNovedadText):
            indexNovedad =  itemNovedad.post_oe_id + 3
            idoeNovedad.append(itemNovedad.post_oe_id)
            for oest in ooees:
                if  str(itemNovedad.post_oe) == str(oest.nombre_oe):

                    sheet7.cell(row = contsheet7, column = 1).value = oest.codigo_oe
                    sheet7.cell(row = contsheet7, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet7.cell(row = contsheet7, column = 1).font = Font(bold=True) 

                    sheet7.cell(row = contsheet7, column = 2).value = str(itemNovedad.post_oe)
                    sheet7.cell(row = contsheet7, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet7.cell(row = contsheet7, column = 2).font = Font(bold=True) 

                    sheet7.cell(row = contsheet7, column = 3).value = str(itemNovedad.novedad)
                    sheet7.cell(row = contsheet7, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet7.cell(row = contsheet7, column = 3).font = Font(bold=True)

                    sheet7.cell(row = contsheet7, column = 4).value = str(itemNovedad.est_actualiz)
                    sheet7.cell(row = contsheet7, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet7.cell(row = contsheet7, column = 4).font = Font(bold=True)

                    sheet7.cell(row = contsheet7, column = 5).value = str(itemNovedad.descrip_novedad)
                    sheet7.cell(row = contsheet7, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet7.cell(row = contsheet7, column = 5).font = Font(bold=True)

                    sheet7.cell(row = contsheet7, column = 6).value = str(itemNovedad.name_nov)
                    sheet7.cell(row = contsheet7, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet7.cell(row = contsheet7, column = 6).font = Font(bold=True)

                    sheet7.cell(row = contsheet7, column = 7).value = str(itemNovedad.fecha_actualiz)
                    sheet7.cell(row = contsheet7, column = 7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet7.cell(row = contsheet7, column = 7).font = Font(bold=True)

                    contsheet7+=1

            for row in range(4, sheet7.max_row + 1):  ## definir tamaño de rows
                sheet7.row_dimensions[row].height = 90

        ############## End Novedad #############

        file_name = "reporte_ooee.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(file_name)
        response['Content-Disposition'] = content
        wb.save(response)
        return response 

########*********** Reporte por filtros de operaciones estadisticas ##############**************

class reportOOEEfilter_xls(TemplateView):
    def get(self, request, *args, **kwargs):
                     
        id_entidad = request.GET.get('entidad')
        id_area_tematica = request.GET.get('area_tematica')
        id_tema = request.GET.get('tema')
        id_fase = request.GET.get('nombre_est')

        user = request.user
        entidad_cod = Entidades_oe.objects.all()
        
        ## opc 1 es diferente de vacia
        if id_entidad != "" and id_area_tematica == "" and  id_tema == "":
            
            ooees = OperacionEstadistica.objects.filter(nombre_est=5).filter(entidad=id_entidad)

        ## opc 1 y opc 2 es diferente de vacia
        elif  id_entidad != "" and id_area_tematica != "" and  id_tema == "":
            
          ooees = OperacionEstadistica.objects.filter(nombre_est=5).filter(entidad=id_entidad).filter(area_tematica=id_area_tematica)

        ## opc 1, opc 2 y opc 3 es diferente de vacia
        elif id_entidad != "" and id_area_tematica != "" and  id_tema != "":
            
            ooees = OperacionEstadistica.objects.filter(nombre_est=5).filter(entidad=id_entidad).filter(area_tematica=id_area_tematica).filter(tema=id_tema)
        
        ## opcion 1 y opcion 3 es diferente de vacia
        elif id_entidad != "" and id_area_tematica == "" and  id_tema != "":
            
            ooees = OperacionEstadistica.objects.filter(nombre_est=5).filter(entidad=id_entidad).filter(tema=id_tema)
        
        ## opc 2 es diferente de vacia
        elif id_entidad == "" and id_area_tematica != "" and  id_tema == "":
            
            ooees = OperacionEstadistica.objects.filter(nombre_est=5).filter(area_tematica=id_area_tematica)
        
        ## opc 2 y opc 3 es diferente de vacia
        elif id_entidad == "" and id_area_tematica != "" and  id_tema != "":
            
            ooees = OperacionEstadistica.objects.filter(nombre_est=5).filter(area_tematica=id_area_tematica).filter(tema=id_tema)

        # opc 3 es diferente de vacia
        elif id_entidad == "" and id_area_tematica == "" and  id_tema != "":
            
            ooees = OperacionEstadistica.objects.filter(nombre_est=5).filter(tema=id_tema)

        entidadesFasesTextList = []
        normaTextList = []
        requerimientoTextList = []
        usuariosPrinTextList = []
        unidadObserTextList = []
        obtencionDatoTextList = []
        muestProbaTextList = []
        muestNoProbaTextList = []
        tipoMarcoTextList = []
        docsDesarrolloTextList = [] 
        conceptosEstandaTextList = []
        clasificacionesTextList = []
        coberGeogTextList = []
        desagreInfoTextList = []
        fuenteFinanTextList =[]
        listaDeVariablesText = []  # modulo c Pregunta 14
        listaResultEstText = [] # modulo c pregunta 15
        medioDatosTextList = []
        periodicidadTextList = []
        herraProcesamiTextList = []
        analResultTextList = []
        medioDifusTextList = []
        fechaPublicTextList = []
        frecuenciaDifusionTextList = [] 
        productosDifundirTextList = []
        otrosProductosTextList = []
        resultadosSimiTextList = []
        hpSistemaInfoTextList = []
        listaEvaluacionText = [] #Evaluación de calidad
        listaCriticaText = [] #Critica
        listaNovedadText = [] #Novedades

        objArray = list(ooees.values('id'))
        for obj in objArray:
            for key, value in obj.items():
                #print("val", value)
                
                entidadesFasesList = MB_EntidadFases.objects.filter(ooee_id=value)
                unioListEntidadFase = list(chain(entidadesFasesList))
                entidadesFasesTextList.extend(unioListEntidadFase)

                normaList = MB_Norma.objects.filter(ooee_id=value)
                unioListNorma = list(chain(normaList))
                normaTextList.extend(unioListNorma)

                requerimientosList = MB_Requerimientos.objects.filter(ooee_id=value)
                unioListRequerimiento = list(chain(requerimientosList))
                requerimientoTextList.extend(unioListRequerimiento)

                usuariosPrinList = MB_PrinUsuarios.objects.filter(ooee_id=value)
                unioListUsuariosPrin = list(chain(usuariosPrinList))
                usuariosPrinTextList.extend(unioListUsuariosPrin)

                unidadObserList = MC_UnidadObservacion.objects.filter(ooee_id=value)
                unioListUnidadObser = list(chain(unidadObserList))
                unidadObserTextList.extend(unioListUnidadObser)

                obtencionDatoList = MC_ObtencionDato.objects.filter(ooee_id=value)
                unioListObtencionDato = list(chain(obtencionDatoList))
                obtencionDatoTextList.extend(unioListObtencionDato)

                muestProbaList = MC_MuestreoProbabilistico.objects.filter(ooee_id=value)
                unioListmuestProba =  list(chain(muestProbaList))
                muestProbaTextList.extend(unioListmuestProba)

                muestNoProbaList = MC_MuestreoNoProbabilistico.objects.filter(ooee_id=value)
                unioListmuestNoProba = list(chain(muestNoProbaList))
                muestNoProbaTextList.extend(unioListmuestNoProba)

                tipoMarcoList = MC_TipoMarco.objects.filter(ooee_id=value)
                unioListTipoMarco = list(chain(tipoMarcoList))
                tipoMarcoTextList.extend(unioListTipoMarco)

                docsDesarrolloList = MC_DocsDesarrollo.objects.filter(ooee_id=value)
                unioListdocsDesarrollo = list(chain(docsDesarrolloList))
                docsDesarrolloTextList.extend(unioListdocsDesarrollo)  

                conceptosEstandaList = MC_ConceptosEstandarizados.objects.filter(ooee_id=value)
                unionListconceptosEstanda = list(chain(conceptosEstandaList))
                conceptosEstandaTextList.extend(unionListconceptosEstanda)  

                clasificacionesList = MC_Clasificaciones.objects.filter(ooee_id=value)
                unioListclasificaciones = list(chain(clasificacionesList))
                clasificacionesTextList.extend(unioListclasificaciones)  

                coberGeogList = MC_CoberturaGeografica.objects.filter(ooee_id=value)
                unioListCoberGeog = list(chain(coberGeogList))
                coberGeogTextList.extend(unioListCoberGeog)

                desagreInfoList = MC_DesagregacionInformacion.objects.filter(ooee_id=value)
                unioListDesagreInfo = list(chain(desagreInfoList))
                desagreInfoTextList.extend(unioListDesagreInfo)

                fuenteFinanList = MC_FuenteFinanciacion.objects.filter(ooee_id=value)
                unioListFuenteFinan = list(chain(fuenteFinanList))
                fuenteFinanTextList.extend(unioListFuenteFinan)

                listVariables = MC_listaVariable.objects.filter(ooee_id=value)  #modulo c pregunta 14
                unioListVariables = list(chain(listVariables))
                listaDeVariablesText.extend(unioListVariables)

                listaResultEst = MC_ResultadoEstadistico.objects.filter(ooee_id=value)  #modulo c pregunta 15
                unioListResultEst = list(chain(listaResultEst))
                listaResultEstText.extend(unioListResultEst)

                medioDatosList = MD_MedioDatos.objects.filter(ooee_id=value)
                unioListMedioDatos = list(chain(medioDatosList))
                medioDatosTextList.extend(unioListMedioDatos)

                periodicidadList = MD_PeriodicidadOe.objects.filter(ooee_id=value)
                unioListPeriodicidad = list(chain(periodicidadList))
                periodicidadTextList.extend(unioListPeriodicidad)

                herraProcesamiList = MD_HerramProcesamiento.objects.filter(ooee_id=value)
                unioListHerraProcesami = list(chain(herraProcesamiList))
                herraProcesamiTextList.extend(unioListHerraProcesami)

                analResultList = ME_AnalisisResultados.objects.filter(ooee_id=value)
                unioListAnalResult = list(chain(analResultList))
                analResultTextList.extend(unioListAnalResult)

                medioDifusList = MF_MediosDifusion.objects.filter(ooee_id=value)
                unioListMedioDifus = list(chain(medioDifusList))
                medioDifusTextList.extend(unioListMedioDifus)

                fechaPublicList = MF_FechaPublicacion.objects.filter(ooee_id=value)
                unioListFechaPublic = list(chain(fechaPublicList))
                fechaPublicTextList.extend(unioListFechaPublic)

                frecuenciaDifusionList = MF_FrecuenciaDifusion.objects.filter(ooee_id=value)
                unioListFrecuenciaDifusion = list(chain(frecuenciaDifusionList))
                frecuenciaDifusionTextList.extend(unioListFrecuenciaDifusion)

                productosDifundirList = MF_ProductosDifundir.objects.filter(ooee_id=value)
                unioListProductosDifundir  = list(chain(productosDifundirList))
                productosDifundirTextList.extend(unioListProductosDifundir)

                otrosProductosList = MF_OtrosProductos.objects.filter(ooee_id=value)
                unioListOtrosProductos = list(chain(otrosProductosList))
                otrosProductosTextList.extend(unioListOtrosProductos)

                resultadosSimiList = MF_ResultadosSimilares.objects.filter(ooee_id=value)
                unioListResultadosSimi = list(chain(resultadosSimiList))
                resultadosSimiTextList.extend(unioListResultadosSimi)

                hpSistemaInfoList = MF_HPSistemaInfo.objects.filter(ooee_id=value)
                unioListhpSistemaInfo = list(chain(hpSistemaInfoList))
                hpSistemaInfoTextList.extend(unioListhpSistemaInfo)

                listEvaluacion = EvaluacionCalidad.objects.filter(post_oe_id=value) # evaluacion de calidad
                unioListEvaluacion = list(chain(listEvaluacion))
                listaEvaluacionText.extend(unioListEvaluacion)
                
                listCritica = Critica.objects.filter(post_oe_id=value) #critica de la oe
                unioListCritica = list(chain(listCritica))
                listaCriticaText.extend(unioListCritica)

                listNovedad = NovedadActualizacion.objects.filter(post_oe_id=value) #Novedad de la oe
                unioListNovedad = list(chain(listNovedad))
                listaNovedadText.extend(unioListNovedad)

        wb = Workbook()
        ws = wb.active
        ws.title = "Directorio OOEE"
        
        sheet2 = wb.create_sheet('Entidades Fases')
        sheet3 = wb.create_sheet('Variables')
        sheet4 = wb.create_sheet('Resultados')
        sheet5 = wb.create_sheet('Eval Calidad')
        sheet6 = wb.create_sheet('Critica OE')
        sheet7 = wb.create_sheet('Novedad OE')

        def set_border(ws, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = ws[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(ws,'A1:IC'+str(ooees.count()+3))

		## size rows

        ws.row_dimensions[1].height = 55
        ws.row_dimensions[2].height = 40
        ws.row_dimensions[3].height = 80
       
		## size column
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 15
        ws.column_dimensions['O'].width = 15
        ws.column_dimensions['P'].width = 15
        ws.column_dimensions['Q'].width = 15
        ws.column_dimensions['R'].width = 22
        ws.column_dimensions['S'].width = 15
        ws.column_dimensions['T'].width = 15
        ws.column_dimensions['U'].width = 15
        ws.column_dimensions['V'].width = 15
        ws.column_dimensions['W'].width = 15
        ws.column_dimensions['X'].width = 15
        ws.column_dimensions['Y'].width = 20
        ws.column_dimensions['Z'].width = 15

        ws.column_dimensions['AA'].width = 15
        ws.column_dimensions['AB'].width = 15
        ws.column_dimensions['AC'].width = 15
        ws.column_dimensions['AD'].width = 15
        ws.column_dimensions['AE'].width = 15
        ws.column_dimensions['AF'].width = 15
        ws.column_dimensions['AG'].width = 15
        ws.column_dimensions['AH'].width = 15
        ws.column_dimensions['AI'].width = 15
        ws.column_dimensions['AJ'].width = 15
        ws.column_dimensions['AK'].width = 15
        ws.column_dimensions['AL'].width = 15
        ws.column_dimensions['AM'].width = 15
        ws.column_dimensions['AN'].width = 15
        ws.column_dimensions['AO'].width = 15
        ws.column_dimensions['AP'].width = 15
        ws.column_dimensions['AQ'].width = 15
        ws.column_dimensions['AR'].width = 15
        ws.column_dimensions['AS'].width = 15
        ws.column_dimensions['AT'].width = 15
        ws.column_dimensions['AU'].width = 15
        ws.column_dimensions['AV'].width = 15
        ws.column_dimensions['AW'].width = 15
        ws.column_dimensions['AX'].width = 15
        ws.column_dimensions['AY'].width = 15
        ws.column_dimensions['AZ'].width = 15

        ws.column_dimensions['BA'].width = 15
        ws.column_dimensions['BB'].width = 15
        ws.column_dimensions['BC'].width = 15
        ws.column_dimensions['BD'].width = 15
        ws.column_dimensions['BE'].width = 15
        ws.column_dimensions['BF'].width = 15
        ws.column_dimensions['BG'].width = 15
        ws.column_dimensions['BH'].width = 15
        ws.column_dimensions['BI'].width = 15
        ws.column_dimensions['BJ'].width = 15
        ws.column_dimensions['BK'].width = 15
        ws.column_dimensions['BL'].width = 15
        ws.column_dimensions['BM'].width = 15
        ws.column_dimensions['BN'].width = 15
        ws.column_dimensions['BO'].width = 15
        ws.column_dimensions['BP'].width = 15
        ws.column_dimensions['BQ'].width = 15
        ws.column_dimensions['BR'].width = 15
        ws.column_dimensions['BS'].width = 15
        ws.column_dimensions['BT'].width = 15
        ws.column_dimensions['BU'].width = 15
        ws.column_dimensions['BV'].width = 15
        ws.column_dimensions['BW'].width = 15
        ws.column_dimensions['BX'].width = 15
        ws.column_dimensions['BY'].width = 15
        ws.column_dimensions['BZ'].width = 15

        ws.column_dimensions['CA'].width = 15
        ws.column_dimensions['CB'].width = 15
        ws.column_dimensions['CC'].width = 15
        ws.column_dimensions['CD'].width = 15
        ws.column_dimensions['CE'].width = 15
        ws.column_dimensions['CF'].width = 15
        ws.column_dimensions['CG'].width = 15
        ws.column_dimensions['CH'].width = 15
        ws.column_dimensions['CI'].width = 15
        ws.column_dimensions['CJ'].width = 15
        ws.column_dimensions['CK'].width = 15
        ws.column_dimensions['CL'].width = 15
        ws.column_dimensions['CM'].width = 15
        ws.column_dimensions['CN'].width = 15
        ws.column_dimensions['CO'].width = 15
        ws.column_dimensions['CP'].width = 15
        ws.column_dimensions['CQ'].width = 15
        ws.column_dimensions['CR'].width = 15
        ws.column_dimensions['CS'].width = 15
        ws.column_dimensions['CT'].width = 15
        ws.column_dimensions['CU'].width = 15
        ws.column_dimensions['CV'].width = 15
        ws.column_dimensions['CW'].width = 15
        ws.column_dimensions['CX'].width = 15
        ws.column_dimensions['CY'].width = 15
        ws.column_dimensions['CZ'].width = 15

        ws.column_dimensions['DA'].width = 15
        ws.column_dimensions['DB'].width = 15
        ws.column_dimensions['DC'].width = 15
        ws.column_dimensions['DD'].width = 15
        ws.column_dimensions['DE'].width = 15
        ws.column_dimensions['DF'].width = 15
        ws.column_dimensions['DG'].width = 15
        ws.column_dimensions['DH'].width = 15
        ws.column_dimensions['DI'].width = 15
        ws.column_dimensions['DJ'].width = 15
        ws.column_dimensions['DK'].width = 15
        ws.column_dimensions['DL'].width = 15
        ws.column_dimensions['DM'].width = 15
        ws.column_dimensions['DN'].width = 15
        ws.column_dimensions['DO'].width = 15
        ws.column_dimensions['DP'].width = 15
        ws.column_dimensions['DQ'].width = 15
        ws.column_dimensions['DR'].width = 15
        ws.column_dimensions['DS'].width = 15
        ws.column_dimensions['DT'].width = 15
        ws.column_dimensions['DU'].width = 15
        ws.column_dimensions['DV'].width = 15
        ws.column_dimensions['DW'].width = 15
        ws.column_dimensions['DX'].width = 15
        ws.column_dimensions['DY'].width = 15
        ws.column_dimensions['DZ'].width = 15

        ws.column_dimensions['EA'].width = 15
        ws.column_dimensions['EB'].width = 15
        ws.column_dimensions['EC'].width = 15
        ws.column_dimensions['ED'].width = 15
        ws.column_dimensions['EE'].width = 15
        ws.column_dimensions['EF'].width = 15
        ws.column_dimensions['EG'].width = 15
        ws.column_dimensions['EH'].width = 15
        ws.column_dimensions['EI'].width = 15
        ws.column_dimensions['EJ'].width = 15
        ws.column_dimensions['EK'].width = 15
        ws.column_dimensions['EL'].width = 15
        ws.column_dimensions['EM'].width = 15
        ws.column_dimensions['EN'].width = 15
        ws.column_dimensions['EO'].width = 15
        ws.column_dimensions['EP'].width = 15
        ws.column_dimensions['EQ'].width = 15
        ws.column_dimensions['ER'].width = 15
        ws.column_dimensions['ES'].width = 15
        ws.column_dimensions['ET'].width = 15
        ws.column_dimensions['EU'].width = 15
        ws.column_dimensions['EV'].width = 15
        ws.column_dimensions['EW'].width = 15
        ws.column_dimensions['EX'].width = 15
        ws.column_dimensions['EY'].width = 15
        ws.column_dimensions['EZ'].width = 15

        ws.column_dimensions['FA'].width = 20
        ws.column_dimensions['FB'].width = 20
        ws.column_dimensions['FC'].width = 15
        ws.column_dimensions['FD'].width = 15
        ws.column_dimensions['FE'].width = 15
        ws.column_dimensions['FF'].width = 15
        ws.column_dimensions['FG'].width = 15
        ws.column_dimensions['FH'].width = 15
        ws.column_dimensions['FI'].width = 15
        ws.column_dimensions['FJ'].width = 15
        ws.column_dimensions['FK'].width = 15
        ws.column_dimensions['FL'].width = 15
        ws.column_dimensions['FM'].width = 15
        ws.column_dimensions['FN'].width = 15
        ws.column_dimensions['FO'].width = 15
        ws.column_dimensions['FP'].width = 15
        ws.column_dimensions['FQ'].width = 15
        ws.column_dimensions['FR'].width = 15
        ws.column_dimensions['FS'].width = 15
        ws.column_dimensions['FT'].width = 15
        ws.column_dimensions['FU'].width = 15
        ws.column_dimensions['FV'].width = 15
        ws.column_dimensions['FW'].width = 15
        ws.column_dimensions['FX'].width = 15
        ws.column_dimensions['FY'].width = 15
        ws.column_dimensions['FZ'].width = 15

        ws.column_dimensions['GA'].width = 20
        ws.column_dimensions['GB'].width = 20
        ws.column_dimensions['GC'].width = 40
        ws.column_dimensions['GD'].width = 15
        ws.column_dimensions['GE'].width = 15
        ws.column_dimensions['GF'].width = 15
        ws.column_dimensions['GG'].width = 15
        ws.column_dimensions['GH'].width = 15
        ws.column_dimensions['GI'].width = 15
        ws.column_dimensions['GJ'].width = 15
        ws.column_dimensions['GK'].width = 15
        ws.column_dimensions['GL'].width = 15
        ws.column_dimensions['GM'].width = 15
        ws.column_dimensions['GN'].width = 15
        ws.column_dimensions['GO'].width = 15
        ws.column_dimensions['GP'].width = 15
        ws.column_dimensions['GQ'].width = 15
        ws.column_dimensions['GR'].width = 15
        ws.column_dimensions['GS'].width = 15
        ws.column_dimensions['GT'].width = 15
        ws.column_dimensions['GU'].width = 15
        ws.column_dimensions['GV'].width = 15
        ws.column_dimensions['GW'].width = 15
        ws.column_dimensions['GX'].width = 15
        ws.column_dimensions['GY'].width = 15
        ws.column_dimensions['GZ'].width = 15

        ws.column_dimensions['HA'].width = 20
        ws.column_dimensions['HB'].width = 20
        ws.column_dimensions['HC'].width = 20
        ws.column_dimensions['HD'].width = 15
        ws.column_dimensions['HE'].width = 15
        ws.column_dimensions['HF'].width = 15
        ws.column_dimensions['HG'].width = 15
        ws.column_dimensions['HH'].width = 15
        ws.column_dimensions['HI'].width = 15
        ws.column_dimensions['HJ'].width = 15
        ws.column_dimensions['HK'].width = 15
        ws.column_dimensions['HL'].width = 15
        ws.column_dimensions['HM'].width = 15
        ws.column_dimensions['HN'].width = 15
        ws.column_dimensions['HO'].width = 15
        ws.column_dimensions['HP'].width = 15
        ws.column_dimensions['HQ'].width = 15
        ws.column_dimensions['HR'].width = 15
        ws.column_dimensions['HS'].width = 15
        ws.column_dimensions['HT'].width = 20
        ws.column_dimensions['HU'].width = 15
        ws.column_dimensions['HV'].width = 20
        ws.column_dimensions['HW'].width = 15
        ws.column_dimensions['HX'].width = 15
        ws.column_dimensions['HY'].width = 40
        ws.column_dimensions['HZ'].width = 15

        ws.column_dimensions['IA'].width = 15
        ws.column_dimensions['IB'].width = 15
        ws.column_dimensions['IC'].width = 25
        
		##styles
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
		
        title_cell = ws['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de las Operaciones Estadísticas'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')
    
        ws.merge_cells('A1:IC1')

        ws.merge_cells('A2:C2')
        ws.merge_cells('D2:E2')
        ws.merge_cells('F2:H2')
        ws.merge_cells('I2:Q2')
        ws.merge_cells('R2:R3')
        ws.merge_cells('S2:W2')
        ws.merge_cells('X2:AE2')
        ws.merge_cells('AF2:AP2')
        ws.merge_cells('AQ2:AQ3')
        ws.merge_cells('AR2:AY2')

        ws.merge_cells('AZ2:BD2')
        ws.merge_cells('BE2:BL2') 
        ws.merge_cells('BM2:BR2')
        ws.merge_cells('BS2:BX2') 
        ws.merge_cells('BY2:CD2')  
        ws.merge_cells('CE2:CL2')  
        ws.merge_cells('CM2:CX2')
        ws.merge_cells('CY2:DB2')
        ws.merge_cells('DC2:DR2')
        ws.merge_cells('DS2:EH2')
        ws.merge_cells('EI2:EK2')
        ws.merge_cells('EL2:ES2')
        ws.merge_cells('ET2:EU2')
        ws.merge_cells('EV2:EZ2')
        ws.merge_cells('FA2:FA3')
        ws.merge_cells('FB2:FB3')
        ws.merge_cells('FC2:FL2')
        ws.merge_cells('FM2:FS2')
        ws.merge_cells('FT2:GB2')
        ws.merge_cells('GC2:GC3')
        ws.merge_cells('GD2:GH2')
        ws.merge_cells('GI2:GM2')
        ws.merge_cells('GN2:GN3')
        ws.merge_cells('GO2:GP2')
        ws.merge_cells('GQ2:GU2')
        ws.merge_cells('GV2:HB2')
        ws.merge_cells('HC2:HI2')
        ws.merge_cells('HJ2:HS2')
        ws.merge_cells('HT2:HV2')
        ws.merge_cells('HW2:HX2')
        ws.merge_cells('HY2:HY3')
        ws.merge_cells('HZ2:HZ3')
        ws.merge_cells('IA2:IA3')
        ws.merge_cells('IB2:IB3')
        ws.merge_cells('IC2:IC3')

        ## insert heads groups
        codigo_cell = ws['A2']
        codigo_cell.value = 'OPERACIÓN ESTADÍSTICA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['D2']
        codigo_cell.value = 'ENTIDAD'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['F2']
        codigo_cell.value = 'ÁREA TEMÁTICA / TEMA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['I2']
        codigo_cell.value = ' A. IDENTIFICACIÓN '
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['R2']
        codigo_cell.value = 'Indique si hay otra(s) entidad(es) que participa(n) en alguna de las fases del proceso estadístico  (ver hoja 2)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['S2']
        codigo_cell.value = 'Bajo cuál(es) de las siguiente(s) norma(s), se soporta la producción de información de la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['X2']
        codigo_cell.value = 'La operación estadística satisface requerimientos de información de:'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = ws['AF2']
        codigo_cell.value = 'Señale cuáles son los principales usuarios  de la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AQ2']
        codigo_cell.value = '¿Cuál es la población objetivo de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AR2']
        codigo_cell.value = '¿Cuál es la unidad de observación de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AZ2']
        codigo_cell.value = '¿Cuál es el tipo de operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BE2']
        codigo_cell.value = 'Indique de donde se obtienen los datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BM2']
        codigo_cell.value = 'Muestreo probabilístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BS2']
        codigo_cell.value = 'Muestreo No probabilístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BY2']
        codigo_cell.value = ' ¿La operación estadística cuenta con un marco estadístico para identificar y ubicar las unidades de observación?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CE2']
        codigo_cell.value = 'Indique cuáles de los siguientes documentos se elaboran para el desarrollo de la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CM2']
        codigo_cell.value = 'Indique si la operación estadística utiliza conceptos estandarizados de'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CY2']
        codigo_cell.value = '¿La operación estadística utiliza nomenclaturas y/o clasificaciones?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DC2']
        codigo_cell.value = '¿Cuál es la cobertura geográfica de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DS2']
        codigo_cell.value = 'Desagregación geográfica '
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EI2']
        codigo_cell.value = 'Desagregación por zona'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EL2']
        codigo_cell.value = 'Desagregación por grupos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['ET2']
        codigo_cell.value = '¿Cuál es el costo anual de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EV2']
        codigo_cell.value = '¿Cuál(es) son las fuentes de financiación de la operación estadística ?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
    
        codigo_cell = ws['FA2']
        codigo_cell.value = 'Liste todas las variables que maneja la operación estadística (Ver hoja 3)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FB2']
        codigo_cell.value = '¿Cuáles son los resultados agregados o indicadores calculados? (Ver hoja 4)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = ws['FC2']
        codigo_cell.value = '¿Cuál es el medio de obtención de los datos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FM2']
        codigo_cell.value = '¿Cuál es la periodicidad de recolección o acopio de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['FT2']
        codigo_cell.value = 'Indique cuáles de las siguientes herramientas son utilizadas en el procesamiento de los datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = ws['GC2']
        codigo_cell.value = 'Haga una  breve descripción de la manera cómo se realiza el procesamiento de los datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GD2']
        codigo_cell.value = '¿Qué tipo de análisis realiza a los resultados obtenidos en la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GI2']
        codigo_cell.value = '¿A través de qué medio(s) difunde los resultados estadísticos (Agregados, indicadores), a los usuarios?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['GN2']
        codigo_cell.value = 'Indique la dirección de la página web donde se encuentran los resultados estadísticos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GO2']
        codigo_cell.value = 'Fechas de disponibilidad de los resultados estadísticos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GQ2']
        codigo_cell.value = '¿Cuál es la próxima fecha de publicación de resultados estadísticos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['GV2']
        codigo_cell.value = '¿Cuál es la frecuencia de difusión de los resultados estadísticos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HC2']
        codigo_cell.value = '¿Cuáles productos utiliza para difundir los resultados estadísticos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HJ2']
        codigo_cell.value = '¿Qué otros productos estadísticos de la OE están disponibles para consulta de los usuarios?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HT2']
        codigo_cell.value = '¿Conoce si otra entidad produce resultados similares a los de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
      
        codigo_cell = ws['HW2']
        codigo_cell.value = '¿La operación estadística hace parte de algún sistema de información?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HY2']
        codigo_cell.value = 'Observaciones'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['HZ2']
        codigo_cell.value = 'Anexos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['IA2']
        codigo_cell.value = 'Estado de la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['IB2']
        codigo_cell.value = 'Validación de la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['IC2']
        codigo_cell.value = 'Estado del proceso de la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ####### insert heads fields

        codigo_cell = ws['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['B3']
        codigo_cell.value = 'Nombre Operación Estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['C3']
        codigo_cell.value = 'Objetivo de la Operación Estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['D3']
        codigo_cell.value = 'Nombre de la Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center',  wrap_text=True)
       
        codigo_cell = ws['E3']
        codigo_cell.value = 'Código de la Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['F3']
        codigo_cell.value = 'Área Tematica'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = ws['G3']
        codigo_cell.value = 'Tema'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['H3']
        codigo_cell.value = 'Tema Compartido'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
      
        codigo_cell = ws['I3']
        codigo_cell.value = 'Nombre de la Dependencia'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['J3']
        codigo_cell.value = 'Nombre del Director'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['K3']
        codigo_cell.value = 'Cargo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['L3']
        codigo_cell.value = 'Correo Electrónico del Director'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = ws['M3']
        codigo_cell.value = 'Teléfono del director'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['N3']
        codigo_cell.value = 'Nombre del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['O3']
        codigo_cell.value = 'Cargo del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['P3']
        codigo_cell.value = 'Correo Electrónico del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['Q3']
        codigo_cell.value = 'Teléfono del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
         
        codigo_cell = ws['S3']
        codigo_cell.value = 'a. Constitución Política'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['T3']
        codigo_cell.value = 'b. Ley'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['U3']
        codigo_cell.value = 'c. Decreto (nacional, departamental, municipal)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['V3']
        codigo_cell.value = 'd. Otra (Resolución, ordenanza, acuerdo)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center',  wrap_text=True)

        codigo_cell = ws['W3']
        codigo_cell.value = 'e. Ninguna'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['X3']
        codigo_cell.value = 'a. Objetivos de desarrollo Sostenible (ODS)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['Y3']
        codigo_cell.value = 'b. Organización para la Cooperación y el Desarrollo Económico (OCDE)   '
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['Z3']
        codigo_cell.value = 'c. Otros compromisos internacionales'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['AA3']
        codigo_cell.value = 'd. Plan Nacional de Desarrollo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AB3']
        codigo_cell.value = 'e. Cuentas económicas y macroeconómicas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['AC3']
        codigo_cell.value = 'f. Plan Sectorial, Territorial o CONPES'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['AD3']
        codigo_cell.value = 'g. Otro (s)  ¿cuál(es) ?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AE3']
        codigo_cell.value = 'h. Ninguno'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        codigo_cell = ws['AF3']
        codigo_cell.value = 'a. Organismos internacionales'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AG3']
        codigo_cell.value = 'b. Presidencia de la República'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AH3']
        codigo_cell.value = 'c. Ministerios'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AI3']
        codigo_cell.value = 'd. Organismos de Control'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AJ3']
        codigo_cell.value = 'e. Otras entidades del orden Nacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AK3']
        codigo_cell.value = 'f. Entidades de orden Territorial'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AL3']
        codigo_cell.value = 'g. Gremios'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AM3']
        codigo_cell.value = 'h. Entidades privadas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AN3']
        codigo_cell.value = 'i. Dependencias de la misma entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AO3']
        codigo_cell.value = 'j. Academia'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AP3']
        codigo_cell.value = 'k. Público en General'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AR3']
        codigo_cell.value = 'a. Empresa'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        codigo_cell = ws['AS3']
        codigo_cell.value = 'b. Establecimiento'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AT3']
        codigo_cell.value = 'c. Hogar'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AU3']
        codigo_cell.value = 'd. Persona'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AV3']
        codigo_cell.value = 'e. Unidad productora agropecuaria'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AW3']
        codigo_cell.value = 'f. Predio'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AX3']
        codigo_cell.value = 'g. Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AY3']
        codigo_cell.value = 'Cuál(es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AZ3']
        codigo_cell.value = 'a. Aprovechamiento de registro administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BA3']
        codigo_cell.value = 'b. Estadística derivada'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BB3']
        codigo_cell.value = 'c. Muestreo probabilístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BC3']
        codigo_cell.value = 'd. Muestreo no probabilístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BD3']
        codigo_cell.value = 'e. Censo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BE3']
        codigo_cell.value = 'Registros Administrativos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BF3']
        codigo_cell.value = 'Listado de RRAA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BG3']
        codigo_cell.value = 'Cuál(es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BH3']
        codigo_cell.value = 'Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BI3']
        codigo_cell.value = 'Operación Estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BJ3']
        codigo_cell.value = 'Listado de OOEE'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BK3']
        codigo_cell.value = 'Cuál(es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BL3']
        codigo_cell.value =  'Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BM3']
        codigo_cell.value =  'Muestreo aleatorio simple'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BN3']
        codigo_cell.value =  'Muestreo aleatorio sistemático'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BO3']
        codigo_cell.value =  'Muestreo aleatorio estratificado'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BP3']
        codigo_cell.value =  'Muestreo aleatorio por conglomerados'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BQ3']
        codigo_cell.value =  'Otro'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BR3']
        codigo_cell.value =  '¿Cual?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BS3']
        codigo_cell.value =  'Muestreo por cuotas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BT3']
        codigo_cell.value =  'Muestreo intencional o de conveniencia'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BU3']
        codigo_cell.value =  'Bola de nieve'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BV3']
        codigo_cell.value =  'Muestreo discrecional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BW3']
        codigo_cell.value =  'Otro'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BX3']
        codigo_cell.value =  '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BY3']
        codigo_cell.value =  'Se cuenta con un marco'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BZ3']
        codigo_cell.value =  'Marco de lista'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CA3']
        codigo_cell.value =  'Marco de área'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CB3']
        codigo_cell.value =  'Marco geoestadístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CC3']
        codigo_cell.value =  'Otro(s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CD3']
        codigo_cell.value =  '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CE3']
        codigo_cell.value =  'Metodología'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CF3']
        codigo_cell.value =  'Ficha Metodológica (técnica)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CG3']
        codigo_cell.value =  'Hojas de vida de indicadores'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CH3']
        codigo_cell.value =  'Manual operativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CI3']
        codigo_cell.value =  'Diccionario de datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CJ3']
        codigo_cell.value =  'Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CK3']
        codigo_cell.value = '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CL3']
        codigo_cell.value = 'Ninguno'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CM3']
        codigo_cell.value = 'DANE'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CN3']
        codigo_cell.value = 'Organismos internacionales'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CO3']
        codigo_cell.value = '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CP3']
        codigo_cell.value = 'Otra entidad de orden nacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CQ3']
        codigo_cell.value = '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CR3']
        codigo_cell.value = 'Leyes, decretos, etc'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CS3']
        codigo_cell.value = '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CT3']
        codigo_cell.value = 'Creación propia'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CU3']
        codigo_cell.value = 'Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CV3']
        codigo_cell.value = '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CW3']
        codigo_cell.value = 'Ninguno'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CX3']
        codigo_cell.value = '¿Por qué?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CY3']
        codigo_cell.value = '¿se utilizan nomenclaturas y/o clasificaciones estandarizadas?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CZ3']
        codigo_cell.value = 'Si: ¿Cuáles?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DA3']
        codigo_cell.value = 'Si: Otras'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DB3']
        codigo_cell.value = 'No: ¿Por qué?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DC3']
        codigo_cell.value = 'a. Nacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DD3']
        codigo_cell.value = 'b. Regional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DE3']
        codigo_cell.value = 'b. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DF3']
        codigo_cell.value = 'b. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DG3']
        codigo_cell.value = 'c. Departamental'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DH3']
        codigo_cell.value = 'c. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DI3']
        codigo_cell.value = 'c. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DJ3']
        codigo_cell.value = 'd. Áreas metropolitanas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DK3']
        codigo_cell.value = 'd. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DL3']
        codigo_cell.value = 'd. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DM3']
        codigo_cell.value = 'e. Municipal'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DN3']
        codigo_cell.value = 'e. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DO3']
        codigo_cell.value = 'e. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DP3']
        codigo_cell.value = 'f. Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DQ3']
        codigo_cell.value = 'f. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DR3']
        codigo_cell.value = 'f. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DS3']
        codigo_cell.value = 'a. Nacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DT3']
        codigo_cell.value = 'b. Regional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DU3']
        codigo_cell.value = 'b. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        
        codigo_cell = ws['DV3']
        codigo_cell.value = 'b. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DW3']
        codigo_cell.value = 'c. Departamental'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DX3']
        codigo_cell.value = 'c. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DY3']
        codigo_cell.value = 'c. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DZ3']
        codigo_cell.value = 'd. Áreas metropolitanas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EA3']
        codigo_cell.value = 'd. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EB3']
        codigo_cell.value = 'd. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EC3']
        codigo_cell.value = 'e. Municipal'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['ED3']
        codigo_cell.value = 'e. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EE3']
        codigo_cell.value = 'e. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EF3']
        codigo_cell.value = 'f. Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EG3']
        codigo_cell.value = 'f. ¿Cuántos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EH3']
        codigo_cell.value = 'f. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EI3']
        codigo_cell.value = 'Total'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EJ3']
        codigo_cell.value = 'Urbano'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EK3']
        codigo_cell.value = 'Rural'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EL3']
        codigo_cell.value = 'Sexo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EM3']
        codigo_cell.value = 'Edad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EN3']
        codigo_cell.value = 'Grupo étnico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EO3']
        codigo_cell.value = 'Discapacidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EP3']
        codigo_cell.value = 'Estrato'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EQ3']
        codigo_cell.value = 'Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['ER3']
        codigo_cell.value = '¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['ES3']
        codigo_cell.value = 'Ninguno'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['ET3']
        codigo_cell.value = 'a. Costo anuaL'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EU3']
        codigo_cell.value = 'b. No sabe'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EV3']
        codigo_cell.value = 'Recursos propios'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EW3']
        codigo_cell.value = 'Aportes de otra entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EX3']
        codigo_cell.value = 'Cooperación internacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EY3']
        codigo_cell.value = 'Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EZ3']
        codigo_cell.value = '¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FC3']
        codigo_cell.value = 'a. Formulario físico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FD3']
        codigo_cell.value = 'b. Formulario electrónico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FE3']
        codigo_cell.value = 'c. Dispositivo Móvil de Captura [DMC]'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FF3']
        codigo_cell.value = 'd. Sistema de Información'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FG3']
        codigo_cell.value = 'd. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FH3']
        codigo_cell.value = 'e. Percepción remota (imágenes satelitales, fotos, sensores, etc)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FI3']
        codigo_cell.value = 'f. Base de datos de registro administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FJ3']
        codigo_cell.value = 'g. Resultados estadísticos de otra operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FK3']
        codigo_cell.value = 'h. Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FL3']
        codigo_cell.value = 'h. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FM3']
        codigo_cell.value = 'a. Anual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FN3']
        codigo_cell.value = 'b. Semestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FO3']
        codigo_cell.value = 'c. Trimestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FP3']
        codigo_cell.value = 'd. Mensual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FQ3']
        codigo_cell.value = 'e. Diaria'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FR3']
        codigo_cell.value = 'f. Otra'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FS3']
        codigo_cell.value = 'f. ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FT3']
        codigo_cell.value = 'a. Excel'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FU3']
        codigo_cell.value = 'b. Acces'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FV3']
        codigo_cell.value = 'c. R'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FW3']
        codigo_cell.value = 'd. SAS'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FX3']
        codigo_cell.value = 'e. SPSS'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FY3']
        codigo_cell.value = 'f. Oracle'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FZ3']
        codigo_cell.value = 'g. Stata'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GA3']
        codigo_cell.value = 'h. Otra'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GB3']
        codigo_cell.value = 'h. ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GD3']
        codigo_cell.value = 'a Consistencia y validación de datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GE3']
        codigo_cell.value = 'b Análisis de tendencias y series de tiempo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GF3']
        codigo_cell.value = 'c Análisis de contexto de los resultados'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GG3']
        codigo_cell.value = 'd Otro.'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GH3']
        codigo_cell.value = 'd. ¿cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GI3']
        codigo_cell.value = 'a. Página web'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GJ3']
        codigo_cell.value = 'b. Medio físico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GK3']
        codigo_cell.value = 'c. Medio electrónico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GL3']
        codigo_cell.value = 'd. Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GM3']
        codigo_cell.value = 'd. ¿cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GO3']
        codigo_cell.value = 'Desde mes y año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GP3']
        codigo_cell.value = 'Hasta mes y año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GQ3']
        codigo_cell.value = 'a. Fecha publicación'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GR3']
        codigo_cell.value = 'a. mes/año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GS3']
        codigo_cell.value = 'b. No sabe'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GT3']
        codigo_cell.value = 'c. No hay'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GU3']
        codigo_cell.value = 'c. ¿por qué?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GV3']
        codigo_cell.value = 'a. Anual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GW3']
        codigo_cell.value = 'b. Semestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GX3']
        codigo_cell.value = 'c. Trimestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GY3']
        codigo_cell.value = 'd. Mensual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GZ3']
        codigo_cell.value = 'e. Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HA3']
        codigo_cell.value = 'e. ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HB3']
        codigo_cell.value = 'f. No está definido'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HC3']
        codigo_cell.value = 'a. Cuadros de salida (tablas)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HD3']
        codigo_cell.value = 'b. Boletín estadístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HE3']
        codigo_cell.value = 'c. Anuario'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HF3']
        codigo_cell.value = 'd. Mapas estadísticos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HG3']
        codigo_cell.value = 'e. Bases de datos interactivas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HH3']
        codigo_cell.value = 'f. Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HI3']
        codigo_cell.value = 'f. ¿cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HJ3']
        codigo_cell.value = 'a. Series históricas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HK3']
        codigo_cell.value = 'a. Desde: mes y año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HL3']
        codigo_cell.value = 'a. Hasta: mes y año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HM3']
        codigo_cell.value = 'b. Microdatos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HN3']
        codigo_cell.value = 'b. Desde: mes y año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HO3']
        codigo_cell.value = 'b. Hasta: mes y año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HP3']
        codigo_cell.value = 'c. Documentos metodológicos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HQ3']
        codigo_cell.value = 'c. URL'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HR3']
        codigo_cell.value = 'd. Calendario de difusión'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HS3']
        codigo_cell.value = 'e. Ninguna'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HT3']
        codigo_cell.value = '¿Conoce si otra entidad produce resultados similares a los de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HU3']
        codigo_cell.value = 'Si: Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HV3']
        codigo_cell.value = 'Si: Operación estadística/Indicadores'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HW3']
        codigo_cell.value = '¿La entidad hace parte de un sistema de información estadística de uso  interinstitucional ?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HX3']
        codigo_cell.value = 'Si: ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        cont = 4
           
        for ooee in ooees:
             
            ws.cell(row = cont, column = 1).value = ooee.codigo_oe
            ws.cell(row = cont, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 1).font = Font(size = "8", name='Barlow')
            ws.cell(row = cont, column = 1).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            ws.cell(row = cont, column = 2).value = ooee.nombre_oe
            ws.cell(row = cont, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 2).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 3).value = ooee.objetivo_oe
            ws.cell(row = cont, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 3).font = Font(size = "8", name='Barlow')

            
            ws.cell(row = cont, column = 4).value = str(ooee.entidad)
            ws.cell(row = cont, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 4).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 5).value =  ooee.entidad.codigo
            ws.cell(row = cont, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 5).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 6).value = str(ooee.area_tematica)
            ws.cell(row = cont, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 6).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 7).value = str(ooee.tema)
            ws.cell(row = cont, column = 7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 7).font = Font(size = "8", name='Barlow')
            
            ## TEMA COMPARTIDO

            listaTemaCompartido = ooee.tema_compartido.all()
            temaCompartido_array = []
            for indexTemaCompa, itemTemaCompa in enumerate(listaTemaCompartido):
                temaCompartido_array.append(str(itemTemaCompa))
                ws.cell(row = cont, column = 8).value = str(temaCompartido_array)
                ws.cell(row = cont, column = 8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 8).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 9).value = ooee.nombre_dep
            ws.cell(row = cont, column = 9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = cont, column = 9).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 10).value = ooee.nombre_dir
            ws.cell(row = cont, column = 10).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 10).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 11).value = ooee.cargo_dir
            ws.cell(row = cont, column = 11).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 11).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 12).value = ooee.correo_dir
            ws.cell(row = cont, column = 12).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 12).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 13).value = ooee.tel_dir
            ws.cell(row = cont, column = 13).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 13).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 14).value = ooee.nombre_resp
            ws.cell(row = cont, column = 14).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 14).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 15).value = ooee.cargo_resp
            ws.cell(row = cont, column = 15).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 15).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 16).value = ooee.correo_resp
            ws.cell(row = cont, column = 16).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 16).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 17).value = ooee.tel_resp
            ws.cell(row = cont, column = 17).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 17).font = Font(size = "8", name='Barlow')
            
            if str(ooee.fase) == "True":
                ws.cell(row = cont, column = 18).value = "Si"
                ws.cell(row = cont, column = 18).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 18).font = Font(size = "8", name='Barlow')
            elif str(ooee.fase) == "False":
                ws.cell(row = cont, column = 18).value = "No"
                ws.cell(row = cont, column = 18).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 18).font = Font(size = "8", name='Barlow')
            
            #*****lista de OOEE Y RRAA PREGUNTA 4 DEL modulo c
            listaRRAAobtencion = ooee.rraa_lista.all()
            listaRRAAobt_array = []
            for indexListaRRAAobt, itemListaRRAAobt in enumerate(listaRRAAobtencion):
                listaRRAAobt_array.append(str(itemListaRRAAobt))
                ws.cell(row = cont, column = 58).value = str(listaRRAAobt_array).replace('[','').replace(']','')
                ws.cell(row = cont, column = 58).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 58).font = Font(size = "8", name='Barlow')

            listaOOEEobtencion = ooee.ooee_lista.all()
            listaOOEEobt_array = []
            for indexListaOOEEobt, itemListaOOEEobt in enumerate(listaOOEEobtencion):
                listaOOEEobt_array.append(str(itemListaOOEEobt))
                ws.cell(row = cont, column = 62).value = str(listaOOEEobt_array)
                ws.cell(row = cont, column = 62).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 62).font = Font(size = "8", name='Barlow')

            listaNorma = ooee.norma.all()
            listaRequerimiento = ooee.requerimientos.all()
            listaPriUser = ooee.pri_usuarios.all()
            listaUnidadObs = ooee.uni_observacion.all()
            listaTipoOper = ooee.tipo_operacion.all()
            listaObtDato =  ooee.obt_dato.all()
            listaTipoProbabilistico = ooee.tipo_probabilistico.all()
            listaTipoNoProbabilistico = ooee.tipo_no_probabilistico.all()
            listaTipoMarco = ooee.tipo_marco.all()
            listaDocsDesarrollo = ooee.docs_des.all()
            ListaListaConc = ooee.lista_conc.all()
            listaClasificaciones =  ooee.nombre_cla.all()
            listaCobGeog = ooee.cob_geo.all()
            listaDesgeo = ooee.des_geo.all()
            listaDesZona = ooee.des_zona.all()
            listaDesGrupo = ooee.des_grupo.all()
            listaFuentes = ooee.fuentes.all()
            listaMediObtDato = ooee.med_obt.all()
            listaPeriodicidad = ooee.periodicidad.all()
            listaHerramProcesa = ooee.h_proc.all()
            listaAnalisResultados = ooee.a_resul.all()
            listaMedioDifusion = ooee.m_dif.all()
            listaFechaPublicacion = ooee.f_publi.all()
            listaFrecuenciaDifusion = ooee.fre_dif.all()
            listaProductosDifundir = ooee.pro_dif.all()
            listaOtrosProductos = ooee.otro_prod.all()

            for index,item in enumerate(listaNorma):
                indice = index
                if  str(item) == 'Ninguna' and index == indice:
                    #print("item", item)
                    ws.cell(row = cont, column = 23).value = str(item)
                    ws.cell(row = cont, column = 23).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 23).font = Font(size = "8", name='Barlow')
                   

            for indexReq, itemReq in enumerate(listaRequerimiento):
                indiceReq = indexReq
                if indexReq == indiceReq and str(itemReq) == 'Ninguno':
                    ws.cell(row = cont, column = 31).value = str(itemReq)
                    ws.cell(row = cont, column = 31).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 31).font = Font(size = "8", name='Barlow')
                   
            
            for indexPriUs, itemPrius in enumerate(listaPriUser):
                indicePriUser = indexPriUs
                if indexPriUs == indicePriUser and str(itemPrius) == 'Público en General':
                    ws.cell(row = cont, column = 42).value = str(itemPrius)
                    ws.cell(row = cont, column = 42).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 42).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 43).value = ooee.pob_obje
            ws.cell(row = cont, column = 43).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 43).font = Font(size = "8", name='Barlow')

            for indexUnidadObs, itemUnidadObs, in enumerate(listaUnidadObs):   
                if  str(itemUnidadObs) == "Empresa":
                    ws.cell(row = cont, column = 44).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 44).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 44).font = Font(size = "8", name='Barlow')
                if  str(itemUnidadObs) == "Establecimiento":   
                    ws.cell(row = cont, column = 45).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 45).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 45).font = Font(size = "8", name='Barlow')
                if  str(itemUnidadObs) == "Hogar":   
                    ws.cell(row = cont, column = 46).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 46).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 46).font = Font(size = "8", name='Barlow')
                if  str(itemUnidadObs) == "Persona":   
                    ws.cell(row = cont, column = 47).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 47).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 47).font = Font(size = "8", name='Barlow')
                if  str(itemUnidadObs) == "Unidad productora agropecuaria":   
                    ws.cell(row = cont, column = 48).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 48).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 48).font = Font(size = "8", name='Barlow')
                if  str(itemUnidadObs) == "Predio":   
                    ws.cell(row = cont, column = 49).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 49).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 49).font = Font(size = "8", name='Barlow')
                if  str(itemUnidadObs) == "Otra (s)":   
                    ws.cell(row = cont, column = 50).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 50).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 50).font = Font(size = "8", name='Barlow')

            for indexListaOper, itemListaOper in enumerate(listaTipoOper):
                if str(itemListaOper) == "Aprovechamiento de registro administrativo":
                    ws.cell(row = cont, column = 52).value = str(itemListaOper)
                    ws.cell(row = cont, column = 52).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 52).font = Font(size = "8", name='Barlow')
                if str(itemListaOper) == "Estadística derivada":
                    ws.cell(row = cont, column = 53).value = str(itemListaOper)
                    ws.cell(row = cont, column = 53).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 53).font = Font(size = "8", name='Barlow')
                if str(itemListaOper) == "Muestreo probabilístico":
                    ws.cell(row = cont, column = 54).value = str(itemListaOper)
                    ws.cell(row = cont, column = 54).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 54).font = Font(size = "8", name='Barlow')
                if str(itemListaOper) == "Muestreo no probabilístico":
                    ws.cell(row = cont, column = 55).value = str(itemListaOper)
                    ws.cell(row = cont, column = 55).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 55).font = Font(size = "8", name='Barlow')
                if str(itemListaOper) == "Censo":
                    ws.cell(row = cont, column = 56).value = str(itemListaOper)
                    ws.cell(row = cont, column = 56).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 56).font = Font(size = "8", name='Barlow')
            
            for indexObtDato, itemObtDato in enumerate(listaObtDato):
                if str(itemObtDato) == "Registro Administrativos":
                    ws.cell(row = cont, column = 57).value = str(itemObtDato)
                    ws.cell(row = cont, column = 57).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 57).font = Font(size = "8", name='Barlow')
                if str(itemObtDato) == "Operación Estadística":
                    ws.cell(row = cont, column = 61).value = str(itemObtDato)
                    ws.cell(row = cont, column = 61).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 61).font = Font(size = "8", name='Barlow')            

            for indexTipoProba, itemTipoProba in enumerate(listaTipoProbabilistico):
                if str(itemTipoProba) == "Muestreo aleatorio simple":
                    ws.cell(row = cont, column = 65).value = str(itemTipoProba)
                    ws.cell(row = cont, column = 65).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 65).font = Font(size = "8", name='Barlow')
                if str(itemTipoProba) == "Muestreo aleatorio sistemático":
                    ws.cell(row = cont, column = 66).value = str(itemTipoProba)
                    ws.cell(row = cont, column = 66).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 66).font = Font(size = "8", name='Barlow')
                if str(itemTipoProba) == "Muestreo aleatorio estratificado":
                    ws.cell(row = cont, column = 67).value = str(itemTipoProba)
                    ws.cell(row = cont, column = 67).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 67).font = Font(size = "8", name='Barlow')
                if str(itemTipoProba) == "Muestreo aleatorio por conglomerados":
                    ws.cell(row = cont, column = 68).value = str(itemTipoProba)
                    ws.cell(row = cont, column = 68).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 68).font = Font(size = "8", name='Barlow')
                if str(itemTipoProba) == "Otro":
                    ws.cell(row = cont, column = 69).value = str(itemTipoProba)
                    ws.cell(row = cont, column = 69).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 69).font = Font(size = "8", name='Barlow')

            for indexTipoNoProba, itemTipoNoProba in enumerate(listaTipoNoProbabilistico):
                if str(itemTipoNoProba) == "Muestreo por cuotas":
                    ws.cell(row = cont, column = 71).value = str(itemTipoNoProba)
                    ws.cell(row = cont, column = 71).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 71).font = Font(size = "8", name='Barlow')
                if str(itemTipoNoProba) == "Muestreo intencional o de conveniencia":
                    ws.cell(row = cont, column = 72).value = str(itemTipoNoProba)
                    ws.cell(row = cont, column = 72).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 72).font = Font(size = "8", name='Barlow')
                if str(itemTipoNoProba) == "Bola de nieve":
                    ws.cell(row = cont, column = 73).value = str(itemTipoNoProba)
                    ws.cell(row = cont, column = 73).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)                               
                    ws.cell(row = cont, column = 73).font = Font(size = "8", name='Barlow')
                if str(itemTipoNoProba) == "Muestreo discrecional":
                    ws.cell(row = cont, column = 74).value = str(itemTipoNoProba)
                    ws.cell(row = cont, column = 74).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 74).font = Font(size = "8", name='Barlow')
                if str(itemTipoNoProba) == "Otro":
                    ws.cell(row = cont, column = 75).value = str(itemTipoNoProba)
                    ws.cell(row = cont, column = 75).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 75).font = Font(size = "8", name='Barlow')

            if str(ooee.marco_estad) == "True":

                ws.cell(row = cont, column = 77).value = "Si"
                ws.cell(row = cont, column = 77).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 77).font = Font(size = "8", name='Barlow')
            
            elif str(ooee.marco_estad) == "False":
                ws.cell(row = cont, column = 77).value = "No"
                ws.cell(row = cont, column = 77).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 77).font = Font(size = "8", name='Barlow')
            
            for indexTipoMarco, itemTipoMarco in enumerate(listaTipoMarco):
                if str(itemTipoMarco) == "Marco de lista":
                    ws.cell(row = cont, column = 78).value = str(itemTipoMarco)
                    ws.cell(row = cont, column = 78).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 78).font = Font(size = "8", name='Barlow')
                if str(itemTipoMarco) == "Marco de área":
                    ws.cell(row = cont, column = 79).value = str(itemTipoMarco)
                    ws.cell(row = cont, column = 79).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 79).font = Font(size = "8", name='Barlow')
                if str(itemTipoMarco) == "Marco geoestadístico":
                    ws.cell(row = cont, column = 80).value = str(itemTipoMarco)
                    ws.cell(row = cont, column = 80).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 80).font = Font(size = "8", name='Barlow')
                if str(itemTipoMarco) == "Otro(s)":
                    ws.cell(row = cont, column = 81).value = str(itemTipoMarco)
                    ws.cell(row = cont, column = 81).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)                        
                    ws.cell(row = cont, column = 81).font = Font(size = "8", name='Barlow')
            
            for indexDocsDesarrollo, itemDocsDesarrollo in enumerate(listaDocsDesarrollo):
                if str(itemDocsDesarrollo) == "Metodología":
                    ws.cell(row = cont, column = 83).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 83).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 83).font = Font(size = "8", name='Barlow')
                if str(itemDocsDesarrollo) == "Ficha Metodológica (técnica)":
                    ws.cell(row = cont, column = 84).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 84).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 84).font = Font(size = "8", name='Barlow')
                if str(itemDocsDesarrollo) == "Hojas de vida de indicadores":
                    ws.cell(row = cont, column = 85).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 85).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 85).font = Font(size = "8", name='Barlow')
                if str(itemDocsDesarrollo) == "Manual operativo":
                    ws.cell(row = cont, column = 86).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 86).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 86).font = Font(size = "8", name='Barlow')
                if str(itemDocsDesarrollo) == "Diccionario de datos":
                    ws.cell(row = cont, column = 87).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 87).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 87).font = Font(size = "8", name='Barlow')
                if str(itemDocsDesarrollo) == "Otro (s)":
                    ws.cell(row = cont, column = 88).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 88).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 88).font = Font(size = "8", name='Barlow')
                if str(itemDocsDesarrollo) == "Ninguno":
                    ws.cell(row = cont, column = 90).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 90).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 90).font = Font(size = "8", name='Barlow')
            
            for indexListaConc, itemListaConc in enumerate(ListaListaConc):
                if str(itemListaConc) == "DANE":
                    ws.cell(row = cont, column = 91).value = str(itemListaConc)
                    ws.cell(row = cont, column = 91).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 91).font = Font(size = "8", name='Barlow')
                if str(itemListaConc) == "Organismos internacionales":
                    ws.cell(row = cont, column = 92).value = str(itemListaConc)
                    ws.cell(row = cont, column = 92).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 92).font = Font(size = "8", name='Barlow')
                if str(itemListaConc) == "Otra entidad de orden nacional":
                    ws.cell(row = cont, column = 94).value = str(itemListaConc)
                    ws.cell(row = cont, column = 94).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 94).font = Font(size = "8", name='Barlow')
                if str(itemListaConc) == "Leyes, decretos, etc.":
                    ws.cell(row = cont, column = 96).value = str(itemListaConc)
                    ws.cell(row = cont, column = 96).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 96).font = Font(size = "8", name='Barlow')
                if str(itemListaConc) == "Creación propia":
                    ws.cell(row = cont, column = 98).value = str(itemListaConc)
                    ws.cell(row = cont, column = 98).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 98).font = Font(size = "8", name='Barlow')
                if str(itemListaConc) == "Otra (s)":
                    ws.cell(row = cont, column = 99).value = str(itemListaConc)
                    ws.cell(row = cont, column = 99).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 99).font = Font(size = "8", name='Barlow')
                if str(itemListaConc) == "Ninguno":
                    ws.cell(row = cont, column = 101).value = str(itemListaConc)
                    ws.cell(row = cont, column = 101).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 101).font = Font(size = "8", name='Barlow')

            if str(ooee.nome_clas) == "True":
                ws.cell(row = cont, column = 103).value = "Si"
                ws.cell(row = cont, column = 103).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 103).font = Font(size = "8", name='Barlow')
            
            elif str(ooee.nome_clas) == "False":
                ws.cell(row = cont, column = 103).value = "No"
                ws.cell(row = cont, column = 103).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 103).font = Font(size = "8", name='Barlow')
            
            
            clasifi_array = []
            for indexClasif, itemClasif in enumerate(listaClasificaciones):
                clasifi_array.append(str(itemClasif))
                ws.cell(row = cont, column = 104).value = str(clasifi_array)
                ws.cell(row = cont, column = 104).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 104).font = Font(size = "8", name='Barlow')
        
            for indexCobGeo, itemCobGeo in enumerate(listaCobGeog):
                
                if str(itemCobGeo) == "Nacional":
                    ws.cell(row = cont, column = 107).value = str(itemCobGeo)
                    ws.cell(row = cont, column = 107).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 107).font = Font(size = "8", name='Barlow')

                if str(itemCobGeo) == "Regional":
                    ws.cell(row = cont, column = 108).value = str(itemCobGeo)
                    ws.cell(row = cont, column = 108).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 108).font = Font(size = "8", name='Barlow')
                
                if str(itemCobGeo) == "Departamental":
                    ws.cell(row = cont, column = 111).value = str(itemCobGeo)
                    ws.cell(row = cont, column = 111).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 111).font = Font(size = "8", name='Barlow')
                
                if str(itemCobGeo) == "Áreas metropolitanas":
                    ws.cell(row = cont, column = 114).value = str(itemCobGeo)
                    ws.cell(row = cont, column = 114).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 114).font = Font(size = "8", name='Barlow')
                
                if str(itemCobGeo) == "Municipal":
                    ws.cell(row = cont, column = 117).value = str(itemCobGeo)
                    ws.cell(row = cont, column = 117).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 117).font = Font(size = "8", name='Barlow')
                
                if str(itemCobGeo) == "Otro (s)":
                    ws.cell(row = cont, column = 120).value = str(itemCobGeo)
                    ws.cell(row = cont, column = 120).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 120).font = Font(size = "8", name='Barlow')


                for indexDesgeo, itemDesgeo in enumerate(listaDesgeo): 
                    
                    if str(itemDesgeo) == "Nacional":
                        ws.cell(row = cont, column = 123).value = str(itemDesgeo)
                        ws.cell(row = cont, column = 123).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 123).font = Font(size = "8", name='Barlow')
                    
                    if str(itemDesgeo) == "Regional":
                        ws.cell(row = cont, column = 124).value = str(itemDesgeo)
                        ws.cell(row = cont, column = 124).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 124).font = Font(size = "8", name='Barlow')

                    if str(itemDesgeo) == "Departamental":
                        ws.cell(row = cont, column = 127).value = str(itemDesgeo)
                        ws.cell(row = cont, column = 127).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 127).font = Font(size = "8", name='Barlow')
                    
                    if str(itemDesgeo) == "Áreas metropolitanas":
                        ws.cell(row = cont, column = 130).value = str(itemDesgeo)
                        ws.cell(row = cont, column = 130).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 130).font = Font(size = "8", name='Barlow')
                    
                    if str(itemDesgeo) == "Municipal":
                        ws.cell(row = cont, column = 133).value = str(itemDesgeo)
                        ws.cell(row = cont, column = 133).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 133).font = Font(size = "8", name='Barlow')
                    
                    if str(itemDesgeo) == "Otro (s)":
                        ws.cell(row = cont, column = 136).value = str(itemDesgeo)
                        ws.cell(row = cont, column = 136).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 136).font = Font(size = "8", name='Barlow')
                
                
                for indexDesZona, itemDesZona in enumerate(listaDesZona):
                    if str(itemDesZona) == "Total":
                        ws.cell(row = cont, column = 139).value = str(itemDesZona)
                        ws.cell(row = cont, column = 139).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 138).font = Font(size = "8", name='Barlow')

                    if str(itemDesZona) == "Urbano":
                        ws.cell(row = cont, column = 140).value = str(itemDesZona)
                        ws.cell(row = cont, column = 140).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 140).font = Font(size = "8", name='Barlow')
                    
                    if str(itemDesZona) == "Rural":
                        ws.cell(row = cont, column = 141).value = str(itemDesZona)
                        ws.cell(row = cont, column = 141).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 141).font = Font(size = "8", name='Barlow')
                    
                
                for indexDesGrupo, itemDesGrupo in enumerate(listaDesGrupo):
                    
                    if str(itemDesGrupo) == "Sexo":
                        ws.cell(row = cont, column = 142).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 142).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 142).font = Font(size = "8", name='Barlow')

                    if str(itemDesGrupo) == "Edad":
                        ws.cell(row = cont, column = 143).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 143).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 143).font = Font(size = "8", name='Barlow')

                    if str(itemDesGrupo) == "Grupo étnico":
                        ws.cell(row = cont, column = 144).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 144).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 144).font = Font(size = "8", name='Barlow')

                    if str(itemDesGrupo) == "Discapacidad":
                        ws.cell(row = cont, column = 145).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 145).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 145).font = Font(size = "8", name='Barlow')

                    if str(itemDesGrupo) == "Estrato":
                        ws.cell(row = cont, column = 146).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 146).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 146).font = Font(size = "8", name='Barlow')

                    if str(itemDesGrupo) == "Otro, ¿cuál?":
                        ws.cell(row = cont, column = 147).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 147).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 147).font = Font(size = "8", name='Barlow')

                    if str(itemDesGrupo) == "Ninguno":
                        ws.cell(row = cont, column = 149).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 149).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 149).font = Font(size = "8", name='Barlow')

                ws.cell(row = cont, column = 150).value = ooee.ca_anual
                ws.cell(row = cont, column = 150).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 150).font = Font(size = "8", name='Barlow') 

                ws.cell(row = cont, column = 151).value = ooee.cb_anual
                ws.cell(row = cont, column = 151).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 151).font = Font(size = "8", name='Barlow')  
               
                for indexFuentes, itemFuentes in enumerate(listaFuentes):
                    
                    if str(itemFuentes) == "Recursos propios":
                        ws.cell(row = cont, column = 152).value = str(itemFuentes)
                        ws.cell(row = cont, column = 152).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 152).font = Font(size = "8", name='Barlow')

                    if str(itemFuentes) == "Aportes de otra entidad":
                        ws.cell(row = cont, column = 153).value = str(itemFuentes)
                        ws.cell(row = cont, column = 153).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 153).font = Font(size = "8", name='Barlow')

                    if str(itemFuentes) == "Cooperación internacional":
                        ws.cell(row = cont, column = 154).value = str(itemFuentes)
                        ws.cell(row = cont, column = 154).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 154).font = Font(size = "8", name='Barlow')

                    if str(itemFuentes) == "Otro (s)":
                        ws.cell(row = cont, column = 155).value = str(itemFuentes)
                        ws.cell(row = cont, column = 155).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 155).font = Font(size = "8", name='Barlow')
                
                ## 158 y 159 si se desea agregar preguntas 14 y 15 del modulo C

                for indexMediObtDato, itemMediObtDato in enumerate(listaMediObtDato):

                    if str(itemMediObtDato) == "Formulario físico":
                        ws.cell(row = cont, column = 159).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 159).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 159).font = Font(size = "8", name='Barlow')
                    
                    if str(itemMediObtDato) == "Formulario electrónico":
                        ws.cell(row = cont, column = 160).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 160).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 160).font = Font(size = "8", name='Barlow')

                    if str(itemMediObtDato) == "Dispositivo Móvil de Captura [DMC]":
                        ws.cell(row = cont, column = 161).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 161).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 161).font = Font(size = "8", name='Barlow')

                    if str(itemMediObtDato) == "Sistema de Información":
                        ws.cell(row = cont, column = 162).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 162).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 162).font = Font(size = "8", name='Barlow')

                    if str(itemMediObtDato) == "Percepción remota (imágenes satelitales, fotos, sensores, etc)":
                        ws.cell(row = cont, column = 164).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 164).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 164).font = Font(size = "8", name='Barlow')

                    if str(itemMediObtDato) == "Base de datos de registro administrativo":
                        ws.cell(row = cont, column = 165).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 165).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 165).font = Font(size = "8", name='Barlow')

                    if str(itemMediObtDato) == "Resultados estadísticos de otra operación estadística":
                        ws.cell(row = cont, column = 166).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 166).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 166).font = Font(size = "8", name='Barlow')

                    if str(itemMediObtDato) == "Otro (s)":
                        ws.cell(row = cont, column = 167).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 167).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 167).font = Font(size = "8", name='Barlow')

                for indexPeriodicidad, itemPeriodicidad in enumerate(listaPeriodicidad):
                    if str(itemPeriodicidad) == "Anual":
                        ws.cell(row = cont, column = 169).value = str(itemPeriodicidad)
                        ws.cell(row = cont, column = 169).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 169).font = Font(size = "8", name='Barlow')
                    
                    if str(itemPeriodicidad) == "Semestral":
                        ws.cell(row = cont, column = 170).value = str(itemPeriodicidad)
                        ws.cell(row = cont, column = 170).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 170).font = Font(size = "8", name='Barlow')

                    if str(itemPeriodicidad) == "Trimestral":
                        ws.cell(row = cont, column = 171).value = str(itemPeriodicidad)
                        ws.cell(row = cont, column = 171).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 171).font = Font(size = "8", name='Barlow')

                    if str(itemPeriodicidad) == "Mensual":
                        ws.cell(row = cont, column = 172).value = str(itemPeriodicidad)
                        ws.cell(row = cont, column = 172).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 172).font = Font(size = "8", name='Barlow')
                    
                    if str(itemPeriodicidad) == "Diaria":
                        ws.cell(row = cont, column = 173).value = str(itemPeriodicidad)
                        ws.cell(row = cont, column = 173).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 173).font = Font(size = "8", name='Barlow')
                    
                    if str(itemPeriodicidad) == "Otra":
                        ws.cell(row = cont, column = 174).value = str(itemPeriodicidad)
                        ws.cell(row = cont, column = 174).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 174).font = Font(size = "8", name='Barlow')

                for indexHerramProcesa, itemHerramProcesa in enumerate(listaHerramProcesa):
                    
                    if str(itemHerramProcesa) == "Excel" :
                        ws.cell(row = cont, column = 176).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 176).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 176).font = Font(size = "8", name='Barlow')

                    if str(itemHerramProcesa) == "Access" :
                        ws.cell(row = cont, column = 177).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 177).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 177).font = Font(size = "8", name='Barlow')

                    if str(itemHerramProcesa) == "R" :
                        ws.cell(row = cont, column = 178).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 178).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 178).font = Font(size = "8", name='Barlow')

                    if str(itemHerramProcesa) == "SAS" :
                        ws.cell(row = cont, column = 179).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 179).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 179).font = Font(size = "8", name='Barlow')

                    if str(itemHerramProcesa) == "SPSS" :
                        ws.cell(row = cont, column = 180).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 180).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 180).font = Font(size = "8", name='Barlow')
                    
                    if str(itemHerramProcesa) == "Oracle" :
                        ws.cell(row = cont, column = 181).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 181).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 181).font = Font(size = "8", name='Barlow')

                    if str(itemHerramProcesa) == "Stata" :
                        ws.cell(row = cont, column = 182).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 182).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 182).font = Font(size = "8", name='Barlow')
                    
                    if str(itemHerramProcesa) == "Otra (s)" :
                        ws.cell(row = cont, column = 183).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 183).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 183).font = Font(size = "8", name='Barlow')

                ws.cell(row = cont, column = 185).value = ooee.descrip_proces
                ws.cell(row = cont, column = 185).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 185).font = Font(size = "8", name='Barlow')  
               
                for indexAnalisResultados, itemAnalisResultados in enumerate(listaAnalisResultados):
                    
                    if str(itemAnalisResultados) == "Consistencia y validación de datos":
                        ws.cell(row = cont, column = 186).value = str(itemAnalisResultados)
                        ws.cell(row = cont, column = 186).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 186).font = Font(size = "8", name='Barlow')

                    if str(itemAnalisResultados) == "Análisis de tendencias y series de tiempo":
                        ws.cell(row = cont, column = 187).value = str(itemAnalisResultados)
                        ws.cell(row = cont, column = 187).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 187).font = Font(size = "8", name='Barlow')

                    if str(itemAnalisResultados) == "Análisis de contexto de los resultados":
                        ws.cell(row = cont, column = 188).value = str(itemAnalisResultados)
                        ws.cell(row = cont, column = 188).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 188).font = Font(size = "8", name='Barlow') 

                    if str(itemAnalisResultados) == "Otro":
                        ws.cell(row = cont, column = 189).value = str(itemAnalisResultados)
                        ws.cell(row = cont, column = 189).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 189).font = Font(size = "8", name='Barlow')  

                for indexMedioDifusion, itemMedioDifusion in enumerate(listaMedioDifusion):

                    if str(itemMedioDifusion) == "Página web":
                        ws.cell(row = cont, column = 191).value = str(itemMedioDifusion)
                        ws.cell(row = cont, column = 191).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 191).font = Font(size = "8", name='Barlow') 
                    
                    if str(itemMedioDifusion) == "Medio físico":
                        ws.cell(row = cont, column = 192).value = str(itemMedioDifusion)
                        ws.cell(row = cont, column = 192).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 192).font = Font(size = "8", name='Barlow') 

                    if str(itemMedioDifusion) == "Medio electrónico":
                        ws.cell(row = cont, column = 193).value = str(itemMedioDifusion)
                        ws.cell(row = cont, column = 193).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 193).font = Font(size = "8", name='Barlow') 

                    if str(itemMedioDifusion) == "Otro":
                        ws.cell(row = cont, column = 194).value = str(itemMedioDifusion)
                        ws.cell(row = cont, column = 194).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 194).font = Font(size = "8", name='Barlow')  
                
                
                ws.cell(row = cont, column = 196).value = str(ooee.res_est_url)
                ws.cell(row = cont, column = 196).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 196).font = Font(size = "8", name='Barlow')  

                ws.cell(row = cont, column = 197).value = str(ooee.dispo_desde)
                ws.cell(row = cont, column = 197).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 197).font = Font(size = "8", name='Barlow')  

                ws.cell(row = cont, column = 198).value = str(ooee.dispo_hasta)
                ws.cell(row = cont, column = 198).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 198).font = Font(size = "8", name='Barlow')  
                
            for indexFechaPublicacion, itemFechaPublicacion in enumerate(listaFechaPublicacion):  
                
                if str(itemFechaPublicacion) == "fecha Publicación":
                    ws.cell(row = cont, column = 199).value = str(itemFechaPublicacion)
                    ws.cell(row = cont, column = 199).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 199).font = Font(size = "8", name='Barlow')

                if str(itemFechaPublicacion) == "No sabe":
                    ws.cell(row = cont, column = 201).value = str(itemFechaPublicacion)
                    ws.cell(row = cont, column = 201).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 201).font = Font(size = "8", name='Barlow')

                if str(itemFechaPublicacion) == "No hay":
                    ws.cell(row = cont, column = 202).value = str(itemFechaPublicacion)
                    ws.cell(row = cont, column = 202).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 202).font = Font(size = "8", name='Barlow')

            for indexFrecuenciaDifusion, itemFrecuenciaDifusion in enumerate(listaFrecuenciaDifusion):
                
                if str(itemFrecuenciaDifusion) == "Anual":
                    ws.cell(row = cont, column = 204).value = str(itemFrecuenciaDifusion)
                    ws.cell(row = cont, column = 204).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 204).font = Font(size = "8", name='Barlow')

                if str(itemFrecuenciaDifusion) == "Semestral":
                    ws.cell(row = cont, column = 205).value = str(itemFrecuenciaDifusion)
                    ws.cell(row = cont, column = 205).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 205).font = Font(size = "8", name='Barlow')

                if str(itemFrecuenciaDifusion) == "Trimestral":
                    ws.cell(row = cont, column = 206).value = str(itemFrecuenciaDifusion)
                    ws.cell(row = cont, column = 206).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 206).font = Font(size = "8", name='Barlow')

                if str(itemFrecuenciaDifusion) == "Mensual":
                    ws.cell(row = cont, column = 207).value = str(itemFrecuenciaDifusion)
                    ws.cell(row = cont, column = 207).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 207).font = Font(size = "8", name='Barlow')

                if str(itemFrecuenciaDifusion) == "Otro (s)":
                    ws.cell(row = cont, column = 208).value = str(itemFrecuenciaDifusion)
                    ws.cell(row = cont, column = 208).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 208).font = Font(size = "8", name='Barlow')

                if str(itemFrecuenciaDifusion) == "No está definido":
                    ws.cell(row = cont, column = 209).value = str(itemFrecuenciaDifusion)
                    ws.cell(row = cont, column = 209).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 209).font = Font(size = "8", name='Barlow')

            for indexProductosDifundir, itemProductosDifundir in enumerate(listaProductosDifundir):

                if str(itemProductosDifundir) == "Cuadros de salida (tablas)":
                    ws.cell(row = cont, column = 211).value = str(itemProductosDifundir)
                    ws.cell(row = cont, column = 211).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 211).font = Font(size = "8", name='Barlow')
                
                if str(itemProductosDifundir) == "Boletín estadístico":
                    ws.cell(row = cont, column = 212).value = str(itemProductosDifundir)
                    ws.cell(row = cont, column = 212).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 212).font = Font(size = "8", name='Barlow')

                if str(itemProductosDifundir) == "Anuario":
                    ws.cell(row = cont, column = 213).value = str(itemProductosDifundir)
                    ws.cell(row = cont, column = 213).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 213).font = Font(size = "8", name='Barlow')

                if str(itemProductosDifundir) == "Mapas estadísticos":
                    ws.cell(row = cont, column = 214).value = str(itemProductosDifundir)
                    ws.cell(row = cont, column = 214).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 214).font = Font(size = "8", name='Barlow')

                if str(itemProductosDifundir) == "Bases de datos interactivas":
                    ws.cell(row = cont, column = 215).value = str(itemProductosDifundir)
                    ws.cell(row = cont, column = 215).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 215).font = Font(size = "8", name='Barlow')

                if str(itemProductosDifundir) == "Otro (s)":
                    ws.cell(row = cont, column = 216).value = str(itemProductosDifundir)
                    ws.cell(row = cont, column = 216).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 216).font = Font(size = "8", name='Barlow')

            for indexOtrosProductos, itemOtrosProductos in enumerate(listaOtrosProductos):

                if str(itemOtrosProductos) == "Series históricas":
                    ws.cell(row = cont, column = 218).value = str(itemOtrosProductos)
                    ws.cell(row = cont, column = 218).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 218).font = Font(size = "8", name='Barlow')

                if str(itemOtrosProductos) == "Microdatos":
                    ws.cell(row = cont, column = 221).value = str(itemOtrosProductos)
                    ws.cell(row = cont, column = 221).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 221).font = Font(size = "8", name='Barlow')
                
                if str(itemOtrosProductos) == "Documentos metodológicos":
                    ws.cell(row = cont, column = 224).value = str(itemOtrosProductos)
                    ws.cell(row = cont, column = 224).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 224).font = Font(size = "8", name='Barlow')
                
                if str(itemOtrosProductos) == "Calendario de difusión":
                    ws.cell(row = cont, column = 226).value = str(itemOtrosProductos)
                    ws.cell(row = cont, column = 226).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 226).font = Font(size = "8", name='Barlow')
                    
                if str(itemOtrosProductos) == "Ninguna":
                    ws.cell(row = cont, column = 227).value = str(itemOtrosProductos)
                    ws.cell(row = cont, column = 227).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 227).font = Font(size = "8", name='Barlow')

            if str(ooee.conoce_otra) == "True":
                ws.cell(row = cont, column = 228).value = "Si"
                ws.cell(row = cont, column = 228).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 228).font = Font(size = "8", name='Barlow')
            elif str(ooee.conoce_otra) == "False":
                ws.cell(row = cont, column = 228).value = "No"
                ws.cell(row = cont, column = 228).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 228).font = Font(size = "8", name='Barlow') 

            if str(ooee.hp_siste_infor) == "True":
                ws.cell(row = cont, column = 231).value = "Si"
                ws.cell(row = cont, column = 231).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 231).font = Font(size = "8", name='Barlow')
            elif str(ooee.hp_siste_infor) == "False":
                ws.cell(row = cont, column = 231).value = "No"
                ws.cell(row = cont, column = 231).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 231).font = Font(size = "8", name='Barlow') 

            ws.cell(row = cont, column = 233).value = ooee.observaciones
            ws.cell(row = cont, column = 233).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 233).font = Font(size = "8", name='Barlow') 

            ws.cell(row = cont, column = 234).value = str(ooee.anexos)
            ws.cell(row = cont, column = 234).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 234).font = Font(size = "8", name='Barlow')    
            
            ws.cell(row = cont, column = 235).value = str(ooee.estado_oe_tematico)
            ws.cell(row = cont, column = 235).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 235).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 236).value = str(ooee.validacion_oe_tematico)
            ws.cell(row = cont, column = 236).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 236).font = Font(size = "8", name='Barlow')
           

            ws.cell(row = cont, column = 237).value = str(ooee.nombre_est)
            ws.cell(row = cont, column = 237).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 237).font = Font(size = "8", name='Barlow')
        
            ## Hoja 1 SI desea continuar despues de la linea 237  ------------------------->  

            cont+=1
        #print("----------filter ------------>", normaTextList)
        for indexNorma, itemNorma in enumerate(normaTextList):
            
            indexNorma = indexNorma + 4

            ws.cell(row = indexNorma, column = 19).value = str(itemNorma.cp_d)
            ws.cell(row = indexNorma, column = 19).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 19).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexNorma, column = 20).value = str(itemNorma.ley_d)
            ws.cell(row = indexNorma, column = 20).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 20).font = Font(size = "8", name='Barlow')
                    
            ws.cell(row = indexNorma, column = 21).value = str(itemNorma.decreto_d)
            ws.cell(row = indexNorma, column = 21).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 21).font = Font(size = "8", name='Barlow')
                    
            ws.cell(row = indexNorma, column = 22).value = str(itemNorma.otra_d)
            ws.cell(row = indexNorma, column = 22).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 22).font = Font(size = "8", name='Barlow') 

        for indexRequerimiento, itemRequerimiento in enumerate(requerimientoTextList):

            indexRequerimiento = indexRequerimiento + 4

            ws.cell(row = indexRequerimiento, column = 24).value = str(itemRequerimiento.ri_ods)
            ws.cell(row = indexRequerimiento, column = 24).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexRequerimiento, column = 24).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexRequerimiento, column = 25).value = str(itemRequerimiento.ri_ocde)
            ws.cell(row = indexRequerimiento, column = 25).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexRequerimiento, column = 25).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexRequerimiento, column = 26).value = str(itemRequerimiento.ri_ci)
            ws.cell(row = indexRequerimiento, column = 26).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexRequerimiento, column = 26).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexRequerimiento, column = 27).value = str(itemRequerimiento.ri_pnd)
            ws.cell(row = indexRequerimiento, column = 27).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexRequerimiento, column = 27).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexRequerimiento, column = 28).value = str(itemRequerimiento.ri_cem)
            ws.cell(row = indexRequerimiento, column = 28).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  
            ws.cell(row = indexRequerimiento, column = 28).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexRequerimiento, column = 29).value = str(itemRequerimiento.ri_pstc)
            ws.cell(row = indexRequerimiento, column = 29).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexRequerimiento, column = 29).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexRequerimiento, column = 30).value = str(itemRequerimiento.ri_otro)
            ws.cell(row = indexRequerimiento, column = 30).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexRequerimiento, column = 30).font = Font(size = "8", name='Barlow')

        for indexPriUser, itemPriUser in enumerate(usuariosPrinTextList):

            indexPriUser = indexPriUser + 4

            ws.cell(row = indexPriUser, column = 32).value = str(itemPriUser.org_int)
            ws.cell(row = indexPriUser, column = 32).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 32).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexPriUser, column = 33).value = str(itemPriUser.pres_rep)
            ws.cell(row = indexPriUser, column = 33).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 33).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexPriUser, column = 34).value = str(itemPriUser.misnit)
            ws.cell(row = indexPriUser, column = 34).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 34).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexPriUser, column = 35).value = str(itemPriUser.org_cont)
            ws.cell(row = indexPriUser, column = 35).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 35).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexPriUser, column = 36).value = str(itemPriUser.o_ent_o_nac)
            ws.cell(row = indexPriUser, column = 36).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 36).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexPriUser, column = 37).value = str(itemPriUser.ent_o_terr)
            ws.cell(row = indexPriUser, column = 37).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 37).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexPriUser, column = 38).value = str(itemPriUser.gremios)
            ws.cell(row = indexPriUser, column = 38).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 38).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexPriUser, column = 39).value = str(itemPriUser.ent_privadas)
            ws.cell(row = indexPriUser, column = 39).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 39).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexPriUser, column = 40).value = str(itemPriUser.dep_misma_entidad)
            ws.cell(row = indexPriUser, column = 40).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 40).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexPriUser, column = 41).value = str(itemPriUser.academia)
            ws.cell(row = indexPriUser, column = 41).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 41).font = Font(size = "8", name='Barlow')

        for indexUniObs, itemUniObs in enumerate(unidadObserTextList):

            indexUniObs = indexUniObs + 4

            ws.cell(row = indexUniObs, column = 51).value = str(itemUniObs.mc_otra)
            ws.cell(row = indexUniObs, column = 51).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexUniObs, column = 51).font = Font(size = "8", name='Barlow')

        for indexObtencionDato, itemObtencionDato in enumerate(obtencionDatoTextList):
    
            indexObtencionDato = indexObtencionDato + 4

            ## si la opcion marcada es Registro Administrativo
            ws.cell(row = indexObtencionDato, column = 59).value = str(itemObtencionDato.mc_ra_cual)
            ws.cell(row = indexObtencionDato, column = 59).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexObtencionDato, column = 59).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexObtencionDato, column = 60).value = str(itemObtencionDato.mc_ra_entidad)
            ws.cell(row = indexObtencionDato, column = 60).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexObtencionDato, column = 60).font = Font(size = "8", name='Barlow')
                        
            ## si la opcion marcada es operacion estadística
            ws.cell(row = indexObtencionDato, column = 63).value = str(itemObtencionDato.mc_oe_cual)
            ws.cell(row = indexObtencionDato, column = 63).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexObtencionDato, column = 63).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexObtencionDato, column = 64).value = str(itemObtencionDato.mc_oe_entidad)
            ws.cell(row = indexObtencionDato, column = 64).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexObtencionDato, column = 64).font = Font(size = "8", name='Barlow')

        for indexMuesProbabilistico, itemMuesProbabilistico  in enumerate(muestProbaTextList):

            indexMuesProbabilistico = indexMuesProbabilistico + 4

            ws.cell(row = indexMuesProbabilistico, column = 70).value = str(itemMuesProbabilistico.prob_otro)
            ws.cell(row = indexMuesProbabilistico, column = 70).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexMuesProbabilistico, column = 70).font = Font(size = "8", name='Barlow')

        for indexMuesNoProbabilistico, itemMuesNoProbabilistico in enumerate(muestNoProbaTextList):

            indexMuesNoProbabilistico = indexMuesNoProbabilistico + 4

            ws.cell(row = indexMuesNoProbabilistico, column = 76).value = str(itemMuesNoProbabilistico.no_prob_otro)
            ws.cell(row = indexMuesNoProbabilistico, column = 76).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexMuesNoProbabilistico, column = 76).font = Font(size = "8", name='Barlow')

        for indexTipMar, itemTipMar in enumerate(tipoMarcoTextList):

            indexTipMar = indexTipMar + 4
                
            ws.cell(row = indexTipMar, column = 82).value = str(itemTipMar.otro_tipo_marco)
            ws.cell(row = indexTipMar, column = 82).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexTipMar, column = 82).font = Font(size = "8", name='Barlow')

        for indexdocsDesa, itemdocsDesa in enumerate(docsDesarrolloTextList):

            indexdocsDesa = indexdocsDesa + 4
                
            ws.cell(row = indexdocsDesa, column = 89).value = str(itemdocsDesa.otro_docs)
            ws.cell(row = indexdocsDesa, column = 89).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdocsDesa, column = 89).font = Font(size = "8", name='Barlow')

        for indexConceptosEstanda, itemConceptosEstanda in enumerate(conceptosEstandaTextList):
            
            indexConceptosEstanda = indexConceptosEstanda + 4 

            ws.cell(row = indexConceptosEstanda, column = 93).value = str(itemConceptosEstanda.org_in_cuales)
            ws.cell(row = indexConceptosEstanda, column = 93).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConceptosEstanda, column = 93).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexConceptosEstanda, column = 95).value = str(itemConceptosEstanda.ent_ordnac_cuales)
            ws.cell(row = indexConceptosEstanda, column = 95).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConceptosEstanda, column = 95).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexConceptosEstanda, column = 97).value = str(itemConceptosEstanda.leye_dec_cuales)
            ws.cell(row = indexConceptosEstanda, column = 97).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConceptosEstanda, column = 97).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexConceptosEstanda, column = 100).value = str(itemConceptosEstanda.otra_cual_conp)
            ws.cell(row = indexConceptosEstanda, column = 100).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConceptosEstanda, column = 100).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexConceptosEstanda, column = 102).value = str(itemConceptosEstanda.ningu_pq)
            ws.cell(row = indexConceptosEstanda, column = 102).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConceptosEstanda, column = 102).font = Font(size = "8", name='Barlow')

        for indexClasificaciones, itemClasificaciones in enumerate(clasificacionesTextList):
    
            indexClasificaciones = indexClasificaciones + 4

            ws.cell(row = indexClasificaciones, column = 105).value = str(itemClasificaciones.otra_cual_clas)
            ws.cell(row = indexClasificaciones, column = 105).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexClasificaciones, column = 105).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexClasificaciones, column = 106).value = str(itemClasificaciones.no_pq)
            ws.cell(row = indexClasificaciones, column = 106).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexClasificaciones, column = 106).font = Font(size = "8", name='Barlow')

        for indexCoberturaGeog, itemCoberturaGeog in enumerate(coberGeogTextList):

            indexCoberturaGeog = indexCoberturaGeog + 4
                
            ws.cell(row = indexCoberturaGeog, column = 109).value = str(itemCoberturaGeog.tot_regional)
            ws.cell(row = indexCoberturaGeog, column = 109).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 109).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 110).value = str(itemCoberturaGeog.cual_regional)
            ws.cell(row = indexCoberturaGeog, column = 110).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 110).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 112).value = str(itemCoberturaGeog.tot_dep)
            ws.cell(row = indexCoberturaGeog, column = 112).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 112).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 113).value = str(itemCoberturaGeog.cual_dep)
            ws.cell(row = indexCoberturaGeog, column = 113).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 113).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 115).value = str(itemCoberturaGeog.tot_are_metr)
            ws.cell(row = indexCoberturaGeog, column = 115).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 115).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 116).value = str(itemCoberturaGeog.cual_are_metr)
            ws.cell(row = indexCoberturaGeog, column = 116).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 116).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 118).value = str(itemCoberturaGeog.tot_mun)
            ws.cell(row = indexCoberturaGeog, column = 118).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 118).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 119).value = str(itemCoberturaGeog.cual_mun)
            ws.cell(row = indexCoberturaGeog, column = 119).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 119).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 121).value = str(itemCoberturaGeog.tot_otro)
            ws.cell(row = indexCoberturaGeog, column = 121).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 121).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 122).value = str(itemCoberturaGeog.cual_otro)
            ws.cell(row = indexCoberturaGeog, column = 122).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 122).font = Font(size = "8", name='Barlow')

        for indexdesagreInfo, itemdesagreInfo in enumerate(desagreInfoTextList):

            indexdesagreInfo = indexdesagreInfo + 4

            ws.cell(row = indexdesagreInfo, column = 125).value = str(itemdesagreInfo.des_tot_regional)
            ws.cell(row = indexdesagreInfo, column = 125).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 125).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexdesagreInfo, column = 126).value = str(itemdesagreInfo.des_cual_regional)
            ws.cell(row = indexdesagreInfo, column = 126).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 126).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexdesagreInfo, column = 128).value = str(itemdesagreInfo.des_tot_dep)
            ws.cell(row = indexdesagreInfo, column = 128).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 128).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 129).value = str(itemdesagreInfo.des_cual_dep)
            ws.cell(row = indexdesagreInfo, column = 129).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 129).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 131).value = str(itemdesagreInfo.des_tot_are_metr)
            ws.cell(row = indexdesagreInfo, column = 131).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 131).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 132).value = str(itemdesagreInfo.des_cual_are_metr)
            ws.cell(row = indexdesagreInfo, column = 132).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 132).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 134).value = str(itemdesagreInfo.des_tot_mun)
            ws.cell(row = indexdesagreInfo, column = 134).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 134).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 135).value = str(itemdesagreInfo.des_cual_mun)
            ws.cell(row = indexdesagreInfo, column = 135).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 135).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 137).value = str(itemdesagreInfo.des_tot_otro)
            ws.cell(row = indexdesagreInfo, column = 137).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 137).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 138).value = str(itemdesagreInfo.des_cual_otro)
            ws.cell(row = indexdesagreInfo, column = 138).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 138).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 148).value = str(itemdesagreInfo.des_grupo_otro)
            ws.cell(row = indexdesagreInfo, column = 148).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 148).font = Font(size = "8", name='Barlow')

        for indexFuenteFinan, itemFuenteFinan in enumerate(fuenteFinanTextList):

            indexFuenteFinan = indexFuenteFinan + 4

            ws.cell(row = indexFuenteFinan, column = 156).value = str(itemFuenteFinan.r_otros)
            ws.cell(row = indexFuenteFinan, column = 156).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexFuenteFinan, column = 156).font = Font(size = "8", name='Barlow')

        for indexMedioDatos, itemMedioDatos in enumerate(medioDatosTextList):

            indexMedioDatos = indexMedioDatos + 4
                
            ws.cell(row = indexMedioDatos, column = 163).value = str(itemMedioDatos.sis_info)
            ws.cell(row = indexMedioDatos, column = 163).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexMedioDatos, column = 163).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexMedioDatos, column = 168).value = str(itemMedioDatos.md_otro)
            ws.cell(row = indexMedioDatos, column = 168).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexMedioDatos, column = 168).font = Font(size = "8", name='Barlow')

        for indexPeriodicidadOE, itemPeriodicidadOE in enumerate(periodicidadTextList):

            indexPeriodicidadOE = indexPeriodicidadOE + 4

            ws.cell(row = indexMedioDatos, column = 175).value = str(itemPeriodicidadOE.per_otro)
            ws.cell(row = indexMedioDatos, column = 175).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexMedioDatos, column = 175).font = Font(size = "8", name='Barlow')

        for indexHerraProcesami, itemHerraProcesami in enumerate(herraProcesamiTextList):

            indexHerraProcesami = indexHerraProcesami + 4

            ws.cell(row = indexHerraProcesami, column = 184).value = str(itemHerraProcesami.herr_otro)
            ws.cell(row = indexHerraProcesami, column = 184).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexHerraProcesami, column = 184).font = Font(size = "8", name='Barlow')


        for indexAnalResul, itemAnalResul in enumerate(analResultTextList):
           
            indexAnalResul = indexAnalResul + 4

            ws.cell(row = indexAnalResul, column = 190).value = str(itemAnalResul.ana_otro)
            ws.cell(row = indexAnalResul, column = 190).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexAnalResul, column = 190).font = Font(size = "8", name='Barlow')

        for indexMedioDifus, itemMedioDifus in enumerate(medioDifusTextList):
            
            indexMedioDifus = indexMedioDifus + 4

            ws.cell(row = indexMedioDifus, column = 195).value = str(itemMedioDifus.dif_otro)
            ws.cell(row = indexMedioDifus, column = 195).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexMedioDifus, column = 195).font = Font(size = "8", name='Barlow')

        for indexFechaPublic, itemFechaPublic in enumerate(fechaPublicTextList):

            indexFechaPublic = indexFechaPublic + 4

            ws.cell(row = indexFechaPublic, column = 200).value = str(itemFechaPublic.fecha)
            ws.cell(row = indexFechaPublic, column = 200).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexFechaPublic, column = 200).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexFechaPublic, column = 203).value = str(itemFechaPublic.no_hay)
            ws.cell(row = indexFechaPublic, column = 203).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexFechaPublic, column = 203).font = Font(size = "8", name='Barlow')

        for indexFrecuencDifus, itemFrecuencDifus in enumerate(frecuenciaDifusionTextList):

            indexFrecuencDifus = indexFrecuencDifus + 4

            ws.cell(row = indexFrecuencDifus, column = 210).value = str(itemFrecuencDifus.no_definido)
            ws.cell(row = indexFrecuencDifus, column = 210).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexFrecuencDifus, column = 210).font = Font(size = "8", name='Barlow')

        for indexProductDifund, itemProductDifund in enumerate(productosDifundirTextList):

            indexProductDifund = indexProductDifund + 4

            ws.cell(row = indexProductDifund, column = 217).value = str(itemProductDifund.difundir_otro)
            ws.cell(row = indexProductDifund, column = 217).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexProductDifund, column = 217).font = Font(size = "8", name='Barlow')

        for indexOtrosProduct, itemOtrosProduct in enumerate(otrosProductosTextList):

            indexOtrosProduct = indexOtrosProduct + 4

            ws.cell(row = indexOtrosProduct, column = 219).value = str(itemOtrosProduct.ser_hist_desde)
            ws.cell(row = indexOtrosProduct, column = 219).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexOtrosProduct, column = 219).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexOtrosProduct, column = 220).value = str(itemOtrosProduct.ser_hist_hasta)
            ws.cell(row = indexOtrosProduct, column = 220).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexOtrosProduct, column = 220).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexOtrosProduct, column = 222).value = str(itemOtrosProduct.microdatos_desde)
            ws.cell(row = indexOtrosProduct, column = 222).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexOtrosProduct, column = 222).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexOtrosProduct, column = 223).value = str(itemOtrosProduct.microdatos_hasta)
            ws.cell(row = indexOtrosProduct, column = 223).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexOtrosProduct, column = 223).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexOtrosProduct, column = 225).value = str(itemOtrosProduct.op_url)
            ws.cell(row = indexOtrosProduct, column = 225).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexOtrosProduct, column = 225).font = Font(size = "8", name='Barlow')

        for indexResultSimi, itemResultSimi in enumerate(resultadosSimiTextList):
            
            indexResultSimi = indexResultSimi + 4

            ws.cell(row = indexResultSimi, column = 229).value = str(itemResultSimi.rs_entidad)
            ws.cell(row = indexResultSimi, column = 229).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexResultSimi, column = 229).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexResultSimi, column = 230).value = str(itemResultSimi.rs_oe)
            ws.cell(row = indexResultSimi, column = 230).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexResultSimi, column = 230).font = Font(size = "8", name='Barlow')
            
        for indexhpSistemaInfo, itemhpSistemaInfo in enumerate(hpSistemaInfoTextList):
            
            indexhpSistemaInfo = indexhpSistemaInfo + 4

            ws.cell(row = indexhpSistemaInfo, column = 232).value = str(itemhpSistemaInfo.si_cual)
            ws.cell(row = indexhpSistemaInfo, column = 232).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexhpSistemaInfo, column = 232).font = Font(size = "8", name='Barlow')

        for row in range(4, ws.max_row + 1):
            ws.row_dimensions[row].height = 90
            
        #############  Hoja 2 entidadades y fases  ##################

        sheet2.merge_cells('A2:B2')
        sheet2.merge_cells('C2:HZ2')
        sheet2.merge_cells('A1:HZ1')

        def set_border(sheet2, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet2[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet2,'A1:HZ'+str(ooees.count()+3))
    
        #row dimensions
        sheet2.row_dimensions[1].height = 55
        sheet2.row_dimensions[2].height = 40

        # column width
        sheet2.column_dimensions['A'].width = 20
        sheet2.column_dimensions['B'].width = 20
        sheet2.column_dimensions['C'].width = 20
        sheet2.column_dimensions['D'].width = 20
        sheet2.column_dimensions['E'].width = 20
        sheet2.column_dimensions['F'].width = 20
        sheet2.column_dimensions['G'].width = 20
        sheet2.column_dimensions['H'].width = 20
        sheet2.column_dimensions['I'].width = 20
        sheet2.column_dimensions['J'].width = 20
        sheet2.column_dimensions['K'].width = 20
        sheet2.column_dimensions['L'].width = 20
        sheet2.column_dimensions['M'].width = 20
        sheet2.column_dimensions['N'].width = 20
        sheet2.column_dimensions['O'].width = 20
        sheet2.column_dimensions['P'].width = 20
        sheet2.column_dimensions['Q'].width = 20
        sheet2.column_dimensions['R'].width = 20
        sheet2.column_dimensions['S'].width = 20
        sheet2.column_dimensions['T'].width = 20
        sheet2.column_dimensions['U'].width = 20
        sheet2.column_dimensions['V'].width = 20
        sheet2.column_dimensions['W'].width = 20
        sheet2.column_dimensions['X'].width = 20
        sheet2.column_dimensions['Y'].width = 20
        sheet2.column_dimensions['Z'].width = 20
    
        title_cell = sheet2['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de las Operaciones Estadísticas'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')
	
        codigo_cell = sheet2['A2']
        codigo_cell.value = 'A. OPERACIÓN ESTADÍSTICA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet2['C2']
        codigo_cell.value = '¿Otra entidad es responsable de una o varias fases de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet2['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = sheet2['B3']
        codigo_cell.value = 'Nombre Operación Estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet2['C3']
        codigo_cell.value = 'Indique si hay otra(s) entidad(es) que participa(n) en alguna de las fases del proceso estadístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        contSheet2 = 4
        for ooee in ooees: 
            sheet2.cell(row = contSheet2, column = 1).value = ooee.codigo_oe
            sheet2.cell(row = contSheet2, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet2.cell(row = contSheet2, column = 1).font = Font(size = "8", name='Barlow')

            sheet2.cell(row = contSheet2, column = 2).value = ooee.nombre_oe
            sheet2.cell(row = contSheet2, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet2.cell(row = contSheet2, column = 2).font = Font(size = "8", name='Barlow')
            
            if str(ooee.fase) == "True":
                sheet2.cell(row = contSheet2, column = 3).value = "Si"
                sheet2.cell(row = contSheet2, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet2.cell(row = contSheet2, column = 3).font = Font(size = "8", name='Barlow')
            elif str(ooee.fase) == "False":
                sheet2.cell(row = contSheet2, column = 3).value = "No"
                sheet2.cell(row = contSheet2, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet2.cell(row = contSheet2, column = 3).font = Font(size = "8", name='Barlow')


            contSheet2+= 1
            
        entidadoeid = []
        for indexEnFas, itemEnfas in enumerate(entidadesFasesTextList):
            indexEnFas = itemEnfas.ooee_id + 3
                    
            entidadoeid.append(itemEnfas.ooee_id)
            #print("array",set(entidadoeid))
            #print("sinset", entidadoeid )
            count_id = entidadoeid.count(itemEnfas.ooee_id)
            c=0 #inicializamos el contador  
            n=6*count_id
            for i in range(1,n+1):  
                if i%6 == 0:  
                    i = i - 2
                    c+=1

            for indexoe, itemoe in enumerate(ooees):
                if itemEnfas.ooee_id == itemoe.pk:
                    posRow = indexoe + 4
                     
                    sheet2.cell(row = 3, column = i ).value = "Nombre de Entidad"
                    sheet2.cell(row = 3, column = i ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = 3, column = i ).font = Font(bold=True) 

                    sheet2.cell(row = posRow, column = i ).value = str(itemEnfas.nombre_entifas) 
                    sheet2.cell(row = posRow, column = i ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = posRow, column = i ).font = Font(size = "8", name='Barlow') 
            
            listaFases = list(itemEnfas.fases.all())  ##iterar para traer las fases seleccionadas

            for indexFase, itemFase in enumerate(listaFases):

                if str(itemFase) == "Detección y análisis de requerimientos":
                    
                    countFase=0 #inicializamos el contador  
                    faseSelected=6*count_id 
                    for posFase in range(1,n+1):  
                        if posFase%6 == 0:
                            posFase = posFase - 1
                            countFase+=1

                    sheet2.cell(row = 3, column = posFase ).value = "a. Detección y análisis de requerimientos"
                    sheet2.cell(row = 3, column = posFase ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = 3, column = posFase ).font = Font(bold=True)

                    sheet2.cell(row = posRow, column = posFase).value = str(itemFase)
                    sheet2.cell(row = posRow, column = posFase).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = posRow, column = posFase).font = Font(size = "8", name='Barlow')
                                
                if str(itemFase) == "Diseño y pruebas":

                    countFase=0 #inicializamos el contador  
                    faseSelected=6*count_id 
                    for posFase in range(1,n+1):  
                        if posFase%6 == 0:  
                            posFase = posFase
                            countFase+=1

                    sheet2.cell(row = 3, column = posFase ).value = "b. Diseño y pruebas"
                    sheet2.cell(row = 3, column = posFase ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = 3, column = posFase ).font = Font(bold=True)

                    sheet2.cell(row = posRow, column = posFase).value = str(itemFase)
                    sheet2.cell(row = posRow, column = posFase).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = posRow, column = posFase).font = Font(size = "8", name='Barlow')

                if str(itemFase) == "Ejecución":

                    countFase=0 #inicializamos el contador  
                    faseSelected=6*count_id 
                    for posFase in range(1,n+1):  
                        if posFase%6 == 0:  
                            posFase = posFase + 1
                            countFase+=1

                    sheet2.cell(row = 3, column = posFase ).value = "c. Ejecución"
                    sheet2.cell(row = 3, column = posFase ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = 3, column = posFase ).font = Font(bold=True)

                    sheet2.cell(row = posRow, column = posFase).value = str(itemFase)
                    sheet2.cell(row = posRow, column = posFase).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = posRow, column = posFase).font = Font(size = "8", name='Barlow')
                
                if str(itemFase) == "Análisis":

                    countFase=0 #inicializamos el contador  
                    faseSelected=6*count_id 
                    for posFase in range(1,n+1):  
                        if posFase%6 == 0:  
                            posFase = posFase + 2
                            countFase+=1

                    sheet2.cell(row = 3, column = posFase ).value = "d. Análisis"
                    sheet2.cell(row = 3, column = posFase ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = 3, column = posFase ).font = Font(bold=True)

                    sheet2.cell(row = posRow, column = posFase).value = str(itemFase)
                    sheet2.cell(row = posRow, column = posFase).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = posRow, column = posFase).font = Font(size = "8", name='Barlow')
                
                if str(itemFase) == "Difusión":

                    countFase=0 #inicializamos el contador  
                    faseSelected=6*count_id 
                    for posFase in range(1,n+1):  
                        if posFase%6 == 0:  
                            posFase = posFase + 3
                            countFase+=1

                    sheet2.cell(row = 3, column = posFase ).value = "e. Difusión"
                    sheet2.cell(row = 3, column = posFase ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = 3, column = posFase ).font = Font(bold=True)
                    
                    sheet2.cell(row = posRow, column = posFase).value = str(itemFase)
                    sheet2.cell(row = posRow, column = posFase).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = posRow, column = posFase).font = Font(size = "8", name='Barlow')
                               
        for row in range(4, sheet2.max_row + 1):
            sheet2.row_dimensions[row].height = 90   
        
                    
        #############  Hoja 3 lista de variables  ##################

        sheet3.merge_cells('A1:C1')
        sheet3.merge_cells('A2:C2')

        def set_border(sheet3, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet3[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet3,'A1:C1')
    
        #row dimensions
        sheet3.row_dimensions[1].height = 55
        sheet3.row_dimensions[2].height = 40

        # column width
        sheet3.column_dimensions['A'].width = 20
        sheet3.column_dimensions['B'].width = 20
        sheet3.column_dimensions['C'].width = 20
        
        title_cell = sheet3['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de las Operaciones Estadísticas'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        title_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
	
        codigo_cell = sheet3['A2']
        codigo_cell.value = 'Lista de variables de la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)


        codigo_cell = sheet3['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = sheet3['B3']
        codigo_cell.value = 'Nombre Operación Estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet3['C3']
        codigo_cell.value = 'Variables que maneja la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        contsheet3 = 4

        arrayoeId = []
        for indexVariab, itemVariab in enumerate(listaDeVariablesText):
            indexVariab =  itemVariab.ooee_id + 3
            arrayoeId.append(itemVariab.ooee_id)
            
            for oest in ooees:
                if str(itemVariab.ooee) == str(oest.nombre_oe):
            
                    sheet3.cell(row = contsheet3, column = 1).value = oest.codigo_oe
                    sheet3.cell(row = contsheet3, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet3.cell(row = contsheet3, column = 1).font = Font(size = "8", name='Barlow')

                    sheet3.cell(row = contsheet3, column = 2).value = str(itemVariab.ooee)
                    sheet3.cell(row = contsheet3, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet3.cell(row = contsheet3, column = 2).font = Font(size = "8", name='Barlow')

                    sheet3.cell(row = contsheet3, column = 3).value = str(itemVariab.lista_var)
                    sheet3.cell(row = contsheet3, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet3.cell(row = contsheet3, column = 3).font = Font(size = "8", name='Barlow')
                    contsheet3+=1
           
        for row in range(4, sheet3.max_row + 1):  ## definir tamaño de rows
            sheet3.row_dimensions[row].height = 90

         #############  end Hoja 3 variables  ##################

        #############  Hoja 4 resultados  ##################

        sheet4.merge_cells('A1:C1')
        sheet4.merge_cells('A2:C2')

        def set_border(sheet4, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet4[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet4,'A1:C1')
    
        #row dimensions
        sheet4.row_dimensions[1].height = 55
        sheet4.row_dimensions[2].height = 40
        sheet4.row_dimensions[3].height = 40

        # column width
        sheet4.column_dimensions['A'].width = 20
        sheet4.column_dimensions['B'].width = 20
        sheet4.column_dimensions['C'].width = 20
    
        title_cell = sheet4['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de las Operaciones Estadísticas'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
	
        codigo_cell = sheet4['A2']
        codigo_cell.value = 'Resultados agregados o indicadores calculados'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet4['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = sheet4['B3']
        codigo_cell.value = 'Nombre Operación Estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet4['C3']
        codigo_cell.value = 'Lista de resultados agregados o indicadores calculados'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        contsheet4 = 4
        listaIdoe = []
        for indexResultEsta, itemResultEsta in enumerate(listaResultEstText):
            indexResultEsta =  itemResultEsta.ooee_id + 3
            listaIdoe.append(itemResultEsta.ooee_id)

            for oest in ooees:
                if str(itemResultEsta.ooee) == str(oest.nombre_oe):
            
                    sheet4.cell(row = contsheet4, column = 1).value = oest.codigo_oe
                    sheet4.cell(row = contsheet4, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = contsheet4, column = 1).font = Font(size = "8", name='Barlow')

                    sheet4.cell(row = contsheet4, column = 2).value = str(itemResultEsta.ooee)
                    sheet4.cell(row = contsheet4, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = contsheet4, column = 2).font = Font(size = "8", name='Barlow')

                    sheet4.cell(row = contsheet4, column = 3).value = str(itemResultEsta.resultEstad)
                    sheet4.cell(row = contsheet4, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = contsheet4, column = 3).font = Font(size = "8", name='Barlow')
                    contsheet4+=1

        for row in range(4, sheet4.max_row + 1):  ## definir tamaño de rows
            sheet4.row_dimensions[row].height = 90

        ############# END Hoja 4 resultados  ##################


        ############## Hoja 5 Evaluación de calidad #############

        
        sheet5.merge_cells('A1:AZ1')
        sheet5.merge_cells('A2:F2')
        sheet5.merge_cells('G2:AZ2')

        def set_border(sheet5, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet5[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet5,'A1:AZ'+str(ooees.count()+3))
        

        #row dimensions
        sheet5.row_dimensions[1].height = 55
        sheet5.row_dimensions[2].height = 40
        sheet5.row_dimensions[3].height = 40

        # column width
        sheet5.column_dimensions['A'].width = 20
        sheet5.column_dimensions['B'].width = 18
        sheet5.column_dimensions['C'].width = 15
        sheet5.column_dimensions['D'].width = 24
        sheet5.column_dimensions['E'].width = 15
        sheet5.column_dimensions['F'].width = 24
        sheet5.column_dimensions['G'].width = 22
        sheet5.column_dimensions['H'].width = 15
        sheet5.column_dimensions['I'].width = 15
        sheet5.column_dimensions['J'].width = 15
        sheet5.column_dimensions['K'].width = 15
        sheet5.column_dimensions['L'].width = 15
        sheet5.column_dimensions['M'].width = 15
        sheet5.column_dimensions['N'].width = 15
        sheet5.column_dimensions['O'].width = 15
        sheet5.column_dimensions['P'].width = 15
        sheet5.column_dimensions['Q'].width = 15
        sheet5.column_dimensions['R'].width = 22
        sheet5.column_dimensions['S'].width = 15
        sheet5.column_dimensions['T'].width = 15
        sheet5.column_dimensions['U'].width = 15
        sheet5.column_dimensions['V'].width = 15
        sheet5.column_dimensions['W'].width = 15
        sheet5.column_dimensions['X'].width = 15
        sheet5.column_dimensions['Y'].width = 20
        sheet5.column_dimensions['Z'].width = 15

        sheet5.column_dimensions['AA'].width = 15
        sheet5.column_dimensions['AB'].width = 15
        sheet5.column_dimensions['AC'].width = 15
        sheet5.column_dimensions['AD'].width = 15
        sheet5.column_dimensions['AE'].width = 15
        sheet5.column_dimensions['AF'].width = 15
        sheet5.column_dimensions['AG'].width = 15
        sheet5.column_dimensions['AH'].width = 15
        sheet5.column_dimensions['AI'].width = 15
        sheet5.column_dimensions['AJ'].width = 15
        sheet5.column_dimensions['AK'].width = 15
        sheet5.column_dimensions['AL'].width = 15
        sheet5.column_dimensions['AM'].width = 15
        sheet5.column_dimensions['AN'].width = 15
        sheet5.column_dimensions['AO'].width = 15
        sheet5.column_dimensions['AP'].width = 15
        sheet5.column_dimensions['AQ'].width = 15
        sheet5.column_dimensions['AR'].width = 15
        sheet5.column_dimensions['AS'].width = 15
        sheet5.column_dimensions['AT'].width = 15
        sheet5.column_dimensions['AU'].width = 15
        sheet5.column_dimensions['AV'].width = 15
        sheet5.column_dimensions['AW'].width = 15
        sheet5.column_dimensions['AX'].width = 15
        sheet5.column_dimensions['AY'].width = 15
        sheet5.column_dimensions['AZ'].width = 15

        title_cell = sheet5['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de las Operaciones Estadísticas'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet5['A2']
        codigo_cell.value = 'IDENTIFICACIÓN'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet5['G2']
        codigo_cell.value = 'EVALUACIÓN DE CALIDAD'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['A3']
        resultados_cell.value = 'Área Temática'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['B3']
        resultados_cell.value = 'Tema'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['C3']
        resultados_cell.value = 'Código Entidad'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['D3']
        resultados_cell.value = 'Nombre Entidad'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['E3']
        resultados_cell.value = 'Código OOEE'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['F3']
        resultados_cell.value = 'Nombre de la Operación Estadística'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        resultados_cell = sheet5['G3']
        resultados_cell.value = 'Número de Evaluaciones'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        contsheet5 = 4

        for ooee in ooees: 


            sheet5.cell(row = contsheet5, column = 1).value = str(ooee.area_tematica)
            sheet5.cell(row = contsheet5, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 1).font = Font(size = "8", name='Barlow') 
            
            sheet5.cell(row = contsheet5, column = 2).value = str(ooee.tema)
            sheet5.cell(row = contsheet5, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 2).font = Font(size = "8", name='Barlow') 
            
            sheet5.cell(row = contsheet5, column = 3).value = ooee.entidad.codigo
            sheet5.cell(row = contsheet5, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 3).font = Font(size = "8", name='Barlow')

            sheet5.cell(row = contsheet5, column = 4).value = str(ooee.entidad)
            sheet5.cell(row = contsheet5, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 4).font = Font(size = "8", name='Barlow')

            sheet5.cell(row = contsheet5, column = 5).value = ooee.codigo_oe
            sheet5.cell(row = contsheet5, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 5).font = Font(size = "8", name='Barlow')

            sheet5.cell(row = contsheet5, column = 6).value = ooee.nombre_oe
            sheet5.cell(row = contsheet5, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 6).font = Font(size = "8", name='Barlow') 
            contsheet5+=1

        idoeEval = []
        for indexEval, itemEval in enumerate(listaEvaluacionText):
            indexEval =  itemEval.post_oe_id + 3
            idoeEval.append(itemEval.post_oe_id)

            ### 1 Estado de la evaluación
            
            count_est_cert = idoeEval.count(itemEval.post_oe_id)
            #print("total evaluaciones",count_est_cert)
            contrField1=0 #inicializamos el contador  
            pos_est_cert = 10*count_est_cert
            for incField1 in range(1,pos_est_cert+1):  
                if incField1%10 == 0:  
                    incField1 = incField1 - 2
                    contrField1+=1

            for indexoe, itemoe in enumerate(ooees):
                if itemEval.post_oe_id == itemoe.pk:
                    posiRow = indexoe + 4

                    sheet5.cell(row = posiRow, column = 7).value = count_est_cert ## número de evaluaciones
                    sheet5.cell(row = posiRow, column = 7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = 7).font = Font(bold=True) 

                    sheet5.cell(row = 3, column = incField1).value = "Estado de la Evaluación"
                    sheet5.cell(row = 3, column = incField1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField1).font = Font(bold=True) 

                    sheet5.cell(row = posiRow, column = incField1).value = str(itemEval.est_evaluacion)
                    sheet5.cell(row = posiRow, column = incField1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField1).font = Font(size = "8", name='Barlow')
  
            ### 2 Resultado año de vigencia


            count_id_oe = idoeEval.count(itemEval.post_oe_id)
            contadorEval=0 #inicializamos el contador  
            posFields = 10*count_id_oe
            #print("______tttt________", posFields)
            for incEval in range(1,posFields+1):  
                if incEval%10 == 0:
                    #print("inceval ",incEval) 
                    incEval = incEval - 1
                    #print("sumado  ",incEval) 
                    contadorEval+=1

            for indexoe, itemoe in enumerate(ooees):
                if itemEval.post_oe_id == itemoe.pk:
                    posiRow = indexoe + 4

                    sheet5.cell(row = 3, column = incEval).value = "Año"
                    sheet5.cell(row = 3, column = incEval ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incEval ).font = Font(bold=True) 

                    if itemEval.year_eva == None:

                        sheet5.cell(row = posiRow, column = incEval).value = ""
                        sheet5.cell(row = posiRow, column = incEval).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incEval).font = Font(size = "8", name='Barlow')
                    else:
                        sheet5.cell(row = posiRow, column = incEval).value = itemEval.year_eva.strftime('%Y')
                        sheet5.cell(row = posiRow, column = incEval).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incEval).font = Font(size = "8", name='Barlow')
            
            ### 3 Observaciones de la evaluación

            contadorField3=0 #inicializamos el contador  
            posFields3 = 10*count_id_oe
            for incField3 in range(1,posFields3+1):  
                if incField3%10 == 0:  
                    incField3 = incField3
                    contadorField3+=1
                
            for indexoe, itemoe in enumerate(ooees):
                if itemEval.post_oe_id == itemoe.pk:
                    posiRow = indexoe + 4

                    sheet5.cell(row = 3, column = incField3).value = "Observaciones"
                    sheet5.cell(row = 3, column = incField3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField3).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField3).value = str(itemEval.observ_est)
                    sheet5.cell(row = posiRow, column = incField3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField3).font = Font(size = "8", name='Barlow')

            ### 4 Metodologia

            contadorField4=0 #inicializamos el contador  
            posFields4 = 10*count_id_oe
            for incField4 in range(1,posFields4+1):  
                if incField4%10 == 0:
                    incField4 = incField4 + 1
                    contadorField4+=1

            for indexoe, itemoe in enumerate(ooees):
                if itemEval.post_oe_id == itemoe.pk:
                    posiRow = indexoe + 4

                    sheet5.cell(row = 3, column = incField4).value = "Metodología"
                    sheet5.cell(row = 3, column = incField4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField4).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField4).value = str(itemEval.metodologia)
                    sheet5.cell(row = posiRow, column = incField4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField4).font = Font(size = "8", name='Barlow')
            
        ### 5 Resultado de la evaluación según metodologia

            contadorField5=0 #inicializamos el contador  
            posFields5 = 10*count_id_oe
            for incField5 in range(1,posFields5+1):  
                if incField5%10 == 0:  
                    incField5 = incField5 + 2
                    contadorField5+=1

            for indexoe, itemoe in enumerate(ooees):
                if itemEval.post_oe_id == itemoe.pk:
                    posiRow = indexoe + 4

                    sheet5.cell(row = 3, column = incField5).value = "Resultado de la evaluación"
                    sheet5.cell(row = 3, column = incField5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField5).font = Font(bold=True)

                    if str(itemEval.metodologia) == "ntcpe 1000":
                        sheet5.cell(row = posiRow, column = incField5).value = str(itemEval.res_evaluacion)
                        sheet5.cell(row = posiRow, column = incField5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField5).font = Font(size = "8", name='Barlow')
                    
                    elif str(itemEval.metodologia) == "matriz de requisitos":
                        sheet5.cell(row = posiRow, column = incField5).value = str(itemEval.res_mzrequi)
                        sheet5.cell(row = posiRow, column = incField5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField5).font = Font(size = "8", name='Barlow')
                    
        ### 6 Observaciones de resultado de la evaluación

            contadorField6=0 #inicializamos el contador  
            posFields6 = 10*count_id_oe
            for incField6 in range(1,posFields6+1):  
                if incField6%10 == 0:  
                    incField6 = incField6 + 3
                    contadorField6+=1

            for indexoe, itemoe in enumerate(ooees):
                if itemEval.post_oe_id == itemoe.pk:
                    posiRow = indexoe + 4

                    sheet5.cell(row = 3, column = incField6).value = "Observaciones"
                    sheet5.cell(row = 3, column = incField6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField6).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField6).value = str(itemEval.observ_resul)
                    sheet5.cell(row = posiRow, column = incField6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField6).font = Font(size = "8", name='Barlow')

        ### 7 Vigencia

            contadorField7=0 #inicializamos el contador  
            posFields7 = 10*count_id_oe
            for incField7 in range(1,posFields7+1):  
                if incField7%10 == 0:  
                    incField7 = incField7 + 4
                    contadorField7+=1

            for indexoe, itemoe in enumerate(ooees):
                if itemEval.post_oe_id == itemoe.pk:
                    posiRow = indexoe + 4

                    sheet5.cell(row = 3, column = incField7).value = "Vigencia"
                    sheet5.cell(row = 3, column = incField7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField7).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField7).value = "Desde: " + str(itemEval.vigencia_desde) + " " + "\n\n Hasta: " + str(itemEval.vigencia_hasta)
                    sheet5.cell(row = posiRow, column = incField7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField7).font = Font(size = "8", name='Barlow')

         ### 8 Plan de Mejoramiento   
            
            contadorField8=0 #inicializamos el contador  
            posFields8 = 10*count_id_oe
            for incField8 in range(1,posFields8+1):  
                if incField8%10 == 0:  
                    incField8 = incField8 + 5
                    contadorField8+=1

            for indexoe, itemoe in enumerate(ooees):
                if itemEval.post_oe_id == itemoe.pk:
                    posiRow = indexoe + 4

                    sheet5.cell(row = 3, column = incField8).value = "Plan de mejoramiento"
                    sheet5.cell(row = 3, column = incField8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField8).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField8).value = str(itemEval.pla_mejoramiento)
                    sheet5.cell(row = posiRow, column = incField8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField8).font = Font(size = "8", name='Barlow')

        ###  9 Seguimiento anual (Vigilancia):   
            
            contadorField9=0 #inicializamos el contador  
            posFields9 = 10*count_id_oe
            for incField9 in range(1,posFields9+1):  
                if incField9%10 == 0:  
                    incField9 = incField9 + 6
                    contadorField9+=1

            for indexoe, itemoe in enumerate(ooees):
                if itemEval.post_oe_id == itemoe.pk:
                    posiRow = indexoe + 4

                    sheet5.cell(row = 3, column = incField9).value = "Seguimiento anual (Vigilancia)"
                    sheet5.cell(row = 3, column = incField9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField9).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField9).value = str(itemEval.seg_vig)
                    sheet5.cell(row = posiRow, column = incField9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField9).font = Font(size = "8", name='Barlow')

        ###  Observaciones seguimiento anual
            
            contadorField10=0 #inicializamos el contador  
            posFields10 = 10*count_id_oe
            for incField10 in range(1,posFields10+1):  
                if incField10%10 == 0:  
                    incField10 = incField10 + 7
                    contadorField10+=1

            for indexoe, itemoe in enumerate(ooees):
                if itemEval.post_oe_id == itemoe.pk:
                    posiRow = indexoe + 4
                
                    sheet5.cell(row = 3, column = incField10).value = "Observaciones"
                    sheet5.cell(row = 3, column = incField10).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField10).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField10).value = str(itemEval.obs_seg_anual)
                    sheet5.cell(row = posiRow, column = incField10).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField10).font = Font(size = "8", name='Barlow')
   
        for row in range(4, sheet5.max_row + 1):  ## definir tamaño de rows
            sheet5.row_dimensions[row].height = 90

            
        ############## End Evaluación de calidad #############
        
        ## ocultar información si no esta autenticado

        if user.is_authenticated == True and user.profile.entidad.pk != 2:
        ############# Hoja 6 Criticas #######################

            sheet6.merge_cells('A1:F1')
            sheet6.merge_cells('A2:F2')

            def set_border(sheet6, cell_range):
                border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                rows = sheet6[cell_range]
                for row in rows:
                    for cell in row:
                        cell.border = border

            set_border(sheet6,'A1:F1')
            
            #row dimensions
            sheet6.row_dimensions[1].height = 55
            sheet6.row_dimensions[2].height = 40
            sheet6.row_dimensions[3].height = 40
            sheet6.row_dimensions[4].height = 40
            sheet6.row_dimensions[5].height = 40
            sheet6.row_dimensions[6].height = 40

            # column width
            sheet6.column_dimensions['A'].width = 20
            sheet6.column_dimensions['B'].width = 20
            sheet6.column_dimensions['C'].width = 20
            sheet6.column_dimensions['D'].width = 20
            sheet6.column_dimensions['E'].width = 20
            sheet6.column_dimensions['F'].width = 20

            title_cell = sheet6['A1']
            title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
            title_cell.value = 'Directorio de las Operaciones Estadísticas'
            title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
            title_cell.alignment =  Alignment(horizontal='center', vertical='center')

            codigo_cell = sheet6['A2']
            codigo_cell.value = 'Críticas'
            codigo_cell.font = Font(bold=True)
            codigo_cell.alignment = Alignment(horizontal='center')
            codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')


            resultados_cell = sheet6['A3']
            resultados_cell.value = 'Código'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet6['B3']
            resultados_cell.value = 'Nombre de la Operación Estadística'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet6['C3']
            resultados_cell.value = 'Estado de la crítica'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet6['D3']
            resultados_cell.value = 'Observaciones de la crítica'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet6['E3']
            resultados_cell.value = 'Funcionario que realiza la crítica'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet6['F3']
            resultados_cell.value = 'Fecha en que se realiza la crítica'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            idoeCritica = []
            contsheet6 = 4
            for indexCritica, itemCritica in enumerate(listaCriticaText):
                indexCritica =  itemCritica.post_oe_id + 3
                idoeCritica.append(itemCritica.post_oe_id)

                for oest in ooees:
                    if str(itemCritica.post_oe) == str(oest.nombre_oe):

                        sheet6.cell(row = contsheet6, column = 1).value = oest.codigo_oe
                        sheet6.cell(row = contsheet6, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet6.cell(row = contsheet6, column = 1).font = Font(bold=True) 

                        sheet6.cell(row = contsheet6, column = 2).value = str(itemCritica.post_oe)
                        sheet6.cell(row = contsheet6, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet6.cell(row = contsheet6, column = 2).font = Font(bold=True) 

                        sheet6.cell(row = contsheet6, column = 3).value = str(itemCritica.estado_crit)
                        sheet6.cell(row = contsheet6, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet6.cell(row = contsheet6, column = 3).font = Font(bold=True)

                        sheet6.cell(row = contsheet6, column = 4).value = str(itemCritica.descrip_critica)
                        sheet6.cell(row = contsheet6, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet6.cell(row = contsheet6, column = 4).font = Font(bold=True)

                        sheet6.cell(row = contsheet6, column = 5).value = str(itemCritica.name_cri)
                        sheet6.cell(row = contsheet6, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet6.cell(row = contsheet6, column = 5).font = Font(bold=True)

                        sheet6.cell(row = contsheet6, column = 6).value = str(itemCritica.fecha_critica)
                        sheet6.cell(row = contsheet6, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet6.cell(row = contsheet6, column = 6).font = Font(bold=True)

                        contsheet6+=1

            for row in range(4, sheet6.max_row + 1):  ## definir tamaño de rows
                sheet6.row_dimensions[row].height = 90

                
            ############## End critica ############# 

            ############## Hoja 7 Novedad #############

            
            sheet7.merge_cells('A1:G1')
            sheet7.merge_cells('A2:G2')

            def set_border(sheet7, cell_range):
                border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                rows = sheet7[cell_range]
                for row in rows:
                    for cell in row:
                        cell.border = border

            set_border(sheet7,'A1:G1')
            

            #row dimensions
            sheet7.row_dimensions[1].height = 55
            sheet7.row_dimensions[2].height = 40
            sheet7.row_dimensions[3].height = 40
            sheet7.row_dimensions[4].height = 40
            sheet7.row_dimensions[5].height = 40
            sheet7.row_dimensions[6].height = 40
            sheet7.row_dimensions[7].height = 40

            # column width
            sheet7.column_dimensions['A'].width = 20
            sheet7.column_dimensions['B'].width = 20
            sheet7.column_dimensions['C'].width = 20
            sheet7.column_dimensions['D'].width = 20
            sheet7.column_dimensions['E'].width = 20
            sheet7.column_dimensions['F'].width = 20
            sheet7.column_dimensions['G'].width = 20

            title_cell = sheet7['A1']
            title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
            title_cell.value = 'Directorio de las Operaciones Estadísticas'
            title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
            title_cell.alignment =  Alignment(horizontal='center', vertical='center')

            codigo_cell = sheet7['A2']
            codigo_cell.value = 'Novedades'
            codigo_cell.font = Font(bold=True)
            codigo_cell.alignment = Alignment(horizontal='center')
            codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')


            resultados_cell = sheet7['A3']
            resultados_cell.value = 'Código'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet7['B3']
            resultados_cell.value = 'Nombre de la Operación Estadística'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet7['C3']
            resultados_cell.value = 'Novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet7['D3']
            resultados_cell.value = 'Estado de la actualización'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet7['E3']
            resultados_cell.value = 'Descripción de la novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet7['F3']
            resultados_cell.value = 'Funcionario que realiza la novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet7['G3']
            resultados_cell.value = 'Fecha en que se realiza la novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            idoeNovedad = []
            contsheet7 = 4
            for indexNovedad, itemNovedad in enumerate(listaNovedadText):
                indexNovedad =  itemNovedad.post_oe_id + 3
                idoeNovedad.append(itemNovedad.post_oe_id)

                for oest in ooees:
                    if str(itemNovedad.post_oe) == str(oest.nombre_oe):

                        sheet7.cell(row = contsheet7, column = 1).value = oest.codigo_oe
                        sheet7.cell(row = contsheet7, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet7.cell(row = contsheet7, column = 1).font = Font(bold=True) 

                        sheet7.cell(row = contsheet7, column = 2).value = str(itemNovedad.post_oe)
                        sheet7.cell(row = contsheet7, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet7.cell(row = contsheet7, column = 2).font = Font(bold=True) 

                        sheet7.cell(row = contsheet7, column = 3).value = str(itemNovedad.novedad)
                        sheet7.cell(row = contsheet7, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet7.cell(row = contsheet7, column = 3).font = Font(bold=True)

                        sheet7.cell(row = contsheet7, column = 4).value = str(itemNovedad.est_actualiz)
                        sheet7.cell(row = contsheet7, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet7.cell(row = contsheet7, column = 4).font = Font(bold=True)

                        sheet7.cell(row = contsheet7, column = 5).value = str(itemNovedad.descrip_novedad)
                        sheet7.cell(row = contsheet7, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet7.cell(row = contsheet7, column = 5).font = Font(bold=True)

                        sheet7.cell(row = contsheet7, column = 6).value = str(itemNovedad.name_nov)
                        sheet7.cell(row = contsheet7, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet7.cell(row = contsheet7, column = 6).font = Font(bold=True)

                        sheet7.cell(row = contsheet7, column = 7).value = str(itemNovedad.fecha_actualiz)
                        sheet7.cell(row = contsheet7, column = 7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet7.cell(row = contsheet7, column = 7).font = Font(bold=True)

                        contsheet7+=1

            for row in range(4, sheet7.max_row + 1):  ## definir tamaño de rows
                sheet7.row_dimensions[row].height = 90

        ############# end Hoja 7 Novedades #####################


        file_name = "reporte_ooee.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(file_name)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


########### reporte de OOEE publicadas ################################################

class reportOOEEPublicadas_xls(TemplateView):
    def get(self, request, *args, **kwargs):

        ooees = OperacionEstadistica.objects.filter(nombre_est=5)
        user = request.user         
        entidadesFasesTextList = []
        normaTextList = []
        requerimientoTextList = []
        usuariosPrinTextList = []
        unidadObserTextList = []
        obtencionDatoTextList = []
        muestProbaTextList = []
        muestNoProbaTextList = []
        tipoMarcoTextList = []
        docsDesarrolloTextList = [] 
        conceptosEstandaTextList = []
        clasificacionesTextList = []
        coberGeogTextList = []
        desagreInfoTextList = []
        fuenteFinanTextList =[]
        listaDeVariablesText = []  # modulo c Pregunta 14
        listaResultEstText = [] # modulo c pregunta 15
        medioDatosTextList = []
        periodicidadTextList = []
        herraProcesamiTextList = []
        analResultTextList = []
        medioDifusTextList = []
        fechaPublicTextList = []
        frecuenciaDifusionTextList = [] 
        productosDifundirTextList = []
        otrosProductosTextList = []
        resultadosSimiTextList = []
        hpSistemaInfoTextList = []
        listaEvaluacionText = [] #Evaluación de calidad
        listaCriticaText = [] #Critica
        listaNovedadText = [] #Novedades

        objArray = list(ooees.values('id'))
        for obj in objArray:
            for key, value in obj.items():
                #print("val", value)
                
                entidadesFasesList = MB_EntidadFases.objects.filter(ooee_id=value)
                unioListEntidadFase = list(chain(entidadesFasesList))
                entidadesFasesTextList.extend(unioListEntidadFase)

                normaList = MB_Norma.objects.filter(ooee_id=value)
                unioListNorma = list(chain(normaList))
                normaTextList.extend(unioListNorma)

                requerimientosList = MB_Requerimientos.objects.filter(ooee_id=value)
                unioListRequerimiento = list(chain(requerimientosList))
                requerimientoTextList.extend(unioListRequerimiento)

                usuariosPrinList = MB_PrinUsuarios.objects.filter(ooee_id=value)
                unioListUsuariosPrin = list(chain(usuariosPrinList))
                usuariosPrinTextList.extend(unioListUsuariosPrin)

                unidadObserList = MC_UnidadObservacion.objects.filter(ooee_id=value)
                unioListUnidadObser = list(chain(unidadObserList))
                unidadObserTextList.extend(unioListUnidadObser)

                obtencionDatoList = MC_ObtencionDato.objects.filter(ooee_id=value)
                unioListObtencionDato = list(chain(obtencionDatoList))
                obtencionDatoTextList.extend(unioListObtencionDato)

                muestProbaList = MC_MuestreoProbabilistico.objects.filter(ooee_id=value)
                unioListmuestProba =  list(chain(muestProbaList))
                muestProbaTextList.extend(unioListmuestProba)

                muestNoProbaList = MC_MuestreoNoProbabilistico.objects.filter(ooee_id=value)
                unioListmuestNoProba = list(chain(muestNoProbaList))
                muestNoProbaTextList.extend(unioListmuestNoProba)

                tipoMarcoList = MC_TipoMarco.objects.filter(ooee_id=value)
                unioListTipoMarco = list(chain(tipoMarcoList))
                tipoMarcoTextList.extend(unioListTipoMarco)

                docsDesarrolloList = MC_DocsDesarrollo.objects.filter(ooee_id=value)
                unioListdocsDesarrollo = list(chain(docsDesarrolloList))
                docsDesarrolloTextList.extend(unioListdocsDesarrollo)  

                conceptosEstandaList = MC_ConceptosEstandarizados.objects.filter(ooee_id=value)
                unionListconceptosEstanda = list(chain(conceptosEstandaList))
                conceptosEstandaTextList.extend(unionListconceptosEstanda)  

                clasificacionesList = MC_Clasificaciones.objects.filter(ooee_id=value)
                unioListclasificaciones = list(chain(clasificacionesList))
                clasificacionesTextList.extend(unioListclasificaciones)  

                coberGeogList = MC_CoberturaGeografica.objects.filter(ooee_id=value)
                unioListCoberGeog = list(chain(coberGeogList))
                coberGeogTextList.extend(unioListCoberGeog)

                desagreInfoList = MC_DesagregacionInformacion.objects.filter(ooee_id=value)
                unioListDesagreInfo = list(chain(desagreInfoList))
                desagreInfoTextList.extend(unioListDesagreInfo)

                fuenteFinanList = MC_FuenteFinanciacion.objects.filter(ooee_id=value)
                unioListFuenteFinan = list(chain(fuenteFinanList))
                fuenteFinanTextList.extend(unioListFuenteFinan)

                listVariables = MC_listaVariable.objects.filter(ooee_id=value)  #modulo c pregunta 14
                unioListVariables = list(chain(listVariables))
                listaDeVariablesText.extend(unioListVariables)

                listaResultEst = MC_ResultadoEstadistico.objects.filter(ooee_id=value)  #modulo c pregunta 15
                unioListResultEst = list(chain(listaResultEst))
                listaResultEstText.extend(unioListResultEst)

                medioDatosList = MD_MedioDatos.objects.filter(ooee_id=value)
                unioListMedioDatos = list(chain(medioDatosList))
                medioDatosTextList.extend(unioListMedioDatos)

                periodicidadList = MD_PeriodicidadOe.objects.filter(ooee_id=value)
                unioListPeriodicidad = list(chain(periodicidadList))
                periodicidadTextList.extend(unioListPeriodicidad)

                herraProcesamiList = MD_HerramProcesamiento.objects.filter(ooee_id=value)
                unioListHerraProcesami = list(chain(herraProcesamiList))
                herraProcesamiTextList.extend(unioListHerraProcesami)

                analResultList = ME_AnalisisResultados.objects.filter(ooee_id=value)
                unioListAnalResult = list(chain(analResultList))
                analResultTextList.extend(unioListAnalResult)

                medioDifusList = MF_MediosDifusion.objects.filter(ooee_id=value)
                unioListMedioDifus = list(chain(medioDifusList))
                medioDifusTextList.extend(unioListMedioDifus)

                fechaPublicList = MF_FechaPublicacion.objects.filter(ooee_id=value)
                unioListFechaPublic = list(chain(fechaPublicList))
                fechaPublicTextList.extend(unioListFechaPublic)

                frecuenciaDifusionList = MF_FrecuenciaDifusion.objects.filter(ooee_id=value)
                unioListFrecuenciaDifusion = list(chain(frecuenciaDifusionList))
                frecuenciaDifusionTextList.extend(unioListFrecuenciaDifusion)

                productosDifundirList = MF_ProductosDifundir.objects.filter(ooee_id=value)
                unioListProductosDifundir  = list(chain(productosDifundirList))
                productosDifundirTextList.extend(unioListProductosDifundir)

                otrosProductosList = MF_OtrosProductos.objects.filter(ooee_id=value)
                unioListOtrosProductos = list(chain(otrosProductosList))
                otrosProductosTextList.extend(unioListOtrosProductos)

                resultadosSimiList = MF_ResultadosSimilares.objects.filter(ooee_id=value)
                unioListResultadosSimi = list(chain(resultadosSimiList))
                resultadosSimiTextList.extend(unioListResultadosSimi)

                hpSistemaInfoList = MF_HPSistemaInfo.objects.filter(ooee_id=value)
                unioListhpSistemaInfo = list(chain(hpSistemaInfoList))
                hpSistemaInfoTextList.extend(unioListhpSistemaInfo)

                listEvaluacion = EvaluacionCalidad.objects.filter(post_oe_id=value) # evalaucion de calidad
                unioListEvaluacion = list(chain(listEvaluacion))
                listaEvaluacionText.extend(unioListEvaluacion)

                listCritica = Critica.objects.filter(post_oe_id=value) #critica de la oe
                unioListCritica = list(chain(listCritica))
                listaCriticaText.extend(unioListCritica)

                listNovedad = NovedadActualizacion.objects.filter(post_oe_id=value) #critica de la oe
                unioListNovedad = list(chain(listNovedad))
                listaNovedadText.extend(unioListNovedad)

        wb = Workbook()
        ws = wb.active
        ws.title = "Directorio OOEE"
        
        sheet2 = wb.create_sheet('Entidades Fases')
        sheet3 = wb.create_sheet('Variables')
        sheet4 = wb.create_sheet('Resultados')
        sheet5 = wb.create_sheet('Eval Calidad')
        sheet6 = wb.create_sheet('Critica OE')
        sheet7 = wb.create_sheet('Novedad OE')

        def set_border(ws, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = ws[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(ws,'A1:IC'+str(ooees.count()+3))

		## size rows

        ws.row_dimensions[1].height = 55
        ws.row_dimensions[2].height = 40
        ws.row_dimensions[3].height = 80
       
		## size column
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 15
        ws.column_dimensions['O'].width = 15
        ws.column_dimensions['P'].width = 15
        ws.column_dimensions['Q'].width = 15
        ws.column_dimensions['R'].width = 22
        ws.column_dimensions['S'].width = 15
        ws.column_dimensions['T'].width = 15
        ws.column_dimensions['U'].width = 15
        ws.column_dimensions['V'].width = 15
        ws.column_dimensions['W'].width = 15
        ws.column_dimensions['X'].width = 15
        ws.column_dimensions['Y'].width = 20
        ws.column_dimensions['Z'].width = 15

        ws.column_dimensions['AA'].width = 15
        ws.column_dimensions['AB'].width = 15
        ws.column_dimensions['AC'].width = 15
        ws.column_dimensions['AD'].width = 15
        ws.column_dimensions['AE'].width = 15
        ws.column_dimensions['AF'].width = 15
        ws.column_dimensions['AG'].width = 15
        ws.column_dimensions['AH'].width = 15
        ws.column_dimensions['AI'].width = 15
        ws.column_dimensions['AJ'].width = 15
        ws.column_dimensions['AK'].width = 15
        ws.column_dimensions['AL'].width = 15
        ws.column_dimensions['AM'].width = 15
        ws.column_dimensions['AN'].width = 15
        ws.column_dimensions['AO'].width = 15
        ws.column_dimensions['AP'].width = 15
        ws.column_dimensions['AQ'].width = 15
        ws.column_dimensions['AR'].width = 15
        ws.column_dimensions['AS'].width = 15
        ws.column_dimensions['AT'].width = 15
        ws.column_dimensions['AU'].width = 15
        ws.column_dimensions['AV'].width = 15
        ws.column_dimensions['AW'].width = 15
        ws.column_dimensions['AX'].width = 15
        ws.column_dimensions['AY'].width = 15
        ws.column_dimensions['AZ'].width = 15

        ws.column_dimensions['BA'].width = 15
        ws.column_dimensions['BB'].width = 15
        ws.column_dimensions['BC'].width = 15
        ws.column_dimensions['BD'].width = 15
        ws.column_dimensions['BE'].width = 15
        ws.column_dimensions['BF'].width = 15
        ws.column_dimensions['BG'].width = 15
        ws.column_dimensions['BH'].width = 15
        ws.column_dimensions['BI'].width = 15
        ws.column_dimensions['BJ'].width = 15
        ws.column_dimensions['BK'].width = 15
        ws.column_dimensions['BL'].width = 15
        ws.column_dimensions['BM'].width = 15
        ws.column_dimensions['BN'].width = 15
        ws.column_dimensions['BO'].width = 15
        ws.column_dimensions['BP'].width = 15
        ws.column_dimensions['BQ'].width = 15
        ws.column_dimensions['BR'].width = 15
        ws.column_dimensions['BS'].width = 15
        ws.column_dimensions['BT'].width = 15
        ws.column_dimensions['BU'].width = 15
        ws.column_dimensions['BV'].width = 15
        ws.column_dimensions['BW'].width = 15
        ws.column_dimensions['BX'].width = 15
        ws.column_dimensions['BY'].width = 15
        ws.column_dimensions['BZ'].width = 15

        ws.column_dimensions['CA'].width = 15
        ws.column_dimensions['CB'].width = 15
        ws.column_dimensions['CC'].width = 15
        ws.column_dimensions['CD'].width = 15
        ws.column_dimensions['CE'].width = 15
        ws.column_dimensions['CF'].width = 15
        ws.column_dimensions['CG'].width = 15
        ws.column_dimensions['CH'].width = 15
        ws.column_dimensions['CI'].width = 15
        ws.column_dimensions['CJ'].width = 15
        ws.column_dimensions['CK'].width = 15
        ws.column_dimensions['CL'].width = 15
        ws.column_dimensions['CM'].width = 15
        ws.column_dimensions['CN'].width = 15
        ws.column_dimensions['CO'].width = 15
        ws.column_dimensions['CP'].width = 15
        ws.column_dimensions['CQ'].width = 15
        ws.column_dimensions['CR'].width = 15
        ws.column_dimensions['CS'].width = 15
        ws.column_dimensions['CT'].width = 15
        ws.column_dimensions['CU'].width = 15
        ws.column_dimensions['CV'].width = 15
        ws.column_dimensions['CW'].width = 15
        ws.column_dimensions['CX'].width = 15
        ws.column_dimensions['CY'].width = 15
        ws.column_dimensions['CZ'].width = 15

        ws.column_dimensions['DA'].width = 15
        ws.column_dimensions['DB'].width = 15
        ws.column_dimensions['DC'].width = 15
        ws.column_dimensions['DD'].width = 15
        ws.column_dimensions['DE'].width = 15
        ws.column_dimensions['DF'].width = 15
        ws.column_dimensions['DG'].width = 15
        ws.column_dimensions['DH'].width = 15
        ws.column_dimensions['DI'].width = 15
        ws.column_dimensions['DJ'].width = 15
        ws.column_dimensions['DK'].width = 15
        ws.column_dimensions['DL'].width = 15
        ws.column_dimensions['DM'].width = 15
        ws.column_dimensions['DN'].width = 15
        ws.column_dimensions['DO'].width = 15
        ws.column_dimensions['DP'].width = 15
        ws.column_dimensions['DQ'].width = 15
        ws.column_dimensions['DR'].width = 15
        ws.column_dimensions['DS'].width = 15
        ws.column_dimensions['DT'].width = 15
        ws.column_dimensions['DU'].width = 15
        ws.column_dimensions['DV'].width = 15
        ws.column_dimensions['DW'].width = 15
        ws.column_dimensions['DX'].width = 15
        ws.column_dimensions['DY'].width = 15
        ws.column_dimensions['DZ'].width = 15

        ws.column_dimensions['EA'].width = 15
        ws.column_dimensions['EB'].width = 15
        ws.column_dimensions['EC'].width = 15
        ws.column_dimensions['ED'].width = 15
        ws.column_dimensions['EE'].width = 15
        ws.column_dimensions['EF'].width = 15
        ws.column_dimensions['EG'].width = 15
        ws.column_dimensions['EH'].width = 15
        ws.column_dimensions['EI'].width = 15
        ws.column_dimensions['EJ'].width = 15
        ws.column_dimensions['EK'].width = 15
        ws.column_dimensions['EL'].width = 15
        ws.column_dimensions['EM'].width = 15
        ws.column_dimensions['EN'].width = 15
        ws.column_dimensions['EO'].width = 15
        ws.column_dimensions['EP'].width = 15
        ws.column_dimensions['EQ'].width = 15
        ws.column_dimensions['ER'].width = 15
        ws.column_dimensions['ES'].width = 15
        ws.column_dimensions['ET'].width = 15
        ws.column_dimensions['EU'].width = 15
        ws.column_dimensions['EV'].width = 15
        ws.column_dimensions['EW'].width = 15
        ws.column_dimensions['EX'].width = 15
        ws.column_dimensions['EY'].width = 15
        ws.column_dimensions['EZ'].width = 15

        ws.column_dimensions['FA'].width = 20
        ws.column_dimensions['FB'].width = 20
        ws.column_dimensions['FC'].width = 15
        ws.column_dimensions['FD'].width = 15
        ws.column_dimensions['FE'].width = 15
        ws.column_dimensions['FF'].width = 15
        ws.column_dimensions['FG'].width = 15
        ws.column_dimensions['FH'].width = 15
        ws.column_dimensions['FI'].width = 15
        ws.column_dimensions['FJ'].width = 15
        ws.column_dimensions['FK'].width = 15
        ws.column_dimensions['FL'].width = 15
        ws.column_dimensions['FM'].width = 15
        ws.column_dimensions['FN'].width = 15
        ws.column_dimensions['FO'].width = 15
        ws.column_dimensions['FP'].width = 15
        ws.column_dimensions['FQ'].width = 15
        ws.column_dimensions['FR'].width = 15
        ws.column_dimensions['FS'].width = 15
        ws.column_dimensions['FT'].width = 15
        ws.column_dimensions['FU'].width = 15
        ws.column_dimensions['FV'].width = 15
        ws.column_dimensions['FW'].width = 15
        ws.column_dimensions['FX'].width = 15
        ws.column_dimensions['FY'].width = 15
        ws.column_dimensions['FZ'].width = 15

        ws.column_dimensions['GA'].width = 20
        ws.column_dimensions['GB'].width = 20
        ws.column_dimensions['GC'].width = 40
        ws.column_dimensions['GD'].width = 15
        ws.column_dimensions['GE'].width = 15
        ws.column_dimensions['GF'].width = 15
        ws.column_dimensions['GG'].width = 15
        ws.column_dimensions['GH'].width = 15
        ws.column_dimensions['GI'].width = 15
        ws.column_dimensions['GJ'].width = 15
        ws.column_dimensions['GK'].width = 15
        ws.column_dimensions['GL'].width = 15
        ws.column_dimensions['GM'].width = 15
        ws.column_dimensions['GN'].width = 15
        ws.column_dimensions['GO'].width = 15
        ws.column_dimensions['GP'].width = 15
        ws.column_dimensions['GQ'].width = 15
        ws.column_dimensions['GR'].width = 15
        ws.column_dimensions['GS'].width = 15
        ws.column_dimensions['GT'].width = 15
        ws.column_dimensions['GU'].width = 15
        ws.column_dimensions['GV'].width = 15
        ws.column_dimensions['GW'].width = 15
        ws.column_dimensions['GX'].width = 15
        ws.column_dimensions['GY'].width = 15
        ws.column_dimensions['GZ'].width = 15

        ws.column_dimensions['HA'].width = 20
        ws.column_dimensions['HB'].width = 20
        ws.column_dimensions['HC'].width = 20
        ws.column_dimensions['HD'].width = 15
        ws.column_dimensions['HE'].width = 15
        ws.column_dimensions['HF'].width = 15
        ws.column_dimensions['HG'].width = 15
        ws.column_dimensions['HH'].width = 15
        ws.column_dimensions['HI'].width = 15
        ws.column_dimensions['HJ'].width = 15
        ws.column_dimensions['HK'].width = 15
        ws.column_dimensions['HL'].width = 15
        ws.column_dimensions['HM'].width = 15
        ws.column_dimensions['HN'].width = 15
        ws.column_dimensions['HO'].width = 15
        ws.column_dimensions['HP'].width = 15
        ws.column_dimensions['HQ'].width = 15
        ws.column_dimensions['HR'].width = 15
        ws.column_dimensions['HS'].width = 15
        ws.column_dimensions['HT'].width = 20
        ws.column_dimensions['HU'].width = 15
        ws.column_dimensions['HV'].width = 20
        ws.column_dimensions['HW'].width = 15
        ws.column_dimensions['HX'].width = 15
        ws.column_dimensions['HY'].width = 40
        ws.column_dimensions['HZ'].width = 15
        ws.column_dimensions['IA'].width = 15
        ws.column_dimensions['IB'].width = 15
        ws.column_dimensions['IC'].width = 25
        

		##insert image
        #img = openpyxl.drawing.image.Image('media/pictures/logoSEN.png')
        #img.anchor = 'A1'	
        #ws.add_image(img)


		##styles
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
		
        title_cell = ws['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de las Operaciones Estadísticas'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')
    
        ws.merge_cells('A1:IC1')

        ws.merge_cells('A2:C2')
        ws.merge_cells('D2:E2')
        ws.merge_cells('F2:H2')
        ws.merge_cells('I2:Q2')
        ws.merge_cells('R2:R3')
        ws.merge_cells('S2:W2')
        ws.merge_cells('X2:AE2')
        ws.merge_cells('AF2:AP2')
        ws.merge_cells('AQ2:AQ3')
        ws.merge_cells('AR2:AY2')

        ws.merge_cells('AZ2:BD2')
        ws.merge_cells('BE2:BL2') 
        ws.merge_cells('BM2:BR2')
        ws.merge_cells('BS2:BX2') 
        ws.merge_cells('BY2:CD2')  
        ws.merge_cells('CE2:CL2')  
        ws.merge_cells('CM2:CX2')
        ws.merge_cells('CY2:DB2')
        ws.merge_cells('DC2:DR2')
        ws.merge_cells('DS2:EH2')
        ws.merge_cells('EI2:EK2')
        ws.merge_cells('EL2:ES2')
        ws.merge_cells('ET2:EU2')
        ws.merge_cells('EV2:EZ2')
        ws.merge_cells('FA2:FA3')
        ws.merge_cells('FB2:FB3')
        ws.merge_cells('FC2:FL2')
        ws.merge_cells('FM2:FS2')
        ws.merge_cells('FT2:GB2')
        ws.merge_cells('GC2:GC3')
        ws.merge_cells('GD2:GH2')
        ws.merge_cells('GI2:GM2')
        ws.merge_cells('GN2:GN3')
        ws.merge_cells('GO2:GP2')
        ws.merge_cells('GQ2:GU2')
        ws.merge_cells('GV2:HB2')
        ws.merge_cells('HC2:HI2')
        ws.merge_cells('HJ2:HS2')
        ws.merge_cells('HT2:HV2')
        ws.merge_cells('HW2:HX2')
        ws.merge_cells('HY2:HY3')
        ws.merge_cells('HZ2:HZ3')

        ws.merge_cells('IA2:IA3')
        ws.merge_cells('IB2:IB3')
        ws.merge_cells('IC2:IC3')



        ## insert heads groups
        codigo_cell = ws['A2']
        codigo_cell.value = 'OPERACIÓN ESTADÍSTICA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['D2']
        codigo_cell.value = 'ENTIDAD'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['F2']
        codigo_cell.value = 'ÁREA TEMÁTICA / TEMA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['I2']
        codigo_cell.value = ' A. IDENTIFICACIÓN '
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['R2']
        codigo_cell.value = 'Indique si hay otra(s) entidad(es) que participa(n) en alguna de las fases del proceso estadístico  (ver hoja 2)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['S2']
        codigo_cell.value = 'Bajo cuál(es) de las siguiente(s) norma(s), se soporta la producción de información de la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['X2']
        codigo_cell.value = 'La operación estadística satisface requerimientos de información de:'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = ws['AF2']
        codigo_cell.value = 'Señale cuáles son los principales usuarios  de la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AQ2']
        codigo_cell.value = '¿Cuál es la población objetivo de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AR2']
        codigo_cell.value = '¿Cuál es la unidad de observación de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AZ2']
        codigo_cell.value = '¿Cuál es el tipo de operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BE2']
        codigo_cell.value = 'Indique de donde se obtienen los datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BM2']
        codigo_cell.value = 'Muestreo probabilístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BS2']
        codigo_cell.value = 'Muestreo No probabilístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BY2']
        codigo_cell.value = ' ¿La operación estadística cuenta con un marco estadístico para identificar y ubicar las unidades de observación?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CE2']
        codigo_cell.value = 'Indique cuáles de los siguientes documentos se elaboran para el desarrollo de la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CM2']
        codigo_cell.value = 'Indique si la operación estadística utiliza conceptos estandarizados de'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CY2']
        codigo_cell.value = '¿La operación estadística utiliza nomenclaturas y/o clasificaciones?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DC2']
        codigo_cell.value = '¿Cuál es la cobertura geográfica de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DS2']
        codigo_cell.value = 'Desagregación geográfica '
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EI2']
        codigo_cell.value = 'Desagregación por zona'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EL2']
        codigo_cell.value = 'Desagregación por grupos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['ET2']
        codigo_cell.value = '¿Cuál es el costo anual de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['EV2']
        codigo_cell.value = '¿Cuál(es) son las fuentes de financiación de la operación estadística ?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
    
        codigo_cell = ws['FA2']
        codigo_cell.value = 'Liste todas las variables que maneja la operación estadística (Ver hoja 3)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FB2']
        codigo_cell.value = '¿Cuáles son los resultados agregados o indicadores calculados? (Ver hoja 4)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = ws['FC2']
        codigo_cell.value = '¿Cuál es el medio de obtención de los datos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['FM2']
        codigo_cell.value = '¿Cuál es la periodicidad de recolección o acopio de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['FT2']
        codigo_cell.value = 'Indique cuáles de las siguientes herramientas son utilizadas en el procesamiento de los datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = ws['GC2']
        codigo_cell.value = 'Haga una  breve descripción de la manera cómo se realiza el procesamiento de los datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GD2']
        codigo_cell.value = '¿Qué tipo de análisis realiza a los resultados obtenidos en la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GI2']
        codigo_cell.value = '¿A través de qué medio(s) difunde los resultados estadísticos (Agregados, indicadores), a los usuarios?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['GN2']
        codigo_cell.value = 'Indique la dirección de la página web donde se encuentran los resultados estadísticos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GO2']
        codigo_cell.value = 'Fechas de disponibilidad de los resultados estadísticos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['GQ2']
        codigo_cell.value = '¿Cuál es la próxima fecha de publicación de resultados estadísticos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['GV2']
        codigo_cell.value = '¿Cuál es la frecuencia de difusión de los resultados estadísticos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HC2']
        codigo_cell.value = '¿Cuáles productos utiliza para difundir los resultados estadísticos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HJ2']
        codigo_cell.value = '¿Qué otros productos estadísticos de la OE están disponibles para consulta de los usuarios?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HT2']
        codigo_cell.value = '¿Conoce si otra entidad produce resultados similares a los de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
      
        codigo_cell = ws['HW2']
        codigo_cell.value = '¿La operación estadística hace parte de algún sistema de información?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['HY2']
        codigo_cell.value = 'Observaciones'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['HZ2']
        codigo_cell.value = 'Anexos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['IA2']
        codigo_cell.value = 'Estado de la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['IB2']
        codigo_cell.value = 'Validación de la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['IC2']
        codigo_cell.value = 'Estado del proceso de la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


    ####### insert heads fields

        codigo_cell = ws['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['B3']
        codigo_cell.value = 'Nombre Operación Estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['C3']
        codigo_cell.value = 'Objetivo de la Operación Estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['D3']
        codigo_cell.value = 'Nombre de la Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center',  wrap_text=True)
       
        codigo_cell = ws['E3']
        codigo_cell.value = 'Código de la Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['F3']
        codigo_cell.value = 'Área Tematica'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = ws['G3']
        codigo_cell.value = 'Tema'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['H3']
        codigo_cell.value = 'Tema Compartido'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
      
        codigo_cell = ws['I3']
        codigo_cell.value = 'Nombre de la Dependencia'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['J3']
        codigo_cell.value = 'Nombre del Director'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['K3']
        codigo_cell.value = 'Cargo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['L3']
        codigo_cell.value = 'Correo Electrónico del Director'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = ws['M3']
        codigo_cell.value = 'Teléfono del director'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['N3']
        codigo_cell.value = 'Nombre del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['O3']
        codigo_cell.value = 'Cargo del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['P3']
        codigo_cell.value = 'Correo Electrónico del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['Q3']
        codigo_cell.value = 'Teléfono del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
         
        codigo_cell = ws['S3']
        codigo_cell.value = 'a. Constitución Política'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['T3']
        codigo_cell.value = 'b. Ley'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['U3']
        codigo_cell.value = 'c. Decreto (nacional, departamental, municipal)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['V3']
        codigo_cell.value = 'd. Otra (Resolución, ordenanza, acuerdo)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center',  wrap_text=True)

        codigo_cell = ws['W3']
        codigo_cell.value = 'e. Ninguna'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['X3']
        codigo_cell.value = 'a. Objetivos de desarrollo Sostenible (ODS)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['Y3']
        codigo_cell.value = 'b. Organización para la Cooperación y el Desarrollo Económico (OCDE)   '
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['Z3']
        codigo_cell.value = 'c. Otros compromisos internacionales'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['AA3']
        codigo_cell.value = 'd. Plan Nacional de Desarrollo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AB3']
        codigo_cell.value = 'e. Cuentas económicas y macroeconómicas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['AC3']
        codigo_cell.value = 'f. Plan Sectorial, Territorial o CONPES'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['AD3']
        codigo_cell.value = 'g. Otro (s)  ¿cuál(es) ?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AE3']
        codigo_cell.value = 'h. Ninguno'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        codigo_cell = ws['AF3']
        codigo_cell.value = 'a. Organismos internacionales'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AG3']
        codigo_cell.value = 'b. Presidencia de la República'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AH3']
        codigo_cell.value = 'c. Ministerios'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AI3']
        codigo_cell.value = 'd. Organismos de Control'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AJ3']
        codigo_cell.value = 'e. Otras entidades del orden Nacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AK3']
        codigo_cell.value = 'f. Entidades de orden Territorial'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AL3']
        codigo_cell.value = 'g. Gremios'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AM3']
        codigo_cell.value = 'h. Entidades privadas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AN3']
        codigo_cell.value = 'i. Dependencias de la misma entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AO3']
        codigo_cell.value = 'j. Academia'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AP3']
        codigo_cell.value = 'k. Público en General'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AR3']
        codigo_cell.value = 'a. Empresa'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        codigo_cell = ws['AS3']
        codigo_cell.value = 'b. Establecimiento'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AT3']
        codigo_cell.value = 'c. Hogar'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AU3']
        codigo_cell.value = 'd. Persona'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AV3']
        codigo_cell.value = 'e. Unidad productora agropecuaria'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AW3']
        codigo_cell.value = 'f. Predio'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AX3']
        codigo_cell.value = 'g. Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AY3']
        codigo_cell.value = 'Cuál(es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['AZ3']
        codigo_cell.value = 'a. Aprovechamiento de registro administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BA3']
        codigo_cell.value = 'b. Estadística derivada'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BB3']
        codigo_cell.value = 'c. Muestreo probabilístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BC3']
        codigo_cell.value = 'd. Muestreo no probabilístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BD3']
        codigo_cell.value = 'e. Censo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BE3']
        codigo_cell.value = 'Registros Administrativos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BF3']
        codigo_cell.value = 'Listado de RRAA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BG3']
        codigo_cell.value = 'Cuál(es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BH3']
        codigo_cell.value = 'Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BI3']
        codigo_cell.value = 'Operación Estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BJ3']
        codigo_cell.value = 'Listado de OOEE'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BK3']
        codigo_cell.value = 'Cuál(es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BL3']
        codigo_cell.value =  'Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BM3']
        codigo_cell.value =  'Muestreo aleatorio simple'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BN3']
        codigo_cell.value =  'Muestreo aleatorio sistemático'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BO3']
        codigo_cell.value =  'Muestreo aleatorio estratificado'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BP3']
        codigo_cell.value =  'Muestreo aleatorio por conglomerados'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BQ3']
        codigo_cell.value =  'Otro'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BR3']
        codigo_cell.value =  '¿Cual?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BS3']
        codigo_cell.value =  'Muestreo por cuotas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BT3']
        codigo_cell.value =  'Muestreo intencional o de conveniencia'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BU3']
        codigo_cell.value =  'Bola de nieve'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BV3']
        codigo_cell.value =  'Muestreo discrecional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BW3']
        codigo_cell.value =  'Otro'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BX3']
        codigo_cell.value =  '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BY3']
        codigo_cell.value =  'Se cuenta con un marco'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['BZ3']
        codigo_cell.value =  'Marco de lista'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CA3']
        codigo_cell.value =  'Marco de área'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CB3']
        codigo_cell.value =  'Marco geoestadístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CC3']
        codigo_cell.value =  'Otro(s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CD3']
        codigo_cell.value =  '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CE3']
        codigo_cell.value =  'Metodología'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CF3']
        codigo_cell.value =  'Ficha Metodológica (técnica)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CG3']
        codigo_cell.value =  'Hojas de vida de indicadores'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CH3']
        codigo_cell.value =  'Manual operativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CI3']
        codigo_cell.value =  'Diccionario de datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CJ3']
        codigo_cell.value =  'Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CK3']
        codigo_cell.value = '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CL3']
        codigo_cell.value = 'Ninguno'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CM3']
        codigo_cell.value = 'DANE'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CN3']
        codigo_cell.value = 'Organismos internacionales'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CO3']
        codigo_cell.value = '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CP3']
        codigo_cell.value = 'Otra entidad de orden nacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CQ3']
        codigo_cell.value = '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CR3']
        codigo_cell.value = 'Leyes, decretos, etc'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CS3']
        codigo_cell.value = '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CT3']
        codigo_cell.value = 'Creación propia'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CU3']
        codigo_cell.value = 'Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CV3']
        codigo_cell.value = '¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CW3']
        codigo_cell.value = 'Ninguno'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CX3']
        codigo_cell.value = '¿Por qué?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CY3']
        codigo_cell.value = '¿se utilizan nomenclaturas y/o clasificaciones estandarizadas?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['CZ3']
        codigo_cell.value = 'Si: ¿Cuáles?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DA3']
        codigo_cell.value = 'Si: Otras'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DB3']
        codigo_cell.value = 'No: ¿Por qué?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DC3']
        codigo_cell.value = 'a. Nacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DD3']
        codigo_cell.value = 'b. Regional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DE3']
        codigo_cell.value = 'b. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DF3']
        codigo_cell.value = 'b. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DG3']
        codigo_cell.value = 'c. Departamental'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DH3']
        codigo_cell.value = 'c. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DI3']
        codigo_cell.value = 'c. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DJ3']
        codigo_cell.value = 'd. Áreas metropolitanas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DK3']
        codigo_cell.value = 'd. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DL3']
        codigo_cell.value = 'd. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DM3']
        codigo_cell.value = 'e. Municipal'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DN3']
        codigo_cell.value = 'e. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DO3']
        codigo_cell.value = 'e. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DP3']
        codigo_cell.value = 'f. Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DQ3']
        codigo_cell.value = 'f. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DR3']
        codigo_cell.value = 'f. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DS3']
        codigo_cell.value = 'a. Nacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DT3']
        codigo_cell.value = 'b. Regional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DU3']
        codigo_cell.value = 'b. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        
        codigo_cell = ws['DV3']
        codigo_cell.value = 'b. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DW3']
        codigo_cell.value = 'c. Departamental'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DX3']
        codigo_cell.value = 'c. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DY3']
        codigo_cell.value = 'c. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['DZ3']
        codigo_cell.value = 'd. Áreas metropolitanas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EA3']
        codigo_cell.value = 'd. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EB3']
        codigo_cell.value = 'd. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EC3']
        codigo_cell.value = 'e. Municipal'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['ED3']
        codigo_cell.value = 'e. ¿Cuantos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EE3']
        codigo_cell.value = 'e. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EF3']
        codigo_cell.value = 'f. Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EG3']
        codigo_cell.value = 'f. ¿Cuántos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EH3']
        codigo_cell.value = 'f. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EI3']
        codigo_cell.value = 'Total'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EJ3']
        codigo_cell.value = 'Urbano'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EK3']
        codigo_cell.value = 'Rural'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EL3']
        codigo_cell.value = 'Sexo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EM3']
        codigo_cell.value = 'Edad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EN3']
        codigo_cell.value = 'Grupo étnico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EO3']
        codigo_cell.value = 'Discapacidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EP3']
        codigo_cell.value = 'Estrato'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EQ3']
        codigo_cell.value = 'Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['ER3']
        codigo_cell.value = '¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['ES3']
        codigo_cell.value = 'Ninguno'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['ET3']
        codigo_cell.value = 'a. Costo anuaL'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EU3']
        codigo_cell.value = 'b. No sabe'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EV3']
        codigo_cell.value = 'Recursos propios'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EW3']
        codigo_cell.value = 'Aportes de otra entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EX3']
        codigo_cell.value = 'Cooperación internacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EY3']
        codigo_cell.value = 'Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['EZ3']
        codigo_cell.value = '¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FC3']
        codigo_cell.value = 'a. Formulario físico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FD3']
        codigo_cell.value = 'b. Formulario electrónico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FE3']
        codigo_cell.value = 'c. Dispositivo Móvil de Captura [DMC]'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FF3']
        codigo_cell.value = 'd. Sistema de Información'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FG3']
        codigo_cell.value = 'd. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FH3']
        codigo_cell.value = 'e. Percepción remota (imágenes satelitales, fotos, sensores, etc)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FI3']
        codigo_cell.value = 'f. Base de datos de registro administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FJ3']
        codigo_cell.value = 'g. Resultados estadísticos de otra operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FK3']
        codigo_cell.value = 'h. Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FL3']
        codigo_cell.value = 'h. ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FM3']
        codigo_cell.value = 'a. Anual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FN3']
        codigo_cell.value = 'b. Semestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FO3']
        codigo_cell.value = 'c. Trimestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FP3']
        codigo_cell.value = 'd. Mensual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FQ3']
        codigo_cell.value = 'e. Diaria'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FR3']
        codigo_cell.value = 'f. Otra'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FS3']
        codigo_cell.value = 'f. ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FT3']
        codigo_cell.value = 'a. Excel'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FU3']
        codigo_cell.value = 'b. Acces'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FV3']
        codigo_cell.value = 'c. R'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FW3']
        codigo_cell.value = 'd. SAS'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FX3']
        codigo_cell.value = 'e. SPSS'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FY3']
        codigo_cell.value = 'f. Oracle'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['FZ3']
        codigo_cell.value = 'g. Stata'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GA3']
        codigo_cell.value = 'h. Otra'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GB3']
        codigo_cell.value = 'h. ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GD3']
        codigo_cell.value = 'a Consistencia y validación de datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GE3']
        codigo_cell.value = 'b Análisis de tendencias y series de tiempo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GF3']
        codigo_cell.value = 'c Análisis de contexto de los resultados'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GG3']
        codigo_cell.value = 'd Otro.'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GH3']
        codigo_cell.value = 'd. ¿cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GI3']
        codigo_cell.value = 'a. Página web'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GJ3']
        codigo_cell.value = 'b. Medio físico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GK3']
        codigo_cell.value = 'c. Medio electrónico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GL3']
        codigo_cell.value = 'd. Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GM3']
        codigo_cell.value = 'd. ¿cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GO3']
        codigo_cell.value = 'Desde mes y año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GP3']
        codigo_cell.value = 'Hasta mes y año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GQ3']
        codigo_cell.value = 'a. Fecha publicación'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GR3']
        codigo_cell.value = 'a. mes/año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GS3']
        codigo_cell.value = 'b. No sabe'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GT3']
        codigo_cell.value = 'c. No hay'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GU3']
        codigo_cell.value = 'c. ¿por qué?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GV3']
        codigo_cell.value = 'a. Anual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GW3']
        codigo_cell.value = 'b. Semestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GX3']
        codigo_cell.value = 'c. Trimestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GY3']
        codigo_cell.value = 'd. Mensual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['GZ3']
        codigo_cell.value = 'e. Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HA3']
        codigo_cell.value = 'e. ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HB3']
        codigo_cell.value = 'f. No está definido'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HC3']
        codigo_cell.value = 'a. Cuadros de salida (tablas)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HD3']
        codigo_cell.value = 'b. Boletín estadístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HE3']
        codigo_cell.value = 'c. Anuario'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HF3']
        codigo_cell.value = 'd. Mapas estadísticos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HG3']
        codigo_cell.value = 'e. Bases de datos interactivas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HH3']
        codigo_cell.value = 'f. Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HI3']
        codigo_cell.value = 'f. ¿cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HJ3']
        codigo_cell.value = 'a. Series históricas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HK3']
        codigo_cell.value = 'a. Desde: mes y año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HL3']
        codigo_cell.value = 'a. Hasta: mes y año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HM3']
        codigo_cell.value = 'b. Microdatos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HN3']
        codigo_cell.value = 'b. Desde: mes y año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HO3']
        codigo_cell.value = 'b. Hasta: mes y año'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HP3']
        codigo_cell.value = 'c. Documentos metodológicos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HQ3']
        codigo_cell.value = 'c. URL'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HR3']
        codigo_cell.value = 'd. Calendario de difusión'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HS3']
        codigo_cell.value = 'e. Ninguna'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HT3']
        codigo_cell.value = '¿Conoce si otra entidad produce resultados similares a los de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HU3']
        codigo_cell.value = 'Si: Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HV3']
        codigo_cell.value = 'Si: Operación estadística/Indicadores'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HW3']
        codigo_cell.value = '¿La entidad hace parte de un sistema de información estadística de uso  interinstitucional ?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        codigo_cell = ws['HX3']
        codigo_cell.value = 'Si: ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        cont = 4
           
        for ooee in ooees:
            
            ws.cell(row = cont, column = 1).value = ooee.codigo_oe
            ws.cell(row = cont, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 1).font = Font(size = "8", name='Barlow')
            ws.cell(row = cont, column = 1).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            ws.cell(row = cont, column = 2).value = ooee.nombre_oe
            ws.cell(row = cont, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 2).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 3).value = ooee.objetivo_oe
            ws.cell(row = cont, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 3).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 4).value = str(ooee.entidad)
            ws.cell(row = cont, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 4).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 5).value =  ooee.entidad.codigo
            ws.cell(row = cont, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 5).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 6).value = str(ooee.area_tematica)
            ws.cell(row = cont, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 6).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 7).value = str(ooee.tema)
            ws.cell(row = cont, column = 7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 7).font = Font(size = "8", name='Barlow')
            
            ## TEMA COMPARTIDO

            listaTemaCompartido = ooee.tema_compartido.all()
            temaCompartido_array = []
            for indexTemaCompa, itemTemaCompa in enumerate(listaTemaCompartido):
                temaCompartido_array.append(str(itemTemaCompa))
                ws.cell(row = cont, column = 8).value = str(temaCompartido_array)
                ws.cell(row = cont, column = 8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 8).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 9).value = ooee.nombre_dep
            ws.cell(row = cont, column = 9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = cont, column = 9).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 10).value = ooee.nombre_dir
            ws.cell(row = cont, column = 10).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 10).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 11).value = ooee.cargo_dir
            ws.cell(row = cont, column = 11).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 11).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 12).value = ooee.correo_dir
            ws.cell(row = cont, column = 12).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 12).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 13).value = ooee.tel_dir
            ws.cell(row = cont, column = 13).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 13).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 14).value = ooee.nombre_resp
            ws.cell(row = cont, column = 14).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 14).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 15).value = ooee.cargo_resp
            ws.cell(row = cont, column = 15).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 15).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 16).value = ooee.correo_resp
            ws.cell(row = cont, column = 16).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 16).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 17).value = ooee.tel_resp
            ws.cell(row = cont, column = 17).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 17).font = Font(size = "8", name='Barlow')
            
            if str(ooee.fase) == "True":
                ws.cell(row = cont, column = 18).value = "Si"
                ws.cell(row = cont, column = 18).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 18).font = Font(size = "8", name='Barlow')
            elif str(ooee.fase) == "False":
                ws.cell(row = cont, column = 18).value = "No"
                ws.cell(row = cont, column = 18).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 18).font = Font(size = "8", name='Barlow')
            
   
            #*****lista de OOEE Y RRAA PREGUNTA 4 DEL modulo c
            listaRRAAobtencion = ooee.rraa_lista.all()
            listaRRAAobt_array = []
            for indexListaRRAAobt, itemListaRRAAobt in enumerate(listaRRAAobtencion):
                listaRRAAobt_array.append(str(itemListaRRAAobt))
                ws.cell(row = cont, column = 58).value = str(listaRRAAobt_array).replace('[','').replace(']','')
                ws.cell(row = cont, column = 58).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 58).font = Font(size = "8", name='Barlow')

            listaOOEEobtencion = ooee.ooee_lista.all()
            listaOOEEobt_array = []
            for indexListaOOEEobt, itemListaOOEEobt in enumerate(listaOOEEobtencion):
                listaOOEEobt_array.append(str(itemListaOOEEobt))
                ws.cell(row = cont, column = 62).value = str(listaOOEEobt_array)
                ws.cell(row = cont, column = 62).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 62).font = Font(size = "8", name='Barlow')
            
            listaNorma = ooee.norma.all()
            listaRequerimiento = ooee.requerimientos.all()
            listaPriUser = ooee.pri_usuarios.all()
            listaUnidadObs = ooee.uni_observacion.all()
            listaTipoOper = ooee.tipo_operacion.all()
            listaObtDato =  ooee.obt_dato.all()
            listaTipoProbabilistico = ooee.tipo_probabilistico.all()
            listaTipoNoProbabilistico = ooee.tipo_no_probabilistico.all()
            listaTipoMarco = ooee.tipo_marco.all()
            listaDocsDesarrollo = ooee.docs_des.all()
            ListaListaConc = ooee.lista_conc.all()
            listaClasificaciones =  ooee.nombre_cla.all()
            listaCobGeog = ooee.cob_geo.all()
            listaDesgeo = ooee.des_geo.all()
            listaDesZona = ooee.des_zona.all()
            listaDesGrupo = ooee.des_grupo.all()
            listaFuentes = ooee.fuentes.all()
            listaMediObtDato = ooee.med_obt.all()
            listaPeriodicidad = ooee.periodicidad.all()
            listaHerramProcesa = ooee.h_proc.all()
            listaAnalisResultados = ooee.a_resul.all()
            listaMedioDifusion = ooee.m_dif.all()
            listaFechaPublicacion = ooee.f_publi.all()
            listaFrecuenciaDifusion = ooee.fre_dif.all()
            listaProductosDifundir = ooee.pro_dif.all()
            listaOtrosProductos = ooee.otro_prod.all()

            for index,item in enumerate(listaNorma):
                indice = index
                if  str(item) == 'Ninguna' and index == indice:
                    #print("item", item)
                    ws.cell(row = cont, column = 23).value = str(item)
                    ws.cell(row = cont, column = 23).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 23).font = Font(size = "8", name='Barlow')
                   

            for indexReq, itemReq in enumerate(listaRequerimiento):
                indiceReq = indexReq
                if indexReq == indiceReq and str(itemReq) == 'Ninguno':
                    ws.cell(row = cont, column = 31).value = str(itemReq)
                    ws.cell(row = cont, column = 31).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 31).font = Font(size = "8", name='Barlow')
                   
            
            for indexPriUs, itemPrius in enumerate(listaPriUser):
                indicePriUser = indexPriUs
                if indexPriUs == indicePriUser and str(itemPrius) == 'Público en General':
                    ws.cell(row = cont, column = 42).value = str(itemPrius)
                    ws.cell(row = cont, column = 42).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 42).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 43).value = ooee.pob_obje
            ws.cell(row = cont, column = 43).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 43).font = Font(size = "8", name='Barlow')

            for indexUnidadObs, itemUnidadObs, in enumerate(listaUnidadObs):   
                if  str(itemUnidadObs) == "Empresa":
                    ws.cell(row = cont, column = 44).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 44).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 44).font = Font(size = "8", name='Barlow')
                if  str(itemUnidadObs) == "Establecimiento":   
                    ws.cell(row = cont, column = 45).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 45).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 45).font = Font(size = "8", name='Barlow')
                if  str(itemUnidadObs) == "Hogar":   
                    ws.cell(row = cont, column = 46).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 46).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 46).font = Font(size = "8", name='Barlow')
                if  str(itemUnidadObs) == "Persona":   
                    ws.cell(row = cont, column = 47).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 47).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 47).font = Font(size = "8", name='Barlow')
                if  str(itemUnidadObs) == "Unidad productora agropecuaria":   
                    ws.cell(row = cont, column = 48).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 48).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 48).font = Font(size = "8", name='Barlow')
                if  str(itemUnidadObs) == "Predio":   
                    ws.cell(row = cont, column = 49).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 49).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 49).font = Font(size = "8", name='Barlow')
                if  str(itemUnidadObs) == "Otra (s)":   
                    ws.cell(row = cont, column = 50).value = str(itemUnidadObs)
                    ws.cell(row = cont, column = 50).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 50).font = Font(size = "8", name='Barlow')

            for indexListaOper, itemListaOper in enumerate(listaTipoOper):
                if str(itemListaOper) == "Aprovechamiento de registro administrativo":
                    ws.cell(row = cont, column = 52).value = str(itemListaOper)
                    ws.cell(row = cont, column = 52).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 52).font = Font(size = "8", name='Barlow')
                if str(itemListaOper) == "Estadística derivada":
                    ws.cell(row = cont, column = 53).value = str(itemListaOper)
                    ws.cell(row = cont, column = 53).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 53).font = Font(size = "8", name='Barlow')
                if str(itemListaOper) == "Muestreo probabilístico":
                    ws.cell(row = cont, column = 54).value = str(itemListaOper)
                    ws.cell(row = cont, column = 54).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 54).font = Font(size = "8", name='Barlow')
                if str(itemListaOper) == "Muestreo no probabilístico":
                    ws.cell(row = cont, column = 55).value = str(itemListaOper)
                    ws.cell(row = cont, column = 55).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 55).font = Font(size = "8", name='Barlow')
                if str(itemListaOper) == "Censo":
                    ws.cell(row = cont, column = 56).value = str(itemListaOper)
                    ws.cell(row = cont, column = 56).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 56).font = Font(size = "8", name='Barlow')
            
            for indexObtDato, itemObtDato in enumerate(listaObtDato):
                if str(itemObtDato) == "Registro Administrativos":
                    ws.cell(row = cont, column = 57).value = str(itemObtDato)
                    ws.cell(row = cont, column = 57).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 57).font = Font(size = "8", name='Barlow')
                if str(itemObtDato) == "Operación Estadística":
                    ws.cell(row = cont, column = 61).value = str(itemObtDato)
                    ws.cell(row = cont, column = 61).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 61).font = Font(size = "8", name='Barlow')            

            for indexTipoProba, itemTipoProba in enumerate(listaTipoProbabilistico):
                if str(itemTipoProba) == "Muestreo aleatorio simple":
                    ws.cell(row = cont, column = 65).value = str(itemTipoProba)
                    ws.cell(row = cont, column = 65).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 65).font = Font(size = "8", name='Barlow')
                if str(itemTipoProba) == "Muestreo aleatorio sistemático":
                    ws.cell(row = cont, column = 66).value = str(itemTipoProba)
                    ws.cell(row = cont, column = 66).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 66).font = Font(size = "8", name='Barlow')
                if str(itemTipoProba) == "Muestreo aleatorio estratificado":
                    ws.cell(row = cont, column = 67).value = str(itemTipoProba)
                    ws.cell(row = cont, column = 67).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 67).font = Font(size = "8", name='Barlow')
                if str(itemTipoProba) == "Muestreo aleatorio por conglomerados":
                    ws.cell(row = cont, column = 68).value = str(itemTipoProba)
                    ws.cell(row = cont, column = 68).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 68).font = Font(size = "8", name='Barlow')
                if str(itemTipoProba) == "Otro":
                    ws.cell(row = cont, column = 69).value = str(itemTipoProba)
                    ws.cell(row = cont, column = 69).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 69).font = Font(size = "8", name='Barlow')

            for indexTipoNoProba, itemTipoNoProba in enumerate(listaTipoNoProbabilistico):
                if str(itemTipoNoProba) == "Muestreo por cuotas":
                    ws.cell(row = cont, column = 71).value = str(itemTipoNoProba)
                    ws.cell(row = cont, column = 71).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 71).font = Font(size = "8", name='Barlow')
                if str(itemTipoNoProba) == "Muestreo intencional o de conveniencia":
                    ws.cell(row = cont, column = 72).value = str(itemTipoNoProba)
                    ws.cell(row = cont, column = 72).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 72).font = Font(size = "8", name='Barlow')
                if str(itemTipoNoProba) == "Bola de nieve":
                    ws.cell(row = cont, column = 73).value = str(itemTipoNoProba)
                    ws.cell(row = cont, column = 73).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)                               
                    ws.cell(row = cont, column = 73).font = Font(size = "8", name='Barlow')
                if str(itemTipoNoProba) == "Muestreo discrecional":
                    ws.cell(row = cont, column = 74).value = str(itemTipoNoProba)
                    ws.cell(row = cont, column = 74).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 74).font = Font(size = "8", name='Barlow')
                if str(itemTipoNoProba) == "Otro":
                    ws.cell(row = cont, column = 75).value = str(itemTipoNoProba)
                    ws.cell(row = cont, column = 75).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)            
                    ws.cell(row = cont, column = 75).font = Font(size = "8", name='Barlow')

            if str(ooee.marco_estad) == "True":

                ws.cell(row = cont, column = 77).value = "Si"
                ws.cell(row = cont, column = 77).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 77).font = Font(size = "8", name='Barlow')
            
            elif str(ooee.marco_estad) == "False":
                ws.cell(row = cont, column = 77).value = "No"
                ws.cell(row = cont, column = 77).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 77).font = Font(size = "8", name='Barlow')
            
            for indexTipoMarco, itemTipoMarco in enumerate(listaTipoMarco):
                if str(itemTipoMarco) == "Marco de lista":
                    ws.cell(row = cont, column = 78).value = str(itemTipoMarco)
                    ws.cell(row = cont, column = 78).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 78).font = Font(size = "8", name='Barlow')
                if str(itemTipoMarco) == "Marco de área":
                    ws.cell(row = cont, column = 79).value = str(itemTipoMarco)
                    ws.cell(row = cont, column = 79).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 79).font = Font(size = "8", name='Barlow')
                if str(itemTipoMarco) == "Marco geoestadístico":
                    ws.cell(row = cont, column = 80).value = str(itemTipoMarco)
                    ws.cell(row = cont, column = 80).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 80).font = Font(size = "8", name='Barlow')
                if str(itemTipoMarco) == "Otro(s)":
                    ws.cell(row = cont, column = 81).value = str(itemTipoMarco)
                    ws.cell(row = cont, column = 81).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)                        
                    ws.cell(row = cont, column = 81).font = Font(size = "8", name='Barlow')
            
            for indexDocsDesarrollo, itemDocsDesarrollo in enumerate(listaDocsDesarrollo):
                if str(itemDocsDesarrollo) == "Metodología":
                    ws.cell(row = cont, column = 83).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 83).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 83).font = Font(size = "8", name='Barlow')
                if str(itemDocsDesarrollo) == "Ficha Metodológica (técnica)":
                    ws.cell(row = cont, column = 84).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 84).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 84).font = Font(size = "8", name='Barlow')
                if str(itemDocsDesarrollo) == "Hojas de vida de indicadores":
                    ws.cell(row = cont, column = 85).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 85).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 85).font = Font(size = "8", name='Barlow')
                if str(itemDocsDesarrollo) == "Manual operativo":
                    ws.cell(row = cont, column = 86).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 86).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 86).font = Font(size = "8", name='Barlow')
                if str(itemDocsDesarrollo) == "Diccionario de datos":
                    ws.cell(row = cont, column = 87).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 87).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 87).font = Font(size = "8", name='Barlow')
                if str(itemDocsDesarrollo) == "Otro (s)":
                    ws.cell(row = cont, column = 88).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 88).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 88).font = Font(size = "8", name='Barlow')
                if str(itemDocsDesarrollo) == "Ninguno":
                    ws.cell(row = cont, column = 90).value = str(itemDocsDesarrollo)
                    ws.cell(row = cont, column = 90).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 90).font = Font(size = "8", name='Barlow')
            
            for indexListaConc, itemListaConc in enumerate(ListaListaConc):
                if str(itemListaConc) == "DANE":
                    ws.cell(row = cont, column = 91).value = str(itemListaConc)
                    ws.cell(row = cont, column = 91).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 91).font = Font(size = "8", name='Barlow')
                if str(itemListaConc) == "Organismos internacionales":
                    ws.cell(row = cont, column = 92).value = str(itemListaConc)
                    ws.cell(row = cont, column = 92).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 92).font = Font(size = "8", name='Barlow')
                if str(itemListaConc) == "Otra entidad de orden nacional":
                    ws.cell(row = cont, column = 94).value = str(itemListaConc)
                    ws.cell(row = cont, column = 94).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 94).font = Font(size = "8", name='Barlow')
                if str(itemListaConc) == "Leyes, decretos, etc.":
                    ws.cell(row = cont, column = 96).value = str(itemListaConc)
                    ws.cell(row = cont, column = 96).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 96).font = Font(size = "8", name='Barlow')
                if str(itemListaConc) == "Creación propia":
                    ws.cell(row = cont, column = 98).value = str(itemListaConc)
                    ws.cell(row = cont, column = 98).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 98).font = Font(size = "8", name='Barlow')
                if str(itemListaConc) == "Otra (s)":
                    ws.cell(row = cont, column = 99).value = str(itemListaConc)
                    ws.cell(row = cont, column = 99).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 99).font = Font(size = "8", name='Barlow')
                if str(itemListaConc) == "Ninguno":
                    ws.cell(row = cont, column = 101).value = str(itemListaConc)
                    ws.cell(row = cont, column = 101).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 101).font = Font(size = "8", name='Barlow')

            if str(ooee.nome_clas) == "True":
                ws.cell(row = cont, column = 103).value = "Si"
                ws.cell(row = cont, column = 103).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 103).font = Font(size = "8", name='Barlow')
            
            elif str(ooee.nome_clas) == "False":
                ws.cell(row = cont, column = 103).value = "No"
                ws.cell(row = cont, column = 103).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 103).font = Font(size = "8", name='Barlow')
            
            
            clasifi_array = []
            for indexClasif, itemClasif in enumerate(listaClasificaciones):
                clasifi_array.append(str(itemClasif))
                ws.cell(row = cont, column = 104).value = str(clasifi_array)
                ws.cell(row = cont, column = 104).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 104).font = Font(size = "8", name='Barlow')
        
            for indexCobGeo, itemCobGeo in enumerate(listaCobGeog):
                
                if str(itemCobGeo) == "Nacional":
                    ws.cell(row = cont, column = 107).value = str(itemCobGeo)
                    ws.cell(row = cont, column = 107).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 107).font = Font(size = "8", name='Barlow')

                if str(itemCobGeo) == "Regional":
                    ws.cell(row = cont, column = 108).value = str(itemCobGeo)
                    ws.cell(row = cont, column = 108).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 108).font = Font(size = "8", name='Barlow')
                
                if str(itemCobGeo) == "Departamental":
                    ws.cell(row = cont, column = 111).value = str(itemCobGeo)
                    ws.cell(row = cont, column = 111).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 111).font = Font(size = "8", name='Barlow')
                
                if str(itemCobGeo) == "Áreas metropolitanas":
                    ws.cell(row = cont, column = 114).value = str(itemCobGeo)
                    ws.cell(row = cont, column = 114).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 114).font = Font(size = "8", name='Barlow')
                
                if str(itemCobGeo) == "Municipal":
                    ws.cell(row = cont, column = 117).value = str(itemCobGeo)
                    ws.cell(row = cont, column = 117).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 117).font = Font(size = "8", name='Barlow')
                
                if str(itemCobGeo) == "Otro (s)":
                    ws.cell(row = cont, column = 120).value = str(itemCobGeo)
                    ws.cell(row = cont, column = 120).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 120).font = Font(size = "8", name='Barlow')


                for indexDesgeo, itemDesgeo in enumerate(listaDesgeo): 
                    
                    if str(itemDesgeo) == "Nacional":
                        ws.cell(row = cont, column = 123).value = str(itemDesgeo)
                        ws.cell(row = cont, column = 123).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 123).font = Font(size = "8", name='Barlow')
                    
                    if str(itemDesgeo) == "Regional":
                        ws.cell(row = cont, column = 124).value = str(itemDesgeo)
                        ws.cell(row = cont, column = 124).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 124).font = Font(size = "8", name='Barlow')

                    if str(itemDesgeo) == "Departamental":
                        ws.cell(row = cont, column = 127).value = str(itemDesgeo)
                        ws.cell(row = cont, column = 127).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 127).font = Font(size = "8", name='Barlow')
                    
                    if str(itemDesgeo) == "Áreas metropolitanas":
                        ws.cell(row = cont, column = 130).value = str(itemDesgeo)
                        ws.cell(row = cont, column = 130).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 130).font = Font(size = "8", name='Barlow')
                    
                    if str(itemDesgeo) == "Municipal":
                        ws.cell(row = cont, column = 133).value = str(itemDesgeo)
                        ws.cell(row = cont, column = 133).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 133).font = Font(size = "8", name='Barlow')
                    
                    if str(itemDesgeo) == "Otro (s)":
                        ws.cell(row = cont, column = 136).value = str(itemDesgeo)
                        ws.cell(row = cont, column = 136).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 136).font = Font(size = "8", name='Barlow')
                
                
                for indexDesZona, itemDesZona in enumerate(listaDesZona):
                    if str(itemDesZona) == "Total":
                        ws.cell(row = cont, column = 139).value = str(itemDesZona)
                        ws.cell(row = cont, column = 139).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 138).font = Font(size = "8", name='Barlow')

                    if str(itemDesZona) == "Urbano":
                        ws.cell(row = cont, column = 140).value = str(itemDesZona)
                        ws.cell(row = cont, column = 140).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 140).font = Font(size = "8", name='Barlow')
                    
                    if str(itemDesZona) == "Rural":
                        ws.cell(row = cont, column = 141).value = str(itemDesZona)
                        ws.cell(row = cont, column = 141).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 141).font = Font(size = "8", name='Barlow')
                    
                
                for indexDesGrupo, itemDesGrupo in enumerate(listaDesGrupo):
                    
                    if str(itemDesGrupo) == "Sexo":
                        ws.cell(row = cont, column = 142).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 142).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 142).font = Font(size = "8", name='Barlow')

                    if str(itemDesGrupo) == "Edad":
                        ws.cell(row = cont, column = 143).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 143).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 143).font = Font(size = "8", name='Barlow')

                    if str(itemDesGrupo) == "Grupo étnico":
                        ws.cell(row = cont, column = 144).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 144).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 144).font = Font(size = "8", name='Barlow')

                    if str(itemDesGrupo) == "Discapacidad":
                        ws.cell(row = cont, column = 145).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 145).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 145).font = Font(size = "8", name='Barlow')

                    if str(itemDesGrupo) == "Estrato":
                        ws.cell(row = cont, column = 146).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 146).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 146).font = Font(size = "8", name='Barlow')

                    if str(itemDesGrupo) == "Otro, ¿cuál?":
                        ws.cell(row = cont, column = 147).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 147).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 147).font = Font(size = "8", name='Barlow')

                    if str(itemDesGrupo) == "Ninguno":
                        ws.cell(row = cont, column = 149).value = str(itemDesGrupo)
                        ws.cell(row = cont, column = 149).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 149).font = Font(size = "8", name='Barlow')

                ws.cell(row = cont, column = 150).value = ooee.ca_anual
                ws.cell(row = cont, column = 150).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 150).font = Font(size = "8", name='Barlow') 

                ws.cell(row = cont, column = 151).value = ooee.cb_anual
                ws.cell(row = cont, column = 151).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 151).font = Font(size = "8", name='Barlow')  
               
                for indexFuentes, itemFuentes in enumerate(listaFuentes):
                    
                    if str(itemFuentes) == "Recursos propios":
                        ws.cell(row = cont, column = 152).value = str(itemFuentes)
                        ws.cell(row = cont, column = 152).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 152).font = Font(size = "8", name='Barlow')

                    if str(itemFuentes) == "Aportes de otra entidad":
                        ws.cell(row = cont, column = 153).value = str(itemFuentes)
                        ws.cell(row = cont, column = 153).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 153).font = Font(size = "8", name='Barlow')

                    if str(itemFuentes) == "Cooperación internacional":
                        ws.cell(row = cont, column = 154).value = str(itemFuentes)
                        ws.cell(row = cont, column = 154).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 154).font = Font(size = "8", name='Barlow')

                    if str(itemFuentes) == "Otro (s)":
                        ws.cell(row = cont, column = 155).value = str(itemFuentes)
                        ws.cell(row = cont, column = 155).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 155).font = Font(size = "8", name='Barlow')
                
                ## 158 y 159 si se desea agregar preguntas 14 y 15 del modulo C

                for indexMediObtDato, itemMediObtDato in enumerate(listaMediObtDato):

                    if str(itemMediObtDato) == "Formulario físico":
                        ws.cell(row = cont, column = 159).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 159).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 159).font = Font(size = "8", name='Barlow')
                    
                    if str(itemMediObtDato) == "Formulario electrónico":
                        ws.cell(row = cont, column = 160).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 160).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 160).font = Font(size = "8", name='Barlow')

                    if str(itemMediObtDato) == "Dispositivo Móvil de Captura [DMC]":
                        ws.cell(row = cont, column = 161).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 161).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 161).font = Font(size = "8", name='Barlow')

                    if str(itemMediObtDato) == "Sistema de Información":
                        ws.cell(row = cont, column = 162).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 162).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 162).font = Font(size = "8", name='Barlow')

                    if str(itemMediObtDato) == "Percepción remota (imágenes satelitales, fotos, sensores, etc)":
                        ws.cell(row = cont, column = 164).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 164).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 164).font = Font(size = "8", name='Barlow')

                    if str(itemMediObtDato) == "Base de datos de registro administrativo":
                        ws.cell(row = cont, column = 165).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 165).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 165).font = Font(size = "8", name='Barlow')

                    if str(itemMediObtDato) == "Resultados estadísticos de otra operación estadística":
                        ws.cell(row = cont, column = 166).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 166).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 166).font = Font(size = "8", name='Barlow')

                    if str(itemMediObtDato) == "Otro (s)":
                        ws.cell(row = cont, column = 167).value = str(itemMediObtDato)
                        ws.cell(row = cont, column = 167).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 167).font = Font(size = "8", name='Barlow')

                for indexPeriodicidad, itemPeriodicidad in enumerate(listaPeriodicidad):
                    if str(itemPeriodicidad) == "Anual":
                        ws.cell(row = cont, column = 169).value = str(itemPeriodicidad)
                        ws.cell(row = cont, column = 169).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 169).font = Font(size = "8", name='Barlow')
                    
                    if str(itemPeriodicidad) == "Semestral":
                        ws.cell(row = cont, column = 170).value = str(itemPeriodicidad)
                        ws.cell(row = cont, column = 170).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 170).font = Font(size = "8", name='Barlow')

                    if str(itemPeriodicidad) == "Trimestral":
                        ws.cell(row = cont, column = 171).value = str(itemPeriodicidad)
                        ws.cell(row = cont, column = 171).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 171).font = Font(size = "8", name='Barlow')

                    if str(itemPeriodicidad) == "Mensual":
                        ws.cell(row = cont, column = 172).value = str(itemPeriodicidad)
                        ws.cell(row = cont, column = 172).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 172).font = Font(size = "8", name='Barlow')
                    
                    if str(itemPeriodicidad) == "Diaria":
                        ws.cell(row = cont, column = 173).value = str(itemPeriodicidad)
                        ws.cell(row = cont, column = 173).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 173).font = Font(size = "8", name='Barlow')
                    
                    if str(itemPeriodicidad) == "Otra":
                        ws.cell(row = cont, column = 174).value = str(itemPeriodicidad)
                        ws.cell(row = cont, column = 174).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 174).font = Font(size = "8", name='Barlow')

                for indexHerramProcesa, itemHerramProcesa in enumerate(listaHerramProcesa):
                    
                    if str(itemHerramProcesa) == "Excel" :
                        ws.cell(row = cont, column = 176).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 176).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 176).font = Font(size = "8", name='Barlow')

                    if str(itemHerramProcesa) == "Access" :
                        ws.cell(row = cont, column = 177).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 177).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 177).font = Font(size = "8", name='Barlow')

                    if str(itemHerramProcesa) == "R" :
                        ws.cell(row = cont, column = 178).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 178).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 178).font = Font(size = "8", name='Barlow')

                    if str(itemHerramProcesa) == "SAS" :
                        ws.cell(row = cont, column = 179).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 179).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 179).font = Font(size = "8", name='Barlow')

                    if str(itemHerramProcesa) == "SPSS" :
                        ws.cell(row = cont, column = 180).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 180).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 180).font = Font(size = "8", name='Barlow')
                    
                    if str(itemHerramProcesa) == "Oracle" :
                        ws.cell(row = cont, column = 181).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 181).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 181).font = Font(size = "8", name='Barlow')

                    if str(itemHerramProcesa) == "Stata" :
                        ws.cell(row = cont, column = 182).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 182).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 182).font = Font(size = "8", name='Barlow')
                    
                    if str(itemHerramProcesa) == "Otra (s)" :
                        ws.cell(row = cont, column = 183).value = str(itemHerramProcesa)
                        ws.cell(row = cont, column = 183).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 183).font = Font(size = "8", name='Barlow')

                ws.cell(row = cont, column = 185).value = ooee.descrip_proces
                ws.cell(row = cont, column = 185).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 185).font = Font(size = "8", name='Barlow')  
               
                for indexAnalisResultados, itemAnalisResultados in enumerate(listaAnalisResultados):
                    
                    if str(itemAnalisResultados) == "Consistencia y validación de datos":
                        ws.cell(row = cont, column = 186).value = str(itemAnalisResultados)
                        ws.cell(row = cont, column = 186).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 186).font = Font(size = "8", name='Barlow')

                    if str(itemAnalisResultados) == "Análisis de tendencias y series de tiempo":
                        ws.cell(row = cont, column = 187).value = str(itemAnalisResultados)
                        ws.cell(row = cont, column = 187).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 187).font = Font(size = "8", name='Barlow')

                    if str(itemAnalisResultados) == "Análisis de contexto de los resultados":
                        ws.cell(row = cont, column = 188).value = str(itemAnalisResultados)
                        ws.cell(row = cont, column = 188).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 188).font = Font(size = "8", name='Barlow') 

                    if str(itemAnalisResultados) == "Otro":
                        ws.cell(row = cont, column = 189).value = str(itemAnalisResultados)
                        ws.cell(row = cont, column = 189).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 189).font = Font(size = "8", name='Barlow')  

                for indexMedioDifusion, itemMedioDifusion in enumerate(listaMedioDifusion):

                    if str(itemMedioDifusion) == "Página web":
                        ws.cell(row = cont, column = 191).value = str(itemMedioDifusion)
                        ws.cell(row = cont, column = 191).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 191).font = Font(size = "8", name='Barlow') 
                    
                    if str(itemMedioDifusion) == "Medio físico":
                        ws.cell(row = cont, column = 192).value = str(itemMedioDifusion)
                        ws.cell(row = cont, column = 192).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 192).font = Font(size = "8", name='Barlow') 

                    if str(itemMedioDifusion) == "Medio electrónico":
                        ws.cell(row = cont, column = 193).value = str(itemMedioDifusion)
                        ws.cell(row = cont, column = 193).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 193).font = Font(size = "8", name='Barlow') 

                    if str(itemMedioDifusion) == "Otro":
                        ws.cell(row = cont, column = 194).value = str(itemMedioDifusion)
                        ws.cell(row = cont, column = 194).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        ws.cell(row = cont, column = 194).font = Font(size = "8", name='Barlow')  
                
                
                ws.cell(row = cont, column = 196).value = str(ooee.res_est_url)
                ws.cell(row = cont, column = 196).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 196).font = Font(size = "8", name='Barlow')  

                ws.cell(row = cont, column = 197).value = str(ooee.dispo_desde)
                ws.cell(row = cont, column = 197).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 197).font = Font(size = "8", name='Barlow')  

                ws.cell(row = cont, column = 198).value = str(ooee.dispo_hasta)
                ws.cell(row = cont, column = 198).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 198).font = Font(size = "8", name='Barlow')  
                
            for indexFechaPublicacion, itemFechaPublicacion in enumerate(listaFechaPublicacion):  
                
                if str(itemFechaPublicacion) == "fecha Publicación":
                    ws.cell(row = cont, column = 199).value = str(itemFechaPublicacion)
                    ws.cell(row = cont, column = 199).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 199).font = Font(size = "8", name='Barlow')

                if str(itemFechaPublicacion) == "No sabe":
                    ws.cell(row = cont, column = 201).value = str(itemFechaPublicacion)
                    ws.cell(row = cont, column = 201).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 201).font = Font(size = "8", name='Barlow')

                if str(itemFechaPublicacion) == "No hay":
                    ws.cell(row = cont, column = 202).value = str(itemFechaPublicacion)
                    ws.cell(row = cont, column = 202).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 202).font = Font(size = "8", name='Barlow')

            for indexFrecuenciaDifusion, itemFrecuenciaDifusion in enumerate(listaFrecuenciaDifusion):
                
                if str(itemFrecuenciaDifusion) == "Anual":
                    ws.cell(row = cont, column = 204).value = str(itemFrecuenciaDifusion)
                    ws.cell(row = cont, column = 204).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 204).font = Font(size = "8", name='Barlow')

                if str(itemFrecuenciaDifusion) == "Semestral":
                    ws.cell(row = cont, column = 205).value = str(itemFrecuenciaDifusion)
                    ws.cell(row = cont, column = 205).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 205).font = Font(size = "8", name='Barlow')

                if str(itemFrecuenciaDifusion) == "Trimestral":
                    ws.cell(row = cont, column = 206).value = str(itemFrecuenciaDifusion)
                    ws.cell(row = cont, column = 206).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 206).font = Font(size = "8", name='Barlow')

                if str(itemFrecuenciaDifusion) == "Mensual":
                    ws.cell(row = cont, column = 207).value = str(itemFrecuenciaDifusion)
                    ws.cell(row = cont, column = 207).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 207).font = Font(size = "8", name='Barlow')

                if str(itemFrecuenciaDifusion) == "Otro (s)":
                    ws.cell(row = cont, column = 208).value = str(itemFrecuenciaDifusion)
                    ws.cell(row = cont, column = 208).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 208).font = Font(size = "8", name='Barlow')

                if str(itemFrecuenciaDifusion) == "No está definido":
                    ws.cell(row = cont, column = 209).value = str(itemFrecuenciaDifusion)
                    ws.cell(row = cont, column = 209).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 209).font = Font(size = "8", name='Barlow')

            for indexProductosDifundir, itemProductosDifundir in enumerate(listaProductosDifundir):

                if str(itemProductosDifundir) == "Cuadros de salida (tablas)":
                    ws.cell(row = cont, column = 211).value = str(itemProductosDifundir)
                    ws.cell(row = cont, column = 211).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 211).font = Font(size = "8", name='Barlow')
                
                if str(itemProductosDifundir) == "Boletín estadístico":
                    ws.cell(row = cont, column = 212).value = str(itemProductosDifundir)
                    ws.cell(row = cont, column = 212).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 212).font = Font(size = "8", name='Barlow')

                if str(itemProductosDifundir) == "Anuario":
                    ws.cell(row = cont, column = 213).value = str(itemProductosDifundir)
                    ws.cell(row = cont, column = 213).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 213).font = Font(size = "8", name='Barlow')

                if str(itemProductosDifundir) == "Mapas estadísticos":
                    ws.cell(row = cont, column = 214).value = str(itemProductosDifundir)
                    ws.cell(row = cont, column = 214).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 214).font = Font(size = "8", name='Barlow')

                if str(itemProductosDifundir) == "Bases de datos interactivas":
                    ws.cell(row = cont, column = 215).value = str(itemProductosDifundir)
                    ws.cell(row = cont, column = 215).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 215).font = Font(size = "8", name='Barlow')

                if str(itemProductosDifundir) == "Otro (s)":
                    ws.cell(row = cont, column = 216).value = str(itemProductosDifundir)
                    ws.cell(row = cont, column = 216).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 216).font = Font(size = "8", name='Barlow')

            for indexOtrosProductos, itemOtrosProductos in enumerate(listaOtrosProductos):

                if str(itemOtrosProductos) == "Series históricas":
                    ws.cell(row = cont, column = 218).value = str(itemOtrosProductos)
                    ws.cell(row = cont, column = 218).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 218).font = Font(size = "8", name='Barlow')

                if str(itemOtrosProductos) == "Microdatos":
                    ws.cell(row = cont, column = 221).value = str(itemOtrosProductos)
                    ws.cell(row = cont, column = 221).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 221).font = Font(size = "8", name='Barlow')
                
                if str(itemOtrosProductos) == "Documentos metodológicos":
                    ws.cell(row = cont, column = 224).value = str(itemOtrosProductos)
                    ws.cell(row = cont, column = 224).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 224).font = Font(size = "8", name='Barlow')
                
                if str(itemOtrosProductos) == "Calendario de difusión":
                    ws.cell(row = cont, column = 226).value = str(itemOtrosProductos)
                    ws.cell(row = cont, column = 226).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 226).font = Font(size = "8", name='Barlow')
                    
                if str(itemOtrosProductos) == "Ninguna":
                    ws.cell(row = cont, column = 227).value = str(itemOtrosProductos)
                    ws.cell(row = cont, column = 227).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 227).font = Font(size = "8", name='Barlow')

            if str(ooee.conoce_otra) == "True":
                ws.cell(row = cont, column = 228).value = "Si"
                ws.cell(row = cont, column = 228).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 228).font = Font(size = "8", name='Barlow')
            elif str(ooee.conoce_otra) == "False":
                ws.cell(row = cont, column = 228).value = "No"
                ws.cell(row = cont, column = 228).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 228).font = Font(size = "8", name='Barlow') 

            if str(ooee.hp_siste_infor) == "True":
                ws.cell(row = cont, column = 231).value = "Si"
                ws.cell(row = cont, column = 231).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 231).font = Font(size = "8", name='Barlow')
            elif str(ooee.hp_siste_infor) == "False":
                ws.cell(row = cont, column = 231).value = "No"
                ws.cell(row = cont, column = 231).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 231).font = Font(size = "8", name='Barlow') 

            ws.cell(row = cont, column = 233).value = ooee.observaciones
            ws.cell(row = cont, column = 233).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 233).font = Font(size = "8", name='Barlow') 

            ws.cell(row = cont, column = 234).value = str(ooee.anexos)
            ws.cell(row = cont, column = 234).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 234).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 235).value = str(ooee.estado_oe_tematico)
            ws.cell(row = cont, column = 235).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 235).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 236).value = str(ooee.validacion_oe_tematico)
            ws.cell(row = cont, column = 236).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 236).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 237).value = str(ooee.nombre_est)
            ws.cell(row = cont, column = 237).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 237).font = Font(size = "8", name='Barlow')
            ## Hoja 1 SI desea continuar despues de la linea 236  ------------------------->  
        
            cont+=1
        #print("----------filter ------------>", normaTextList)
        for indexNorma, itemNorma in enumerate(normaTextList):
            
            indexNorma = indexNorma + 4

            ws.cell(row = indexNorma, column = 19).value = str(itemNorma.cp_d)
            ws.cell(row = indexNorma, column = 19).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 19).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexNorma, column = 20).value = str(itemNorma.ley_d)
            ws.cell(row = indexNorma, column = 20).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 20).font = Font(size = "8", name='Barlow')
                    
            ws.cell(row = indexNorma, column = 21).value = str(itemNorma.decreto_d)
            ws.cell(row = indexNorma, column = 21).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 21).font = Font(size = "8", name='Barlow')
                    
            ws.cell(row = indexNorma, column = 22).value = str(itemNorma.otra_d)
            ws.cell(row = indexNorma, column = 22).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 22).font = Font(size = "8", name='Barlow') 

        for indexRequerimiento, itemRequerimiento in enumerate(requerimientoTextList):

            indexRequerimiento = indexRequerimiento + 4

            ws.cell(row = indexRequerimiento, column = 24).value = str(itemRequerimiento.ri_ods)
            ws.cell(row = indexRequerimiento, column = 24).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexRequerimiento, column = 24).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexRequerimiento, column = 25).value = str(itemRequerimiento.ri_ocde)
            ws.cell(row = indexRequerimiento, column = 25).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexRequerimiento, column = 25).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexRequerimiento, column = 26).value = str(itemRequerimiento.ri_ci)
            ws.cell(row = indexRequerimiento, column = 26).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexRequerimiento, column = 26).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexRequerimiento, column = 27).value = str(itemRequerimiento.ri_pnd)
            ws.cell(row = indexRequerimiento, column = 27).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexRequerimiento, column = 27).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexRequerimiento, column = 28).value = str(itemRequerimiento.ri_cem)
            ws.cell(row = indexRequerimiento, column = 28).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  
            ws.cell(row = indexRequerimiento, column = 28).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexRequerimiento, column = 29).value = str(itemRequerimiento.ri_pstc)
            ws.cell(row = indexRequerimiento, column = 29).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexRequerimiento, column = 29).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexRequerimiento, column = 30).value = str(itemRequerimiento.ri_otro)
            ws.cell(row = indexRequerimiento, column = 30).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexRequerimiento, column = 30).font = Font(size = "8", name='Barlow')

        for indexPriUser, itemPriUser in enumerate(usuariosPrinTextList):

            indexPriUser = indexPriUser + 4

            ws.cell(row = indexPriUser, column = 32).value = str(itemPriUser.org_int)
            ws.cell(row = indexPriUser, column = 32).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 32).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexPriUser, column = 33).value = str(itemPriUser.pres_rep)
            ws.cell(row = indexPriUser, column = 33).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 33).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexPriUser, column = 34).value = str(itemPriUser.misnit)
            ws.cell(row = indexPriUser, column = 34).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 34).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexPriUser, column = 35).value = str(itemPriUser.org_cont)
            ws.cell(row = indexPriUser, column = 35).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 35).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexPriUser, column = 36).value = str(itemPriUser.o_ent_o_nac)
            ws.cell(row = indexPriUser, column = 36).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 36).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexPriUser, column = 37).value = str(itemPriUser.ent_o_terr)
            ws.cell(row = indexPriUser, column = 37).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 37).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexPriUser, column = 38).value = str(itemPriUser.gremios)
            ws.cell(row = indexPriUser, column = 38).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 38).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexPriUser, column = 39).value = str(itemPriUser.ent_privadas)
            ws.cell(row = indexPriUser, column = 39).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 39).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexPriUser, column = 40).value = str(itemPriUser.dep_misma_entidad)
            ws.cell(row = indexPriUser, column = 40).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 40).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexPriUser, column = 41).value = str(itemPriUser.academia)
            ws.cell(row = indexPriUser, column = 41).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = indexPriUser, column = 41).font = Font(size = "8", name='Barlow')

        for indexUniObs, itemUniObs in enumerate(unidadObserTextList):

            indexUniObs = indexUniObs + 4

            ws.cell(row = indexUniObs, column = 51).value = str(itemUniObs.mc_otra)
            ws.cell(row = indexUniObs, column = 51).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexUniObs, column = 51).font = Font(size = "8", name='Barlow')

        for indexObtencionDato, itemObtencionDato in enumerate(obtencionDatoTextList):
    
            indexObtencionDato = indexObtencionDato + 4

            ## si la opcion marcada es Registro Administrativo
            ws.cell(row = indexObtencionDato, column = 59).value = str(itemObtencionDato.mc_ra_cual)
            ws.cell(row = indexObtencionDato, column = 59).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexObtencionDato, column = 59).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexObtencionDato, column = 60).value = str(itemObtencionDato.mc_ra_entidad)
            ws.cell(row = indexObtencionDato, column = 60).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexObtencionDato, column = 60).font = Font(size = "8", name='Barlow')
                        
            ## si la opcion marcada es operacion estadística
            ws.cell(row = indexObtencionDato, column = 63).value = str(itemObtencionDato.mc_oe_cual)
            ws.cell(row = indexObtencionDato, column = 63).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexObtencionDato, column = 63).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexObtencionDato, column = 64).value = str(itemObtencionDato.mc_oe_entidad)
            ws.cell(row = indexObtencionDato, column = 64).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexObtencionDato, column = 64).font = Font(size = "8", name='Barlow')

        for indexMuesProbabilistico, itemMuesProbabilistico  in enumerate(muestProbaTextList):

            indexMuesProbabilistico = indexMuesProbabilistico + 4

            ws.cell(row = indexMuesProbabilistico, column = 70).value = str(itemMuesProbabilistico.prob_otro)
            ws.cell(row = indexMuesProbabilistico, column = 70).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexMuesProbabilistico, column = 70).font = Font(size = "8", name='Barlow')

        for indexMuesNoProbabilistico, itemMuesNoProbabilistico in enumerate(muestNoProbaTextList):

            indexMuesNoProbabilistico = indexMuesNoProbabilistico + 4

            ws.cell(row = indexMuesNoProbabilistico, column = 76).value = str(itemMuesNoProbabilistico.no_prob_otro)
            ws.cell(row = indexMuesNoProbabilistico, column = 76).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexMuesNoProbabilistico, column = 76).font = Font(size = "8", name='Barlow')

        for indexTipMar, itemTipMar in enumerate(tipoMarcoTextList):

            indexTipMar = indexTipMar + 4
                
            ws.cell(row = indexTipMar, column = 82).value = str(itemTipMar.otro_tipo_marco)
            ws.cell(row = indexTipMar, column = 82).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexTipMar, column = 82).font = Font(size = "8", name='Barlow')

        for indexdocsDesa, itemdocsDesa in enumerate(docsDesarrolloTextList):

            indexdocsDesa = indexdocsDesa + 4
                
            ws.cell(row = indexdocsDesa, column = 89).value = str(itemdocsDesa.otro_docs)
            ws.cell(row = indexdocsDesa, column = 89).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdocsDesa, column = 89).font = Font(size = "8", name='Barlow')

        for indexConceptosEstanda, itemConceptosEstanda in enumerate(conceptosEstandaTextList):
            
            indexConceptosEstanda = indexConceptosEstanda + 4 

            ws.cell(row = indexConceptosEstanda, column = 93).value = str(itemConceptosEstanda.org_in_cuales)
            ws.cell(row = indexConceptosEstanda, column = 93).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConceptosEstanda, column = 93).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexConceptosEstanda, column = 95).value = str(itemConceptosEstanda.ent_ordnac_cuales)
            ws.cell(row = indexConceptosEstanda, column = 95).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConceptosEstanda, column = 95).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexConceptosEstanda, column = 97).value = str(itemConceptosEstanda.leye_dec_cuales)
            ws.cell(row = indexConceptosEstanda, column = 97).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConceptosEstanda, column = 97).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexConceptosEstanda, column = 100).value = str(itemConceptosEstanda.otra_cual_conp)
            ws.cell(row = indexConceptosEstanda, column = 100).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConceptosEstanda, column = 100).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexConceptosEstanda, column = 102).value = str(itemConceptosEstanda.ningu_pq)
            ws.cell(row = indexConceptosEstanda, column = 102).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConceptosEstanda, column = 102).font = Font(size = "8", name='Barlow')

        for indexClasificaciones, itemClasificaciones in enumerate(clasificacionesTextList):
    
            indexClasificaciones = indexClasificaciones + 4

            ws.cell(row = indexClasificaciones, column = 105).value = str(itemClasificaciones.otra_cual_clas)
            ws.cell(row = indexClasificaciones, column = 105).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexClasificaciones, column = 105).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexClasificaciones, column = 106).value = str(itemClasificaciones.no_pq)
            ws.cell(row = indexClasificaciones, column = 106).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexClasificaciones, column = 106).font = Font(size = "8", name='Barlow')

        for indexCoberturaGeog, itemCoberturaGeog in enumerate(coberGeogTextList):

            indexCoberturaGeog = indexCoberturaGeog + 4
                
            ws.cell(row = indexCoberturaGeog, column = 109).value = str(itemCoberturaGeog.tot_regional)
            ws.cell(row = indexCoberturaGeog, column = 109).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 109).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 110).value = str(itemCoberturaGeog.cual_regional)
            ws.cell(row = indexCoberturaGeog, column = 110).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 110).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 112).value = str(itemCoberturaGeog.tot_dep)
            ws.cell(row = indexCoberturaGeog, column = 112).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 112).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 113).value = str(itemCoberturaGeog.cual_dep)
            ws.cell(row = indexCoberturaGeog, column = 113).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 113).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 115).value = str(itemCoberturaGeog.tot_are_metr)
            ws.cell(row = indexCoberturaGeog, column = 115).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 115).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 116).value = str(itemCoberturaGeog.cual_are_metr)
            ws.cell(row = indexCoberturaGeog, column = 116).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 116).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 118).value = str(itemCoberturaGeog.tot_mun)
            ws.cell(row = indexCoberturaGeog, column = 118).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 118).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 119).value = str(itemCoberturaGeog.cual_mun)
            ws.cell(row = indexCoberturaGeog, column = 119).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 119).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 121).value = str(itemCoberturaGeog.tot_otro)
            ws.cell(row = indexCoberturaGeog, column = 121).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 121).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 122).value = str(itemCoberturaGeog.cual_otro)
            ws.cell(row = indexCoberturaGeog, column = 122).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 122).font = Font(size = "8", name='Barlow')

        for indexdesagreInfo, itemdesagreInfo in enumerate(desagreInfoTextList):

            indexdesagreInfo = indexdesagreInfo + 4

            ws.cell(row = indexdesagreInfo, column = 125).value = str(itemdesagreInfo.des_tot_regional)
            ws.cell(row = indexdesagreInfo, column = 125).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 125).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexdesagreInfo, column = 126).value = str(itemdesagreInfo.des_cual_regional)
            ws.cell(row = indexdesagreInfo, column = 126).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 126).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexdesagreInfo, column = 128).value = str(itemdesagreInfo.des_tot_dep)
            ws.cell(row = indexdesagreInfo, column = 128).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 128).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 129).value = str(itemdesagreInfo.des_cual_dep)
            ws.cell(row = indexdesagreInfo, column = 129).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 129).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 131).value = str(itemdesagreInfo.des_tot_are_metr)
            ws.cell(row = indexdesagreInfo, column = 131).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 131).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 132).value = str(itemdesagreInfo.des_cual_are_metr)
            ws.cell(row = indexdesagreInfo, column = 132).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 132).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 134).value = str(itemdesagreInfo.des_tot_mun)
            ws.cell(row = indexdesagreInfo, column = 134).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 134).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 135).value = str(itemdesagreInfo.des_cual_mun)
            ws.cell(row = indexdesagreInfo, column = 135).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 135).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 137).value = str(itemdesagreInfo.des_tot_otro)
            ws.cell(row = indexdesagreInfo, column = 137).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 137).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 138).value = str(itemdesagreInfo.des_cual_otro)
            ws.cell(row = indexdesagreInfo, column = 138).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 138).font = Font(size = "8", name='Barlow')
                
            ws.cell(row = indexdesagreInfo, column = 148).value = str(itemdesagreInfo.des_grupo_otro)
            ws.cell(row = indexdesagreInfo, column = 148).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexdesagreInfo, column = 148).font = Font(size = "8", name='Barlow')

        for indexFuenteFinan, itemFuenteFinan in enumerate(fuenteFinanTextList):

            indexFuenteFinan = indexFuenteFinan + 4

            ws.cell(row = indexFuenteFinan, column = 156).value = str(itemFuenteFinan.r_otros)
            ws.cell(row = indexFuenteFinan, column = 156).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexFuenteFinan, column = 156).font = Font(size = "8", name='Barlow')

        for indexMedioDatos, itemMedioDatos in enumerate(medioDatosTextList):

            indexMedioDatos = indexMedioDatos + 4
                
            ws.cell(row = indexMedioDatos, column = 163).value = str(itemMedioDatos.sis_info)
            ws.cell(row = indexMedioDatos, column = 163).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexMedioDatos, column = 163).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexMedioDatos, column = 168).value = str(itemMedioDatos.md_otro)
            ws.cell(row = indexMedioDatos, column = 168).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexMedioDatos, column = 168).font = Font(size = "8", name='Barlow')

        for indexPeriodicidadOE, itemPeriodicidadOE in enumerate(periodicidadTextList):

            indexPeriodicidadOE = indexPeriodicidadOE + 4

            ws.cell(row = indexMedioDatos, column = 175).value = str(itemPeriodicidadOE.per_otro)
            ws.cell(row = indexMedioDatos, column = 175).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexMedioDatos, column = 175).font = Font(size = "8", name='Barlow')

        for indexHerraProcesami, itemHerraProcesami in enumerate(herraProcesamiTextList):

            indexHerraProcesami = indexHerraProcesami + 4

            ws.cell(row = indexHerraProcesami, column = 184).value = str(itemHerraProcesami.herr_otro)
            ws.cell(row = indexHerraProcesami, column = 184).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexHerraProcesami, column = 184).font = Font(size = "8", name='Barlow')


        for indexAnalResul, itemAnalResul in enumerate(analResultTextList):
           
            indexAnalResul = indexAnalResul + 4

            ws.cell(row = indexAnalResul, column = 190).value = str(itemAnalResul.ana_otro)
            ws.cell(row = indexAnalResul, column = 190).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexAnalResul, column = 190).font = Font(size = "8", name='Barlow')

        for indexMedioDifus, itemMedioDifus in enumerate(medioDifusTextList):
            
            indexMedioDifus = indexMedioDifus + 4

            ws.cell(row = indexMedioDifus, column = 195).value = str(itemMedioDifus.dif_otro)
            ws.cell(row = indexMedioDifus, column = 195).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexMedioDifus, column = 195).font = Font(size = "8", name='Barlow')

        for indexFechaPublic, itemFechaPublic in enumerate(fechaPublicTextList):

            indexFechaPublic = indexFechaPublic + 4

            ws.cell(row = indexFechaPublic, column = 200).value = str(itemFechaPublic.fecha)
            ws.cell(row = indexFechaPublic, column = 200).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexFechaPublic, column = 200).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexFechaPublic, column = 203).value = str(itemFechaPublic.no_hay)
            ws.cell(row = indexFechaPublic, column = 203).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexFechaPublic, column = 203).font = Font(size = "8", name='Barlow')

        for indexFrecuencDifus, itemFrecuencDifus in enumerate(frecuenciaDifusionTextList):

            indexFrecuencDifus = indexFrecuencDifus + 4

            ws.cell(row = indexFrecuencDifus, column = 210).value = str(itemFrecuencDifus.no_definido)
            ws.cell(row = indexFrecuencDifus, column = 210).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexFrecuencDifus, column = 210).font = Font(size = "8", name='Barlow')

        for indexProductDifund, itemProductDifund in enumerate(productosDifundirTextList):

            indexProductDifund = indexProductDifund + 4

            ws.cell(row = indexProductDifund, column = 217).value = str(itemProductDifund.difundir_otro)
            ws.cell(row = indexProductDifund, column = 217).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexProductDifund, column = 217).font = Font(size = "8", name='Barlow')

        for indexOtrosProduct, itemOtrosProduct in enumerate(otrosProductosTextList):

            indexOtrosProduct = indexOtrosProduct + 4

            ws.cell(row = indexOtrosProduct, column = 219).value = str(itemOtrosProduct.ser_hist_desde)
            ws.cell(row = indexOtrosProduct, column = 219).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexOtrosProduct, column = 219).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexOtrosProduct, column = 220).value = str(itemOtrosProduct.ser_hist_hasta)
            ws.cell(row = indexOtrosProduct, column = 220).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexOtrosProduct, column = 220).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexOtrosProduct, column = 222).value = str(itemOtrosProduct.microdatos_desde)
            ws.cell(row = indexOtrosProduct, column = 222).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexOtrosProduct, column = 222).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexOtrosProduct, column = 223).value = str(itemOtrosProduct.microdatos_hasta)
            ws.cell(row = indexOtrosProduct, column = 223).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexOtrosProduct, column = 223).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexOtrosProduct, column = 225).value = str(itemOtrosProduct.op_url)
            ws.cell(row = indexOtrosProduct, column = 225).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexOtrosProduct, column = 225).font = Font(size = "8", name='Barlow')

        for indexResultSimi, itemResultSimi in enumerate(resultadosSimiTextList):
            
            indexResultSimi = indexResultSimi + 4

            ws.cell(row = indexResultSimi, column = 229).value = str(itemResultSimi.rs_entidad)
            ws.cell(row = indexResultSimi, column = 229).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexResultSimi, column = 229).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexResultSimi, column = 230).value = str(itemResultSimi.rs_oe)
            ws.cell(row = indexResultSimi, column = 230).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexResultSimi, column = 230).font = Font(size = "8", name='Barlow')
            
        for indexhpSistemaInfo, itemhpSistemaInfo in enumerate(hpSistemaInfoTextList):
            
            indexhpSistemaInfo = indexhpSistemaInfo + 4

            ws.cell(row = indexhpSistemaInfo, column = 232).value = str(itemhpSistemaInfo.si_cual)
            ws.cell(row = indexhpSistemaInfo, column = 232).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexhpSistemaInfo, column = 232).font = Font(size = "8", name='Barlow')

        for row in range(4, ws.max_row + 1):
            ws.row_dimensions[row].height = 90
            
        #############  Hoja 2 entidadades y fases  ##################

        sheet2.merge_cells('A2:B2')
        sheet2.merge_cells('C2:HZ2')
        sheet2.merge_cells('A1:HZ1')

        def set_border(sheet2, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet2[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet2,'A1:HZ'+str(ooees.count()+3))
    
        #row dimensions
        sheet2.row_dimensions[1].height = 55
        sheet2.row_dimensions[2].height = 40

        # column width
        sheet2.column_dimensions['A'].width = 20
        sheet2.column_dimensions['B'].width = 20
        sheet2.column_dimensions['C'].width = 20
        sheet2.column_dimensions['D'].width = 20
        sheet2.column_dimensions['E'].width = 20
        sheet2.column_dimensions['F'].width = 20
        sheet2.column_dimensions['G'].width = 20
        sheet2.column_dimensions['H'].width = 20
        sheet2.column_dimensions['I'].width = 20
        sheet2.column_dimensions['J'].width = 20
        sheet2.column_dimensions['K'].width = 20
        sheet2.column_dimensions['L'].width = 20
        sheet2.column_dimensions['M'].width = 20
        sheet2.column_dimensions['N'].width = 20
        sheet2.column_dimensions['O'].width = 20
        sheet2.column_dimensions['P'].width = 20
        sheet2.column_dimensions['Q'].width = 20
        sheet2.column_dimensions['R'].width = 20
        sheet2.column_dimensions['S'].width = 20
        sheet2.column_dimensions['T'].width = 20
        sheet2.column_dimensions['U'].width = 20
        sheet2.column_dimensions['V'].width = 20
        sheet2.column_dimensions['W'].width = 20
        sheet2.column_dimensions['X'].width = 20
        sheet2.column_dimensions['Y'].width = 20
        sheet2.column_dimensions['Z'].width = 20
    
        title_cell = sheet2['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de las Operaciones Estadísticas'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')
	
        codigo_cell = sheet2['A2']
        codigo_cell.value = 'A. OPERACIÓN ESTADÍSTICA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet2['C2']
        codigo_cell.value = '¿Otra entidad es responsable de una o varias fases de la operación estadística?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet2['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = sheet2['B3']
        codigo_cell.value = 'Nombre Operación Estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet2['C3']
        codigo_cell.value = 'Indique si hay otra(s) entidad(es) que participa(n) en alguna de las fases del proceso estadístico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        contSheet2 = 4
        for ooee in ooees: 
            sheet2.cell(row = contSheet2, column = 1).value = ooee.codigo_oe
            sheet2.cell(row = contSheet2, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet2.cell(row = contSheet2, column = 1).font = Font(size = "8", name='Barlow')

            sheet2.cell(row = contSheet2, column = 2).value = ooee.nombre_oe
            sheet2.cell(row = contSheet2, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet2.cell(row = contSheet2, column = 2).font = Font(size = "8", name='Barlow')
            
            if str(ooee.fase) == "True":
                sheet2.cell(row = contSheet2, column = 3).value = "Si"
                sheet2.cell(row = contSheet2, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet2.cell(row = contSheet2, column = 3).font = Font(size = "8", name='Barlow')
            elif str(ooee.fase) == "False":
                sheet2.cell(row = contSheet2, column = 3).value = "No"
                sheet2.cell(row = contSheet2, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet2.cell(row = contSheet2, column = 3).font = Font(size = "8", name='Barlow')


            contSheet2+= 1
            
        entidadoeid = []
        for indexEnFas, itemEnfas in enumerate(entidadesFasesTextList):
            indexEnFas = itemEnfas.ooee_id + 3
                    
            entidadoeid.append(itemEnfas.ooee_id)
            #print("array",set(entidadoeid))
            #print("sinset", entidadoeid )
            count_id = entidadoeid.count(itemEnfas.ooee_id)
            c=0 #inicializamos el contador  
            n=6*count_id
            for i in range(1,n+1):  
                if i%6 == 0:  
                    i = i - 2
                    c+=1

            for indexoe, itemoe in enumerate(ooees):
                if itemEnfas.ooee_id == itemoe.pk:
                    posRow = indexoe + 4
                     
                    sheet2.cell(row = 3, column = i ).value = "Nombre de Entidad"
                    sheet2.cell(row = 3, column = i ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = 3, column = i ).font = Font(bold=True) 

                    sheet2.cell(row = posRow, column = i ).value = str(itemEnfas.nombre_entifas) 
                    sheet2.cell(row = posRow, column = i ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = posRow, column = i ).font = Font(size = "8", name='Barlow') 
            
            listaFases = list(itemEnfas.fases.all())  ##iterar para traer las fases seleccionadas

            for indexFase, itemFase in enumerate(listaFases):

                if str(itemFase) == "Detección y análisis de requerimientos":
                    
                    countFase=0 #inicializamos el contador  
                    faseSelected=6*count_id 
                    for posFase in range(1,n+1):  
                        if posFase%6 == 0:
                            posFase = posFase - 1
                            countFase+=1

                    sheet2.cell(row = 3, column = posFase ).value = "a. Detección y análisis de requerimientos"
                    sheet2.cell(row = 3, column = posFase ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = 3, column = posFase ).font = Font(bold=True)

                    sheet2.cell(row = posRow, column = posFase).value = str(itemFase)
                    sheet2.cell(row = posRow, column = posFase).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = posRow, column = posFase).font = Font(size = "8", name='Barlow')
                                
                if str(itemFase) == "Diseño y pruebas":

                    countFase=0 #inicializamos el contador  
                    faseSelected=6*count_id 
                    for posFase in range(1,n+1):  
                        if posFase%6 == 0:  
                            posFase = posFase
                            countFase+=1

                    sheet2.cell(row = 3, column = posFase ).value = "b. Diseño y pruebas"
                    sheet2.cell(row = 3, column = posFase ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = 3, column = posFase ).font = Font(bold=True)

                    sheet2.cell(row = posRow, column = posFase).value = str(itemFase)
                    sheet2.cell(row = posRow, column = posFase).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = posRow, column = posFase).font = Font(size = "8", name='Barlow')

                if str(itemFase) == "Ejecución":

                    countFase=0 #inicializamos el contador  
                    faseSelected=6*count_id 
                    for posFase in range(1,n+1):  
                        if posFase%6 == 0:  
                            posFase = posFase + 1
                            countFase+=1

                    sheet2.cell(row = 3, column = posFase ).value = "c. Ejecución"
                    sheet2.cell(row = 3, column = posFase ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = 3, column = posFase ).font = Font(bold=True)

                    sheet2.cell(row = posRow, column = posFase).value = str(itemFase)
                    sheet2.cell(row = posRow, column = posFase).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = posRow, column = posFase).font = Font(size = "8", name='Barlow')
                
                if str(itemFase) == "Análisis":

                    countFase=0 #inicializamos el contador  
                    faseSelected=6*count_id 
                    for posFase in range(1,n+1):  
                        if posFase%6 == 0:  
                            posFase = posFase + 2
                            countFase+=1

                    sheet2.cell(row = 3, column = posFase ).value = "d. Análisis"
                    sheet2.cell(row = 3, column = posFase ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = 3, column = posFase ).font = Font(bold=True)

                    sheet2.cell(row = posRow, column = posFase).value = str(itemFase)
                    sheet2.cell(row = posRow, column = posFase).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = posRow, column = posFase).font = Font(size = "8", name='Barlow')
                
                if str(itemFase) == "Difusión":

                    countFase=0 #inicializamos el contador  
                    faseSelected=6*count_id 
                    for posFase in range(1,n+1):  
                        if posFase%6 == 0:  
                            posFase = posFase + 3
                            countFase+=1

                    sheet2.cell(row = 3, column = posFase ).value = "e. Difusión"
                    sheet2.cell(row = 3, column = posFase ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = 3, column = posFase ).font = Font(bold=True)
                    
                    sheet2.cell(row = posRow, column = posFase).value = str(itemFase)
                    sheet2.cell(row = posRow, column = posFase).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = posRow, column = posFase).font = Font(size = "8", name='Barlow')
                               
        for row in range(4, sheet2.max_row + 1):
            sheet2.row_dimensions[row].height = 90   
        
                    
        #############  Hoja 3 lista de variables  ##################

        
        sheet3.merge_cells('A1:C1')
        sheet3.merge_cells('A2:C2')

        def set_border(sheet3, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet3[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet3,'A1:C1')
    
        #row dimensions
        sheet3.row_dimensions[1].height = 55
        sheet3.row_dimensions[2].height = 40

        # column width
        sheet3.column_dimensions['A'].width = 20
        sheet3.column_dimensions['B'].width = 20
        sheet3.column_dimensions['C'].width = 20
        
        title_cell = sheet3['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de las Operaciones Estadísticas'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        title_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
	
        codigo_cell = sheet3['A2']
        codigo_cell.value = 'Lista de variables de la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)


        codigo_cell = sheet3['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = sheet3['B3']
        codigo_cell.value = 'Nombre Operación Estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet3['C3']
        codigo_cell.value = 'Variables que maneja la operación estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        contsheet3 = 4

        arrayoeId = []
        for indexVariab, itemVariab in enumerate(listaDeVariablesText):
            indexVariab =  itemVariab.ooee_id + 3
            arrayoeId.append(itemVariab.ooee_id)
            for oest in ooees:
                if  str(itemVariab.ooee) == str(oest.nombre_oe):
            
                    sheet3.cell(row = contsheet3, column = 1).value = oest.codigo_oe
                    sheet3.cell(row = contsheet3, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet3.cell(row = contsheet3, column = 1).font = Font(size = "8", name='Barlow')

                    sheet3.cell(row = contsheet3, column = 2).value = str(itemVariab.ooee)
                    sheet3.cell(row = contsheet3, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet3.cell(row = contsheet3, column = 2).font = Font(size = "8", name='Barlow')

                    sheet3.cell(row = contsheet3, column = 3).value = str(itemVariab.lista_var)
                    sheet3.cell(row = contsheet3, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet3.cell(row = contsheet3, column = 3).font = Font(size = "8", name='Barlow')
                    contsheet3+=1
           
        for row in range(4, sheet3.max_row + 1):  ## definir tamaño de rows
            sheet3.row_dimensions[row].height = 90

         #############  end Hoja 3 variables  ##################

        #############  Hoja 4 resultados  ##################

        sheet4.merge_cells('A1:C1')
        sheet4.merge_cells('A2:C2')

        def set_border(sheet4, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet4[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet4,'A1:C1')
    
        #row dimensions
        sheet4.row_dimensions[1].height = 55
        sheet4.row_dimensions[2].height = 40
        sheet4.row_dimensions[3].height = 40

        # column width
        sheet4.column_dimensions['A'].width = 20
        sheet4.column_dimensions['B'].width = 20
        sheet4.column_dimensions['C'].width = 20
    
        title_cell = sheet4['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de las Operaciones Estadísticas'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
	
        codigo_cell = sheet4['A2']
        codigo_cell.value = 'Resultados agregados o indicadores calculados'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet4['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = sheet4['B3']
        codigo_cell.value = 'Nombre Operación Estadística'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet4['C3']
        codigo_cell.value = 'Lista de resultados agregados o indicadores calculados'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        contsheet4 = 4
        listaIdoe = []
        for indexResultEsta, itemResultEsta in enumerate(listaResultEstText):
            indexResultEsta =  itemResultEsta.ooee_id + 3
            listaIdoe.append(itemResultEsta.ooee_id)
            
            for oest in ooees:
                if  str(itemResultEsta.ooee) == str(oest.nombre_oe):
                    sheet4.cell(row = contsheet4, column = 1).value = oest.codigo_oe
                    sheet4.cell(row = contsheet4, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = contsheet4, column = 1).font = Font(size = "8", name='Barlow')

                    sheet4.cell(row = contsheet4, column = 2).value = str(itemResultEsta.ooee)
                    sheet4.cell(row = contsheet4, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = contsheet4, column = 2).font = Font(size = "8", name='Barlow')

                    sheet4.cell(row = contsheet4, column = 3).value = str(itemResultEsta.resultEstad)
                    sheet4.cell(row = contsheet4, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = contsheet4, column = 3).font = Font(size = "8", name='Barlow')
                    contsheet4+=1

        for row in range(4, sheet4.max_row + 1):  ## definir tamaño de rows
            sheet4.row_dimensions[row].height = 90

        ############# END Hoja 4 resultados  ##################

        ############## Hoja 5 Evaluación de calidad #############

        
        sheet5.merge_cells('A1:AZ1')
        sheet5.merge_cells('A2:F2')
        sheet5.merge_cells('G2:AZ2')

        def set_border(sheet5, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet5[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet5,'A1:AZ'+str(ooees.count()+3))
        

        #row dimensions
        sheet5.row_dimensions[1].height = 55
        sheet5.row_dimensions[2].height = 40
        sheet5.row_dimensions[3].height = 40

        # column width
        sheet5.column_dimensions['A'].width = 20
        sheet5.column_dimensions['B'].width = 18
        sheet5.column_dimensions['C'].width = 15
        sheet5.column_dimensions['D'].width = 24
        sheet5.column_dimensions['E'].width = 15
        sheet5.column_dimensions['F'].width = 24
        sheet5.column_dimensions['G'].width = 22
        sheet5.column_dimensions['H'].width = 15
        sheet5.column_dimensions['I'].width = 15
        sheet5.column_dimensions['J'].width = 15
        sheet5.column_dimensions['K'].width = 15
        sheet5.column_dimensions['L'].width = 15
        sheet5.column_dimensions['M'].width = 15
        sheet5.column_dimensions['N'].width = 15
        sheet5.column_dimensions['O'].width = 15
        sheet5.column_dimensions['P'].width = 15
        sheet5.column_dimensions['Q'].width = 15
        sheet5.column_dimensions['R'].width = 22
        sheet5.column_dimensions['S'].width = 15
        sheet5.column_dimensions['T'].width = 15
        sheet5.column_dimensions['U'].width = 15
        sheet5.column_dimensions['V'].width = 15
        sheet5.column_dimensions['W'].width = 15
        sheet5.column_dimensions['X'].width = 15
        sheet5.column_dimensions['Y'].width = 20
        sheet5.column_dimensions['Z'].width = 15

        sheet5.column_dimensions['AA'].width = 15
        sheet5.column_dimensions['AB'].width = 15
        sheet5.column_dimensions['AC'].width = 15
        sheet5.column_dimensions['AD'].width = 15
        sheet5.column_dimensions['AE'].width = 15
        sheet5.column_dimensions['AF'].width = 15
        sheet5.column_dimensions['AG'].width = 15
        sheet5.column_dimensions['AH'].width = 15
        sheet5.column_dimensions['AI'].width = 15
        sheet5.column_dimensions['AJ'].width = 15
        sheet5.column_dimensions['AK'].width = 15
        sheet5.column_dimensions['AL'].width = 15
        sheet5.column_dimensions['AM'].width = 15
        sheet5.column_dimensions['AN'].width = 15
        sheet5.column_dimensions['AO'].width = 15
        sheet5.column_dimensions['AP'].width = 15
        sheet5.column_dimensions['AQ'].width = 15
        sheet5.column_dimensions['AR'].width = 15
        sheet5.column_dimensions['AS'].width = 15
        sheet5.column_dimensions['AT'].width = 15
        sheet5.column_dimensions['AU'].width = 15
        sheet5.column_dimensions['AV'].width = 15
        sheet5.column_dimensions['AW'].width = 15
        sheet5.column_dimensions['AX'].width = 15
        sheet5.column_dimensions['AY'].width = 15
        sheet5.column_dimensions['AZ'].width = 15

        title_cell = sheet5['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de las Operaciones Estadísticas'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet5['A2']
        codigo_cell.value = 'IDENTIFICACIÓN'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet5['G2']
        codigo_cell.value = 'EVALUACIÓN DE CALIDAD'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['A3']
        resultados_cell.value = 'Área Temática'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['B3']
        resultados_cell.value = 'Tema'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['C3']
        resultados_cell.value = 'Código Entidad'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['D3']
        resultados_cell.value = 'Nombre Entidad'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['E3']
        resultados_cell.value = 'Código OOEE'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['F3']
        resultados_cell.value = 'Nombre de la Operación Estadística'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        resultados_cell = sheet5['G3']
        resultados_cell.value = 'Número de Evaluaciones'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        contsheet5 = 4

        for ooee in ooees:

            sheet5.cell(row = contsheet5, column = 1).value = str(ooee.area_tematica)
            sheet5.cell(row = contsheet5, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 1).font = Font(size = "8", name='Barlow') 
            
            sheet5.cell(row = contsheet5, column = 2).value = str(ooee.tema)
            sheet5.cell(row = contsheet5, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 2).font = Font(size = "8", name='Barlow') 
            
            sheet5.cell(row = contsheet5, column = 3).value = ooee.entidad.codigo
            sheet5.cell(row = contsheet5, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 3).font = Font(size = "8", name='Barlow')

            sheet5.cell(row = contsheet5, column = 4).value = str(ooee.entidad)
            sheet5.cell(row = contsheet5, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 4).font = Font(size = "8", name='Barlow')

            sheet5.cell(row = contsheet5, column = 5).value = ooee.codigo_oe
            sheet5.cell(row = contsheet5, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 5).font = Font(size = "8", name='Barlow')

            sheet5.cell(row = contsheet5, column = 6).value = ooee.nombre_oe
            sheet5.cell(row = contsheet5, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 6).font = Font(size = "8", name='Barlow') 
            contsheet5+=1

        idoeEval = []
        for indexEval, itemEval in enumerate(listaEvaluacionText):
            indexEval =  itemEval.post_oe_id + 3
            idoeEval.append(itemEval.post_oe_id)

            ### 1 Estado de la evaluación
            
            count_est_cert = idoeEval.count(itemEval.post_oe_id)
            #print("total evaluaciones",count_est_cert)
            contrField1=0 #inicializamos el contador  
            pos_est_cert = 10*count_est_cert
            for incField1 in range(1,pos_est_cert+1):  
                if incField1%10 == 0:  
                    incField1 = incField1 - 2
                    contrField1+=1

            for indexoe, itemoe in enumerate(ooees):
                if itemEval.post_oe_id == itemoe.pk:
                    posiRow = indexoe + 4

                    sheet5.cell(row = posiRow, column = 7).value = count_est_cert ## número de evaluaciones
                    sheet5.cell(row = posiRow, column = 7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = 7).font = Font(bold=True) 

                    sheet5.cell(row = 3, column = incField1).value = "Estado de la Evaluación"
                    sheet5.cell(row = 3, column = incField1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField1).font = Font(bold=True) 

                    sheet5.cell(row = posiRow, column = incField1).value = str(itemEval.est_evaluacion)
                    sheet5.cell(row = posiRow, column = incField1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField1).font = Font(size = "8", name='Barlow')
  
            ### 2 Resultado año de vigencia


            count_id_oe = idoeEval.count(itemEval.post_oe_id)
            contadorEval=0 #inicializamos el contador  
            posFields = 10*count_id_oe
            #print("______tttt________", posFields)
            for incEval in range(1,posFields+1):  
                if incEval%10 == 0:
                    #print("inceval ",incEval) 
                    incEval = incEval - 1
                    #print("sumado  ",incEval) 
                    contadorEval+=1

            for indexoe, itemoe in enumerate(ooees):
                if itemEval.post_oe_id == itemoe.pk:
                    posiRow = indexoe + 4

                    sheet5.cell(row = 3, column = incEval).value = "Año"
                    sheet5.cell(row = 3, column = incEval ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incEval ).font = Font(bold=True) 

                    if itemEval.year_eva == None:

                        sheet5.cell(row = posiRow, column = incEval).value = ""
                        sheet5.cell(row = posiRow, column = incEval).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incEval).font = Font(size = "8", name='Barlow')
                    else:
                        sheet5.cell(row = posiRow, column = incEval).value = itemEval.year_eva.strftime('%Y')
                        sheet5.cell(row = posiRow, column = incEval).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incEval).font = Font(size = "8", name='Barlow')
            
            ### 3 Observaciones de la evaluación

            contadorField3=0 #inicializamos el contador  
            posFields3 = 10*count_id_oe
            for incField3 in range(1,posFields3+1):  
                if incField3%10 == 0:  
                    incField3 = incField3
                    contadorField3+=1
                
            for indexoe, itemoe in enumerate(ooees):
                if itemEval.post_oe_id == itemoe.pk:
                    posiRow = indexoe + 4

                    sheet5.cell(row = 3, column = incField3).value = "Observaciones"
                    sheet5.cell(row = 3, column = incField3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField3).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField3).value = str(itemEval.observ_est)
                    sheet5.cell(row = posiRow, column = incField3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField3).font = Font(size = "8", name='Barlow')

            ### 4 Metodologia

            contadorField4=0 #inicializamos el contador  
            posFields4 = 10*count_id_oe
            for incField4 in range(1,posFields4+1):  
                if incField4%10 == 0:
                    incField4 = incField4 + 1
                    contadorField4+=1

            for indexoe, itemoe in enumerate(ooees):
                if itemEval.post_oe_id == itemoe.pk:
                    posiRow = indexoe + 4

                    sheet5.cell(row = 3, column = incField4).value = "Metodología"
                    sheet5.cell(row = 3, column = incField4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField4).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField4).value = str(itemEval.metodologia)
                    sheet5.cell(row = posiRow, column = incField4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField4).font = Font(size = "8", name='Barlow')
            
        ### 5 Resultado de la evaluación según metodologia

            contadorField5=0 #inicializamos el contador  
            posFields5 = 10*count_id_oe
            for incField5 in range(1,posFields5+1):  
                if incField5%10 == 0:  
                    incField5 = incField5 + 2
                    contadorField5+=1

            for indexoe, itemoe in enumerate(ooees):
                if itemEval.post_oe_id == itemoe.pk:
                    posiRow = indexoe + 4

                    sheet5.cell(row = 3, column = incField5).value = "Resultado de la evaluación"
                    sheet5.cell(row = 3, column = incField5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField5).font = Font(bold=True)

                    if str(itemEval.metodologia) == "ntcpe 1000":
                        sheet5.cell(row = posiRow, column = incField5).value = str(itemEval.res_evaluacion)
                        sheet5.cell(row = posiRow, column = incField5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField5).font = Font(size = "8", name='Barlow')
                    
                    elif str(itemEval.metodologia) == "matriz de requisitos":
                        sheet5.cell(row = posiRow, column = incField5).value = str(itemEval.res_mzrequi)
                        sheet5.cell(row = posiRow, column = incField5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField5).font = Font(size = "8", name='Barlow')
                    
        ### 6 Observaciones de resultado de la evaluación

            contadorField6=0 #inicializamos el contador  
            posFields6 = 10*count_id_oe
            for incField6 in range(1,posFields6+1):  
                if incField6%10 == 0:  
                    incField6 = incField6 + 3
                    contadorField6+=1

            for indexoe, itemoe in enumerate(ooees):
                if itemEval.post_oe_id == itemoe.pk:
                    posiRow = indexoe + 4

                    sheet5.cell(row = 3, column = incField6).value = "Observaciones"
                    sheet5.cell(row = 3, column = incField6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField6).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField6).value = str(itemEval.observ_resul)
                    sheet5.cell(row = posiRow, column = incField6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField6).font = Font(size = "8", name='Barlow')

        ### 7 Vigencia

            contadorField7=0 #inicializamos el contador  
            posFields7 = 10*count_id_oe
            for incField7 in range(1,posFields7+1):  
                if incField7%10 == 0:  
                    incField7 = incField7 + 4
                    contadorField7+=1

            for indexoe, itemoe in enumerate(ooees):
                if itemEval.post_oe_id == itemoe.pk:
                    posiRow = indexoe + 4

                    sheet5.cell(row = 3, column = incField7).value = "Vigencia"
                    sheet5.cell(row = 3, column = incField7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField7).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField7).value = "Desde: " + str(itemEval.vigencia_desde) + " " + "\n\n Hasta: " + str(itemEval.vigencia_hasta)
                    sheet5.cell(row = posiRow, column = incField7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField7).font = Font(size = "8", name='Barlow')

         ### 8 Plan de Mejoramiento   
            
            contadorField8=0 #inicializamos el contador  
            posFields8 = 10*count_id_oe
            for incField8 in range(1,posFields8+1):  
                if incField8%10 == 0:  
                    incField8 = incField8 + 5
                    contadorField8+=1

            for indexoe, itemoe in enumerate(ooees):
                if itemEval.post_oe_id == itemoe.pk:
                    posiRow = indexoe + 4

                    sheet5.cell(row = 3, column = incField8).value = "Plan de mejoramiento"
                    sheet5.cell(row = 3, column = incField8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField8).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField8).value = str(itemEval.pla_mejoramiento)
                    sheet5.cell(row = posiRow, column = incField8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField8).font = Font(size = "8", name='Barlow')

        ###  9 Seguimiento anual (Vigilancia):   
            
            contadorField9=0 #inicializamos el contador  
            posFields9 = 10*count_id_oe
            for incField9 in range(1,posFields9+1):  
                if incField9%10 == 0:  
                    incField9 = incField9 + 6
                    contadorField9+=1

            for indexoe, itemoe in enumerate(ooees):
                if itemEval.post_oe_id == itemoe.pk:
                    posiRow = indexoe + 4

                    sheet5.cell(row = 3, column = incField9).value = "Seguimiento anual (Vigilancia)"
                    sheet5.cell(row = 3, column = incField9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField9).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField9).value = str(itemEval.seg_vig)
                    sheet5.cell(row = posiRow, column = incField9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField9).font = Font(size = "8", name='Barlow')

        ###  Observaciones seguimiento anual
            
            contadorField10=0 #inicializamos el contador  
            posFields10 = 10*count_id_oe
            for incField10 in range(1,posFields10+1):  
                if incField10%10 == 0:  
                    incField10 = incField10 + 7
                    contadorField10+=1

            for indexoe, itemoe in enumerate(ooees):
                if itemEval.post_oe_id == itemoe.pk:
                    posiRow = indexoe + 4
                
                    sheet5.cell(row = 3, column = incField10).value = "Observaciones"
                    sheet5.cell(row = 3, column = incField10).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField10).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField10).value = str(itemEval.obs_seg_anual)
                    sheet5.cell(row = posiRow, column = incField10).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField10).font = Font(size = "8", name='Barlow')
   
        for row in range(4, sheet5.max_row + 1):  ## definir tamaño de rows
            sheet5.row_dimensions[row].height = 90

        ############## End Evaluación de calidad ############# 

        ## ocultar información si no esta autenticado
        if user.is_authenticated == True and user.profile.role.id != 2:
        ############## Hoja 6 Critica #############

        
            sheet6.merge_cells('A1:F1')
            sheet6.merge_cells('A2:F2')

            def set_border(sheet6, cell_range):
                border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                rows = sheet6[cell_range]
                for row in rows:
                    for cell in row:
                        cell.border = border

            set_border(sheet6,'A1:F1')
            

            #row dimensions
            sheet6.row_dimensions[1].height = 55
            sheet6.row_dimensions[2].height = 40
            sheet6.row_dimensions[3].height = 40
            sheet6.row_dimensions[4].height = 40
            sheet6.row_dimensions[5].height = 40
            sheet6.row_dimensions[6].height = 40

            # column width
            sheet6.column_dimensions['A'].width = 20
            sheet6.column_dimensions['B'].width = 20
            sheet6.column_dimensions['C'].width = 20
            sheet6.column_dimensions['D'].width = 20
            sheet6.column_dimensions['E'].width = 20
            sheet6.column_dimensions['F'].width = 20

            title_cell = sheet6['A1']
            title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
            title_cell.value = 'Directorio de las Operaciones Estadísticas'
            title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
            title_cell.alignment =  Alignment(horizontal='center', vertical='center')

            codigo_cell = sheet6['A2']
            codigo_cell.value = 'Críticas'
            codigo_cell.font = Font(bold=True)
            codigo_cell.alignment = Alignment(horizontal='center')
            codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')


            resultados_cell = sheet6['A3']
            resultados_cell.value = 'Código'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet6['B3']
            resultados_cell.value = 'Nombre de la Operación Estadística'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet6['C3']
            resultados_cell.value = 'Estado de la crítica'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet6['D3']
            resultados_cell.value = 'Observaciones de la crítica'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet6['E3']
            resultados_cell.value = 'Funcionario que realiza la crítica'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet6['F3']
            resultados_cell.value = 'Fecha en que se realiza la crítica'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            idoeCritica = []
            contsheet6 = 4
            for indexCritica, itemCritica in enumerate(listaCriticaText):
                indexCritica =  itemCritica.post_oe_id + 3
                idoeCritica.append(itemCritica.post_oe_id)
                for oest in ooees:
                    if  str(itemCritica.post_oe) == str(oest.nombre_oe):

                        sheet6.cell(row = contsheet6, column = 1).value = oest.codigo_oe
                        sheet6.cell(row = contsheet6, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet6.cell(row = contsheet6, column = 1).font = Font(bold=True) 

                        sheet6.cell(row = contsheet6, column = 2).value = str(itemCritica.post_oe)
                        sheet6.cell(row = contsheet6, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet6.cell(row = contsheet6, column = 2).font = Font(bold=True) 

                        sheet6.cell(row = contsheet6, column = 3).value = str(itemCritica.estado_crit)
                        sheet6.cell(row = contsheet6, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet6.cell(row = contsheet6, column = 3).font = Font(bold=True)

                        sheet6.cell(row = contsheet6, column = 4).value = str(itemCritica.descrip_critica)
                        sheet6.cell(row = contsheet6, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet6.cell(row = contsheet6, column = 4).font = Font(bold=True)

                        sheet6.cell(row = contsheet6, column = 5).value = str(itemCritica.name_cri)
                        sheet6.cell(row = contsheet6, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet6.cell(row = contsheet6, column = 5).font = Font(bold=True)

                        sheet6.cell(row = contsheet6, column = 6).value = str(itemCritica.fecha_critica)
                        sheet6.cell(row = contsheet6, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet6.cell(row = contsheet6, column = 6).font = Font(bold=True)

                        contsheet6+=1

            for row in range(4, sheet6.max_row + 1):  ## definir tamaño de rows
                sheet6.row_dimensions[row].height = 90

            
        ############## End critica ############# 

         ############## Hoja 7 Novedad #############
        
        
            sheet7.merge_cells('A1:G1')
            sheet7.merge_cells('A2:G2')

            def set_border(sheet7, cell_range):
                border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                rows = sheet7[cell_range]
                for row in rows:
                    for cell in row:
                        cell.border = border

            set_border(sheet7,'A1:G1')
            

            #row dimensions
            sheet7.row_dimensions[1].height = 55
            sheet7.row_dimensions[2].height = 40
            sheet7.row_dimensions[3].height = 40
            sheet7.row_dimensions[4].height = 40
            sheet7.row_dimensions[5].height = 40
            sheet7.row_dimensions[6].height = 40
            sheet7.row_dimensions[7].height = 40

            # column width
            sheet7.column_dimensions['A'].width = 20
            sheet7.column_dimensions['B'].width = 20
            sheet7.column_dimensions['C'].width = 20
            sheet7.column_dimensions['D'].width = 20
            sheet7.column_dimensions['E'].width = 20
            sheet7.column_dimensions['F'].width = 20
            sheet7.column_dimensions['G'].width = 20

            title_cell = sheet7['A1']
            title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
            title_cell.value = 'Directorio de las Operaciones Estadísticas'
            title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
            title_cell.alignment =  Alignment(horizontal='center', vertical='center')

            codigo_cell = sheet7['A2']
            codigo_cell.value = 'Novedades'
            codigo_cell.font = Font(bold=True)
            codigo_cell.alignment = Alignment(horizontal='center')
            codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')


            resultados_cell = sheet7['A3']
            resultados_cell.value = 'Código'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet7['B3']
            resultados_cell.value = 'Nombre de la Operación Estadística'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet7['C3']
            resultados_cell.value = 'Novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet7['D3']
            resultados_cell.value = 'Estado de la actualización'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet7['E3']
            resultados_cell.value = 'Descripción de la novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet7['F3']
            resultados_cell.value = 'Funcionario que realiza la novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet7['G3']
            resultados_cell.value = 'Fecha en que se realiza la novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            idoeNovedad = []
            contsheet7 = 4
            for indexNovedad, itemNovedad in enumerate(listaNovedadText):
                indexNovedad =  itemNovedad.post_oe_id + 3
                idoeNovedad.append(itemNovedad.post_oe_id)
                for oest in ooees:
                    if  str(itemNovedad.post_oe) == str(oest.nombre_oe):

                        sheet7.cell(row = contsheet7, column = 1).value = oest.codigo_oe
                        sheet7.cell(row = contsheet7, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet7.cell(row = contsheet7, column = 1).font = Font(bold=True) 

                        sheet7.cell(row = contsheet7, column = 2).value = str(itemNovedad.post_oe)
                        sheet7.cell(row = contsheet7, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet7.cell(row = contsheet7, column = 2).font = Font(bold=True) 

                        sheet7.cell(row = contsheet7, column = 3).value = str(itemNovedad.novedad)
                        sheet7.cell(row = contsheet7, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet7.cell(row = contsheet7, column = 3).font = Font(bold=True)

                        sheet7.cell(row = contsheet7, column = 4).value = str(itemNovedad.est_actualiz)
                        sheet7.cell(row = contsheet7, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet7.cell(row = contsheet7, column = 4).font = Font(bold=True)

                        sheet7.cell(row = contsheet7, column = 5).value = str(itemNovedad.descrip_novedad)
                        sheet7.cell(row = contsheet7, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet7.cell(row = contsheet7, column = 5).font = Font(bold=True)

                        sheet7.cell(row = contsheet7, column = 6).value = str(itemNovedad.name_nov)
                        sheet7.cell(row = contsheet7, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet7.cell(row = contsheet7, column = 6).font = Font(bold=True)

                        sheet7.cell(row = contsheet7, column = 7).value = str(itemNovedad.fecha_actualiz)
                        sheet7.cell(row = contsheet7, column = 7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet7.cell(row = contsheet7, column = 7).font = Font(bold=True)

                        contsheet7+=1

            for row in range(4, sheet7.max_row + 1):  ## definir tamaño de rows
                sheet7.row_dimensions[row].height = 90

            
        ############## End Novedad #############
       

        file_name = "reporte_ooee.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(file_name)
        response['Content-Disposition'] = content
        response['Set-Cookie'] = 'fileDownload=true; Path=/'
        wb.save(response)
        return response 



############ reporte complemenatrio para administrador ######################

@method_decorator(login_required, name='dispatch')
class reporteUltimaNovedadOE(View):
    def get(self, request):
        
        # Get some data to write to the spreadsheet.
        operaciones_est = OperacionEstadistica.objects.all()
        ##NovedadActualizacionRRAA.filter(post_ra=1).order_by('-id')[:1] 
        # Create an in-memory output file for the new workbook.
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet(name="Ultima novedad oooe")
        
        # Get some data to write to the spreadsheet.
        row = 4
        col = 0
        ## altura de celda
        worksheet.set_row(0, 40)
        worksheet.set_row(1, 30)
        worksheet.set_row(2, 30)
        worksheet.set_row(3, 40)
        worksheet.set_row(4, 30)
        worksheet.set_row(5, 30)
        worksheet.set_row(6, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet.set_column('A:A', 30)
        worksheet.set_column('B:B', 30)
        worksheet.set_column('C:C', 30)
        worksheet.set_column('D:D', 30)
        worksheet.set_column('E:E', 30)
        worksheet.set_column('F:F', 30)
        worksheet.set_column('G:G', 30)
        
        
        ## agregar estilos
        Backgroundcolor = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 18,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        Backgroundcolor2 = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 14,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        Backgroundcolor3 = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 12,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        cell_format = workbook.add_format()
        cell_format.set_font_color('white')

        textInfo_format = workbook.add_format({'text_wrap': True, 'align': 'center', 'valign': 'vcenter'})
        

        formatfecha = workbook.add_format({'num_format':'yyyy-mm-dd', 'align': 'center', 'valign': 'vcenter'})

        ##FILA 1
        worksheet.merge_range('A1:G1', "Inventario de Operaciones Estadísticas", Backgroundcolor)
        worksheet.conditional_format('A1:G1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Operaciones Estadísticas',
                                                'format': cell_format})

        ##FILA 2

        worksheet.merge_range('A2:G3', "ÚLTIMAS NOVEDADES", Backgroundcolor2)
        worksheet.conditional_format('A2:G3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'ÚLTIMAS NOVEDADES',
                                                'format': cell_format })

        worksheet.write(3, 0, "Código", Backgroundcolor3)
        worksheet.write(3, 1, "Nombre de la Operación Estadística", Backgroundcolor3)
        worksheet.write(3, 2, "Fase en el sistema", Backgroundcolor3)
        worksheet.write(3, 3, "Área temática", Backgroundcolor3)
        worksheet.write(3, 4, "Tema", Backgroundcolor3)
        worksheet.write(3, 5, "Estado de la novedad", Backgroundcolor3)
        worksheet.write(3, 6, "Fecha en que se realiza la novedad", Backgroundcolor3)
         
        for item in operaciones_est:
            item.codigo_oe

            for oper in NovedadActualizacion.objects.filter(post_oe__codigo_oe=item.codigo_oe).order_by('-id')[:1]:

                worksheet.write(row, col, oper.post_oe.codigo_oe, textInfo_format)
                worksheet.write(row, col + 1, oper.post_oe.nombre_oe, textInfo_format)
                worksheet.write(row, col + 2, str(oper.post_oe.nombre_est), textInfo_format)
                worksheet.write(row, col + 3, str(oper.post_oe.area_tematica), textInfo_format)
                worksheet.write(row, col + 4, str(oper.post_oe.tema), textInfo_format)
                worksheet.write(row, col + 5, str(oper.est_actualiz), textInfo_format)
                worksheet.write(row, col + 6, str(oper.fecha_actualiz), formatfecha)
                row += 1

        # Close the workbook before sending the data.
        workbook.close()

        # Rewind the buffer.
        output.seek(0)

        filename = 'operaciones_estadisticas.xlsx'
        # Set up the Http response.
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        return response

