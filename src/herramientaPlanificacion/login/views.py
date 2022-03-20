#-- encoding:utf-8 --
from django.shortcuts import render, redirect, get_object_or_404
from django.http import Http404
from django.contrib import messages, auth
from django.contrib.auth.models import User
from django.http import HttpResponseRedirect, HttpResponse
from django.template.loader import get_template
from django.urls import reverse_lazy
from django.template.context_processors import csrf
from django.contrib.auth.decorators import login_required
import json
from django.http import HttpResponse
from django.core import serializers
from django.views.generic import TemplateView, ListView, FormView, CreateView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from login.forms import LoginForm
from django.contrib.auth import logout
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .forms import FormEditUser, EditProfileForm, UserForm, UserProfileForm
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import UserCreationForm
from django.core.mail import send_mail
import xlwt
from .models import Role, User, Profile, Entidades_oe
from ooee.models import OperacionEstadistica, NovedadActualizacion
from rraa.models import RegistroAdministrativo, NovedadActualizacionRRAA
from demandas.models import demandaInfor, NovedadActualizacionddi
from django.utils.crypto import get_random_string
from django.db.models import Q

## images report


##  report by excel ##
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Font, Fill, Border, Side,  GradientFill, Alignment, Color, PatternFill
from django.views.generic import TemplateView



#import librarys by create user

from django.contrib.sites.shortcuts import get_current_site
from django.template.loader import render_to_string
#password = get_random_string(length=8)


# Create your views here.


def login(request):

	message = None
	if request.method == "POST":
		form = LoginForm(request.POST)
		if not request.POST.get('remember_me', None):
			request.session.set_expiry(0)
		if form.is_valid():
			username = request.POST['username']
			password = request.POST['password']
			user = auth.authenticate(username=username, password=password)
			if user is not None:
				if user.is_active:
					auth.login(request, user)
					users = User.objects.filter()
					return HttpResponseRedirect('loggedin')
                    
				else:
					messages.warning(request, "tu usuario esta inactivo")                 
			else:
				messages.warning(request, "nombre de usuario y/o password incorrectos")
		else:
			messages.info(request, "los campos son obligatorios") 
	else: 
		form = LoginForm()
	return render(request, 'login/login.html',{'message': message, 'form': form} ) 


## function send email by user created
def createEmailNewUser(request, pk, passw):
	user = get_object_or_404(User, pk=pk)
	subject = "SISTEMA DE IDENTIFICACIÓN Y CARACTERIZACIÓN DE OFERTA Y DEMANDA ESTADÍSTICA DEL SEN: Creación de usuario"
	recipients = []
	recipients.append(user.email)
	body = 'SISTEMA DE IDENTIFICACIÓN Y CARACTERIZACIÓN DE OFERTA Y DEMANDA ESTADÍSTICA DEL SEN' \
		'\n\n Bienvenido   '+ user.username + \
			'\n\n Ahora tiene acceso al Sistema de Información SEN: (https://inventariosen.dane.gov.co/login/inicioSesion) '\
			 '\n\n Se le ha asignado el Rol de: '+ user.profile.role.name + \
				 '\n\n para acceder al aplicativo debe ingresar las siguientes credenciales:' \
					'\n Username: ' + user.username + \
						'\n Contraseña: '+ passw + \
							'\n\n Al ingresar podra crear y actualizar las operaciones estadísticas y registros administrativos que la entidad ' \
								 + user.profile.entidad.nombre +' produce.'
	
	return subject, recipients, body

## function send email update rol
def createEmailUpdateRole(request, pk):
	user = get_object_or_404(User, pk=pk)
	subject = "SISTEMA DE IDENTIFICACIÓN Y CARACTERIZACIÓN DE OFERTA Y DEMANDA ESTADÍSTICA DEL SEN: Cambio de Rol"
	recipients = []       
	recipients.append(user.email)

	body =  'SISTEMA DE IDENTIFICACIÓN Y CARACTERIZACIÓN DE OFERTA Y DEMANDA ESTADÍSTICA DEL SEN' \
		'\n\n Buen Día '+ user.username + \
			'\n\nQueremos informarle que su rol en el sistema ha cambiado a: ' +user.profile.role.name + \
				'\n\n Acceso al Sistema de Información SEN: (https://inventariosen.dane.gov.co/login/inicioSesion)'
    
	return subject, recipients, body


def sendEmail(subject, recipients, body):    
    send_mail(subject, body, 'sen@dane.gov.co', recipients,  fail_silently = False)  


@login_required
def user_detail(request, pk):
	user = get_object_or_404(User, pk=pk)
	#print("usuarios",user.pk)
	return render(request, 'login/user_detail.html', {'user': user})


@login_required
def user_edit(request, pk):
    user = get_object_or_404(User, pk=pk)
    #print("user ddd",user, "id", user.pk, "profile", user.profile)
    args = {}
	
    if request.method == "POST":
        u_form = EditProfileForm(request.POST, instance=user)
        p_form = FormEditUser(request.POST, instance=user.profile)
        if u_form.is_valid() and p_form.is_valid():
            user_form = u_form.save()
            custom_form = p_form.save(False)
            custom_form.user = user_form
            custom_form.save()
            subject, recipients, body = createEmailUpdateRole(user, user.pk)
            sendEmail(subject, recipients, body)
            return redirect('login:adminUsuarios')
    else:
        u_form = EditProfileForm(instance=user)
        p_form = FormEditUser(instance=user.profile)
        count_entities = Entidades_oe.objects.count()
        args['u_form'] = u_form
        args['p_form'] = p_form
        args['count_entities'] = count_entities
        #print("argumento",args)
    return render(request, 'login/edit_user.html', args)


@login_required   
def createUserProfile(request):
	
	username = request.POST.get('usuario')
	email = request.POST.get('email')
	role_id = request.POST.get('rol')
	entidad_id = request.POST.get('entidad')	
	new_profile = User(username=username, email=email, is_active=True, is_staff=True)
	new_profile.save()
	#user_edit(request, new_profile)


@login_required
def createUser(request):
    count_entities = Entidades_oe.objects.count()
    registered = False 
    if request.method == 'POST':
        user_form = UserForm(data=request.POST)
        profile_form = UserProfileForm(data=request.POST)
        print("has profile",profile_form)
        if user_form.is_valid() and profile_form.is_valid():
            user = user_form.save()
            passw = get_random_string(length=8)
            user.set_password(passw)
            user.save()
            profile = profile_form.save(commit=False)
            profile.user = user
            user_edit(request, user.pk)
            subject, recipients,  body = createEmailNewUser(user, user.pk, passw)
            sendEmail(subject, recipients, body) 
            registered = True
        else:
            print(user_form.errors,profile_form.errors)
    else:
        user_form = UserForm()
        profile_form = UserProfileForm()
        count_entities = Entidades_oe.objects.count()
    return render(request,'login/register.html',
                          {'user_form':user_form,
                           'profile_form':profile_form,
						   'count_entities': count_entities,
                           'registered':registered})


@login_required
def loggedin(request):
    message = 'Bienvenido al inventario de operaciones Estadísticas y registros administrativos'
    count_entities = Entidades_oe.objects.filter(estado_id=1).count()
    user = request.user
    entidadByUser = user.profile.entidad.id

    if user.is_authenticated == True and str(user) == "sdazag": # 1 sdazag

        ###ooee
        count_borrador = OperacionEstadistica.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28)).filter(nombre_est_id=1).count()
        count_enviado = OperacionEstadistica.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28)).filter(nombre_est_id=2).count()
        count_devuelto = OperacionEstadistica.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28)).filter(nombre_est_id=3).count() 
        count_validado = OperacionEstadistica.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28)).filter(nombre_est_id=4).count()
        count_publicado = OperacionEstadistica.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28)).filter(nombre_est_id=5).count()  
        count_archivado = OperacionEstadistica.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28)).filter(nombre_est_id=6).count()
        count_oe = OperacionEstadistica.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28)).count()

        ###rraa
     
        count_borrador_ra = RegistroAdministrativo.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28)).filter(sist_estado_id=1).count()
        count_enviado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28)).filter(sist_estado_id=2).count()
        count_devuelto_ra = RegistroAdministrativo.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28)).filter(sist_estado_id=3).count() 
        count_validado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28)).filter(sist_estado_id=4).count()
        count_publicado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28)).filter(sist_estado_id=5).count()
        count_archivado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28)).filter(sist_estado_id=6).count()
        
        count_ra = RegistroAdministrativo.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28)).count()

        ##ddi
        count_borrador_ddi = demandaInfor.objects.filter(Q (tema_prin_id=2) | Q (tema_prin_id=7) | Q (tema_prin_id=19) | Q (tema_prin_id=28)).filter(nombre_est_id=1).count()
        count_enviado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=2) | Q (tema_prin_id=7) | Q (tema_prin_id=19) | Q (tema_prin_id=28)).filter(nombre_est_id=2).count()
        count_devuelto_ddi = demandaInfor.objects.filter(Q (tema_prin_id=2) | Q (tema_prin_id=7) | Q (tema_prin_id=19) | Q (tema_prin_id=28)).filter(nombre_est_id=3).count()
        count_validado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=2) | Q (tema_prin_id=7) | Q (tema_prin_id=19) | Q (tema_prin_id=28)).filter(nombre_est_id=4).count()
        count_publicado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=2) | Q (tema_prin_id=7) | Q (tema_prin_id=19) | Q (tema_prin_id=28)).filter(nombre_est_id=5).count()
        count_archivado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=2) | Q (tema_prin_id=7) | Q (tema_prin_id=19) | Q (tema_prin_id=28)).filter(nombre_est_id=6).count()
        count_ddi = demandaInfor.objects.filter(Q (tema_prin_id=2) | Q (tema_prin_id=7) | Q (tema_prin_id=19) | Q (tema_prin_id=28)).exclude(nombre_est_id=6).count()

        #### novedad ooee
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28))
        id_oes = []
        novedad_ooee = []
        for index, item  in enumerate(filtros_tema_oe):
            id_oes.append(item.pk)
        for y in id_oes:
            novedades_oe_new = NovedadActualizacion.objects.filter(post_oe_id=y).last()
            if novedades_oe_new != None:
                novedad_ooee.append(str(novedades_oe_new.est_actualiz))
        count_oe_nueva = novedad_ooee.count('Nueva')
        

        #### novedad rraa
        filtros_tema_ra = RegistroAdministrativo.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28))
        id_ras = []
        novedad_rraa = []
        for index, item  in enumerate(filtros_tema_ra):
            id_ras.append(item.pk)
        for x in id_ras:
            novedades_ra_new = NovedadActualizacionRRAA.objects.filter(post_ra_id=x).last()
            if novedades_ra_new != None:
                novedad_rraa.append(str(novedades_ra_new.est_actualiz))
        count_ra_nueva = novedad_rraa.count('Nueva')

        ### novedad ddi
        filtros_tema_ddi = demandaInfor.objects.filter(Q (tema_prin_id=2) | Q (tema_prin_id=7) | Q (tema_prin_id=19) | Q (tema_prin_id=28))
        id_ddis = []
        novedad_ddi = []
        for index, item  in enumerate(filtros_tema_ddi):
            id_ddis.append(item.pk)
        for x in id_ddis:
            novedades_ddi_new = NovedadActualizacionddi.objects.filter(post_ddi_id=x).last()
            if novedades_ddi_new != None:
                novedad_ddi.append(str(novedades_ddi_new.est_actualiz))
        count_ddi_nueva = novedad_ddi.count('Nueva')

    
        """ elif user.is_authenticated == True and str(user) == "aobandor": # 3 

        ##ooee
        oe = OperacionEstadistica.objects.filter(Q (tema=5) | Q (tema=14) | Q (tema=15) | Q (tema=16) | Q (tema=17) | Q (tema=18)).order_by('nombre_oe')
        
        borrador = OperacionEstadistica.objects.filter(Q (tema=5) | Q (tema=14) | Q (tema=15) | Q (tema=16) | Q (tema=17) | Q (tema=18)).filter(nombre_est=1)
        count_borrador = borrador.count()
        
        enviado = OperacionEstadistica.objects.filter(Q (tema=5) | Q (tema=14) | Q (tema=15) | Q (tema=16) | Q (tema=17) | Q (tema=18)).filter(nombre_est=2)
        count_enviado = enviado.count()

        devuelto = OperacionEstadistica.objects.filter(Q (tema=5) | Q (tema=14) | Q (tema=15) | Q (tema=16) | Q (tema=17) | Q (tema=18)).filter(nombre_est=3)
        count_devuelto = devuelto.count()

        validado = OperacionEstadistica.objects.filter(Q (tema=5) | Q (tema=14) | Q (tema=15) | Q (tema=16) | Q (tema=17) | Q (tema=18)).filter(nombre_est=4)
        count_validado = validado.count()

        publicado = OperacionEstadistica.objects.filter(Q (tema=5) | Q (tema=14) | Q (tema=15) | Q (tema=16) | Q (tema=17) | Q (tema=18)).filter(nombre_est=5)
        count_publicado = publicado.count()

        archivado = OperacionEstadistica.objects.filter(Q (tema=5) | Q (tema=14) | Q (tema=15) | Q (tema=16) | Q (tema=17) | Q (tema=18)).filter(nombre_est=6)
        count_archivado = archivado.count()

        count_oe = oe.count()

        #rraa
        ra = RegistroAdministrativo.objects.filter(Q (tema=5) | Q (tema=14) | Q (tema=15) | Q (tema=16) | Q (tema=17) | Q (tema=18)).order_by('nombre_ra')
        
        borrador_ra = RegistroAdministrativo.objects.filter(Q (tema=5) | Q (tema=14) | Q (tema=15) | Q (tema=16) | Q (tema=17) | Q (tema=18)).filter(sist_estado=1)
        count_borrador_ra = borrador_ra.count()
        
        enviado_ra = RegistroAdministrativo.objects.filter(Q (tema=5) | Q (tema=14) | Q (tema=15) | Q (tema=16) | Q (tema=17) | Q (tema=18)).filter(sist_estado=2)
        count_enviado_ra = enviado_ra.count()

        devuelto_ra = RegistroAdministrativo.objects.filter(Q (tema=5) | Q (tema=14) | Q (tema=15) | Q (tema=16) | Q (tema=17) | Q (tema=18)).filter(sist_estado=3)
        count_devuelto_ra = devuelto_ra.count()

        validado_ra = RegistroAdministrativo.objects.filter(Q (tema=5) | Q (tema=14) | Q (tema=15) | Q (tema=16) | Q (tema=17) | Q (tema=18)).filter(sist_estado=4)
        count_validado_ra = validado_ra.count()

        publicado_ra = RegistroAdministrativo.objects.filter(Q (tema=5) | Q (tema=14) | Q (tema=15) | Q (tema=16) | Q (tema=17) | Q (tema=18)).filter(sist_estado=5)
        count_publicado_ra = publicado_ra.count()

        archivado_ra = RegistroAdministrativo.objects.filter(Q (tema=5) | Q (tema=14) | Q (tema=15) | Q (tema=16) | Q (tema=17) | Q (tema=18)).filter(sist_estado=6)
        count_archivado_ra = archivado_ra.count()

        count_ra = ra.count() 
        
        #### novedad ooee
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (tema=5) | Q (tema=14) | Q (tema=15) | Q (tema=16) | Q (tema=17) | Q (tema=18))
        id_oes = []
        novedad_ooee = []
        for index, item  in enumerate(filtros_tema_oe):
            id_oes.append(item.pk)
        for y in id_oes:
            novedades_oe_new = NovedadActualizacion.objects.filter(post_oe_id=y).last()
            if novedades_oe_new != None:
                novedad_ooee.append(str(novedades_oe_new.est_actualiz))
        count_oe_nueva = novedad_ooee.count('Nueva')
        

        #### novedad rraa
        filtros_tema_ra = RegistroAdministrativo.objects.filter(Q (tema=5) | Q (tema=14) | Q (tema=15) | Q (tema=16) | Q (tema=17) | Q (tema=18))
        id_ras = []
        novedad_rraa = []
        for index, item  in enumerate(filtros_tema_ra):
            id_ras.append(item.pk)
        for x in id_ras:
            novedades_ra_new = NovedadActualizacionRRAA.objects.filter(post_ra_id=x).last()
            if novedades_ra_new != None:
                novedad_rraa.append(str(novedades_ra_new.est_actualiz))
        count_ra_nueva = novedad_rraa.count('Nueva')"""

    elif user.is_authenticated == True and str(user) == "dvlizarazog": # 4 dvlizarazog
        ##ooee 
        count_borrador = OperacionEstadistica.objects.filter(Q (tema_id=9) | Q (tema_id=29) | Q (tema_id=30)).filter(nombre_est_id=1).count()
        count_enviado = OperacionEstadistica.objects.filter(Q (tema_id=9) | Q (tema_id=29) | Q (tema_id=30)).filter(nombre_est_id=2).count()
        count_devuelto = OperacionEstadistica.objects.filter(Q (tema_id=9) | Q (tema_id=29) | Q (tema_id=30)).filter(nombre_est_id=3).count()
        count_validado = OperacionEstadistica.objects.filter(Q (tema_id=9) | Q (tema_id=29) | Q (tema_id=30)).filter(nombre_est_id=4).count() 
        count_publicado = OperacionEstadistica.objects.filter(Q (tema_id=9) | Q (tema_id=29) | Q (tema_id=30)).filter(nombre_est_id=5).count()
        count_archivado = OperacionEstadistica.objects.filter(Q (tema_id=9) | Q (tema_id=29) | Q (tema_id=30)).filter(nombre_est_id=6).count()
        count_oe = OperacionEstadistica.objects.filter(Q (tema_id=9) | Q (tema_id=29) | Q (tema_id=30)).count()

        ##rraa      
        count_borrador_ra = RegistroAdministrativo.objects.filter(Q (tema_id=9) | Q (tema_id=29) | Q (tema_id=30)).filter(sist_estado_id=1).count()
        count_enviado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=9) | Q (tema_id=29) | Q (tema_id=30)).filter(sist_estado_id=2).count()
        count_devuelto_ra = RegistroAdministrativo.objects.filter(Q (tema_id=9) | Q (tema_id=29) | Q (tema_id=30)).filter(sist_estado_id=3).count()
        count_validado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=9) | Q (tema_id=29) | Q (tema_id=30)).filter(sist_estado_id=4).count()
        count_publicado_ra = RegistroAdministrativo.objects.filter(Q (tema=9) | Q (tema=29) | Q (tema=30)).filter(sist_estado=5).count()
        count_archivado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=9) | Q (tema_id=29) | Q (tema_id=30)).filter(sist_estado_id=6).count()
        count_ra = RegistroAdministrativo.objects.filter(Q (tema_id=9) | Q (tema_id=29) | Q (tema_id=30)).count()

        ##ddi
        count_borrador_ddi = demandaInfor.objects.filter(Q (tema_prin_id=9) | Q (tema_prin_id=29) | Q (tema_prin_id=30)).filter(nombre_est_id=1).count()
        count_enviado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=9) | Q (tema_prin_id=29) | Q (tema_prin_id=30)).filter(nombre_est_id=2).count()
        count_devuelto_ddi = demandaInfor.objects.filter(Q (tema_prin_id=9) | Q (tema_prin_id=29) | Q (tema_prin_id=30)).filter(nombre_est_id=3).count()
        count_validado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=9) | Q (tema_prin_id=29) | Q (tema_prin_id=30)).filter(nombre_est_id=4).count()
        count_publicado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=9) | Q (tema_prin_id=29) | Q (tema_prin_id=30)).filter(nombre_est_id=5).count()
        count_archivado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=9) | Q (tema_prin_id=29) | Q (tema_prin_id=30)).filter(nombre_est_id=6).count()
        count_ddi = demandaInfor.objects.filter(Q (tema_prin_id=9) | Q (tema_prin_id=29) | Q (tema_prin_id=30)).exclude(nombre_est_id=6).count()

        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (tema_id=9) | Q (tema_id=29) | Q (tema_id=30))
        id_oes = []
        novedad_ooee = []
        for index, item  in enumerate(filtros_tema_oe):
            id_oes.append(item.pk)
        for y in id_oes:
            novedades_oe_new = NovedadActualizacion.objects.filter(post_oe_id=y).last()
            if novedades_oe_new != None:
                novedad_ooee.append(str(novedades_oe_new.est_actualiz))
        count_oe_nueva = novedad_ooee.count('Nueva')
        

        #### novedad rraa
        filtros_tema_ra = RegistroAdministrativo.objects.filter(Q (tema_id=9) | Q (tema_id=29) | Q (tema_id=30))
        id_ras = []
        novedad_rraa = []
        for index, item  in enumerate(filtros_tema_ra):
            id_ras.append(item.pk)
        for x in id_ras:
            novedades_ra_new = NovedadActualizacionRRAA.objects.filter(post_ra_id=x).last()
            if novedades_ra_new != None:
                novedad_rraa.append(str(novedades_ra_new.est_actualiz))
        count_ra_nueva = novedad_rraa.count('Nueva')

        ### novedad ddi
        filtros_tema_ddi = demandaInfor.objects.filter(Q (tema_prin_id=9) | Q (tema_prin_id=29) | Q (tema_prin_id=30))
        id_ddis = []
        novedad_ddi = []
        for index, item  in enumerate(filtros_tema_ddi):
            id_ddis.append(item.pk)
        for x in id_ddis:
            novedades_ddi_new = NovedadActualizacionddi.objects.filter(post_ddi_id=x).last()
            if novedades_ddi_new != None:
                novedad_ddi.append(str(novedades_ddi_new.est_actualiz))
        count_ddi_nueva = novedad_ddi.count('Nueva')

    elif user.is_authenticated == True and str(user) == "ancardenasc": # 5 frherreran freddy es administrador se cambia por gavargasr
        ##ooee
        count_borrador = OperacionEstadistica.objects.filter(tema_id=10).filter(nombre_est_id=1).count()
        count_enviado = OperacionEstadistica.objects.filter(tema_id=10).filter(nombre_est_id=2).count()
        count_devuelto = OperacionEstadistica.objects.filter(tema=10).filter(nombre_est_id=3).count()
        count_validado = OperacionEstadistica.objects.filter(tema_id=10).filter(nombre_est_id=4).count()
        count_publicado = OperacionEstadistica.objects.filter(tema_id=10).filter(nombre_est_id=5).count()
        count_archivado = OperacionEstadistica.objects.filter(tema_id=10).filter(nombre_est_id=6).count()
        count_oe = OperacionEstadistica.objects.filter(tema_id=10).count()

        ##rraa
        count_borrador_ra = RegistroAdministrativo.objects.filter(tema_id=10).filter(sist_estado_id=1).count()
        count_enviado_ra = RegistroAdministrativo.objects.filter(tema_id=10).filter(sist_estado_id=2).count()
        count_devuelto_ra = RegistroAdministrativo.objects.filter(tema_id=10).filter(sist_estado_id=3).count()
        count_validado_ra = RegistroAdministrativo.objects.filter(tema_id=10).filter(sist_estado_id=4).count()
        count_publicado_ra =  RegistroAdministrativo.objects.filter(tema_id=10).filter(sist_estado_id=5).count()
        count_archivado_ra = RegistroAdministrativo.objects.filter(tema_id=10).filter(sist_estado_id=6).count()
        count_ra = RegistroAdministrativo.objects.filter(tema_id=10).count()

        ##ddi
        count_borrador_ddi = demandaInfor.objects.filter(tema_prin_id=10).filter(nombre_est_id=1).count()
        count_enviado_ddi = demandaInfor.objects.filter(tema_prin_id=10).filter(nombre_est_id=2).count()
        count_devuelto_ddi = demandaInfor.objects.filter(tema_prin_id=10).filter(nombre_est_id=3).count()
        count_validado_ddi = demandaInfor.objects.filter(tema_prin_id=10).filter(nombre_est_id=4).count()
        count_publicado_ddi = demandaInfor.objects.filter(tema_prin_id=10).filter(nombre_est_id=5).count()
        count_archivado_ddi = demandaInfor.objects.filter(tema_prin_id=10).filter(nombre_est_id=6).count()
        count_ddi = demandaInfor.objects.filter(tema_prin_id=10).exclude(nombre_est_id=6).count()

        #### novedad ooee
        filtros_tema_oe = OperacionEstadistica.objects.filter(tema_id=10)
        id_oes = []
        novedad_ooee = []
        for index, item  in enumerate(filtros_tema_oe):
            id_oes.append(item.pk)
        for y in id_oes:
            novedades_oe_new = NovedadActualizacion.objects.filter(post_oe_id=y).last()
            if novedades_oe_new != None:
                novedad_ooee.append(str(novedades_oe_new.est_actualiz))
        count_oe_nueva = novedad_ooee.count('Nueva')
        

        #### novedad rraa
        filtros_tema_ra = RegistroAdministrativo.objects.filter(tema_id=10)
        id_ras = []
        novedad_rraa = []
        for index, item  in enumerate(filtros_tema_ra):
            id_ras.append(item.pk)
        for x in id_ras:
            novedades_ra_new = NovedadActualizacionRRAA.objects.filter(post_ra_id=x).last()
            if novedades_ra_new != None:
                novedad_rraa.append(str(novedades_ra_new.est_actualiz))
        count_ra_nueva = novedad_rraa.count('Nueva')
        
        ### novedad ddi
        filtros_tema_ddi = demandaInfor.objects.filter(tema_prin_id=10)
        id_ddis = []
        novedad_ddi = []
        for index, item  in enumerate(filtros_tema_ddi):
            id_ddis.append(item.pk)
        for x in id_ddis:
            novedades_ddi_new = NovedadActualizacionddi.objects.filter(post_ddi_id=x).last()
            if novedades_ddi_new != None:
                novedad_ddi.append(str(novedades_ddi_new.est_actualiz))
        count_ddi_nueva = novedad_ddi.count('Nueva')
            
    elif user.is_authenticated == True and str(user) == "pczambranog": # 6 lhsanchezz se remplaza por pczambranog
        ##ooee
        count_borrador = OperacionEstadistica.objects.filter(Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).filter(nombre_est=1).count()
        count_enviado = OperacionEstadistica.objects.filter(Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).filter(nombre_est_id=2).count()
        count_devuelto = OperacionEstadistica.objects.filter(Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).filter(nombre_est_id=3).count()
        count_validado = OperacionEstadistica.objects.filter(Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).filter(nombre_est_id=4).count()
        count_publicado = OperacionEstadistica.objects.filter(Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).filter(nombre_est_id=5).count()
        count_archivado = OperacionEstadistica.objects.filter(Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).filter(nombre_est_id=6).count()
        count_oe = OperacionEstadistica.objects.filter(Q (tema=11) | Q (tema=12) | Q (tema=25)).count()

        ##rraa
        count_borrador_ra = RegistroAdministrativo.objects.filter(Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).filter(sist_estado_id=1).count()
        count_enviado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).filter(sist_estado_id=2).count()
        count_devuelto_ra = RegistroAdministrativo.objects.filter(Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).filter(sist_estado_id=3).count()
        count_validado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).filter(sist_estado_id=4).count()
        count_publicado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).filter(sist_estado_id=5).count()
        count_archivado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).filter(sist_estado_id=6).count()
        count_ra =  RegistroAdministrativo.objects.filter(Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).count()

        ##ddi
        count_borrador_ddi = demandaInfor.objects.filter(Q (tema_prin_id=11) | Q (tema_prin_id=12) | Q (tema_prin_id=25)).filter(nombre_est_id=1).count()
        count_enviado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=11) | Q (tema_prin_id=12) | Q (tema_prin_id=25)).filter(nombre_est_id=2).count()
        count_devuelto_ddi = demandaInfor.objects.filter(Q (tema_prin_id=11) | Q (tema_prin_id=12) | Q (tema_prin_id=25)).filter(nombre_est_id=3).count()
        count_validado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=11) | Q (tema_prin_id=12) | Q (tema_prin_id=25)).filter(nombre_est_id=4).count()
        count_publicado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=11) | Q (tema_prin_id=12) | Q (tema_prin_id=25)).filter(nombre_est_id=5).count()
        count_archivado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=11) | Q (tema_prin_id=12) | Q (tema_prin_id=25)).filter(nombre_est_id=6).count()
        count_ddi = demandaInfor.objects.filter(Q (tema_prin_id=11) | Q (tema_prin_id=12) | Q (tema_prin_id=25)).exclude(nombre_est_id=6).count()

        #### novedad ooee
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25))
        id_oes = []
        novedad_ooee = []
        for index, item  in enumerate(filtros_tema_oe):
            id_oes.append(item.pk)
        for y in id_oes:
            novedades_oe_new = NovedadActualizacion.objects.filter(post_oe_id=y).last()
            if novedades_oe_new != None:
                novedad_ooee.append(str(novedades_oe_new.est_actualiz))
        count_oe_nueva = novedad_ooee.count('Nueva')
        

        #### novedad rraa
        filtros_tema_ra = RegistroAdministrativo.objects.filter(Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25))
        id_ras = []
        novedad_rraa = []
        for index, item  in enumerate(filtros_tema_ra):
            id_ras.append(item.pk)
        for x in id_ras:
            novedades_ra_new = NovedadActualizacionRRAA.objects.filter(post_ra_id=x).last()
            if novedades_ra_new != None:
                novedad_rraa.append(str(novedades_ra_new.est_actualiz))
        count_ra_nueva = novedad_rraa.count('Nueva')

        ### novedad ddi
        filtros_tema_ddi = demandaInfor.objects.filter(Q (tema_prin_id=11) | Q (tema_prin_id=12) | Q (tema_prin_id=25))
        id_ddis = []
        novedad_ddi = []
        for index, item  in enumerate(filtros_tema_ddi):
            id_ddis.append(item.pk)
        for x in id_ddis:
            novedades_ddi_new = NovedadActualizacionddi.objects.filter(post_ddi_id=x).last()
            if novedades_ddi_new != None:
                novedad_ddi.append(str(novedades_ddi_new.est_actualiz))
        count_ddi_nueva = novedad_ddi.count('Nueva')

    elif user.is_authenticated == True and str(user) == "mjpradag":   # 7 mppulidor se cambia por mjpradag
        ##ooee
        count_borrador = OperacionEstadistica.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27)).filter(nombre_est_id=1).count()
        count_enviado =  OperacionEstadistica.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27)).filter(nombre_est_id=2).count()
        count_devuelto =  OperacionEstadistica.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27)).filter(nombre_est_id=3).count()
        count_validado = OperacionEstadistica.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27)).filter(nombre_est_id=4).count()
        count_publicado = OperacionEstadistica.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27)).filter(nombre_est_id=5).count()
        count_archivado = OperacionEstadistica.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27)).filter(nombre_est_id=6).count()
        count_oe = OperacionEstadistica.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27)).count()

        ##rraa
        count_borrador_ra = RegistroAdministrativo.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27)).filter(sist_estado_id=1).count()
        count_enviado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27)).filter(sist_estado=2).count()
        count_devuelto_ra =  RegistroAdministrativo.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27)).filter(sist_estado_id=3).count()
        count_validado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27)).filter(sist_estado_id=4).count()
        count_publicado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27)).filter(sist_estado_id=5).count() 
        count_archivado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27)).filter(sist_estado_id=6).count()
        count_ra = RegistroAdministrativo.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27)).count()

        ##ddi
        count_borrador_ddi = demandaInfor.objects.filter(Q (tema_prin_id=13) | Q (tema_prin_id=24) | Q (tema_prin_id=26) | Q (tema_prin_id=27)).filter(nombre_est_id=1).count()
        count_enviado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=13) | Q (tema_prin_id=24) | Q (tema_prin_id=26) | Q (tema_prin_id=27)).filter(nombre_est_id=2).count()
        count_devuelto_ddi = demandaInfor.objects.filter(Q (tema_prin_id=13) | Q (tema_prin_id=24) | Q (tema_prin_id=26) | Q (tema_prin_id=27)).filter(nombre_est_id=3).count()
        count_validado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=13) | Q (tema_prin_id=24) | Q (tema_prin_id=26) | Q (tema_prin_id=27)).filter(nombre_est_id=4).count()
        count_publicado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=13) | Q (tema_prin_id=24) | Q (tema_prin_id=26) | Q (tema_prin_id=27)).filter(nombre_est_id=5).count()
        count_archivado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=13) | Q (tema_prin_id=24) | Q (tema_prin_id=26) | Q (tema_prin_id=27)).filter(nombre_est_id=6).count()
        count_ddi = demandaInfor.objects.exclude(nombre_est_id=6).count()
        
        #### novedad ooee
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27))
        id_oes = []
        novedad_ooee = []
        for index, item  in enumerate(filtros_tema_oe):
            id_oes.append(item.pk)
        for y in id_oes:
            novedades_oe_new = NovedadActualizacion.objects.filter(post_oe_id=y).last()
            if novedades_oe_new != None:
                novedad_ooee.append(str(novedades_oe_new.est_actualiz))
        count_oe_nueva = novedad_ooee.count('Nueva')
        

        #### novedad rraa
        filtros_tema_ra = RegistroAdministrativo.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27))
        id_ras = []
        novedad_rraa = []
        for index, item  in enumerate(filtros_tema_ra):
            id_ras.append(item.pk)
        for x in id_ras:
            novedades_ra_new = NovedadActualizacionRRAA.objects.filter(post_ra_id=x).last()
            if novedades_ra_new != None:
                novedad_rraa.append(str(novedades_ra_new.est_actualiz))
        count_ra_nueva = novedad_rraa.count('Nueva')

        ### novedad ddi
        filtros_tema_ddi = demandaInfor.objects.filter(Q (tema_prin_id=13) | Q (tema_prin_id=24) | Q (tema_prin_id=26) | Q (tema_prin_id=27))
        id_ddis = []
        novedad_ddi = []
        for index, item  in enumerate(filtros_tema_ddi):
            id_ddis.append(item.pk)
        for x in id_ddis:
            novedades_ddi_new = NovedadActualizacionddi.objects.filter(post_ddi_id=x).last()
            if novedades_ddi_new != None:
                novedad_ddi.append(str(novedades_ddi_new.est_actualiz))
        count_ddi_nueva = novedad_ddi.count('Nueva')

    elif user.is_authenticated == True and str(user) == "eeguayazans": # 8 eeguayazans Edgar Eduardo Guayazan Sierra
        ##ooee
        count_borrador = OperacionEstadistica.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20)).filter(nombre_est_id=1).count()
        count_enviado = OperacionEstadistica.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20)).filter(nombre_est_id=2).count()
        count_devuelto = OperacionEstadistica.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20)).filter(nombre_est_id=3).count()
        count_validado = OperacionEstadistica.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20)).filter(nombre_est_id=4).count()
        count_publicado = OperacionEstadistica.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20)).filter(nombre_est_id=5).count()
        count_archivado = OperacionEstadistica.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20)).filter(nombre_est_id=6).count()
        count_oe = OperacionEstadistica.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20)).count()

        ##rraa
        count_borrador_ra = RegistroAdministrativo.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20)).filter(sist_estado_id=1).count()
        count_enviado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20)).filter(sist_estado_id=2).count()
        count_devuelto_ra = RegistroAdministrativo.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20)).filter(sist_estado_id=3).count()
        count_validado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20)).filter(sist_estado_id=4).count()
        count_publicado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20)).filter(sist_estado_id=5).count()
        count_archivado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20)).filter(sist_estado=6).count()
        count_ra = RegistroAdministrativo.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20)).count()

        ##ddi
        count_borrador_ddi = demandaInfor.objects.filter(Q (tema_prin_id=4) | Q (tema_prin_id=8) | Q (tema_prin_id=20)).filter(nombre_est_id=1).count()
        count_enviado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=4) | Q (tema_prin_id=8) | Q (tema_prin_id=20)).filter(nombre_est_id=2).count()
        count_devuelto_ddi = demandaInfor.objects.filter(Q (tema_prin_id=4) | Q (tema_prin_id=8) | Q (tema_prin_id=20)).filter(nombre_est_id=3).count()
        count_validado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=4) | Q (tema_prin_id=8) | Q (tema_prin_id=20)).filter(nombre_est_id=4).count()
        count_publicado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=4) | Q (tema_prin_id=8) | Q (tema_prin_id=20)).filter(nombre_est_id=5).count()
        count_archivado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=4) | Q (tema_prin_id=8) | Q (tema_prin_id=20)).filter(nombre_est_id=6).count()
        count_ddi = demandaInfor.objects.filter(Q (tema_prin_id=4) | Q (tema_prin_id=8) | Q (tema_prin_id=20)).exclude(nombre_est_id=6).count()
        
        #### novedad ooee
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20))
        id_oes = []
        novedad_ooee = []
        for index, item  in enumerate(filtros_tema_oe):
            id_oes.append(item.pk)
        for y in id_oes:
            novedades_oe_new = NovedadActualizacion.objects.filter(post_oe_id=y).last()
            if novedades_oe_new != None:
                novedad_ooee.append(str(novedades_oe_new.est_actualiz))
        count_oe_nueva = novedad_ooee.count('Nueva')
        

        #### novedad rraa
        filtros_tema_ra = RegistroAdministrativo.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20))
        id_ras = []
        novedad_rraa = []
        for index, item  in enumerate(filtros_tema_ra):
            id_ras.append(item.pk)
        for x in id_ras:
            novedades_ra_new = NovedadActualizacionRRAA.objects.filter(post_ra_id=x).last()
            if novedades_ra_new != None:
                novedad_rraa.append(str(novedades_ra_new.est_actualiz))
        count_ra_nueva = novedad_rraa.count('Nueva')

        ### novedad ddi
        filtros_tema_ddi = demandaInfor.objects.filter(Q (tema_prin_id=4) | Q (tema_prin_id=8) | Q (tema_prin_id=20))
        id_ddis = []
        novedad_ddi = []
        for index, item  in enumerate(filtros_tema_ddi):
            id_ddis.append(item.pk)
        for x in id_ddis:
            novedades_ddi_new = NovedadActualizacionddi.objects.filter(post_ddi_id=x).last()
            if novedades_ddi_new != None:
                novedad_ddi.append(str(novedades_ddi_new.est_actualiz))
        count_ddi_nueva = novedad_ddi.count('Nueva')

    elif user.is_authenticated == True and str(user) == "mlbarretob": # 9 mlbarretob
        ##ooee
        count_borrador = OperacionEstadistica.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23)).filter(nombre_est_id=1).count()
        count_enviado = OperacionEstadistica.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23)).filter(nombre_est_id=2).count()
        count_devuelto = OperacionEstadistica.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23)).filter(nombre_est_id=3).count()
        count_validado = OperacionEstadistica.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23)).filter(nombre_est_id=4).count()
        count_publicado =  OperacionEstadistica.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23)).filter(nombre_est_id=5).count()
        count_archivado = OperacionEstadistica.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23)).filter(nombre_est_id=6).count()
        count_oe =  OperacionEstadistica.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23)).count()

        ##rraa  
        count_borrador_ra = RegistroAdministrativo.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23)).filter(sist_estado_id=1).count()
        count_enviado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23)).filter(sist_estado_id=2).count()
        count_devuelto_ra = RegistroAdministrativo.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23)).filter(sist_estado_id=3).count() 
        count_validado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23)).filter(sist_estado_id=4).count() 
        count_publicado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23)).filter(sist_estado_id=5).count() 
        count_archivado_ra = RegistroAdministrativo.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23)).filter(sist_estado_id=6).count()

        count_ra = RegistroAdministrativo.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23)).count()

        ##ddi  tema_prin
        count_borrador_ddi = demandaInfor.objects.filter(Q (tema_prin_id=21) | Q (tema_prin_id=22) | Q (tema_prin_id=23)).filter(nombre_est_id=1).count()
        count_enviado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=21) | Q (tema_prin_id=22) | Q (tema_prin_id=23)).filter(nombre_est_id=2).count()
        count_devuelto_ddi = demandaInfor.objects.filter(Q (tema_prin_id=21) | Q (tema_prin_id=22) | Q (tema_prin_id=23)).filter(nombre_est_id=3).count()
        count_validado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=21) | Q (tema_prin_id=22) | Q (tema_prin_id=23)).filter(nombre_est_id=4).count()
        count_publicado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=21) | Q (tema_prin_id=22) | Q (tema_prin_id=23)).filter(nombre_est_id=5).count()
        count_archivado_ddi = demandaInfor.objects.filter(Q (tema_prin_id=21) | Q (tema_prin_id=22) | Q (tema_prin_id=23)).filter(nombre_est_id=6).count()
        count_ddi = demandaInfor.objects.filter(Q (tema_prin_id=21) | Q (tema_prin_id=22) | Q (tema_prin_id=23)).exclude(nombre_est_id=6).count()

        #### novedad ooee
        filtros_tema_oe = OperacionEstadistica.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23))
        id_oes = []
        novedad_ooee = []
        for index, item  in enumerate(filtros_tema_oe):
            id_oes.append(item.pk)
        for y in id_oes:
            novedades_oe_new = NovedadActualizacion.objects.filter(post_oe_id=y).last()
            if novedades_oe_new != None:
                novedad_ooee.append(str(novedades_oe_new.est_actualiz))
        count_oe_nueva = novedad_ooee.count('Nueva')
        

        #### novedad rraa
        filtros_tema_ra = RegistroAdministrativo.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23))
        id_ras = []
        novedad_rraa = []
        for index, item  in enumerate(filtros_tema_ra):
            id_ras.append(item.pk)
        for x in id_ras:
            novedades_ra_new = NovedadActualizacionRRAA.objects.filter(post_ra_id=x).last()
            if novedades_ra_new != None:
                novedad_rraa.append(str(novedades_ra_new.est_actualiz))
        count_ra_nueva = novedad_rraa.count('Nueva')

        ### novedad ddi
        filtros_tema_ddi = demandaInfor.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23))
        id_ddis = []
        novedad_ddi = []
        for index, item  in enumerate(filtros_tema_ddi):
            id_ddis.append(item.pk)
        for x in id_ddis:
            novedades_ddi_new = NovedadActualizacionddi.objects.filter(post_ddi_id=x).last()
            if novedades_ddi_new != None:
                novedad_ddi.append(str(novedades_ddi_new.est_actualiz))
        count_ddi_nueva = novedad_ddi.count('Nueva')

    elif user.is_authenticated == True and user.profile.role.id == 1 or user.profile.role.id == 5 or user.profile.role.id == 6 or user.profile.role.id == 7: 

        ##ooee
        count_borrador = OperacionEstadistica.objects.filter(nombre_est_id=1).count()
        count_enviado = OperacionEstadistica.objects.filter(nombre_est_id=2).count()
        count_devuelto = OperacionEstadistica.objects.filter(nombre_est_id=3).count()
        count_validado = OperacionEstadistica.objects.filter(nombre_est_id=4).count()
        count_publicado = OperacionEstadistica.objects.filter(nombre_est_id=5).count()
        count_archivado = OperacionEstadistica.objects.filter(nombre_est=6).count()
        count_oe = OperacionEstadistica.objects.count()
        ##rraa
        count_borrador_ra = RegistroAdministrativo.objects.filter(sist_estado=1).count() 
        count_enviado_ra = RegistroAdministrativo.objects.filter(sist_estado=2).count()
        count_devuelto_ra = RegistroAdministrativo.objects.filter(sist_estado=3).count()
        count_validado_ra = RegistroAdministrativo.objects.filter(sist_estado=4).count()
        count_publicado_ra = RegistroAdministrativo.objects.filter(sist_estado=5).count()
        count_archivado_ra = RegistroAdministrativo.objects.filter(sist_estado=6).count()

        count_ra = RegistroAdministrativo.objects.count()

        ##ddi
        count_borrador_ddi = demandaInfor.objects.filter(nombre_est_id=1).count()
        count_enviado_ddi = demandaInfor.objects.filter(nombre_est_id=2).count()
        count_devuelto_ddi = demandaInfor.objects.filter(nombre_est_id=3).count()
        count_validado_ddi = demandaInfor.objects.filter(nombre_est_id=4).count()
        count_publicado_ddi = demandaInfor.objects.filter(nombre_est_id=5).count()
        count_archivado_ddi = demandaInfor.objects.filter(nombre_est_id=6).count()
        count_ddi = demandaInfor.objects.exclude(nombre_est_id=6).count()

        #### novedad ooee

        novedad = []
        unique_oe = NovedadActualizacion.objects.values('post_oe').order_by('post_oe').distinct()
       
        for nov in unique_oe:
            
            lista_oe = list(nov.values())
            oe_nov = str(lista_oe)[1:-1]
            novedad_last = NovedadActualizacion.objects.filter(post_oe_id=oe_nov).last()
            novedad.append(str(novedad_last.est_actualiz))
        count_oe_nueva = novedad.count('Nueva')

        #### novedad rraa
        novedad_rraa = []
        unique_ra = NovedadActualizacionRRAA.objects.values('post_ra').order_by('post_ra').distinct()
        for nove in unique_ra: 
            lista_ra = list(nove.values())
            ra_nov = str(lista_ra)[1:-1]
            novedad_last_ra = NovedadActualizacionRRAA.objects.filter(post_ra_id=ra_nov).last()
            novedad_rraa.append(str(novedad_last_ra.est_actualiz)) 
        count_ra_nueva = novedad_rraa.count('Nueva')


        ### novedad ddi
        unique_ddi = NovedadActualizacionddi.objects.values('post_ddi').order_by('post_ddi').distinct()
        novedad_ddi = []
        for noved in unique_ddi: 
            lista_ddi = list(noved.values())
            ddi_nov = str(lista_ddi)[1:-1]
            novedad_last_ddi = NovedadActualizacionddi.objects.filter(post_ddi_id=ddi_nov).last()
            novedad_ddi.append(str(novedad_last_ddi.est_actualiz)) 
        count_ddi_nueva = novedad_ddi.count('Nueva')

    else:  ## rol fuente
        ##ooee
        count_borrador = OperacionEstadistica.objects.filter(entidad_id=user.profile.entidad.id).filter(nombre_est_id=1).count()
        count_enviado = OperacionEstadistica.objects.filter(entidad_id=user.profile.entidad.id).filter(nombre_est_id=2).count()
        count_devuelto = OperacionEstadistica.objects.filter(entidad_id=user.profile.entidad.id).filter(nombre_est_id=3).count()
        count_validado = OperacionEstadistica.objects.filter(entidad_id=user.profile.entidad.id).filter(nombre_est_id=4).count() 
        count_publicado = OperacionEstadistica.objects.filter(entidad_id=user.profile.entidad.id).filter(nombre_est_id=5).count()
        count_archivado = OperacionEstadistica.objects.filter(entidad_id=user.profile.entidad.id).filter(nombre_est_id=6).count()
        count_oe = OperacionEstadistica.objects.filter(entidad_id=entidadByUser).count()
        ##rraa         
        count_borrador_ra = RegistroAdministrativo.objects.filter(entidad_pri_id=user.profile.entidad.id).filter(sist_estado_id=1).count()
        count_enviado_ra = RegistroAdministrativo.objects.filter(entidad_pri_id=user.profile.entidad.id).filter(sist_estado_id=2).count()
        count_devuelto_ra = RegistroAdministrativo.objects.filter(entidad_pri_id=user.profile.entidad.id).filter(sist_estado_id=3).count() 
        count_validado_ra = RegistroAdministrativo.objects.filter(entidad_pri_id=user.profile.entidad.id).filter(sist_estado_id=4).count() 
        count_publicado_ra = RegistroAdministrativo.objects.filter(entidad_pri_id=user.profile.entidad.id).filter(sist_estado_id=5).count()
        count_archivado_ra = RegistroAdministrativo.objects.filter(entidad_pri_id=user.profile.entidad.id).filter(sist_estado_id=6).count()
        count_ra = RegistroAdministrativo.objects.filter(entidad_pri_id=entidadByUser).count()

        ##ddi
        count_borrador_ddi = demandaInfor.objects.filter(entidad_sol=user.profile.entidad.id).filter(nombre_est_id=1).count()
        count_enviado_ddi = demandaInfor.objects.filter(entidad_sol=user.profile.entidad.id).filter(nombre_est_id=2).count()
        count_devuelto_ddi = demandaInfor.objects.filter(entidad_sol=user.profile.entidad.id).filter(nombre_est_id=3).count()
        count_validado_ddi = demandaInfor.objects.filter(entidad_sol=user.profile.entidad.id).filter(nombre_est_id=4).count()
        count_publicado_ddi = demandaInfor.objects.filter(entidad_sol=user.profile.entidad.id).filter(nombre_est_id=5).count()
        count_archivado_ddi = demandaInfor.objects.filter(entidad_sol=user.profile.entidad.id).filter(nombre_est_id=6).count()
        count_ddi = demandaInfor.objects.filter(entidad_sol=entidadByUser).exclude(nombre_est_id=6).count()

        #### novedad ooee
        filtros_tema_oe = OperacionEstadistica.objects.filter(entidad_id=user.profile.entidad.id)
        id_oes = []
        novedad_ooee = []
        for index, item  in enumerate(filtros_tema_oe):
            id_oes.append(item.pk)
        for y in id_oes:
            novedades_oe_new = NovedadActualizacion.objects.filter(post_oe_id=y).last()
            if novedades_oe_new != None:
                novedad_ooee.append(str(novedades_oe_new.est_actualiz))
        count_oe_nueva = novedad_ooee.count('Nueva')
        

        #### novedad rraa
        filtros_tema_ra = RegistroAdministrativo.objects.filter(entidad_pri_id=user.profile.entidad.id)
        id_ras = []
        novedad_rraa = []
        for index, item  in enumerate(filtros_tema_ra):
            id_ras.append(item.pk)
        for x in id_ras:
            novedades_ra_new = NovedadActualizacionRRAA.objects.filter(post_ra_id=x).last()
            if novedades_ra_new != None:
                novedad_rraa.append(str(novedades_ra_new.est_actualiz))
        count_ra_nueva = novedad_rraa.count('Nueva')

        ### novedad ddi
        filtros_tema_ddi = demandaInfor.objects.filter(entidad_sol_id=user.profile.entidad.id)
        id_ddis = []
        novedad_ddi = []
        for index, item  in enumerate(filtros_tema_ddi):
            id_ddis.append(item.pk)
        for x in id_ddis:
            novedades_ddi_new = NovedadActualizacionddi.objects.filter(post_ddi_id=x).last()
            if novedades_ddi_new != None:
                novedad_ddi.append(str(novedades_ddi_new.est_actualiz))
        count_ddi_nueva = novedad_ddi.count('Nueva')

    return render(request, 'login/bienvenido.html', {'message': message, "count_oe": count_oe, 
	'count_borrador': count_borrador, 'count_enviado': count_enviado, 'count_devuelto': count_devuelto,
	'count_validado': count_validado, 'count_publicado': count_publicado, 'count_archivado': count_archivado, 
    "count_ra": count_ra, 
	'count_borrador_ra': count_borrador_ra, 'count_enviado_ra': count_enviado_ra, 'count_devuelto_ra': count_devuelto_ra,
	'count_validado_ra': count_validado_ra, 'count_publicado_ra': count_publicado_ra, 'count_archivado_ra': count_archivado_ra,
    'count_entities': count_entities, 'count_oe_nueva': count_oe_nueva,  'count_ra_nueva':  count_ra_nueva,
    'count_borrador_ddi': count_borrador_ddi, 'count_enviado_ddi': count_enviado_ddi, 'count_devuelto_ddi': count_devuelto_ddi,
    'count_validado_ddi': count_validado_ddi, 'count_publicado_ddi': count_publicado_ddi, 'count_archivado_ddi': count_archivado_ddi,
    'count_ddi': count_ddi, 'count_ddi_nueva': count_ddi_nueva }) 
	

def logout(request):
    auth.logout(request)
    count_entities =  Entidades_oe.objects.filter(estado=1).count()
    
    #login/login.html
    return render(request,'login/logout.html', {'count_entities': count_entities } )

@login_required
def userAdministration(request):
	
    users = User.objects.all()
    count_entities = Entidades_oe.objects.count()     
    return render(request, 'login/allUsers.html',  {'users': users, 'count_entities':count_entities})

# ----------------------------report Users xlwt no esta implementado ---------------------

def export_users_xls(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="reporte_usuarios.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Usuarios')
    row_num = 0
    row_two_num = 2
    row_four_num = 4
    titleStyle = xlwt.easyxf('font: name Roboto, bold 1,height 280;')
    style_text_align_vert_center_horiz_center = xlwt.easyxf("align: wrap on,horiz centre; font: name Roboto, bold 1; borders: top_color black, bottom_color black, right_color black, left_color black,\
                              left thin, right thin, top thin, bottom thin;\
                     pattern: pattern solid, fore_color white;")
    style_text_align_vert_bottom_horiz_left = xlwt.easyxf("align: wrap on,horiz Left; font: name Roboto; borders: top_color black, bottom_color black, right_color black, left_color black,\
                              left thin, right thin, top thin, bottom thin;\
                     pattern: pattern solid, fore_color white;")
    SubTitleStyle = xlwt.easyxf('font: name Roboto, bold 1,height 180;')

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    col_width = 256*30
   

    columns = [     'Nombre',
                    'Apellidos',
                    'username', 
                    'Email',
					'Rol'
                    
                ]

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], style_text_align_vert_center_horiz_center)

    font_style = xlwt.XFStyle()
    rowsEnt = User.objects.all().values_list('first_name')
    #print( rowsEnt )
	
    rows = User.objects.all().values_list('first_name',
                                          'last_name', 
                                          'username', 
                                          'email')
	 

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.col(col_num).width = col_width
            ws.write(row_num, col_num, str(row[col_num]), style_text_align_vert_bottom_horiz_left)

    wb.save(response)
    return response


# ----------------------------report Users by openpyxl---------------------


class report_users_xls(TemplateView):
    def get(self, request, *args, **kwargs):
        profiles = Profile.objects.all()
        users = User.objects.all()	    
        wb = Workbook()
        ws = wb.active

		## size rows

        ws.row_dimensions[1].height = 55

		## size column 
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 70
        ws.column_dimensions['G'].width = 70

		##insert image
		#img = openpyxl.drawing.image.Image('media/pictures/logoSEN.png')
		#img.anchor = 'A1'	
		#ws.add_image(img)


		##styles
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        title_cell = ws['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Reporte de Usuarios'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')
        title_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
	
        ws.merge_cells('A1:F1')

        first_name_cell = ws['A2']
        first_name_cell.value = 'Nombre'
        first_name_cell.font = Font(bold=True)
        first_name_cell.alignment = Alignment(horizontal='center')
        first_name_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        last_name_cell = ws['B2']
        last_name_cell.value = 'Apellido'
        last_name_cell.font = Font(bold=True)
        last_name_cell.alignment = Alignment(horizontal='center')
        last_name_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        email_cell = ws['C2'] 
        email_cell.value = 'Email'
        email_cell.font = Font(bold=True)
        email_cell.alignment = Alignment(horizontal='center')
        email_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
		
        username_cell = ws['D2'] 
        username_cell.value = 'Username'
        username_cell.font = Font(bold=True)
        username_cell.alignment = Alignment(horizontal='center')
        username_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

		
        role_cell = ws['E2']
        role_cell.value = 'Rol'
        role_cell.font = Font(bold=True)
        role_cell.alignment = Alignment(horizontal='center')
        role_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
		
        entitie_cell = ws['F2'] 
        entitie_cell.value = 'Entidad'
        entitie_cell.font = Font(bold=True)
        entitie_cell.alignment = Alignment(horizontal='center')	
        entitie_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))		
		

        cont = 3
        for user in users:
            ws.cell(row = cont, column = 1).value = user.first_name
            ws.cell(row = cont, column = 2).value = user.last_name
            ws.cell(row = cont, column = 3).value = user.email
            ws.cell(row = cont, column = 1).border = thin_border
            ws.cell(row = cont, column = 2).border = thin_border
            ws.cell(row = cont, column = 3).border = thin_border
            cont+=1

        conta = 3
        for profile in profiles:
            ws.cell(row = conta, column = 4).value = str(profile.user)
            ws.cell(row = conta, column = 5).value = str(profile.role)
            ws.cell(row = conta, column = 6).value = str(profile.entidad)
            ws.cell(row = conta, column = 4).border = thin_border
            ws.cell(row = conta, column = 5).border = thin_border
            ws.cell(row = conta, column = 6).border = thin_border
			
            conta+=1			

        file_name = "reporte_usuarios.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(file_name)
        response['Content-Disposition'] = content
        response['Set-Cookie'] = 'fileDownload=true; Path=/'
        wb.save(response)
        return response 


 




