from django.forms.formsets import formset_factory
from login.models import Role, User, Profile 
from entities.models import Entidades_oe
from .models import demandaInfor, UtilizaInfEstText, listaVariables, \
    NormasText, InfSolRespRequerimientoText, UsuariosPrincInfEstText, DesagregacionReqText, TipoRequerimientoText, DesagregacionGeoReqGeograficaText, \
        PeriodicidadDifusionText, consumidores_info, SolReqInformacion, Comentariosddi, NovedadActualizacionddi, Criticaddi
   
from .forms import  CreateDDIForm, EditDDIForm, CrearUtilizaInfEstTextForm, CrearUsPrincInfEstTextForm, CrearNormasTextForm,\
        CrearInfSolRespRequerimientoTextForm, TipoRequerimientoTextForm, listaVariablesFormset, DesagregacionReqTextForm, \
            DesagregacionGeoReqGeograficaTextForm, PeriodicidadDifusionTextForm, NovedadDDIForm,  \
                    EditUtilizaInfEstTextForm, EditUsPrincInfEstTextForm, EditNormasTextForm, EditInfSolRespRequerimientoTextForm, \
                    EditTipoRequerimientoTextForm, EditDesagregacionReqTextForm, EditDesaReqGeogTextForm, EditPerDifusionTextForm, \
                        EditSolReqInformacionForm, CrearSolReqInformacionForm, ComentariosddiForm, CriticaddiForm, \
                            CrearConsumidorInfoForm, EditarConsumidorInfoForm
                          

from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator

from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render, redirect, get_object_or_404
##  report by excel ##
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Font, Fill, Border, Side, GradientFill, Alignment, Color, PatternFill
##  end report by excel ##
from django.contrib import messages 
from django.views.generic import TemplateView
from django.forms import inlineformset_factory
from django.core import serializers
from .filters import TopicsFilterddi, NameFilterDDI
from itertools import chain
from django.core.mail import send_mail
from django.utils import timezone

from django.db.models import Q

from django.http import JsonResponse
from itertools import groupby
from django.db.models import Count

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
# Create your views here.

####libreria reporte excel
import xlsxwriter
import io
from django.views.generic import View
import json


## vista de modulo de consulta para demandas en estado publicado
def consultaModuloddi(request):
    
    results = demandaInfor.objects.filter(nombre_est_id=5).count()
    count_entities = Entidades_oe.objects.filter(estado_id=1).count()
    ddi = demandaInfor.objects.filter(nombre_est_id=5).only('entidad_sol', 'area_tem', 'tema_prin').order_by('entidad_sol')
    count_total_ddi = ddi.count()


    ddi_filter = TopicsFilterddi(request.GET, queryset=ddi)
    nameddiFilter = NameFilterDDI(request.GET, queryset=ddi)

    page = request.GET.get('page', 1)
    paginator = Paginator(ddi, 10)

    try:
        demandasInf = paginator.page(page)
    except PageNotAnInteger:
        demandasInf = paginator.page(1)
    except EmptyPage:
        demandasInf = paginator.page(paginator.num_pages)

    index = demandasInf.number - 1
    max_index = len(paginator.page_range)
    start_index = index - 3 if index >= 3 else 0
    end_index = index + 3 if index <= max_index - 3 else max_index
    page_start = 1
    page_end = max_index
    page_range = paginator.page_range[start_index:end_index]
    page_number= int(page)
    #result_show = show_operaciones*page_number if show_operaciones == 12 else results
    #show_incre = page_number if page_number == 1 else result_show - show_operaciones + 1

    
    return render(request, 'ddi/consulta/modulo_consulta_ddi.html',  {'ddi': ddi, 'filter': ddi_filter,
    'count_entities': count_entities, 'count_total_ddi': count_total_ddi, 'demandasInf': demandasInf,
    'page_range': page_range, 'page_end': page_end, 'page_start': page_start, 'results': results, 
    'nameddiFilter': nameddiFilter })



### modulo de consulta ***  filtro por entidad, tema y area tematica con base completa demandas publicados
class SearchAjaxddiView(TemplateView):
    def get(self, request, *args, **kwargs):
        
        id_requerimiento = request.GET.get('requerimiento') #opc 1
        id_area_tematica = request.GET.get('area_tematica') #opc 2
        id_tema = request.GET.get('tema') #opc 3

        ## opc 1 es diferente de vacia
        if id_requerimiento != "" and id_area_tematica == "" and  id_tema == "":

            ddi = demandaInfor.objects.filter(nombre_est=5)
            ddi_filter = TopicsFilterddi(request.GET, queryset=ddi)
            demdeInf = ddi_filter.qs.filter(pm_b_6=id_requerimiento)

        ## opc 1 y opc 2 es diferente de vacia
        elif  id_requerimiento != "" and id_area_tematica != "" and  id_tema == "":
        
            ddi = demandaInfor.objects.filter(nombre_est=5)
            ddi_filter = TopicsFilterddi(request.GET, queryset=ddi)
            demdeInf = ddi_filter.qs.filter(pm_b_6=id_requerimiento).filter(area_tem=id_area_tematica)
        
        ## opc 1, opc 2 y opc 3 es diferente de vacia
        elif id_requerimiento != "" and id_area_tematica != "" and  id_tema != "":

            ddi = demandaInfor.objects.filter(nombre_est=5)
            ddi_filter = TopicsFilterddi(request.GET, queryset=ddi)
            demdeInf = ddi_filter.qs.filter(pm_b_6=id_requerimiento).filter(area_tem=id_area_tematica).filter(tema_prin=id_tema)
        
        ## opcion 1 y opcion 3 es diferente de vacia
        elif id_requerimiento != "" and id_area_tematica == "" and  id_tema != "":
        
            ddi = demandaInfor.objects.filter(nombre_est=5)
            ddi_filter = TopicsFilterddi(request.GET, queryset=ddi)
            demdeInf = ddi_filter.qs.filter(pm_b_6=id_requerimiento).filter(tema_prin=id_tema)

        #opcion 2 es diferente de vacia
        elif id_requerimiento == "" and id_area_tematica != "" and  id_tema == "":

            ddi = demandaInfor.objects.filter(nombre_est=5)
            ddi_filter = TopicsFilterddi(request.GET, queryset=ddi)
            demdeInf = ddi_filter.qs.filter(area_tem=id_area_tematica)

        ## opc 2 y opc 3 es diferente de vacia
        elif id_requerimiento == "" and id_area_tematica != "" and  id_tema != "":

            ddi = demandaInfor.objects.filter(nombre_est=5)
            ddi_filter = TopicsFilterddi(request.GET, queryset=ddi)
            demdeInf = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema)

        #opcion 3 es diferente de vacia
        elif id_requerimiento == "" and id_area_tematica == "" and  id_tema != "":

            ddi = demandaInfor.objects.filter(nombre_est=5)
            ddi_filter = TopicsFilterddi(request.GET, queryset=ddi)
            demdeInf = ddi_filter.qs.filter(tema_prin=id_tema)

        else:
            demdeInf = []
        
        data = serializers.serialize('json', demdeInf,
                    fields=('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'codigo_ddi', 'nombre_est', 'pm_b_6', 'compl_dem_a'), use_natural_foreign_keys=True, use_natural_primary_keys=True)
        return HttpResponse(data, content_type='application/json')


## filtros por nombre para que funcione con paginación django modulo_cosulta_oe
class filterByNameDdi(TemplateView):
    def get(self, request, *args, **kwargs):
        
        nombre = request.GET.get('nombre_ddi') #opc 1
        #print("que estoy recibiendo", nombre)
        deman_inf = demandaInfor.objects.filter(nombre_est=5).filter(pm_b_1__icontains=nombre)
        data = serializers.serialize('json', deman_inf,
                    fields=('pm_b_1', 'area_tem', 'tema_prin', 'nombre_est', 'pm_b_6', 'compl_dem_a'), use_natural_foreign_keys=True, 
                    use_natural_primary_keys=True)
        return HttpResponse(data, content_type='application/json')


###### formulario crear ddi
@login_required
def createDDIView(request):

    user = request.user
    registered = False
    count_ddi = demandaInfor.objects.count()
    entidad_creadora = user.profile.entidad.pk
    entities = Entidades_oe.objects.filter(estado_id=1) #entidades publicadas
    count_entities = entities.count()
    createddi_form = CreateDDIForm(request.POST or None, request.FILES or None)   #form crear DDI
     
    ## campos de texto preguntas de selección multiple
    crearUtilInfEstText_form = CrearUtilizaInfEstTextForm(request.POST or None)
    crearUsPrincInfEstText_form = CrearUsPrincInfEstTextForm(request.POST or None)
    crearNormasText_form = CrearNormasTextForm(request.POST or None)
    crearInfSolRespRequerimientoText_form = CrearInfSolRespRequerimientoTextForm(request.POST or None)
    crearTipoRequerimientoText_form = TipoRequerimientoTextForm(request.POST or None)
    crearDesagregacionReqText_form = DesagregacionReqTextForm(request.POST or None)
    crearDesagregacionGeoReqGeograficaText_form = DesagregacionGeoReqGeograficaTextForm(request.POST or None)
    crearPeriodicidadDifusionText_form = PeriodicidadDifusionTextForm(request.POST or None)

    ## pregunta 11 sub preguntas 
    crearSolReqInformacion_form = CrearSolReqInformacionForm ## test
    new_subquestion_eleven = None
    ## end pregunta 11 sub preguntas

    ### novedad para cuando se crea una nueva demanda
    novedad_ddi_form = NovedadDDIForm
    #novedades = post_oe.novedadactualizacions.filter(active=True)
    new_novedad = None

    if request.method == 'GET':  ##formset
        formvariableReqset = listaVariablesFormset(queryset=listaVariables.objects.none(), prefix='variablerequerimiento')

    if request.method == "POST": 
        
        createddi_form = CreateDDIForm(request.POST, request.FILES)

        crearUtilInfEstText_form = CrearUtilizaInfEstTextForm(request.POST)
        crearUsPrincInfEstText_form = CrearUsPrincInfEstTextForm(request.POST)
        crearNormasText_form = CrearNormasTextForm(request.POST)
        crearInfSolRespRequerimientoText_form = CrearInfSolRespRequerimientoTextForm(request.POST)
        crearTipoRequerimientoText_form = TipoRequerimientoTextForm(request.POST)
        formvariableReqset = listaVariablesFormset(request.POST, prefix='variablerequerimiento')
        crearDesagregacionReqText_form = DesagregacionReqTextForm(request.POST)
        crearDesagregacionGeoReqGeograficaText_form = DesagregacionGeoReqGeograficaTextForm(request.POST)
        crearPeriodicidadDifusionText_form = PeriodicidadDifusionTextForm(request.POST)

        ## pregunta 11 sub preguntas 
        crearSolReqInformacion_form = CrearSolReqInformacionForm(data=request.POST) ## test
        ## end pregunta 11 sub preguntas    

        ## Novedades para cuando es creada
        novedad_ddi_form = NovedadDDIForm(data=request.POST)
    
        if createddi_form.is_valid() and  crearUtilInfEstText_form.is_valid() and crearUsPrincInfEstText_form.is_valid() and \
            crearNormasText_form.is_valid() and crearInfSolRespRequerimientoText_form.is_valid() and crearTipoRequerimientoText_form.is_valid() and \
                formvariableReqset.is_valid() and crearDesagregacionReqText_form.is_valid() and crearDesagregacionGeoReqGeograficaText_form.is_valid() and \
                    crearPeriodicidadDifusionText_form.is_valid() and novedad_ddi_form.is_valid() and crearSolReqInformacion_form.is_valid():
             
            instancia = createddi_form.save(commit=False)

            generate_cod = count_ddi + 870 ## generar nuevas DDI codigo a partir de 
            instancia.codigo_ddi = "D" + str(generate_cod)
           
            if 'pm_b_9_anexar' in request.FILES:
                instancia.pm_b_9_anexar = request.FILES['pm_b_9_anexar']

            if 'pm_d_1_anexos' in request.FILES:
                instancia.pm_d_1_anexos = request.FILES['pm_d_1_anexos']

            instancia.save()
            obj_id = instancia.id  ##obtener id del objeto que estoy creando

            addUtilInfEstText = crearUtilInfEstText_form.save(commit=False)
            addUtilInfEstText.ddi_id = obj_id
            addUtilInfEstText.save()

            addUsPrincInfEstText = crearUsPrincInfEstText_form.save(commit=False)
            addUsPrincInfEstText.ddi_id = obj_id
            addUsPrincInfEstText.save()

            addNormasText = crearNormasText_form.save(commit=False)
            addNormasText.ddi_id = obj_id
            addNormasText.save()

            addInfSolRespRequerimientoText = crearInfSolRespRequerimientoText_form.save(commit=False)
            addInfSolRespRequerimientoText.ddi_id = obj_id
            addInfSolRespRequerimientoText.save()

            addTipoRequerimientoText = crearTipoRequerimientoText_form.save(commit=False)
            addTipoRequerimientoText.ddi_id = obj_id
            addTipoRequerimientoText.save()

            ## formset pregunta 9
            for formvariableReq in formvariableReqset:
                formVariableRequerimiento = formvariableReq.save(commit=False)
                formVariableRequerimiento.ddi_id = obj_id
                formVariableRequerimiento.instancia = instancia
                if formVariableRequerimiento.lista_varia == "":
                    formVariableRequerimiento = formvariableReq.save(commit=False)
                else: 
                    formVariableRequerimiento.save()

            ## pregunta 11 sub preguntas 
            ## pregunta 11
            new_subquestion_eleven = crearSolReqInformacion_form.save(commit=False)
            new_subquestion_eleven.post_ddi_id = obj_id
            if new_subquestion_eleven.inc_var_cual == "" and new_subquestion_eleven.cam_preg_cual == "" and new_subquestion_eleven.am_des_tem_cual == "" and \
                    new_subquestion_eleven.am_des_geo_cual == "" and new_subquestion_eleven.dif_resul_cual == "" and new_subquestion_eleven.opc_aprov_cual == "" and \
                        new_subquestion_eleven.inc_varia_cual == "" and new_subquestion_eleven.camb_pregu_cual == "" and new_subquestion_eleven.otros_aprov_ra == "" and \
                            new_subquestion_eleven.nueva_oe == "" and new_subquestion_eleven.indi_cual == "" and new_subquestion_eleven.gen_nueva == "":

                new_subquestion_eleven = crearSolReqInformacion_form.save(commit=False)
            else:
                new_subquestion_eleven.save()
                crearSolReqInformacion_form.save_m2m()

            ## end pregunta 11 sub preguntas    

            addDesagregacionReqText = crearDesagregacionReqText_form.save(commit=False)
            addDesagregacionReqText.ddi_id = obj_id
            addDesagregacionReqText.save()

            addDesagregacionGeoReqGeograficaText = crearDesagregacionGeoReqGeograficaText_form.save(commit=False)
            addDesagregacionGeoReqGeograficaText.ddi_id = obj_id
            addDesagregacionGeoReqGeograficaText.save()

            
            addPeriodicidadDifusionText = crearPeriodicidadDifusionText_form.save(commit=False)
            addPeriodicidadDifusionText.ddi_id = obj_id
            addPeriodicidadDifusionText.save()

            ##  novedades
            new_novedad = novedad_ddi_form.save(commit=False)
            new_novedad.post_ddi_id = obj_id
            new_novedad.name_nov = str(request.user)
            new_novedad.est_actualiz_id = 4
            new_novedad.descrip_novedad = "Se crea demanda de información."
            new_novedad.save()

            createddi_form.save_m2m()

            registered = True
            print("save")
        else:    
            print("no guardo", createddi_form.errors.as_data())
            messages.error(request, "Las respuestas NO se han guardado verifique si tiene errores")
        
    return render(request, 'ddi/admin/create_ddi.html', { 'createddi_form': createddi_form,
        'crearUtilInfEstText_form': crearUtilInfEstText_form, 'crearUsPrincInfEstText_form': crearUsPrincInfEstText_form,
        'crearNormasText_form': crearNormasText_form, 'crearInfSolRespRequerimientoText_form': crearInfSolRespRequerimientoText_form,
        'crearTipoRequerimientoText_form': crearTipoRequerimientoText_form, 'formvariableReqset': formvariableReqset,
        'crearDesagregacionReqText_form': crearDesagregacionReqText_form, 'crearDesagregacionGeoReqGeograficaText_form': crearDesagregacionGeoReqGeograficaText_form,
        'crearPeriodicidadDifusionText_form': crearPeriodicidadDifusionText_form,  'novedad_ddi_form':  novedad_ddi_form,
        'registered': registered, 'count_entities': count_entities,
        'crearSolReqInformacion_form': crearSolReqInformacion_form
        
    })

## correo para notificar a la fuente cuando se cambia el estado a devuelto
def sendEmailStatusChangeDDI(subject, recipients, body):    
    send_mail(subject, body, 'sen@dane.gov.co', recipients,  fail_silently = False)  

def emailNotificationDDIStateView(request, ddi, listaUsuarios):
    
    subject = 'Cambio de estado en formulario de Demandas de Información '+ str(ddi).replace('\n', '').replace('\r', '')

    recipients = listaUsuarios
      
    body = 'SISTEMA DE IDENTIFICACIÓN Y CARACTERIZACIÓN DE OFERTA Y DEMANDA ESTADÍSTICA DEL SEN\n\n' \
            'La demanda de información: '+ str(ddi) + \
			    '\n\n Ha sido DEVUELTO(A)\n' \
                    '\n Por favor ingrese al aplicativo, revise los comentarios generados por el temático y realice los ajustes correspondientes.'

    return subject, recipients, body

def sendEmailNotification(subject, recipients, body):    
    send_mail(subject, body, 'sen@dane.gov.co', recipients,  fail_silently = False)  


## envio de notificaciones al editar DDI
def createEmailEditDDI(request, ddi, temaRes, fieldsEdited_ddi, fieldsFormset1):

    subject = 'Cambios en formulario de demandas de información '+ str(ddi).replace('\n', '').replace('\r', '')
   
    if temaRes == 2 or temaRes == 7 or temaRes == 19 or temaRes == 28: 

        responsable_ddi = 'sdazag@dane.gov.co' #sdazag
           
    elif temaRes == 5 or temaRes == 14 or temaRes ==  15 or temaRes ==  16 or temaRes ==  17 or temaRes ==  18: 

        responsable_ddi = 'aobandor@dane.gov.co' # aobandor
          
    elif temaRes == 9 or temaRes == 29 or temaRes == 30:
        
        responsable_ddi = 'dvlizarazog@dane.gov.co' # dvlizarazog
    
    elif temaRes == 10 or temaRes == 6 or temaRes == 1:
        
        responsable_ddi = 'gavargasr@dane.gov.co' # gavargasr  
        
    #elif temaRes == 11 or temaRes == 12 or temaRes == 25 or temaRes == 3: 
        
    #    responsable_ddi = 'lhsanchezz@dane.gov.co' # lhsanchezz

    elif temaRes == 11 or temaRes == 12 or temaRes == 25: 
        
        responsable_ddi = 'pczambranog@dane.gov.co' # pczambranog
         
    elif temaRes == 13 or temaRes == 24 or temaRes == 26 or temaRes == 27: 
        
        responsable_ddi = 'mjpradag@dane.gov.co'  # 7 mjpradag

    elif temaRes == 11 or temaRes == 12 or temaRes == 25 or temaRes == 3: 
        responsable_ddi = 'rctrianaa@dane.gov.co' # rctrianaa
          
    elif temaRes == 4 or temaRes == 8 or temaRes == 20: 

        responsable_ddi = 'eeguayazans@dane.gov.co' # eeguayazans
      
    elif temaRes == 21 or temaRes == 22 or temaRes == 23:  
        
        responsable_ddi = 'mlbarretob@dane.gov.co' # mlbarretob
                 
    recipients = [ responsable_ddi ]
      
    body = 'SISTEMA DE IDENTIFICACIÓN Y CARACTERIZACIÓN DE OFERTA Y DEMANDA ESTADÍSTICA DEL SEN\n\n' \
            'La demanda de información: '+ str(ddi) + \
			    '\n\nha sido editado en los siguientes campos:\n' \
                   '\n' + str(fieldsEdited_ddi).replace('[','').replace(']','').replace(',', '\n').replace("'", "") + \
                    '\n' + str(fieldsFormset1).replace('[','').replace(']','').replace(',', '\n').replace("'", "")
                                                                        
    return subject, recipients, body


@login_required
def ddi_edit(request, pk):

    entities = Entidades_oe.objects.filter(estado_id=1)
    count_entities = entities.count() 
    ##___***id para instancia de editar los formularios
    post_ddi = get_object_or_404(demandaInfor, pk=pk)
    ddi_id =  demandaInfor.objects.get(pk=pk)  
    
    user = request.user
    roleUser = user.profile.role.id
    #entidad_creadora = user.profile.entidad.pk

    ## traer listado de usuarios responsable de ddi

    


    user_email = []

    if Profile.objects.filter(entidad=ddi_id.entidad_sol_id).exists():
        user_responsable = Profile.objects.filter(entidad=ddi_id.entidad_sol_id)
        ##print("entidad", user_responsable)
        for use in list(chain(user_responsable)):
            user_email.append(str(use.user.email))
        ##print("___________",user_email)
            
    
    ## traer listado de usuarios responsable de la ddi
    
    
    #entidad_responsable = ddi_id.entidad.pk
    #test = ddi_id.entidad_sol
    #print(test)
    
    ### id's para traer información de los campos de texto
    q_mb_3_id = UtilizaInfEstText.objects.get(ddi_id=ddi_id.pk)
    q_mb_4_id = UsuariosPrincInfEstText.objects.get(ddi_id=ddi_id.pk)
    q_mb_5_id = NormasText.objects.get(ddi_id=ddi_id.pk)
    q_mb_6_id = InfSolRespRequerimientoText.objects.get(ddi_id=ddi_id.pk)
    q_mb_8_id = TipoRequerimientoText.objects.get(ddi_id=ddi_id.pk)
    
    q_mb_12_id = DesagregacionReqText.objects.get(ddi_id=ddi_id.pk)
    q_mb_13_id = DesagregacionGeoReqGeograficaText.objects.get(ddi_id=ddi_id.pk)
    q_mb_14_id = PeriodicidadDifusionText.objects.get(ddi_id=ddi_id.pk)
    ### end id's para traer información de los campos de texto

    editddi_form = EditDDIForm(instance=ddi_id)
    editUtilizaInfEstText_form = EditUtilizaInfEstTextForm(instance=q_mb_3_id) #pregunta 3
    editUsPrincInfEstText_form = EditUsPrincInfEstTextForm(instance=q_mb_4_id) #pregunta 4
    editNormasText_form = EditNormasTextForm(instance=q_mb_5_id) #pregunta 5
    editInfSolRespRequerimientoText_form = EditInfSolRespRequerimientoTextForm(instance=q_mb_6_id)# pregunta 6
    editTipoRequerimientoText_form = EditTipoRequerimientoTextForm(instance=q_mb_8_id)# pregunta 8
    editDesagregacionReqText_form = EditDesagregacionReqTextForm(instance=q_mb_12_id)# pregunta 12
    editDesaReqGeogText_form = EditDesaReqGeogTextForm(instance=q_mb_13_id)# pregunta 13
    editPerDifusionText_form = EditPerDifusionTextForm(instance=q_mb_14_id)# pregunta 14
    
    ## pregunta 11
    crearSolReqInformacion_form = CrearSolReqInformacionForm() ##crear
    pregOnce = post_ddi.pregOnce.filter(active=True)
    new_answer_11 = None

    ## comentarios
    comment_ddi_form = ComentariosddiForm()
    comments_ddi = post_ddi.comentarios_ddi.filter(active=True)
    new_comment_ddi = None

    ## novedades
    novedad_ddi_form = NovedadDDIForm()
    novedades_ddi = post_ddi.novedadactddi.filter(active=True)
    new_novedad_ddi = None

    ## Critica
    critica_ddi_form = CriticaddiForm()
    criticas_ddi = post_ddi.critica_ddi.filter(active=True)
    new_critica_ddi = None

    if request.method == 'GET':
        editformlisVar_form  = listaVariablesFormset(queryset=listaVariables.objects.filter(ddi_id=ddi_id.pk), prefix='listVarddi')# pregunta 9 formset
       
    if request.method == "POST":

        editddi_form = EditDDIForm(request.POST, request.FILES, instance=ddi_id)
        editUtilizaInfEstText_form = EditUtilizaInfEstTextForm(request.POST, instance=q_mb_3_id) #pregunta 3
        editUsPrincInfEstText_form = EditUsPrincInfEstTextForm(request.POST, instance=q_mb_4_id) #pregunta 4
        editNormasText_form = EditNormasTextForm(request.POST, instance=q_mb_5_id) #pregunta 5
        editInfSolRespRequerimientoText_form = EditInfSolRespRequerimientoTextForm(request.POST, instance=q_mb_6_id)# pregunta 6
        editTipoRequerimientoText_form = EditTipoRequerimientoTextForm(request.POST, instance=q_mb_8_id)# pregunta 8
        editformlisVar_form  = listaVariablesFormset(request.POST, prefix='listVarddi')# pregunta 9 formset
        editDesagregacionReqText_form = EditDesagregacionReqTextForm(request.POST, instance=q_mb_12_id)# pregunta 12
        editDesaReqGeogText_form = EditDesaReqGeogTextForm(request.POST, instance=q_mb_13_id)# pregunta 13
        editPerDifusionText_form = EditPerDifusionTextForm(request.POST, instance=q_mb_14_id)# pregunta 14

        #editSolReqInformacion_form = EditSolReqInformacionForm(data=request.POST, instance=ddi_id) ## pregunta 11 editar
        crearSolReqInformacion_form = CrearSolReqInformacionForm(data=request.POST) ##  pregunta 11

        ## comentarios
        comment_ddi_form = ComentariosddiForm(data=request.POST)
        
        ##Novedades
        novedad_ddi_form = NovedadDDIForm(data=request.POST)

        ##Critica
        critica_ddi_form = CriticaddiForm(data=request.POST)

        ## detectar cambios
        fieldsEdited_ddi = []
        fieldsFormset1 = []

        if editddi_form.has_changed() or editformlisVar_form.has_changed():
            #print("The following fields changed: %s" % ", ".join(editddi_form.changed_data))
            #print("contador de cambios", editddi_form.changed_data)
              
            for index, item in enumerate(editddi_form.changed_data):

                editddi_form.fields[item].widget.attrs['title']
                questionParamTitle = editddi_form.fields[item].widget.attrs['title']
                fieldsEdited_ddi.append(questionParamTitle)# array que almacena lista de campos editados
            if editformlisVar_form.has_changed() == True:
                fieldsFormset1 = ["Módulo B Pregunta 9: ¿Qué variables necesita para suplir el requerimiento?"]

        if editddi_form.is_valid() and editUtilizaInfEstText_form.is_valid() and editUsPrincInfEstText_form.is_valid() and \
            editNormasText_form.is_valid() and editInfSolRespRequerimientoText_form.is_valid() and editTipoRequerimientoText_form.is_valid() and \
                editformlisVar_form.is_valid() and editDesagregacionReqText_form.is_valid() and editDesaReqGeogText_form.is_valid() and \
                    editPerDifusionText_form.is_valid() and crearSolReqInformacion_form.is_valid()  and \
                        novedad_ddi_form.is_valid() and comment_ddi_form.is_valid() and critica_ddi_form.is_valid():            
            
            edit_ddi_form = editddi_form.save(commit=False)

            ## Si rol fuente  --> edita ddi publicada
            if roleUser == 3 and edit_ddi_form.nombre_est_id == 5:
                edit_ddi_form.nombre_est_id = 2

            elif roleUser == 3 and edit_ddi_form.nombre_est_id == 4:
                edit_ddi_form.nombre_est_id = 2
            ## end validaciones según el rol para cambio de estado

            temaRes = edit_ddi_form.tema_prin_id

            #envio de correo para notificar cuando el estado cambia a devuelto ****
        
            #if edit_ddi_form.nombre_est_id == 3 and edit_ddi_form.entidad_sol_id != 1:
            #    subject, recipients, body = emailNotificationDDIStateView(user, post_ddi, user_email) 
            #    sendEmailStatusChangeDDI(subject, recipients, body)
            #end envio de correo para notificar cuando el estado cambia a devuelto

            if 'pm_b_9_anexar' in request.FILES:
                edit_ddi_form.pm_b_9_anexar = request.FILES['pm_b_9_anexar']

            if 'pm_d_1_anexos' in request.FILES:
                edit_ddi_form.pm_d_1_anexos = request.FILES['pm_d_1_anexos']

            
            ## send correo Notificaciones ***
            ##subject, recipients, body = createEmailEditDDI(user, post_ddi, temaRes, fieldsEdited_ddi, fieldsFormset1 )
            ##sendEmailNotification(subject, recipients, body)
            ## end send correo Notificaciones

            edit_ddi_form.save()
            obj_id = edit_ddi_form.id

            ## Modulo B

            editUtilizaInfEstText_ddi_form = editUtilizaInfEstText_form.save(commit=False)
            editUtilizaInfEstText_ddi_form.save()

            editUsPrincInfEstText_ddi_form = editUsPrincInfEstText_form.save(commit=False)
            editUsPrincInfEstText_ddi_form.save()
            
            editNormasText_ddi_form = editNormasText_form.save(commit=False)
            editNormasText_ddi_form.save()
            
            editInfSolRespRequerimientoText_ddi_form = editInfSolRespRequerimientoText_form.save(commit=False)
            editInfSolRespRequerimientoText_ddi_form.save()
            
            editTipoRequerimientoText_ddi_form = editTipoRequerimientoText_form.save(commit=False)
            editTipoRequerimientoText_ddi_form.save()

            editDesagregacionReqText_ddi_form = editDesagregacionReqText_form.save(commit=False)
            editDesagregacionReqText_ddi_form.save()
            
            editDesaReqGeogText_ddi_form = editDesaReqGeogText_form.save(commit=False)
            editDesaReqGeogText_ddi_form.save()

            editPerDifusionText_ddi_form = editPerDifusionText_form.save(commit=False)
            editPerDifusionText_ddi_form.save()

            #*** formsets         
            formListaVar = editformlisVar_form.save(commit=False)
           
            if editformlisVar_form.deleted_forms:
                for obj in editformlisVar_form.deleted_objects:
                    obj.delete() 
            else:
                for formListaVar in editformlisVar_form:
                # so that `ooee` instance can be attached.
                    formListaVariableddi = formListaVar.save(commit=False)
                    formListaVariableddi.ddi_id = obj_id
                    formListaVariableddi.editddi_form = editddi_form
                    if formListaVariableddi.lista_varia == "":
                        formListaVariableddi = formListaVar.save(commit=False)
                    else:
                        # Save listvar to the database
                        formListaVariableddi.save()

            ## pregunta 11
            new_answer_11 = crearSolReqInformacion_form.save(commit=False)
            new_answer_11.post_ddi = post_ddi
            if new_answer_11.inc_var_cual == "" and new_answer_11.cam_preg_cual == "" and new_answer_11.am_des_tem_cual == "" and \
                    new_answer_11.am_des_geo_cual == "" and new_answer_11.dif_resul_cual == "" and new_answer_11.opc_aprov_cual == "" and \
                        new_answer_11.inc_varia_cual == "" and new_answer_11.camb_pregu_cual == "" and new_answer_11.otros_aprov_ra == "" and \
                            new_answer_11.nueva_oe == "" and new_answer_11.indi_cual == "" and new_answer_11.gen_nueva == "":

                new_answer_11 = crearSolReqInformacion_form.save(commit=False)
            else:
                new_answer_11.save()
                crearSolReqInformacion_form.save_m2m()

            ## comentarios
            new_comment_ddi = comment_ddi_form.save(commit=False)
            new_comment_ddi.name = str(request.user)
            new_comment_ddi.post_ddi = post_ddi
            if new_comment_ddi.body == "":
                new_comment_ddi = comment_ddi_form.save(commit=False)
            else:
            # Save the comment to the database
                new_comment_ddi.save()

            ## novedades -modulo de actualización
            new_novedad_ddi = novedad_ddi_form.save(commit=False)
            new_novedad_ddi.name_nov = str(request.user)
            new_novedad_ddi.post_ddi = post_ddi
            if new_novedad_ddi.descrip_novedad == "":
                new_novedad_ddi = novedad_ddi_form.save(commit=False)
            else:
                new_novedad_ddi.save()

            ## Criticas
            new_critica_ddi = critica_ddi_form.save(commit=False)
            new_critica_ddi.name_cri = str(request.user)
            new_critica_ddi.post_ddi = post_ddi
            if new_critica_ddi.descrip_critica == "":
                new_critica_ddi = critica_ddi_form.save(commit=False)
            else:
                new_critica_ddi.save()

            editddi_form.save_m2m() ## metodo para guardar relaciones manyTomany
           
            messages.success(request, 'Las respuestas se han guardado con éxito')
            
            return HttpResponseRedirect(request.path_info)
            
        else:
            print("fom prin",editddi_form.errors.as_data())
            print("preg_11",crearSolReqInformacion_form.errors.as_data())
            messages.error(request, "Las respuestas NO se han guardado verifique si tiene errores")
            

    return render(request, 'ddi/admin/edit_ddi.html', {'editddi_form': editddi_form, 'count_entities': count_entities,
    'editUtilizaInfEstText_form': editUtilizaInfEstText_form, 'editUsPrincInfEstText_form': editUsPrincInfEstText_form,
    'editNormasText_form': editNormasText_form, 'editInfSolRespRequerimientoText_form': editInfSolRespRequerimientoText_form,
    'editTipoRequerimientoText_form': editTipoRequerimientoText_form, 'editformlisVar_form': editformlisVar_form , 
    'editDesagregacionReqText_form': editDesagregacionReqText_form, 'editDesaReqGeogText_form': editDesaReqGeogText_form,
    'editPerDifusionText_form': editPerDifusionText_form,
    'pregOnce': pregOnce, 'crearSolReqInformacion_form': crearSolReqInformacion_form, 
    'novedad_ddi_form': novedad_ddi_form, 'new_novedad_ddi': new_novedad_ddi, 'novedades_ddi': novedades_ddi,
    'comment_ddi_form': comment_ddi_form, 'new_comment_ddi': new_comment_ddi, 'comments_ddi': comments_ddi,
    'critica_ddi_form': critica_ddi_form, 'new_critica_ddi': new_critica_ddi, 'criticas_ddi': criticas_ddi, 'user': user
    })


def EditSolReqInformacionView(request, pk):
    entities = Entidades_oe.objects.filter(estado_id=1)
    count_entities = entities.count() 
    ##___***id para instancia de editar los formularios
    post_ddi = get_object_or_404(SolReqInformacion, pk=pk)
    ddi_id =  SolReqInformacion.objects.get(pk=pk)

    editSolReqInformacion_form = EditSolReqInformacionForm(instance=ddi_id) ## editar pregunta 11
    
    if request.method == "POST":
        editSolReqInformacion_form = EditSolReqInformacionForm(data=request.POST, instance=ddi_id) ## pregunta 11 editar
        
        if editSolReqInformacion_form.is_valid():
        
            editSolReqInformacion_ddi_form = editSolReqInformacion_form.save(commit=False)
            editSolReqInformacion_ddi_form.save()

            editSolReqInformacion_form.save_m2m() ## metodo para guardar relaciones manyTomany
           
            messages.success(request, 'Las respuestas se han guardado con éxito')
            
            return HttpResponseRedirect(request.path_info)
            
        else:
            #print("fom prin", editSolReqInformacion_ddi_form.as_data())
            
            messages.error(request, "Las respuestas NO se han guardado verifique si tiene errores")
            
    return render(request, 'ddi/admin/edit_preguntaOnce.html', {'editSolReqInformacion_form': editSolReqInformacion_form, 
    'count_entities': count_entities})


### ficha técnica
def ficha_tecnica_ddi(request, pk):
    
    entities = Entidades_oe.objects.filter(estado_id=1) #entidades publicadas
    count_entities = entities.count()
    
    ddi = get_object_or_404(demandaInfor, pk=pk)
    #print(ddi.pk)
    
    normasText_ft = NormasText.objects.get(ddi_id=ddi.pk)
    infSolRespRequerimientoText_ft = InfSolRespRequerimientoText.objects.get(ddi_id=ddi.pk)
    listaVariablesText_ft = listaVariables.objects.filter(ddi_id=ddi.pk)
    tipoRequerimientoText_ft = TipoRequerimientoText.objects.get(ddi_id=ddi.pk)
    desagregacionReqText_ft = DesagregacionReqText.objects.get(ddi_id=ddi.pk)
    desGeoReqGeograficaText_ft = DesagregacionGeoReqGeograficaText.objects.get(ddi_id=ddi.pk)
    perioDifusionText_ft = PeriodicidadDifusionText.objects.get(ddi_id=ddi.pk)

    for x in ddi.total_si_no_pmb7.all():
        preg7 = x.total_si_no_pmb7

    ##estado de la novedad
    if  NovedadActualizacionddi.objects.filter(post_ddi=ddi.pk).exists():
        estadoNovedad = NovedadActualizacionddi.objects.filter(post_ddi=ddi.pk)
        for stateNov in estadoNovedad:
            stateNov.est_actualiz
        if stateNov.est_actualiz != None:
            estadoNovDDI = stateNov.est_actualiz
            fechaEstado = stateNov.fecha_actualiz
        else:
            estadoNovDDI = ""
            fechaEstado = ""
    else:
        estadoNovDDI = ""
        fechaEstado = ""
       
    return render(request, 'ddi/consulta/ddi_ficha.html', {'ddi': ddi, 'count_entities': count_entities, 'preg7': preg7,
        'normasText_ft': normasText_ft, 'infSolRespRequerimientoText_ft': infSolRespRequerimientoText_ft, 
        'tipoRequerimientoText_ft': tipoRequerimientoText_ft, 
        'listaVariablesText_ft': listaVariablesText_ft, 'desagregacionReqText_ft': desagregacionReqText_ft,
        'desGeoReqGeograficaText_ft': desGeoReqGeograficaText_ft, 'perioDifusionText_ft': perioDifusionText_ft,
        'estadoNovDDI': estadoNovDDI, 'fechaEstado': fechaEstado})


@login_required
def allDDI(request):
    
    count_entities = Entidades_oe.objects.filter(estado_id=1).count()
    user = request.user
    
    if user.is_authenticated == True and str(user) == "sdazag": # 1 sdazag
        ddinfo = demandaInfor.objects.filter(Q (tema_prin_id=2) | Q (tema_prin_id=7) | Q (tema_prin_id=19) | Q (tema_prin_id=28)).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'nombre_est', 'pm_b_6', 'compl_dem_a').order_by('pm_b_1')
        ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)
        nameddiFilter = NameFilterDDI(request.GET, queryset=ddinfo)
    
        """ elif user.is_authenticated == True and str(user) == "aobandor": # 3 se cambia usuaaria a rol administrador

        ddinfo = demandaInfor.objects.filter(Q (tema_prin_id=5) | Q (tema_prin_id=14) | Q (tema_prin_id=15) | Q (tema_prin_id=16) | Q (tema_prin_id=17) | Q (tema_prin_id=18)).order_by('pm_b_1')
        ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo) """

    elif user.is_authenticated == True and str(user) == "dvlizarazog": # 4
        
        ddinfo = demandaInfor.objects.filter(Q (tema_prin_id=9) | Q (tema_prin_id=29) | Q (tema_prin_id=30)).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'nombre_est', 'pm_b_6', 'compl_dem_a').order_by('pm_b_1')
        ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)
        nameddiFilter = NameFilterDDI(request.GET, queryset=ddinfo)

    elif user.is_authenticated == True and str(user) == "ancardenasc": # 5frherreran  se cambia por gavargasr
        
        ddinfo = demandaInfor.objects.filter(tema_prin_id=10).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'nombre_est', 'pm_b_6', 'compl_dem_a').order_by('pm_b_1')
        ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)
        nameddiFilter = NameFilterDDI(request.GET, queryset=ddinfo)

    elif user.is_authenticated == True and str(user) == "lhsanchezz": # 6

        ddinfo = demandaInfor.objects.filter(Q (tema_prin_id=3) | Q (tema_prin_id=11) | Q (tema_prin_id=12) | Q (tema_prin_id=25)).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'nombre_est', 'pm_b_6', 'compl_dem_a').order_by('pm_b_1')
        ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)
        nameddiFilter = NameFilterDDI(request.GET, queryset=ddinfo)

    elif user.is_authenticated == True and str(user) == "pczambranog": # 6
        ddinfo = demandaInfor.objects.filter(Q (tema_prin_id=11) | Q (tema_prin_id=12) | Q (tema_prin_id=25)).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'nombre_est', 'pm_b_6', 'compl_dem_a').order_by('pm_b_1')
        ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)
        nameddiFilter = NameFilterDDI(request.GET, queryset=ddinfo)
        
    elif user.is_authenticated == True and str(user) == "mjpradag": # 7  mppulidor se cambia por mjpradag

        ddinfo = demandaInfor.objects.filter(Q (tema_prin_id=13) | Q (tema_prin_id=24) | Q (tema_prin_id=26) | Q (tema_prin_id=27)).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'nombre_est', 'pm_b_6', 'compl_dem_a').order_by('pm_b_1')
        ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)
        nameddiFilter = NameFilterDDI(request.GET, queryset=ddinfo)
       
    elif user.is_authenticated == True and str(user) == "mpriveras": # 8

        ddinfo = demandaInfor.objects.filter(Q (tema_prin_id=4) | Q (tema_prin_id=8) | Q (tema_prin_id=20)).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'nombre_est', 'pm_b_6', 'compl_dem_a').order_by('pm_b_1')
        ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)
        nameddiFilter = NameFilterDDI(request.GET, queryset=ddinfo)

    elif user.is_authenticated == True and str(user) == "eeguayazans": # 8 eeguayazans Edgar Eduardo Guayazan Sierra

        ddinfo = demandaInfor.objects.filter(Q (tema_prin_id=4) | Q (tema_prin_id=8) | Q (tema_prin_id=20)).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'nombre_est', 'pm_b_6', 'compl_dem_a').order_by('pm_b_1')
        ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)
        nameddiFilter = NameFilterDDI(request.GET, queryset=ddinfo)
        
    elif user.is_authenticated == True and str(user) == "mlbarretob": # 9

        ddinfo = demandaInfor.objects.filter(Q (tema_prin_id=21) | Q (tema_prin_id=22) | Q (tema_prin_id=23)).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'nombre_est', 'pm_b_6', 'compl_dem_a').order_by('pm_b_1')
        ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)
        nameddiFilter = NameFilterDDI(request.GET, queryset=ddinfo)

    elif user.is_authenticated == True and user.profile.role.id == 1 or user.profile.role.id == 5: ##administrador o rol revisor
            
        ddinfo = demandaInfor.objects.all().only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'nombre_est', 'pm_b_6', 'compl_dem_a').order_by('pm_b_1')
        ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)
        nameddiFilter = NameFilterDDI(request.GET, queryset=ddinfo)  

    else:
        
        ddinfo = demandaInfor.objects.filter(entidad_sol=user.profile.entidad.id).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'nombre_est', 'pm_b_6', 'compl_dem_a').exclude(nombre_est_id=6).order_by('pm_b_1')
        ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)
        nameddiFilter = NameFilterDDI(request.GET, queryset=ddinfo)
        
    page = request.GET.get('page', 1)
    paginator = Paginator(ddinfo, 10)
    results = ddinfo.count()

    try:
        demanda = paginator.page(page)
    except PageNotAnInteger:
        demanda = paginator.page(1)
    except EmptyPage:
        demanda = paginator.page(paginator.num_pages)

    index = demanda.number - 1
    max_index = len(paginator.page_range)
    start_index = index - 3 if index >= 3 else 0
    end_index = index + 3 if index <= max_index - 3 else max_index
    page_start = 1
    page_end = max_index
    page_range = paginator.page_range[start_index:end_index]
    page_number= int(page)
        
    
    return render(request, 'ddi/admin/all_ddi.html',  {'ddinfo': ddinfo, 'filter': ddi_filter, 'count_entities': count_entities,
    'demanda': demanda, 'page_range': page_range, 'page_end': page_end, 'page_start': page_start, 'nameddiFilter': nameddiFilter, 
    'results': results})

### función ajax para vista all_ddi por filtros
class FilterDDIAjaxView(TemplateView):
    def get(self, request, *args, **kwargs):
        
        id_area_tematica = request.GET.get('area_tematica_id') #opc 1
        id_tema = request.GET.get('tema_id') #opc 2
        id_fase = request.GET.get('fase_id') #opc 3
        id_requi = request.GET.get('requeri_id') # opc 4
        user = request.user
        
        ########## user sdazag ################
        if user.is_authenticated == True and str(user) == "sdazag": # 1 sdazag
            ddinfo =  demandaInfor.objects.filter(Q (tema_prin=2) | Q (tema_prin=7) | Q (tema_prin=19) | Q (tema_prin=28)).order_by('pm_b_1')
            ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)

            ## opc 1 es diferente de vacia
            if  id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica)

            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(nombre_est=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(pm_b_6=id_requi)

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema)  
    
            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).filter(nombre_est=id_fase)

            ## opc 1, opc 2, opc 3 es y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_requi != "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).filter(nombre_est=id_fase).filter(pm_b_6=id_requi)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(nombre_est=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(id_requi_id=id_requi)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema).filter(nombre_est=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema).filter(pm_b_6=id_requi)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and id_requi != "":
                demanda = ddi_filter.qs.filter(nombre_est=id_fase).filter(pm_b_6=id_requi)

        #######################  user aobandor se cambia usuario a administrador ###########################  
            """ elif user.is_authenticated == True and str(user) == "aobandor": # 3

            ddinfo =  demandaInfor.filter(Q (tema_prin=5) | Q (tema_prin=14) | Q (tema_prin=15) | Q (tema_prin=16) | Q (tema_prin=17) | Q (tema_prin=18)).order_by('pm_b_1')
            ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo) """
            
        ####################### user dvlizarazog #############################
        elif user.is_authenticated == True and str(user) == "dvlizarazog": # 4
            
            ddinfo = demandaInfor.objects.filter(Q (tema_prin=9) | Q (tema_prin=29) | Q (tema_prin=30)).order_by('pm_b_1')
            ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)

            ## opc 1 es diferente de vacia
            if  id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica)

            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(nombre_est=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(pm_b_6=id_requi)

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema)  
    
            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).filter(nombre_est=id_fase)

            ## opc 1, opc 2, opc 3 es y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_requi != "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).filter(nombre_est=id_fase).filter(pm_b_6=id_requi)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(nombre_est=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(pm_b_6=id_requi)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema).filter(nombre_est=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema).filter(pm_b_6=id_requi)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and id_requi != "":
                demanda = ddi_filter.qs.filter(nombre_est=id_fase).filter(pm_b_6=id_requi)

        #################### user fherreran ##########################
        elif user.is_authenticated == True and str(user) == "ancardenasc": # 5frherreran se cambia por gavargasr
            
            ddinfo = demandaInfor.objects.filter(tema_prin=10).order_by('pm_b_1')
            ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)

            ## opc 1 es diferente de vacia
            if  id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica)

            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(nombre_est=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(pm_b_6=id_requi)

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema)  
    
            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).filter(nombre_est=id_fase)

            ## opc 1, opc 2, opc 3 es y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_requi != "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).filter(nombre_est=id_fase).filter(pm_b_6=id_requi)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(nombre_est=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(pm_b_6=id_requi)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema).filter(nombre_est=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema).filter(pm_b_6=id_requi)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and id_requi != "":
                demanda = ddi_filter.qs.filter(nombre_est=id_fase).filter(pm_b_6=id_requi)

        ########################### user lhsanchezz se remplaza por  pczambranog ####################  
        elif user.is_authenticated == True and str(user) == "pczambranog": # 6
            
            ddinfo = demandaInfor.objects.filter(Q (tema_prin=11) | Q (tema_prin=12) | Q (tema_prin=25)).order_by('pm_b_1')
            ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)

            ## opc 1 es diferente de vacia
            if  id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica)

            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(nombre_est=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_requi != "":
               demanda = ddi_filter.qs.filter(pm_b_6=id_requi)

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and id_requi == "": 
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema)  
    
            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).filter(nombre_est=id_fase)

            ## opc 1, opc 2, opc 3 es y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_requi != "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).filter(nombre_est=id_fase).filter(pm_b_6=id_requi)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and id_requi == "":
               demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(nombre_est=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(pm_b_6=id_requi)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema).filter(nombre_est=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and id_requi != "":
               demanda = ddi_filter.qs.filter(tema_prin=id_tema).filter(pm_b_6=id_requi)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and id_requi != "":
               demanda = ddi_filter.qs.filter(nombre_est=id_fase).filter(pm_b_6=id_requi)

        ####################### user mjpradag ###########################3
        elif user.is_authenticated == True and str(user) == "mjpradag": # 7  mppulidor se cambia por mjpradag 

            ddinfo = demandaInfor.objects.filter(Q (tema_prin=13) | Q (tema_prin=24) | Q (tema_prin=26) | Q (tema_prin=27)).order_by('pm_b_1')
            ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)

            ## opc 1 es diferente de vacia
            if  id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_requi == "":
               demanda = ddi_filter.qs.filter(area_tem=id_area_tematica)

            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_requi == "":  
                demanda = ddi_filter.qs.filter(nombre_est=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(pm_b_6=id_requi)

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema)  
    
            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).filter(nombre_est=id_fase)

            ## opc 1, opc 2, opc 3 es y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_requi != "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).filter(nombre_est=id_fase).filter(pm_b_6=id_requi)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(nombre_est=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(pm_b_6=id_requi)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema).filter(nombre_est=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema).filter(pm_b_6=id_requi)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and id_requi != "":
                demanda = ddi_filter.qs.filter(nombre_est=id_fase).filter(pm_b_6=id_requi)

        #################### user mpriveras ########################
        elif user.is_authenticated == True and str(user) == "mpriveras": # 8

            ddinfo = demandaInfor.objects.filter(Q (tema_prin=4) | Q (tema_prin=8) | Q (tema_prin=20)).order_by('pm_b_1')
            ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)

            ## opc 1 es diferente de vacia
            if  id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica)

            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(nombre_est=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(pm_b_6=id_requi)

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema)  
    
            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).filter(nombre_est=id_fase)

            ## opc 1, opc 2, opc 3 es y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_requi != "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).filter(nombre_est=id_fase).filter(pm_b_6=id_requi)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(nombre_est=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(pm_b_6=id_requi)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema).filter(nombre_est=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema).filter(pm_b_6=id_requi)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and id_requi != "":
                demanda = ddi_filter.qs.filter(nombre_est=id_fase).filter(pm_b_6=id_requi)

        #################### user eeguayazans ########################
        elif user.is_authenticated == True and str(user) == "eeguayazans": # 8 eeguayazans Edgar Eduardo Guayazan Sierra

            ddinfo = demandaInfor.objects.filter(Q (tema_prin=4) | Q (tema_prin=8) | Q (tema_prin=20)).order_by('pm_b_1')
            ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)

            ## opc 1 es diferente de vacia
            if  id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica)

            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(nombre_est=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(pm_b_6=id_requi)

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema)  
    
            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).filter(nombre_est=id_fase)

            ## opc 1, opc 2, opc 3 es y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_requi != "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).filter(nombre_est=id_fase).filter(pm_b_6=id_requi)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(nombre_est=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(pm_b_6=id_requi)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema).filter(nombre_est=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema).filter(pm_b_6=id_requi)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and id_requi != "":
                demanda = ddi_filter.qs.filter(nombre_est=id_fase).filter(pm_b_6=id_requi)

        ############################user mlbarretob ######################
        elif user.is_authenticated == True and str(user) == "mlbarretob": # 9

            ddinfo = demandaInfor.objects.filter(Q (tema_prin=21) | Q (tema_prin=22) | Q (tema_prin=23)).order_by('pm_b_1')
            ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)

            ## opc 1 es diferente de vacia
            if  id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica)

            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(nombre_est=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(pm_b_6=id_requi)

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema)  
    
            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).filter(nombre_est=id_fase)

            ## opc 1, opc 2, opc 3 es y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_requi != "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).filter(nombre_est=id_fase).filter(pm_b_6=id_requi)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(nombre_est=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(pm_b_6=id_requi)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema).filter(nombre_est=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema).filter(pm_b_6=id_requi)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and id_requi != "":
                demanda = ddi_filter.qs.filter(nombre_est=id_fase).filter(pm_b_6=id_requi)

        ############################ Rol administrador y revisor #######################
        elif user.is_authenticated == True and user.profile.role.id == 1 or user.profile.role.id == 5: 
            ddinfo = demandaInfor.objects.all().order_by('pm_b_1')
            ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)

            ## opc 1 es diferente de vacia
            if  id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica)

            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema)   
                
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(nombre_est=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(pm_b_6=id_requi)

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema)  
    
            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).filter(nombre_est=id_fase)

            ## opc 1, opc 2, opc 3 es y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_requi != "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).filter(nombre_est=id_fase).filter(pm_b_6=id_requi)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(nombre_est=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(pm_b_6=id_requi)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema).filter(nombre_est=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema).filter(pm_b_6=id_requi)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and id_requi != "":
                demanda = ddi_filter.qs.filter(nombre_est=id_fase).filter(pm_b_6=id_requi)

         ###########################  Rol fuente ###############################
        else:
            ddinfo = demandaInfor.objects.filter(entidad_sol=user.profile.entidad.id).exclude(nombre_est=6).order_by('pm_b_1')
            ddi_filter = TopicsFilterddi(request.GET, queryset=ddinfo)
            ## opc 1 es diferente de vacia
            if  id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).exclude(nombre_est=6)

            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema).exclude(nombre_est=6)  
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(nombre_est=id_fase).exclude(nombre_est=6)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(pm_b_6=id_requi).exclude(nombre_est_id=6)

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).exclude(nombre_est=6)  
    
            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).filter(nombre_est=id_fase).exclude(nombre_est=6)

            ## opc 1, opc 2, opc 3 es y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and id_requi != "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(tema_prin=id_tema).filter(nombre_est=id_fase).filter(pm_b_6=id_requi).exclude(nombre_est=6)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(nombre_est=id_fase).exclude(nombre_est=6)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(area_tem=id_area_tematica).filter(pm_b_6=id_requi).exclude(nombre_est=6)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and id_requi == "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema).filter(nombre_est=id_fase).exclude(nombre_est=6)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and id_requi != "":
                demanda = ddi_filter.qs.filter(tema_prin=id_tema).filter(pm_b_6=id_requi).exclude(nombre_est=6)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and id_requi != "":
                demanda = ddi_filter.qs.filter(nombre_est=id_fase).filter(pm_b_6=id_requi).exclude(nombre_est=6)

        data = serializers.serialize('json', demanda,
                fields=('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'codigo_ddi', 'nombre_est', 'pm_b_6', 'compl_dem_a'), use_natural_foreign_keys=True, use_natural_primary_keys=True)
        return HttpResponse(data, content_type='application/json')

## filtros por nombre para que funcione con paginación django all_ooee
class FilterNameAdminDDI(TemplateView):
    def get(self, request, *args, **kwargs):
        
        nombre = request.GET.get('nombre_dda') 
        user = request.user

        ### usuarios planificación
    
        if user.is_authenticated == True and str(user) == "sdazag": # 1 sdazag
            demanda = demandaInfor.objects.filter(Q (tema_prin_id=2) | Q (tema_prin_id=7) | Q (tema_prin_id=19) | Q (tema_prin_id=28)).filter(Q (pm_b_1__icontains=nombre)).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'codigo_ddi', 'nombre_est', 'pm_b_6', 'compl_dem_a').order_by('pm_b_1')
            nameddiFilter = NameFilterDDI(request.GET, queryset=demanda)

        elif user.is_authenticated == True and str(user) == "dvlizarazog": # 4
        
            demanda = demandaInfor.objects.filter(Q (tema_prin_id=9) | Q (tema_prin_id=29) | Q (tema_prin_id=30)).filter(pm_b_1__icontains=nombre).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'codigo_ddi', 'nombre_est', 'pm_b_6', 'compl_dem_a').order_by('pm_b_1')
            nameddiFilter = NameFilterDDI(request.GET, queryset=demanda)

        elif user.is_authenticated == True and str(user) == "ancardenasc": # 5frherreran se cambia por gavargasr
        
            demanda = demandaInfor.objects.filter(tema_prin_id=10).filter(pm_b_1__icontains=nombre).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'codigo_ddi', 'nombre_est', 'pm_b_6', 'compl_dem_a').order_by('pm_b_1')
            nameddiFilter = NameFilterDDI(request.GET, queryset=demanda)

        elif user.is_authenticated == True and str(user) == "lhsanchezz": # 6 no continua con contrato
        
            demanda = demandaInfor.objects.filter(Q (tema_prin_id=3) | Q (tema_prin_id=11) | Q (tema_prin_id=12) | Q (tema_prin_id=25)).filter(pm_b_1__icontains=nombre).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'codigo_ddi', 'nombre_est', 'pm_b_6', 'compl_dem_a').order_by('pm_b_1')
            nameddiFilter = NameFilterDDI(request.GET, queryset=demanda)

        elif user.is_authenticated == True and str(user) == "pczambranog": # 6
        
            demanda = demandaInfor.objects.filter(Q (tema_prin_id=11) | Q (tema_prin_id=12) | Q (tema_prin_id=25)).filter(pm_b_1__icontains=nombre).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'codigo_ddi', 'nombre_est', 'pm_b_6', 'compl_dem_a').order_by('pm_b_1')
            nameddiFilter = NameFilterDDI(request.GET, queryset=demanda)
        
        elif user.is_authenticated == True and str(user) == "mjpradag": # 7 mppulidor se cambia por mjpradag 

            demanda = demandaInfor.objects.filter(Q (tema_prin_id=13) | Q (tema_prin_id=24) | Q (tema_prin_id=26) | Q (tema_prin_id=27)).filter(pm_b_1__icontains=nombre).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'codigo_ddi', 'nombre_est', 'pm_b_6', 'compl_dem_a').order_by('pm_b_1')
            nameddiFilter = NameFilterDDI(request.GET, queryset=demanda)
       
        elif user.is_authenticated == True and str(user) == "mpriveras": # 8

            demanda = demandaInfor.objects.filter(Q (tema_prin_id=4) | Q (tema_prin_id=8) | Q (tema_prin_id=20)).filter(pm_b_1__icontains=nombre).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'codigo_ddi', 'nombre_est', 'pm_b_6', 'compl_dem_a').order_by('pm_b_1')
            nameddiFilter = NameFilterDDI(request.GET, queryset=demanda)

        elif user.is_authenticated == True and str(user) == "eeguayazans": # 8 eeguayazans Edgar Eduardo Guayazan Sierra

            demanda = demandaInfor.objects.filter(Q (tema_prin_id=4) | Q (tema_prin_id=8) | Q (tema_prin_id=20)).filter(pm_b_1__icontains=nombre).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'codigo_ddi', 'nombre_est', 'pm_b_6', 'compl_dem_a').order_by('pm_b_1')
            nameddiFilter = NameFilterDDI(request.GET, queryset=demanda)
        
        elif user.is_authenticated == True and str(user) == "mlbarretob": # 9

            demanda = demandaInfor.objects.filter(Q (tema_prin_id=21) | Q (tema_prin_id=22) | Q (tema_prin_id=23)).filter(pm_b_1__icontains=nombre).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'codigo_ddi', 'nombre_est', 'pm_b_6', 'compl_dem_a').order_by('pm_b_1')
            nameddiFilter = NameFilterDDI(request.GET, queryset=demanda)
        
    ### end usuarios planificación
    
        elif user.is_authenticated == True and user.profile.role.id == 1 or user.profile.role.id == 5 or user.profile.role.id == 6: ##administrador o rol revisor o evaluador de calidad
            
            demanda = demandaInfor.objects.filter(pm_b_1__icontains=nombre).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'codigo_ddi', 'nombre_est', 'pm_b_6', 'compl_dem_a').order_by('pm_b_1')
            nameddiFilter = NameFilterDDI(request.GET, queryset=demanda)   

        else:
            demanda = demandaInfor.objects.filter(Q (entidad_sol_id=user.profile.entidad.id)).filter(Q (pm_b_1__icontains=nombre)).only('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'codigo_ddi', 'nombre_est', 'pm_b_6', 'compl_dem_a').exclude(nombre_est_id=6).order_by('pm_b_1')
            nameddiFilter = NameFilterDDI(request.GET, queryset=demanda)

        data = serializers.serialize('json', demanda,
                    fields=('pm_b_1', 'entidad_sol', 'area_tem', 'tema_prin', 'codigo_ddi', 'nombre_est', 'pm_b_6', 'compl_dem_a'), use_natural_foreign_keys=True, use_natural_primary_keys=True)
        return HttpResponse(data, content_type='application/json')


@login_required
def crearConsumidorInfoView(request):
    count_entities = Entidades_oe.objects.count()
    registered = False
    create_cons_form = CrearConsumidorInfoForm(request.POST or None)
    if request.method == "POST": 
        create_cons_form = CrearConsumidorInfoForm(request.POST)
        if create_cons_form.is_valid():
            instancia = create_cons_form.save(commit=False)
            instancia.save()
            registered = True
            print("save")
        else:        
            messages.error(request, "Las respuestas NO se han guardado verifique si tiene errores")
        
    return render(request, 'ddi/admin/crear_consumidorInfo.html', {'create_cons_form': create_cons_form, 
    'count_entities': count_entities, 'registered': registered  })


@login_required
def EditarConsumidorInfoView(request, pk):

    count_entities = Entidades_oe.objects.count()

    consumidor_info = consumidores_info.objects.get(pk=pk)
    edit_cons_form = EditarConsumidorInfoForm(instance=consumidor_info)
    if request.method == "POST":
        edit_cons_form = EditarConsumidorInfoForm(request.POST, instance=consumidor_info)
        if edit_cons_form.is_valid():
            editCons_form = edit_cons_form.save(commit=False)
            editCons_form.save()
            messages.success(request, 'Las respuestas se han guardado con éxito')
            
            return HttpResponseRedirect(request.path_info)
        else:
            messages.error(request, "Las respuestas NO se han guardado verifique si tiene errores")
    return render(request, 'ddi/admin/editar_consumidorInfo.html', {'edit_cons_form': edit_cons_form, 'count_entities': count_entities,
     'consumidor_info': consumidor_info})


@login_required
def allConsumidorInfoView(request):

    count_entities = Entidades_oe.objects.count()
    consoInfo = consumidores_info.objects.filter(estado_id=5)
    total_consInfo = consoInfo.count()

    return render(request, 'ddi/admin/all_consumidorInfo.html', { 'count_entities': count_entities,
        'consoInfo': consoInfo, 'total_consInfo': total_consInfo
    })



### reporte inventarios ddi completo
@method_decorator(login_required, name='dispatch')
class reporteInventariosDDI(View):
    def get(self, request):
        
        # Get some data to write to the spreadsheet.
        demandas = demandaInfor.objects.all()

        # Create an in-memory output file for the new workbook.
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet(name="Inventario de demandas")
        
        # Get some data to write to the spreadsheet.
        row = 4
        col = 0
        ## altura de celda
        worksheet.set_row(0, 40)
        worksheet.set_row(1, 30)
        worksheet.set_row(2, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet.set_column('A:A', 30)
        worksheet.set_column('B:B', 30)
        worksheet.set_column('C:C', 30)
        worksheet.set_column('D:D', 30)
        worksheet.set_column('E:E', 30)
        worksheet.set_column('F:F', 30)
        worksheet.set_column('G:G', 40)
        worksheet.set_column('H:H', 30)
        worksheet.set_column('I:I', 30)
        worksheet.set_column('J:J', 30)
        worksheet.set_column('K:K', 30)
        worksheet.set_column('L:L', 30)
        worksheet.set_column('M:M', 30)
        worksheet.set_column('N:N', 30)
        worksheet.set_column('O:O', 30)
        worksheet.set_column('P:P', 30)
        worksheet.set_column('Q:Q', 30)
        worksheet.set_column('R:R', 30)
        worksheet.set_column('S:S', 30)
        worksheet.set_column('T:T', 30)
        worksheet.set_column('U:U', 40)
        worksheet.set_column('V:V', 40)
        worksheet.set_column('W:W', 30)
        worksheet.set_column('X:X', 30)
        worksheet.set_column('Y:Y', 30)
        worksheet.set_column('Z:Z', 30)

        worksheet.set_column('AA:AA', 40)
        worksheet.set_column('AB:AB', 40)
        worksheet.set_column('AC:AC', 30)
        worksheet.set_column('AD:AD', 30)
        worksheet.set_column('AE:AE', 30)
        worksheet.set_column('AF:AF', 30)
        worksheet.set_column('AG:AG', 30)
        worksheet.set_column('AH:AH', 30)
        worksheet.set_column('AI:AI', 30)
        worksheet.set_column('AJ:AJ', 30)
        worksheet.set_column('AK:AK', 30)
        worksheet.set_column('AL:AL', 30)
        worksheet.set_column('AM:AM', 30)
        worksheet.set_column('AN:AN', 30)
        worksheet.set_column('AO:AO', 30)
        worksheet.set_column('AP:AP', 30)
        worksheet.set_column('AQ:AQ', 30)
        worksheet.set_column('AR:AR', 30)
        worksheet.set_column('AS:AS', 30)
        worksheet.set_column('AT:AT', 30)
        worksheet.set_column('AU:AU', 40)
        worksheet.set_column('AV:AV', 40)
        worksheet.set_column('AW:AW', 40)
        worksheet.set_column('AX:AX', 30)
        worksheet.set_column('AY:AY', 30)
        worksheet.set_column('AZ:AZ', 30)

        worksheet.set_column('BA:BA', 40)
        worksheet.set_column('BB:BB', 40)
        worksheet.set_column('BC:BC', 30)
        worksheet.set_column('BD:BD', 30)
        worksheet.set_column('BE:BE', 30)
        worksheet.set_column('BF:BF', 30)
        worksheet.set_column('BG:BG', 40)
        worksheet.set_column('BH:BH', 30)
        worksheet.set_column('BI:BI', 30)
        worksheet.set_column('BJ:BJ', 30)
        worksheet.set_column('BK:BK', 30)
        worksheet.set_column('BL:BL', 30)
        worksheet.set_column('BM:BM', 30)
        worksheet.set_column('BN:BN', 30)
        worksheet.set_column('BO:BO', 30)
        worksheet.set_column('BP:BP', 30)
        worksheet.set_column('BQ:BQ', 30)
        worksheet.set_column('BR:BR', 30)
        worksheet.set_column('BS:BS', 30)
        worksheet.set_column('BT:BT', 30)
        worksheet.set_column('BU:BU', 30)
        worksheet.set_column('BV:BV', 30)
        worksheet.set_column('BW:BW', 30)
        worksheet.set_column('BX:BX', 30)
        worksheet.set_column('BY:BY', 30)
        worksheet.set_column('BZ:BZ', 30)

        worksheet.set_column('CA:CA', 40)
        worksheet.set_column('CB:CB', 40)
        worksheet.set_column('CC:CC', 30)
        worksheet.set_column('CD:CD', 30)
        worksheet.set_column('CE:CE', 30)
        worksheet.set_column('CF:CF', 40)
        worksheet.set_column('CG:CG', 40)
        worksheet.set_column('CH:CH', 30)
        worksheet.set_column('CI:CI', 30)
        worksheet.set_column('CJ:CJ', 30)
        worksheet.set_column('CK:CK', 30)
        worksheet.set_column('CL:CL', 30)
        worksheet.set_column('CM:CM', 30)
        worksheet.set_column('CN:CN', 30)
        worksheet.set_column('CO:CO', 30)
        worksheet.set_column('CP:CP', 30)
        worksheet.set_column('CQ:CQ', 30)
        worksheet.set_column('CR:CR', 30)
        worksheet.set_column('CS:CS', 30)
        worksheet.set_column('CT:CT', 30)
        worksheet.set_column('CU:CU', 30)
        worksheet.set_column('CV:CV', 30)
        worksheet.set_column('CW:CW', 30)
        worksheet.set_column('CX:CX', 30)
        worksheet.set_column('CY:CY', 30)
        worksheet.set_column('CZ:CZ', 30)

        worksheet.set_column('DA:DA', 30)
        worksheet.set_column('DB:DB', 30)
        worksheet.set_column('DC:DC', 30)
       
        
        ## agregar estilos
        Backgroundcolor = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 18,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        Backgroundcolor2 = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 14,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        Backgroundcolor3 = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 12,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        cell_format = workbook.add_format()
        cell_format.set_font_color('white')

        textInfo_format = workbook.add_format({'text_wrap': True, 'align': 'center', 'valign': 'vcenter'})
        

        formatfecha = workbook.add_format({'num_format': 'dd/mm/yy', 'align': 'center', 'valign': 'vcenter'})

        ##FILA 1
        worksheet.merge_range('A1:DC1', "Inventario de Demandas de Información", Backgroundcolor)
        worksheet.conditional_format('A1:DC1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Demandas de Información',
                                                'format': cell_format})

        ##FILA 2

        worksheet.merge_range('A2:A3', " ", Backgroundcolor2)
        worksheet.conditional_format('A2:A3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': ' ',
                                                'format': cell_format })

        worksheet.merge_range('B2:B3', " ", Backgroundcolor2)
        worksheet.conditional_format('B2:B3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': ' ',
                                                'format': cell_format })

        worksheet.merge_range('C2:T2', "MÓDULO A. IDENTIFICACIÓN ", Backgroundcolor2)
        worksheet.conditional_format('C2:T2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'MÓDULO A. IDENTIFICACIÓN',
                                                'format': cell_format })

        
        worksheet.merge_range('U2:CU2', "MÓDULO B. DETECCIÓN Y ANÁLISIS DE REQUERIMIENTOS", Backgroundcolor2)
        worksheet.conditional_format('U2:CU2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'MÓDULO B. DETECCIÓN Y ANÁLISIS DE REQUERIMIENTOS',
                                                'format': cell_format })


        worksheet.merge_range('CX2:DA2', "COMPLEMENTO DE DEMANDA", Backgroundcolor3)
        worksheet.conditional_format('CX2:DA2', {'type': 'text',
                                        'criteria': 'begins with',
                                        'value': 'COMPLEMENTO DE DEMANDA',
                                        'format': cell_format })


        worksheet.merge_range('DB2:DC2', "MÓDULO VALIDACIÓN DE LA DEMANDA", Backgroundcolor3)
        worksheet.conditional_format('DB2:DC2', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'MÓDULO VALIDACIÓN DE LA DEMANDA',
                                                'format': cell_format })


        ##FILA 3
        worksheet.merge_range('C3:E3', "ÁREA TEMÁTICA / TEMA", Backgroundcolor2)
        worksheet.conditional_format('C3:E3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'ÁREA TEMÁTICA / TEMA',
                                                'format': cell_format })

        
        worksheet.merge_range('F3:F4',  "Seleccione el comité estadístico sectorial al que pertenece la demanda de información estadística", Backgroundcolor3)
        worksheet.conditional_format('F3:F4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Seleccione el comité estadístico sectorial al que pertenece la demanda de información estadística',
                                                'format': cell_format })

        worksheet.merge_range('G3:N3',  "¿Quién identificó la demanda de información?", Backgroundcolor3)
        worksheet.conditional_format('G3:N3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '¿Quién identificó la demanda de información?',
                                                'format': cell_format })

        worksheet.merge_range('O3:Z3',  "A. Datos del Solicitante", Backgroundcolor3)
        worksheet.conditional_format('O3:AZ3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'A. Datos del Solicitante',
                                                'format': cell_format })

        worksheet.merge_range('AA3:AA4',  "1. ¿Cuál es el indicador o requerimiento de información estadística?", Backgroundcolor3)
        worksheet.conditional_format('AA3:AA4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '1. ¿Cuál es el indicador o requerimiento de información estadística?',
                                                'format': cell_format })

        worksheet.merge_range('AB3:AB4',  "2. Descripción general de la información que requiere", Backgroundcolor3)
        worksheet.conditional_format('AB3:AB4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '2. Descripción general de la información que requiere',
                                                'format': cell_format })

        worksheet.merge_range('AC3:AG3',  "3. ¿Para qué se utiliza la información estadística?", Backgroundcolor3)
        worksheet.conditional_format('AC3:AG3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '3. ¿Para qué se utiliza la información estadística?',
                                                'format': cell_format })

        worksheet.merge_range('AH3:AS3',  "4. ¿Cuáles podrían ser los usuarios principales de la información solicitada?", Backgroundcolor3)
        worksheet.conditional_format('AH3:AS3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '4. ¿Cuáles podrían ser los usuarios principales de la información solicitada?',
                                                'format': cell_format })

        worksheet.merge_range('AT3:AX3',  "5. La información solicitada responde a las siguientes normas", Backgroundcolor3)
        worksheet.conditional_format('AT3:AX3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '5. La información solicitada responde a las siguientes normas',
                                                'format': cell_format })

        worksheet.merge_range('AY3:BE3',  "6. La información solicitada responde a los siguientes requerimientos", Backgroundcolor3)
        worksheet.conditional_format('AY3:BE3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '6. La información solicitada responde a los siguientes requerimientos',
                                                'format': cell_format })

        worksheet.merge_range('BF3:BL3',  "7. ¿La información requerida está siendo producida en su totalidad?", Backgroundcolor3)
        worksheet.conditional_format('BF3:BL3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '7. ¿La información requerida está siendo producida en su totalidad?',
                                                'format': cell_format })

        worksheet.merge_range('BM3:BQ3',  "8. ¿Qué tipo de requerimiento es?", Backgroundcolor3)
        worksheet.conditional_format('BF3:BQ3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '8. ¿Qué tipo de requerimiento es?',
                                                'format': cell_format })


        worksheet.merge_range('BR3:BR4',  "9. ¿Qué variables necesita para suplir el requerimiento? (Ver hoja 2 - Lista de variables)", Backgroundcolor3)
        worksheet.conditional_format('BR3:BR4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '9. ¿Qué variables necesita para suplir el requerimiento? (Ver hoja 2 - Lista de variables)',
                                                'format': cell_format })

        worksheet.merge_range('BS3:BT3',  "10. ¿Cuál entidad considera que debe producir la información requerida?", Backgroundcolor3)
        worksheet.conditional_format('BS3:BT3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '10. ¿Cuál entidad considera que debe producir la información requerida?',
                                                'format': cell_format })


        worksheet.merge_range('BU3:BX3',  "11. ¿Cuál(es) de las siguientes opciones podrían dar solución al requerimiento de información? (Ver hoja 3 - pregunta 11)", Backgroundcolor3)
        worksheet.conditional_format('BU3:BX3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '11. ¿Cuál(es) de las siguientes opciones podrían dar solución al requerimiento de información? (Ver hoja 3 - pregunta 11)',
                                                'format': cell_format })

        
        worksheet.merge_range('BY3:CE3',  "12. Indique la desagregación requerida:", Backgroundcolor3)
        worksheet.conditional_format('BY3:CE3', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '12. Indique la desagregación requerida:',
                                                'format': cell_format })


        worksheet.merge_range('CF3:CP3', "13. Indique la desagregación requerida", Backgroundcolor3)
        worksheet.conditional_format('CF3:CP3', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '13. Indique la desagregación requerida',
                                                'format': cell_format })

        worksheet.merge_range('CQ3:CU3', "14. Periodicidad de difusión requerida", Backgroundcolor3)
        worksheet.conditional_format('CQ3:CU3', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '14. Periodicidad de difusión requerida',
                                                'format': cell_format })


        worksheet.merge_range('CV2:CV4', "MÓDULO C. OBSERVACIONES", Backgroundcolor3)
        worksheet.conditional_format('CV2:CV4', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'MÓDULO C. OBSERVACIONES',
                                                'format': cell_format })


        worksheet.merge_range('CW2:CW4', "MÓDULO D. ANEXOS", Backgroundcolor3)
        worksheet.conditional_format('CW2:CW4', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'MÓDULO D. ANEXOS',
                                                'format': cell_format })


        worksheet.merge_range('CX3:CY3', "¿Es un demanda insatisfecha priorizada?", Backgroundcolor3)
        worksheet.conditional_format('CX3:CY3', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '¿Es un demanda insatisfecha priorizada?',
                                                'format': cell_format })

        worksheet.merge_range('CZ3:DA3', "¿Es un demanda insatisfecha priorizada?", Backgroundcolor3)
        worksheet.conditional_format('CZ3:DA3', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '¿Es un demanda insatisfecha priorizada?',
                                                'format': cell_format })


        worksheet.merge_range('DB3:DC3', "¿El requerimiento reportado es una demanda de información estadística?", Backgroundcolor3)
        worksheet.conditional_format('DB3:DC3', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '¿El requerimiento reportado es una demanda de información estadística?',
                                                'format': cell_format })


        worksheet.write(3, 0, "Código", Backgroundcolor3)
        worksheet.write(3, 1, "Estado", Backgroundcolor3)
        worksheet.write(3, 2, "Área temática", Backgroundcolor3)
        worksheet.write(3, 3, "Tema principal", Backgroundcolor3)
        worksheet.write(3, 4, "Tema compartido", Backgroundcolor3)


        worksheet.write(3, 6, "Consejo Asesor Técnico del SEN", Backgroundcolor3)
        worksheet.write(3, 7, "Comité Estadístico Sectorial", Backgroundcolor3)
        worksheet.write(3, 8, "Mesa Estadística Sectorial", Backgroundcolor3)
        worksheet.write(3, 9, "Entidad", Backgroundcolor3)
        worksheet.write(3, 10, "¿Cuales?", Backgroundcolor3)
        worksheet.write(3, 11, "Otra entidad", Backgroundcolor3)
        worksheet.write(3, 12, "DANE", Backgroundcolor3)
        worksheet.write(3, 13, "Otras instancias de coordinación", Backgroundcolor3)
        

        worksheet.write(3, 14, "Código de la entidad", Backgroundcolor3)
        worksheet.write(3, 15, "Nombre de la entidad", Backgroundcolor3)
        worksheet.write(3, 16, "Entidad consumidora de información", Backgroundcolor3)
        worksheet.write(3, 17, "Dependencia", Backgroundcolor3)
        worksheet.write(3, 18, "Nombre del jefe de la dependencia", Backgroundcolor3)
        worksheet.write(3, 19, "Cargo", Backgroundcolor3)
        worksheet.write(3, 20, "Correo electrónico", Backgroundcolor3)
        worksheet.write(3, 21, "Teléfono jefe dependencia", Backgroundcolor3)
        worksheet.write(3, 22, "Persona que realiza el requerimiento", Backgroundcolor3)
        worksheet.write(3, 23, "Cargo", Backgroundcolor3)
        worksheet.write(3, 24, "Correo electrónico", Backgroundcolor3)
        worksheet.write(3, 25, "Teléfono", Backgroundcolor3)

        worksheet.write(3, 28, "a. Análisis de contexto", Backgroundcolor3)
        worksheet.write(3, 29, "b. Diseño, formulación o seguimiento de políticas", Backgroundcolor3)
        worksheet.write(3, 30, "c. Reportes internacionales", Backgroundcolor3)
        worksheet.write(3, 31, "d. Otro ¿Cuál?", Backgroundcolor3)
        worksheet.write(3, 32, "¿Cuál?", Backgroundcolor3)

        worksheet.write(3, 33, "a. Organismos internacionales", Backgroundcolor3)
        worksheet.write(3, 34, "b. Presidencia de la República", Backgroundcolor3)
        worksheet.write(3, 35, "c. Ministerios", Backgroundcolor3)
        worksheet.write(3, 36, "d. Organismos de Control", Backgroundcolor3)
        worksheet.write(3, 37, "e. Otras entidades del orden Nacional", Backgroundcolor3)
        worksheet.write(3, 38, "f. Entidades de orden Territorial", Backgroundcolor3)
        worksheet.write(3, 39, "g. Gremios", Backgroundcolor3)
        worksheet.write(3, 40, "h. Entidades privadas", Backgroundcolor3)
        worksheet.write(3, 41, "i. Dependencias de la misma entidad", Backgroundcolor3)
        worksheet.write(3, 42, "j. Academia", Backgroundcolor3)
        worksheet.write(3, 43, "k. Público en General", Backgroundcolor3)
        worksheet.write(3, 44, "l. Otra ¿cuál?", Backgroundcolor3)

        worksheet.write(3, 45, "a. Constitución Política", Backgroundcolor3)
        worksheet.write(3, 46, "b. Ley", Backgroundcolor3)
        worksheet.write(3, 47, "c. Decreto (Nacional, departamental, etc.)", Backgroundcolor3)
        worksheet.write(3, 48, "d. Otra (Resolución, ordenanza, acuerdo)", Backgroundcolor3)
        worksheet.write(3, 49, "e. Ninguna", Backgroundcolor3)

        worksheet.write(3, 50, "a. Plan Nacional de Desarrollo (Capítulo y línea)", Backgroundcolor3)
        worksheet.write(3, 51, "b. Cuentas económicas y macroeconómicas", Backgroundcolor3)
        worksheet.write(3, 52, "c. Plan Sectorial, Territorial o CONPES", Backgroundcolor3)
        worksheet.write(3, 53, "d. Objetivos de Desarrollo Sostenible (ODS) ¿Cuál es el número del Objetivo? ¿Cuál es el número del Indicador?", Backgroundcolor3)
        worksheet.write(3, 54, "e. Organización para la Cooperación y el Desarrollo Económicos (OCDE)", Backgroundcolor3)
        worksheet.write(3, 55, "f. Otros compromisos internacionales", Backgroundcolor3)
        worksheet.write(3, 56, "g. Otro(s)", Backgroundcolor3)

        worksheet.write(3, 57, "Si/No", Backgroundcolor3)
        worksheet.write(3, 58, "¿Cuál entidad?", Backgroundcolor3)
        worksheet.write(3, 59, "Otra entidad", Backgroundcolor3)
        worksheet.write(3, 60, "Nombre de la operación estadística", Backgroundcolor3)
        worksheet.write(3, 61, "Otra operación estadística", Backgroundcolor3)
        worksheet.write(3, 62, "Nombre del registro administrativo", Backgroundcolor3)
        worksheet.write(3, 63, "Otra registro administrativo", Backgroundcolor3)

        worksheet.write(3, 64, "a. Agregado estadístico o indicador", Backgroundcolor3)
        worksheet.write(3, 65, "b. Ampliación cobertura geográfica", Backgroundcolor3)
        worksheet.write(3, 66, "c. Ampliación cobertura temática", Backgroundcolor3)
        worksheet.write(3, 67, "d. Ajustes en la difusión", Backgroundcolor3)
        worksheet.write(3, 68, "e. Otro. ¿Cuál?", Backgroundcolor3)

        worksheet.write(3, 70, "Cual entidad", Backgroundcolor3)
        worksheet.write(3, 71, "Otra ¿Cuál?", Backgroundcolor3)

        worksheet.write(3, 72, "a. Aprovechamiento de una operación (es) estadística (s)", Backgroundcolor3)
        worksheet.write(3, 73, "b. Aprovechamiento de un registro administrativo", Backgroundcolor3)
        worksheet.write(3, 74, "c. Generación de nueva información", Backgroundcolor3)
        worksheet.write(3, 75, "d. Otra (s) ¿Cuáles?", Backgroundcolor3)

        worksheet.write(3, 76, "a. Sexo", Backgroundcolor3)
        worksheet.write(3, 77, "b. Edades", Backgroundcolor3)
        worksheet.write(3, 78, "c. Grupos étnicos", Backgroundcolor3)
        worksheet.write(3, 79, "d. Discapacidad", Backgroundcolor3)
        worksheet.write(3, 80, "e. Estrato", Backgroundcolor3)
        worksheet.write(3, 81, "f. Otra ¿Cuál?", Backgroundcolor3)
        worksheet.write(3, 82, "g. Ninguna", Backgroundcolor3)

        worksheet.write(3, 83, "Geográfica", Backgroundcolor3)
        worksheet.write(3, 84, "Nacional", Backgroundcolor3)
        worksheet.write(3, 85, "Regional", Backgroundcolor3)
        worksheet.write(3, 86, "Departamental", Backgroundcolor3)
        worksheet.write(3, 87, "Áreas metropolitanas", Backgroundcolor3)
        worksheet.write(3, 88, "Municipal", Backgroundcolor3)
        worksheet.write(3, 89, "Otra, ¿cuál?", Backgroundcolor3)
        worksheet.write(3, 90, "Zona", Backgroundcolor3)
        worksheet.write(3, 91, "Total", Backgroundcolor3)
        worksheet.write(3, 92, "Urbano", Backgroundcolor3)
        worksheet.write(3, 93, "Rural", Backgroundcolor3)

        worksheet.write(3, 94, "a. Anual", Backgroundcolor3)
        worksheet.write(3, 95, "b. Semestral", Backgroundcolor3)
        worksheet.write(3, 96, "c. Trimestral", Backgroundcolor3)
        worksheet.write(3, 97, "d. Mensual", Backgroundcolor3)
        worksheet.write(3, 98, "e. Otra ¿Cuál?", Backgroundcolor3)

        ##COMPLEMENTO DE DEMANDA
        worksheet.write(3, 101, "SI/NO", Backgroundcolor3)
        worksheet.write(3, 102, "¿Instancia de coordinación o entidad que lo priorizó?", Backgroundcolor3)
        worksheet.write(3, 103, "SI/NO", Backgroundcolor3)
        worksheet.write(3, 104, "Describa de forma breve la acción que suple la demanda insatisfecha", Backgroundcolor3)

        ##VALIDACIÓN DE LA DEMANDA
        worksheet.write(3, 105, "SI/NO", Backgroundcolor3)
        worksheet.write(3, 106, "Describa de forma breve", Backgroundcolor3)

        for dem in demandas:
            worksheet.write(row, col, dem.codigo_ddi, textInfo_format)
            worksheet.write(row, col + 1, str(dem.nombre_est), textInfo_format)
            worksheet.write(row, col + 2, str(dem.area_tem), textInfo_format)
            worksheet.write(row, col + 3, str(dem.tema_prin), textInfo_format)
            
            lista_temaComp = dem.tema_comp.all()
            lista_temaComp_array = []
            for index, item in enumerate (lista_temaComp):
                lista_temaComp_array.append(item)
                worksheet.write(row, col + 4, str(lista_temaComp_array).replace('[','').replace(']','').replace("<TemaCompartido:", "").replace(">", ""), textInfo_format)
               
            lista_comiteEst = dem.comite_est.all()
            lista_comiteEst_array = []
            for index, item in enumerate (lista_comiteEst):
                lista_comiteEst_array.append(item)
                worksheet.write(row, col + 5, str(lista_comiteEst_array).replace('[','').replace(']','').replace("<ComiteEstSect:", "").replace(">", ""), textInfo_format) 
            
            lista_quien_identddi = dem.quien_identddi.all()
            for index, item in enumerate(lista_quien_identddi):
                if str(item) == "Consejo Asesor Técnico del SEN":
                    worksheet.write(row, col + 6, str(item), textInfo_format)

                if str(item) == "Comité Estadístico Sectorial":
                    worksheet.write(row, col + 7, str(item), textInfo_format)

                if str(item) == "Mesa Estadística Sectorial":
                    worksheet.write(row, col + 8, str(item), textInfo_format)

                if str(item) == "Entidad":
                    worksheet.write(row, col + 9, str(item), textInfo_format)

                if str(item) == "DANE":
                    worksheet.write(row, col + 12, str(item), textInfo_format)

                if str(item) == "Otras instancias de coordinación":
                    worksheet.write(row, col + 13, str(item), textInfo_format)

            lista_entidad_qiddi = dem.entidad_qiddi.all()
            lista_entidad_qiddi_array = []
            for index, item in enumerate (lista_entidad_qiddi):
                lista_entidad_qiddi_array .append(item)
                worksheet.write(row, col + 10, str(lista_entidad_qiddi_array ).replace('[','').replace(']','').replace("<entidad_service:", "").replace(">", ""), textInfo_format)

            worksheet.write(row, col + 11, str(dem.otra_entidad_qiddi), textInfo_format)

            if dem.entidad_sol == None:
                worksheet.write(row, col + 14, "", textInfo_format)
            else:
                worksheet.write(row, col + 14, str(dem.entidad_sol.codigo), textInfo_format)

            if dem.entidad_sol == None:
                worksheet.write(row, col + 15, "", textInfo_format)
            else:
                worksheet.write(row, col + 15, str(dem.entidad_sol), textInfo_format)

            if dem.entidad_cons_sol == None:
                worksheet.write(row, col + 16, "", textInfo_format)
            else:
                worksheet.write(row, col + 16, str(dem.entidad_cons_sol), textInfo_format)


            worksheet.write(row, col + 17, str(dem.dependencia), textInfo_format)
            worksheet.write(row, col + 18, str(dem.nombre_jef_dep), textInfo_format)
            worksheet.write(row, col + 19, str(dem.cargo_jef_dep), textInfo_format)
            worksheet.write(row, col + 20, str(dem.correo_jef_dep), textInfo_format)

            if dem.telefono_jef_dep == None:
                worksheet.write(row, col + 21, "", textInfo_format)
            else:
                worksheet.write(row, col + 21, str(dem.telefono_jef_dep), textInfo_format)

            worksheet.write(row, col + 22, str(dem.pers_req), textInfo_format)
            worksheet.write(row, col + 23, str(dem.cargo_pers_req), textInfo_format)
            worksheet.write(row, col + 24, str(dem.correo_pers_req), textInfo_format)

            if dem.telefono_pers_req == None:
                worksheet.write(row, col + 25, "", textInfo_format)
            else:
                worksheet.write(row, col + 25, str(dem.telefono_pers_req), textInfo_format)

            #MÓDULO B
            worksheet.write(row, col + 26, str(dem.pm_b_1), textInfo_format)
            worksheet.write(row, col + 27, str(dem.pm_b_2), textInfo_format)
            
            lista_pm_b_3 = dem.pm_b_3.all()
            for index_pmb3, item_pmb3 in enumerate(lista_pm_b_3):
                if str(item_pmb3) == "Análisis de contexto":
                    worksheet.write(row, col + 28, str(item_pmb3), textInfo_format)

                if str(item_pmb3) == "Diseño, formulación o seguimiento de políticas":
                    worksheet.write(row, col + 29, str(item_pmb3), textInfo_format)

                if str(item_pmb3) == "Reportes internacionales":
                    worksheet.write(row, col + 30, str(item_pmb3), textInfo_format)

                if str(item_pmb3) == "Otro ¿Cuál?":
                    worksheet.write(row, col + 31, str(item_pmb3), textInfo_format)

                
            lista_pm_b_4 = dem.pm_b_4.all()
            for index_pm_b_4, item_pm_b_4 in enumerate(lista_pm_b_4):
                
                if str(item_pm_b_4) == "Presidencia de la República":
                    worksheet.write(row, col + 34, str(item_pm_b_4), textInfo_format)

                if str(item_pm_b_4) == "Público en General":
                    worksheet.write(row, col + 43, str(item_pm_b_4), textInfo_format)


            lista_pm_b_5 = dem.pm_b_5.all()
            for index_pm_b_5, item_pm_b_5 in enumerate(lista_pm_b_5):
                
                if str(item_pm_b_5) == "Ninguna":
                    worksheet.write(row, col + 49, str(item_pm_b_5), textInfo_format)

            lista_pm_b_7 = dem.total_si_no_pmb7.all()
            lista_pm_b_7_array = []
            for index_pm_b_7, item_pm_b_7 in enumerate(lista_pm_b_7):
                lista_pm_b_7_array.append(item_pm_b_7)
                worksheet.write(row, col + 57, str(lista_pm_b_7_array).replace('[','').replace(']','').replace("<InfoProdTotalidad:", "").replace(">", ""), textInfo_format)

            lista_entidad_pm_b7 = dem.entidad_pm_b7.all()
            lista_entidad_pm_b7_array = []
            for index, item in enumerate(lista_entidad_pm_b7):
                lista_entidad_pm_b7_array.append(item)
                worksheet.write(row, col + 58, str(lista_entidad_pm_b7_array).replace('[','').replace(']','').replace("<entidad_service:", "").replace(">", ""), textInfo_format)
            
            worksheet.write(row, col + 59, str(dem.otro_entidad_pm_b7), textInfo_format)

            lista_ooee_pm_b7 = dem.ooee_pm_b7.all()
            lista_ooee_pm_b7_array = []
            for index, item in enumerate(lista_ooee_pm_b7):
                lista_ooee_pm_b7_array.append(item)
                worksheet.write(row, col + 60, str(lista_ooee_pm_b7_array).replace('[','').replace(']','').replace("<ooee_service:", "").replace(">", ""), textInfo_format)
            
            worksheet.write(row, col + 61, str(dem.otro_ooee_pm_b7), textInfo_format)
            
            lista_rraa_pm_b7 = dem.rraa_pm_b7.all()
            lista_rraa_pm_b7_array = []
            for index, item in enumerate(lista_rraa_pm_b7):
                lista_rraa_pm_b7_array.append(item)
                worksheet.write(row, col + 62, str(lista_rraa_pm_b7_array).replace('[','').replace(']','').replace("<rraa_service:", "").replace(">", ""), textInfo_format)

            worksheet.write(row, col + 63, str(dem.otro_rraa_pm_b7), textInfo_format)
                
            lista_pm_b_8 = dem.pm_b_8.all()
            for index_pm_b_8, item_pm_b_8 in enumerate(lista_pm_b_8):
                
                if str(item_pm_b_8) == "Agregado estadístico o indicador":
                    worksheet.write(row, col + 64, str(item_pm_b_8), textInfo_format)

                if str(item_pm_b_8) == "Ampliación cobertura geográfica":
                    worksheet.write(row, col + 65, str(item_pm_b_8), textInfo_format)

                if str(item_pm_b_8) == "Ampliación cobertura temática":
                    worksheet.write(row, col + 66, str(item_pm_b_8), textInfo_format)

                if str(item_pm_b_8) == "Ajustes en la difusión":
                    worksheet.write(row, col + 67, str(item_pm_b_8), textInfo_format)


            lista_pm_b_10 = dem.pm_b_10.all()
            lista_pm_b_10_array = []
            for index, item in enumerate(lista_pm_b_10):
                lista_pm_b_10_array.append(item)
                worksheet.write(row, col + 70, str(lista_pm_b_10_array).replace('[','').replace(']','').replace("<entidad_service:", "").replace(">", ""), textInfo_format)
            
            worksheet.write(row, col + 71, str(dem.pm_b_10_otro), textInfo_format)

            lista_pm_b_11_1 = dem.pm_b_11_1.all()
            for index_pm_b_11_1, item_pm_b_11_1 in enumerate(lista_pm_b_11_1):
                
                if str(item_pm_b_11_1) == "Aprovechamiento de una operación (es) estadística (s)":
                    worksheet.write(row, col + 72, str(item_pm_b_11_1), textInfo_format)

                if str(item_pm_b_11_1) == "Aprovechamiento de un registro administrativo":
                    worksheet.write(row, col + 73, str(item_pm_b_11_1), textInfo_format)

                if str(item_pm_b_11_1) == "Generación de nueva información":
                    worksheet.write(row, col + 74, str(item_pm_b_11_1), textInfo_format)

            worksheet.write(row, col + 75, str(dem.otra_cual_11_1_d), textInfo_format)

            lista_pm_b_12 = dem.pm_b_12.all()
            for index_pm_b_12, item_pm_b_12 in enumerate(lista_pm_b_12):
                
                if str(item_pm_b_12) == "Sexo":
                    worksheet.write(row, col + 76, str(item_pm_b_12), textInfo_format)

                if str(item_pm_b_12) == "Edades":
                    worksheet.write(row, col + 77, str(item_pm_b_12), textInfo_format)

                if str(item_pm_b_12) == "Grupos étnicos":
                    worksheet.write(row, col + 78, str(item_pm_b_12), textInfo_format)

                if str(item_pm_b_12) == "Discapacidad":
                    worksheet.write(row, col + 79, str(item_pm_b_12), textInfo_format)

                if str(item_pm_b_12) == "Estrato":
                    worksheet.write(row, col + 80, str(item_pm_b_12), textInfo_format)

                if str(item_pm_b_12) == "Ninguna":
                    worksheet.write(row, col + 82, str(item_pm_b_12), textInfo_format)


            lista_pm_b_13 = dem.pm_b_13.all()
            for index_pm_b_13, item_pm_b_13 in enumerate(lista_pm_b_13):
                
                if str(item_pm_b_13) == "Geográfica":
                    worksheet.write(row, col + 83, str(item_pm_b_13), textInfo_format)

                if str(item_pm_b_13) == "Zona":
                    worksheet.write(row, col + 90, str(item_pm_b_13), textInfo_format)


            lista_pm_b_13_geo = dem.pm_b_13_geo.all()
            for index_pm_b_13_geo, item_pm_b_13_geo in enumerate(lista_pm_b_13_geo):

                if str(item_pm_b_13_geo) == "Nacional":
                    worksheet.write(row, col + 84, str(item_pm_b_13_geo), textInfo_format)

                if str(item_pm_b_13_geo) == "Regional":
                    worksheet.write(row, col + 85, str(item_pm_b_13_geo), textInfo_format)

                if str(item_pm_b_13_geo) == "Departamental":
                    worksheet.write(row, col + 86, str(item_pm_b_13_geo), textInfo_format)

                if str(item_pm_b_13_geo) == "Áreas metropolitanas":
                    worksheet.write(row, col + 87, str(item_pm_b_13_geo), textInfo_format)

                if str(item_pm_b_13_geo) == "Municipal":
                    worksheet.write(row, col + 88, str(item_pm_b_13_geo), textInfo_format)

            lista_pm_b_13_zona = dem.pm_b_13_zona.all()
            for index_pm_b_13_zona, item_pm_b_13_zona in enumerate(lista_pm_b_13_zona):

                if str(item_pm_b_13_zona) == "Total":
                    worksheet.write(row, col + 91, str(item_pm_b_13_zona), textInfo_format)

                if str(item_pm_b_13_zona) == "Urbano":
                    worksheet.write(row, col + 92, str(item_pm_b_13_zona), textInfo_format)

                if str(item_pm_b_13_zona) == "Rural":
                    worksheet.write(row, col + 93, str(item_pm_b_13_zona), textInfo_format)


            lista_pm_b_14 = dem.pm_b_14.all()
            for index_pm_b_14, item_pm_b_14 in enumerate(lista_pm_b_14):
                
                if str(item_pm_b_14) == "Anual":
                    worksheet.write(row, col + 94, str(item_pm_b_14), textInfo_format)

                if str(item_pm_b_14) == "Semestral":
                    worksheet.write(row, col + 95, str(item_pm_b_14), textInfo_format)

                if str(item_pm_b_14) == "Trimestral":
                    worksheet.write(row, col + 96, str(item_pm_b_14), textInfo_format)

                if str(item_pm_b_14) == "Mensual":
                    worksheet.write(row, col + 97, str(item_pm_b_14), textInfo_format)
            
            ##MÓDULO C
            worksheet.write(row, col + 99, str(dem.pm_c_1), textInfo_format)
            ##MÓDULO D
            worksheet.write(row, col + 100, str(dem.pm_d_1_anexos), textInfo_format)
            ###Complemento de la demanda
            lista_compl_dem_a = dem.compl_dem_a.all()
            lista_compl_dem_a_array = []
            for index, item in enumerate(lista_compl_dem_a):
                lista_compl_dem_a_array.append(item)
                worksheet.write(row, col + 101, str(lista_compl_dem_a_array).replace('[','').replace(']','').replace("<DemandaInsaprio:", "").replace(">", ""), textInfo_format)
            
            worksheet.write(row, col + 102, str(dem.compl_dem_a_text), textInfo_format)
            
            lista_comple_dem_b = dem.comple_dem_b.all()
            lista_comple_dem_b_array = []
            for index, item in enumerate(lista_comple_dem_b):
                lista_comple_dem_b_array.append(item)
                worksheet.write(row, col + 103, str(lista_comple_dem_b_array).replace('[','').replace(']','').replace("<PlanaccionSuplirDem:", "").replace(">", ""), textInfo_format)
            
            worksheet.write(row, col + 104, str(dem.compl_dem_b_text), textInfo_format)

            ###Complemento de la demanda
            lista_validacion_ddi = dem.validacion_ddi.all()
            lista_validacion_ddi_array = []
            for index, item in enumerate(lista_validacion_ddi):
                lista_validacion_ddi_array.append(item)
                worksheet.write(row, col + 105, str(lista_validacion_ddi_array).replace('[','').replace(']','').replace("<EsDemandadeInfor:", "").replace(">", ""), textInfo_format)
                        
            worksheet.write(row, col + 106, str(dem.validacion_ddi_text), textInfo_format)

            row += 1

        # modulo B campos de textos
        utilizaInfEstTextList = list(UtilizaInfEstText.objects.all())
        for index, item in enumerate(utilizaInfEstTextList):
            
            index = item.ddi_id + 3
            worksheet.write(index, col + 32, str(item.otro_cual), textInfo_format)

        UsuariosPrincInfEstTextList = list(UsuariosPrincInfEstText.objects.all())
        for index, item in enumerate(UsuariosPrincInfEstTextList):
            
            index = item.ddi_id + 3
            worksheet.write(index, col + 33, str(item.orginter_text), textInfo_format)
            worksheet.write(index, col + 35, str(item.ministerios_text), textInfo_format)
            worksheet.write(index, col + 36, str(item.orgcontrol_text), textInfo_format)
            worksheet.write(index, col + 37, str(item.oentidadesordenal_text), textInfo_format)
            worksheet.write(index, col + 38, str(item.entidadesordenterr_text), textInfo_format)
            worksheet.write(index, col + 39, str(item.gremios_text), textInfo_format)
            worksheet.write(index, col + 40, str(item.entiprivadas_text), textInfo_format)
            worksheet.write(index, col + 41, str(item.dependenentidad_text), textInfo_format)
            worksheet.write(index, col + 42, str(item.academia_text), textInfo_format)
            worksheet.write(index, col + 44, str(item.otro_cual_text), textInfo_format)

        
        NormasTextList = list(NormasText.objects.all())
        for index, item in enumerate(NormasTextList):
            index = item.ddi_id + 3
            worksheet.write(index, col + 45, str(item.const_pol_text), textInfo_format)
            worksheet.write(index, col + 46, str(item.ley_text), textInfo_format)
            worksheet.write(index, col + 47, str(item.decreto_text), textInfo_format)
            worksheet.write(index, col + 48, str(item.otra_text), textInfo_format)

        InfSolRespRequerimientoTextList = list(InfSolRespRequerimientoText.objects.all())
        for index, item in enumerate(InfSolRespRequerimientoTextList):
            index = item.ddi_id + 3
            worksheet.write(index, col + 50, str(item.planalDes_text), textInfo_format)
            worksheet.write(index, col + 51, str(item.cuentasecomacroec_text), textInfo_format)
            worksheet.write(index, col + 52, str(item.plansecterrcom_text), textInfo_format)
            worksheet.write(index, col + 53, str(item.objdessost_text), textInfo_format)
            worksheet.write(index, col + 54, str(item.orgcooper_text), textInfo_format)
            worksheet.write(index, col + 55, str(item.otroscomprInt_text), textInfo_format)
            worksheet.write(index, col + 56, str(item.otros_text), textInfo_format)


        TipoRequerimientoTextList = list(TipoRequerimientoText.objects.all())
        for index, item in enumerate(TipoRequerimientoTextList):
            index = item.ddi_id + 3
            worksheet.write(index, col + 68, str(item.otros_c_text), textInfo_format)

        DesagregacionReqTextList = list(DesagregacionReqText.objects.all())
        for index, item in enumerate(DesagregacionReqTextList):
            index = item.ddi_id + 3
            worksheet.write(index, col + 81, str(item.otra_cual_text_a), textInfo_format)
            

        DesagregacionGeoReqGeograficaTextList = list(DesagregacionGeoReqGeograficaText.objects.all())
        for index, item in enumerate(DesagregacionGeoReqGeograficaTextList):
            index = item.ddi_id + 3
            worksheet.write(index, col + 89, str(item.otra_cual_text_b), textInfo_format)

        PeriodicidadDifusionTextList = list(PeriodicidadDifusionText.objects.all())
        for index, item in enumerate(PeriodicidadDifusionTextList):
            index = item.ddi_id + 3
            worksheet.write(index, col + 98, str(item.otra_cual_text_c), textInfo_format)


        ################################## Hoja 2 pregunta 9 lista de variables ###################3 
        worksheet2 = workbook.add_worksheet(name="Lista de variables")
        # Get some data to write to the spreadsheet.
        
        preg9 = listaVariables.objects.all()
        
        row2 = 4
        col2 = 0
        ## altura de celda
        worksheet2.set_row(0, 40)
        worksheet2.set_row(1, 30)
        worksheet2.set_row(2, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet2.set_column('A:A', 40)
        worksheet2.set_column('B:B', 40)
        worksheet2.set_column('C:C', 40)
        

        worksheet2.merge_range('A1:C1', "Inventario de Demandas de Información", Backgroundcolor)
        worksheet2.conditional_format('A1:C1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Demandas de Información',
                                                'format': cell_format})


        worksheet2.merge_range('A2:C2', "MÓDULO B Detección y Análisis de Requerimientos", Backgroundcolor2)
        worksheet2.conditional_format('A2:C2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "MÓDULO B Detección y Análisis de Requerimientos",
                                                'format': cell_format})


        worksheet2.merge_range('A3:C3', "¿Qué variables necesita para suplir el requerimiento?", Backgroundcolor2)
        worksheet2.conditional_format('A2:C2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "¿Qué variables necesita para suplir el requerimiento?",
                                                'format': cell_format})


        
        worksheet2.write(3, 0, "Código", Backgroundcolor3)
        worksheet2.write(3, 1, "¿Cuál es el indicador o requerimiento de información estadística?", Backgroundcolor3)
        worksheet2.write(3, 2, "variables que necesita para suplir el requerimiento", Backgroundcolor3)
       

        for p9 in preg9:
            worksheet2.write(row2, col2, str(p9.ddi.codigo_ddi), textInfo_format)
            worksheet2.write(row2, col2 + 1, str(p9.ddi), textInfo_format)
            worksheet2.write(row2, col2 + 2, p9.lista_varia, textInfo_format)

            row2 += 1

        ################################## Hoja 3 pregunta 11 ###################
        worksheet3 = workbook.add_worksheet(name="Pregunta 11")

        preg11 = SolReqInformacion.objects.all()

        row3 = 5
        col3 = 0
        ## altura de celda
        worksheet3.set_row(0, 40)
        worksheet3.set_row(1, 30)
        worksheet3.set_row(2, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet3.set_column('A:A', 40)
        worksheet3.set_column('B:B', 40)
        worksheet3.set_column('C:C', 40)
        worksheet3.set_column('D:D', 40)
        worksheet3.set_column('E:E', 40)
        worksheet3.set_column('F:F', 40)
        worksheet3.set_column('G:G', 40)
        worksheet3.set_column('H:H', 40)
        worksheet3.set_column('I:I', 40)
        worksheet3.set_column('J:J', 40)
        worksheet3.set_column('K:K', 40)
        worksheet3.set_column('L:L', 40)
        worksheet3.set_column('M:M', 40)
        worksheet3.set_column('N:N', 40)
        worksheet3.set_column('O:O', 40)
        worksheet3.set_column('P:P', 40)
        
        
        worksheet3.merge_range('A1:P1', "Inventario de Demandas de Información", Backgroundcolor)
        worksheet3.conditional_format('A1:P1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Demandas de Información',
                                                'format': cell_format})

        worksheet3.merge_range('A2:A4', "", Backgroundcolor)
        worksheet3.conditional_format('A2:A4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '',
                                                'format': cell_format})

        worksheet3.merge_range('B2:B4', "", Backgroundcolor)
        worksheet3.conditional_format('B2:B4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '',
                                                'format': cell_format})


        worksheet3.merge_range('C2:P2', "MÓDULO B Detección y Análisis de Requerimientos", Backgroundcolor2)
        worksheet3.conditional_format('C2:P2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "MÓDULO B Detección y Análisis de Requerimientos",
                                                'format': cell_format})


        worksheet3.merge_range('C3:P3', "¿Cuál(es) de las siguientes opciones podrían dar solución al requerimiento de información?", Backgroundcolor2)
        worksheet3.conditional_format('C3:P3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "¿Cuál(es) de las siguientes opciones podrían dar solución al requerimiento de información?",
                                                'format': cell_format})

        worksheet3.merge_range('C4:I4', "a) Aprovechamiento de una operación (es) estadística (s)", Backgroundcolor2)
        worksheet3.conditional_format('C4:I4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "a) Aprovechamiento de una operación (es) estadística (s)",
                                                'format': cell_format})


        worksheet3.merge_range('J4:M4', "b) Aprovechamiento de un registro administrativo", Backgroundcolor2)
        worksheet3.conditional_format('J4:M4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "b) Aprovechamiento de un registro administrativo",
                                                'format': cell_format})

        worksheet3.merge_range('N4:P4', "c) Generación de nueva información", Backgroundcolor2)
        worksheet3.conditional_format('N4:P4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "c) Generación de nueva información",
                                                'format': cell_format})


        
        worksheet3.write(4, 0, "Código", Backgroundcolor3)
        worksheet3.write(4, 1, "¿Cuál es el indicador o requerimiento de información estadística?", Backgroundcolor3)
        
        worksheet3.write(4, 2, "a) 1. ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 3, "a)2.1. Incluir variables/preguntas ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 4, "a)2.2. Cambiar la formulación de alguna(s) pregunta(s) ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 5, "a)2.3. Ampliar la desagregación temática ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 6, "a)2.4. Ampliar la desagregación geográfica ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 7, "a)2.5. Aumentar la frecuencia en la difusión de los resultados ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 8, "a)2.6. Otra ¿Cuál(es)?", Backgroundcolor3)

        worksheet3.write(4, 9, "b)1.  ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 10, "b)2.1.  Inclusión de variables/preguntas", Backgroundcolor3)
        worksheet3.write(4, 11, "b)2.2. Cambios en la formulación de alguna(s) pregunta(s)", Backgroundcolor3)
        worksheet3.write(4, 12, "b)2.3. Otro ", Backgroundcolor3)

        
        worksheet3.write(4, 13, "c)1. Operación estadística nueva", Backgroundcolor3)
        worksheet3.write(4, 14, "c)2. Indicador ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 15, "c)3. Otra  ¿Cuál?", Backgroundcolor3)
        				
					
        for p11 in preg11:
            worksheet3.write(row3, col3, str(p11.post_ddi.codigo_ddi), textInfo_format)
            worksheet3.write(row3, col3 + 1, str(p11.post_ddi), textInfo_format)

            if p11.ooee_pmb11 == None:
                worksheet3.write(row3, col3 + 2, "", textInfo_format)
            else:
                worksheet3.write(row3, col3 + 2, str(p11.ooee_pmb11), textInfo_format)


            worksheet3.write(row3, col3 + 3, p11.inc_var_cual, textInfo_format)
            worksheet3.write(row3, col3 + 4, p11.cam_preg_cual, textInfo_format)
            worksheet3.write(row3, col3 + 5, p11.am_des_tem_cual, textInfo_format)
            worksheet3.write(row3, col3 + 6, p11.am_des_geo_cual, textInfo_format)
            worksheet3.write(row3, col3 + 7, p11.dif_resul_cual, textInfo_format)
            worksheet3.write(row3, col3 + 8, p11.opc_aprov_cual, textInfo_format)

            if p11.rraa_pmb11 == None:
                worksheet3.write(row3, col3 + 9, "", textInfo_format)
            else:
                worksheet3.write(row3, col3 + 9, str(p11.rraa_pmb11), textInfo_format)

            worksheet3.write(row3, col3 + 10, p11.inc_varia_cual, textInfo_format)
            worksheet3.write(row3, col3 + 11, p11.camb_pregu_cual, textInfo_format)
            worksheet3.write(row3, col3 + 12, p11.otros_aprov_ra, textInfo_format)

            worksheet3.write(row3, col3 + 13, p11.nueva_oe, textInfo_format)
            worksheet3.write(row3, col3 + 14, p11.indi_cual, textInfo_format)
            worksheet3.write(row3, col3 + 15, p11.gen_nueva, textInfo_format)
            
            row3 += 1

        ################################## Hoja 4 Comentarios ###################
        worksheet4 = workbook.add_worksheet(name="Comentarios")

        comenta = Comentariosddi.objects.all()

        row4 = 3
        col4 = 0
        ## altura de celda
        worksheet4.set_row(0, 40)
        worksheet4.set_row(1, 30)
        worksheet4.set_row(2, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet4.set_column('A:A', 40)
        worksheet4.set_column('B:B', 40)
        worksheet4.set_column('C:C', 40)
        worksheet4.set_column('D:D', 40)
        worksheet4.set_column('E:E', 40)
        

        worksheet4.merge_range('A1:E1', "Inventario de Demandas de Información", Backgroundcolor)
        worksheet4.conditional_format('A1:E1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Demandas de Información',
                                                'format': cell_format})


        worksheet4.merge_range('A2:E2', "COMENTARIOS", Backgroundcolor2)
        worksheet4.conditional_format('A2:E2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "COMENTARIOS",
                                                'format': cell_format})
    
        worksheet4.write(2, 0, "Código", Backgroundcolor3)
        worksheet4.write(2, 1, "¿Cuál es el indicador o requerimiento de información estadística?", Backgroundcolor3)
        worksheet4.write(2, 2, "Usuario que realiza el comentario", Backgroundcolor3)
        worksheet4.write(2, 3, "Comentario", Backgroundcolor3)
        worksheet4.write(2, 4, "Fecha en que se realiza el comentario", Backgroundcolor3)
       
        for comen in comenta:
            worksheet4.write(row4, col4, str(comen.post_ddi.codigo_ddi), textInfo_format)
            worksheet4.write(row4, col4 + 1, str(comen.post_ddi), textInfo_format)
            worksheet4.write(row4, col4 + 2, comen.name, textInfo_format)
            worksheet4.write(row4, col4 + 3, comen.body, textInfo_format)
            worksheet4.write(row4, col4 + 4, comen.created_on.replace(tzinfo=None), formatfecha)

            row4 += 1

        ################################## Hoja 5 MÓDULO DE ACTUALIZACIÓN - Novedades ################### 
        worksheet5 = workbook.add_worksheet(name="Actualizacion-Novedades")

        nove = NovedadActualizacionddi.objects.all()

        row5 = 3
        col5 = 0
        ## altura de celda
        worksheet5.set_row(0, 40)
        worksheet5.set_row(1, 30)
        worksheet5.set_row(2, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet5.set_column('A:A', 40)
        worksheet5.set_column('B:B', 40)
        worksheet5.set_column('C:C', 40)
        worksheet5.set_column('D:D', 40)
        worksheet5.set_column('E:E', 40)
        worksheet5.set_column('F:F', 40)
        worksheet5.set_column('G:G', 40)
        

        worksheet5.merge_range('A1:G1', "Inventario de Demandas de Información", Backgroundcolor)
        worksheet5.conditional_format('A1:G1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Demandas de Información',
                                                'format': cell_format})


        worksheet5.merge_range('A2:G2', "MÓDULO DE ACTUALIZACIÓN - NOVEDADES", Backgroundcolor2)
        worksheet5.conditional_format('A2:G2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "MÓDULO DE ACTUALIZACIÓN - NOVEDADES",
                                                'format': cell_format})
    
        worksheet5.write(2, 0, "Código", Backgroundcolor3)
        worksheet5.write(2, 1, "¿Cuál es el indicador o requerimiento de información estadística?", Backgroundcolor3)
        worksheet5.write(2, 2, "Novedad", Backgroundcolor3)
        worksheet5.write(2, 3, "Estado de la actualización", Backgroundcolor3)
        worksheet5.write(2, 4, "Descripción de la novedad", Backgroundcolor3)
        worksheet5.write(2, 5, "Funcionario que realiza la novedad", Backgroundcolor3)
        worksheet5.write(2, 6, "Fecha en que se realiza la novedad", Backgroundcolor3)
       
        for nov in nove:
            worksheet5.write(row5, col5, str(nov.post_ddi.codigo_ddi), textInfo_format)
            worksheet5.write(row5, col5 + 1, str(nov.post_ddi), textInfo_format)
            worksheet5.write(row5, col5 + 2, str(nov.novedad), textInfo_format)
            worksheet5.write(row5, col5 + 3, str(nov.est_actualiz), textInfo_format)
            worksheet5.write(row5, col5 + 4, nov.descrip_novedad, textInfo_format)
            worksheet5.write(row5, col5 + 5, nov.name_nov, textInfo_format)
            worksheet5.write(row5, col5 + 6, nov.fecha_actualiz.replace(tzinfo=None), formatfecha)

            row5 += 1

        ################################## Hoja 6 Critica ################### 
        worksheet6 = workbook.add_worksheet(name="Criticas")

        criti = Criticaddi.objects.all()

        row6 = 3
        col6 = 0
        ## altura de celda
        worksheet6.set_row(0, 40)
        worksheet6.set_row(1, 30)
        worksheet6.set_row(2, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet6.set_column('A:A', 40)
        worksheet6.set_column('B:B', 40)
        worksheet6.set_column('C:C', 40)
        worksheet6.set_column('D:D', 40)
        worksheet6.set_column('E:E', 40)
        worksheet6.set_column('F:F', 40)
    
        worksheet6.merge_range('A1:F1', "Inventario de Demandas de Información", Backgroundcolor)
        worksheet6.conditional_format('A1:F1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Demandas de Información',
                                                'format': cell_format})


        worksheet6.merge_range('A2:F2', "CRÍTICAS", Backgroundcolor2)
        worksheet6.conditional_format('A2:F2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "CRÍTICAS",
                                                'format': cell_format})
    
        worksheet6.write(2, 0, "Código", Backgroundcolor3)
        worksheet6.write(2, 1, "¿Cuál es el indicador o requerimiento de información estadística?", Backgroundcolor3)
        worksheet6.write(2, 2, "Estado de la crítica", Backgroundcolor3)
        worksheet6.write(2, 3, "Observaciones de la crítica", Backgroundcolor3)
        worksheet6.write(2, 4, "Funcionario que realiza la crítica", Backgroundcolor3)
        worksheet6.write(2, 5, "Fecha en que se realiza la crítica", Backgroundcolor3)
        
        for cri in criti:
            worksheet6.write(row6, col6, str(cri.post_ddi.codigo_ddi), textInfo_format)
            worksheet6.write(row6, col6 + 1, str(cri.post_ddi), textInfo_format)
            worksheet6.write(row6, col6 + 2, str(cri.estado_crit), textInfo_format)
            worksheet6.write(row6, col6 + 3, str(cri.descrip_critica), textInfo_format)
            worksheet6.write(row6, col6 + 4, cri.name_cri, textInfo_format)
            worksheet6.write(row6, col6 + 5, cri.fecha_critica.replace(tzinfo=None), formatfecha)

            row6 += 1

        # Close the workbook before sending the data.
        workbook.close()

        # Rewind the buffer.
        output.seek(0)

        # Set up the Http response.
        filename = 'inventario_ddi.xlsx'
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        return response


### reporte inventarios ddi en estado publicado
@method_decorator(login_required, name='dispatch')
class reporteDDIPublicados(View):
    def get(self, request):
        
        # Get some data to write to the spreadsheet.
        demandas = demandaInfor.objects.filter(nombre_est=5)
        user = request.user
        
        preg3_array = []
        preg4_array = []
        preg5_array = []
        preg6_array = []
        preg8_array = []
        preg9_array = []
        preg11_array = []
        preg12_array = []
        preg13_array = []
        preg14_array = []
        comentarios_array = []
        novedades_array = []
        criticas_array = []


        objArray = list(demandas.values('id'))
        for obj in objArray:
            for key, value in obj.items():

                utilizaInfEstTextList = UtilizaInfEstText.objects.filter(ddi_id=value)
                union_preg3 = list(chain(utilizaInfEstTextList))
                preg3_array.extend(union_preg3)

                UsuariosPrincInfEstTextList = UsuariosPrincInfEstText.objects.filter(ddi_id=value)
                union_preg4 = list(chain(UsuariosPrincInfEstTextList))
                preg4_array.extend(union_preg4)

                NormasTextList = NormasText.objects.filter(ddi_id=value)
                union_preg5 = list(chain(NormasTextList))
                preg5_array.extend(union_preg5)

                InfSolRespRequerimientoTextList = InfSolRespRequerimientoText.objects.filter(ddi_id=value)
                union_preg6 = list(chain(InfSolRespRequerimientoTextList))
                preg6_array.extend(union_preg6)

                TipoRequerimientoTextList = TipoRequerimientoText.objects.filter(ddi_id=value)
                union_preg8 = list(chain(TipoRequerimientoTextList))
                preg8_array.extend(union_preg8)

                listaVariablesList = listaVariables.objects.filter(ddi_id=value)
                union_preg9 = list(chain(listaVariablesList))
                preg9_array.extend(union_preg9)

                SolReqInformacionList = SolReqInformacion.objects.filter(post_ddi_id=value)
                union_preg11 = list(chain(SolReqInformacionList))
                preg11_array.extend(union_preg11)

                DesagregacionReqTextList = DesagregacionReqText.objects.filter(ddi_id=value)
                union_preg12 = list(chain(DesagregacionReqTextList))
                preg12_array.extend(union_preg12)

                DesagregacionGeoReqGeograficaTextList = DesagregacionGeoReqGeograficaText.objects.filter(ddi_id=value)
                union_preg13 = list(chain(DesagregacionGeoReqGeograficaTextList))
                preg13_array.extend(union_preg13)

                PeriodicidadDifusionTextList = PeriodicidadDifusionText.objects.filter(ddi_id=value)
                union_preg14 = list(chain(PeriodicidadDifusionTextList))
                preg14_array.extend(union_preg14)

                ComentariosList = Comentariosddi.objects.filter(post_ddi_id=value)
                union_comentarios = list(chain(ComentariosList))
                comentarios_array.extend(union_comentarios)

                novedadesList = NovedadActualizacionddi.objects.filter(post_ddi_id=value)
                union_novedades = list(chain(novedadesList))
                novedades_array.extend(union_novedades)

                criticasList = Criticaddi.objects.filter(post_ddi_id=value)
                union_criticas = list(chain(novedadesList))
                criticas_array.extend(union_criticas)

        # Create an in-memory output file for the new workbook.
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet(name="Inventario de demandas")
        
        # Get some data to write to the spreadsheet.
        row = 4
        col = 0
        ## altura de celda
        worksheet.set_row(0, 40)
        worksheet.set_row(1, 30)
        worksheet.set_row(2, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet.set_column('A:A', 30)
        worksheet.set_column('B:B', 30)
        worksheet.set_column('C:C', 30)
        worksheet.set_column('D:D', 30)
        worksheet.set_column('E:E', 30)
        worksheet.set_column('F:F', 30)
        worksheet.set_column('G:G', 40)
        worksheet.set_column('H:H', 30)
        worksheet.set_column('I:I', 30)
        worksheet.set_column('J:J', 30)
        worksheet.set_column('K:K', 30)
        worksheet.set_column('L:L', 30)
        worksheet.set_column('M:M', 30)
        worksheet.set_column('N:N', 30)
        worksheet.set_column('O:O', 30)
        worksheet.set_column('P:P', 30)
        worksheet.set_column('Q:Q', 30)
        worksheet.set_column('R:R', 30)
        worksheet.set_column('S:S', 30)
        worksheet.set_column('T:T', 30)
        worksheet.set_column('U:U', 40)
        worksheet.set_column('V:V', 40)
        worksheet.set_column('W:W', 30)
        worksheet.set_column('X:X', 30)
        worksheet.set_column('Y:Y', 30)
        worksheet.set_column('Z:Z', 30)

        worksheet.set_column('AA:AA', 40)
        worksheet.set_column('AB:AB', 40)
        worksheet.set_column('AC:AC', 30)
        worksheet.set_column('AD:AD', 30)
        worksheet.set_column('AE:AE', 30)
        worksheet.set_column('AF:AF', 30)
        worksheet.set_column('AG:AG', 30)
        worksheet.set_column('AH:AH', 30)
        worksheet.set_column('AI:AI', 30)
        worksheet.set_column('AJ:AJ', 30)
        worksheet.set_column('AK:AK', 30)
        worksheet.set_column('AL:AL', 30)
        worksheet.set_column('AM:AM', 30)
        worksheet.set_column('AN:AN', 30)
        worksheet.set_column('AO:AO', 30)
        worksheet.set_column('AP:AP', 30)
        worksheet.set_column('AQ:AQ', 30)
        worksheet.set_column('AR:AR', 30)
        worksheet.set_column('AS:AS', 30)
        worksheet.set_column('AT:AT', 30)
        worksheet.set_column('AU:AU', 40)
        worksheet.set_column('AV:AV', 40)
        worksheet.set_column('AW:AW', 40)
        worksheet.set_column('AX:AX', 30)
        worksheet.set_column('AY:AY', 30)
        worksheet.set_column('AZ:AZ', 30)

        worksheet.set_column('BA:BA', 40)
        worksheet.set_column('BB:BB', 40)
        worksheet.set_column('BC:BC', 30)
        worksheet.set_column('BD:BD', 30)
        worksheet.set_column('BE:BE', 30)
        worksheet.set_column('BF:BF', 30)
        worksheet.set_column('BG:BG', 40)
        worksheet.set_column('BH:BH', 30)
        worksheet.set_column('BI:BI', 30)
        worksheet.set_column('BJ:BJ', 30)
        worksheet.set_column('BK:BK', 30)
        worksheet.set_column('BL:BL', 30)
        worksheet.set_column('BM:BM', 30)
        worksheet.set_column('BN:BN', 30)
        worksheet.set_column('BO:BO', 30)
        worksheet.set_column('BP:BP', 30)
        worksheet.set_column('BQ:BQ', 30)
        worksheet.set_column('BR:BR', 30)
        worksheet.set_column('BS:BS', 30)
        worksheet.set_column('BT:BT', 30)
        worksheet.set_column('BU:BU', 30)
        worksheet.set_column('BV:BV', 30)
        worksheet.set_column('BW:BW', 30)
        worksheet.set_column('BX:BX', 30)
        worksheet.set_column('BY:BY', 30)
        worksheet.set_column('BZ:BZ', 30)

        worksheet.set_column('CA:CA', 40)
        worksheet.set_column('CB:CB', 40)
        worksheet.set_column('CC:CC', 30)
        worksheet.set_column('CD:CD', 30)
        worksheet.set_column('CE:CE', 30)
        worksheet.set_column('CF:CF', 40)
        worksheet.set_column('CG:CG', 40)
        worksheet.set_column('CH:CH', 30)
        worksheet.set_column('CI:CI', 30)
        worksheet.set_column('CJ:CJ', 30)
        worksheet.set_column('CK:CK', 30)
        worksheet.set_column('CL:CL', 30)
        worksheet.set_column('CM:CM', 30)
        worksheet.set_column('CN:CN', 30)
        worksheet.set_column('CO:CO', 30)
        worksheet.set_column('CP:CP', 30)
        worksheet.set_column('CQ:CQ', 30)
        worksheet.set_column('CR:CR', 30)
        worksheet.set_column('CS:CS', 30)
        worksheet.set_column('CT:CT', 30)
        worksheet.set_column('CU:CU', 30)
        worksheet.set_column('CV:CV', 30)
        worksheet.set_column('CW:CW', 30)
        worksheet.set_column('CX:CX', 30)
        worksheet.set_column('CY:CY', 30)
        worksheet.set_column('CZ:CZ', 30)

        worksheet.set_column('DA:DA', 30)
        worksheet.set_column('DB:DB', 30)
        worksheet.set_column('DC:DC', 30)
       
        ## agregar estilos
        Backgroundcolor = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 18,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        Backgroundcolor2 = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 14,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        Backgroundcolor3 = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 12,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        cell_format = workbook.add_format()
        cell_format.set_font_color('white')

        textInfo_format = workbook.add_format({'text_wrap': True, 'align': 'center', 'valign': 'vcenter'})
        formatfecha = workbook.add_format({'num_format': 'dd/mm/yy', 'align': 'center', 'valign': 'vcenter'})

        ##FILA 1
        worksheet.merge_range('A1:DC1', "Inventario de Demandas de Información", Backgroundcolor)
        worksheet.conditional_format('A1:DC1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Demandas de Información',
                                                'format': cell_format})

        ##FILA 2
        worksheet.merge_range('A2:A3', " ", Backgroundcolor2)
        worksheet.conditional_format('A2:A3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': ' ',
                                                'format': cell_format })

        worksheet.merge_range('B2:B3', " ", Backgroundcolor2)
        worksheet.conditional_format('B2:B3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': ' ',
                                                'format': cell_format })

        worksheet.merge_range('C2:T2', "MÓDULO A. IDENTIFICACIÓN ", Backgroundcolor2)
        worksheet.conditional_format('C2:T2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'MÓDULO A. IDENTIFICACIÓN',
                                                'format': cell_format })

        
        worksheet.merge_range('U2:CU2', "MÓDULO B. DETECCIÓN Y ANÁLISIS DE REQUERIMIENTOS", Backgroundcolor2)
        worksheet.conditional_format('U2:CU2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'MÓDULO B. DETECCIÓN Y ANÁLISIS DE REQUERIMIENTOS',
                                                'format': cell_format })


        worksheet.merge_range('CX2:DA2', "COMPLEMENTO DE DEMANDA", Backgroundcolor3)
        worksheet.conditional_format('CX2:DA2', {'type': 'text',
                                        'criteria': 'begins with',
                                        'value': 'COMPLEMENTO DE DEMANDA',
                                        'format': cell_format })


        worksheet.merge_range('DB2:DC2', "MÓDULO VALIDACIÓN DE LA DEMANDA", Backgroundcolor3)
        worksheet.conditional_format('DB2:DC2', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'MÓDULO VALIDACIÓN DE LA DEMANDA',
                                                'format': cell_format })


        ##FILA 3
        worksheet.merge_range('C3:E3', "ÁREA TEMÁTICA / TEMA", Backgroundcolor2)
        worksheet.conditional_format('C3:E3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'ÁREA TEMÁTICA / TEMA',
                                                'format': cell_format })

        
        worksheet.merge_range('F3:F4',  "Seleccione el comité estadístico sectorial al que pertenece la demanda de información estadística", Backgroundcolor3)
        worksheet.conditional_format('F3:F4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Seleccione el comité estadístico sectorial al que pertenece la demanda de información estadística',
                                                'format': cell_format })

        worksheet.merge_range('G3:N3',  "¿Quién identificó la demanda de información?", Backgroundcolor3)
        worksheet.conditional_format('G3:N3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '¿Quién identificó la demanda de información?',
                                                'format': cell_format })

        worksheet.merge_range('O3:Z3',  "A. Datos del Solicitante", Backgroundcolor3)
        worksheet.conditional_format('O3:AZ3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'A. Datos del Solicitante',
                                                'format': cell_format })

        worksheet.merge_range('AA3:AA4',  "1. ¿Cuál es el indicador o requerimiento de información estadística?", Backgroundcolor3)
        worksheet.conditional_format('AA3:AA4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '1. ¿Cuál es el indicador o requerimiento de información estadística?',
                                                'format': cell_format })

        worksheet.merge_range('AB3:AB4',  "2. Descripción general de la información que requiere", Backgroundcolor3)
        worksheet.conditional_format('AB3:AB4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '2. Descripción general de la información que requiere',
                                                'format': cell_format })

        worksheet.merge_range('AC3:AG3',  "3. ¿Para qué se utiliza la información estadística?", Backgroundcolor3)
        worksheet.conditional_format('AC3:AG3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '3. ¿Para qué se utiliza la información estadística?',
                                                'format': cell_format })

        worksheet.merge_range('AH3:AS3',  "4. ¿Cuáles podrían ser los usuarios principales de la información solicitada?", Backgroundcolor3)
        worksheet.conditional_format('AH3:AS3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '4. ¿Cuáles podrían ser los usuarios principales de la información solicitada?',
                                                'format': cell_format })

        worksheet.merge_range('AT3:AX3',  "5. La información solicitada responde a las siguientes normas", Backgroundcolor3)
        worksheet.conditional_format('AT3:AX3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '5. La información solicitada responde a las siguientes normas',
                                                'format': cell_format })

        worksheet.merge_range('AY3:BE3',  "6. La información solicitada responde a los siguientes requerimientos", Backgroundcolor3)
        worksheet.conditional_format('AY3:BE3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '6. La información solicitada responde a los siguientes requerimientos',
                                                'format': cell_format })

        worksheet.merge_range('BF3:BL3',  "7. ¿La información requerida está siendo producida en su totalidad?", Backgroundcolor3)
        worksheet.conditional_format('BF3:BL3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '7. ¿La información requerida está siendo producida en su totalidad?',
                                                'format': cell_format })

        worksheet.merge_range('BM3:BQ3',  "8. ¿Qué tipo de requerimiento es?", Backgroundcolor3)
        worksheet.conditional_format('BF3:BQ3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '8. ¿Qué tipo de requerimiento es?',
                                                'format': cell_format })


        worksheet.merge_range('BR3:BR4',  "9. ¿Qué variables necesita para suplir el requerimiento? (Ver hoja 2 - Lista de variables)", Backgroundcolor3)
        worksheet.conditional_format('BR3:BR4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '9. ¿Qué variables necesita para suplir el requerimiento? (Ver hoja 2 - Lista de variables)',
                                                'format': cell_format })

        worksheet.merge_range('BS3:BT3',  "10. ¿Cuál entidad considera que debe producir la información requerida?", Backgroundcolor3)
        worksheet.conditional_format('BS3:BT3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '10. ¿Cuál entidad considera que debe producir la información requerida?',
                                                'format': cell_format })


        worksheet.merge_range('BU3:BX3',  "11. ¿Cuál(es) de las siguientes opciones podrían dar solución al requerimiento de información? (Ver hoja 3 - pregunta 11)", Backgroundcolor3)
        worksheet.conditional_format('BU3:BX3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '11. ¿Cuál(es) de las siguientes opciones podrían dar solución al requerimiento de información? (Ver hoja 3 - pregunta 11)',
                                                'format': cell_format })

        
        worksheet.merge_range('BY3:CE3',  "12. Indique la desagregación requerida:", Backgroundcolor3)
        worksheet.conditional_format('BY3:CE3', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '12. Indique la desagregación requerida:',
                                                'format': cell_format })


        worksheet.merge_range('CF3:CP3', "13. Indique la desagregación requerida", Backgroundcolor3)
        worksheet.conditional_format('CF3:CP3', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '13. Indique la desagregación requerida',
                                                'format': cell_format })

        worksheet.merge_range('CQ3:CU3', "14. Periodicidad de difusión requerida", Backgroundcolor3)
        worksheet.conditional_format('CQ3:CU3', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '14. Periodicidad de difusión requerida',
                                                'format': cell_format })


        worksheet.merge_range('CV2:CV4', "MÓDULO C. OBSERVACIONES", Backgroundcolor3)
        worksheet.conditional_format('CV2:CV4', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'MÓDULO C. OBSERVACIONES',
                                                'format': cell_format })


        worksheet.merge_range('CW2:CW4', "MÓDULO D. ANEXOS", Backgroundcolor3)
        worksheet.conditional_format('CW2:CW4', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'MÓDULO D. ANEXOS',
                                                'format': cell_format })


        worksheet.merge_range('CX3:CY3', "¿Es un demanda insatisfecha priorizada?", Backgroundcolor3)
        worksheet.conditional_format('CX3:CY3', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '¿Es un demanda insatisfecha priorizada?',
                                                'format': cell_format })

        worksheet.merge_range('CZ3:DA3', "¿Es un demanda insatisfecha priorizada?", Backgroundcolor3)
        worksheet.conditional_format('CZ3:DA3', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '¿Es un demanda insatisfecha priorizada?',
                                                'format': cell_format })


        worksheet.merge_range('DB3:DC3', "¿El requerimiento reportado es una demanda de información estadística?", Backgroundcolor3)
        worksheet.conditional_format('DB3:DC3', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '¿El requerimiento reportado es una demanda de información estadística?',
                                                'format': cell_format })

        worksheet.write(3, 0, "Código", Backgroundcolor3)
        worksheet.write(3, 1, "Estado", Backgroundcolor3)
        worksheet.write(3, 2, "Área temática", Backgroundcolor3)
        worksheet.write(3, 3, "Tema principal", Backgroundcolor3)
        worksheet.write(3, 4, "Tema compartido", Backgroundcolor3)

        worksheet.write(3, 6, "Consejo Asesor Técnico del SEN", Backgroundcolor3)
        worksheet.write(3, 7, "Comité Estadístico Sectorial", Backgroundcolor3)
        worksheet.write(3, 8, "Mesa Estadística Sectorial", Backgroundcolor3)
        worksheet.write(3, 9, "Entidad", Backgroundcolor3)
        worksheet.write(3, 10, "¿Cuales?", Backgroundcolor3)
        worksheet.write(3, 11, "Otra entidad", Backgroundcolor3)
        worksheet.write(3, 12, "DANE", Backgroundcolor3)
        worksheet.write(3, 13, "Otras instancias de coordinación", Backgroundcolor3)
        
        worksheet.write(3, 14, "Código de la entidad", Backgroundcolor3)
        worksheet.write(3, 15, "Nombre de la entidad", Backgroundcolor3)
        worksheet.write(3, 16, "Entidad consumidora de información", Backgroundcolor3)
        worksheet.write(3, 17, "Dependencia", Backgroundcolor3)
        worksheet.write(3, 18, "Nombre del jefe de la dependencia", Backgroundcolor3)
        worksheet.write(3, 19, "Cargo", Backgroundcolor3)
        worksheet.write(3, 20, "Correo electrónico", Backgroundcolor3)
        worksheet.write(3, 21, "Teléfono jefe dependencia", Backgroundcolor3)
        worksheet.write(3, 22, "Persona que realiza el requerimiento", Backgroundcolor3)
        worksheet.write(3, 23, "Cargo", Backgroundcolor3)
        worksheet.write(3, 24, "Correo electrónico", Backgroundcolor3)
        worksheet.write(3, 25, "Teléfono", Backgroundcolor3)

        worksheet.write(3, 28, "a. Análisis de contexto", Backgroundcolor3)
        worksheet.write(3, 29, "b. Diseño, formulación o seguimiento de políticas", Backgroundcolor3)
        worksheet.write(3, 30, "c. Reportes internacionales", Backgroundcolor3)
        worksheet.write(3, 31, "d. Otro ¿Cuál?", Backgroundcolor3)
        worksheet.write(3, 32, "¿Cuál?", Backgroundcolor3)

        worksheet.write(3, 33, "a. Organismos internacionales", Backgroundcolor3)
        worksheet.write(3, 34, "b. Presidencia de la República", Backgroundcolor3)
        worksheet.write(3, 35, "c. Ministerios", Backgroundcolor3)
        worksheet.write(3, 36, "d. Organismos de Control", Backgroundcolor3)
        worksheet.write(3, 37, "e. Otras entidades del orden Nacional", Backgroundcolor3)
        worksheet.write(3, 38, "f. Entidades de orden Territorial", Backgroundcolor3)
        worksheet.write(3, 39, "g. Gremios", Backgroundcolor3)
        worksheet.write(3, 40, "h. Entidades privadas", Backgroundcolor3)
        worksheet.write(3, 41, "i. Dependencias de la misma entidad", Backgroundcolor3)
        worksheet.write(3, 42, "j. Academia", Backgroundcolor3)
        worksheet.write(3, 43, "k. Público en General", Backgroundcolor3)
        worksheet.write(3, 44, "l. Otra ¿cuál?", Backgroundcolor3)

        worksheet.write(3, 45, "a. Constitución Política", Backgroundcolor3)
        worksheet.write(3, 46, "b. Ley", Backgroundcolor3)
        worksheet.write(3, 47, "c. Decreto (Nacional, departamental, etc.)", Backgroundcolor3)
        worksheet.write(3, 48, "d. Otra (Resolución, ordenanza, acuerdo)", Backgroundcolor3)
        worksheet.write(3, 49, "e. Ninguna", Backgroundcolor3)

        worksheet.write(3, 50, "a. Plan Nacional de Desarrollo (Capítulo y línea)", Backgroundcolor3)
        worksheet.write(3, 51, "b. Cuentas económicas y macroeconómicas", Backgroundcolor3)
        worksheet.write(3, 52, "c. Plan Sectorial, Territorial o CONPES", Backgroundcolor3)
        worksheet.write(3, 53, "d. Objetivos de Desarrollo Sostenible (ODS) ¿Cuál es el número del Objetivo? ¿Cuál es el número del Indicador?", Backgroundcolor3)
        worksheet.write(3, 54, "e. Organización para la Cooperación y el Desarrollo Económicos (OCDE)", Backgroundcolor3)
        worksheet.write(3, 55, "f. Otros compromisos internacionales", Backgroundcolor3)
        worksheet.write(3, 56, "g. Otro(s)", Backgroundcolor3)

        worksheet.write(3, 57, "Si/No", Backgroundcolor3)
        worksheet.write(3, 58, "¿Cuál entidad?", Backgroundcolor3)
        worksheet.write(3, 59, "Otra entidad", Backgroundcolor3)
        worksheet.write(3, 60, "Nombre de la operación estadística", Backgroundcolor3)
        worksheet.write(3, 61, "Otra operación estadística", Backgroundcolor3)
        worksheet.write(3, 62, "Nombre del registro administrativo", Backgroundcolor3)
        worksheet.write(3, 63, "Otra registro administrativo", Backgroundcolor3)

        worksheet.write(3, 64, "a. Agregado estadístico o indicador", Backgroundcolor3)
        worksheet.write(3, 65, "b. Ampliación cobertura geográfica", Backgroundcolor3)
        worksheet.write(3, 66, "c. Ampliación cobertura temática", Backgroundcolor3)
        worksheet.write(3, 67, "d. Ajustes en la difusión", Backgroundcolor3)
        worksheet.write(3, 68, "e. Otro. ¿Cuál?", Backgroundcolor3)

        worksheet.write(3, 70, "Cual entidad", Backgroundcolor3)
        worksheet.write(3, 71, "Otra ¿Cuál?", Backgroundcolor3)

        worksheet.write(3, 72, "a. Aprovechamiento de una operación (es) estadística (s)", Backgroundcolor3)
        worksheet.write(3, 73, "b. Aprovechamiento de un registro administrativo", Backgroundcolor3)
        worksheet.write(3, 74, "c. Generación de nueva información", Backgroundcolor3)
        worksheet.write(3, 75, "d. Otra (s) ¿Cuáles?", Backgroundcolor3)

        worksheet.write(3, 76, "a. Sexo", Backgroundcolor3)
        worksheet.write(3, 77, "b. Edades", Backgroundcolor3)
        worksheet.write(3, 78, "c. Grupos étnicos", Backgroundcolor3)
        worksheet.write(3, 79, "d. Discapacidad", Backgroundcolor3)
        worksheet.write(3, 80, "e. Estrato", Backgroundcolor3)
        worksheet.write(3, 81, "f. Otra ¿Cuál?", Backgroundcolor3)
        worksheet.write(3, 82, "g. Ninguna", Backgroundcolor3)

        worksheet.write(3, 83, "Geográfica", Backgroundcolor3)
        worksheet.write(3, 84, "Nacional", Backgroundcolor3)
        worksheet.write(3, 85, "Regional", Backgroundcolor3)
        worksheet.write(3, 86, "Departamental", Backgroundcolor3)
        worksheet.write(3, 87, "Áreas metropolitanas", Backgroundcolor3)
        worksheet.write(3, 88, "Municipal", Backgroundcolor3)
        worksheet.write(3, 89, "Otra, ¿cuál?", Backgroundcolor3)
        worksheet.write(3, 90, "Zona", Backgroundcolor3)
        worksheet.write(3, 91, "Total", Backgroundcolor3)
        worksheet.write(3, 92, "Urbano", Backgroundcolor3)
        worksheet.write(3, 93, "Rural", Backgroundcolor3)

        worksheet.write(3, 94, "a. Anual", Backgroundcolor3)
        worksheet.write(3, 95, "b. Semestral", Backgroundcolor3)
        worksheet.write(3, 96, "c. Trimestral", Backgroundcolor3)
        worksheet.write(3, 97, "d. Mensual", Backgroundcolor3)
        worksheet.write(3, 98, "e. Otra ¿Cuál?", Backgroundcolor3)

        ##COMPLEMENTO DE DEMANDA
        worksheet.write(3, 101, "SI/NO", Backgroundcolor3)
        worksheet.write(3, 102, "¿Instancia de coordinación o entidad que lo priorizó?", Backgroundcolor3)
        worksheet.write(3, 103, "SI/NO", Backgroundcolor3)
        worksheet.write(3, 104, "Describa de forma breve la acción que suple la demanda insatisfecha", Backgroundcolor3)

        ##VALIDACIÓN DE LA DEMANDA
        worksheet.write(3, 105, "SI/NO", Backgroundcolor3)
        worksheet.write(3, 106, "Describa de forma breve", Backgroundcolor3)

        for dem in demandas:
            worksheet.write(row, col, dem.codigo_ddi, textInfo_format)
            worksheet.write(row, col + 1, str(dem.nombre_est), textInfo_format)
            worksheet.write(row, col + 2, str(dem.area_tem), textInfo_format)
            worksheet.write(row, col + 3, str(dem.tema_prin), textInfo_format)
            
            lista_temaComp = dem.tema_comp.all()
            lista_temaComp_array = []
            for index, item in enumerate (lista_temaComp):
                lista_temaComp_array.append(item)
                worksheet.write(row, col + 4, str(lista_temaComp_array).replace('[','').replace(']','').replace("<TemaCompartido:", "").replace(">", ""), textInfo_format)
               
            lista_comiteEst = dem.comite_est.all()
            lista_comiteEst_array = []
            for index, item in enumerate (lista_comiteEst):
                lista_comiteEst_array.append(item)
                worksheet.write(row, col + 5, str(lista_comiteEst_array).replace('[','').replace(']','').replace("<ComiteEstSect:", "").replace(">", ""), textInfo_format) 
            
            lista_quien_identddi = dem.quien_identddi.all()
            for index, item in enumerate(lista_quien_identddi):
                if str(item) == "Consejo Asesor Técnico del SEN":
                    worksheet.write(row, col + 6, str(item), textInfo_format)

                if str(item) == "Comité Estadístico Sectorial":
                    worksheet.write(row, col + 7, str(item), textInfo_format)

                if str(item) == "Mesa Estadística Sectorial":
                    worksheet.write(row, col + 8, str(item), textInfo_format)

                if str(item) == "Entidad":
                    worksheet.write(row, col + 9, str(item), textInfo_format)

                if str(item) == "DANE":
                    worksheet.write(row, col + 12, str(item), textInfo_format)

                if str(item) == "Otras instancias de coordinación":
                    worksheet.write(row, col + 13, str(item), textInfo_format)

            lista_entidad_qiddi = dem.entidad_qiddi.all()
            lista_entidad_qiddi_array = []
            for index, item in enumerate (lista_entidad_qiddi):
                lista_entidad_qiddi_array .append(item)
                worksheet.write(row, col + 10, str(lista_entidad_qiddi_array ).replace('[','').replace(']','').replace("<entidad_service:", "").replace(">", ""), textInfo_format)

            worksheet.write(row, col + 11, str(dem.otra_entidad_qiddi), textInfo_format)

            if dem.entidad_sol == None:
                worksheet.write(row, col + 14, "", textInfo_format)
            else:
                worksheet.write(row, col + 14, str(dem.entidad_sol.codigo), textInfo_format)

            if dem.entidad_sol == None:
                worksheet.write(row, col + 15, "", textInfo_format)
            else: 
                worksheet.write(row, col + 15, str(dem.entidad_sol), textInfo_format)

            if dem.entidad_cons_sol == None:
                worksheet.write(row, col + 16, "", textInfo_format)
            else:
                worksheet.write(row, col + 16, str(dem.entidad_cons_sol), textInfo_format)


            worksheet.write(row, col + 17, str(dem.dependencia), textInfo_format)
            worksheet.write(row, col + 18, str(dem.nombre_jef_dep), textInfo_format)
            worksheet.write(row, col + 19, str(dem.cargo_jef_dep), textInfo_format)
            worksheet.write(row, col + 20, str(dem.correo_jef_dep), textInfo_format)

            if dem.telefono_jef_dep == None:
                worksheet.write(row, col + 21, "", textInfo_format)
            else:
                worksheet.write(row, col + 21, str(dem.telefono_jef_dep), textInfo_format)


            worksheet.write(row, col + 22, str(dem.pers_req), textInfo_format)
            worksheet.write(row, col + 23, str(dem.cargo_pers_req), textInfo_format)
            worksheet.write(row, col + 24, str(dem.correo_pers_req), textInfo_format)

            if dem.telefono_pers_req == None:
                worksheet.write(row, col + 25, "", textInfo_format)
            else:
                worksheet.write(row, col + 25, str(dem.telefono_pers_req), textInfo_format)
            #MÓDULO B
            worksheet.write(row, col + 26, str(dem.pm_b_1), textInfo_format)
            worksheet.write(row, col + 27, str(dem.pm_b_2), textInfo_format)
            
            lista_pm_b_3 = dem.pm_b_3.all()
            for index_pmb3, item_pmb3 in enumerate(lista_pm_b_3):
                if str(item_pmb3) == "Análisis de contexto":
                    worksheet.write(row, col + 28, str(item_pmb3), textInfo_format)

                if str(item_pmb3) == "Diseño, formulación o seguimiento de políticas":
                    worksheet.write(row, col + 29, str(item_pmb3), textInfo_format)

                if str(item_pmb3) == "Reportes internacionales":
                    worksheet.write(row, col + 30, str(item_pmb3), textInfo_format)

                if str(item_pmb3) == "Otro ¿Cuál?":
                    worksheet.write(row, col + 31, str(item_pmb3), textInfo_format)

            lista_pm_b_4 = dem.pm_b_4.all()
            for index_pm_b_4, item_pm_b_4 in enumerate(lista_pm_b_4):
                
                if str(item_pm_b_4) == "Presidencia de la República":
                    worksheet.write(row, col + 34, str(item_pm_b_4), textInfo_format)

                if str(item_pm_b_4) == "Público en General":
                    worksheet.write(row, col + 43, str(item_pm_b_4), textInfo_format)

            lista_pm_b_5 = dem.pm_b_5.all()
            for index_pm_b_5, item_pm_b_5 in enumerate(lista_pm_b_5):
                
                if str(item_pm_b_5) == "Ninguna":
                    worksheet.write(row, col + 49, str(item_pm_b_5), textInfo_format)

            lista_pm_b_7 = dem.total_si_no_pmb7.all()
            lista_pm_b_7_array = []
            for index_pm_b_7, item_pm_b_7 in enumerate(lista_pm_b_7):
                lista_pm_b_7_array.append(item_pm_b_7)
                worksheet.write(row, col + 57, str(lista_pm_b_7_array).replace('[','').replace(']','').replace("<InfoProdTotalidad:", "").replace(">", ""), textInfo_format)

            lista_entidad_pm_b7 = dem.entidad_pm_b7.all()
            lista_entidad_pm_b7_array = []
            for index, item in enumerate(lista_entidad_pm_b7):
                lista_entidad_pm_b7_array.append(item)
                worksheet.write(row, col + 58, str(lista_entidad_pm_b7_array).replace('[','').replace(']','').replace("<entidad_service:", "").replace(">", ""), textInfo_format)
            
            worksheet.write(row, col + 59, str(dem.otro_entidad_pm_b7), textInfo_format)

            lista_ooee_pm_b7 = dem.ooee_pm_b7.all()
            lista_ooee_pm_b7_array = []
            for index, item in enumerate(lista_ooee_pm_b7):
                lista_ooee_pm_b7_array.append(item)
                worksheet.write(row, col + 60, str(lista_ooee_pm_b7_array).replace('[','').replace(']','').replace("<ooee_service:", "").replace(">", ""), textInfo_format)
            
            worksheet.write(row, col + 61, str(dem.otro_ooee_pm_b7), textInfo_format)
            
            lista_rraa_pm_b7 = dem.rraa_pm_b7.all()
            lista_rraa_pm_b7_array = []
            for index, item in enumerate(lista_rraa_pm_b7):
                lista_rraa_pm_b7_array.append(item)
                worksheet.write(row, col + 62, str(lista_rraa_pm_b7_array).replace('[','').replace(']','').replace("<rraa_service:", "").replace(">", ""), textInfo_format)

            worksheet.write(row, col + 63, str(dem.otro_rraa_pm_b7), textInfo_format)
                
            lista_pm_b_8 = dem.pm_b_8.all()
            for index_pm_b_8, item_pm_b_8 in enumerate(lista_pm_b_8):
                
                if str(item_pm_b_8) == "Agregado estadístico o indicador":
                    worksheet.write(row, col + 64, str(item_pm_b_8), textInfo_format)

                if str(item_pm_b_8) == "Ampliación cobertura geográfica":
                    worksheet.write(row, col + 65, str(item_pm_b_8), textInfo_format)

                if str(item_pm_b_8) == "Ampliación cobertura temática":
                    worksheet.write(row, col + 66, str(item_pm_b_8), textInfo_format)

                if str(item_pm_b_8) == "Ajustes en la difusión":
                    worksheet.write(row, col + 67, str(item_pm_b_8), textInfo_format)

            lista_pm_b_10 = dem.pm_b_10.all()
            lista_pm_b_10_array = []
            for index, item in enumerate(lista_pm_b_10):
                lista_pm_b_10_array.append(item)
                worksheet.write(row, col + 70, str(lista_pm_b_10_array).replace('[','').replace(']','').replace("<entidad_service:", "").replace(">", ""), textInfo_format)
            
            worksheet.write(row, col + 71, str(dem.pm_b_10_otro), textInfo_format)

            lista_pm_b_11_1 = dem.pm_b_11_1.all()
            for index_pm_b_11_1, item_pm_b_11_1 in enumerate(lista_pm_b_11_1):
                
                if str(item_pm_b_11_1) == "Aprovechamiento de una operación (es) estadística (s)":
                    worksheet.write(row, col + 72, str(item_pm_b_11_1), textInfo_format)

                if str(item_pm_b_11_1) == "Aprovechamiento de un registro administrativo":
                    worksheet.write(row, col + 73, str(item_pm_b_11_1), textInfo_format)

                if str(item_pm_b_11_1) == "Generación de nueva información":
                    worksheet.write(row, col + 74, str(item_pm_b_11_1), textInfo_format)

            worksheet.write(row, col + 75, str(dem.otra_cual_11_1_d), textInfo_format)

            lista_pm_b_12 = dem.pm_b_12.all()
            for index_pm_b_12, item_pm_b_12 in enumerate(lista_pm_b_12):
                
                if str(item_pm_b_12) == "Sexo":
                    worksheet.write(row, col + 76, str(item_pm_b_12), textInfo_format)

                if str(item_pm_b_12) == "Edades":
                    worksheet.write(row, col + 77, str(item_pm_b_12), textInfo_format)

                if str(item_pm_b_12) == "Grupos étnicos":
                    worksheet.write(row, col + 78, str(item_pm_b_12), textInfo_format)

                if str(item_pm_b_12) == "Discapacidad":
                    worksheet.write(row, col + 79, str(item_pm_b_12), textInfo_format)

                if str(item_pm_b_12) == "Estrato":
                    worksheet.write(row, col + 80, str(item_pm_b_12), textInfo_format)

                if str(item_pm_b_12) == "Ninguna":
                    worksheet.write(row, col + 82, str(item_pm_b_12), textInfo_format)

            lista_pm_b_13 = dem.pm_b_13.all()
            for index_pm_b_13, item_pm_b_13 in enumerate(lista_pm_b_13):
                
                if str(item_pm_b_13) == "Geográfica":
                    worksheet.write(row, col + 83, str(item_pm_b_13), textInfo_format)

                if str(item_pm_b_13) == "Zona":
                    worksheet.write(row, col + 90, str(item_pm_b_13), textInfo_format)


            lista_pm_b_13_geo = dem.pm_b_13_geo.all()
            for index_pm_b_13_geo, item_pm_b_13_geo in enumerate(lista_pm_b_13_geo):

                if str(item_pm_b_13_geo) == "Nacional":
                    worksheet.write(row, col + 84, str(item_pm_b_13_geo), textInfo_format)

                if str(item_pm_b_13_geo) == "Regional":
                    worksheet.write(row, col + 85, str(item_pm_b_13_geo), textInfo_format)

                if str(item_pm_b_13_geo) == "Departamental":
                    worksheet.write(row, col + 86, str(item_pm_b_13_geo), textInfo_format)

                if str(item_pm_b_13_geo) == "Áreas metropolitanas":
                    worksheet.write(row, col + 87, str(item_pm_b_13_geo), textInfo_format)

                if str(item_pm_b_13_geo) == "Municipal":
                    worksheet.write(row, col + 88, str(item_pm_b_13_geo), textInfo_format)

            lista_pm_b_13_zona = dem.pm_b_13_zona.all()
            for index_pm_b_13_zona, item_pm_b_13_zona in enumerate(lista_pm_b_13_zona):

                if str(item_pm_b_13_zona) == "Total":
                    worksheet.write(row, col + 91, str(item_pm_b_13_zona), textInfo_format)

                if str(item_pm_b_13_zona) == "Urbano":
                    worksheet.write(row, col + 92, str(item_pm_b_13_zona), textInfo_format)

                if str(item_pm_b_13_zona) == "Rural":
                    worksheet.write(row, col + 93, str(item_pm_b_13_zona), textInfo_format)
            

            lista_pm_b_14 = dem.pm_b_14.all()
            for index_pm_b_14, item_pm_b_14 in enumerate(lista_pm_b_14):
                
                if str(item_pm_b_14) == "Anual":
                    worksheet.write(row, col + 94, str(item_pm_b_14), textInfo_format)

                if str(item_pm_b_14) == "Semestral":
                    worksheet.write(row, col + 95, str(item_pm_b_14), textInfo_format)

                if str(item_pm_b_14) == "Trimestral":
                    worksheet.write(row, col + 96, str(item_pm_b_14), textInfo_format)

                if str(item_pm_b_14) == "Mensual":
                    worksheet.write(row, col + 97, str(item_pm_b_14), textInfo_format)
            
            ##MÓDULO C
            worksheet.write(row, col + 99, str(dem.pm_c_1), textInfo_format)

            ##MÓDULO D
            worksheet.write(row, col + 100, str(dem.pm_d_1_anexos), textInfo_format)

            ###Complemento de la demanda
            lista_compl_dem_a = dem.compl_dem_a.all()
            lista_compl_dem_a_array = []
            for index, item in enumerate(lista_compl_dem_a):
                lista_compl_dem_a_array.append(item)
                worksheet.write(row, col + 101, str(lista_compl_dem_a_array).replace('[','').replace(']','').replace("<DemandaInsaprio:", "").replace(">", ""), textInfo_format)
            
            worksheet.write(row, col + 102, str(dem.compl_dem_a_text), textInfo_format)
            
            lista_comple_dem_b = dem.comple_dem_b.all()
            lista_comple_dem_b_array = []
            for index, item in enumerate(lista_comple_dem_b):
                lista_comple_dem_b_array.append(item)
                worksheet.write(row, col + 103, str(lista_comple_dem_b_array).replace('[','').replace(']','').replace("<PlanaccionSuplirDem:", "").replace(">", ""), textInfo_format)
            
            worksheet.write(row, col + 104, str(dem.compl_dem_b_text), textInfo_format)

            ###validación de la demanda
            lista_validacion_ddi = dem.validacion_ddi.all()
            lista_validacion_ddi_array = []
            for index, item in enumerate(lista_validacion_ddi):
                lista_validacion_ddi_array.append(item)
                worksheet.write(row, col + 105, str(lista_validacion_ddi_array).replace('[','').replace(']','').replace("<EsDemandadeInfor:", "").replace(">", ""), textInfo_format)
                        
            worksheet.write(row, col + 106, str(dem.validacion_ddi_text), textInfo_format)

            row += 1

        # modulo B campos de textos
        
        for index, item in enumerate(utilizaInfEstTextList):
            
            index = item.ddi_id + 3
            worksheet.write(index, col + 32, str(item.otro_cual), textInfo_format)

        for index, item in enumerate(UsuariosPrincInfEstTextList):
            
            index = item.ddi_id + 3
            worksheet.write(index, col + 33, str(item.orginter_text), textInfo_format)
            worksheet.write(index, col + 35, str(item.ministerios_text), textInfo_format)
            worksheet.write(index, col + 36, str(item.orgcontrol_text), textInfo_format)
            worksheet.write(index, col + 37, str(item.oentidadesordenal_text), textInfo_format)
            worksheet.write(index, col + 38, str(item.entidadesordenterr_text), textInfo_format)
            worksheet.write(index, col + 39, str(item.gremios_text), textInfo_format)
            worksheet.write(index, col + 40, str(item.entiprivadas_text), textInfo_format)
            worksheet.write(index, col + 41, str(item.dependenentidad_text), textInfo_format)
            worksheet.write(index, col + 42, str(item.academia_text), textInfo_format)
            worksheet.write(index, col + 44, str(item.otro_cual_text), textInfo_format)

        for index, item in enumerate(NormasTextList):
            index = item.ddi_id + 3
            worksheet.write(index, col + 45, str(item.const_pol_text), textInfo_format)
            worksheet.write(index, col + 46, str(item.ley_text), textInfo_format)
            worksheet.write(index, col + 47, str(item.decreto_text), textInfo_format)
            worksheet.write(index, col + 48, str(item.otra_text), textInfo_format)

        for index, item in enumerate(InfSolRespRequerimientoTextList):
            index = item.ddi_id + 3
            worksheet.write(index, col + 50, str(item.planalDes_text), textInfo_format)
            worksheet.write(index, col + 51, str(item.cuentasecomacroec_text), textInfo_format)
            worksheet.write(index, col + 52, str(item.plansecterrcom_text), textInfo_format)
            worksheet.write(index, col + 53, str(item.objdessost_text), textInfo_format)
            worksheet.write(index, col + 54, str(item.orgcooper_text), textInfo_format)
            worksheet.write(index, col + 55, str(item.otroscomprInt_text), textInfo_format)
            worksheet.write(index, col + 56, str(item.otros_text), textInfo_format)
        
        for index, item in enumerate(TipoRequerimientoTextList):
            index = item.ddi_id + 3
            worksheet.write(index, col + 68, str(item.otros_c_text), textInfo_format)

        for index, item in enumerate(DesagregacionReqTextList):
            index = item.ddi_id + 3
            worksheet.write(index, col + 81, str(item.otra_cual_text_a), textInfo_format)
            
        for index, item in enumerate(DesagregacionGeoReqGeograficaTextList):
            index = item.ddi_id + 3
            worksheet.write(index, col + 89, str(item.otra_cual_text_b), textInfo_format)

        for index, item in enumerate(PeriodicidadDifusionTextList):
            index = item.ddi_id + 3
            worksheet.write(index, col + 98, str(item.otra_cual_text_c), textInfo_format)
        
        ################################## Hoja 2 pregunta 9 lista de variables ###################3 
        worksheet2 = workbook.add_worksheet(name="Lista de variables")
        # Get some data to write to the spreadsheet.
        
        preg9 = listaVariables.objects.all()
        
        row2 = 4
        col2 = 0
        ## altura de celda
        worksheet2.set_row(0, 40)
        worksheet2.set_row(1, 30)
        worksheet2.set_row(2, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet2.set_column('A:A', 40)
        worksheet2.set_column('B:B', 40)
        worksheet2.set_column('C:C', 40)
        

        worksheet2.merge_range('A1:C1', "Inventario de Demandas de Información", Backgroundcolor)
        worksheet2.conditional_format('A1:C1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Demandas de Información',
                                                'format': cell_format})


        worksheet2.merge_range('A2:C2', "MÓDULO B Detección y Análisis de Requerimientos", Backgroundcolor2)
        worksheet2.conditional_format('A2:C2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "MÓDULO B Detección y Análisis de Requerimientos",
                                                'format': cell_format})


        worksheet2.merge_range('A3:C3', "¿Qué variables necesita para suplir el requerimiento?", Backgroundcolor2)
        worksheet2.conditional_format('A2:C2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "¿Qué variables necesita para suplir el requerimiento?",
                                                'format': cell_format})


        
        worksheet2.write(3, 0, "Código", Backgroundcolor3)
        worksheet2.write(3, 1, "¿Cuál es el indicador o requerimiento de información estadística?", Backgroundcolor3)
        worksheet2.write(3, 2, "variables que necesita para suplir el requerimiento", Backgroundcolor3)
       

        for p9 in preg9:
            if p9.ddi.nombre_est_id == 5:
                worksheet2.write(row2, col2, str(p9.ddi.codigo_ddi), textInfo_format)
                worksheet2.write(row2, col2 + 1, str(p9.ddi), textInfo_format)
                worksheet2.write(row2, col2 + 2, p9.lista_varia, textInfo_format)

                row2 += 1
        
        
        ################################## Hoja 3 pregunta 11 ###################
        worksheet3 = workbook.add_worksheet(name="Pregunta 11")

        preg11 = SolReqInformacion.objects.all()

        row3 = 5
        col3 = 0
        ## altura de celda
        worksheet3.set_row(0, 40)
        worksheet3.set_row(1, 30)
        worksheet3.set_row(2, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet3.set_column('A:A', 40)
        worksheet3.set_column('B:B', 40)
        worksheet3.set_column('C:C', 40)
        worksheet3.set_column('D:D', 40)
        worksheet3.set_column('E:E', 40)
        worksheet3.set_column('F:F', 40)
        worksheet3.set_column('G:G', 40)
        worksheet3.set_column('H:H', 40)
        worksheet3.set_column('I:I', 40)
        worksheet3.set_column('J:J', 40)
        worksheet3.set_column('K:K', 40)
        worksheet3.set_column('L:L', 40)
        worksheet3.set_column('M:M', 40)
        worksheet3.set_column('N:N', 40)
        worksheet3.set_column('O:O', 40)
        worksheet3.set_column('P:P', 40)
        
        
        worksheet3.merge_range('A1:P1', "Inventario de Demandas de Información", Backgroundcolor)
        worksheet3.conditional_format('A1:P1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Demandas de Información',
                                                'format': cell_format})

        worksheet3.merge_range('A2:A4', "", Backgroundcolor)
        worksheet3.conditional_format('A2:A4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '',
                                                'format': cell_format})

        worksheet3.merge_range('B2:B4', "", Backgroundcolor)
        worksheet3.conditional_format('B2:B4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '',
                                                'format': cell_format})


        worksheet3.merge_range('C2:P2', "MÓDULO B Detección y Análisis de Requerimientos", Backgroundcolor2)
        worksheet3.conditional_format('C2:P2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "MÓDULO B Detección y Análisis de Requerimientos",
                                                'format': cell_format})


        worksheet3.merge_range('C3:P3', "¿Cuál(es) de las siguientes opciones podrían dar solución al requerimiento de información?", Backgroundcolor2)
        worksheet3.conditional_format('C3:P3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "¿Cuál(es) de las siguientes opciones podrían dar solución al requerimiento de información?",
                                                'format': cell_format})

        worksheet3.merge_range('C4:I4', "a) Aprovechamiento de una operación (es) estadística (s)", Backgroundcolor2)
        worksheet3.conditional_format('C4:I4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "a) Aprovechamiento de una operación (es) estadística (s)",
                                                'format': cell_format})


        worksheet3.merge_range('J4:M4', "b) Aprovechamiento de un registro administrativo", Backgroundcolor2)
        worksheet3.conditional_format('J4:M4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "b) Aprovechamiento de un registro administrativo",
                                                'format': cell_format})

        worksheet3.merge_range('N4:P4', "c) Generación de nueva información", Backgroundcolor2)
        worksheet3.conditional_format('N4:P4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "c) Generación de nueva información",
                                                'format': cell_format})


        
        worksheet3.write(4, 0, "Código", Backgroundcolor3)
        worksheet3.write(4, 1, "¿Cuál es el indicador o requerimiento de información estadística?", Backgroundcolor3)
        
        worksheet3.write(4, 2, "a) 1. ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 3, "a)2.1. Incluir variables/preguntas ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 4, "a)2.2. Cambiar la formulación de alguna(s) pregunta(s) ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 5, "a)2.3. Ampliar la desagregación temática ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 6, "a)2.4. Ampliar la desagregación geográfica ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 7, "a)2.5. Aumentar la frecuencia en la difusión de los resultados ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 8, "a)2.6. Otra ¿Cuál(es)?", Backgroundcolor3)

        worksheet3.write(4, 9, "b)1.  ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 10, "b)2.1.  Inclusión de variables/preguntas", Backgroundcolor3)
        worksheet3.write(4, 11, "b)2.2. Cambios en la formulación de alguna(s) pregunta(s)", Backgroundcolor3)
        worksheet3.write(4, 12, "b)2.3. Otro ", Backgroundcolor3)

        
        worksheet3.write(4, 13, "c)1. Operación estadística nueva", Backgroundcolor3)
        worksheet3.write(4, 14, "c)2. Indicador ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 15, "c)3. Otra  ¿Cuál?", Backgroundcolor3)
			
        for p11 in preg11:

            if p11.post_ddi.nombre_est_id == 5:
                worksheet3.write(row3, col3, str(p11.post_ddi.codigo_ddi), textInfo_format)
                worksheet3.write(row3, col3 + 1, str(p11.post_ddi), textInfo_format)

                if p11.ooee_pmb11 == None:
                    worksheet3.write(row3, col3 + 2, "", textInfo_format)
                else:
                    worksheet3.write(row3, col3 + 2, str(p11.ooee_pmb11), textInfo_format)

                worksheet3.write(row3, col3 + 3, p11.inc_var_cual, textInfo_format)
                worksheet3.write(row3, col3 + 4, p11.cam_preg_cual, textInfo_format)
                worksheet3.write(row3, col3 + 5, p11.am_des_tem_cual, textInfo_format)
                worksheet3.write(row3, col3 + 6, p11.am_des_geo_cual, textInfo_format)
                worksheet3.write(row3, col3 + 7, p11.dif_resul_cual, textInfo_format)
                worksheet3.write(row3, col3 + 8, p11.opc_aprov_cual, textInfo_format)

                if p11.rraa_pmb11 == None:
                    worksheet3.write(row3, col3 + 9, "", textInfo_format)
                else:
                    worksheet3.write(row3, col3 + 9, str(p11.rraa_pmb11), textInfo_format)

                worksheet3.write(row3, col3 + 10, p11.inc_varia_cual, textInfo_format)
                worksheet3.write(row3, col3 + 11, p11.camb_pregu_cual, textInfo_format)
                worksheet3.write(row3, col3 + 12, p11.otros_aprov_ra, textInfo_format)
                


                worksheet3.write(row3, col3 + 13, p11.nueva_oe, textInfo_format)
                worksheet3.write(row3, col3 + 14, p11.indi_cual, textInfo_format)
                worksheet3.write(row3, col3 + 15, p11.gen_nueva, textInfo_format)
                
                row3 += 1

        ################################## Hoja 4 Comentarios ###################
        worksheet4 = workbook.add_worksheet(name="Comentarios")

        comenta = Comentariosddi.objects.all()

        row4 = 3
        col4 = 0
        ## altura de celda
        worksheet4.set_row(0, 40)
        worksheet4.set_row(1, 30)
        worksheet4.set_row(2, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet4.set_column('A:A', 40)
        worksheet4.set_column('B:B', 40)
        worksheet4.set_column('C:C', 40)
        worksheet4.set_column('D:D', 40)
        worksheet4.set_column('E:E', 40)
        

        worksheet4.merge_range('A1:E1', "Inventario de Demandas de Información", Backgroundcolor)
        worksheet4.conditional_format('A1:E1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Demandas de Información',
                                                'format': cell_format})


        worksheet4.merge_range('A2:E2', "COMENTARIOS", Backgroundcolor2)
        worksheet4.conditional_format('A2:E2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "COMENTARIOS",
                                                'format': cell_format})
    
        worksheet4.write(2, 0, "Código", Backgroundcolor3)
        worksheet4.write(2, 1, "¿Cuál es el indicador o requerimiento de información estadística?", Backgroundcolor3)
        worksheet4.write(2, 2, "Usuario que realiza el comentario", Backgroundcolor3)
        worksheet4.write(2, 3, "Comentario", Backgroundcolor3)
        worksheet4.write(2, 4, "Fecha en que se realiza el comentario", Backgroundcolor3)
       
        for comen in comenta:
            if comen.post_ddi.nombre_est_id == 5:
                worksheet4.write(row4, col4, str(comen.post_ddi.codigo_ddi), textInfo_format)
                worksheet4.write(row4, col4 + 1, str(comen.post_ddi), textInfo_format)
                worksheet4.write(row4, col4 + 2, comen.name, textInfo_format)
                worksheet4.write(row4, col4 + 3, comen.body, textInfo_format)
                worksheet4.write(row4, col4 + 4, comen.created_on.replace(tzinfo=None), formatfecha)

                row4 += 1

        ################################## Hoja 5 MÓDULO DE ACTUALIZACIÓN - Novedades ################### 
        worksheet5 = workbook.add_worksheet(name="Actualizacion-Novedades")

        nove = NovedadActualizacionddi.objects.all()

        row5 = 3
        col5 = 0
        ## altura de celda
        worksheet5.set_row(0, 40)
        worksheet5.set_row(1, 30)
        worksheet5.set_row(2, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet5.set_column('A:A', 40)
        worksheet5.set_column('B:B', 40)
        worksheet5.set_column('C:C', 40)
        worksheet5.set_column('D:D', 40)
        worksheet5.set_column('E:E', 40)
        worksheet5.set_column('F:F', 40)
        worksheet5.set_column('G:G', 40)
        

        worksheet5.merge_range('A1:G1', "Inventario de Demandas de Información", Backgroundcolor)
        worksheet5.conditional_format('A1:G1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Demandas de Información',
                                                'format': cell_format})


        worksheet5.merge_range('A2:G2', "MÓDULO DE ACTUALIZACIÓN - NOVEDADES", Backgroundcolor2)
        worksheet5.conditional_format('A2:G2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "MÓDULO DE ACTUALIZACIÓN - NOVEDADES",
                                                'format': cell_format})
    
        worksheet5.write(2, 0, "Código", Backgroundcolor3)
        worksheet5.write(2, 1, "¿Cuál es el indicador o requerimiento de información estadística?", Backgroundcolor3)
        worksheet5.write(2, 2, "Novedad", Backgroundcolor3)
        worksheet5.write(2, 3, "Estado de la actualización", Backgroundcolor3)
        worksheet5.write(2, 4, "Descripción de la novedad", Backgroundcolor3)
        worksheet5.write(2, 5, "Funcionario que realiza la novedad", Backgroundcolor3)
        worksheet5.write(2, 6, "Fecha en que se realiza la novedad", Backgroundcolor3)
       
        for nov in nove:
            if nov.post_ddi.nombre_est_id == 5:
                worksheet5.write(row5, col5, str(nov.post_ddi.codigo_ddi), textInfo_format)
                worksheet5.write(row5, col5 + 1, str(nov.post_ddi), textInfo_format)
                worksheet5.write(row5, col5 + 2, str(nov.novedad), textInfo_format)
                worksheet5.write(row5, col5 + 3, str(nov.est_actualiz), textInfo_format)
                worksheet5.write(row5, col5 + 4, nov.descrip_novedad, textInfo_format)
                worksheet5.write(row5, col5 + 5, nov.name_nov, textInfo_format)
                worksheet5.write(row5, col5 + 6, nov.fecha_actualiz.replace(tzinfo=None), formatfecha)

                row5 += 1

        ################################## Hoja 6 Critica ################### 
        worksheet6 = workbook.add_worksheet(name="Criticas")

        criti = Criticaddi.objects.all()

        row6 = 3
        col6 = 0
        ## altura de celda
        worksheet6.set_row(0, 40)
        worksheet6.set_row(1, 30)
        worksheet6.set_row(2, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet6.set_column('A:A', 40)
        worksheet6.set_column('B:B', 40)
        worksheet6.set_column('C:C', 40)
        worksheet6.set_column('D:D', 40)
        worksheet6.set_column('E:E', 40)
        worksheet6.set_column('F:F', 40)
    
        worksheet6.merge_range('A1:F1', "Inventario de Demandas de Información", Backgroundcolor)
        worksheet6.conditional_format('A1:F1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Demandas de Información',
                                                'format': cell_format})


        worksheet6.merge_range('A2:F2', "CRÍTICAS", Backgroundcolor2)
        worksheet6.conditional_format('A2:F2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "CRÍTICAS",
                                                'format': cell_format})
    
        worksheet6.write(2, 0, "Código", Backgroundcolor3)
        worksheet6.write(2, 1, "¿Cuál es el indicador o requerimiento de información estadística?", Backgroundcolor3)
        worksheet6.write(2, 2, "Estado de la crítica", Backgroundcolor3)
        worksheet6.write(2, 3, "Observaciones de la crítica", Backgroundcolor3)
        worksheet6.write(2, 4, "Funcionario que realiza la crítica", Backgroundcolor3)
        worksheet6.write(2, 5, "Fecha en que se realiza la crítica", Backgroundcolor3)
        
        for cri in criti:
            if cri.post_ddi.nombre_est_id == 5:
                worksheet6.write(row6, col6, str(cri.post_ddi.codigo_ddi), textInfo_format)
                worksheet6.write(row6, col6 + 1, str(cri.post_ddi), textInfo_format)
                worksheet6.write(row6, col6 + 2, str(cri.estado_crit), textInfo_format)
                worksheet6.write(row6, col6 + 3, str(cri.descrip_critica), textInfo_format)
                worksheet6.write(row6, col6 + 4, cri.name_cri, textInfo_format)
                worksheet6.write(row6, col6 + 5, cri.fecha_critica.replace(tzinfo=None), formatfecha)

                row6 += 1

        # Close the workbook before sending the data.
        workbook.close()

        # Rewind the buffer.
        output.seek(0)

        # Set up the Http response.
        filename = 'ddi_publicadas.xlsx'
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        return response


### reporte demandas por filtros
@method_decorator(login_required, name='dispatch')
class reporteDDIporFiltros(View):
    def get(self, request, *args, **kwargs):

        id_requerimiento = request.GET.get('pm_b_6')
        id_area_tematica = request.GET.get('area_tem')
        id_tema = request.GET.get('tema_prin')

        print("que me llega", id_tema)
        user = request.user
        entidad_cod = Entidades_oe.objects.all()

         ## opc 1 es diferente de vacia
        if id_requerimiento != "" and id_area_tematica == "" and  id_tema == "":
            
            demandas = demandaInfor.objects.filter(nombre_est=5).filter(pm_b_6=id_requerimiento)

        ## opc 1 y opc 2 es diferente de vacia
        elif id_requerimiento != "" and id_area_tematica != "" and  id_tema == "":

            demandas = demandaInfor.objects.filter(nombre_est=5).filter(pm_b_6=id_requerimiento).filter(area_tem=id_area_tematica)  
            
        ## opc 1, opc 2 y opc 3 es diferente de vacia
        elif id_requerimiento != "" and id_area_tematica != "" and  id_tema != "":
            
            demandas = demandaInfor.objects.filter(nombre_est=5).filter(pm_b_6=id_requerimiento).filter(area_tem=id_area_tematica).filter(tema_prin=id_tema)

        ## opcion 1 y opcion 3 es diferente de vacia
        elif id_requerimiento != "" and id_area_tematica == "" and  id_tema != "":
            
            demandas = demandaInfor.objects.filter(nombre_est=5).filter(pm_b_6=id_requerimiento).filter(tema_prin=id_tema)
            
        ## opc 2 es diferente de vacia
        elif id_requerimiento == "" and id_area_tematica != "" and  id_tema == "":
            
            demandas = demandaInfor.objects.filter(nombre_est=5).filter(area_tem=id_area_tematica)
        
        ## opc 2 y opc 3 es diferente de vacia
        elif id_requerimiento == "" and id_area_tematica != "" and  id_tema != "":
            
            demandas = demandaInfor.objects.filter(nombre_est=5).filter(area_tem=id_area_tematica).filter(tema_prin=id_tema)

        # opc 3 es diferente de vacia
        elif id_requerimiento == "" and id_area_tematica == "" and  id_tema != "":
           
            demandas = demandaInfor.objects.filter(nombre_est=5).filter(tema_prin=id_tema)
           
        # Get some data to write to the spreadsheet.
        
        preg3_array = []
        preg4_array = []
        preg5_array = []
        preg6_array = []
        preg8_array = []
        preg9_array = []
        preg11_array = []
        preg12_array = []
        preg13_array = []
        preg14_array = []
        comentarios_array = []
        novedades_array = []
        criticas_array = []


        objArray = list(demandas.values('id'))
        for obj in objArray:
            for key, value in obj.items():

                utilizaInfEstTextList = UtilizaInfEstText.objects.filter(ddi_id=value)
                union_preg3 = list(chain(utilizaInfEstTextList))
                preg3_array.extend(union_preg3)

                UsuariosPrincInfEstTextList = UsuariosPrincInfEstText.objects.filter(ddi_id=value)
                union_preg4 = list(chain(UsuariosPrincInfEstTextList))
                preg4_array.extend(union_preg4)

                NormasTextList = NormasText.objects.filter(ddi_id=value)
                union_preg5 = list(chain(NormasTextList))
                preg5_array.extend(union_preg5)

                InfSolRespRequerimientoTextList = InfSolRespRequerimientoText.objects.filter(ddi_id=value)
                union_preg6 = list(chain(InfSolRespRequerimientoTextList))
                preg6_array.extend(union_preg6)

                TipoRequerimientoTextList = TipoRequerimientoText.objects.filter(ddi_id=value)
                union_preg8 = list(chain(TipoRequerimientoTextList))
                preg8_array.extend(union_preg8)

                listaVariablesList = listaVariables.objects.filter(ddi_id=value)
                union_preg9 = list(chain(listaVariablesList))
                preg9_array.extend(union_preg9)

                SolReqInformacionList = SolReqInformacion.objects.filter(post_ddi_id=value)
                union_preg11 = list(chain(SolReqInformacionList))
                preg11_array.extend(union_preg11)

                DesagregacionReqTextList = DesagregacionReqText.objects.filter(ddi_id=value)
                union_preg12 = list(chain(DesagregacionReqTextList))
                preg12_array.extend(union_preg12)

                DesagregacionGeoReqGeograficaTextList = DesagregacionGeoReqGeograficaText.objects.filter(ddi_id=value)
                union_preg13 = list(chain(DesagregacionGeoReqGeograficaTextList))
                preg13_array.extend(union_preg13)

                PeriodicidadDifusionTextList = PeriodicidadDifusionText.objects.filter(ddi_id=value)
                union_preg14 = list(chain(PeriodicidadDifusionTextList))
                preg14_array.extend(union_preg14)

                ComentariosList = Comentariosddi.objects.filter(post_ddi_id=value)
                union_comentarios = list(chain(ComentariosList))
                comentarios_array.extend(union_comentarios)

                novedadesList = NovedadActualizacionddi.objects.filter(post_ddi_id=value)
                union_novedades = list(chain(novedadesList))
                novedades_array.extend(union_novedades)

                criticasList = Criticaddi.objects.filter(post_ddi_id=value)
                union_criticas = list(chain(criticasList))
                criticas_array.extend(union_criticas)

        # Create an in-memory output file for the new workbook.
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet(name="Inventario de demandas")
        
        # Get some data to write to the spreadsheet.
        row = 4
        col = 0
        ## altura de celda
        worksheet.set_row(0, 40)
        worksheet.set_row(1, 30)
        worksheet.set_row(2, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet.set_column('A:A', 30)
        worksheet.set_column('B:B', 30)
        worksheet.set_column('C:C', 30)
        worksheet.set_column('D:D', 30)
        worksheet.set_column('E:E', 30)
        worksheet.set_column('F:F', 30)
        worksheet.set_column('G:G', 40)
        worksheet.set_column('H:H', 30)
        worksheet.set_column('I:I', 30)
        worksheet.set_column('J:J', 30)
        worksheet.set_column('K:K', 30)
        worksheet.set_column('L:L', 30)
        worksheet.set_column('M:M', 30)
        worksheet.set_column('N:N', 30)
        worksheet.set_column('O:O', 30)
        worksheet.set_column('P:P', 30)
        worksheet.set_column('Q:Q', 30)
        worksheet.set_column('R:R', 30)
        worksheet.set_column('S:S', 30)
        worksheet.set_column('T:T', 30)
        worksheet.set_column('U:U', 40)
        worksheet.set_column('V:V', 40)
        worksheet.set_column('W:W', 30)
        worksheet.set_column('X:X', 30)
        worksheet.set_column('Y:Y', 30)
        worksheet.set_column('Z:Z', 30)

        worksheet.set_column('AA:AA', 40)
        worksheet.set_column('AB:AB', 40)
        worksheet.set_column('AC:AC', 30)
        worksheet.set_column('AD:AD', 30)
        worksheet.set_column('AE:AE', 30)
        worksheet.set_column('AF:AF', 30)
        worksheet.set_column('AG:AG', 30)
        worksheet.set_column('AH:AH', 30)
        worksheet.set_column('AI:AI', 30)
        worksheet.set_column('AJ:AJ', 30)
        worksheet.set_column('AK:AK', 30)
        worksheet.set_column('AL:AL', 30)
        worksheet.set_column('AM:AM', 30)
        worksheet.set_column('AN:AN', 30)
        worksheet.set_column('AO:AO', 30)
        worksheet.set_column('AP:AP', 30)
        worksheet.set_column('AQ:AQ', 30)
        worksheet.set_column('AR:AR', 30)
        worksheet.set_column('AS:AS', 30)
        worksheet.set_column('AT:AT', 30)
        worksheet.set_column('AU:AU', 40)
        worksheet.set_column('AV:AV', 40)
        worksheet.set_column('AW:AW', 40)
        worksheet.set_column('AX:AX', 30)
        worksheet.set_column('AY:AY', 30)
        worksheet.set_column('AZ:AZ', 30)

        worksheet.set_column('BA:BA', 40)
        worksheet.set_column('BB:BB', 40)
        worksheet.set_column('BC:BC', 30)
        worksheet.set_column('BD:BD', 30)
        worksheet.set_column('BE:BE', 30)
        worksheet.set_column('BF:BF', 30)
        worksheet.set_column('BG:BG', 40)
        worksheet.set_column('BH:BH', 30)
        worksheet.set_column('BI:BI', 30)
        worksheet.set_column('BJ:BJ', 30)
        worksheet.set_column('BK:BK', 30)
        worksheet.set_column('BL:BL', 30)
        worksheet.set_column('BM:BM', 30)
        worksheet.set_column('BN:BN', 30)
        worksheet.set_column('BO:BO', 30)
        worksheet.set_column('BP:BP', 30)
        worksheet.set_column('BQ:BQ', 30)
        worksheet.set_column('BR:BR', 30)
        worksheet.set_column('BS:BS', 30)
        worksheet.set_column('BT:BT', 30)
        worksheet.set_column('BU:BU', 30)
        worksheet.set_column('BV:BV', 30)
        worksheet.set_column('BW:BW', 30)
        worksheet.set_column('BX:BX', 30)
        worksheet.set_column('BY:BY', 30)
        worksheet.set_column('BZ:BZ', 30)

        worksheet.set_column('CA:CA', 40)
        worksheet.set_column('CB:CB', 40)
        worksheet.set_column('CC:CC', 30)
        worksheet.set_column('CD:CD', 30)
        worksheet.set_column('CE:CE', 30)
        worksheet.set_column('CF:CF', 40)
        worksheet.set_column('CG:CG', 40)
        worksheet.set_column('CH:CH', 30)
        worksheet.set_column('CI:CI', 30)
        worksheet.set_column('CJ:CJ', 30)
        worksheet.set_column('CK:CK', 30)
        worksheet.set_column('CL:CL', 30)
        worksheet.set_column('CM:CM', 30)
        worksheet.set_column('CN:CN', 30)
        worksheet.set_column('CO:CO', 30)
        worksheet.set_column('CP:CP', 30)
        worksheet.set_column('CQ:CQ', 30)
        worksheet.set_column('CR:CR', 30)
        worksheet.set_column('CS:CS', 30)
        worksheet.set_column('CT:CT', 30)
        worksheet.set_column('CU:CU', 30)
        worksheet.set_column('CV:CV', 30)
        worksheet.set_column('CW:CW', 30)
        worksheet.set_column('CX:CX', 30)
        worksheet.set_column('CY:CY', 30)
        worksheet.set_column('CZ:CZ', 30)

        worksheet.set_column('DA:DA', 30)
        worksheet.set_column('DB:DB', 30)
        worksheet.set_column('DC:DC', 30)
       
        ## agregar estilos
        Backgroundcolor = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 18,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        Backgroundcolor2 = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 14,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        Backgroundcolor3 = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 12,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        cell_format = workbook.add_format()
        cell_format.set_font_color('white')

        textInfo_format = workbook.add_format({'text_wrap': True, 'align': 'center', 'valign': 'vcenter'})
        formatfecha = workbook.add_format({'num_format': 'dd/mm/yy', 'align': 'center', 'valign': 'vcenter'})

        ##FILA 1
        worksheet.merge_range('A1:DC1', "Inventario de Demandas de Información", Backgroundcolor)
        worksheet.conditional_format('A1:DC1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Demandas de Información',
                                                'format': cell_format})

        ##FILA 2

        worksheet.merge_range('A2:A3', "", Backgroundcolor)
        worksheet.conditional_format('A2:A3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '',
                                                'format': cell_format})

        worksheet.merge_range('B2:B3', "", Backgroundcolor)
        worksheet.conditional_format('B2:B3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '',
                                                'format': cell_format})

        worksheet.merge_range('C2:T2', "MÓDULO A. IDENTIFICACIÓN ", Backgroundcolor2)
        worksheet.conditional_format('C2:T2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'MÓDULO A. IDENTIFICACIÓN',
                                                'format': cell_format })

        
        worksheet.merge_range('U2:CU2', "MÓDULO B. DETECCIÓN Y ANÁLISIS DE REQUERIMIENTOS", Backgroundcolor2)
        worksheet.conditional_format('U2:CU2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'MÓDULO B. DETECCIÓN Y ANÁLISIS DE REQUERIMIENTOS',
                                                'format': cell_format })


        worksheet.merge_range('CX2:DA2', "COMPLEMENTO DE DEMANDA", Backgroundcolor3)
        worksheet.conditional_format('CX2:DA2', {'type': 'text',
                                        'criteria': 'begins with',
                                        'value': 'COMPLEMENTO DE DEMANDA',
                                        'format': cell_format })


        worksheet.merge_range('DB2:DC2', "MÓDULO VALIDACIÓN DE LA DEMANDA", Backgroundcolor3)
        worksheet.conditional_format('DB2:DC2', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'MÓDULO VALIDACIÓN DE LA DEMANDA',
                                                'format': cell_format })


        ##FILA 3
        worksheet.merge_range('C3:E3', "ÁREA TEMÁTICA / TEMA", Backgroundcolor2)
        worksheet.conditional_format('C3:E3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'ÁREA TEMÁTICA / TEMA',
                                                'format': cell_format })

        
        worksheet.merge_range('F3:F4',  "Seleccione el comité estadístico sectorial al que pertenece la demanda de información estadística", Backgroundcolor3)
        worksheet.conditional_format('F3:F4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Seleccione el comité estadístico sectorial al que pertenece la demanda de información estadística',
                                                'format': cell_format })

        worksheet.merge_range('G3:N3',  "¿Quién identificó la demanda de información?", Backgroundcolor3)
        worksheet.conditional_format('G3:N3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '¿Quién identificó la demanda de información?',
                                                'format': cell_format })

        worksheet.merge_range('O3:Z3',  "A. Datos del Solicitante", Backgroundcolor3)
        worksheet.conditional_format('O3:AZ3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'A. Datos del Solicitante',
                                                'format': cell_format })

        worksheet.merge_range('AA3:AA4',  "1. ¿Cuál es el indicador o requerimiento de información estadística?", Backgroundcolor3)
        worksheet.conditional_format('AA3:AA4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '1. ¿Cuál es el indicador o requerimiento de información estadística?',
                                                'format': cell_format })

        worksheet.merge_range('AB3:AB4',  "2. Descripción general de la información que requiere", Backgroundcolor3)
        worksheet.conditional_format('AB3:AB4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '2. Descripción general de la información que requiere',
                                                'format': cell_format })

        worksheet.merge_range('AC3:AG3',  "3. ¿Para qué se utiliza la información estadística?", Backgroundcolor3)
        worksheet.conditional_format('AC3:AG3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '3. ¿Para qué se utiliza la información estadística?',
                                                'format': cell_format })

        worksheet.merge_range('AH3:AS3',  "4. ¿Cuáles podrían ser los usuarios principales de la información solicitada?", Backgroundcolor3)
        worksheet.conditional_format('AH3:AS3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '4. ¿Cuáles podrían ser los usuarios principales de la información solicitada?',
                                                'format': cell_format })

        worksheet.merge_range('AT3:AX3',  "5. La información solicitada responde a las siguientes normas", Backgroundcolor3)
        worksheet.conditional_format('AT3:AX3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '5. La información solicitada responde a las siguientes normas',
                                                'format': cell_format })

        worksheet.merge_range('AY3:BE3',  "6. La información solicitada responde a los siguientes requerimientos", Backgroundcolor3)
        worksheet.conditional_format('AY3:BE3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '6. La información solicitada responde a los siguientes requerimientos',
                                                'format': cell_format })

        worksheet.merge_range('BF3:BL3',  "7. ¿La información requerida está siendo producida en su totalidad?", Backgroundcolor3)
        worksheet.conditional_format('BF3:BL3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '7. ¿La información requerida está siendo producida en su totalidad?',
                                                'format': cell_format })

        worksheet.merge_range('BM3:BQ3',  "8. ¿Qué tipo de requerimiento es?", Backgroundcolor3)
        worksheet.conditional_format('BF3:BQ3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '8. ¿Qué tipo de requerimiento es?',
                                                'format': cell_format })


        worksheet.merge_range('BR3:BR4',  "9. ¿Qué variables necesita para suplir el requerimiento? (Ver hoja 2 - Lista de variables)", Backgroundcolor3)
        worksheet.conditional_format('BR3:BR4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '9. ¿Qué variables necesita para suplir el requerimiento? (Ver hoja 2 - Lista de variables)',
                                                'format': cell_format })

        worksheet.merge_range('BS3:BT3',  "10. ¿Cuál entidad considera que debe producir la información requerida?", Backgroundcolor3)
        worksheet.conditional_format('BS3:BT3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '10. ¿Cuál entidad considera que debe producir la información requerida?',
                                                'format': cell_format })


        worksheet.merge_range('BU3:BX3',  "11. ¿Cuál(es) de las siguientes opciones podrían dar solución al requerimiento de información? (Ver hoja 3 - pregunta 11)", Backgroundcolor3)
        worksheet.conditional_format('BU3:BX3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '11. ¿Cuál(es) de las siguientes opciones podrían dar solución al requerimiento de información? (Ver hoja 3 - pregunta 11)',
                                                'format': cell_format })

        
        worksheet.merge_range('BY3:CE3',  "12. Indique la desagregación requerida:", Backgroundcolor3)
        worksheet.conditional_format('BY3:CE3', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '12. Indique la desagregación requerida:',
                                                'format': cell_format })


        worksheet.merge_range('CF3:CP3', "13. Indique la desagregación requerida", Backgroundcolor3)
        worksheet.conditional_format('CF3:CP3', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '13. Indique la desagregación requerida',
                                                'format': cell_format })

        worksheet.merge_range('CQ3:CU3', "14. Periodicidad de difusión requerida", Backgroundcolor3)
        worksheet.conditional_format('CQ3:CU3', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '14. Periodicidad de difusión requerida',
                                                'format': cell_format })


        worksheet.merge_range('CV2:CV4', "MÓDULO C. OBSERVACIONES", Backgroundcolor3)
        worksheet.conditional_format('CV2:CV4', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'MÓDULO C. OBSERVACIONES',
                                                'format': cell_format })


        worksheet.merge_range('CW2:CW4', "MÓDULO D. ANEXOS", Backgroundcolor3)
        worksheet.conditional_format('CW2:CW4', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'MÓDULO D. ANEXOS',
                                                'format': cell_format })


        worksheet.merge_range('CX3:CY3', "¿Es un demanda insatisfecha priorizada?", Backgroundcolor3)
        worksheet.conditional_format('CX3:CY3', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '¿Es un demanda insatisfecha priorizada?',
                                                'format': cell_format })

        worksheet.merge_range('CZ3:DA3', "¿Es un demanda insatisfecha priorizada?", Backgroundcolor3)
        worksheet.conditional_format('CZ3:DA3', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '¿Es un demanda insatisfecha priorizada?',
                                                'format': cell_format })


        worksheet.merge_range('DB3:DC3', "¿El requerimiento reportado es una demanda de información estadística?", Backgroundcolor3)
        worksheet.conditional_format('DB3:DC3', {'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '¿El requerimiento reportado es una demanda de información estadística?',
                                                'format': cell_format })

        worksheet.write(3, 0, "Código", Backgroundcolor3)
        worksheet.write(3, 1, "Estado", Backgroundcolor3)
        worksheet.write(3, 2, "Área temática", Backgroundcolor3)
        worksheet.write(3, 3, "Tema principal", Backgroundcolor3)
        worksheet.write(3, 4, "Tema compartido", Backgroundcolor3)

        worksheet.write(3, 6, "Consejo Asesor Técnico del SEN", Backgroundcolor3)
        worksheet.write(3, 7, "Comité Estadístico Sectorial", Backgroundcolor3)
        worksheet.write(3, 8, "Mesa Estadística Sectorial", Backgroundcolor3)
        worksheet.write(3, 9, "Entidad", Backgroundcolor3)
        worksheet.write(3, 10, "¿Cuales?", Backgroundcolor3)
        worksheet.write(3, 11, "Otra entidad", Backgroundcolor3)
        worksheet.write(3, 12, "DANE", Backgroundcolor3)
        worksheet.write(3, 13, "Otras instancias de coordinación", Backgroundcolor3)
        
        worksheet.write(3, 14, "Código de la entidad", Backgroundcolor3)
        worksheet.write(3, 15, "Nombre de la entidad", Backgroundcolor3)
        worksheet.write(3, 16, "Entidad consumidora de información", Backgroundcolor3)
        worksheet.write(3, 17, "Dependencia", Backgroundcolor3)
        worksheet.write(3, 18, "Nombre del jefe de la dependencia", Backgroundcolor3)
        worksheet.write(3, 19, "Cargo", Backgroundcolor3)
        worksheet.write(3, 20, "Correo electrónico", Backgroundcolor3)
        worksheet.write(3, 21, "Teléfono jefe dependencia", Backgroundcolor3)
        worksheet.write(3, 22, "Persona que realiza el requerimiento", Backgroundcolor3)
        worksheet.write(3, 23, "Cargo", Backgroundcolor3)
        worksheet.write(3, 24, "Correo electrónico", Backgroundcolor3)
        worksheet.write(3, 25, "Teléfono", Backgroundcolor3)

        worksheet.write(3, 28, "a. Análisis de contexto", Backgroundcolor3)
        worksheet.write(3, 29, "b. Diseño, formulación o seguimiento de políticas", Backgroundcolor3)
        worksheet.write(3, 30, "c. Reportes internacionales", Backgroundcolor3)
        worksheet.write(3, 31, "d. Otro ¿Cuál?", Backgroundcolor3)
        worksheet.write(3, 32, "¿Cuál?", Backgroundcolor3)

        worksheet.write(3, 33, "a. Organismos internacionales", Backgroundcolor3)
        worksheet.write(3, 34, "b. Presidencia de la República", Backgroundcolor3)
        worksheet.write(3, 35, "c. Ministerios", Backgroundcolor3)
        worksheet.write(3, 36, "d. Organismos de Control", Backgroundcolor3)
        worksheet.write(3, 37, "e. Otras entidades del orden Nacional", Backgroundcolor3)
        worksheet.write(3, 38, "f. Entidades de orden Territorial", Backgroundcolor3)
        worksheet.write(3, 39, "g. Gremios", Backgroundcolor3)
        worksheet.write(3, 40, "h. Entidades privadas", Backgroundcolor3)
        worksheet.write(3, 41, "i. Dependencias de la misma entidad", Backgroundcolor3)
        worksheet.write(3, 42, "j. Academia", Backgroundcolor3)
        worksheet.write(3, 43, "k. Público en General", Backgroundcolor3)
        worksheet.write(3, 44, "l. Otra ¿cuál?", Backgroundcolor3)

        worksheet.write(3, 45, "a. Constitución Política", Backgroundcolor3)
        worksheet.write(3, 46, "b. Ley", Backgroundcolor3)
        worksheet.write(3, 47, "c. Decreto (Nacional, departamental, etc.)", Backgroundcolor3)
        worksheet.write(3, 48, "d. Otra (Resolución, ordenanza, acuerdo)", Backgroundcolor3)
        worksheet.write(3, 49, "e. Ninguna", Backgroundcolor3)

        worksheet.write(3, 50, "a. Plan Nacional de Desarrollo (Capítulo y línea)", Backgroundcolor3)
        worksheet.write(3, 51, "b. Cuentas económicas y macroeconómicas", Backgroundcolor3)
        worksheet.write(3, 52, "c. Plan Sectorial, Territorial o CONPES", Backgroundcolor3)
        worksheet.write(3, 53, "d. Objetivos de Desarrollo Sostenible (ODS) ¿Cuál es el número del Objetivo? ¿Cuál es el número del Indicador?", Backgroundcolor3)
        worksheet.write(3, 54, "e. Organización para la Cooperación y el Desarrollo Económicos (OCDE)", Backgroundcolor3)
        worksheet.write(3, 55, "f. Otros compromisos internacionales", Backgroundcolor3)
        worksheet.write(3, 56, "g. Otro(s)", Backgroundcolor3)

        worksheet.write(3, 57, "Si/No", Backgroundcolor3)
        worksheet.write(3, 58, "¿Cuál entidad?", Backgroundcolor3)
        worksheet.write(3, 59, "Otra entidad", Backgroundcolor3)
        worksheet.write(3, 60, "Nombre de la operación estadística", Backgroundcolor3)
        worksheet.write(3, 61, "Otra operación estadística", Backgroundcolor3)
        worksheet.write(3, 62, "Nombre del registro administrativo", Backgroundcolor3)
        worksheet.write(3, 63, "Otra registro administrativo", Backgroundcolor3)

        worksheet.write(3, 64, "a. Agregado estadístico o indicador", Backgroundcolor3)
        worksheet.write(3, 65, "b. Ampliación cobertura geográfica", Backgroundcolor3)
        worksheet.write(3, 66, "c. Ampliación cobertura temática", Backgroundcolor3)
        worksheet.write(3, 67, "d. Ajustes en la difusión", Backgroundcolor3)
        worksheet.write(3, 68, "e. Otro. ¿Cuál?", Backgroundcolor3)

        worksheet.write(3, 70, "Cual entidad", Backgroundcolor3)
        worksheet.write(3, 71, "Otra ¿Cuál?", Backgroundcolor3)

        worksheet.write(3, 72, "a. Aprovechamiento de una operación (es) estadística (s)", Backgroundcolor3)
        worksheet.write(3, 73, "b. Aprovechamiento de un registro administrativo", Backgroundcolor3)
        worksheet.write(3, 74, "c. Generación de nueva información", Backgroundcolor3)
        worksheet.write(3, 75, "d. Otra (s) ¿Cuáles?", Backgroundcolor3)

        worksheet.write(3, 76, "a. Sexo", Backgroundcolor3)
        worksheet.write(3, 77, "b. Edades", Backgroundcolor3)
        worksheet.write(3, 78, "c. Grupos étnicos", Backgroundcolor3)
        worksheet.write(3, 79, "d. Discapacidad", Backgroundcolor3)
        worksheet.write(3, 80, "e. Estrato", Backgroundcolor3)
        worksheet.write(3, 81, "f. Otra ¿Cuál?", Backgroundcolor3)
        worksheet.write(3, 82, "g. Ninguna", Backgroundcolor3)

        worksheet.write(3, 83, "Geográfica", Backgroundcolor3)
        worksheet.write(3, 84, "Nacional", Backgroundcolor3)
        worksheet.write(3, 85, "Regional", Backgroundcolor3)
        worksheet.write(3, 86, "Departamental", Backgroundcolor3)
        worksheet.write(3, 87, "Áreas metropolitanas", Backgroundcolor3)
        worksheet.write(3, 88, "Municipal", Backgroundcolor3)
        worksheet.write(3, 89, "Otra, ¿cuál?", Backgroundcolor3)
        worksheet.write(3, 90, "Zona", Backgroundcolor3)
        worksheet.write(3, 91, "Total", Backgroundcolor3)
        worksheet.write(3, 92, "Urbano", Backgroundcolor3)
        worksheet.write(3, 93, "Rural", Backgroundcolor3)

        worksheet.write(3, 94, "a. Anual", Backgroundcolor3)
        worksheet.write(3, 95, "b. Semestral", Backgroundcolor3)
        worksheet.write(3, 96, "c. Trimestral", Backgroundcolor3)
        worksheet.write(3, 97, "d. Mensual", Backgroundcolor3)
        worksheet.write(3, 98, "e. Otra ¿Cuál?", Backgroundcolor3)

        ##COMPLEMENTO DE DEMANDA
        worksheet.write(3, 101, "SI/NO", Backgroundcolor3)
        worksheet.write(3, 102, "¿Instancia de coordinación o entidad que lo priorizó?", Backgroundcolor3)
        worksheet.write(3, 103, "SI/NO", Backgroundcolor3)
        worksheet.write(3, 104, "Describa de forma breve la acción que suple la demanda insatisfecha", Backgroundcolor3)

        ##VALIDACIÓN DE LA DEMANDA
        worksheet.write(3, 105, "SI/NO", Backgroundcolor3)
        worksheet.write(3, 106, "Describa de forma breve", Backgroundcolor3)

        for dem in demandas:
            worksheet.write(row, col, dem.codigo_ddi, textInfo_format)
            worksheet.write(row, col + 1, str(dem.nombre_est), textInfo_format)
            worksheet.write(row, col + 2, str(dem.area_tem), textInfo_format)
            worksheet.write(row, col + 3, str(dem.tema_prin), textInfo_format)
            
            lista_temaComp = dem.tema_comp.all()
            lista_temaComp_array = []
            for index, item in enumerate (lista_temaComp):
                lista_temaComp_array.append(item)
                worksheet.write(row, col + 4, str(lista_temaComp_array).replace('[','').replace(']','').replace("<TemaCompartido:", "").replace(">", ""), textInfo_format)
               
            lista_comiteEst = dem.comite_est.all()
            lista_comiteEst_array = []
            for index, item in enumerate (lista_comiteEst):
                lista_comiteEst_array.append(item)
                worksheet.write(row, col + 5, str(lista_comiteEst_array).replace('[','').replace(']','').replace("<ComiteEstSect:", "").replace(">", ""), textInfo_format) 
            
            lista_quien_identddi = dem.quien_identddi.all()
            for index, item in enumerate(lista_quien_identddi):
                if str(item) == "Consejo Asesor Técnico del SEN":
                    worksheet.write(row, col + 6, str(item), textInfo_format)

                if str(item) == "Comité Estadístico Sectorial":
                    worksheet.write(row, col + 7, str(item), textInfo_format)

                if str(item) == "Mesa Estadística Sectorial":
                    worksheet.write(row, col + 8, str(item), textInfo_format)

                if str(item) == "Entidad":
                    worksheet.write(row, col + 9, str(item), textInfo_format)

                if str(item) == "DANE":
                    worksheet.write(row, col + 12, str(item), textInfo_format)

                if str(item) == "Otras instancias de coordinación":
                    worksheet.write(row, col + 13, str(item), textInfo_format)

            lista_entidad_qiddi = dem.entidad_qiddi.all()
            lista_entidad_qiddi_array = []
            for index, item in enumerate (lista_entidad_qiddi):
                lista_entidad_qiddi_array .append(item)
                worksheet.write(row, col + 10, str(lista_entidad_qiddi_array ).replace('[','').replace(']','').replace("<entidad_service:", "").replace(">", ""), textInfo_format)

            worksheet.write(row, col + 11, str(dem.otra_entidad_qiddi), textInfo_format)

            if dem.entidad_sol == None:
                worksheet.write(row, col + 14, "", textInfo_format)
            else:
                worksheet.write(row, col + 14, str(dem.entidad_sol.codigo), textInfo_format)

            if dem.entidad_sol == None:
                worksheet.write(row, col + 15, "", textInfo_format)
            else:
                worksheet.write(row, col + 15, str(dem.entidad_sol), textInfo_format)

            if dem.entidad_cons_sol == None:
                worksheet.write(row, col + 16, "", textInfo_format)
            else:
                worksheet.write(row, col + 16, str(dem.entidad_cons_sol), textInfo_format)


            worksheet.write(row, col + 17, str(dem.dependencia), textInfo_format)
            worksheet.write(row, col + 18, str(dem.nombre_jef_dep), textInfo_format)
            worksheet.write(row, col + 19, str(dem.cargo_jef_dep), textInfo_format)
            worksheet.write(row, col + 20, str(dem.correo_jef_dep), textInfo_format)

            if dem.telefono_jef_dep == None:
                worksheet.write(row, col + 21, "", textInfo_format)
            else:
                worksheet.write(row, col + 21, str(dem.telefono_jef_dep), textInfo_format)

            worksheet.write(row, col + 22, str(dem.pers_req), textInfo_format)
            worksheet.write(row, col + 23, str(dem.cargo_pers_req), textInfo_format)
            worksheet.write(row, col + 24, str(dem.correo_pers_req), textInfo_format)

            if dem.telefono_pers_req == None:
                worksheet.write(row, col + 25, "", textInfo_format)
            else:
                worksheet.write(row, col + 25, str(dem.telefono_pers_req), textInfo_format)

            #MÓDULO B
            worksheet.write(row, col + 26, str(dem.pm_b_1), textInfo_format)
            worksheet.write(row, col + 27, str(dem.pm_b_2), textInfo_format)
            
            lista_pm_b_3 = dem.pm_b_3.all()
            for index_pmb3, item_pmb3 in enumerate(lista_pm_b_3):
                if str(item_pmb3) == "Análisis de contexto":
                    worksheet.write(row, col + 28, str(item_pmb3), textInfo_format)

                if str(item_pmb3) == "Diseño, formulación o seguimiento de políticas":
                    worksheet.write(row, col + 29, str(item_pmb3), textInfo_format)

                if str(item_pmb3) == "Reportes internacionales":
                    worksheet.write(row, col + 30, str(item_pmb3), textInfo_format)

                if str(item_pmb3) == "Otro ¿Cuál?":
                    worksheet.write(row, col + 31, str(item_pmb3), textInfo_format)

            lista_pm_b_4 = dem.pm_b_4.all()
            for index_pm_b_4, item_pm_b_4 in enumerate(lista_pm_b_4):
                
                if str(item_pm_b_4) == "Presidencia de la República":
                    worksheet.write(row, col + 34, str(item_pm_b_4), textInfo_format)

                if str(item_pm_b_4) == "Público en General":
                    worksheet.write(row, col + 43, str(item_pm_b_4), textInfo_format)

            lista_pm_b_5 = dem.pm_b_5.all()
            for index_pm_b_5, item_pm_b_5 in enumerate(lista_pm_b_5):
                
                if str(item_pm_b_5) == "Ninguna":
                    worksheet.write(row, col + 49, str(item_pm_b_5), textInfo_format)

            lista_pm_b_7 = dem.total_si_no_pmb7.all()
            lista_pm_b_7_array = []
            for index_pm_b_7, item_pm_b_7 in enumerate(lista_pm_b_7):
                lista_pm_b_7_array.append(item_pm_b_7)
                worksheet.write(row, col + 57, str(lista_pm_b_7_array).replace('[','').replace(']','').replace("<InfoProdTotalidad:", "").replace(">", ""), textInfo_format)

            lista_entidad_pm_b7 = dem.entidad_pm_b7.all()
            lista_entidad_pm_b7_array = []
            for index, item in enumerate(lista_entidad_pm_b7):
                lista_entidad_pm_b7_array.append(item)
                worksheet.write(row, col + 58, str(lista_entidad_pm_b7_array).replace('[','').replace(']','').replace("<entidad_service:", "").replace(">", ""), textInfo_format)
            
            worksheet.write(row, col + 59, str(dem.otro_entidad_pm_b7), textInfo_format)

            lista_ooee_pm_b7 = dem.ooee_pm_b7.all()
            lista_ooee_pm_b7_array = []
            for index, item in enumerate(lista_ooee_pm_b7):
                lista_ooee_pm_b7_array.append(item)
                worksheet.write(row, col + 60, str(lista_ooee_pm_b7_array).replace('[','').replace(']','').replace("<ooee_service:", "").replace(">", ""), textInfo_format)
            
            worksheet.write(row, col + 61, str(dem.otro_ooee_pm_b7), textInfo_format)
            
            lista_rraa_pm_b7 = dem.rraa_pm_b7.all()
            lista_rraa_pm_b7_array = []
            for index, item in enumerate(lista_rraa_pm_b7):
                lista_rraa_pm_b7_array.append(item)
                worksheet.write(row, col + 62, str(lista_rraa_pm_b7_array).replace('[','').replace(']','').replace("<rraa_service:", "").replace(">", ""), textInfo_format)

            worksheet.write(row, col + 63, str(dem.otro_rraa_pm_b7), textInfo_format)
                
            lista_pm_b_8 = dem.pm_b_8.all()
            for index_pm_b_8, item_pm_b_8 in enumerate(lista_pm_b_8):
                
                if str(item_pm_b_8) == "Agregado estadístico o indicador":
                    worksheet.write(row, col + 64, str(item_pm_b_8), textInfo_format)

                if str(item_pm_b_8) == "Ampliación cobertura geográfica":
                    worksheet.write(row, col + 65, str(item_pm_b_8), textInfo_format)

                if str(item_pm_b_8) == "Ampliación cobertura temática":
                    worksheet.write(row, col + 66, str(item_pm_b_8), textInfo_format)

                if str(item_pm_b_8) == "Ajustes en la difusión":
                    worksheet.write(row, col + 67, str(item_pm_b_8), textInfo_format)

            lista_pm_b_10 = dem.pm_b_10.all()
            lista_pm_b_10_array = []
            for index, item in enumerate(lista_pm_b_10):
                lista_pm_b_10_array.append(item)
                worksheet.write(row, col + 70, str(lista_pm_b_10_array).replace('[','').replace(']','').replace("<entidad_service:", "").replace(">", ""), textInfo_format)
            
            worksheet.write(row, col + 71, str(dem.pm_b_10_otro), textInfo_format)

            lista_pm_b_11_1 = dem.pm_b_11_1.all()
            for index_pm_b_11_1, item_pm_b_11_1 in enumerate(lista_pm_b_11_1):
                
                if str(item_pm_b_11_1) == "Aprovechamiento de una operación (es) estadística (s)":
                    worksheet.write(row, col + 72, str(item_pm_b_11_1), textInfo_format)

                if str(item_pm_b_11_1) == "Aprovechamiento de un registro administrativo":
                    worksheet.write(row, col + 73, str(item_pm_b_11_1), textInfo_format)

                if str(item_pm_b_11_1) == "Generación de nueva información":
                    worksheet.write(row, col + 74, str(item_pm_b_11_1), textInfo_format)

            worksheet.write(row, col + 75, str(dem.otra_cual_11_1_d), textInfo_format)

            lista_pm_b_12 = dem.pm_b_12.all()
            for index_pm_b_12, item_pm_b_12 in enumerate(lista_pm_b_12):
                
                if str(item_pm_b_12) == "Sexo":
                    worksheet.write(row, col + 76, str(item_pm_b_12), textInfo_format)

                if str(item_pm_b_12) == "Edades":
                    worksheet.write(row, col + 77, str(item_pm_b_12), textInfo_format)

                if str(item_pm_b_12) == "Grupos étnicos":
                    worksheet.write(row, col + 78, str(item_pm_b_12), textInfo_format)

                if str(item_pm_b_12) == "Discapacidad":
                    worksheet.write(row, col + 79, str(item_pm_b_12), textInfo_format)

                if str(item_pm_b_12) == "Estrato":
                    worksheet.write(row, col + 80, str(item_pm_b_12), textInfo_format)

                if str(item_pm_b_12) == "Ninguna":
                    worksheet.write(row, col + 82, str(item_pm_b_12), textInfo_format)

            lista_pm_b_13 = dem.pm_b_13.all()
            for index_pm_b_13, item_pm_b_13 in enumerate(lista_pm_b_13):
                
                if str(item_pm_b_13) == "Geográfica":
                    worksheet.write(row, col + 83, str(item_pm_b_13), textInfo_format)

                if str(item_pm_b_13) == "Zona":
                    worksheet.write(row, col + 90, str(item_pm_b_13), textInfo_format)


            lista_pm_b_13_geo = dem.pm_b_13_geo.all()
            for index_pm_b_13_geo, item_pm_b_13_geo in enumerate(lista_pm_b_13_geo):

                if str(item_pm_b_13_geo) == "Nacional":
                    worksheet.write(row, col + 84, str(item_pm_b_13_geo), textInfo_format)

                if str(item_pm_b_13_geo) == "Regional":
                    worksheet.write(row, col + 85, str(item_pm_b_13_geo), textInfo_format)

                if str(item_pm_b_13_geo) == "Departamental":
                    worksheet.write(row, col + 86, str(item_pm_b_13_geo), textInfo_format)

                if str(item_pm_b_13_geo) == "Áreas metropolitanas":
                    worksheet.write(row, col + 87, str(item_pm_b_13_geo), textInfo_format)

                if str(item_pm_b_13_geo) == "Municipal":
                    worksheet.write(row, col + 88, str(item_pm_b_13_geo), textInfo_format)

            lista_pm_b_13_zona = dem.pm_b_13_zona.all()
            for index_pm_b_13_zona, item_pm_b_13_zona in enumerate(lista_pm_b_13_zona):

                if str(item_pm_b_13_zona) == "Total":
                    worksheet.write(row, col + 91, str(item_pm_b_13_zona), textInfo_format)

                if str(item_pm_b_13_zona) == "Urbano":
                    worksheet.write(row, col + 92, str(item_pm_b_13_zona), textInfo_format)

                if str(item_pm_b_13_zona) == "Rural":
                    worksheet.write(row, col + 93, str(item_pm_b_13_zona), textInfo_format)

            lista_pm_b_14 = dem.pm_b_14.all()
            for index_pm_b_14, item_pm_b_14 in enumerate(lista_pm_b_14):
                
                if str(item_pm_b_14) == "Anual":
                    worksheet.write(row, col + 94, str(item_pm_b_14), textInfo_format)

                if str(item_pm_b_14) == "Semestral":
                    worksheet.write(row, col + 95, str(item_pm_b_14), textInfo_format)

                if str(item_pm_b_14) == "Trimestral":
                    worksheet.write(row, col + 96, str(item_pm_b_14), textInfo_format)

                if str(item_pm_b_14) == "Mensual":
                    worksheet.write(row, col + 97, str(item_pm_b_14), textInfo_format)
            
            ##MÓDULO C
            worksheet.write(row, col + 99, str(dem.pm_c_1), textInfo_format)

            ##MÓDULO D
            worksheet.write(row, col + 100, str(dem.pm_d_1_anexos), textInfo_format)

            ###Complemento de la demanda
            lista_compl_dem_a = dem.compl_dem_a.all()
            lista_compl_dem_a_array = []
            for index, item in enumerate(lista_compl_dem_a):
                lista_compl_dem_a_array.append(item)
                worksheet.write(row, col + 101, str(lista_compl_dem_a_array).replace('[','').replace(']','').replace("<DemandaInsaprio:", "").replace(">", ""), textInfo_format)
            
            worksheet.write(row, col + 102, str(dem.compl_dem_a_text), textInfo_format)
            
            lista_comple_dem_b = dem.comple_dem_b.all()
            lista_comple_dem_b_array = []
            for index, item in enumerate(lista_comple_dem_b):
                lista_comple_dem_b_array.append(item)
                worksheet.write(row, col + 103, str(lista_comple_dem_b_array).replace('[','').replace(']','').replace("<PlanaccionSuplirDem:", "").replace(">", ""), textInfo_format)
            
            worksheet.write(row, col + 104, str(dem.compl_dem_b_text), textInfo_format)

            ###validación de la demanda
            lista_validacion_ddi = dem.validacion_ddi.all()
            lista_validacion_ddi_array = []
            for index, item in enumerate(lista_validacion_ddi):
                lista_validacion_ddi_array.append(item)
                worksheet.write(row, col + 105, str(lista_validacion_ddi_array).replace('[','').replace(']','').replace("<EsDemandadeInfor:", "").replace(">", ""), textInfo_format)
                        
            worksheet.write(row, col + 106, str(dem.validacion_ddi_text), textInfo_format)

            row += 1

        # modulo B campos de textos

        for indexp3, item in enumerate(preg3_array):  
            
            indexp3 = indexp3 + 4
            worksheet.write(indexp3, col + 32, str(item.otro_cual), textInfo_format)

        for indexp4, itemp4 in enumerate(preg4_array):
            
            indexp4 = indexp4 + 4
            worksheet.write(indexp4, col + 33, str(itemp4.orginter_text), textInfo_format)
            worksheet.write(indexp4, col + 35, str(itemp4.ministerios_text), textInfo_format)
            worksheet.write(indexp4, col + 36, str(itemp4.orgcontrol_text), textInfo_format)
            worksheet.write(indexp4, col + 37, str(itemp4.oentidadesordenal_text), textInfo_format)
            worksheet.write(indexp4, col + 38, str(itemp4.entidadesordenterr_text), textInfo_format)
            worksheet.write(indexp4, col + 39, str(itemp4.gremios_text), textInfo_format)
            worksheet.write(indexp4, col + 40, str(itemp4.entiprivadas_text), textInfo_format)
            worksheet.write(indexp4, col + 41, str(itemp4.dependenentidad_text), textInfo_format)
            worksheet.write(indexp4, col + 42, str(itemp4.academia_text), textInfo_format)
            worksheet.write(indexp4, col + 44, str(itemp4.otro_cual_text), textInfo_format)

        for index, item in enumerate(preg5_array):
            index = index + 4
            worksheet.write(index, col + 45, str(item.const_pol_text), textInfo_format)
            worksheet.write(index, col + 46, str(item.ley_text), textInfo_format)
            worksheet.write(index, col + 47, str(item.decreto_text), textInfo_format)
            worksheet.write(index, col + 48, str(item.otra_text), textInfo_format)

        for index, item in enumerate(preg6_array):
            index = index + 4
            worksheet.write(index, col + 50, str(item.planalDes_text), textInfo_format)
            worksheet.write(index, col + 51, str(item.cuentasecomacroec_text), textInfo_format)
            worksheet.write(index, col + 52, str(item.plansecterrcom_text), textInfo_format)
            worksheet.write(index, col + 53, str(item.objdessost_text), textInfo_format)
            worksheet.write(index, col + 54, str(item.orgcooper_text), textInfo_format)
            worksheet.write(index, col + 55, str(item.otroscomprInt_text), textInfo_format)
            worksheet.write(index, col + 56, str(item.otros_text), textInfo_format)
        
        for index, item in enumerate(preg8_array):
            index = index + 4
            worksheet.write(index, col + 68, str(item.otros_c_text), textInfo_format)

        for index, item in enumerate(preg12_array):
            index = index + 4
            worksheet.write(index, col + 81, str(item.otra_cual_text_a), textInfo_format)
            
        for index, item in enumerate(preg13_array):
            index = index + 4
            worksheet.write(index, col + 89, str(item.otra_cual_text_b), textInfo_format)

        for index, item in enumerate(preg14_array):
            index = index + 4
            worksheet.write(index, col + 98, str(item.otra_cual_text_c), textInfo_format)

        ################################## Hoja 2 pregunta 9 lista de variables ###################3 
        worksheet2 = workbook.add_worksheet(name="Lista de variables")
        # Get some data to write to the spreadsheet.
         
        row2 = 4
        col2 = 0
        ## altura de celda
        worksheet2.set_row(0, 40)
        worksheet2.set_row(1, 30)
        worksheet2.set_row(2, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet2.set_column('A:A', 40)
        worksheet2.set_column('B:B', 40)
        worksheet2.set_column('C:C', 40)
        

        worksheet2.merge_range('A1:C1', "Inventario de Demandas de Información", Backgroundcolor)
        worksheet2.conditional_format('A1:C1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Demandas de Información',
                                                'format': cell_format})


        worksheet2.merge_range('A2:C2', "MÓDULO B Detección y Análisis de Requerimientos", Backgroundcolor2)
        worksheet2.conditional_format('A2:C2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "MÓDULO B Detección y Análisis de Requerimientos",
                                                'format': cell_format})


        worksheet2.merge_range('A3:C3', "¿Qué variables necesita para suplir el requerimiento?", Backgroundcolor2)
        worksheet2.conditional_format('A2:C2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "¿Qué variables necesita para suplir el requerimiento?",
                                                'format': cell_format})


        
        worksheet2.write(3, 0, "Código", Backgroundcolor3)
        worksheet2.write(3, 1, "¿Cuál es el indicador o requerimiento de información estadística?", Backgroundcolor3)
        worksheet2.write(3, 2, "variables que necesita para suplir el requerimiento", Backgroundcolor3)
       

        arrayoeId = []
        for indexVariab, itemVariab in enumerate(preg9_array):
            indexVariab =  itemVariab.ddi_id + 3
            arrayoeId.append(itemVariab.ddi_id)
            
            for dema in demandas:
                if str(itemVariab.ddi) == str(dema.pm_b_1):

                    worksheet2.write(row2, col2, str(dema.codigo_ddi), textInfo_format)
                    worksheet2.write(row2, col2 + 1, str(itemVariab.ddi), textInfo_format)
                    worksheet2.write(row2, col2 + 2, itemVariab.lista_varia, textInfo_format)
                    
                    row2 += 1
        
         ################################## Hoja 3 pregunta 11 ###################
        worksheet3 = workbook.add_worksheet(name="Pregunta 11")

        row3 = 5
        col3 = 0
        ## altura de celda
        worksheet3.set_row(0, 40)
        worksheet3.set_row(1, 30)
        worksheet3.set_row(2, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet3.set_column('A:A', 40)
        worksheet3.set_column('B:B', 40)
        worksheet3.set_column('C:C', 40)
        worksheet3.set_column('D:D', 40)
        worksheet3.set_column('E:E', 40)
        worksheet3.set_column('F:F', 40)
        worksheet3.set_column('G:G', 40)
        worksheet3.set_column('H:H', 40)
        worksheet3.set_column('I:I', 40)
        worksheet3.set_column('J:J', 40)
        worksheet3.set_column('K:K', 40)
        worksheet3.set_column('L:L', 40)
        worksheet3.set_column('M:M', 40)
        worksheet3.set_column('N:N', 40)
        worksheet3.set_column('O:O', 40)
        worksheet3.set_column('P:P', 40)
        
        
        worksheet3.merge_range('A1:P1', "Inventario de Demandas de Información", Backgroundcolor)
        worksheet3.conditional_format('A1:P1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Demandas de Información',
                                                'format': cell_format})

        worksheet3.merge_range('A2:A4', "", Backgroundcolor)
        worksheet3.conditional_format('A2:A4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '',
                                                'format': cell_format})

        worksheet3.merge_range('B2:B4', "", Backgroundcolor)
        worksheet3.conditional_format('B2:B4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': '',
                                                'format': cell_format})


        worksheet3.merge_range('C2:P2', "MÓDULO B Detección y Análisis de Requerimientos", Backgroundcolor2)
        worksheet3.conditional_format('C2:P2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "MÓDULO B Detección y Análisis de Requerimientos",
                                                'format': cell_format})


        worksheet3.merge_range('C3:P3', "¿Cuál(es) de las siguientes opciones podrían dar solución al requerimiento de información?", Backgroundcolor2)
        worksheet3.conditional_format('C3:P3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "¿Cuál(es) de las siguientes opciones podrían dar solución al requerimiento de información?",
                                                'format': cell_format})

        worksheet3.merge_range('C4:I4', "a) Aprovechamiento de una operación (es) estadística (s)", Backgroundcolor2)
        worksheet3.conditional_format('C4:I4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "a) Aprovechamiento de una operación (es) estadística (s)",
                                                'format': cell_format})


        worksheet3.merge_range('J4:M4', "b) Aprovechamiento de un registro administrativo", Backgroundcolor2)
        worksheet3.conditional_format('J4:M4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "b) Aprovechamiento de un registro administrativo",
                                                'format': cell_format})

        worksheet3.merge_range('N4:P4', "c) Generación de nueva información", Backgroundcolor2)
        worksheet3.conditional_format('N4:P4', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "c) Generación de nueva información",
                                                'format': cell_format})


        
        worksheet3.write(4, 0, "Código", Backgroundcolor3)
        worksheet3.write(4, 1, "¿Cuál es el indicador o requerimiento de información estadística?", Backgroundcolor3)
        
        worksheet3.write(4, 2, "a) 1. ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 3, "a)2.1. Incluir variables/preguntas ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 4, "a)2.2. Cambiar la formulación de alguna(s) pregunta(s) ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 5, "a)2.3. Ampliar la desagregación temática ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 6, "a)2.4. Ampliar la desagregación geográfica ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 7, "a)2.5. Aumentar la frecuencia en la difusión de los resultados ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 8, "a)2.6. Otra ¿Cuál(es)?", Backgroundcolor3)

        worksheet3.write(4, 9, "b)1.  ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 10, "b)2.1.  Inclusión de variables/preguntas", Backgroundcolor3)
        worksheet3.write(4, 11, "b)2.2. Cambios en la formulación de alguna(s) pregunta(s)", Backgroundcolor3)
        worksheet3.write(4, 12, "b)2.3. Otro ", Backgroundcolor3)

        
        worksheet3.write(4, 13, "c)1. Operación estadística nueva", Backgroundcolor3)
        worksheet3.write(4, 14, "c)2. Indicador ¿Cuál(es)?", Backgroundcolor3)
        worksheet3.write(4, 15, "c)3. Otra  ¿Cuál?", Backgroundcolor3)
			
        idoePreg11 = []
        for indexPreg11, itemPreg11 in enumerate(preg11_array):
            indexPreg11 = itemPreg11.post_ddi_id + 5
            idoePreg11.append(itemPreg11.post_ddi_id)

            for dema in demandas:
                if str(itemPreg11.post_ddi) == str(dema.pm_b_1):

                    worksheet3.write(row3, col3, str(dema.codigo_ddi), textInfo_format)
                    worksheet3.write(row3, col3 + 1, str(itemPreg11.post_ddi), textInfo_format)

                    if itemPreg11.ooee_pmb11 == None:
                        worksheet3.write(row3, col3 + 2, "", textInfo_format)
                    else:
                        worksheet3.write(row3, col3 + 2, str(itemPreg11.ooee_pmb11), textInfo_format)

                    worksheet3.write(row3, col3 + 3, itemPreg11.inc_var_cual, textInfo_format)
                    worksheet3.write(row3, col3 + 4, itemPreg11.cam_preg_cual, textInfo_format)
                    worksheet3.write(row3, col3 + 5, itemPreg11.am_des_tem_cual, textInfo_format)
                    worksheet3.write(row3, col3 + 6, itemPreg11.am_des_geo_cual, textInfo_format)
                    worksheet3.write(row3, col3 + 7, itemPreg11.dif_resul_cual, textInfo_format)
                    worksheet3.write(row3, col3 + 8, itemPreg11.opc_aprov_cual, textInfo_format)

                    if itemPreg11.rraa_pmb11 == None:
                        worksheet3.write(row3, col3 + 9, "", textInfo_format)
                    else:
                        worksheet3.write(row3, col3 + 9, str(itemPreg11.rraa_pmb11), textInfo_format)

                    worksheet3.write(row3, col3 + 10, itemPreg11.inc_varia_cual, textInfo_format)
                    worksheet3.write(row3, col3 + 11, itemPreg11.camb_pregu_cual, textInfo_format)
                    worksheet3.write(row3, col3 + 12, itemPreg11.otros_aprov_ra, textInfo_format)

                    worksheet3.write(row3, col3 + 13, itemPreg11.nueva_oe, textInfo_format)
                    worksheet3.write(row3, col3 + 14, itemPreg11.indi_cual, textInfo_format)
                    worksheet3.write(row3, col3 + 15, itemPreg11.gen_nueva, textInfo_format)
                    
                    row3 += 1

        ################################## Hoja 4 Comentarios ###################
        worksheet4 = workbook.add_worksheet(name="Comentarios")

        comenta = Comentariosddi.objects.all()

        row4 = 3
        col4 = 0
        ## altura de celda
        worksheet4.set_row(0, 40)
        worksheet4.set_row(1, 30)
        worksheet4.set_row(2, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet4.set_column('A:A', 40)
        worksheet4.set_column('B:B', 40)
        worksheet4.set_column('C:C', 40)
        worksheet4.set_column('D:D', 40)
        worksheet4.set_column('E:E', 40)
        

        worksheet4.merge_range('A1:E1', "Inventario de Demandas de Información", Backgroundcolor)
        worksheet4.conditional_format('A1:E1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Demandas de Información',
                                                'format': cell_format})


        worksheet4.merge_range('A2:E2', "COMENTARIOS", Backgroundcolor2)
        worksheet4.conditional_format('A2:E2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "COMENTARIOS",
                                                'format': cell_format})
    
        worksheet4.write(2, 0, "Código", Backgroundcolor3)
        worksheet4.write(2, 1, "¿Cuál es el indicador o requerimiento de información estadística?", Backgroundcolor3)
        worksheet4.write(2, 2, "Usuario que realiza el comentario", Backgroundcolor3)
        worksheet4.write(2, 3, "Comentario", Backgroundcolor3)
        worksheet4.write(2, 4, "Fecha en que se realiza el comentario", Backgroundcolor3)

        idoeComentarios = []
        for indexCom, itemCom in enumerate(comentarios_array):
            indexCom = itemCom.post_ddi_id + 3
            idoeComentarios.append(itemCom.post_ddi_id)

            for dema in demandas:
                if str(itemCom.post_ddi) == str(dema.pm_b_1):

                    worksheet4.write(row4, col4, str(dema.codigo_ddi), textInfo_format)
                    worksheet4.write(row4, col4 + 1, str(itemCom.post_ddi), textInfo_format)
                    worksheet4.write(row4, col4 + 2, itemCom.name, textInfo_format)
                    worksheet4.write(row4, col4 + 3, itemCom.body, textInfo_format)
                    worksheet4.write(row4, col4 + 4, itemCom.created_on.replace(tzinfo=None), formatfecha)

                    row4 += 1

        ################################## Hoja 5 MÓDULO DE ACTUALIZACIÓN - Novedades ################### 
        worksheet5 = workbook.add_worksheet(name="Actualizacion-Novedades")

        nove = NovedadActualizacionddi.objects.all()

        row5 = 3
        col5 = 0
        ## altura de celda
        worksheet5.set_row(0, 40)
        worksheet5.set_row(1, 30)
        worksheet5.set_row(2, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet5.set_column('A:A', 40)
        worksheet5.set_column('B:B', 40)
        worksheet5.set_column('C:C', 40)
        worksheet5.set_column('D:D', 40)
        worksheet5.set_column('E:E', 40)
        worksheet5.set_column('F:F', 40)
        worksheet5.set_column('G:G', 40)
        

        worksheet5.merge_range('A1:G1', "Inventario de Demandas de Información", Backgroundcolor)
        worksheet5.conditional_format('A1:G1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Demandas de Información',
                                                'format': cell_format})


        worksheet5.merge_range('A2:G2', "MÓDULO DE ACTUALIZACIÓN - NOVEDADES", Backgroundcolor2)
        worksheet5.conditional_format('A2:G2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "MÓDULO DE ACTUALIZACIÓN - NOVEDADES",
                                                'format': cell_format})
    
        worksheet5.write(2, 0, "Código", Backgroundcolor3)
        worksheet5.write(2, 1, "¿Cuál es el indicador o requerimiento de información estadística?", Backgroundcolor3)
        worksheet5.write(2, 2, "Novedad", Backgroundcolor3)
        worksheet5.write(2, 3, "Estado de la actualización", Backgroundcolor3)
        worksheet5.write(2, 4, "Descripción de la novedad", Backgroundcolor3)
        worksheet5.write(2, 5, "Funcionario que realiza la novedad", Backgroundcolor3)
        worksheet5.write(2, 6, "Fecha en que se realiza la novedad", Backgroundcolor3)


        iddiNovedades = []
        for indexNov, itemNov in enumerate(novedades_array):
            indexNov = itemNov.post_ddi_id + 3
            iddiNovedades.append(itemNov.post_ddi_id)

            for dema in demandas:
                if str(itemNov.post_ddi) == str(dema.pm_b_1):

                    worksheet5.write(row5, col5, str(dema.codigo_ddi), textInfo_format)
                    worksheet5.write(row5, col5 + 1, str(itemNov.post_ddi), textInfo_format)
                    worksheet5.write(row5, col5 + 2, str(itemNov.novedad), textInfo_format)
                    worksheet5.write(row5, col5 + 3, str(itemNov.est_actualiz), textInfo_format)
                    worksheet5.write(row5, col5 + 4, itemNov.descrip_novedad, textInfo_format)
                    worksheet5.write(row5, col5 + 5, itemNov.name_nov, textInfo_format)
                    worksheet5.write(row5, col5 + 6, itemNov.fecha_actualiz.replace(tzinfo=None), formatfecha)

                    row5 += 1

        ################################## Hoja 6 Critica ################### 
        worksheet6 = workbook.add_worksheet(name="Criticas")

        criti = Criticaddi.objects.all()

        row6 = 3
        col6 = 0
        ## altura de celda
        worksheet6.set_row(0, 40)
        worksheet6.set_row(1, 30)
        worksheet6.set_row(2, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet6.set_column('A:A', 40)
        worksheet6.set_column('B:B', 40)
        worksheet6.set_column('C:C', 40)
        worksheet6.set_column('D:D', 40)
        worksheet6.set_column('E:E', 40)
        worksheet6.set_column('F:F', 40)
    
        worksheet6.merge_range('A1:F1', "Inventario de Demandas de Información", Backgroundcolor)
        worksheet6.conditional_format('A1:F1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Demandas de Información',
                                                'format': cell_format})


        worksheet6.merge_range('A2:F2', "CRÍTICAS", Backgroundcolor2)
        worksheet6.conditional_format('A2:F2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': "CRÍTICAS",
                                                'format': cell_format})
    
        worksheet6.write(2, 0, "Código", Backgroundcolor3)
        worksheet6.write(2, 1, "¿Cuál es el indicador o requerimiento de información estadística?", Backgroundcolor3)
        worksheet6.write(2, 2, "Estado de la crítica", Backgroundcolor3)
        worksheet6.write(2, 3, "Observaciones de la crítica", Backgroundcolor3)
        worksheet6.write(2, 4, "Funcionario que realiza la crítica", Backgroundcolor3)
        worksheet6.write(2, 5, "Fecha en que se realiza la crítica", Backgroundcolor3)
        
        iddiCriticas = []
        for indexCri, itemCri in enumerate(criticas_array):
            indexCri = itemCri.post_ddi_id + 3
            iddiCriticas.append(itemCri.post_ddi_id)

            for dema in demandas:
                if str(itemCri.post_ddi) == str(dema.pm_b_1):

                    worksheet6.write(row6, col6, str(dema.codigo_ddi), textInfo_format)
                    worksheet6.write(row6, col6 + 1, str(itemCri.post_ddi), textInfo_format)
                    worksheet6.write(row6, col6 + 2, str(itemCri.estado_crit), textInfo_format)
                    worksheet6.write(row6, col6 + 3, str(itemCri.descrip_critica), textInfo_format)
                    worksheet6.write(row6, col6 + 4, itemCri.name_cri, textInfo_format)
                    worksheet6.write(row6, col6 + 5, itemCri.fecha_critica.replace(tzinfo=None), formatfecha)

                    row6 += 1

        # Close the workbook before sending the data.
        workbook.close()

        # Rewind the buffer.
        output.seek(0)

        # Set up the Http response.
        filename = 'ddi_filter.xlsx'
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        return response



### reporte entidades consumidoras de información
@method_decorator(login_required, name='dispatch')
class reporteEntidadesConsumidoras(View):
    def get(self, request):
        
        # Get some data to write to the spreadsheet.
        consumidores = consumidores_info.objects.all()

        # Create an in-memory output file for the new workbook.
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet(name="consumidores informacion")
        
        # Get some data to write to the spreadsheet.
        row = 3
        col = 0
        ## altura de celda
        worksheet.set_row(0, 40)
        worksheet.set_row(1, 30)
        worksheet.set_row(2, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet.set_column('A:A', 30)
        worksheet.set_column('B:B', 30)
        worksheet.set_column('C:C', 30)
        worksheet.set_column('D:D', 30)
        worksheet.set_column('E:E', 30)
        worksheet.set_column('F:F', 30)
        worksheet.set_column('G:G', 40)
        worksheet.set_column('H:H', 30)
        worksheet.set_column('I:I', 30)
        worksheet.set_column('J:J', 30)
        worksheet.set_column('K:K', 30)
        worksheet.set_column('L:L', 30)
        
            
        ## agregar estilos
        Backgroundcolor = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 18,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        Backgroundcolor2 = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 14,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        Backgroundcolor3 = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 12,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        cell_format = workbook.add_format()
        cell_format.set_font_color('white')

        textInfo_format = workbook.add_format({'text_wrap': True, 'align': 'center', 'valign': 'vcenter'})
        

        formatfecha = workbook.add_format({'num_format': 'dd/mm/yy', 'align': 'center', 'valign': 'vcenter'})

        ##FILA 1
        worksheet.merge_range('A1:L1', "Inventario entidades consumidoras de información", Backgroundcolor)
        worksheet.conditional_format('A1:L1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario entidades consumidoras de información',
                                                'format': cell_format})

        ##FILA 2
        worksheet.merge_range('A2:G2', "Datos de la entidad", Backgroundcolor2)
        worksheet.conditional_format('A2:G2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Datos de la entidad',
                                                'format': cell_format })

    
        ##FILA 3
        worksheet.merge_range('H2:L2', "Responsable de la solicitud", Backgroundcolor2)
        worksheet.conditional_format('H2:L2', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Responsable de la solicitud',
                                                'format': cell_format })

        worksheet.write(2, 0, "Nombre de la entidad", Backgroundcolor3)
        worksheet.write(2, 1, "Nit", Backgroundcolor3)
        worksheet.write(2, 2, "Tipo de entidad", Backgroundcolor3)
        worksheet.write(2, 3, "Dirección", Backgroundcolor3)
        worksheet.write(2, 4, "Teléfono", Backgroundcolor3)
        worksheet.write(2, 5, "Página web", Backgroundcolor3)
        worksheet.write(2, 6, "Estado", Backgroundcolor3)

        worksheet.write(2, 7, "Nombre", Backgroundcolor3)
        worksheet.write(2, 8, "Cargo", Backgroundcolor3)
        worksheet.write(2, 9, "Correo ", Backgroundcolor3)
        worksheet.write(2, 10, "Teléfono", Backgroundcolor3)
        worksheet.write(2, 11, "Extensión", Backgroundcolor3)
       
        for con in consumidores:
            
            worksheet.write(row, col, con.nombre_ec, textInfo_format)
            worksheet.write(row, col + 1, con.nit_ec, textInfo_format)
            worksheet.write(row, col + 2, str(con.tipo_entidad_ec), textInfo_format)
            worksheet.write(row, col + 3, con.direccion_ec, textInfo_format)
            worksheet.write(row, col + 4, con.telefono_ec, textInfo_format)
            worksheet.write(row, col + 5, con.pagina_web_ec, textInfo_format)
            worksheet.write(row, col + 6, str(con.estado), textInfo_format)
            worksheet.write(row, col + 7, con.nombre_resp, textInfo_format)
            worksheet.write(row, col + 8, con.cargo_resp, textInfo_format)
            worksheet.write(row, col + 9, con.correo_resp, textInfo_format)
            worksheet.write(row, col + 10, con.telefono_resp, textInfo_format)
            worksheet.write(row, col + 11, con.extension_resp, textInfo_format)

            row += 1

        # Close the workbook before sending the data.
        workbook.close()

        # Rewind the buffer.
        output.seek(0)

        # Set up the Http response.
        filename = 'entidades_consumidoras.xlsx'
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        return response


############ reporte complemenatrio para administrador ######################

@method_decorator(login_required, name='dispatch')
class reporteUltimaNovedadDDI(View):
    def get(self, request):
        
        # Get some data to write to the spreadsheet.
        demandas_infor = demandaInfor.objects.all()
        ##NovedadActualizacionRRAA.filter(post_ra=1).order_by('-id')[:1] 
        # Create an in-memory output file for the new workbook.
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet(name="Ultima novedad ddi")
        
        # Get some data to write to the spreadsheet.
        row = 4
        col = 0
        ## altura de celda
        worksheet.set_row(0, 40)
        worksheet.set_row(1, 30)
        worksheet.set_row(2, 30)
        worksheet.set_row(3, 30)
        worksheet.set_row(4, 30)
        worksheet.set_row(5, 30)
        worksheet.set_row(6, 30)
        worksheet.set_row(7, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet.set_column('A:A', 30)
        worksheet.set_column('B:B', 50)
        worksheet.set_column('C:C', 30)
        worksheet.set_column('D:D', 30)
        worksheet.set_column('E:E', 30)
        worksheet.set_column('F:F', 30)
        worksheet.set_column('G:G', 30)
        
        ## agregar estilos
        Backgroundcolor = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 18,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        Backgroundcolor2 = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 14,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        Backgroundcolor3 = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 12,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        cell_format = workbook.add_format()
        cell_format.set_font_color('white')

        textInfo_format = workbook.add_format({'text_wrap': True, 'align': 'center', 'valign': 'vcenter'})
        

        formatfecha = workbook.add_format({'num_format':'yyyy-mm-dd', 'align': 'center', 'valign': 'vcenter'})

        ##FILA 1
        worksheet.merge_range('A1:G1', "Inventario de Demandas de Información", Backgroundcolor)
        worksheet.conditional_format('A1:G1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Demandas de Información',
                                                'format': cell_format})

        ##FILA 2

        worksheet.merge_range('A2:G3', "ÚLTIMAS NOVEDADES", Backgroundcolor2)
        worksheet.conditional_format('A2:G3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'ÚLTIMAS NOVEDADES',
                                                'format': cell_format })

        worksheet.write(3, 0, "Código", Backgroundcolor3)
        worksheet.write(3, 1, "¿Cuál es el indicador o requerimiento de información estadística?", Backgroundcolor3)
        worksheet.write(3, 2, "Fase en el sistema", Backgroundcolor3)
        worksheet.write(3, 3, "Área temática", Backgroundcolor3)
        worksheet.write(3, 4, "Tema", Backgroundcolor3)
        worksheet.write(3, 5, "Estado de la novedad", Backgroundcolor3)
        worksheet.write(3, 6, "Fecha en que se realiza la novedad", Backgroundcolor3)
         
        for item in demandas_infor:
            item.codigo_ddi

            for demanda in NovedadActualizacionddi.objects.filter(post_ddi__codigo_ddi=item.codigo_ddi).order_by('-id')[:1]:

                worksheet.write(row, col, demanda.post_ddi.codigo_ddi, textInfo_format)
                worksheet.write(row, col + 1, demanda.post_ddi.pm_b_1, textInfo_format)
                worksheet.write(row, col + 2, str(demanda.post_ddi.nombre_est), textInfo_format)
                worksheet.write(row, col + 3, str(demanda.post_ddi.area_tem), textInfo_format)
                worksheet.write(row, col + 4, str(demanda.post_ddi.tema_prin), textInfo_format)
                worksheet.write(row, col + 5, str(demanda.est_actualiz), textInfo_format)
                worksheet.write(row, col + 6, str(demanda.fecha_actualiz), formatfecha)
                row += 1

        # Close the workbook before sending the data.
        workbook.close()

        # Rewind the buffer.
        output.seek(0)

        filename = 'demandas_informacion.xlsx'
        # Set up the Http response.
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        return response