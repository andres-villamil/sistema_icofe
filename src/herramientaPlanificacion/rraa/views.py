from django.shortcuts import render
from login.models import Role, User, Profile

from .forms import CreateRAForm, MB_NormaRRAAForm, MB_DocumentoMetodologRRAAForm, MB_VariableRecolectadaFormset, \
    MB_ConceptosEstandarizadosRRAAForm, MB_ClasificacionesRRAAForm, MB_RecoleccionDatosRRAAForm, MB_FrecuenciaRecoleccionDatoForm, MB_HerramientasUtilProcesaForm, \
        MB_SeguridadInformForm, MB_FrecuenciaAlmacenamientobdForm, MB_CoberturaGeograficaRRAAForm, MB_IndicadorResultadoAgregadoFormset, \
            MB_EntidadesAccesoFormset, MB_NoAccesoDatosForm, \
                EditRAForm, EditMB_NormaRRAAForm, EditMB_DocumentoMetodologRRAAForm, \
                    EditMB_ConceptosEstandarizadosRRAAForm, EditMB_ClasificacionesRRAAForm, EditMB_RecoleccionDatosRRAAForm, EditMB_FrecuenciaRecoleccionDatoForm, EditMB_HerramientasUtilProcesaForm, \
                        EditMB_SeguridadInformForm, EditMB_FrecuenciaAlmacenamientobdForm, EditMB_CoberturaGeograficaRRAAForm, \
                            EditMB_NoAccesoDatosForm, \
                                CommentRRAAForm, NovedadActualizacionRRAAForm, CreateFortalecimientoRRAAForm, EditFortalecimientoRRAAForm, CriticaRRAAForm

from .models import Entidades_oe, RegistroAdministrativo, MB_NormaRRAA, MB_DocumentoMetodologRRAA, MB_VariableRecolectada, MB_ConceptosEstandarizadosRRAA, \
    MB_ClasificacionesRRAA, MB_RecoleccionDatosRRAA, MB_FrecuenciaRecoleccionDato, MB_HerramientasUtilProcesa, \
        MB_SeguridadInform, MB_FrecuenciaAlmacenamientobd, MB_CoberturaGeograficaRRAA, MB_IndicadorResultadoAgregado, \
            MB_EntidadesAccesoRRAA, MB_NoAccesoDatos, SeguimientoPlanFortalecimiento, \
                CommentRRAA, NovedadActualizacionRRAA, CriticaRRAA, FortalecimientoRRAA

from django.contrib.auth.decorators import login_required


from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
from django.shortcuts import render, redirect, get_object_or_404
##  report by excel ##
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Font, Fill, Border, Side, GradientFill, Alignment, Color, PatternFill
##  end report by excel ##
from django.contrib import messages 
from django.views.generic import TemplateView
from django.forms import inlineformset_factory
from django.core import serializers
from itertools import chain
from django.utils import timezone
from django.core.mail import send_mail

from .filters import ItemsFilter, NameFilterRRAA
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger


####libreria reporte excel
import xlsxwriter
import io
from datetime import datetime
from django.views.generic import View

from django.utils.decorators import method_decorator


# Create your views here.

## vista de RRAA por entidad 
@login_required
def allRRAA(request):

    count_entities = Entidades_oe.objects.filter(estado_id=1).count()
    user = request.user
   
    if user.is_authenticated == True and str(user) == "sdazag": # 1 sdazag 
        ras = RegistroAdministrativo.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28)).only('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'ra_activo', 'sist_estado').order_by('nombre_ra')
        ra_filter = ItemsFilter(request.GET, queryset=ras)
        
        
        """ elif user.is_authenticated == True and str(user) == "aobandor": # 3
            ras = RegistroAdministrativo.objects.filter(Q (tema=5) | Q (tema=14) | Q (tema=15) | Q (tema=16) | Q (tema=17) | Q (tema=18)).order_by('nombre_ra')
            ra_filter = ItemsFilter(request.GET, queryset=ras)  """

    elif user.is_authenticated == True and str(user) == "dvlizarazog": # 4 
        
        ras = RegistroAdministrativo.objects.filter(Q (tema_id=9) | Q (tema_id=29) | Q (tema_id=30)).only('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'ra_activo', 'sist_estado').order_by('nombre_ra')
        ra_filter = ItemsFilter(request.GET, queryset=ras)

    elif user.is_authenticated == True and str(user) == "ancardenasc": # 5frherreran o se cambia por gavargasr
        
        ras = RegistroAdministrativo.objects.filter(tema_id=10).only('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'ra_activo', 'sist_estado').order_by('nombre_ra')
        ra_filter = ItemsFilter(request.GET, queryset=ras)
         
    elif user.is_authenticated == True and str(user) == "lhsanchezz": # 6 lhsanchezz
        
        ras = RegistroAdministrativo.objects.filter(Q (tema_id=3) | Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).only('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'ra_activo', 'sist_estado').order_by('nombre_ra')
        ra_filter = ItemsFilter(request.GET, queryset=ras)

    elif user.is_authenticated == True and str(user) == "pczambranog": # 6 pczambranog
        
        ras = RegistroAdministrativo.objects.filter(Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).only('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'ra_activo', 'sist_estado').order_by('nombre_ra')
        ra_filter = ItemsFilter(request.GET, queryset=ras)

    elif user.is_authenticated == True and str(user) == "mjpradag": # 7 mppulidor se cambia por mjpradag

        ras = RegistroAdministrativo.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27)).only('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'ra_activo', 'sist_estado').order_by('nombre_ra')
        ra_filter = ItemsFilter(request.GET, queryset=ras)

    elif user.is_authenticated == True and str(user) == "eeguayazans": # 8 eeguayazans Edgar Eduardo Guayazan Sierra 

        ras = RegistroAdministrativo.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20)).only('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'ra_activo', 'sist_estado').order_by('nombre_ra')
        ra_filter = ItemsFilter(request.GET, queryset=ras)    

    elif user.is_authenticated == True and str(user) == "mlbarretob": # 9 mlbarretob

        ras = RegistroAdministrativo.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23)).only('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'ra_activo', 'sist_estado').order_by('nombre_ra')
        ra_filter = ItemsFilter(request.GET, queryset=ras)

    elif user.is_authenticated == True and user.profile.role.id == 1 or user.profile.role.id == 5 or  user.profile.role.id == 7: ## si es administrador
            
        ras = RegistroAdministrativo.objects.all().order_by('nombre_ra')
        ra_filter = ItemsFilter(request.GET, queryset=ras)

    else:
        ras = RegistroAdministrativo.objects.filter(entidad_pri_id=user.profile.entidad.id).exclude(sist_estado=6).order_by('nombre_ra')
        ra_filter = ItemsFilter(request.GET, queryset=ras)

    nameRAFilter = NameFilterRRAA(request.GET, queryset=ras)
    page = request.GET.get('page', 1)
    paginator = Paginator(ras, 10)
    count_total_ra = ras.count()

    try:
        registro = paginator.page(page)
    except PageNotAnInteger:
        registro = paginator.page(1)
    except EmptyPage:
        registro = paginator.page(paginator.num_pages)

    index = registro.number - 1
    max_index = len(paginator.page_range)
    start_index = index - 3 if index >= 3 else 0
    end_index = index + 3 if index <= max_index - 3 else max_index
    page_start = 1
    page_end = max_index
    page_range = paginator.page_range[start_index:end_index]
    page_number= int(page)

    return render(request, 'rraa/all_rraa.html',  {'ras': ras, 'filter': ra_filter, 'count_entities': count_entities,
    'registro': registro, 'page_range': page_range, 'page_end': page_end, 'page_start': page_start,  'nameRAFilter': nameRAFilter,
    'count_total_ra': count_total_ra})

######## función ajax para buscador por nombre de registro all_rrraa
class FilterNameAdminRA(TemplateView):
    def get(self, request, *args, **kwargs):
        
        nombre = request.GET.get('nombre_ra') #opc 1
        user = request.user

        if user.is_authenticated == True and str(user) == "sdazag": # 1 sdazag 
            registro = RegistroAdministrativo.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28)).filter(nombre_ra__icontains=nombre).only('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'ra_activo', 'sist_estado').order_by('nombre_ra')
            nameRAFilter = NameFilterRRAA(request.GET, queryset=registro) 
        
        elif user.is_authenticated == True and str(user) == "dvlizarazog": # 4 
        
            registro = RegistroAdministrativo.objects.filter(Q (tema_id=9) | Q (tema_id=29) | Q (tema_id=30)).filter(nombre_ra__icontains=nombre).only('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'ra_activo', 'sist_estado').order_by('nombre_ra')
            nameRAFilter = NameFilterRRAA(request.GET, queryset=registro) 

        elif user.is_authenticated == True and str(user) == "ancardenasc": # 5frherreran se cambia por gavargasr
        
            registro = RegistroAdministrativo.objects.filter(tema_id=10).filter(nombre_ra__icontains=nombre).only('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'ra_activo', 'sist_estado').order_by('nombre_ra')
            nameRAFilter = NameFilterRRAA(request.GET, queryset=registro) 
         
        elif user.is_authenticated == True and str(user) == "lhsanchezz": # 6 lhsanchezz
        
            registro = RegistroAdministrativo.objects.filter(Q (tema_id=3) | Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).filter(nombre_ra__icontains=nombre).only('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'ra_activo', 'sist_estado').order_by('nombre_ra')
            nameRAFilter = NameFilterRRAA(request.GET, queryset=registro) 

        elif user.is_authenticated == True and str(user) == "pczambranog": # 6 pczambranog
        
            registro = RegistroAdministrativo.objects.filter(Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25)).filter(nombre_ra__icontains=nombre).only('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'ra_activo', 'sist_estado').order_by('nombre_ra')
            nameRAFilter = NameFilterRRAA(request.GET, queryset=registro) 

        elif user.is_authenticated == True and str(user) == "mjpradag": # 7 mppulidor se cambia por mjpradag

            registro = RegistroAdministrativo.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27)).filter(nombre_ra__icontains=nombre).only('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'ra_activo', 'sist_estado').order_by('nombre_ra')
            nameRAFilter = NameFilterRRAA(request.GET, queryset=registro) 

        elif user.is_authenticated == True and str(user) == "eeguayazans": # 8 eeguayazans Edgar Eduardo Guayazan Sierra

            registro = RegistroAdministrativo.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20)).filter(nombre_ra__icontains=nombre).only('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'ra_activo', 'sist_estado').order_by('nombre_ra')
            nameRAFilter = NameFilterRRAA(request.GET, queryset=registro)      

        elif user.is_authenticated == True and str(user) == "mlbarretob": # 9 mlbarretob

            registro = RegistroAdministrativo.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23)).filter(nombre_ra__icontains=nombre).only('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'ra_activo', 'sist_estado').order_by('nombre_ra')
            nameRAFilter = NameFilterRRAA(request.GET, queryset=registro) 

        elif user.is_authenticated == True and user.profile.role.id == 1 or user.profile.role.id == 5 or  user.profile.role.id == 7: ## si es administrador
            
            registro = RegistroAdministrativo.objects.filter(nombre_ra__icontains=nombre).order_by('nombre_ra')
            nameRAFilter = NameFilterRRAA(request.GET, queryset=registro) 

        else:
            registro = RegistroAdministrativo.objects.filter(entidad_pri_id=user.profile.entidad.id).filter(nombre_ra__icontains=nombre).exclude(sist_estado_id=6).order_by('nombre_ra')
            nameRAFilter = NameFilterRRAA(request.GET, queryset=registro)

        data = serializers.serialize('json', registro,
                    fields=('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'codigo_rraa', 'sist_estado', 'ra_activo'), use_natural_foreign_keys=True, use_natural_primary_keys=True)
        return HttpResponse(data, content_type='application/json')

############ ajax all_rraa para filtros ####################
class filter_ajaxAllrraa(TemplateView):
    def get(self, request, *args, **kwargs):
        
        id_area_tematica = request.GET.get('area_tematica') #opc 1
        id_tema = request.GET.get('tema') #opc 2
        id_fase = request.GET.get('fase')  #opc 3
        id_entidad = request.GET.get('entidad')  #opc 4
        user = request.user

        ########## user sdazag ################
        if user.is_authenticated == True and str(user) == "sdazag": # 1 sdazag
            ras = RegistroAdministrativo.objects.filter(Q (tema=2) | Q (tema=7) | Q (tema=19) | Q (tema=28)).order_by('nombre_ra')
            ra_filter = ItemsFilter(request.GET, queryset=ras)

            #opcion 1 es diferente de vacia
            if id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica)
            
            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_entidad == "":
                filtered_queryset_data = ra_filter.qs.filter(sist_estado=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_entidad != "":
                filtered_queryset_data = ra_filter.qs.filter(entidad_pri=id_entidad) 

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema)  

            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema).filter(sist_estado=id_fase)

            ## opc 1, opc 2, opc 3 y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != ""  and id_entidad != "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema).filter(sist_estado=id_fase).filter(entidad_pri=id_entidad)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(sist_estado=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and  id_entidad != "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(entidad_pri=id_entidad)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and  id_entidad == "":
        
                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema).filter(sist_estado=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and  id_entidad != "":
        
                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema).filter(entidad_pri=id_entidad)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and  id_entidad != "":
        
                filtered_queryset_data = ra_filter.qs.filter(sist_estado=id_fase).filter(entidad_pri=id_entidad)

        #######################  user aobandor se cambia usuario a rol a administrador ###########################  
            """ elif user.is_authenticated == True and str(user) == "aobandor": # 3

            ras = RegistroAdministrativo.objects.filter(Q (tema=5) | Q (tema=14) | Q (tema=15) | Q (tema=16) | Q (tema=17) | Q (tema=18)).order_by('nombre_ra')
            ra_filter = ItemsFilter(request.GET, queryset=ras) """


        ####################### user dvlizarazog #############################
        elif user.is_authenticated == True and str(user) == "dvlizarazog": # 4
            
            ras = RegistroAdministrativo.objects.filter(Q (tema=9) | Q (tema=29) | Q (tema=30)).order_by('nombre_ra')
            ra_filter = ItemsFilter(request.GET, queryset=ras)

            #opcion 1 es diferente de vacia
            if id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica)
            
            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_entidad == "":
                filtered_queryset_data = ra_filter.qs.filter(sist_estado=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_entidad != "":
                filtered_queryset_data = ra_filter.qs.filter(entidad_pri=id_entidad) 

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema)  

            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema).filter(sist_estado=id_fase)

            ## opc 1, opc 2, opc 3 y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != ""  and id_entidad != "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema).filter(sist_estado=id_fase).filter(entidad_pri=id_entidad)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(sist_estado=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and  id_entidad != "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(entidad_pri=id_entidad)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and  id_entidad == "":
        
                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema).filter(sist_estado=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and  id_entidad != "":
        
                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema).filter(entidad_pri=id_entidad)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and  id_entidad != "":
        
                filtered_queryset_data = ra_filter.qs.filter(sist_estado=id_fase).filter(entidad_pri=id_entidad)

        #################### user fherreran ##########################
        elif user.is_authenticated == True and str(user) == "ancardenasc": # 5frherreran se cambia por gavargasr
            
            ras = RegistroAdministrativo.objects.filter(tema=10).order_by('nombre_ra')
            ra_filter = ItemsFilter(request.GET, queryset=ras)
            
            #opcion 1 es diferente de vacia
            if id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica)
            
            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_entidad == "":
                filtered_queryset_data = ra_filter.qs.filter(sist_estado=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_entidad != "":
                filtered_queryset_data = ra_filter.qs.filter(entidad_pri=id_entidad) 

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema)  

            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema).filter(sist_estado=id_fase)

            ## opc 1, opc 2, opc 3 y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != ""  and id_entidad != "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema).filter(sist_estado=id_fase).filter(entidad_pri=id_entidad)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(sist_estado=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and  id_entidad != "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(entidad_pri=id_entidad)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and  id_entidad == "":
        
                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema).filter(sist_estado=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and  id_entidad != "":
        
                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema).filter(entidad_pri=id_entidad)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and  id_entidad != "":
        
                filtered_queryset_data = ra_filter.qs.filter(sist_estado=id_fase).filter(entidad_pri=id_entidad)
        
        ########################### user lhsanchezz se remplaza por pczambranog ####################  
        elif user.is_authenticated == True and str(user) == "pczambranog": # 6
            
            ras = RegistroAdministrativo.objects.filter(Q (tema=11) | Q (tema=12) | Q (tema=25)).order_by('nombre_ra')
            ra_filter = ItemsFilter(request.GET, queryset=ras)

            #opcion 1 es diferente de vacia
            if id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica)
            
            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_entidad == "":
                filtered_queryset_data = ra_filter.qs.filter(sist_estado=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_entidad != "":
                filtered_queryset_data = ra_filter.qs.filter(entidad_pri=id_entidad) 

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema)  

            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema).filter(sist_estado=id_fase)

            ## opc 1, opc 2, opc 3 y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != ""  and id_entidad != "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema).filter(sist_estado=id_fase).filter(entidad_pri=id_entidad)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(sist_estado=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and  id_entidad != "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(entidad_pri=id_entidad)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and  id_entidad == "":
        
                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema).filter(sist_estado=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and  id_entidad != "":
        
                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema).filter(entidad_pri=id_entidad)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and  id_entidad != "":
        
                filtered_queryset_data = ra_filter.qs.filter(sist_estado=id_fase).filter(entidad_pri=id_entidad)

        ####################### user mjpradag ###########################3
        elif user.is_authenticated == True and str(user) == "mjpradag": # 7 mppulidor se cambia por mjpradag

            ras = RegistroAdministrativo.objects.filter(Q (tema=13) | Q (tema=24) | Q (tema=26) | Q (tema=27)).order_by('nombre_ra')
            ra_filter = ItemsFilter(request.GET, queryset=ras)

            #opcion 1 es diferente de vacia
            if id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica)
            
            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_entidad == "":
                filtered_queryset_data = ra_filter.qs.filter(sist_estado=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_entidad != "":
                filtered_queryset_data = ra_filter.qs.filter(entidad_pri=id_entidad) 

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema)  

            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema).filter(sist_estado=id_fase)

            ## opc 1, opc 2, opc 3 y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != ""  and id_entidad != "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema).filter(sist_estado=id_fase).filter(entidad_pri=id_entidad)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(sist_estado=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and  id_entidad != "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(entidad_pri=id_entidad)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and  id_entidad == "":
        
                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema).filter(sist_estado=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and  id_entidad != "":
        
                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema).filter(entidad_pri=id_entidad)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and  id_entidad != "":
        
                filtered_queryset_data = ra_filter.qs.filter(sist_estado=id_fase).filter(entidad_pri=id_entidad)

        #################### user eeguayazans ########################
        elif user.is_authenticated == True and str(user) == "eeguayazans": # 8 eeguayazans Edgar Eduardo Guayazan Sierra

            ras = RegistroAdministrativo.objects.filter(Q (tema=4) | Q (tema=8) | Q (tema=20)).order_by('nombre_ra')
            ra_filter = ItemsFilter(request.GET, queryset=ras)

            #opcion 1 es diferente de vacia
            if id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica)
            
            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_entidad == "":
                filtered_queryset_data = ra_filter.qs.filter(sist_estado=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_entidad != "":
                filtered_queryset_data = ra_filter.qs.filter(entidad_pri=id_entidad) 

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema)  

            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema).filter(sist_estado=id_fase)

            ## opc 1, opc 2, opc 3 y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != ""  and id_entidad != "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema).filter(sist_estado=id_fase).filter(entidad_pri=id_entidad)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(sist_estado=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and  id_entidad != "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(entidad_pri=id_entidad)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and  id_entidad == "":
        
                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema).filter(sist_estado=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and  id_entidad != "":
        
                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema).filter(entidad_pri=id_entidad)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and  id_entidad != "":
        
                filtered_queryset_data = ra_filter.qs.filter(sist_estado=id_fase).filter(entidad_pri=id_entidad)

        

        ############################user mlbarretob ######################
        elif user.is_authenticated == True and str(user) == "mlbarretob": # 9

            ras = RegistroAdministrativo.objects.filter(Q (tema=21) | Q (tema=22) | Q (tema=23)).order_by('nombre_ra')
            ra_filter = ItemsFilter(request.GET, queryset=ras)

            #opcion 1 es diferente de vacia
            if id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica)
            
            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_entidad == "":
                filtered_queryset_data = ra_filter.qs.filter(sist_estado=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_entidad != "":
                filtered_queryset_data = ra_filter.qs.filter(entidad_pri=id_entidad) 

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema)  

            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema).filter(sist_estado=id_fase)

            ## opc 1, opc 2, opc 3 y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != ""  and id_entidad != "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema).filter(sist_estado=id_fase).filter(entidad_pri=id_entidad)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(sist_estado=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and  id_entidad != "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(entidad_pri=id_entidad)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and  id_entidad == "":
        
                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema).filter(sist_estado=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and  id_entidad != "":
        
                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema).filter(entidad_pri=id_entidad)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and  id_entidad != "":
        
                filtered_queryset_data = ra_filter.qs.filter(sist_estado=id_fase).filter(entidad_pri=id_entidad)

        ############################ Rol administrador y revisor #######################
        elif user.is_authenticated == True and user.profile.role.id == 1 or user.profile.role.id == 5: 
            
            ras = RegistroAdministrativo.objects.all().order_by('nombre_ra')
            ra_filter = ItemsFilter(request.GET, queryset=ras)

            #opcion 1 es diferente de vacia
            if id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica)
            
            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema)   
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_entidad == "":
                filtered_queryset_data = ra_filter.qs.filter(sist_estado=id_fase)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_entidad != "":
                filtered_queryset_data = ra_filter.qs.filter(entidad_pri=id_entidad) 

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema)  

            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema).filter(sist_estado=id_fase)

            ## opc 1, opc 2, opc 3 y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != ""  and id_entidad != "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema).filter(sist_estado=id_fase).filter(entidad_pri=id_entidad)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(sist_estado=id_fase)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and  id_entidad != "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(entidad_pri=id_entidad)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and  id_entidad == "":
        
                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema).filter(sist_estado=id_fase)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and  id_entidad != "":
        
                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema).filter(entidad_pri=id_entidad)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and  id_entidad != "":
        
                filtered_queryset_data = ra_filter.qs.filter(sist_estado=id_fase).filter(entidad_pri=id_entidad)
        
        ###########################  Rol fuente ###############################
        else:
            ras = RegistroAdministrativo.objects.filter(entidad_pri=user.profile.entidad.id).order_by('nombre_ra')
            ra_filter = ItemsFilter(request.GET, queryset=ras)

            #opcion 1 es diferente de vacia
            if id_area_tematica != "" and  id_tema == "" and id_fase == "" and id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).exclude(sist_estado_id=6)
            
            #opcion 2 es diferente de vacia
            elif id_area_tematica == "" and  id_tema != "" and id_fase == "" and id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema).exclude(sist_estado_id=6)  
            
            #opcion 3 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase != "" and id_entidad == "":
                filtered_queryset_data = ra_filter.qs.filter(sist_estado=id_fase).exclude(sist_estado_id=6)

            #opcion 4 es diferente de vacia
            elif id_area_tematica == "" and  id_tema == "" and id_fase == "" and id_entidad != "":
                filtered_queryset_data = ra_filter.qs.filter(entidad_pri=id_entidad).exclude(sist_estado_id=6) 

            ## opc 1 y opc 2 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase == "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema).exclude(sist_estado_id=6)  

            ## opc 1, opc 2 y opc 3 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema).filter(sist_estado=id_fase).exclude(sist_estado_id=6)

            ## opc 1, opc 2, opc 3 y opc 4 es diferente de vacia
            elif  id_area_tematica != "" and  id_tema != "" and id_fase != ""  and id_entidad != "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema).filter(sist_estado=id_fase).filter(entidad_pri=id_entidad).exclude(sist_estado_id=6)

            ## opc 1, opc 3  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase != "" and  id_entidad == "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(sist_estado=id_fase).exclude(sist_estado_id=6)

            ## opc 1, opc 4  diferente de vacia
            elif id_area_tematica != "" and  id_tema == ""  and id_fase == "" and  id_entidad != "":

                filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(entidad_pri=id_entidad).exclude(sist_estado_id=6)
            
            ## opc 2, opc 3  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase != "" and  id_entidad == "":
        
                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema).filter(sist_estado=id_fase).exclude(sist_estado_id=6)

            ## opc 2, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema != ""  and id_fase == "" and  id_entidad != "":
        
                filtered_queryset_data = ra_filter.qs.filter(tema=id_tema).filter(entidad_pri=id_entidad).exclude(sist_estado_id=6)

            ## opc 3, opc 4  diferente de vacia
            elif id_area_tematica == "" and  id_tema == ""  and id_fase != "" and  id_entidad != "":
        
                filtered_queryset_data = ra_filter.qs.filter(sist_estado=id_fase).filter(entidad_pri=id_entidad).exclude(sist_estado_id=6)              

        data = serializers.serialize('json', filtered_queryset_data,
                    fields=('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'codigo_rraa', 'sist_estado', 'ra_activo'), use_natural_foreign_keys=True, use_natural_primary_keys=True)
        return HttpResponse(data, content_type='application/json')


## vista de consulta de todos los RRAA EN ESTADO PUBLICADO
def consultationModuleRRAA(request):
    
    count_entities = Entidades_oe.objects.filter(estado_id=1).count()
    ras = RegistroAdministrativo.objects.filter(sist_estado_id=5).only('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'ra_activo', 'sist_estado').order_by('nombre_ra')
    count_total_ra = ras.count()
    ra_filter = ItemsFilter(request.GET, queryset=ras)

    nameRAFilter = NameFilterRRAA(request.GET, queryset=ras)

    page = request.GET.get('page', 1)
    paginator = Paginator(ras, 10)

    try:
        registro = paginator.page(page)
    except PageNotAnInteger:
        registro = paginator.page(1)
    except EmptyPage:
        registro = paginator.page(paginator.num_pages)

    index = registro.number - 1
    max_index = len(paginator.page_range)
    start_index = index - 3 if index >= 3 else 0
    end_index = index + 3 if index <= max_index - 3 else max_index
    page_start = 1
    page_end = max_index
    page_range = paginator.page_range[start_index:end_index]
    page_number= int(page)
    
    return render(request, 'rraa/modulo_consulta_ra.html',  {'ras': ras, 'filter': ra_filter, 
    'count_entities': count_entities, 'count_total_ra': count_total_ra, 'registro': registro,
    'page_range': page_range, 'page_end': page_end, 'page_start': page_start,  'nameRAFilter': nameRAFilter })


## filtros por nombre
class FilterByNamera(TemplateView):
    def get(self, request, *args, **kwargs):
        
        nombre = request.GET.get('nombre_ra') #opc 1
        registro = RegistroAdministrativo.objects.filter(sist_estado_id=5).filter(nombre_ra__icontains=nombre)
        oe_filter = NameFilterRRAA(request.GET, queryset=registro)
        data = serializers.serialize('json', registro,
                    fields=('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'codigo_rraa', 'sist_estado', 'ra_activo'), use_natural_foreign_keys=True, use_natural_primary_keys=True)
        return HttpResponse(data, content_type='application/json')


############ ajax modulo consulta RRAA ####################
class consulta_ajaxModulorraa(TemplateView):
    def get(self, request, *args, **kwargs):
        
        id_entidad = request.GET.get('entidad') #opc 1
        id_area_tematica = request.GET.get('area_tematica') #opc 2
        id_tema = request.GET.get('tema') #opc 3
       
       ## opc 1 es diferente de vacia
        if id_entidad != "" and id_area_tematica == "" and  id_tema == "":

            ras = RegistroAdministrativo.objects.filter(sist_estado=5)
            ra_filter = ItemsFilter(request.GET, queryset=ras)
            filtered_queryset_data = ra_filter.qs.filter(entidad_pri=id_entidad)
            
        ## opc 1 y opc 2 es diferente de vacia
        elif  id_entidad != "" and id_area_tematica != "" and  id_tema == "":
            
            ras = RegistroAdministrativo.objects.filter(sist_estado=5)
            ra_filter = ItemsFilter(request.GET, queryset=ras)
            filtered_queryset_data = ra_filter.qs.filter(entidad_pri=id_entidad).filter(area_temat=id_area_tematica)
           
        ## opc 1, opc 2 y opc 3 es diferente de vacia
        elif id_entidad != "" and id_area_tematica != "" and  id_tema != "":
            
            ras = RegistroAdministrativo.objects.filter(sist_estado=5)
            ra_filter = ItemsFilter(request.GET, queryset=ras)
            filtered_queryset_data = ra_filter.qs.filter(entidad_pri=id_entidad).filter(area_temat=id_area_tematica).filter(tema=id_tema)
        
        ## opcion 1 y opcion 3 es diferente de vacia
        elif id_entidad != "" and id_area_tematica == "" and  id_tema != "":
            
            ras = RegistroAdministrativo.objects.filter(sist_estado=5)
            ra_filter = ItemsFilter(request.GET, queryset=ras)
            filtered_queryset_data = ra_filter.qs.filter(entidad_pri=id_entidad).filter(tema=id_tema)
        
        ## opc 2 es diferente de vacia
        elif id_entidad == "" and id_area_tematica != "" and  id_tema == "":
            
            ras = RegistroAdministrativo.objects.filter(sist_estado=5)
            ra_filter = ItemsFilter(request.GET, queryset=ras)
            filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica)

        ## opc 2 y opc 3 es diferente de vacia
        elif id_entidad == "" and id_area_tematica != "" and  id_tema != "":
            
            ras = RegistroAdministrativo.objects.filter(sist_estado=5)
            ra_filter = ItemsFilter(request.GET, queryset=ras)
            filtered_queryset_data = ra_filter.qs.filter(area_temat=id_area_tematica).filter(tema=id_tema)

        # opc 3 es diferente de vacia
        elif id_entidad == "" and id_area_tematica == "" and  id_tema != "":
            
            ras = RegistroAdministrativo.objects.filter(sist_estado=5)
            ra_filter = ItemsFilter(request.GET, queryset=ras)
            filtered_queryset_data = ra_filter.qs.filter(tema=id_tema)

        else:
            filtered_queryset_data = []
        
        data = serializers.serialize('json', filtered_queryset_data,
                    fields=('nombre_ra', 'entidad_pri', 'area_temat', 'tema', 'codigo_rraa', 'sist_estado', 'ra_activo'), use_natural_foreign_keys=True, use_natural_primary_keys=True)
        return HttpResponse(data, content_type='application/json')




@login_required
def createRRAAView(request):

    count_entities = Entidades_oe.objects.count()
    count_ra = RegistroAdministrativo.objects.count()

    user = request.user
    entidad_responsable = user.profile.entidad.pk
    ## forms 
    create_ra_form = CreateRAForm(request.POST or None, request.FILES or None)   #form crear ra
    norma_ra_form = MB_NormaRRAAForm(request.POST or None)
    documentoMetodolog_ra_form = MB_DocumentoMetodologRRAAForm(request.POST or None)
    concepEstan_ra_form = MB_ConceptosEstandarizadosRRAAForm(request.POST or None)
    clasificacion_ra_form = MB_ClasificacionesRRAAForm(request.POST or None)
    recolecDatos_ra_form = MB_RecoleccionDatosRRAAForm(request.POST or None)
    frecuencRecole_ra_form = MB_FrecuenciaRecoleccionDatoForm(request.POST or None)
    heramiUtiPr_ra_form = MB_HerramientasUtilProcesaForm(request.POST or None)
    seguInfor_ra_form = MB_SeguridadInformForm(request.POST or None)
    frecuencAlma_ra_form = MB_FrecuenciaAlmacenamientobdForm(request.POST or None)
    cobertGeogra_ra_form =  MB_CoberturaGeograficaRRAAForm(request.POST or None)
    noAccesoDato_ra_form = MB_NoAccesoDatosForm(request.POST or None)

    #novedad
    novedad_rraa_form = NovedadActualizacionRRAAForm()
    new_novedad_rraa = None

    registered = False

    if request.method == 'GET':  ##formset
        formvariableRecolset = MB_VariableRecolectadaFormset(queryset=MB_VariableRecolectada.objects.none(), prefix='variablerecolectada')
        formIndicadorReAgreset = MB_IndicadorResultadoAgregadoFormset(queryset=MB_IndicadorResultadoAgregado.objects.none(), prefix='indicadorresultadoagregado')
        formEntidadAcceset = MB_EntidadesAccesoFormset(queryset=MB_EntidadesAccesoRRAA.objects.none(), prefix='entidadacceso')

    elif request.method == "POST":

        create_ra_form = CreateRAForm(request.POST)   #form crear ra
        norma_ra_form = MB_NormaRRAAForm(request.POST)
        documentoMetodolog_ra_form = MB_DocumentoMetodologRRAAForm(request.POST)
        concepEstan_ra_form = MB_ConceptosEstandarizadosRRAAForm(request.POST)
        clasificacion_ra_form = MB_ClasificacionesRRAAForm(request.POST)
        recolecDatos_ra_form = MB_RecoleccionDatosRRAAForm(request.POST)
        frecuencRecole_ra_form = MB_FrecuenciaRecoleccionDatoForm(request.POST)
        heramiUtiPr_ra_form = MB_HerramientasUtilProcesaForm(request.POST)
        seguInfor_ra_form = MB_SeguridadInformForm(request.POST)
        frecuencAlma_ra_form = MB_FrecuenciaAlmacenamientobdForm(request.POST)
        cobertGeogra_ra_form =  MB_CoberturaGeograficaRRAAForm(request.POST)
        noAccesoDato_ra_form = MB_NoAccesoDatosForm(request.POST)
        formvariableRecolset = MB_VariableRecolectadaFormset(request.POST, prefix='variablerecolectada')
        formIndicadorReAgreset = MB_IndicadorResultadoAgregadoFormset(request.POST, prefix='indicadorresultadoagregado')
        formEntidadAcceset = MB_EntidadesAccesoFormset(request.POST, prefix='entidadacceso')

        ## Novedades para cuando es creada
        novedad_rraaa_form = NovedadActualizacionRRAAForm(data=request.POST)
        
        if create_ra_form.is_valid() and concepEstan_ra_form.is_valid() and norma_ra_form.is_valid() and documentoMetodolog_ra_form.is_valid() and \
            clasificacion_ra_form.is_valid() and recolecDatos_ra_form.is_valid() and frecuencRecole_ra_form.is_valid() and heramiUtiPr_ra_form.is_valid() and \
                seguInfor_ra_form.is_valid() and frecuencAlma_ra_form.is_valid() and cobertGeogra_ra_form.is_valid() and \
                    formEntidadAcceset.is_valid() and noAccesoDato_ra_form.is_valid() and formvariableRecolset.is_valid() and formIndicadorReAgreset.is_valid() and novedad_rraaa_form.is_valid():

            instancia = create_ra_form.save(commit=False)

            generate_cod = count_ra + 148
            instancia.codigo_rraa = "RA" + str(generate_cod)
            instancia.nombre_diligenciador =  request.user

            if 'variableRecol_file' in request.FILES:
                instancia.variableRecol_file = request.FILES['variableRecol_file']

            instancia.save()
            obj_id = instancia.id  ##obtener id del objeto que estoy creando
            ## formset pregunta 7
            for formvariableRecol in formvariableRecolset:
                # so that `rraa` instance can be attached.
                formVariableRecoleccion = formvariableRecol.save(commit=False)
                formVariableRecoleccion.rraa_id = obj_id
                formVariableRecoleccion.instancia = instancia
                if formVariableRecoleccion.variableRec == "":
                    formVariableRecoleccion = formvariableRecol.save(commit=False)
                else: 
                    formVariableRecoleccion.save()
            
            ## formset pregunta 16

            for formIndicadorResultadosAg in formIndicadorReAgreset:
                # so that `rraa` instance can be attached.
                formIndResultadosAgregados = formIndicadorResultadosAg.save(commit=False)
                formIndResultadosAgregados.rraa_id = obj_id
                formIndResultadosAgregados.instancia = instancia
                if formIndResultadosAgregados.ind_res_agre == "":
                    formIndResultadosAgregados = formIndicadorResultadosAg.save(commit=False)
                else: 
                    formIndResultadosAgregados.save()


            ## formset pregunta 18

            for formEntiAcc in formEntidadAcceset:

                formEntidadAcceso_ra = formEntiAcc.save(commit=False)
                formEntidadAcceso_ra.rraa_id = obj_id
                formEntidadAcceso_ra.instancia = instancia
                if formEntidadAcceso_ra.nomb_entidad_acc == "":
                    formEntidadAcceso_ra = formEntiAcc.save(commit=False)
                else:
                    # Save the entitie to the database
                    formEntidadAcceso_ra.save()
                    formEntiAcc.save_m2m() 
                     

            addnorma_ra = norma_ra_form.save(commit=False) 
            addnorma_ra.rraa_id = obj_id
            addnorma_ra.save()

            adddocumentoMetodo_ra = documentoMetodolog_ra_form.save(commit=False)
            adddocumentoMetodo_ra.rraa_id = obj_id
            adddocumentoMetodo_ra.save()

            addconcepEstan_ra = concepEstan_ra_form.save(commit=False)
            addconcepEstan_ra.rraa_id = obj_id
            addconcepEstan_ra.save()

            addclasificacion_ra = clasificacion_ra_form.save(commit=False)
            addclasificacion_ra.rraa_id = obj_id
            addclasificacion_ra.save()

            addrecolecDatos_ra = recolecDatos_ra_form.save(commit=False) 
            addrecolecDatos_ra.rraa_id = obj_id
            addrecolecDatos_ra.save()
            
            addfrecuencRecole_ra = frecuencRecole_ra_form.save(commit=False) 
            addfrecuencRecole_ra.rraa_id = obj_id
            addfrecuencRecole_ra.save()
            
            addheramiUtiPr_ra = heramiUtiPr_ra_form.save(commit=False)
            addheramiUtiPr_ra.rraa_id = obj_id
            addheramiUtiPr_ra.save()

            addseguInfor_ra = seguInfor_ra_form.save(commit=False)
            addseguInfor_ra.rraa_id = obj_id
            addseguInfor_ra.save()

                    
            addfrecuencAlma_ra = frecuencAlma_ra_form.save(commit=False)
            addfrecuencAlma_ra.rraa_id = obj_id
            addfrecuencAlma_ra.save()
                        
            addcobertGeogra_ra = cobertGeogra_ra_form.save(commit=False)
            addcobertGeogra_ra.rraa_id = obj_id
            addcobertGeogra_ra.save()


            addnoAccesoDato_ra = noAccesoDato_ra_form.save(commit=False)
            addnoAccesoDato_ra.rraa_id = obj_id
            addnoAccesoDato_ra.save()


        ##  novedades
            new_novedad_rraa = novedad_rraa_form.save(commit=False)
            new_novedad_rraa.post_ra_id = obj_id
            new_novedad_rraa.name_nov = request.user
            new_novedad_rraa.est_actualiz_id = 4
            new_novedad_rraa.descrip_novedad = "Se crea registro administrativo."
            new_novedad_rraa.save()

        ## end form 
            create_ra_form.save_m2m() ## metodo para guardar relaciones manyTomany
            registered = True
            # redirect to a new URL:
            #print("save")
            #return redirect('rraa:all_rraa')
        else:    
            
            #print("no guardo____________RRAA_______________")
            messages.error(request, "Las respuestas NO se han guardado verifique si tiene errores")
    
        
    return render(request, 'rraa/created_rraa.html', { 'count_entities': count_entities, 'create_ra_form': create_ra_form, 'norma_ra_form' : norma_ra_form,
    'documentoMetodolog_ra_form': documentoMetodolog_ra_form, 'concepEstan_ra_form': concepEstan_ra_form, 'clasificacion_ra_form' : clasificacion_ra_form,
    'recolecDatos_ra_form': recolecDatos_ra_form,'frecuencRecole_ra_form': frecuencRecole_ra_form, 'heramiUtiPr_ra_form': heramiUtiPr_ra_form, 'seguInfor_ra_form': seguInfor_ra_form,
    'frecuencAlma_ra_form' : frecuencAlma_ra_form, 'cobertGeogra_ra_form': cobertGeogra_ra_form, 'formvariableRecolset': formvariableRecolset,
    'formIndicadorReAgreset': formIndicadorReAgreset, 'formEntidadAcceset': formEntidadAcceset, 'noAccesoDato_ra_form':  noAccesoDato_ra_form,
    'registered': registered, 'new_novedad_rraa': new_novedad_rraa, 'novedad_rraa_form': novedad_rraa_form, 
    'entidad_responsable': entidad_responsable })


@login_required
def nuevos_registros(request):

    count_entities = Entidades_oe.objects.filter(estado_id=1).count()
    user = request.user
    entidadByUser = user.profile.entidad.id

    if user.is_authenticated == True and str(user) == "sdazag": # 1

        filtros_tema_oe = RegistroAdministrativo.objects.filter(Q (tema_id=2) | Q (tema_id=7) | Q (tema_id=19) | Q (tema_id=28))

        novedad = []
        unique_ra = NovedadActualizacionRRAA.objects.values('post_ra').order_by('post_ra').distinct()
        for nov in unique_ra: 
            lista_ra = list(nov.values())
            ra_nov = str(lista_ra)[1:-1]
            novedad_last = NovedadActualizacionRRAA.objects.filter(post_ra_id=ra_nov).last()
            if  str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"ra_identify": int(ra_nov), "last_novedad": str(novedad_last.est_actualiz),  "nombre_ra": str(novedad_last.post_ra) })

    elif user.is_authenticated == True and str(user) == "dvlizarazog": # 4 
        
        filtros_tema_ra = RegistroAdministrativo.objects.filter(Q (tema_id=9) | Q (tema_id=29) | Q (tema_id=30))
        novedad = []
        unique_ra = NovedadActualizacionRRAA.objects.values('post_ra').order_by('post_ra').distinct()
        for nov in unique_ra: 
            lista_ra = list(nov.values())
            ra_nov = str(lista_ra)[1:-1]
            novedad_last = NovedadActualizacionRRAA.objects.filter(post_ra_id=ra_nov).last()
            if  str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"ra_identify": int(ra_nov), "last_novedad": str(novedad_last.est_actualiz),  "nombre_ra": str(novedad_last.post_ra) })
       
    elif user.is_authenticated == True and str(user) == "ancardenasc": # 5 frherreran freddy es administrador se cambia por  gavargasr
        
        filtros_tema_ra = RegistroAdministrativo.objects.filter(tema_id=10)
        novedad = []
        unique_ra = NovedadActualizacionRRAA.objects.values('post_ra').order_by('post_ra').distinct()
        for nov in unique_ra: 
            lista_ra = list(nov.values())
            ra_nov = str(lista_ra)[1:-1]
            novedad_last = NovedadActualizacionRRAA.objects.filter(post_ra_id=ra_nov).last()
            if  str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"ra_identify": int(ra_nov), "last_novedad": str(novedad_last.est_actualiz),  "nombre_ra": str(novedad_last.post_ra) })
       

    elif user.is_authenticated == True and str(user) == "lhsanchezz": # 6 no continua
        
        filtros_tema_ra = RegistroAdministrativo.objects.filter(Q (tema_id=3) | Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25))
        novedad = []
        unique_ra = NovedadActualizacionRRAA.objects.values('post_ra').order_by('post_ra').distinct()
        for nov in unique_ra: 
            lista_ra = list(nov.values())
            ra_nov = str(lista_ra)[1:-1]
            novedad_last = NovedadActualizacionRRAA.objects.filter(post_ra_id=ra_nov).last()
            if  str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"ra_identify": int(ra_nov), "last_novedad": str(novedad_last.est_actualiz),  "nombre_ra": str(novedad_last.post_ra) })

    elif user.is_authenticated == True and str(user) == "pczambranog": # 6
        
        filtros_tema_ra = RegistroAdministrativo.objects.filter(Q (tema_id=11) | Q (tema_id=12) | Q (tema_id=25))
        novedad = []
        unique_ra = NovedadActualizacionRRAA.objects.values('post_ra').order_by('post_ra').distinct()
        for nov in unique_ra: 
            lista_ra = list(nov.values())
            ra_nov = str(lista_ra)[1:-1]
            novedad_last = NovedadActualizacionRRAA.objects.filter(post_ra_id=ra_nov).last()
            if  str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"ra_identify": int(ra_nov), "last_novedad": str(novedad_last.est_actualiz),  "nombre_ra": str(novedad_last.post_ra) })


    elif user.is_authenticated == True and str(user) == "mjpradag": # 7 mppulidor se cambia por mjpradag
    
        filtros_tema_ra = RegistroAdministrativo.objects.filter(Q (tema_id=13) | Q (tema_id=24) | Q (tema_id=26) | Q (tema_id=27))
        novedad = []
        unique_ra = NovedadActualizacionRRAA.objects.values('post_ra').order_by('post_ra').distinct()
        for nov in unique_ra: 
            lista_ra = list(nov.values())
            ra_nov = str(lista_ra)[1:-1]
            novedad_last = NovedadActualizacionRRAA.objects.filter(post_ra_id=ra_nov).last()
            if  str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"ra_identify": int(ra_nov), "last_novedad": str(novedad_last.est_actualiz),  "nombre_ra": str(novedad_last.post_ra) })


    elif user.is_authenticated == True and str(user) == "eeguayazans": # 8 eeguayazans Edgar Eduardo Guayazan Sierra
        
        filtros_tema_ra  = RegistroAdministrativo.objects.filter(Q (tema_id=4) | Q (tema_id=8) | Q (tema_id=20))
        novedad = []
        unique_ra = NovedadActualizacionRRAA.objects.values('post_ra').order_by('post_ra').distinct()
        for nov in unique_ra: 
            lista_ra = list(nov.values())
            ra_nov = str(lista_ra)[1:-1]
            novedad_last = NovedadActualizacionRRAA.objects.filter(post_ra_id=ra_nov).last()
            if  str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"ra_identify": int(ra_nov), "last_novedad": str(novedad_last.est_actualiz),  "nombre_ra": str(novedad_last.post_ra) })

    elif user.is_authenticated == True and str(user) == "mlbarretob": # 9

        filtros_tema_ra  = RegistroAdministrativo.objects.filter(Q (tema_id=21) | Q (tema_id=22) | Q (tema_id=23))
        novedad = []
        unique_ra = NovedadActualizacionRRAA.objects.values('post_ra').order_by('post_ra').distinct()
        for nov in unique_ra: 
            lista_ra = list(nov.values())
            ra_nov = str(lista_ra)[1:-1]
            novedad_last = NovedadActualizacionRRAA.objects.filter(post_ra_id=ra_nov).last()
            if  str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"ra_identify": int(ra_nov), "last_novedad": str(novedad_last.est_actualiz),  "nombre_ra": str(novedad_last.post_ra) })

    elif user.is_authenticated == True and user.profile.role.id == 1 or user.profile.role.id == 5 or user.profile.role.id == 6 or user.profile.role.id == 7: 

        filtros_tema_ra = RegistroAdministrativo.objects.all()
        novedad = []
        unique_ra = NovedadActualizacionRRAA.objects.values('post_ra').order_by('post_ra').distinct()
        for nov in unique_ra: 
            lista_ra = list(nov.values())
            ra_nov = str(lista_ra)[1:-1]
            novedad_last = NovedadActualizacionRRAA.objects.filter(post_ra_id=ra_nov).last()
            if  str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"ra_identify": int(ra_nov), "last_novedad": str(novedad_last.est_actualiz),  "nombre_ra": str(novedad_last.post_ra) })
    
    else:
        
        filtros_tema_ra  = RegistroAdministrativo.objects.filter(entidad_pri_id=entidadByUser)
        novedad = []
        unique_ra = NovedadActualizacionRRAA.objects.values('post_ra').order_by('post_ra').distinct()
        for nov in unique_ra: 
            lista_ra = list(nov.values())
            ra_nov = str(lista_ra)[1:-1]
            novedad_last = NovedadActualizacionRRAA.objects.filter(post_ra_id=ra_nov).last()
            if  str(novedad_last.est_actualiz) == 'Nueva':
                novedad.append({"ra_identify": int(ra_nov), "last_novedad": str(novedad_last.est_actualiz),  "nombre_ra": str(novedad_last.post_ra) })

    return render(request, 'rraa/nuevas_rraa.html', { 'novedad': novedad, 'count_entities': count_entities, 'filtros_tema_ra': filtros_tema_ra })





## correo para notificar a la fuente cuando se cambia el estado a devuelto
def sendEmailStatusChangeRa(subject, recipients, body):    
    send_mail(subject, body, 'sen@dane.gov.co', recipients,  fail_silently = False)  

def emailNotificationRRAAStateView(request, ra, listaUsuarios):
    
    subject = 'Cambio de estado en formulario de Registro Administrativo '+ str(ra).replace('\n', '').replace('\r', '')

    recipients = listaUsuarios
      
    body = 'SISTEMA DE IDENTIFICACIÓN Y CARACTERIZACIÓN DE OFERTA Y DEMANDA ESTADÍSTICA DEL SEN\n\n' \
            'El registro administrativo: '+ str(ra) + \
			    '\n\n Ha sido DEVUELTO(A)\n' \
                    '\n Por favor ingrese al aplicativo, revise los comentarios generados por el temático y realice los ajustes correspondientes.'

    return subject, recipients, body

## end correo para notificar a la fuente cuando se cambia el estado a devuelto


def sendEmailNotification(subject, recipients, body):    
    send_mail(subject, body, 'sen@dane.gov.co', recipients,  fail_silently = False)  


## envio de notificaciones al editar RRAA
def createEmailEditRRAA(request, ra, temaRes, fieldsEdited_ra, fieldsFormset1, fieldsFormset2, fieldsFormset3):

    #print("que llega", fieldsEdited_ra)
    subject = 'Cambios en formulario de registro administrativo '+ str(ra).replace('\n', '').replace('\r', '')
   
    if temaRes == 2 or temaRes == 7 or temaRes == 19 or temaRes == 28: 

        responsable_ra = 'sdazag@dane.gov.co' #sdazag
           
    elif temaRes == 5 or temaRes == 14 or temaRes ==  15 or temaRes ==  16 or temaRes ==  17 or temaRes ==  18: 

        responsable_ra = 'aobandor@dane.gov.co' # aobandor
          
    elif temaRes == 9 or temaRes == 29 or temaRes == 30:
        
        responsable_ra = 'dvlizarazog@dane.gov.co' # dvlizarazog
    
    elif temaRes == 10 or temaRes == 6 or temaRes == 1:
        
        responsable_ra = 'gavargasr@dane.gov.co' # (frherreran este usuario no sigue en dane se reemplaza por gavargasr)
        
    #elif temaRes == 11 or temaRes == 12 or temaRes == 25 or temaRes == 3: 
        
    #    responsable_ra = 'lhsanchezz@dane.gov.co' # lhsanchezz

    elif temaRes == 11 or temaRes == 12 or temaRes == 25: 
        
        responsable_ra = 'pczambranog@dane.gov.co' # pczambranog
         
    elif temaRes == 13 or temaRes == 24 or temaRes == 26 or temaRes == 27: 
        
        responsable_ra = 'mjpradag@dane.gov.co'  # 7 mppulidor se cambia por mjpradag

    ### Nota importante:  se pone responsable a ruth mientras ingresan los contratistas lhsanchezz
    elif temaRes == 11 or temaRes == 12 or temaRes == 25 or temaRes == 3: 
        responsable_ra = 'rctrianaa@dane.gov.co' # rctrianaa
    ### Nota importante: se pone responsable a ruth mientras ingresan los contratistas lhsanchezz
          
    elif temaRes == 4 or temaRes == 8 or temaRes == 20: 

        responsable_ra = 'eeguayazans@dane.gov.co' # eeguayazans
      
    elif temaRes == 21 or temaRes == 22 or temaRes == 23:  
        
        responsable_ra = 'mlbarretob@dane.gov.co' # mlbarretob
                 
    recipients = [ responsable_ra ]
      
    body = 'SISTEMA DE IDENTIFICACIÓN Y CARACTERIZACIÓN DE OFERTA Y DEMANDA ESTADÍSTICA DEL SEN\n\n' \
            'El registro administrativo: '+ str(ra) + \
			    '\n\nha sido editado en los siguientes campos:\n' \
                   '\n' + str(fieldsEdited_ra).replace('[','').replace(']','').replace(',', '\n').replace("'", "") + \
                    '\n' + str(fieldsFormset1).replace('[','').replace(']','').replace(',', '\n').replace("'", "") + \
                    '\n' + str(fieldsFormset2).replace('[','').replace(']','').replace(',', '\n').replace("'", "") + \
                    '\n' + str(fieldsFormset3).replace('[','').replace(']','').replace(',', '\n').replace("'", "")                                                       
    
    
    return subject, recipients, body



@login_required
def editRRAAView(request, pk):

    count_entities = Entidades_oe.objects.count()
    ra_id =  RegistroAdministrativo.objects.get(pk=pk)
    post_ra = get_object_or_404(RegistroAdministrativo, pk=pk)
    count_ra = RegistroAdministrativo.objects.count()

    user = request.user
    roleUser = user.profile.role.id
    entidad_responsable = ra_id.entidad_pri.pk
    sistema_estado = ra_id.sist_estado.pk
    #print("rol", roleUser)

    ## traer listado de usuarios responsable de rraa
    user_email = []
    user_responsable = Profile.objects.filter(entidad=ra_id.entidad_pri.pk)
    for use in list(chain(user_responsable)):
        user_email.append(str(use.user.email))
    #print("array de usuarios",user_email)

    ## traer listado de usuarios responsable de la rraa

    ## models  id ra 
    norma_ra_id = MB_NormaRRAA.objects.get(rraa_id=ra_id.pk)
    docuMetodolog_ra_id = MB_DocumentoMetodologRRAA.objects.get(rraa_id=ra_id.pk)
    concepEstand_ra_id = MB_ConceptosEstandarizadosRRAA.objects.get(rraa_id=ra_id.pk)
    clasificacion_ra_id = MB_ClasificacionesRRAA.objects.get(rraa_id=ra_id.pk)
    recolDatos_ra_id = MB_RecoleccionDatosRRAA.objects.get(rraa_id=ra_id.pk)
    frecRecolDato_ra_id = MB_FrecuenciaRecoleccionDato.objects.get(rraa_id=ra_id.pk)
    herraUtilProcesa_ra_id = MB_HerramientasUtilProcesa.objects.get(rraa_id=ra_id.pk)
    seguridadInform_ra_id = MB_SeguridadInform.objects.get(rraa_id=ra_id.pk)
    frecAlmacebd_ra_id = MB_FrecuenciaAlmacenamientobd.objects.get(rraa_id=ra_id.pk)
    coberturaGeografica_ra_id = MB_CoberturaGeograficaRRAA.objects.get(rraa_id=ra_id.pk)
    noAccesoDato_ra_id = MB_NoAccesoDatos.objects.get(rraa_id=ra_id.pk)
    
    ## end models id ra  ///MB_VariableRecolectada, MB_IndicadorResultadoAgregado

    ## edit form by id 
    edit_ra_form = EditRAForm(instance=ra_id)
    editnorma_ra_form = EditMB_NormaRRAAForm(instance=norma_ra_id)
    editdocumentoMetodolog_ra_form = EditMB_DocumentoMetodologRRAAForm(instance=docuMetodolog_ra_id)
    editconcepEstan_ra_form = EditMB_ConceptosEstandarizadosRRAAForm(instance=concepEstand_ra_id)
    editclasificacion_ra_form = EditMB_ClasificacionesRRAAForm(instance=clasificacion_ra_id)
    editrecolecDatos_ra_form = EditMB_RecoleccionDatosRRAAForm(instance=recolDatos_ra_id)
    editfrecuencRecole_ra_form = EditMB_FrecuenciaRecoleccionDatoForm(instance=frecRecolDato_ra_id)
    editheramiUtiPr_ra_form = EditMB_HerramientasUtilProcesaForm(instance=herraUtilProcesa_ra_id)
    editseguInfor_ra_form = EditMB_SeguridadInformForm(instance=seguridadInform_ra_id)
    editfrecuencAlma_ra_form = EditMB_FrecuenciaAlmacenamientobdForm(instance=frecAlmacebd_ra_id)
    editcobertGeogra_ra_form =  EditMB_CoberturaGeograficaRRAAForm(instance=coberturaGeografica_ra_id)
    editnoAccesoDato_ra_form = EditMB_NoAccesoDatosForm(instance=noAccesoDato_ra_id)
    ## end edit form by id

    ## comentarios
    comment_ra_form = CommentRRAAForm
    comments = post_ra.commentrraa.filter(active=True)
    new_comment_ra = None

    ## Novedades actualización
    novedad_ra_form = NovedadActualizacionRRAAForm()
    novedad = post_ra.novedadactualizacionrraa.filter(active=True)
    new_novedad_ra = None

    ##  mostrar el ultimo estado y fecha de actualización
    if post_ra.novedadactualizacionrraa.filter(active=True).exists():
        for nove in novedad:
            nove.est_actualiz
        if nove.est_actualiz != None:
            estadoNovedadRA = nove.est_actualiz
            fechaNovedadRA = nove.fecha_actualiz
        else:
            estadoNovedadRA = ""
            fechaNovedadRA = ""
    else:
        estadoNovedadRA = ""
        fechaNovedadRA = ""

    ##  end mostrar el ultimo estado y fecha de actualización
    
    ## Critica

    critica_ra_form = CriticaRRAAForm()
    criticas = post_ra.critica_ra.filter(active=True)
    new_critica_ra = None

    args = {}
    if request.method == 'GET':

        editformvariableRecolset = MB_VariableRecolectadaFormset(queryset=MB_VariableRecolectada.objects.filter(rraa_id=ra_id.pk), prefix='variablerecolectada')
        editformIndicadorReAgreset = MB_IndicadorResultadoAgregadoFormset(queryset=MB_IndicadorResultadoAgregado.objects.filter(rraa_id=ra_id.pk), prefix='indicadorresultadoagregado')
        editformEntidadAcceset = MB_EntidadesAccesoFormset(queryset=MB_EntidadesAccesoRRAA.objects.filter(rraa_id=ra_id.pk), prefix='entidadacceso')

    if request.method == "POST":

        edit_ra_form = EditRAForm(request.POST, request.FILES, instance=ra_id)
        editnorma_ra_form = EditMB_NormaRRAAForm(request.POST, instance=norma_ra_id)
        editdocumentoMetodolog_ra_form = EditMB_DocumentoMetodologRRAAForm(request.POST, instance=docuMetodolog_ra_id)
        editconcepEstan_ra_form = EditMB_ConceptosEstandarizadosRRAAForm(request.POST, instance=concepEstand_ra_id)
        editclasificacion_ra_form = EditMB_ClasificacionesRRAAForm(request.POST, instance=clasificacion_ra_id)
        editrecolecDatos_ra_form = EditMB_RecoleccionDatosRRAAForm(request.POST, instance=recolDatos_ra_id)
        editfrecuencRecole_ra_form = EditMB_FrecuenciaRecoleccionDatoForm(request.POST, instance=frecRecolDato_ra_id)
        editheramiUtiPr_ra_form = EditMB_HerramientasUtilProcesaForm(request.POST, instance=herraUtilProcesa_ra_id)
        editseguInfor_ra_form = EditMB_SeguridadInformForm(request.POST, instance=seguridadInform_ra_id)
        editfrecuencAlma_ra_form = EditMB_FrecuenciaAlmacenamientobdForm(request.POST, instance=frecAlmacebd_ra_id)
        editcobertGeogra_ra_form =  EditMB_CoberturaGeograficaRRAAForm(request.POST, instance=coberturaGeografica_ra_id)
        editnoAccesoDato_ra_form = EditMB_NoAccesoDatosForm(request.POST, instance=noAccesoDato_ra_id)
        ## formsets ***
        editformvariableRecolset = MB_VariableRecolectadaFormset(request.POST, prefix='variablerecolectada')
        editformIndicadorReAgreset = MB_IndicadorResultadoAgregadoFormset(request.POST, prefix='indicadorresultadoagregado')
        editformEntidadAcceset = MB_EntidadesAccesoFormset(request.POST, prefix='entidadacceso')
        ##Comentarios ***    
        comment_ra_form = CommentRRAAForm(data=request.POST)
        ## Novedad -Actualización
        novedad_ra_form = NovedadActualizacionRRAAForm(data=request.POST)
        ##Critica
        critica_ra_form = CriticaRRAAForm(data=request.POST)

        ## detectar cambios
        fieldsEdited_ra = []
        fieldsFormset1 = []
        fieldsFormset2 = []
        fieldsFormset3 = []

        if edit_ra_form.has_changed() or editformvariableRecolset.has_changed() or editformIndicadorReAgreset.has_changed() or editformEntidadAcceset.has_changed():
            #print("The following fields changed: %s" % ", ".join(edit_ra_form.changed_data))
            #print("contador de cambios", edit_ra_form.changed_data)
              
            for index, item in enumerate(edit_ra_form.changed_data):

                edit_ra_form.fields[item].widget.attrs['title']
                questionParamTitle = edit_ra_form.fields[item].widget.attrs['title']
                fieldsEdited_ra.append(questionParamTitle)# array que almacena lista de campos editados
            if editformvariableRecolset.has_changed() == True:
                fieldsFormset1 = ["Módulo B Pregunta 7: Indique cuáles son las variables recolectadas con el RA"]

            if editformIndicadorReAgreset.has_changed() == True:
                fieldsFormset2 = ["Módulo B Pregunta 16: ¿La entidad hace uso de los datos del RA para generar información estadística ?"]

            if editformEntidadAcceset.has_changed() == True:
                fieldsFormset3 = ["Módulo B Pregunta 18: Relacione las entidades que tienen acceso a los datos del RA"]
            
        if edit_ra_form.is_valid() and editconcepEstan_ra_form.is_valid() and editnorma_ra_form.is_valid() and editdocumentoMetodolog_ra_form.is_valid() and \
            editclasificacion_ra_form.is_valid() and editrecolecDatos_ra_form.is_valid() and editfrecuencRecole_ra_form.is_valid() and editheramiUtiPr_ra_form.is_valid() and \
                editseguInfor_ra_form.is_valid() and editfrecuencAlma_ra_form.is_valid() and editcobertGeogra_ra_form.is_valid() and \
                    editnoAccesoDato_ra_form.is_valid() and editformEntidadAcceset.is_valid() and editformvariableRecolset.is_valid() and editformIndicadorReAgreset.is_valid() and \
                        comment_ra_form.is_valid() and critica_ra_form.is_valid() and novedad_ra_form.is_valid():

            instancia_edit = edit_ra_form.save(commit=False)
            #instancia_edit.codigo_rraa = "RA" + str(ra_id.pk)

            instancia_edit.nombre_diligenciador =  request.user

            ## validaciones según el rol para cambio de estado

            ## Si rol fuente  --> edita rraa publicada
            if roleUser == 3 and instancia_edit.sist_estado_id == 5:
                instancia_edit.sist_estado_id = 2

            elif roleUser == 3 and instancia_edit.sist_estado_id == 4:
                instancia_edit.sist_estado_id = 2

            ## end validaciones según el rol para cambio de estado
            
            temaRes = instancia_edit.tema_id

            #envio de correo para notificar cuando el estado cambia a devuelto ****
        
            if instancia_edit.sist_estado_id == 3 and instancia_edit.entidad_pri_id != 1:
                subject, recipients, body = emailNotificationRRAAStateView(user, post_ra, user_email)
                sendEmailStatusChangeRa(subject, recipients, body)
            #end envio de correo para notificar cuando el estado cambia a devuelto


            if 'variableRecol_file' in request.FILES:
                instancia_edit.variableRecol_file = request.FILES['variableRecol_file']


            ## send correo Notificaciones ***
            subject, recipients, body = createEmailEditRRAA(user, post_ra, temaRes, fieldsEdited_ra, fieldsFormset1, fieldsFormset2, fieldsFormset3 )
            sendEmailNotification(subject, recipients, body)
            ## end send correo Notificaciones


            instancia_edit.save()
            obj_id = instancia_edit.id  ##obtener id del objeto que voy a actualizar
            
            ## forms

            editnorma_ra = editnorma_ra_form.save(commit=False) 
            editnorma_ra.save()

            editdocumentoMetodo_ra = editdocumentoMetodolog_ra_form.save(commit=False)
            editdocumentoMetodo_ra.save()

            editconcepEstan_ra = editconcepEstan_ra_form.save(commit=False)
            editconcepEstan_ra.save()

            editclasific_ra = editclasificacion_ra_form.save(commit=False)
            editclasific_ra.save()

            editrecolecDatos_ra = editrecolecDatos_ra_form.save(commit=False) 
            editrecolecDatos_ra.save()
            
            editfrecuencRecole_ra = editfrecuencRecole_ra_form.save(commit=False)
            editfrecuencRecole_ra.save()
            
            editheramiUtiPr_ra = editheramiUtiPr_ra_form.save(commit=False)
            editheramiUtiPr_ra.save()

            editseguInfor_ra = editseguInfor_ra_form.save(commit=False)
            editseguInfor_ra.save()
    
            editfrecuencAlma_ra = editfrecuencAlma_ra_form.save(commit=False)
            editfrecuencAlma_ra.save()
                        
            editcobertGeogra_ra = editcobertGeogra_ra_form.save(commit=False)
            editcobertGeogra_ra.save()

            editnoAccesoDato_ra = editnoAccesoDato_ra_form.save(commit=False)
            editnoAccesoDato_ra.save()

            ## formset pregunta 7 ****

            formvariableRecol = editformvariableRecolset.save(commit=False)
           
            if editformvariableRecolset.deleted_forms:
                for obj in editformvariableRecolset.deleted_objects:
                    obj.delete() 
            else:
                for formvariableRecol in editformvariableRecolset:
                # so that `ooee` instance can be attached.
                    formVariableRecoleccion = formvariableRecol.save(commit=False)
                    formVariableRecoleccion.rraa_id = obj_id
                    formVariableRecoleccion.instancia_edit = instancia_edit
                    if formVariableRecoleccion.variableRec == "":
                        formVariableRecoleccion = formvariableRecol.save(commit=False)
                    else:
                        # Save listvar to the database
                        formVariableRecoleccion.save() 
            
            ## formset pregunta 16 ***

            formIndicadorResultadosAg = editformIndicadorReAgreset.save(commit=False)
           
            if editformIndicadorReAgreset.deleted_forms:
                for obj in editformIndicadorReAgreset.deleted_objects:
                    obj.delete() 
            else:
                for formIndicadorResultadosAg in editformIndicadorReAgreset:
                # so that `ooee` instance can be attached.
                    formIndResultadosAgregados = formIndicadorResultadosAg.save(commit=False)
                    formIndResultadosAgregados.rraa_id = obj_id
                    formIndResultadosAgregados.instancia_edit = instancia_edit
                    if formIndResultadosAgregados.ind_res_agre == "":
                        formIndResultadosAgregados = formIndicadorResultadosAg.save(commit=False)
                    else:
                        # Save listvar to the database
                        formIndResultadosAgregados.save()


            ## formset pregunta 18 ***

            formEntiAcc = editformEntidadAcceset.save(commit=False)
           
            if editformEntidadAcceset.deleted_forms:
                for obj in editformEntidadAcceset.deleted_objects:
                    obj.delete() 
            else:
                for formEntiAcc in editformEntidadAcceset:

                    formEntidadAcceso_ra = formEntiAcc.save(commit=False)
                    formEntidadAcceso_ra.rraa_id = obj_id
                    formEntidadAcceso_ra.instancia = instancia_edit
                    if formEntidadAcceso_ra.nomb_entidad_acc == "":
                        formEntidadAcceso_ra = formEntiAcc.save(commit=False)
                    else:
                        # Save the entitie to the database
                        formEntidadAcceso_ra.save()
                        formEntiAcc.save_m2m()

            ## comentarios
            new_comment_ra = comment_ra_form.save(commit=False)
            new_comment_ra.name = request.user
            new_comment_ra.post_ra = post_ra
            if new_comment_ra.body == "":
                new_comment_ra = comment_ra_form.save(commit=False)
            else:
            # Save the comment to the database
                new_comment_ra.save()


            ## Novedad Actualización
            new_novedad_ra = novedad_ra_form.save(commit=False)
            new_novedad_ra.name_nov = request.user
            new_novedad_ra.post_ra = post_ra
            if new_novedad_ra.descrip_novedad == "":
                new_novedad_ra = novedad_ra_form.save(commit=False)
            else:
                new_novedad_ra.save()


            ## critica
            new_critica_ra = critica_ra_form.save(commit=False)
            new_critica_ra.user_critico = request.user
            new_critica_ra.post_ra = post_ra
            if new_critica_ra.observa_critica == "":
                new_critica_ra = critica_ra_form.save(commit=False)
            else:
                new_critica_ra.save()

        ## end form 
            edit_ra_form.save_m2m() ## metodo para guardar relaciones manyTomany
            
            # redirect to a new URL:
            #print("save")
            messages.success(request, 'Las respuestas se han guardado con éxito')
            
            #return redirect('rraa:all_rraa')
            return HttpResponseRedirect(request.path_info)
            
        else:    
            
            #print("no guardo____________RRAA_______________")
            messages.error(request, "Las respuestas NO se han guardado verifique si tiene errores")
    
        
    return render(request, 'rraa/edit_rraa.html', { 'count_entities': count_entities, 'edit_ra_form': edit_ra_form, 'editnorma_ra_form' : editnorma_ra_form,
    'editdocumentoMetodolog_ra_form': editdocumentoMetodolog_ra_form, 'editconcepEstan_ra_form': editconcepEstan_ra_form, 'editclasificacion_ra_form': editclasificacion_ra_form,
    'editrecolecDatos_ra_form': editrecolecDatos_ra_form, 'editfrecuencRecole_ra_form': editfrecuencRecole_ra_form, 'editheramiUtiPr_ra_form': editheramiUtiPr_ra_form, 
    'editseguInfor_ra_form': editseguInfor_ra_form, 'editfrecuencAlma_ra_form' : editfrecuencAlma_ra_form, 'editcobertGeogra_ra_form': editcobertGeogra_ra_form,
    'editformvariableRecolset': editformvariableRecolset, 'editformIndicadorReAgreset': editformIndicadorReAgreset, 'editformEntidadAcceset': editformEntidadAcceset, 
    'editnoAccesoDato_ra_form':  editnoAccesoDato_ra_form, 'entidad_responsable': entidad_responsable, 'sistema_estado': sistema_estado,
    'comments': comments, 'new_comment_ra': new_comment_ra, 'comment_ra_form': comment_ra_form,
    'criticas': criticas, 'new_critica_ra': new_critica_ra, 'critica_ra_form': critica_ra_form,
    'novedad': novedad, 'new_novedad_ra': new_novedad_ra, 'novedad_ra_form': novedad_ra_form, 'estadoNovedadRA': estadoNovedadRA, 
    'fechaNovedadRA': fechaNovedadRA
    })


def detailRRAA(request, pk):
    
    entities = Entidades_oe.objects.filter(estado_id=1) #entidades publicadas
    count_entities = entities.count()
    ra = get_object_or_404(RegistroAdministrativo, pk=pk)
    ## obtener informacion de RRAA
    normaFieldText = MB_NormaRRAA.objects.get(rraa_id=ra.pk)
    docuMetodologFieldText = MB_DocumentoMetodologRRAA(rraa_id=ra.pk)
    concepEstandFieldText = MB_ConceptosEstandarizadosRRAA.objects.get(rraa_id=ra.pk)
    recolDatosFieldText = MB_RecoleccionDatosRRAA.objects.get(rraa_id=ra.pk)
    frecRecolDatoFieldText = MB_FrecuenciaRecoleccionDato.objects.get(rraa_id=ra.pk)
    herraUtilProcesaFieldText = MB_HerramientasUtilProcesa.objects.get(rraa_id=ra.pk)
    seguridadInformFieldText = MB_SeguridadInform.objects.get(rraa_id=ra.pk)
    frecAlmacebdFieldText = MB_FrecuenciaAlmacenamientobd.objects.get(rraa_id=ra.pk)
    coberturaGeograficaFieldText = MB_CoberturaGeograficaRRAA.objects.get(rraa_id=ra.pk)
    noAccesoDatoFieldText = MB_NoAccesoDatos.objects.get(rraa_id=ra.pk)

    ##formsets ***
    variableRecolectadaFieldText = MB_VariableRecolectada.objects.filter(rraa_id=ra.pk) #pregunta 7
    indicadorResultadoAgregadoFieldText = MB_IndicadorResultadoAgregado.objects.filter(rraa_id=ra.pk) #pregunta 16

    ##estado de la novedad

    if NovedadActualizacionRRAA.objects.filter(post_ra=ra.pk).exists():
        estadoNovedad = NovedadActualizacionRRAA.objects.filter(post_ra=ra.pk)
        for stateNov in estadoNovedad:
            stateNov.est_actualiz
        if stateNov.est_actualiz != None:
            estadoNovRA = stateNov.est_actualiz
            fechaEstadoRA = stateNov.fecha_actualiz
        else:
            estadoNovRA = ""
            fechaEstadoRA = ""
    else:
        estadoNovRA = ""
        fechaEstadoRA = ""

    return render(request, 'rraa/rraa_detail.html', {'ra': ra, 'count_entities': count_entities, 
    'normaFieldText': normaFieldText, 'docuMetodologFieldText': docuMetodologFieldText,
    'concepEstandFieldText': concepEstandFieldText, 'recolDatosFieldText': recolDatosFieldText,
    'frecRecolDatoFieldText': frecRecolDatoFieldText, 'herraUtilProcesaFieldText': herraUtilProcesaFieldText,
    'seguridadInformFieldText' : seguridadInformFieldText, 'frecAlmacebdFieldText' : frecAlmacebdFieldText,
    'coberturaGeograficaFieldText': coberturaGeograficaFieldText, 'noAccesoDatoFieldText': noAccesoDatoFieldText,
    'variableRecolectadaFieldText': variableRecolectadaFieldText, 'indicadorResultadoAgregadoFieldText': indicadorResultadoAgregadoFieldText,
    'estadoNovRA': estadoNovRA, 'fechaEstadoRA': fechaEstadoRA })


@login_required
def create_FortalecimientoRRAA(request, pk):

    count_entities = Entidades_oe.objects.count() 
    post_ra = get_object_or_404(RegistroAdministrativo, pk=pk)
    #print("RRAA", post_ra)
    createFortalecimiento_ra_form = CreateFortalecimientoRRAAForm()
    planFortalecimiento = post_ra.fortalecimiento_rraa.filter(active=True)
    new_planFortalecimiento = None
    registered = False

    if request.method == "POST":

        createFortalecimiento_ra_form = CreateFortalecimientoRRAAForm(data=request.POST)

        if createFortalecimiento_ra_form.is_valid():

            new_planFortalecimiento = createFortalecimiento_ra_form.save(commit=False)
            new_planFortalecimiento.name_dilige =  request.user
            new_planFortalecimiento.post_ra = post_ra
            if new_planFortalecimiento.year_diagnostico == None:
                new_planFortalecimiento = createFortalecimiento_ra_form.save(commit=False)
            else:
                new_planFortalecimiento.save()
                registered = True
                #return redirect('rraa:all_rraa')
        else:
            
            createFortalecimiento_ra_form =  CreateFortalecimientoRRAAForm()
            messages.error(request, "Las respuestas NO se han guardado verifique si tiene errores")

    return render(request, 'fortalecimiento_rraa/createFortalecimiento_rraa.html', {'count_entities': count_entities, 'post_ra' : post_ra,
    'planFortalecimiento': planFortalecimiento, 'new_planFortalecimiento': new_planFortalecimiento,
    'createFortalecimiento_ra_form': createFortalecimiento_ra_form, 'registered': registered,
    })


@login_required
def edit_FortalecimientoRRAA(request, pk):

    count_entities = Entidades_oe.objects.count()
    fort_id = FortalecimientoRRAA.objects.get(pk=pk)
    post_ra = get_object_or_404(RegistroAdministrativo, pk=pk)

    editFortalecimiento_ra_form = EditFortalecimientoRRAAForm(instance=fort_id)
    
    if request.method == "POST":

        editFortalecimiento_ra_form = EditFortalecimientoRRAAForm(request.POST, instance=fort_id)

        if editFortalecimiento_ra_form.is_valid():

            new_fortalecimiento = editFortalecimiento_ra_form.save(commit=False)
            new_fortalecimiento.fecha_reg_sis = timezone.now()
            new_fortalecimiento.save()
            
            messages.success(request, 'los cambios se han guardado con éxito')
            
            return HttpResponseRedirect(request.path_info)
            
        else:
            messages.error(request, "Las cambios NO se han guardado verifique si tiene errores")

    return render(request, 'fortalecimiento_rraa/editFortalecimiento_rraa.html', {'count_entities': count_entities,
     'editFortalecimiento_ra_form': editFortalecimiento_ra_form, 'post_ra': post_ra })


########################### reporte de registros por estado publicado ##############################################

class reportRRAAfull_xls(TemplateView):  ###reporte por estado publicado
    def get(self, request, *args, **kwargs):
             
        rraas = RegistroAdministrativo.objects.filter(sist_estado=5)
        user = request.user
        
        normaTextList = []
        documentoMetodTextList = []
        variableRecolectadaTextList = []
        conceptosEstandarizadosTextList = []
        clasificacionTextList = []
        recoleccionDatosTextList = []
        FrecuenciaRecoleccionDatoTextList = []
        herramientasUtilProcesaTextList = []
        seguridadInformTextList = []
        frecuenciaAlmacenamientobdTextList = []
        coberturaGeograficaTextList = []
        indicadorResultadoAgregadoTextList = []
        entidadesAccesoRRAATextList = []
        noAccesoDatosTextList = []
        fortalecimientoTextList = []
        criticaTextList = [] #critica
        novedadTextList = [] #novedad
        
        objArray = list(rraas.values('id'))
        for obj in objArray:
            for key, value in obj.items():
                #print("value RRAA", value)
                normaQuestionList = MB_NormaRRAA.objects.filter(rraa_id=value)
                unionListNorma = list(chain(normaQuestionList))
                normaTextList.extend(unionListNorma)

                documentoMetodQuestionList = MB_DocumentoMetodologRRAA.objects.filter(rraa_id=value)
                unionListDocumentoMetod = list(chain(documentoMetodQuestionList))
                documentoMetodTextList.extend(unionListDocumentoMetod)

                variabRecolQuestionList = MB_VariableRecolectada.objects.filter(rraa_id=value)  # formset 7
                unionListVariabRecol = list(chain(variabRecolQuestionList))
                variableRecolectadaTextList.extend(unionListVariabRecol)

                concepEstandQuestionList = MB_ConceptosEstandarizadosRRAA.objects.filter(rraa_id=value)
                unionListConcepEstand = list(chain(concepEstandQuestionList))
                conceptosEstandarizadosTextList.extend(unionListConcepEstand)

                clasificacionQuestionList = MB_ClasificacionesRRAA.objects.filter(rraa_id=value)
                unionListClasificacion = list(chain(clasificacionQuestionList))
                clasificacionTextList.extend(unionListClasificacion)

                recolecDatoQuestionList = MB_RecoleccionDatosRRAA.objects.filter(rraa_id=value)
                unionListRecolecDato = list(chain(recolecDatoQuestionList))
                recoleccionDatosTextList.extend(unionListRecolecDato)

                frecRecolDatoQuestionList = MB_FrecuenciaRecoleccionDato.objects.filter(rraa_id=value)
                unionListFrecRecolDato = list(chain(frecRecolDatoQuestionList))
                FrecuenciaRecoleccionDatoTextList.extend(unionListFrecRecolDato)

                herramUtilProcQuestionList = MB_HerramientasUtilProcesa.objects.filter(rraa_id=value)
                unionListHerramUtilProc = list(chain(herramUtilProcQuestionList))
                herramientasUtilProcesaTextList.extend(unionListHerramUtilProc)

                seguridadInformQuestionList = MB_SeguridadInform.objects.filter(rraa_id=value)
                unionListSeguridadInform = list(chain(seguridadInformQuestionList))
                seguridadInformTextList.extend(unionListSeguridadInform)

                frecAlmacenQuestionList = MB_FrecuenciaAlmacenamientobd.objects.filter(rraa_id=value)
                unionListFrecAlmacen = list(chain(frecAlmacenQuestionList))
                frecuenciaAlmacenamientobdTextList.extend(unionListFrecAlmacen)

                coberGeogQuestionList = MB_CoberturaGeograficaRRAA.objects.filter(rraa_id=value)
                unionListCoberGeog = list(chain(coberGeogQuestionList))
                coberturaGeograficaTextList.extend(unionListCoberGeog)

                indResAgreQuestionList = MB_IndicadorResultadoAgregado.objects.filter(rraa_id=value)  ##formset 16
                unionListIndResAgre = list(chain(indResAgreQuestionList))
                indicadorResultadoAgregadoTextList.extend(unionListIndResAgre)

                entidadesAcceQuestionList = MB_EntidadesAccesoRRAA.objects.filter(rraa_id=value)  ##formset 18
                unionListEntidadesAcce = list(chain(entidadesAcceQuestionList))
                entidadesAccesoRRAATextList.extend(unionListEntidadesAcce)

                noAccesoDatosQuestionList = MB_NoAccesoDatos.objects.filter(rraa_id=value)
                unionListNoAccesoDatos = list(chain(noAccesoDatosQuestionList))
                noAccesoDatosTextList.extend(unionListNoAccesoDatos)

                fortalecimientoQuestionList = FortalecimientoRRAA.objects.filter(post_ra_id=value)  ## otra hoja
                unionListFortalecimiento = list(chain(fortalecimientoQuestionList))
                fortalecimientoTextList.extend(unionListFortalecimiento)

                criticaQuestionList = CriticaRRAA.objects.filter(post_ra_id=value)  ## critica
                unionListCritica = list(chain(criticaQuestionList))
                criticaTextList.extend(unionListCritica)

                novedadQuestionList = NovedadActualizacionRRAA.objects.filter(post_ra_id=value)  ## NovedadActualizacionRRAA
                unionListNovedad = list(chain(novedadQuestionList))
                novedadTextList.extend(unionListNovedad)

        wb = Workbook()
        ws = wb.active
        ws.title = "Directorio RRAA"
        sheet2 = wb.create_sheet('Variables RRAA')
        sheet3 = wb.create_sheet('Resultados RRAA')
        sheet4 = wb.create_sheet('Entidades RRAA')
        sheet5 = wb.create_sheet('Fortalecimiento RRAA')
        sheet6 = wb.create_sheet('Critica RRAA')
        sheet7 = wb.create_sheet('Novedad RRAA')

        
        def set_border(ws, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = ws[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(ws,'A1:DG'+str(rraas.count()+3))

		## size rows

        ws.row_dimensions[1].height = 55
        ws.row_dimensions[2].height = 40
        ws.row_dimensions[3].height = 80

        ## size column
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 15
        ws.column_dimensions['O'].width = 15
        ws.column_dimensions['P'].width = 15
        ws.column_dimensions['Q'].width = 15
        ws.column_dimensions['R'].width = 22
        ws.column_dimensions['S'].width = 15
        ws.column_dimensions['T'].width = 15
        ws.column_dimensions['U'].width = 15
        ws.column_dimensions['V'].width = 15
        ws.column_dimensions['W'].width = 15
        ws.column_dimensions['X'].width = 15
        ws.column_dimensions['Y'].width = 20
        ws.column_dimensions['Z'].width = 15
        ws.column_dimensions['AA'].width = 15
        ws.column_dimensions['AB'].width = 15
        ws.column_dimensions['AC'].width = 15
        ws.column_dimensions['AD'].width = 15
        ws.column_dimensions['AE'].width = 15
        ws.column_dimensions['AF'].width = 15
        ws.column_dimensions['AG'].width = 15
        ws.column_dimensions['AH'].width = 15
        ws.column_dimensions['AI'].width = 15
        ws.column_dimensions['AJ'].width = 15
        ws.column_dimensions['AK'].width = 15
        ws.column_dimensions['AL'].width = 15
        ws.column_dimensions['AM'].width = 15
        ws.column_dimensions['AN'].width = 15
        ws.column_dimensions['AO'].width = 15
        ws.column_dimensions['AP'].width = 15
        ws.column_dimensions['AQ'].width = 15
        ws.column_dimensions['AR'].width = 15
        ws.column_dimensions['AS'].width = 15
        ws.column_dimensions['AT'].width = 15
        ws.column_dimensions['AU'].width = 15
        ws.column_dimensions['AV'].width = 15
        ws.column_dimensions['AW'].width = 15
        ws.column_dimensions['AX'].width = 15
        ws.column_dimensions['AY'].width = 15
        ws.column_dimensions['AZ'].width = 15
        ws.column_dimensions['BA'].width = 15
        ws.column_dimensions['BB'].width = 15
        ws.column_dimensions['BC'].width = 15
        ws.column_dimensions['BD'].width = 15
        ws.column_dimensions['BE'].width = 15
        ws.column_dimensions['BF'].width = 15
        ws.column_dimensions['BG'].width = 15
        ws.column_dimensions['BH'].width = 15
        ws.column_dimensions['BI'].width = 15
        ws.column_dimensions['BJ'].width = 15
        ws.column_dimensions['BK'].width = 15
        ws.column_dimensions['BL'].width = 15
        ws.column_dimensions['BM'].width = 15
        ws.column_dimensions['BN'].width = 15
        ws.column_dimensions['BO'].width = 15
        ws.column_dimensions['BP'].width = 15
        ws.column_dimensions['BQ'].width = 15
        ws.column_dimensions['BR'].width = 15
        ws.column_dimensions['BS'].width = 15
        ws.column_dimensions['BT'].width = 15
        ws.column_dimensions['BU'].width = 15
        ws.column_dimensions['BV'].width = 15
        ws.column_dimensions['BW'].width = 15
        ws.column_dimensions['BX'].width = 15
        ws.column_dimensions['BY'].width = 15
        ws.column_dimensions['BZ'].width = 15
        ws.column_dimensions['CA'].width = 15
        ws.column_dimensions['CB'].width = 15
        ws.column_dimensions['CC'].width = 15
        ws.column_dimensions['CD'].width = 15
        ws.column_dimensions['CE'].width = 15
        ws.column_dimensions['CF'].width = 15
        ws.column_dimensions['CG'].width = 15
        ws.column_dimensions['CH'].width = 15
        ws.column_dimensions['CI'].width = 15
        ws.column_dimensions['CJ'].width = 15
        ws.column_dimensions['CK'].width = 15
        ws.column_dimensions['CL'].width = 15
        ws.column_dimensions['CM'].width = 15
        ws.column_dimensions['CN'].width = 15
        ws.column_dimensions['CO'].width = 15
        ws.column_dimensions['CP'].width = 15
        ws.column_dimensions['CQ'].width = 15
        ws.column_dimensions['CR'].width = 15
        ws.column_dimensions['CS'].width = 15
        ws.column_dimensions['CT'].width = 15
        ws.column_dimensions['CU'].width = 15
        ws.column_dimensions['CV'].width = 15
        ws.column_dimensions['CW'].width = 15
        ws.column_dimensions['CX'].width = 15
        ws.column_dimensions['CY'].width = 15
        ws.column_dimensions['CZ'].width = 15
        ws.column_dimensions['DA'].width = 15
        ws.column_dimensions['DB'].width = 22
        ws.column_dimensions['DC'].width = 22
        ws.column_dimensions['DD'].width = 22
        ws.column_dimensions['DE'].width = 22
        ws.column_dimensions['DF'].width = 22
        ws.column_dimensions['DG'].width = 25


        title_cell = ws['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de Registros Administrativos'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')
    
        ws.merge_cells('A1:DG1')

        ws.merge_cells('A2:C2')
        ws.merge_cells('D2:E2')
        ws.merge_cells('F2:H2')
        ws.merge_cells('I2:Q2')
        ws.merge_cells('R2:V2')
        ws.merge_cells('W2:W3')
        ws.merge_cells('X2:Y2')
        ws.merge_cells('Z2:AD2')
        ws.merge_cells('AE2:AR2')

        ws.merge_cells('AS2:AV2')
        ws.merge_cells('AW2:BC2')
        ws.merge_cells('BD2:BL2')
        ws.merge_cells('BM2:BU2')
        ws.merge_cells('BV2:BZ2')
        ws.merge_cells('CA2:CH2')
        ws.merge_cells('CI2:CS2')
        ws.merge_cells('CT2:CT3')
        ws.merge_cells('CU2:CU3')
        ws.merge_cells('CV2:DA2')
        ws.merge_cells('DB2:DB3') 

        ws.merge_cells('DC2:DC3') 
        ws.merge_cells('DD2:DD3') 
        ws.merge_cells('DE2:DE3') 
        ws.merge_cells('DF2:DF3') 
        ws.merge_cells('DG2:DG3')

        ## heads

        codigo_cell = ws['A2']
        codigo_cell.value = 'REGISTRO ADMINISTRATIVO'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['D2']
        codigo_cell.value = 'ENTIDAD'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['F2']
        codigo_cell.value = 'ÁREA TEMÁTICA / TEMA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['I2']
        codigo_cell.value = ' A. IDENTIFICACIÓN'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['R2']
        codigo_cell.value = 'Bajo cuál(es) de las siguientes normas, se soporta la creación del RA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['W2']
        codigo_cell.value = 'Describa la razón por la cuál se creó el RA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['X2']
        codigo_cell.value = 'Indique desde y hasta cuando se ha recolectado el RA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['Z2']
        codigo_cell.value = '¿El RA cuenta con alguno de los siguientes documentos metodológicos o funcionales?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AE2']
        codigo_cell.value = 'Indique si el RA utiliza conceptos estandarizados provenientes de'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AS2']
        codigo_cell.value = '¿El RA utiliza nomenclaturas y/o clasificaciones?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AW2']
        codigo_cell.value = '¿Cuál es el medio de obtención o recolección de los datos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BD2']
        codigo_cell.value = '¿Con qué frecuencia se recolectan los datos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['BM2']
        codigo_cell.value = 'Indique cuáles de las siguientes herramientas son utilizadas en el procesamiento de los datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BV2']
        codigo_cell.value = '¿Con cuáles herramientas cuenta para garantizar la seguridad de la información del RA?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CA2']
        codigo_cell.value = '¿La información recolectada es acopiada o almacenada en una base de datos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CI2']
        codigo_cell.value = '¿Cuál es la cobertura geográfica del RA?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CT2']
        codigo_cell.value = '¿La entidad hace uso de los datos del RA para generar información estadística (Agregados, indicadores, Resultados estadísticos, etc)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CU2']
        codigo_cell.value = '¿Usuarios externos a la entidad, tienen acceso a los datos del RA?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CV2']
        codigo_cell.value = '¿Cuál es la razón principal por la cual no se permite el acceso a los datos del RA?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DB2']
        codigo_cell.value = 'OBSERVACIONES'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DC2']
        codigo_cell.value = 'Registro administrativo activo (SI / NO)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DD2']
        codigo_cell.value = 'Usuario DANE'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DE2']
        codigo_cell.value = 'Responde a requerimientos ODS (SI / NO)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DF2']
        codigo_cell.value = 'Indicador ODS a que da respuesta'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DG2']
        codigo_cell.value = 'Estado del proceso del Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)



        #--------------------------------------------------------------------------------------------
        
        
        codigo_cell = ws['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['B3']
        codigo_cell.value = 'Nombre del Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['C3']
        codigo_cell.value = 'Objetivo del Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['D3']
        codigo_cell.value = 'Nombre de la Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center',  wrap_text=True)
       
        codigo_cell = ws['E3']
        codigo_cell.value = 'Código de la Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['F3']
        codigo_cell.value = 'Área Tematica'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = ws['G3']
        codigo_cell.value = 'Tema'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['H3']
        codigo_cell.value = 'Tema Compartido'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['I3']
        codigo_cell.value = 'Nombre de la Dependencia'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['J3']
        codigo_cell.value = 'Nombre del Director'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['K3']
        codigo_cell.value = 'Cargo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['L3']
        codigo_cell.value = 'Correo Electrónico del Director'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = ws['M3']
        codigo_cell.value = 'Teléfono del director'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['N3']
        codigo_cell.value = 'Nombre del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['O3']
        codigo_cell.value = 'Cargo del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['P3']
        codigo_cell.value = 'Correo Electrónico del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['Q3']
        codigo_cell.value = 'Teléfono del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['R3']
        codigo_cell.value = 'a. Constitución Política'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['S3']
        codigo_cell.value = 'b. Ley'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['T3']
        codigo_cell.value = 'c. Decreto (nacional, departamental, municipal)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['U3']
        codigo_cell.value = 'd. Otra (Resolución, ordenanza, acuerdo)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center',  wrap_text=True)

        codigo_cell = ws['V3']
        codigo_cell.value = 'e. Ninguna'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['X3']
        codigo_cell.value = 'Fecha de inicio de recolección de información'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['Y3']
        codigo_cell.value = 'Fecha de última recolección de información'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['Z3']
        codigo_cell.value = 'a) Ficha técnica'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AA3']
        codigo_cell.value = 'b) Manual de diligenciamiento'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AB3']
        codigo_cell.value = 'c) Diccionario de datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AC3']
        codigo_cell.value = 'd) Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AD3']
        codigo_cell.value = 'd) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AE3']
        codigo_cell.value = 'a). DANE'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AF3']
        codigo_cell.value = 'b). OCDE'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AG3']
        codigo_cell.value = 'c). ONU'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AH3']
        codigo_cell.value = 'd). EUROSTAT'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AI3']
        codigo_cell.value = 'e). Otro organismo Internacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AJ3']
        codigo_cell.value = 'e) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
  
        codigo_cell = ws['AK3']
        codigo_cell.value = 'f). Otra entidad de orden nacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AL3']
        codigo_cell.value = 'f) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AM3']
        codigo_cell.value = 'g). Leyes, decretos, etc.'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AN3']
        codigo_cell.value = 'g) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AO3']
        codigo_cell.value = 'h). Creación propia'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AP3']
        codigo_cell.value = 'i). Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AQ3']
        codigo_cell.value = 'i) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AR3']
        codigo_cell.value = 'j). No utiliza conceptos estandarizados'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AS3']
        codigo_cell.value = '¿utiliza nomenclaturas y/o clasificaciones?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AT3']
        codigo_cell.value = 'Si: Cuáles?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AU3']
        codigo_cell.value = 'Si: Otras'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


        codigo_cell = ws['AV3']
        codigo_cell.value = 'No: ¿Por qué?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


        codigo_cell = ws['AW3']
        codigo_cell.value = 'a) Formulario físico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AX3']
        codigo_cell.value = 'b) Formulario electrónico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


        codigo_cell = ws['AY3']
        codigo_cell.value = 'c) Dispositivo Móvil de Captura [DMC]'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


        codigo_cell = ws['AZ3']
        codigo_cell.value = 'd) Sistema de información'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BA3']
        codigo_cell.value = 'd) Cuáles?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BB3']
        codigo_cell.value = 'e) Imágenes satelitales'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BC3']
        codigo_cell.value = 'f) Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BD3']
        codigo_cell.value = 'a) Anual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BE3']
        codigo_cell.value = 'b) Semestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BF3']
        codigo_cell.value = 'c) Bimensual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BG3']
        codigo_cell.value = 'd) Trimestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BH3']
        codigo_cell.value = 'e) Mensual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BI3']
        codigo_cell.value = 'f) Diaria'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BJ3']
        codigo_cell.value = 'g) Por evento'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BK3']
        codigo_cell.value = 'h) Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BL3']
        codigo_cell.value = 'h) ¿cuál(es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BM3']
        codigo_cell.value = 'a) Excel'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BN3']
        codigo_cell.value = 'b) Access'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BO3']
        codigo_cell.value = 'c) R.'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BP3']
        codigo_cell.value = 'd) SAS.'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BQ3']
        codigo_cell.value = 'e) SPSS.'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BR3']
        codigo_cell.value = 'f) Oracle'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BS3']
        codigo_cell.value = 'g) Stata.'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BT3']
        codigo_cell.value = 'h) Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BU3']
        codigo_cell.value = 'h) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BV3']
        codigo_cell.value = 'a) Backups'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BW3']
        codigo_cell.value = 'b) Aislamiento del servidor'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BX3']
        codigo_cell.value = 'c) Utilización de perfiles'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BY3']
        codigo_cell.value = 'd) Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BZ3']
        codigo_cell.value = 'd) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CA3']
        codigo_cell.value = '¿La información es almacenada en una base de datos? '
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CB3']
        codigo_cell.value = 'a) Anual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CC3']
        codigo_cell.value = 'b) Semestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CD3']
        codigo_cell.value = 'c) Trimestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CE3']
        codigo_cell.value = 'd) Mensual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CF3']
        codigo_cell.value = 'e) Diaria'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CG3']
        codigo_cell.value = 'f) Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CH3']
        codigo_cell.value = 'f) ¿cuál(es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CI3']
        codigo_cell.value = 'a) Nacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CJ3']
        codigo_cell.value = 'b) Regional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CK3']
        codigo_cell.value = 'b) ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CL3']
        codigo_cell.value = 'c) Departamental'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CM3']
        codigo_cell.value = 'c) ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CN3']
        codigo_cell.value = 'd) Áreas metropolitanas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CO3']
        codigo_cell.value = 'd) ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CP3']
        codigo_cell.value = 'e) Municipal'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CQ3']
        codigo_cell.value = 'e) ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CR3']
        codigo_cell.value = 'f) Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CS3']
        codigo_cell.value = 'f) ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CV3']
        codigo_cell.value = 'a) Políticas de la entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CW3']
        codigo_cell.value = 'b) Disposición normativa'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CX3']
        codigo_cell.value = 'c) Acuerdos con el informante'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CY3']
        codigo_cell.value = 'd) Carácter confidencial'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CZ3']
        codigo_cell.value = 'e) Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DA3']
        codigo_cell.value = 'e) ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
 
        cont = 4
           
        for rraa in rraas:
            
            ws.cell(row = cont, column = 1).value = rraa.codigo_rraa
            ws.cell(row = cont, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 1).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 2).value = rraa.nombre_ra
            ws.cell(row = cont, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 2).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 3).value = rraa.objetivo_ra
            ws.cell(row = cont, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 3).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 4).value = str(rraa.entidad_pri)
            ws.cell(row = cont, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 4).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 5).value =  rraa.entidad_pri.codigo
            ws.cell(row = cont, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 5).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 6).value = str(rraa.area_temat)
            ws.cell(row = cont, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 6).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 7).value = str(rraa.tema)
            ws.cell(row = cont, column = 7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 7).font = Font(size = "8", name='Barlow')

            ## TEMA COMPARTIDO

            listaTemaCompartido = rraa.tema_compart.all()
            temaCompartido_array = []
            for indexTemaCompa, itemTemaCompa in enumerate(listaTemaCompartido):
                temaCompartido_array.append(str(itemTemaCompa))
                ws.cell(row = cont, column = 8).value = str(temaCompartido_array)
                ws.cell(row = cont, column = 8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 8).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 9).value = rraa.nom_dep
            ws.cell(row = cont, column = 9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = cont, column = 9).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 10).value = rraa.nom_dir
            ws.cell(row = cont, column = 10).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 10).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 11).value = rraa.car_dir
            ws.cell(row = cont, column = 11).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 11).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 12).value = rraa.cor_dir
            ws.cell(row = cont, column = 12).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 12).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 13).value = rraa.telef_dir
            ws.cell(row = cont, column = 13).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 13).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 14).value = rraa.nom_resp
            ws.cell(row = cont, column = 14).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 14).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 15).value = rraa.carg_resp
            ws.cell(row = cont, column = 15).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 15).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 16).value = rraa.cor_resp
            ws.cell(row = cont, column = 16).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 16).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 17).value = rraa.telef_resp
            ws.cell(row = cont, column = 17).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 17).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 18).value = rraa.pq_secreo
            ws.cell(row = cont, column = 18).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 18).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 23).value = rraa.pq_secreo
            ws.cell(row = cont, column = 23).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 23).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 24).value = rraa.fecha_ini_rec
            ws.cell(row = cont, column = 24).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 24).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 25).value = rraa.fecha_ult_rec
            ws.cell(row = cont, column = 25).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 25).font = Font(size = "8", name='Barlow')

            if str(rraa.clas_s_n) == "True":
                ws.cell(row = cont, column = 45).value = "Si"
                ws.cell(row = cont, column = 45).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 45).font = Font(size = "8", name='Barlow')
            
            elif str(rraa.clas_s_n) == "False":
                ws.cell(row = cont, column = 45).value = "No"
                ws.cell(row = cont, column = 45).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 45).font = Font(size = "8", name='Barlow')

            ##Clasificaciones
            listaClasificaciones = rraa.nomb_cla.all()
            clasificaciones_array = []

            for indexClasificaciones, itemClasificaciones in enumerate(listaClasificaciones):
                clasificaciones_array.append(str(itemClasificaciones))
                ws.cell(row = cont, column = 46).value = str(clasificaciones_array)
                ws.cell(row = cont, column = 46).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 46).font = Font(size = "8", name='Barlow')

            if str(rraa.almacen_bd_s_n) == "True":
                ws.cell(row = cont, column = 79).value = "Si"
                ws.cell(row = cont, column = 79).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 79).font = Font(size = "8", name='Barlow')
            
            elif str(rraa.almacen_bd_s_n) == "False":
                ws.cell(row = cont, column = 79).value = "No"
                ws.cell(row = cont, column = 79).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 79).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 98).value = str(rraa.uso_de_datos)
            ws.cell(row = cont, column = 98).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 98).font = Font(size = "8", name='Barlow')

            if str(rraa.user_exte_acceso) == "True":
                ws.cell(row = cont, column = 99).value = "Si"
                ws.cell(row = cont, column = 99).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 99).font = Font(size = "8", name='Barlow')
            
            elif str(rraa.user_exte_acceso) == "False":
                ws.cell(row = cont, column = 99).value = "No"
                ws.cell(row = cont, column = 99).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 99).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 106).value = rraa.observacion
            ws.cell(row = cont, column = 106).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 106).font = Font(size = "8", name='Barlow')

            if str(rraa.ra_activo) == "True":
                ws.cell(row = cont, column = 107).value = "Si"
                ws.cell(row = cont, column = 107).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 107).font = Font(size = "8", name='Barlow')
            
            elif str(rraa.ra_activo) == "False":
                ws.cell(row = cont, column = 107).value = "No"
                ws.cell(row = cont, column = 107).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 107).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 108).value = rraa.user_dane
            ws.cell(row = cont, column = 108).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 108).font = Font(size = "8", name='Barlow')

            if str(rraa.responde_ods) == "True":
                ws.cell(row = cont, column = 109).value = "Si"
                ws.cell(row = cont, column = 109).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 109).font = Font(size = "8", name='Barlow')
            
            elif str(rraa.responde_ods) == "False":
                ws.cell(row = cont, column = 109).value = "No"
                ws.cell(row = cont, column = 109).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 109).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 110).value = rraa.indicador_ods
            ws.cell(row = cont, column = 110).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 110).font = Font(size = "8", name='Barlow')

            
            ws.cell(row = cont, column = 111).value = str(rraa.sist_estado)
            ws.cell(row = cont, column = 111).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 111).font = Font(size = "8", name='Barlow')
            
            ##Si desea continuar despues de la columna 111 _____________
        
            ## opciones de preguntas
            listaNorma = rraa.norma_ra.all()
            listaDocumentoMetod = rraa.doc_met_ra.all()
            listaConcepEstand = rraa.con_est_ra.all()
            listaRecolec = rraa.recole_dato.all()
            listaFrecRecol = rraa.fre_rec_dato.all()
            listaHerramUtil = rraa.herr_u_pro.all()
            listaSeguridadIn = rraa.seg_inf.all()
            listaFrecAlm = rraa.frec_alm_bd.all()
            listaCoberGeo = rraa.cob_geograf.all()
            listaNoAccesoDat = rraa.no_hay_acceso.all()

            for index, item in enumerate(listaNorma):
                indice = index
                
                if  str(item) == 'Ninguna' and index == indice:
                    #print("item", item)
                    ws.cell(row = cont, column = 22).value = str(item)
                    ws.cell(row = cont, column = 22).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 22).font = Font(size = "8", name='Barlow')

            for indexDocMet, itemDocMet in enumerate(listaDocumentoMetod):
        
                if  str(itemDocMet) == 'Ficha técnica':
                   
                    ws.cell(row = cont, column = 26).value = str(itemDocMet)
                    ws.cell(row = cont, column = 26).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 26).font = Font(size = "8", name='Barlow')

                if  str(itemDocMet) == 'Manual de diligenciamiento':

                    ws.cell(row = cont, column = 27).value = str(itemDocMet)
                    ws.cell(row = cont, column = 27).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 27).font = Font(size = "8", name='Barlow')
                
                if str(itemDocMet) == 'Diccionario de datos':

                    ws.cell(row = cont, column = 28).value = str(itemDocMet)
                    ws.cell(row = cont, column = 28).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 28).font = Font(size = "8", name='Barlow')

                if str(itemDocMet) == 'Otro (s)':

                    ws.cell(row = cont, column = 29).value = str(itemDocMet)
                    ws.cell(row = cont, column = 29).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 29).font = Font(size = "8", name='Barlow')

            for itemConcepEstand, indexConcepEstand in enumerate(listaConcepEstand):
                
                if str(itemConcepEstand) == 'DANE':

                    ws.cell(row = cont, column = 31).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 31).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 31).font = Font(size = "8", name='Barlow')

                if str(itemConcepEstand) == 'OCDE':
                    ws.cell(row = cont, column = 32).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 32).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 32).font = Font(size = "8", name='Barlow')

                if str(itemConcepEstand) == 'ONU':
                    ws.cell(row = cont, column = 33).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 33).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 33).font = Font(size = "8", name='Barlow')

                if str(itemConcepEstand) == 'EUROSTAT':
                    ws.cell(row = cont, column = 34).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 34).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 34).font = Font(size = "8", name='Barlow')

                if str(itemConcepEstand) == 'Otro organismo Internacional':
                    ws.cell(row = cont, column = 35).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 35).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 35).font = Font(size = "8", name='Barlow')

                if str(itemConcepEstand) == 'Otra entidad de orden nacional':
                    ws.cell(row = cont, column = 37).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 37).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 37).font = Font(size = "8", name='Barlow')
                
                if str(itemConcepEstand) == 'Leyes, decretos, etc':
                    ws.cell(row = cont, column = 39).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 39).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 39).font = Font(size = "8", name='Barlow')   
                
                if str(itemConcepEstand) == 'Creación propia':
                    ws.cell(row = cont, column = 41).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 41).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 41).font = Font(size = "8", name='Barlow')
                
                if str(itemConcepEstand) == 'Otra (s)':
                    ws.cell(row = cont, column = 42).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 42).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 42).font = Font(size = "8", name='Barlow')
                
                if str(itemConcepEstand) == 'No utiliza conceptos estandarizados':
                    ws.cell(row = cont, column = 44).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 44).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 44).font = Font(size = "8", name='Barlow')

            
            for indexRecolec, itemRecolec in enumerate(listaRecolec):

                if str(itemRecolec) == 'Formulario físico':
                    ws.cell(row = cont, column = 49).value = str(itemRecolec)
                    ws.cell(row = cont, column = 49).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 49).font = Font(size = "8", name='Barlow')

                if str(itemRecolec) == 'Formulario electrónico':
                    ws.cell(row = cont, column = 50).value = str(itemRecolec)
                    ws.cell(row = cont, column = 50).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 50).font = Font(size = "8", name='Barlow')

                if str(itemRecolec) == 'Dispositivo Móvil de Captura [DMC]':
                    ws.cell(row = cont, column = 51).value = str(itemRecolec)
                    ws.cell(row = cont, column = 51).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 51).font = Font(size = "8", name='Barlow')

                if str(itemRecolec) == 'Sistema de información':
                    ws.cell(row = cont, column = 52).value = str(itemRecolec)
                    ws.cell(row = cont, column = 52).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 52).font = Font(size = "8", name='Barlow')

                if str(itemRecolec) == 'Imágenes satelitales':
                    ws.cell(row = cont, column = 54).value = str(itemRecolec)
                    ws.cell(row = cont, column = 54).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 54).font = Font(size = "8", name='Barlow')

            for indexFrecRecol, itemFrecRecol in enumerate(listaFrecRecol):

                if str(itemFrecRecol) == 'Anual':
                    ws.cell(row = cont, column = 56).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 56).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 56).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Semestral':
                    ws.cell(row = cont, column = 57).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 57).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 57).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Bimensual':
                    ws.cell(row = cont, column = 58).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 58).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 58).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Trimestral':
                    ws.cell(row = cont, column = 59).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 59).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 59).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Mensual':
                    ws.cell(row = cont, column = 60).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 60).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 60).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Diaria':

                    ws.cell(row = cont, column = 61).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 61).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 61).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Por evento':

                    ws.cell(row = cont, column = 62).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 62).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 62).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Otra (s)':

                    ws.cell(row = cont, column = 63).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 63).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 63).font = Font(size = "8", name='Barlow')

            
            for indexHerramUtil, itemHerramUtil in enumerate(listaHerramUtil):

                if str(itemHerramUtil) == 'Excel':
                    ws.cell(row = cont, column = 65).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 65).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 65).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'Access':
                    ws.cell(row = cont, column = 66).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 66).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 66).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'R':
                    ws.cell(row = cont, column = 67).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 67).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 67).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'SAS':
                    ws.cell(row = cont, column = 68).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 68).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 68).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'SPSS':
                    ws.cell(row = cont, column = 69).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 69).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 69).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'Oracle':
                    ws.cell(row = cont, column = 70).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 70).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 70).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'Stata':
                    ws.cell(row = cont, column = 71).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 71).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 71).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'Otra (s)':
                    ws.cell(row = cont, column = 72).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 72).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 72).font = Font(size = "8", name='Barlow')

            for indexSeguridadIn, itemSeguridadIn in enumerate(listaSeguridadIn):

                if str(itemSeguridadIn) == 'Backups':

                    ws.cell(row = cont, column = 74).value = str(itemSeguridadIn)
                    ws.cell(row = cont, column = 74).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 74).font = Font(size = "8", name='Barlow')

                if str(itemSeguridadIn) == 'Aislamiento del servidor':

                    ws.cell(row = cont, column = 75).value = str(itemSeguridadIn)
                    ws.cell(row = cont, column = 75).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 75).font = Font(size = "8", name='Barlow')

                if str(itemSeguridadIn) == 'Utilización de perfiles':

                    ws.cell(row = cont, column = 76).value = str(itemSeguridadIn)
                    ws.cell(row = cont, column = 76).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 76).font = Font(size = "8", name='Barlow')

                if str(itemSeguridadIn) == 'Otra (s)':

                    ws.cell(row = cont, column = 77).value = str(itemSeguridadIn)
                    ws.cell(row = cont, column = 77).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 77).font = Font(size = "8", name='Barlow')

            for indexFrecAlm, itemFrecAlm in enumerate(listaFrecAlm):

                if str(itemFrecAlm) == 'Anual':
                    ws.cell(row = cont, column = 80).value = str(itemFrecAlm)
                    ws.cell(row = cont, column = 80).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 80).font = Font(size = "8", name='Barlow')

                if str(itemFrecAlm) == 'Semestral':
                    ws.cell(row = cont, column = 81).value = str(itemFrecAlm)
                    ws.cell(row = cont, column = 81).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 81).font = Font(size = "8", name='Barlow')

                if str(itemFrecAlm) == 'Trimestral':
                    ws.cell(row = cont, column = 82).value = str(itemFrecAlm)
                    ws.cell(row = cont, column = 82).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 82).font = Font(size = "8", name='Barlow')

                if str(itemFrecAlm) == 'Mensual':
                    ws.cell(row = cont, column = 83).value = str(itemFrecAlm)
                    ws.cell(row = cont, column = 83).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 83).font = Font(size = "8", name='Barlow')

                if str(itemFrecAlm) == 'Diaria':
                    ws.cell(row = cont, column = 84).value = str(itemFrecAlm)
                    ws.cell(row = cont, column = 84).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 84).font = Font(size = "8", name='Barlow')

                if str(itemFrecAlm) == 'Otra (s)':
                    ws.cell(row = cont, column = 85).value = str(itemFrecAlm)
                    ws.cell(row = cont, column = 85).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 85).font = Font(size = "8", name='Barlow')

            for indexCoberGeo, itemCoberGeo in enumerate(listaCoberGeo):

                if str(itemCoberGeo) == 'Nacional':
                    ws.cell(row = cont, column = 87).value = str(itemCoberGeo)
                    ws.cell(row = cont, column = 87).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 87).font = Font(size = "8", name='Barlow')

                if str(itemCoberGeo) == 'Regional':

                    ws.cell(row = cont, column = 88).value = str(itemCoberGeo)
                    ws.cell(row = cont, column = 88).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 88).font = Font(size = "8", name='Barlow')

                if str(itemCoberGeo) == 'Departamental':

                    ws.cell(row = cont, column = 90).value = str(itemCoberGeo)
                    ws.cell(row = cont, column = 90).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 90).font = Font(size = "8", name='Barlow')

                if str(itemCoberGeo) == 'Áreas metropolitanas':

                    ws.cell(row = cont, column = 92).value = str(itemCoberGeo)
                    ws.cell(row = cont, column = 92).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 92).font = Font(size = "8", name='Barlow')

                if str(itemCoberGeo) == 'Municipal':

                    ws.cell(row = cont, column = 94).value = str(itemCoberGeo)
                    ws.cell(row = cont, column = 94).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 94).font = Font(size = "8", name='Barlow')

                if str(itemCoberGeo) == 'Otro (s)':

                    ws.cell(row = cont, column = 96).value = str(itemCoberGeo)
                    ws.cell(row = cont, column = 96).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 96).font = Font(size = "8", name='Barlow')

            for indexNoAccesoDat, itemNoAccesoDat in enumerate(listaNoAccesoDat):

                if str(itemNoAccesoDat) == 'Políticas de la entidad':

                    ws.cell(row = cont, column = 100).value = str(itemNoAccesoDat)
                    ws.cell(row = cont, column = 100).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 100).font = Font(size = "8", name='Barlow')

                if str(itemNoAccesoDat) == 'Disposición normativa':

                    ws.cell(row = cont, column = 101).value = str(itemNoAccesoDat)
                    ws.cell(row = cont, column = 101).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 101).font = Font(size = "8", name='Barlow')

                if str(itemNoAccesoDat) == 'Acuerdos con el informante':

                    ws.cell(row = cont, column = 102).value = str(itemNoAccesoDat)
                    ws.cell(row = cont, column = 102).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 102).font = Font(size = "8", name='Barlow')

                if str(itemNoAccesoDat) == 'Carácter confidencial':

                    ws.cell(row = cont, column = 103).value = str(itemNoAccesoDat)
                    ws.cell(row = cont, column = 103).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 103).font = Font(size = "8", name='Barlow')

                if str(itemNoAccesoDat) == 'Otra (s)':

                    ws.cell(row = cont, column = 104).value = str(itemNoAccesoDat)
                    ws.cell(row = cont, column = 104).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 104).font = Font(size = "8", name='Barlow')

          
            cont+=1

        ## campos de texto   

        #print("---------------------->", normaTextList)
        for indexNorma, itemNorma in enumerate(normaTextList):
           
            indexNorma =  indexNorma + 4
            ws.cell(row = indexNorma, column = 18).value = str(itemNorma.cp_ra)
            ws.cell(row = indexNorma, column = 18).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 18).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexNorma, column = 19).value = str(itemNorma.ley_ra)
            ws.cell(row = indexNorma, column = 19).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 19).font = Font(size = "8", name='Barlow')
                    
            ws.cell(row = indexNorma, column = 20).value = str(itemNorma.decreto_ra)
            ws.cell(row = indexNorma, column = 20).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 20).font = Font(size = "8", name='Barlow')
                    
            ws.cell(row = indexNorma, column = 21).value = str(itemNorma.otra_ra)
            ws.cell(row = indexNorma, column = 21).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 21).font = Font(size = "8", name='Barlow')


        for indexDocumentoMetod, itemDocumentoMetod in enumerate(documentoMetodTextList):
           
            indexDocumentoMetod =  indexDocumentoMetod + 4
            ws.cell(row = indexDocumentoMetod, column = 30).value = str(itemDocumentoMetod.otra_doc_cual)
            ws.cell(row = indexDocumentoMetod, column = 30).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexDocumentoMetod, column = 30).font = Font(size = "8", name='Barlow')


        for indexConEst, itemConEst in enumerate(conceptosEstandarizadosTextList):
           
            indexConEst =  indexConEst + 4

            ws.cell(row = indexConEst, column = 36).value = str(itemConEst.org_in_cual)
            ws.cell(row = indexConEst, column = 36).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConEst, column = 36).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexConEst, column = 38).value = str(itemConEst.ent_ordnac_cual)
            ws.cell(row = indexConEst, column = 38).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConEst, column = 38).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexConEst, column = 40).value = str(itemConEst.leye_dec_cual)
            ws.cell(row = indexConEst, column = 40).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConEst, column = 40).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexConEst, column = 43).value = str(itemConEst.otra_ce_cual)
            ws.cell(row = indexConEst, column = 43).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConEst, column = 43).font = Font(size = "8", name='Barlow')
        

        for indexClasifi, itemClasifi in enumerate(clasificacionTextList):
           
            indexClasifi =  indexClasifi + 4

            ws.cell(row = indexClasifi, column = 47).value = str(itemClasifi.otra_cual_clas)
            ws.cell(row = indexClasifi, column = 47).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexClasifi, column = 47).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexClasifi, column = 48).value = str(itemClasifi.no_pq)
            ws.cell(row = indexClasifi, column = 48).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexClasifi, column = 48).font = Font(size = "8", name='Barlow')

        
        for indexRecoleccionDatos, itemRecoleccionDatos in enumerate(recoleccionDatosTextList):

            indexRecoleccionDatos = indexRecoleccionDatos + 4

            ws.cell(row = indexRecoleccionDatos, column = 53).value = str(itemRecoleccionDatos.sistema_inf_cual)
            ws.cell(row = indexRecoleccionDatos, column = 53).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexRecoleccionDatos, column = 53).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexRecoleccionDatos, column = 55).value = str(itemRecoleccionDatos.otro_c)
            ws.cell(row = indexRecoleccionDatos, column = 55).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexRecoleccionDatos, column = 55).font = Font(size = "8", name='Barlow')

        for indexFrecuenciaRecoleccionDato, itemFrecuenciaRecoleccionDato in enumerate(FrecuenciaRecoleccionDatoTextList):
            
            indexFrecuenciaRecoleccionDato = indexFrecuenciaRecoleccionDato + 4

            ws.cell(row = indexFrecuenciaRecoleccionDato, column = 64).value = str(itemFrecuenciaRecoleccionDato.otra_cual_fre)
            ws.cell(row = indexFrecuenciaRecoleccionDato, column = 64).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexFrecuenciaRecoleccionDato, column = 64).font = Font(size = "8", name='Barlow')

         
        for indexherramientasUtilPro, itemherramientasUtilPro in enumerate(herramientasUtilProcesaTextList):
            
            indexherramientasUtilPro = indexherramientasUtilPro + 4

            ws.cell(row = indexherramientasUtilPro, column = 73).value = str(itemherramientasUtilPro.otra_herram)
            ws.cell(row = indexherramientasUtilPro, column = 73).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexherramientasUtilPro, column = 73).font = Font(size = "8", name='Barlow')

        
        for indexSeguridadInfo, itemSeguridadInfo in enumerate(seguridadInformTextList):

            indexSeguridadInfo = indexSeguridadInfo + 4

            ws.cell(row = indexSeguridadInfo, column = 78).value = str(itemSeguridadInfo.otra_cual_s)
            ws.cell(row = indexSeguridadInfo, column = 78).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexSeguridadInfo, column = 78).font = Font(size = "8", name='Barlow')

        for indexfrecuenciaAlmacen, itemfrecuenciaAlmacen in enumerate(frecuenciaAlmacenamientobdTextList):

            indexfrecuenciaAlmacen = indexfrecuenciaAlmacen + 4

            ws.cell(row = indexfrecuenciaAlmacen, column = 86).value = str(itemfrecuenciaAlmacen.otra_alm_bd)
            ws.cell(row = indexfrecuenciaAlmacen, column = 86).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexfrecuenciaAlmacen, column = 86).font = Font(size = "8", name='Barlow')

        for indexCoberturaGeog, itemCoberturaGeog in enumerate(coberturaGeograficaTextList):
            
            indexCoberturaGeog = indexCoberturaGeog + 4

            ws.cell(row = indexCoberturaGeog, column = 89).value = str(itemCoberturaGeog.cual_regio)
            ws.cell(row = indexCoberturaGeog, column = 89).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 89).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 91).value = str(itemCoberturaGeog.cual_depa)
            ws.cell(row = indexCoberturaGeog, column = 91).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 91).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 93).value = str(itemCoberturaGeog.cual_are_metrop)
            ws.cell(row = indexCoberturaGeog, column = 93).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 93).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 95).value = str(itemCoberturaGeog.cual_munic)
            ws.cell(row = indexCoberturaGeog, column = 95).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 95).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 97).value = str(itemCoberturaGeog.cual_otro)
            ws.cell(row = indexCoberturaGeog, column = 97).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 97).font = Font(size = "8", name='Barlow')

        for indexNoAccesoDatos, itemNoAccesoDatos in enumerate(noAccesoDatosTextList):

            indexNoAccesoDatos = indexNoAccesoDatos + 4

            ws.cell(row = indexNoAccesoDatos, column = 105).value = str(itemNoAccesoDatos.otra_no_acceso)
            ws.cell(row = indexNoAccesoDatos, column = 105).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNoAccesoDatos, column = 105).font = Font(size = "8", name='Barlow')

        for row in range(4, ws.max_row + 1):
            ws.row_dimensions[row].height = 90


       ####################### HOJA 2 ######################## 

        sheet2.merge_cells('A1:C1')
        sheet2.merge_cells('A2:C2')

        def set_border(sheet2, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet2[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet2,'A1:C1')
    
        #row dimensions
        sheet2.row_dimensions[1].height = 55
        sheet2.row_dimensions[2].height = 40

        # column width
        sheet2.column_dimensions['A'].width = 20
        sheet2.column_dimensions['B'].width = 20
        sheet2.column_dimensions['C'].width = 20
        
        title_cell = sheet2['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de los Registros Administrativos'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        title_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
	
        codigo_cell = sheet2['A2']
        codigo_cell.value = 'Lista de variables de los Registros Administrativos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)


        codigo_cell = sheet2['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = sheet2['B3']
        codigo_cell.value = 'Nombre Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet2['C3']
        codigo_cell.value = 'Variables que maneja el Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        contsheet2 = 4

        arrayraId = []
        for indexVariab, itemVariab in enumerate(variableRecolectadaTextList):
            indexVariab =  itemVariab.rraa_id + 3
            arrayraId.append(itemVariab.rraa_id)
            for radm in rraas:
                if str(itemVariab.rraa) == str(radm.nombre_ra):
                    sheet2.cell(row = contsheet2, column = 1).value = radm.codigo_rraa
                    sheet2.cell(row = contsheet2, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = contsheet2, column = 1).font = Font(size = "8", name='Barlow')

                    sheet2.cell(row = contsheet2, column = 2).value = str(itemVariab.rraa)
                    sheet2.cell(row = contsheet2, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = contsheet2, column = 2).font = Font(size = "8", name='Barlow')

                    sheet2.cell(row = contsheet2, column = 3).value = str(itemVariab.variableRec)
                    sheet2.cell(row = contsheet2, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = contsheet2, column = 3).font = Font(size = "8", name='Barlow')
                    contsheet2+=1
           
        for row in range(4, sheet2.max_row + 1):  ## definir tamaño de rows
            sheet2.row_dimensions[row].height = 90

        ##################### HOJA 3  #################################################

        sheet3.merge_cells('A1:C1')
        sheet3.merge_cells('A2:C2')

        def set_border(sheet3, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet3[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet3,'A1:C1')
    
        #row dimensions
        sheet3.row_dimensions[1].height = 55
        sheet3.row_dimensions[2].height = 40

        # column width
        sheet3.column_dimensions['A'].width = 20
        sheet3.column_dimensions['B'].width = 20
        sheet3.column_dimensions['C'].width = 20
        
        title_cell = sheet3['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de los Registros Administrativos'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        title_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
	
        codigo_cell = sheet3['A2']
        codigo_cell.value = 'La entidad hace uso de los datos del RA para generar información estadística (Agregados, indicadores, Resultados estadísticos, etc)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet3['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = sheet3['B3']
        codigo_cell.value = 'Nombre Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet3['C3']
        codigo_cell.value = 'Lista de resultados agregados o indicadores'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        contsheet3 = 4
        arrayraId = []
        for indexIndicador, itemIndicador in enumerate(indicadorResultadoAgregadoTextList):
            indexIndicador =  itemIndicador.rraa_id + 3
            arrayraId.append(itemIndicador.rraa_id)
            
            sheet3.cell(row = contsheet3, column = 1).value = itemIndicador.rraa.codigo_rraa
            sheet3.cell(row = contsheet3, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet3.cell(row = contsheet3, column = 1).font = Font(size = "8", name='Barlow')

            sheet3.cell(row = contsheet3, column = 2).value = str(itemIndicador.rraa)
            sheet3.cell(row = contsheet3, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet3.cell(row = contsheet3, column = 2).font = Font(size = "8", name='Barlow')

            sheet3.cell(row = contsheet3, column = 3).value = str(itemIndicador.ind_res_agre)
            sheet3.cell(row = contsheet3, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet3.cell(row = contsheet3, column = 3).font = Font(size = "8", name='Barlow')
            contsheet3+=1
           
        for row in range(4, sheet3.max_row + 1):  ## definir tamaño de rows
            sheet3.row_dimensions[row].height = 90

        ####################### Hoja 4 pregunta 18 #############################################


        def set_border(sheet4, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet4[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet4,'A1:AG'+str(rraas.count()+3))

		## size rows

        sheet4.row_dimensions[1].height = 55
        sheet4.row_dimensions[2].height = 40
        sheet4.row_dimensions[3].height = 50

        ## size column
        sheet4.column_dimensions['A'].width = 18
        sheet4.column_dimensions['B'].width = 18
        sheet4.column_dimensions['C'].width = 15
        sheet4.column_dimensions['D'].width = 15
        sheet4.column_dimensions['E'].width = 15
        sheet4.column_dimensions['F'].width = 15
        sheet4.column_dimensions['G'].width = 15
        sheet4.column_dimensions['H'].width = 15
        sheet4.column_dimensions['I'].width = 15
        sheet4.column_dimensions['J'].width = 15
        sheet4.column_dimensions['K'].width = 15
        sheet4.column_dimensions['L'].width = 15
        sheet4.column_dimensions['M'].width = 15
        sheet4.column_dimensions['N'].width = 15
        sheet4.column_dimensions['O'].width = 15
        sheet4.column_dimensions['P'].width = 15
        sheet4.column_dimensions['Q'].width = 15
        sheet4.column_dimensions['R'].width = 15
        sheet4.column_dimensions['S'].width = 15
        sheet4.column_dimensions['T'].width = 15
        sheet4.column_dimensions['U'].width = 15
        sheet4.column_dimensions['V'].width = 15
        sheet4.column_dimensions['W'].width = 15
        sheet4.column_dimensions['X'].width = 15
        sheet4.column_dimensions['Y'].width = 15
        sheet4.column_dimensions['Z'].width = 15
        sheet4.column_dimensions['AA'].width = 15
        sheet4.column_dimensions['AB'].width = 15
        sheet4.column_dimensions['AC'].width = 15
        sheet4.column_dimensions['AD'].width = 15
        sheet4.column_dimensions['AE'].width = 15
        sheet4.column_dimensions['AF'].width = 15
        sheet4.column_dimensions['AG'].width = 15
        sheet4.column_dimensions['AH'].width = 15
        

        title_cell = sheet4['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de Registros Administrativos'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')
    
        sheet4.merge_cells('A1:AG1')
        sheet4.merge_cells('A2:B2')
        sheet4.merge_cells('C2:AG2')

        ## heads

        codigo_cell = sheet4['A2']
        codigo_cell.value = 'REGISTRO ADMINISTRATIVO'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet4['C2']
        codigo_cell.value = 'Relacione las entidades que tienen acceso a los datos del RA:'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        
        codigo_cell = sheet4['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet4['B3']
        codigo_cell.value = 'Nombre del Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        
        contsheet4 = 4
        for rraa in rraas: 
            sheet4.cell(row = contsheet4, column = 1).value = rraa.codigo_rraa
            sheet4.cell(row = contsheet4, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet4.cell(row = contsheet4, column = 1).font = Font(size = "8", name='Barlow')

            sheet4.cell(row = contsheet4, column = 2).value = rraa.nombre_ra
            sheet4.cell(row = contsheet4, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet4.cell(row = contsheet4, column = 2).font = Font(size = "8", name='Barlow')

            contsheet4+= 1
    
        entidadraid = []
        for indexEntidAcc, itemEntidAcc in enumerate(entidadesAccesoRRAATextList): 
            indexEntidAcc = itemEntidAcc.rraa_id + 3
                    
            entidadraid.append(itemEntidAcc.rraa_id)
           
            count_id = entidadraid.count(itemEntidAcc.rraa_id)
            c=0 #inicializamos el contador  
            n=5*count_id
            for i in range(1,n+1):  
                if i%5 == 0:  
                    i = i - 2  ##posición de celda
                    c+=1

            for indexra, itemra in enumerate(rraas):
                if itemEntidAcc.rraa_id == itemra.pk:
                    posRow = indexra + 4
                     
                    sheet4.cell(row = 3, column = i ).value = "Nombre de Entidad"
                    sheet4.cell(row = 3, column = i ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = 3, column = i ).font = Font(bold=True) 

                    sheet4.cell(row = posRow, column = i ).value = str(itemEntidAcc.nomb_entidad_acc) 
                    sheet4.cell(row = posRow, column = i ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = posRow, column = i ).font = Font(size = "8", name='Barlow')

                    sheet4.cell(row = 3, column = i+4).value = "¿cuál?"
                    sheet4.cell(row = 3, column = i+4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = 3, column = i+4).font = Font(bold=True)

                    sheet4.cell(row = posRow, column = i+4).value = str(itemEntidAcc.otro_cual) 
                    sheet4.cell(row = posRow, column = i+4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = posRow, column = i+4).font = Font(size = "8", name='Barlow')


            
            listaFinalidad = list(itemEntidAcc.opcion_pr.all())  ##iterar para traer las fases seleccionadas

            for indexFinalida, itemFinalidad in enumerate(listaFinalidad):

                if str(itemFinalidad) == "Estadístico":
                    
                    countFin=0 #inicializamos el contador  
                    finSelected=5*count_id 
                    for posFin in range(1,n+1):  
                        if posFin%5 == 0:
                            posFin = posFin - 1 ##posicion celda
                            countFin+=1

                    sheet4.cell(row = 3, column = posFin ).value = "Estadístico"
                    sheet4.cell(row = 3, column = posFin ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = 3, column = posFin ).font = Font(bold=True)

                    sheet4.cell(row = posRow, column = posFin).value = str(itemFinalidad)
                    sheet4.cell(row = posRow, column = posFin).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = posRow, column = posFin).font = Font(size = "8", name='Barlow')
                                
                if str(itemFinalidad) == "No sabe":

                    countFin=0 #inicializamos el contador  
                    finSelected=5*count_id 
                    for posFin in range(1,n+1):  
                        if posFin%5 == 0:  
                            posFin = posFin ##posición celda 
                            countFin+=1

                    sheet4.cell(row = 3, column = posFin ).value = "No sabe"
                    sheet4.cell(row = 3, column = posFin ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = 3, column = posFin ).font = Font(bold=True)

                    sheet4.cell(row = posRow, column = posFin).value = str(itemFinalidad)
                    sheet4.cell(row = posRow, column = posFin).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = posRow, column = posFin).font = Font(size = "8", name='Barlow')

                if str(itemFinalidad) == "Otro, ¿cuál?":

                    countFin=0 #inicializamos el contador  
                    finSelected=5*count_id 
                    for posFin in range(1,n+1):  
                        if posFin%5 == 0:  
                            posFin = posFin + 1
                            countFin+=1

                    sheet4.cell(row = 3, column = posFin ).value = "Otro"
                    sheet4.cell(row = 3, column = posFin ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = 3, column = posFin ).font = Font(bold=True)

                    sheet4.cell(row = posRow, column = posFin).value = str(itemFinalidad)
                    sheet4.cell(row = posRow, column = posFin).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = posRow, column = posFin).font = Font(size = "8", name='Barlow')
                


        for row in range(4, sheet4.max_row + 1):
            sheet4.row_dimensions[row].height = 90


        #################################### Hoja 5 ###########################

        def set_border(sheet5, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet5[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet5,'A1:AZ'+str(rraas.count()+3))

		## size rows

        sheet5.row_dimensions[1].height = 55
        sheet5.row_dimensions[2].height = 40
        sheet5.row_dimensions[3].height = 50

        ## size column
        sheet5.column_dimensions['A'].width = 18
        sheet5.column_dimensions['B'].width = 18
        sheet5.column_dimensions['C'].width = 24
        sheet5.column_dimensions['D'].width = 24
        sheet5.column_dimensions['E'].width = 24
        sheet5.column_dimensions['F'].width = 24
        sheet5.column_dimensions['G'].width = 24
        sheet5.column_dimensions['H'].width = 24
        sheet5.column_dimensions['I'].width = 24
        sheet5.column_dimensions['J'].width = 24
        sheet5.column_dimensions['K'].width = 24
        sheet5.column_dimensions['L'].width = 24
        sheet5.column_dimensions['M'].width = 24
        sheet5.column_dimensions['N'].width = 24
        sheet5.column_dimensions['O'].width = 24
        sheet5.column_dimensions['P'].width = 24
        sheet5.column_dimensions['Q'].width = 24
        sheet5.column_dimensions['R'].width = 24
        sheet5.column_dimensions['S'].width = 24
        sheet5.column_dimensions['T'].width = 24
        sheet5.column_dimensions['U'].width = 24
        sheet5.column_dimensions['V'].width = 24
        sheet5.column_dimensions['W'].width = 24
        sheet5.column_dimensions['X'].width = 24
        sheet5.column_dimensions['Y'].width = 24
        sheet5.column_dimensions['Z'].width = 24
        sheet5.column_dimensions['AA'].width = 24
        sheet5.column_dimensions['AB'].width = 24
        sheet5.column_dimensions['AC'].width = 24
        sheet5.column_dimensions['AD'].width = 24
        sheet5.column_dimensions['AE'].width = 24
        sheet5.column_dimensions['AF'].width = 24
        sheet5.column_dimensions['AG'].width = 24
        sheet5.column_dimensions['AH'].width = 24
        
        sheet5.merge_cells('A1:AZ1')
        sheet5.merge_cells('A2:F2')
        sheet5.merge_cells('G2:AZ2')

        ## heads
        
        title_cell = sheet5['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de Registros Administrativos'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet5['A2']
        codigo_cell.value = 'IDENTIFICACIÓN'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet5['G2']
        codigo_cell.value = 'Fortalecimiento RRAA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['A3']
        resultados_cell.value = 'Área Temática'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['B3']
        resultados_cell.value = 'Tema'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['C3']
        resultados_cell.value = 'Código Entidad'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['D3']
        resultados_cell.value = 'Nombre Entidad'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['E3']
        resultados_cell.value = 'Código RRAA'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['F3']
        resultados_cell.value = 'Nombre del registro administrativo'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        
        contsheet5 = 4
        for rraa in rraas: 
            
            sheet5.cell(row = contsheet5, column = 1).value = str(rraa.area_temat)
            sheet5.cell(row = contsheet5, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 1).font = Font(size = "8", name='Barlow') 
            
            sheet5.cell(row = contsheet5, column = 2).value = str(rraa.tema)
            sheet5.cell(row = contsheet5, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 2).font = Font(size = "8", name='Barlow') 
            
            sheet5.cell(row = contsheet5, column = 3).value = rraa.entidad_pri.codigo
            sheet5.cell(row = contsheet5, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 3).font = Font(size = "8", name='Barlow')

            sheet5.cell(row = contsheet5, column = 4).value = str(rraa.entidad_pri)
            sheet5.cell(row = contsheet5, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 4).font = Font(size = "8", name='Barlow')

            sheet5.cell(row = contsheet5, column = 5).value = rraa.codigo_rraa
            sheet5.cell(row = contsheet5, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 5).font = Font(size = "8", name='Barlow')

            sheet5.cell(row = contsheet5, column = 6).value = rraa.nombre_ra
            sheet5.cell(row = contsheet5, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 6).font = Font(size = "8", name='Barlow')
            
            contsheet5+= 1

        idraFort = []
        for indexFort, itemFort in enumerate(fortalecimientoTextList): 
            indexFort = itemFort.post_ra_id + 3
            idraFort.append(itemFort.post_ra_id)

            ### 1 Diagnóstico del RA
            
            count_diagnostico = idraFort.count(itemFort.post_ra_id)
            
            contrField1=0 #inicializamos el contador  
            pos_diagno_rraa = 9*count_diagnostico
            for incField1 in range(1,pos_diagno_rraa+1):  
                if incField1%9 == 0:  
                    incField1 = incField1 - 2
                    contrField1+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField1).value = "Diagnóstico del RRAA"
                    sheet5.cell(row = 3, column = incField1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField1).font = Font(bold=True) 

                    if str(itemFort.diagnostico_ra) == "True":

                        sheet5.cell(row = posiRow, column = incField1).value = "Si"
                        sheet5.cell(row = posiRow, column = incField1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField1).font = Font(size = "8", name='Barlow')

                    elif str(itemFort.diagnostico_ra) == "False":

                        sheet5.cell(row = posiRow, column = incField1).value = "No"
                        sheet5.cell(row = posiRow, column = incField1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField1).font = Font(size = "8", name='Barlow')

  
            ### 2 Año de diagnóstico de rraa

            count_id_ra = idraFort.count(itemFort.post_ra_id)
            contadorFort=0 #inicializamos el contador  
            posFields = 9*count_id_ra
            
            for incFort in range(1,posFields+1):  
                if incFort%9 == 0:
                    incFort = incFort - 1
                    contadorFort+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incFort).value = "Año de diagnóstico de RRAA"
                    sheet5.cell(row = 3, column = incFort).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incFort).font = Font(bold=True) 

                    if itemFort.year_diagnostico == None:

                        sheet5.cell(row = posiRow, column = incFort).value = ""
                        sheet5.cell(row = posiRow, column = incFort).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incFort).font = Font(size = "8", name='Barlow')
                    else:
                        sheet5.cell(row = posiRow, column = incFort).value = itemFort.year_diagnostico.strftime('%Y')
                        sheet5.cell(row = posiRow, column = incFort).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incFort).font = Font(size = "8", name='Barlow')
            
            ### 3 Módulo o sección diagnosticado:

            contadorField3=0 #inicializamos el contador  
            posFields3 = 9*count_id_ra
            for incField3 in range(1,posFields3+1):  
                if incField3%9 == 0:  
                    incField3 = incField3
                    contadorField3+=1
                
            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField3).value = "Módulo o sección diagnosticado:"
                    sheet5.cell(row = 3, column = incField3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField3).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField3).value = str(itemFort.mod_sec_diagn)
                    sheet5.cell(row = posiRow, column = incField3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField3).font = Font(size = "8", name='Barlow')

            ### 4 Plan de fortalecimiento aprobado por la entidad:

            contadorField4=0 #inicializamos el contador  
            posFields4 = 9*count_id_ra
            for incField4 in range(1,posFields4+1):  
                if incField4%9 == 0:
                    incField4 = incField4 + 1
                    contadorField4+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField4).value = "Plan de fortalecimiento aprobado por la entidad:"
                    sheet5.cell(row = 3, column = incField4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField4).font = Font(bold=True)

                    if str(itemFort.plan_fort_aprob) == "True":
                        sheet5.cell(row = posiRow, column = incField4).value = "Si"
                        sheet5.cell(row = posiRow, column = incField4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField4).font = Font(size = "8", name='Barlow')

                    elif str(itemFort.plan_fort_aprob) == "False":
                        sheet5.cell(row = posiRow, column = incField4).value = "No"
                        sheet5.cell(row = posiRow, column = incField4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField4).font = Font(size = "8", name='Barlow')
            
        ### 5 Fecha de aprobación del Plan de Fortalecimiento

            contadorField5=0 #inicializamos el contador  
            posFields5 = 9*count_id_ra
            for incField5 in range(1,posFields5+1):  
                if incField5%9 == 0:  
                    incField5 = incField5 + 2
                    contadorField5+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField5).value = "Fecha de aprobación del Plan de Fortalecimiento"
                    sheet5.cell(row = 3, column = incField5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField5).font = Font(bold=True)

                    if itemFort.fecha_aprobacion == None:

                        sheet5.cell(row = posiRow, column = incField5).value = ""
                        sheet5.cell(row = posiRow, column = incField5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField5).font = Font(size = "8", name='Barlow')
                    else:
                        sheet5.cell(row = posiRow, column = incField5).value = itemFort.fecha_aprobacion
                        sheet5.cell(row = posiRow, column = incField5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField5).font = Font(size = "8", name='Barlow')
                    
        ### 6 Seguimiento a la implementación del Plan de fortalecimiento:

            contadorField6=0 #inicializamos el contador  
            posFields6 = 9*count_id_ra
            for incField6 in range(1,posFields6+1):  
                if incField6%9 == 0:  
                    incField6 = incField6 + 3
                    contadorField6+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField6).value = "Seguimiento a la implementación del Plan de fortalecimiento"
                    sheet5.cell(row = 3, column = incField6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField6).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField6).value = str(itemFort.seg_imple_plan)
                    sheet5.cell(row = posiRow, column = incField6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField6).font = Font(size = "8", name='Barlow')

        ### 7 Fecha de inicio de la implementación del Plan de Fortaleciemiento

            contadorField7=0 #inicializamos el contador  
            posFields7 = 9*count_id_ra
            for incField7 in range(1,posFields7+1):  
                if incField7%9 == 0:  
                    incField7 = incField7 + 4
                    contadorField7+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField7).value = "Fecha de inicio de la implementación del Plan de Fortaleciemiento"
                    sheet5.cell(row = 3, column = incField7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField7).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField7).value = str(itemFort.fecha_inicio_plan)
                    sheet5.cell(row = posiRow, column = incField7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField7).font = Font(size = "8", name='Barlow')

         ### 8 Fecha de último seguimiento a la implementación del Plan de Fortalecimiento   
            
            contadorField8=0 #inicializamos el contador  
            posFields8 = 9*count_id_ra
            for incField8 in range(1,posFields8+1):  
                if incField8%9 == 0:  
                    incField8 = incField8 + 5
                    contadorField8+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField8).value = "Fecha de último seguimiento a la implementación del Plan de Fortalecimiento "
                    sheet5.cell(row = 3, column = incField8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField8).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField8).value = str(itemFort.fecha_ultimo_seguimiento)
                    sheet5.cell(row = posiRow, column = incField8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField8).font = Font(size = "8", name='Barlow')

        ###  9 Fecha de finalización de la implementación del Plan de fortalecimiento
            
            contadorField9=0 #inicializamos el contador  
            posFields9 = 9*count_id_ra
            for incField9 in range(1,posFields9+1):  
                if incField9%9 == 0:  
                    incField9 = incField9 + 6
                    contadorField9+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField9).value = "Fecha de finalización de la implementación del Plan de fortalecimiento"
                    sheet5.cell(row = 3, column = incField9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField9).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField9).value = str(itemFort.fecha_finalizacion)
                    sheet5.cell(row = posiRow, column = incField9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField9).font = Font(size = "8", name='Barlow')

       
        for row in range(4, sheet5.max_row + 1):
            sheet5.row_dimensions[row].height = 90

        ############# no mostrar información si no esta autenticado #######################
        if user.is_authenticated == True and user.profile.role.id != 2:
        ############# Hoja 6 Criticas #######################

            sheet6.merge_cells('A1:F1')
            sheet6.merge_cells('A2:F2')

            def set_border(sheet6, cell_range):
                border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                rows = sheet6[cell_range]
                for row in rows:
                    for cell in row:
                        cell.border = border

            set_border(sheet6,'A1:F1')
            
            #row dimensions
            sheet6.row_dimensions[1].height = 55
            sheet6.row_dimensions[2].height = 40
            sheet6.row_dimensions[3].height = 40
            sheet6.row_dimensions[4].height = 40
            sheet6.row_dimensions[5].height = 40
            sheet6.row_dimensions[6].height = 40
            

            # column width
            sheet6.column_dimensions['A'].width = 20
            sheet6.column_dimensions['B'].width = 20
            sheet6.column_dimensions['C'].width = 20
            sheet6.column_dimensions['D'].width = 20
            sheet6.column_dimensions['E'].width = 20
            sheet6.column_dimensions['F'].width = 20
        

            title_cell = sheet6['A1']
            title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
            title_cell.value = 'Directorio de los Registros Administrativos'
            title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
            title_cell.alignment =  Alignment(horizontal='center', vertical='center')

            codigo_cell = sheet6['A2']
            codigo_cell.value = 'Críticas'
            codigo_cell.font = Font(bold=True)
            codigo_cell.alignment = Alignment(horizontal='center')
            codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')


            resultados_cell = sheet6['A3']
            resultados_cell.value = 'Código'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet6['B3']
            resultados_cell.value = 'Nombre del Registro Administrativo'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet6['C3']
            resultados_cell.value = 'Estado de la crítica'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
                
            resultados_cell = sheet6['D3']
            resultados_cell.value = 'Observaciones de la crítica'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet6['E3']
            resultados_cell.value = 'Funcionario que realiza la crítica'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet6['F3']
            resultados_cell.value = 'Fecha en que se realiza la crítica'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            idraCritica = []
            contsheet6 = 4
            for indexCritica, itemCritica in enumerate(criticaTextList):
                indexCritica =  itemCritica.post_ra_id + 3
                idraCritica.append(itemCritica.post_ra_id)

                sheet6.cell(row = contsheet6, column = 1).value = itemCritica.post_ra.codigo_rraa
                sheet6.cell(row = contsheet6, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet6.cell(row = contsheet6, column = 1).font = Font(bold=True) 

                sheet6.cell(row = contsheet6, column = 2).value = str(itemCritica.post_ra)
                sheet6.cell(row = contsheet6, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet6.cell(row = contsheet6, column = 2).font = Font(bold=True)

                sheet6.cell(row = contsheet6, column = 3).value = str(itemCritica.estado_critica_ra)
                sheet6.cell(row = contsheet6, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet6.cell(row = contsheet6, column = 3).font = Font(bold=True)

                sheet6.cell(row = contsheet6, column = 4).value = str(itemCritica.observa_critica)
                sheet6.cell(row = contsheet6, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet6.cell(row = contsheet6, column = 4).font = Font(bold=True)

                sheet6.cell(row = contsheet6, column = 5).value = str(itemCritica.user_critico)
                sheet6.cell(row = contsheet6, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet6.cell(row = contsheet6, column = 5).font = Font(bold=True)

                sheet6.cell(row = contsheet6, column = 6).value = str(itemCritica.fecha_critica)
                sheet6.cell(row = contsheet6, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet6.cell(row = contsheet6, column = 6).font = Font(bold=True)

                contsheet6+=1

            for row in range(4, sheet6.max_row + 1):  ## definir tamaño de rows
                sheet6.row_dimensions[row].height = 90

            ############# END Hoja 6 Criticas #######################

            ############# Hoja 7 novedades #######################

            sheet7.merge_cells('A1:H1')
            sheet7.merge_cells('A2:H2')

            def set_border(sheet7, cell_range):
                border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                rows = sheet7[cell_range]
                for row in rows:
                    for cell in row:
                        cell.border = border

            set_border(sheet7,'A1:H1')
            

            #row dimensions
            sheet7.row_dimensions[1].height = 55
            sheet7.row_dimensions[2].height = 40
            sheet7.row_dimensions[3].height = 40
            sheet7.row_dimensions[4].height = 40
            sheet7.row_dimensions[5].height = 40
            sheet7.row_dimensions[6].height = 40
            sheet7.row_dimensions[7].height = 40
            sheet7.row_dimensions[8].height = 40

            # column width
            sheet7.column_dimensions['A'].width = 20
            sheet7.column_dimensions['B'].width = 20
            sheet7.column_dimensions['C'].width = 20
            sheet7.column_dimensions['D'].width = 20
            sheet7.column_dimensions['E'].width = 20
            sheet7.column_dimensions['F'].width = 20
            sheet7.column_dimensions['G'].width = 20
            sheet7.column_dimensions['H'].width = 20

            title_cell = sheet7['A1']
            title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
            title_cell.value = 'Directorio de los Registros Administrativos'
            title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
            title_cell.alignment =  Alignment(horizontal='center', vertical='center')

            codigo_cell = sheet7['A2']
            codigo_cell.value = 'Novedades'
            codigo_cell.font = Font(bold=True)
            codigo_cell.alignment = Alignment(horizontal='center')
            codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')


            resultados_cell = sheet7['A3']
            resultados_cell.value = 'Código'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet7['B3']
            resultados_cell.value = 'Nombre del  Registro Administrativo'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet7['C3']
            resultados_cell.value = 'Novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet7['D3']
            resultados_cell.value = 'Estado'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet7['E3']
            resultados_cell.value = 'Descripción de la novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet7['F3']
            resultados_cell.value = 'Observaciones de la novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet7['G3']
            resultados_cell.value = 'Funcionario que realiza la novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet7['H3']
            resultados_cell.value = 'Fecha en que se realiza la novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            idraNovedad = []
            contsheet7 = 4
            for indexNovedad, itemNovedad in enumerate(novedadTextList):
                indexNovedad =  itemNovedad.post_ra_id + 3
                idraNovedad.append(itemNovedad.post_ra_id)

                sheet7.cell(row = contsheet7, column = 1).value = itemNovedad.post_ra.codigo_rraa
                sheet7.cell(row = contsheet7, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 1).font = Font(bold=True) 

                sheet7.cell(row = contsheet7, column = 2).value = str(itemNovedad.post_ra)
                sheet7.cell(row = contsheet7, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 2).font = Font(bold=True) 

                sheet7.cell(row = contsheet7, column = 3).value = str(itemNovedad.novedad)
                sheet7.cell(row = contsheet7, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 3).font = Font(bold=True)

                sheet7.cell(row = contsheet7, column = 4).value = str(itemNovedad.est_actualiz)
                sheet7.cell(row = contsheet7, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 4).font = Font(bold=True)

                sheet7.cell(row = contsheet7, column = 5).value = str(itemNovedad.descrip_novedad)
                sheet7.cell(row = contsheet7, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 5).font = Font(bold=True)

                sheet7.cell(row = contsheet7, column = 6).value = str(itemNovedad.obser_novedad)
                sheet7.cell(row = contsheet7, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 6).font = Font(bold=True)

                sheet7.cell(row = contsheet7, column = 7).value = str(itemNovedad.name_nov)
                sheet7.cell(row = contsheet7, column = 7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 7).font = Font(bold=True)

                sheet7.cell(row = contsheet7, column = 8).value = str(itemNovedad.fecha_actualiz)
                sheet7.cell(row = contsheet7, column = 8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 8).font = Font(bold=True)
                contsheet7+=1

            for row in range(4, sheet7.max_row + 1):  ## definir tamaño de rows
                sheet7.row_dimensions[row].height = 90

        ############# END Hoja 7 novedades #######################


        file_name = "reporte_RRAA.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(file_name)
        response['Content-Disposition'] = content
        wb.save(response)
        return response



#######################  reporte detallado por filtros RRAA #############################

class reportRRAAdetail_xls(TemplateView):  
    def get(self, request, *args, **kwargs):

        id_entidad = request.GET.get('entidad_pri')
        id_area_tematica = request.GET.get('area_temat')
        id_tema = request.GET.get('tema')
        
        ## opc 1 es diferente de vacia
        if id_entidad != "" and id_area_tematica == "" and  id_tema == "":
            
            rraas = RegistroAdministrativo.objects.filter(sist_estado=5).filter(entidad_pri=id_entidad)

        ## opc 1 y opc 2 es diferente de vacia
        elif  id_entidad != "" and id_area_tematica != "" and  id_tema == "":
            
           rraas = RegistroAdministrativo.objects.filter(sist_estado=5).filter(entidad_pri=id_entidad).filter(area_temat=id_area_tematica)

        ## opc 1, opc 2 y opc 3 es diferente de vacia
        elif id_entidad != "" and id_area_tematica != "" and  id_tema != "":
            
            rraas = RegistroAdministrativo.objects.filter(sist_estado=5).filter(entidad_pri=id_entidad).filter(area_temat=id_area_tematica).filter(tema=id_tema)
        
        ## opcion 1 y opcion 3 es diferente de vacia
        elif id_entidad != "" and id_area_tematica == "" and  id_tema != "":
            
            rraas = RegistroAdministrativo.objects.filter(sist_estado=5).filter(entidad_pri=id_entidad).filter(tema=id_tema)
        
        ## opc 2 es diferente de vacia
        elif id_entidad == "" and id_area_tematica != "" and  id_tema == "":
            
            rraas = RegistroAdministrativo.objects.filter(sist_estado=5).filter(area_temat=id_area_tematica)
        ## opc 2 y opc 3 es diferente de vacia
        elif id_entidad == "" and id_area_tematica != "" and  id_tema != "":
            
            rraas = RegistroAdministrativo.objects.filter(sist_estado=5).filter(area_temat=id_area_tematica).filter(tema=id_tema)

        # opc 3 es diferente de vacia
        elif id_entidad == "" and id_area_tematica == "" and  id_tema != "":
            
            rraas = RegistroAdministrativo.objects.filter(sist_estado=5).filter(tema=id_tema)

        normaTextList = []
        documentoMetodTextList = []
        variableRecolectadaTextList = []
        conceptosEstandarizadosTextList = []
        clasificacionTextList = []
        recoleccionDatosTextList = []
        FrecuenciaRecoleccionDatoTextList = []
        herramientasUtilProcesaTextList = []
        seguridadInformTextList = []
        frecuenciaAlmacenamientobdTextList = []
        coberturaGeograficaTextList = []
        indicadorResultadoAgregadoTextList = []
        entidadesAccesoRRAATextList = []
        noAccesoDatosTextList = []
        fortalecimientoTextList = []
        criticaTextList = [] #critica
        novedadTextList = [] #novedad
        user = request.user

        objArray = list(rraas.values('id'))
        for obj in objArray:
            for key, value in obj.items():
                #print("value RRAA", value)
                normaQuestionList = MB_NormaRRAA.objects.filter(rraa_id=value)
                unionListNorma = list(chain(normaQuestionList))
                normaTextList.extend(unionListNorma)

                documentoMetodQuestionList = MB_DocumentoMetodologRRAA.objects.filter(rraa_id=value)
                unionListDocumentoMetod = list(chain(documentoMetodQuestionList))
                documentoMetodTextList.extend(unionListDocumentoMetod)

                variabRecolQuestionList = MB_VariableRecolectada.objects.filter(rraa_id=value)  # formset 7
                unionListVariabRecol = list(chain(variabRecolQuestionList))
                variableRecolectadaTextList.extend(unionListVariabRecol)

                concepEstandQuestionList = MB_ConceptosEstandarizadosRRAA.objects.filter(rraa_id=value)
                unionListConcepEstand = list(chain(concepEstandQuestionList))
                conceptosEstandarizadosTextList.extend(unionListConcepEstand)

                clasificacionQuestionList = MB_ClasificacionesRRAA.objects.filter(rraa_id=value)
                unionListClasificacion = list(chain(clasificacionQuestionList))
                clasificacionTextList.extend(unionListClasificacion)

                recolecDatoQuestionList = MB_RecoleccionDatosRRAA.objects.filter(rraa_id=value)
                unionListRecolecDato = list(chain(recolecDatoQuestionList))
                recoleccionDatosTextList.extend(unionListRecolecDato)

                frecRecolDatoQuestionList = MB_FrecuenciaRecoleccionDato.objects.filter(rraa_id=value)
                unionListFrecRecolDato = list(chain(frecRecolDatoQuestionList))
                FrecuenciaRecoleccionDatoTextList.extend(unionListFrecRecolDato)

                herramUtilProcQuestionList = MB_HerramientasUtilProcesa.objects.filter(rraa_id=value)
                unionListHerramUtilProc = list(chain(herramUtilProcQuestionList))
                herramientasUtilProcesaTextList.extend(unionListHerramUtilProc)

                seguridadInformQuestionList = MB_SeguridadInform.objects.filter(rraa_id=value)
                unionListSeguridadInform = list(chain(seguridadInformQuestionList))
                seguridadInformTextList.extend(unionListSeguridadInform)

                frecAlmacenQuestionList = MB_FrecuenciaAlmacenamientobd.objects.filter(rraa_id=value)
                unionListFrecAlmacen = list(chain(frecAlmacenQuestionList))
                frecuenciaAlmacenamientobdTextList.extend(unionListFrecAlmacen)

                coberGeogQuestionList = MB_CoberturaGeograficaRRAA.objects.filter(rraa_id=value)
                unionListCoberGeog = list(chain(coberGeogQuestionList))
                coberturaGeograficaTextList.extend(unionListCoberGeog)

                indResAgreQuestionList = MB_IndicadorResultadoAgregado.objects.filter(rraa_id=value)  ##formset 16
                unionListIndResAgre = list(chain(indResAgreQuestionList))
                indicadorResultadoAgregadoTextList.extend(unionListIndResAgre)

                entidadesAcceQuestionList = MB_EntidadesAccesoRRAA.objects.filter(rraa_id=value)  ##formset 18
                unionListEntidadesAcce = list(chain(entidadesAcceQuestionList))
                entidadesAccesoRRAATextList.extend(unionListEntidadesAcce)

                noAccesoDatosQuestionList = MB_NoAccesoDatos.objects.filter(rraa_id=value)
                unionListNoAccesoDatos = list(chain(noAccesoDatosQuestionList))
                noAccesoDatosTextList.extend(unionListNoAccesoDatos)

                fortalecimientoQuestionList = FortalecimientoRRAA.objects.filter(post_ra_id=value)  ## otra hoja
                unionListFortalecimiento = list(chain(fortalecimientoQuestionList))
                fortalecimientoTextList.extend(unionListFortalecimiento)

                criticaQuestionList = CriticaRRAA.objects.filter(post_ra_id=value)  ## critica
                unionListCritica = list(chain(criticaQuestionList))
                criticaTextList.extend(unionListCritica)

                novedadQuestionList = NovedadActualizacionRRAA.objects.filter(post_ra_id=value)  ## NovedadActualizacionRRAA
                unionListNovedad = list(chain(novedadQuestionList))
                novedadTextList.extend(unionListNovedad)

        wb = Workbook()
        ws = wb.active
        ws.title = "Directorio RRAA"
        sheet2 = wb.create_sheet('Variables RRAA')
        sheet3 = wb.create_sheet('Resultados RRAA')
        sheet4 = wb.create_sheet('Entidades RRAA')
        sheet5 = wb.create_sheet('Fortalecimiento RRAA')
        sheet6 = wb.create_sheet('Critica RRAA')
        sheet7 = wb.create_sheet('Novedad RRAA')

        
        def set_border(ws, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = ws[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(ws,'A1:DG1'+str(rraas.count()+3))

		## size rows

        ws.row_dimensions[1].height = 55
        ws.row_dimensions[2].height = 40
        ws.row_dimensions[3].height = 80

        ## size column
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 15
        ws.column_dimensions['O'].width = 15
        ws.column_dimensions['P'].width = 15
        ws.column_dimensions['Q'].width = 15
        ws.column_dimensions['R'].width = 22
        ws.column_dimensions['S'].width = 15
        ws.column_dimensions['T'].width = 15
        ws.column_dimensions['U'].width = 15
        ws.column_dimensions['V'].width = 15
        ws.column_dimensions['W'].width = 15
        ws.column_dimensions['X'].width = 15
        ws.column_dimensions['Y'].width = 20
        ws.column_dimensions['Z'].width = 15
        ws.column_dimensions['AA'].width = 15
        ws.column_dimensions['AB'].width = 15
        ws.column_dimensions['AC'].width = 15
        ws.column_dimensions['AD'].width = 15
        ws.column_dimensions['AE'].width = 15
        ws.column_dimensions['AF'].width = 15
        ws.column_dimensions['AG'].width = 15
        ws.column_dimensions['AH'].width = 15
        ws.column_dimensions['AI'].width = 15
        ws.column_dimensions['AJ'].width = 15
        ws.column_dimensions['AK'].width = 15
        ws.column_dimensions['AL'].width = 15
        ws.column_dimensions['AM'].width = 15
        ws.column_dimensions['AN'].width = 15
        ws.column_dimensions['AO'].width = 15
        ws.column_dimensions['AP'].width = 15
        ws.column_dimensions['AQ'].width = 15
        ws.column_dimensions['AR'].width = 15
        ws.column_dimensions['AS'].width = 15
        ws.column_dimensions['AT'].width = 15
        ws.column_dimensions['AU'].width = 15
        ws.column_dimensions['AV'].width = 15
        ws.column_dimensions['AW'].width = 15
        ws.column_dimensions['AX'].width = 15
        ws.column_dimensions['AY'].width = 15
        ws.column_dimensions['AZ'].width = 15
        ws.column_dimensions['BA'].width = 15
        ws.column_dimensions['BB'].width = 15
        ws.column_dimensions['BC'].width = 15
        ws.column_dimensions['BD'].width = 15
        ws.column_dimensions['BE'].width = 15
        ws.column_dimensions['BF'].width = 15
        ws.column_dimensions['BG'].width = 15
        ws.column_dimensions['BH'].width = 15
        ws.column_dimensions['BI'].width = 15
        ws.column_dimensions['BJ'].width = 15
        ws.column_dimensions['BK'].width = 15
        ws.column_dimensions['BL'].width = 15
        ws.column_dimensions['BM'].width = 15
        ws.column_dimensions['BN'].width = 15
        ws.column_dimensions['BO'].width = 15
        ws.column_dimensions['BP'].width = 15
        ws.column_dimensions['BQ'].width = 15
        ws.column_dimensions['BR'].width = 15
        ws.column_dimensions['BS'].width = 15
        ws.column_dimensions['BT'].width = 15
        ws.column_dimensions['BU'].width = 15
        ws.column_dimensions['BV'].width = 15
        ws.column_dimensions['BW'].width = 15
        ws.column_dimensions['BX'].width = 15
        ws.column_dimensions['BY'].width = 15
        ws.column_dimensions['BZ'].width = 15
        ws.column_dimensions['CA'].width = 15
        ws.column_dimensions['CB'].width = 15
        ws.column_dimensions['CC'].width = 15
        ws.column_dimensions['CD'].width = 15
        ws.column_dimensions['CE'].width = 15
        ws.column_dimensions['CF'].width = 15
        ws.column_dimensions['CG'].width = 15
        ws.column_dimensions['CH'].width = 15
        ws.column_dimensions['CI'].width = 15
        ws.column_dimensions['CJ'].width = 15
        ws.column_dimensions['CK'].width = 15
        ws.column_dimensions['CL'].width = 15
        ws.column_dimensions['CM'].width = 15
        ws.column_dimensions['CN'].width = 15
        ws.column_dimensions['CO'].width = 15
        ws.column_dimensions['CP'].width = 15
        ws.column_dimensions['CQ'].width = 15
        ws.column_dimensions['CR'].width = 15
        ws.column_dimensions['CS'].width = 15
        ws.column_dimensions['CT'].width = 15
        ws.column_dimensions['CU'].width = 15
        ws.column_dimensions['CV'].width = 15
        ws.column_dimensions['CW'].width = 15
        ws.column_dimensions['CX'].width = 15
        ws.column_dimensions['CY'].width = 15
        ws.column_dimensions['CZ'].width = 15
        ws.column_dimensions['DA'].width = 15
        ws.column_dimensions['DB'].width = 22
        ws.column_dimensions['DC'].width = 22
        ws.column_dimensions['DD'].width = 22
        ws.column_dimensions['DE'].width = 22
        ws.column_dimensions['DF'].width = 22
        ws.column_dimensions['DG'].width = 25

        title_cell = ws['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de Registros Administrativos'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')
    
        ws.merge_cells('A1:DG1')

        ws.merge_cells('A2:C2')
        ws.merge_cells('D2:E2')
        ws.merge_cells('F2:H2')
        ws.merge_cells('I2:Q2')
        ws.merge_cells('R2:V2')
        ws.merge_cells('W2:W3')
        ws.merge_cells('X2:Y2')
        ws.merge_cells('Z2:AD2')
        ws.merge_cells('AE2:AR2')

        ws.merge_cells('AS2:AV2')
        ws.merge_cells('AW2:BC2')
        ws.merge_cells('BD2:BL2')
        ws.merge_cells('BM2:BU2')
        ws.merge_cells('BV2:BZ2')
        ws.merge_cells('CA2:CH2')
        ws.merge_cells('CI2:CS2')
        ws.merge_cells('CT2:CT3')
        ws.merge_cells('CU2:CU3')
        ws.merge_cells('CV2:DA2')
        ws.merge_cells('DB2:DB3')

        ws.merge_cells('DC2:DC3')
        ws.merge_cells('DD2:DD3')
        ws.merge_cells('DE2:DE3')
        ws.merge_cells('DF2:DF3')
        ws.merge_cells('DG2:DG3')

        ## heads

        codigo_cell = ws['A2']
        codigo_cell.value = 'REGISTRO ADMINISTRATIVO'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['D2']
        codigo_cell.value = 'ENTIDAD'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['F2']
        codigo_cell.value = 'ÁREA TEMÁTICA / TEMA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['I2']
        codigo_cell.value = ' A. IDENTIFICACIÓN '
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['R2']
        codigo_cell.value = 'Bajo cuál(es) de las siguientes normas, se soporta la creación del RA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['W2']
        codigo_cell.value = 'Describa la razón por la cuál se creó el RA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['X2']
        codigo_cell.value = 'Indique desde y hasta cuando se ha recolectado el RA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['Z2']
        codigo_cell.value = '¿El RA cuenta con alguno de los siguientes documentos metodológicos o funcionales?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AE2']
        codigo_cell.value = 'Indique si el RA utiliza conceptos estandarizados provenientes de'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AS2']
        codigo_cell.value = '¿El RA utiliza nomenclaturas y/o clasificaciones?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AW2']
        codigo_cell.value = '¿Cuál es el medio de obtención o recolección de los datos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BD2']
        codigo_cell.value = '¿Con qué frecuencia se recolectan los datos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['BM2']
        codigo_cell.value = 'Indique cuáles de las siguientes herramientas son utilizadas en el procesamiento de los datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BV2']
        codigo_cell.value = '¿Con cuáles herramientas cuenta para garantizar la seguridad de la información del RA?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CA2']
        codigo_cell.value = '¿La información recolectada es acopiada o almacenada en una base de datos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CI2']
        codigo_cell.value = '¿Cuál es la cobertura geográfica del RA?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CT2']
        codigo_cell.value = '¿La entidad hace uso de los datos del RA para generar información estadística (Agregados, indicadores, Resultados estadísticos, etc)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CU2']
        codigo_cell.value = '¿Usuarios externos a la entidad, tienen acceso a los datos del RA?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CV2']
        codigo_cell.value = '¿Cuál es la razón principal por la cual no se permite el acceso a los datos del RA?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DB2']
        codigo_cell.value = 'OBSERVACIONES'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DC2']
        codigo_cell.value = 'Registro administrativo activo (SI / NO)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DD2']
        codigo_cell.value = 'Usuario DANE'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DE2']
        codigo_cell.value = 'Responde a requerimientos ODS (SI / NO)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DF2']
        codigo_cell.value = 'Indicador ODS a que da respuesta'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DG2']
        codigo_cell.value = 'Estado del proceso del Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        #-----------------------------------------------------------------------------------------------------------

        codigo_cell = ws['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['B3']
        codigo_cell.value = 'Nombre del Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['C3']
        codigo_cell.value = 'Objetivo del Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['D3']
        codigo_cell.value = 'Nombre de la Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center',  wrap_text=True)
       
        codigo_cell = ws['E3']
        codigo_cell.value = 'Código de la Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['F3']
        codigo_cell.value = 'Área Tematica'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = ws['G3']
        codigo_cell.value = 'Tema'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['H3']
        codigo_cell.value = 'Tema Compartido'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['I3']
        codigo_cell.value = 'Nombre de la Dependencia'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['J3']
        codigo_cell.value = 'Nombre del Director'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['K3']
        codigo_cell.value = 'Cargo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['L3']
        codigo_cell.value = 'Correo Electrónico del Director'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = ws['M3']
        codigo_cell.value = 'Teléfono del director'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['N3']
        codigo_cell.value = 'Nombre del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['O3']
        codigo_cell.value = 'Cargo del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['P3']
        codigo_cell.value = 'Correo Electrónico del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['Q3']
        codigo_cell.value = 'Teléfono del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['R3']
        codigo_cell.value = 'a. Constitución Política'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['S3']
        codigo_cell.value = 'b. Ley'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['T3']
        codigo_cell.value = 'c. Decreto (nacional, departamental, municipal)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['U3']
        codigo_cell.value = 'd. Otra (Resolución, ordenanza, acuerdo)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center',  wrap_text=True)

        codigo_cell = ws['V3']
        codigo_cell.value = 'e. Ninguna'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['X3']
        codigo_cell.value = 'Fecha de inicio de recolección de información'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['Y3']
        codigo_cell.value = 'Fecha de última recolección de información'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['Z3']
        codigo_cell.value = 'a) Ficha técnica'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AA3']
        codigo_cell.value = 'b) Manual de diligenciamiento'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AB3']
        codigo_cell.value = 'c) Diccionario de datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AC3']
        codigo_cell.value = 'd) Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AD3']
        codigo_cell.value = 'd) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AE3']
        codigo_cell.value = 'a). DANE'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AF3']
        codigo_cell.value = 'b). OCDE'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AG3']
        codigo_cell.value = 'c). ONU'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AH3']
        codigo_cell.value = 'd). EUROSTAT'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AI3']
        codigo_cell.value = 'e). Otro organismo Internacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AJ3']
        codigo_cell.value = 'e) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
  
        codigo_cell = ws['AK3']
        codigo_cell.value = 'f). Otra entidad de orden nacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AL3']
        codigo_cell.value = 'f) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AM3']
        codigo_cell.value = 'g). Leyes, decretos, etc.'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AN3']
        codigo_cell.value = 'g) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AO3']
        codigo_cell.value = 'h). Creación propia'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AP3']
        codigo_cell.value = 'i). Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AQ3']
        codigo_cell.value = 'i) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AR3']
        codigo_cell.value = 'j). No utiliza conceptos estandarizados'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AS3']
        codigo_cell.value = '¿utiliza nomenclaturas y/o clasificaciones?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AT3']
        codigo_cell.value = 'Si: Cuáles?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AU3']
        codigo_cell.value = 'Si: Otras'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


        codigo_cell = ws['AV3']
        codigo_cell.value = 'No: ¿Por qué?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


        codigo_cell = ws['AW3']
        codigo_cell.value = 'a) Formulario físico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AX3']
        codigo_cell.value = 'b) Formulario electrónico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


        codigo_cell = ws['AY3']
        codigo_cell.value = 'c) Dispositivo Móvil de Captura [DMC]'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


        codigo_cell = ws['AZ3']
        codigo_cell.value = 'd) Sistema de información'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BA3']
        codigo_cell.value = 'd) Cuáles?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BB3']
        codigo_cell.value = 'e) Imágenes satelitales'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BC3']
        codigo_cell.value = 'f) Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BD3']
        codigo_cell.value = 'a) Anual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BE3']
        codigo_cell.value = 'b) Semestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BF3']
        codigo_cell.value = 'c) Bimensual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BG3']
        codigo_cell.value = 'd) Trimestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BH3']
        codigo_cell.value = 'e) Mensual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BI3']
        codigo_cell.value = 'f) Diaria'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BJ3']
        codigo_cell.value = 'g) Por evento'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BK3']
        codigo_cell.value = 'h) Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BL3']
        codigo_cell.value = 'h) ¿cuál(es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BM3']
        codigo_cell.value = 'a) Excel'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BN3']
        codigo_cell.value = 'b) Access'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BO3']
        codigo_cell.value = 'c) R.'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BP3']
        codigo_cell.value = 'd) SAS.'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BQ3']
        codigo_cell.value = 'e) SPSS.'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BR3']
        codigo_cell.value = 'f) Oracle'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BS3']
        codigo_cell.value = 'g) Stata.'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BT3']
        codigo_cell.value = 'h) Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BU3']
        codigo_cell.value = 'h) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BV3']
        codigo_cell.value = 'a) Backups'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BW3']
        codigo_cell.value = 'b) Aislamiento del servidor'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BX3']
        codigo_cell.value = 'c) Utilización de perfiles'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BY3']
        codigo_cell.value = 'd) Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BZ3']
        codigo_cell.value = 'd) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CA3']
        codigo_cell.value = '¿La información es almacenada en una base de datos? '
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CB3']
        codigo_cell.value = 'a) Anual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CC3']
        codigo_cell.value = 'b) Semestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CD3']
        codigo_cell.value = 'c) Trimestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CE3']
        codigo_cell.value = 'd) Mensual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CF3']
        codigo_cell.value = 'e) Diaria'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CG3']
        codigo_cell.value = 'f) Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CH3']
        codigo_cell.value = 'f) ¿cuál(es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CI3']
        codigo_cell.value = 'a) Nacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CJ3']
        codigo_cell.value = 'b) Regional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CK3']
        codigo_cell.value = 'b) ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CL3']
        codigo_cell.value = 'c) Departamental'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CM3']
        codigo_cell.value = 'c) ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CN3']
        codigo_cell.value = 'd) Áreas metropolitanas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CO3']
        codigo_cell.value = 'd) ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CP3']
        codigo_cell.value = 'e) Municipal'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CQ3']
        codigo_cell.value = 'e) ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CR3']
        codigo_cell.value = 'f) Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CS3']
        codigo_cell.value = 'f) ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CV3']
        codigo_cell.value = 'a) Políticas de la entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CW3']
        codigo_cell.value = 'b) Disposición normativa'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CX3']
        codigo_cell.value = 'c) Acuerdos con el informante'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CY3']
        codigo_cell.value = 'd) Carácter confidencial'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CZ3']
        codigo_cell.value = 'e) Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DA3']
        codigo_cell.value = 'e) ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
 
        cont = 4
           
        for rraa in rraas:
            
            ws.cell(row = cont, column = 1).value = rraa.codigo_rraa
            ws.cell(row = cont, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 1).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 2).value = rraa.nombre_ra
            ws.cell(row = cont, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 2).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 3).value = rraa.objetivo_ra
            ws.cell(row = cont, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 3).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 4).value = str(rraa.entidad_pri)
            ws.cell(row = cont, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 4).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 5).value =  rraa.entidad_pri.codigo
            ws.cell(row = cont, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 5).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 6).value = str(rraa.area_temat)
            ws.cell(row = cont, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 6).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 7).value = str(rraa.tema)
            ws.cell(row = cont, column = 7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 7).font = Font(size = "8", name='Barlow')

            ## TEMA COMPARTIDO

            listaTemaCompartido = rraa.tema_compart.all()
            temaCompartido_array = []
            for indexTemaCompa, itemTemaCompa in enumerate(listaTemaCompartido):
                temaCompartido_array.append(str(itemTemaCompa))
                ws.cell(row = cont, column = 8).value = str(temaCompartido_array)
                ws.cell(row = cont, column = 8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 8).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 9).value = rraa.nom_dep
            ws.cell(row = cont, column = 9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = cont, column = 9).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 10).value = rraa.nom_dir
            ws.cell(row = cont, column = 10).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 10).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 11).value = rraa.car_dir
            ws.cell(row = cont, column = 11).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 11).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 12).value = rraa.cor_dir
            ws.cell(row = cont, column = 12).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 12).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 13).value = rraa.telef_dir
            ws.cell(row = cont, column = 13).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 13).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 14).value = rraa.nom_resp
            ws.cell(row = cont, column = 14).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 14).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 15).value = rraa.carg_resp
            ws.cell(row = cont, column = 15).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 15).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 16).value = rraa.cor_resp
            ws.cell(row = cont, column = 16).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 16).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 17).value = rraa.telef_resp
            ws.cell(row = cont, column = 17).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 17).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 18).value = rraa.pq_secreo
            ws.cell(row = cont, column = 18).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 18).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 23).value = rraa.pq_secreo
            ws.cell(row = cont, column = 23).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 23).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 24).value = rraa.fecha_ini_rec
            ws.cell(row = cont, column = 24).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 24).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 25).value = rraa.fecha_ult_rec
            ws.cell(row = cont, column = 25).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 25).font = Font(size = "8", name='Barlow')

            if str(rraa.clas_s_n) == "True":
                ws.cell(row = cont, column = 45).value = "Si"
                ws.cell(row = cont, column = 45).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 45).font = Font(size = "8", name='Barlow')
            
            elif str(rraa.clas_s_n) == "False":
                ws.cell(row = cont, column = 45).value = "No"
                ws.cell(row = cont, column = 45).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 45).font = Font(size = "8", name='Barlow')

            ##Clasificaciones
            listaClasificaciones = rraa.nomb_cla.all()
            clasificaciones_array = []

            for indexClasificaciones, itemClasificaciones in enumerate(listaClasificaciones):
                clasificaciones_array.append(str(itemClasificaciones))
                ws.cell(row = cont, column = 46).value = str(clasificaciones_array)
                ws.cell(row = cont, column = 46).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 46).font = Font(size = "8", name='Barlow')

            if str(rraa.almacen_bd_s_n) == "True":
                ws.cell(row = cont, column = 79).value = "Si"
                ws.cell(row = cont, column = 79).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 79).font = Font(size = "8", name='Barlow')
            
            elif str(rraa.almacen_bd_s_n) == "False":
                ws.cell(row = cont, column = 79).value = "No"
                ws.cell(row = cont, column = 79).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 79).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 98).value = str(rraa.uso_de_datos)
            ws.cell(row = cont, column = 98).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 98).font = Font(size = "8", name='Barlow')

            if str(rraa.user_exte_acceso) == "True":
                ws.cell(row = cont, column = 99).value = "Si"
                ws.cell(row = cont, column = 99).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 99).font = Font(size = "8", name='Barlow')
            
            elif str(rraa.user_exte_acceso) == "False":
                ws.cell(row = cont, column = 99).value = "No"
                ws.cell(row = cont, column = 99).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 99).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 106).value = rraa.observacion
            ws.cell(row = cont, column = 106).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 106).font = Font(size = "8", name='Barlow')

            if str(rraa.ra_activo) == "True":
                ws.cell(row = cont, column = 107).value = "Si"
                ws.cell(row = cont, column = 107).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 107).font = Font(size = "8", name='Barlow')
            
            elif str(rraa.ra_activo) == "False":
                ws.cell(row = cont, column = 107).value = "No"
                ws.cell(row = cont, column = 107).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 107).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 108).value = rraa.user_dane
            ws.cell(row = cont, column = 108).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 108).font = Font(size = "8", name='Barlow')

            if str(rraa.responde_ods) == "True":
                ws.cell(row = cont, column = 109).value = "Si"
                ws.cell(row = cont, column = 109).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 109).font = Font(size = "8", name='Barlow')
            
            elif str(rraa.responde_ods) == "False":
                ws.cell(row = cont, column = 109).value = "No"
                ws.cell(row = cont, column = 109).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 109).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 110).value = rraa.indicador_ods
            ws.cell(row = cont, column = 110).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 110).font = Font(size = "8", name='Barlow')

            
            ws.cell(row = cont, column = 111).value = str(rraa.sist_estado)
            ws.cell(row = cont, column = 111).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 111).font = Font(size = "8", name='Barlow')
            
            ##Si desea continuar despues de la columna 111 _____________
            
            ## opciones de preguntas
            listaNorma = rraa.norma_ra.all()
            listaDocumentoMetod = rraa.doc_met_ra.all()
            listaConcepEstand = rraa.con_est_ra.all()
            listaRecolec = rraa.recole_dato.all()
            listaFrecRecol = rraa.fre_rec_dato.all()
            listaHerramUtil = rraa.herr_u_pro.all()
            listaSeguridadIn = rraa.seg_inf.all()
            listaFrecAlm = rraa.frec_alm_bd.all()
            listaCoberGeo = rraa.cob_geograf.all()
            listaNoAccesoDat = rraa.no_hay_acceso.all()

            for index, item in enumerate(listaNorma):
                indice = index
                
                if  str(item) == 'Ninguna' and index == indice:
                    #print("item", item)
                    ws.cell(row = cont, column = 22).value = str(item)
                    ws.cell(row = cont, column = 22).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 22).font = Font(size = "8", name='Barlow')

            for indexDocMet, itemDocMet in enumerate(listaDocumentoMetod):
        
                if  str(itemDocMet) == 'Ficha técnica':
                   
                    ws.cell(row = cont, column = 26).value = str(itemDocMet)
                    ws.cell(row = cont, column = 26).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 26).font = Font(size = "8", name='Barlow')

                if  str(itemDocMet) == 'Manual de diligenciamiento':

                    ws.cell(row = cont, column = 27).value = str(itemDocMet)
                    ws.cell(row = cont, column = 27).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 27).font = Font(size = "8", name='Barlow')
                
                if str(itemDocMet) == 'Diccionario de datos':

                    ws.cell(row = cont, column = 28).value = str(itemDocMet)
                    ws.cell(row = cont, column = 28).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 28).font = Font(size = "8", name='Barlow')

                if str(itemDocMet) == 'Otro (s)':

                    ws.cell(row = cont, column = 29).value = str(itemDocMet)
                    ws.cell(row = cont, column = 29).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 29).font = Font(size = "8", name='Barlow')

            for itemConcepEstand, indexConcepEstand in enumerate(listaConcepEstand):
                
                if str(itemConcepEstand) == 'DANE':

                    ws.cell(row = cont, column = 31).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 31).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 31).font = Font(size = "8", name='Barlow')

                if str(itemConcepEstand) == 'OCDE':
                    ws.cell(row = cont, column = 32).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 32).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 32).font = Font(size = "8", name='Barlow')

                if str(itemConcepEstand) == 'ONU':
                    ws.cell(row = cont, column = 33).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 33).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 33).font = Font(size = "8", name='Barlow')

                if str(itemConcepEstand) == 'EUROSTAT':
                    ws.cell(row = cont, column = 34).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 34).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 34).font = Font(size = "8", name='Barlow')

                if str(itemConcepEstand) == 'Otro organismo Internacional':
                    ws.cell(row = cont, column = 35).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 35).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 35).font = Font(size = "8", name='Barlow')

                if str(itemConcepEstand) == 'Otra entidad de orden nacional':
                    ws.cell(row = cont, column = 37).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 37).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 37).font = Font(size = "8", name='Barlow')
                
                if str(itemConcepEstand) == 'Leyes, decretos, etc':
                    ws.cell(row = cont, column = 39).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 39).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 39).font = Font(size = "8", name='Barlow')   
                
                if str(itemConcepEstand) == 'Creación propia':
                    ws.cell(row = cont, column = 41).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 41).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 41).font = Font(size = "8", name='Barlow')
                
                if str(itemConcepEstand) == 'Otra (s)':
                    ws.cell(row = cont, column = 42).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 42).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 42).font = Font(size = "8", name='Barlow')
                
                if str(itemConcepEstand) == 'No utiliza conceptos estandarizados':
                    ws.cell(row = cont, column = 44).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 44).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 44).font = Font(size = "8", name='Barlow')

            
            for indexRecolec, itemRecolec in enumerate(listaRecolec):

                if str(itemRecolec) == 'Formulario físico':
                    ws.cell(row = cont, column = 49).value = str(itemRecolec)
                    ws.cell(row = cont, column = 49).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 49).font = Font(size = "8", name='Barlow')

                if str(itemRecolec) == 'Formulario electrónico':
                    ws.cell(row = cont, column = 50).value = str(itemRecolec)
                    ws.cell(row = cont, column = 50).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 50).font = Font(size = "8", name='Barlow')

                if str(itemRecolec) == 'Dispositivo Móvil de Captura [DMC]':
                    ws.cell(row = cont, column = 51).value = str(itemRecolec)
                    ws.cell(row = cont, column = 51).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 51).font = Font(size = "8", name='Barlow')

                if str(itemRecolec) == 'Sistema de información':
                    ws.cell(row = cont, column = 52).value = str(itemRecolec)
                    ws.cell(row = cont, column = 52).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 52).font = Font(size = "8", name='Barlow')

                if str(itemRecolec) == 'Imágenes satelitales':
                    ws.cell(row = cont, column = 54).value = str(itemRecolec)
                    ws.cell(row = cont, column = 54).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 54).font = Font(size = "8", name='Barlow')

            for indexFrecRecol, itemFrecRecol in enumerate(listaFrecRecol):

                if str(itemFrecRecol) == 'Anual':
                    ws.cell(row = cont, column = 56).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 56).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 56).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Semestral':
                    ws.cell(row = cont, column = 57).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 57).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 57).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Bimensual':
                    ws.cell(row = cont, column = 58).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 58).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 58).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Trimestral':
                    ws.cell(row = cont, column = 59).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 59).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 59).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Mensual':
                    ws.cell(row = cont, column = 60).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 60).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 60).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Diaria':

                    ws.cell(row = cont, column = 61).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 61).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 61).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Por evento':

                    ws.cell(row = cont, column = 62).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 62).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 62).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Otra (s)':

                    ws.cell(row = cont, column = 63).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 63).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 63).font = Font(size = "8", name='Barlow')

            
            for indexHerramUtil, itemHerramUtil in enumerate(listaHerramUtil):

                if str(itemHerramUtil) == 'Excel':
                    ws.cell(row = cont, column = 65).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 65).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 65).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'Access':
                    ws.cell(row = cont, column = 66).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 66).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 66).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'R':
                    ws.cell(row = cont, column = 67).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 67).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 67).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'SAS':
                    ws.cell(row = cont, column = 68).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 68).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 68).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'SPSS':
                    ws.cell(row = cont, column = 69).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 69).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 69).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'Oracle':
                    ws.cell(row = cont, column = 70).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 70).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 70).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'Stata':
                    ws.cell(row = cont, column = 71).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 71).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 71).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'Otra (s)':
                    ws.cell(row = cont, column = 72).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 72).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 72).font = Font(size = "8", name='Barlow')

            for indexSeguridadIn, itemSeguridadIn in enumerate(listaSeguridadIn):

                if str(itemSeguridadIn) == 'Backups':

                    ws.cell(row = cont, column = 74).value = str(itemSeguridadIn)
                    ws.cell(row = cont, column = 74).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 74).font = Font(size = "8", name='Barlow')

                if str(itemSeguridadIn) == 'Aislamiento del servidor':

                    ws.cell(row = cont, column = 75).value = str(itemSeguridadIn)
                    ws.cell(row = cont, column = 75).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 75).font = Font(size = "8", name='Barlow')

                if str(itemSeguridadIn) == 'Utilización de perfiles':

                    ws.cell(row = cont, column = 76).value = str(itemSeguridadIn)
                    ws.cell(row = cont, column = 76).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 76).font = Font(size = "8", name='Barlow')

                if str(itemSeguridadIn) == 'Otra (s)':

                    ws.cell(row = cont, column = 77).value = str(itemSeguridadIn)
                    ws.cell(row = cont, column = 77).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 77).font = Font(size = "8", name='Barlow')

            for indexFrecAlm, itemFrecAlm in enumerate(listaFrecAlm):

                if str(itemFrecAlm) == 'Anual':
                    ws.cell(row = cont, column = 80).value = str(itemFrecAlm)
                    ws.cell(row = cont, column = 80).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 80).font = Font(size = "8", name='Barlow')

                if str(itemFrecAlm) == 'Semestral':
                    ws.cell(row = cont, column = 81).value = str(itemFrecAlm)
                    ws.cell(row = cont, column = 81).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 81).font = Font(size = "8", name='Barlow')

                if str(itemFrecAlm) == 'Trimestral':
                    ws.cell(row = cont, column = 82).value = str(itemFrecAlm)
                    ws.cell(row = cont, column = 82).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 82).font = Font(size = "8", name='Barlow')

                if str(itemFrecAlm) == 'Mensual':
                    ws.cell(row = cont, column = 83).value = str(itemFrecAlm)
                    ws.cell(row = cont, column = 83).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 83).font = Font(size = "8", name='Barlow')

                if str(itemFrecAlm) == 'Diaria':
                    ws.cell(row = cont, column = 84).value = str(itemFrecAlm)
                    ws.cell(row = cont, column = 84).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 84).font = Font(size = "8", name='Barlow')

                if str(itemFrecAlm) == 'Otra (s)':
                    ws.cell(row = cont, column = 85).value = str(itemFrecAlm)
                    ws.cell(row = cont, column = 85).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 85).font = Font(size = "8", name='Barlow')

            for indexCoberGeo, itemCoberGeo in enumerate(listaCoberGeo):

                if str(itemCoberGeo) == 'Nacional':
                    ws.cell(row = cont, column = 87).value = str(itemCoberGeo)
                    ws.cell(row = cont, column = 87).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 87).font = Font(size = "8", name='Barlow')

                if str(itemCoberGeo) == 'Regional':

                    ws.cell(row = cont, column = 88).value = str(itemCoberGeo)
                    ws.cell(row = cont, column = 88).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 88).font = Font(size = "8", name='Barlow')

                if str(itemCoberGeo) == 'Departamental':

                    ws.cell(row = cont, column = 90).value = str(itemCoberGeo)
                    ws.cell(row = cont, column = 90).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 90).font = Font(size = "8", name='Barlow')

                if str(itemCoberGeo) == 'Áreas metropolitanas':

                    ws.cell(row = cont, column = 92).value = str(itemCoberGeo)
                    ws.cell(row = cont, column = 92).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 92).font = Font(size = "8", name='Barlow')

                if str(itemCoberGeo) == 'Municipal':

                    ws.cell(row = cont, column = 94).value = str(itemCoberGeo)
                    ws.cell(row = cont, column = 94).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 94).font = Font(size = "8", name='Barlow')

                if str(itemCoberGeo) == 'Otro (s)':

                    ws.cell(row = cont, column = 96).value = str(itemCoberGeo)
                    ws.cell(row = cont, column = 96).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 96).font = Font(size = "8", name='Barlow')

            for indexNoAccesoDat, itemNoAccesoDat in enumerate(listaNoAccesoDat):

                if str(itemNoAccesoDat) == 'Políticas de la entidad':

                    ws.cell(row = cont, column = 100).value = str(itemNoAccesoDat)
                    ws.cell(row = cont, column = 100).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 100).font = Font(size = "8", name='Barlow')

                if str(itemNoAccesoDat) == 'Disposición normativa':

                    ws.cell(row = cont, column = 101).value = str(itemNoAccesoDat)
                    ws.cell(row = cont, column = 101).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 101).font = Font(size = "8", name='Barlow')

                if str(itemNoAccesoDat) == 'Acuerdos con el informante':

                    ws.cell(row = cont, column = 102).value = str(itemNoAccesoDat)
                    ws.cell(row = cont, column = 102).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 102).font = Font(size = "8", name='Barlow')

                if str(itemNoAccesoDat) == 'Carácter confidencial':

                    ws.cell(row = cont, column = 103).value = str(itemNoAccesoDat)
                    ws.cell(row = cont, column = 103).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 103).font = Font(size = "8", name='Barlow')

                if str(itemNoAccesoDat) == 'Otra (s)':

                    ws.cell(row = cont, column = 104).value = str(itemNoAccesoDat)
                    ws.cell(row = cont, column = 104).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 104).font = Font(size = "8", name='Barlow')

          
            cont+=1

        ## campos de texto   

        #print("---------------------->", normaTextList)
        for indexNorma, itemNorma in enumerate(normaTextList):
           
            indexNorma =  indexNorma + 4
            ws.cell(row = indexNorma, column = 18).value = str(itemNorma.cp_ra)
            ws.cell(row = indexNorma, column = 18).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 18).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexNorma, column = 19).value = str(itemNorma.ley_ra)
            ws.cell(row = indexNorma, column = 19).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 19).font = Font(size = "8", name='Barlow')
                    
            ws.cell(row = indexNorma, column = 20).value = str(itemNorma.decreto_ra)
            ws.cell(row = indexNorma, column = 20).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 20).font = Font(size = "8", name='Barlow')
                    
            ws.cell(row = indexNorma, column = 21).value = str(itemNorma.otra_ra)
            ws.cell(row = indexNorma, column = 21).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 21).font = Font(size = "8", name='Barlow')


        for indexDocumentoMetod, itemDocumentoMetod in enumerate(documentoMetodTextList):
           
            indexDocumentoMetod =  indexDocumentoMetod + 4
            ws.cell(row = indexDocumentoMetod, column = 30).value = str(itemDocumentoMetod.otra_doc_cual)
            ws.cell(row = indexDocumentoMetod, column = 30).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexDocumentoMetod, column = 30).font = Font(size = "8", name='Barlow')


        for indexConEst, itemConEst in enumerate(conceptosEstandarizadosTextList):
           
            indexConEst =  indexConEst + 4

            ws.cell(row = indexConEst, column = 36).value = str(itemConEst.org_in_cual)
            ws.cell(row = indexConEst, column = 36).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConEst, column = 36).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexConEst, column = 38).value = str(itemConEst.ent_ordnac_cual)
            ws.cell(row = indexConEst, column = 38).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConEst, column = 38).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexConEst, column = 40).value = str(itemConEst.leye_dec_cual)
            ws.cell(row = indexConEst, column = 40).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConEst, column = 40).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexConEst, column = 43).value = str(itemConEst.otra_ce_cual)
            ws.cell(row = indexConEst, column = 43).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConEst, column = 43).font = Font(size = "8", name='Barlow')
        

        for indexClasifi, itemClasifi in enumerate(clasificacionTextList):
           
            indexClasifi =  indexClasifi + 4

            ws.cell(row = indexClasifi, column = 47).value = str(itemClasifi.otra_cual_clas)
            ws.cell(row = indexClasifi, column = 47).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexClasifi, column = 47).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexClasifi, column = 48).value = str(itemClasifi.no_pq)
            ws.cell(row = indexClasifi, column = 48).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexClasifi, column = 48).font = Font(size = "8", name='Barlow')

        
        for indexRecoleccionDatos, itemRecoleccionDatos in enumerate(recoleccionDatosTextList):

            indexRecoleccionDatos = indexRecoleccionDatos + 4

            ws.cell(row = indexRecoleccionDatos, column = 53).value = str(itemRecoleccionDatos.sistema_inf_cual)
            ws.cell(row = indexRecoleccionDatos, column = 53).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexRecoleccionDatos, column = 53).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexRecoleccionDatos, column = 55).value = str(itemRecoleccionDatos.otro_c)
            ws.cell(row = indexRecoleccionDatos, column = 55).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexRecoleccionDatos, column = 55).font = Font(size = "8", name='Barlow')

        for indexFrecuenciaRecoleccionDato, itemFrecuenciaRecoleccionDato in enumerate(FrecuenciaRecoleccionDatoTextList):
            
            indexFrecuenciaRecoleccionDato = indexFrecuenciaRecoleccionDato + 4

            ws.cell(row = indexFrecuenciaRecoleccionDato, column = 64).value = str(itemFrecuenciaRecoleccionDato.otra_cual_fre)
            ws.cell(row = indexFrecuenciaRecoleccionDato, column = 64).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexFrecuenciaRecoleccionDato, column = 64).font = Font(size = "8", name='Barlow')

         
        for indexherramientasUtilPro, itemherramientasUtilPro in enumerate(herramientasUtilProcesaTextList):
            
            indexherramientasUtilPro = indexherramientasUtilPro + 4

            ws.cell(row = indexherramientasUtilPro, column = 73).value = str(itemherramientasUtilPro.otra_herram)
            ws.cell(row = indexherramientasUtilPro, column = 73).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexherramientasUtilPro, column = 73).font = Font(size = "8", name='Barlow')

        
        for indexSeguridadInfo, itemSeguridadInfo in enumerate(seguridadInformTextList):

            indexSeguridadInfo = indexSeguridadInfo + 4

            ws.cell(row = indexSeguridadInfo, column = 78).value = str(itemSeguridadInfo.otra_cual_s)
            ws.cell(row = indexSeguridadInfo, column = 78).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexSeguridadInfo, column = 78).font = Font(size = "8", name='Barlow')

        for indexfrecuenciaAlmacen, itemfrecuenciaAlmacen in enumerate(frecuenciaAlmacenamientobdTextList):

            indexfrecuenciaAlmacen = indexfrecuenciaAlmacen + 4

            ws.cell(row = indexfrecuenciaAlmacen, column = 86).value = str(itemfrecuenciaAlmacen.otra_alm_bd)
            ws.cell(row = indexfrecuenciaAlmacen, column = 86).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexfrecuenciaAlmacen, column = 86).font = Font(size = "8", name='Barlow')

        for indexCoberturaGeog, itemCoberturaGeog in enumerate(coberturaGeograficaTextList):
            
            indexCoberturaGeog = indexCoberturaGeog + 4

            ws.cell(row = indexCoberturaGeog, column = 89).value = str(itemCoberturaGeog.cual_regio)
            ws.cell(row = indexCoberturaGeog, column = 89).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 89).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 91).value = str(itemCoberturaGeog.cual_depa)
            ws.cell(row = indexCoberturaGeog, column = 91).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 91).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 93).value = str(itemCoberturaGeog.cual_are_metrop)
            ws.cell(row = indexCoberturaGeog, column = 93).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 93).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 95).value = str(itemCoberturaGeog.cual_munic)
            ws.cell(row = indexCoberturaGeog, column = 95).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 95).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 97).value = str(itemCoberturaGeog.cual_otro)
            ws.cell(row = indexCoberturaGeog, column = 97).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 97).font = Font(size = "8", name='Barlow')

        for indexNoAccesoDatos, itemNoAccesoDatos in enumerate(noAccesoDatosTextList):

            indexNoAccesoDatos = indexNoAccesoDatos + 4

            ws.cell(row = indexNoAccesoDatos, column = 105).value = str(itemNoAccesoDatos.otra_no_acceso)
            ws.cell(row = indexNoAccesoDatos, column = 105).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNoAccesoDatos, column = 105).font = Font(size = "8", name='Barlow')

        for row in range(4, ws.max_row + 1):
            ws.row_dimensions[row].height = 90

        ####################### HOJA 2 ######################## 

        sheet2.merge_cells('A1:C1')
        sheet2.merge_cells('A2:C2')

        def set_border(sheet2, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet2[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet2,'A1:C1')
    
        #row dimensions
        sheet2.row_dimensions[1].height = 55
        sheet2.row_dimensions[2].height = 40

        # column width
        sheet2.column_dimensions['A'].width = 20
        sheet2.column_dimensions['B'].width = 20
        sheet2.column_dimensions['C'].width = 20
        
        title_cell = sheet2['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de los Registros Administrativos'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        title_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
	
        codigo_cell = sheet2['A2']
        codigo_cell.value = 'Lista de variables de los Registros Administrativos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)


        codigo_cell = sheet2['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = sheet2['B3']
        codigo_cell.value = 'Nombre Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet2['C3']
        codigo_cell.value = 'Variables que maneja el Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        contsheet2 = 4

        arrayraId = []
        for indexVariab, itemVariab in enumerate(variableRecolectadaTextList):
            indexVariab =  itemVariab.rraa_id + 3
            arrayraId.append(itemVariab.rraa_id)
            
            sheet2.cell(row = contsheet2, column = 1).value = itemVariab.rraa.codigo_rraa
            sheet2.cell(row = contsheet2, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet2.cell(row = contsheet2, column = 1).font = Font(size = "8", name='Barlow')

            sheet2.cell(row = contsheet2, column = 2).value = str(itemVariab.rraa)
            sheet2.cell(row = contsheet2, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet2.cell(row = contsheet2, column = 2).font = Font(size = "8", name='Barlow')

            sheet2.cell(row = contsheet2, column = 3).value = str(itemVariab.variableRec)
            sheet2.cell(row = contsheet2, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet2.cell(row = contsheet2, column = 3).font = Font(size = "8", name='Barlow')
            contsheet2+=1
           
        for row in range(4, sheet2.max_row + 1):  ## definir tamaño de rows
            sheet2.row_dimensions[row].height = 90

        ##################### HOJA 3  #################################################

        sheet3.merge_cells('A1:C1')
        sheet3.merge_cells('A2:C2')

        def set_border(sheet3, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet3[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet3,'A1:C1')
    
        #row dimensions
        sheet3.row_dimensions[1].height = 55
        sheet3.row_dimensions[2].height = 40

        # column width
        sheet3.column_dimensions['A'].width = 20
        sheet3.column_dimensions['B'].width = 20
        sheet3.column_dimensions['C'].width = 20
        
        title_cell = sheet3['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de los Registros Administrativos'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        title_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
	
        codigo_cell = sheet3['A2']
        codigo_cell.value = 'La entidad hace uso de los datos del RA para generar información estadística (Agregados, indicadores, Resultados estadísticos, etc)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet3['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = sheet3['B3']
        codigo_cell.value = 'Nombre Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet3['C3']
        codigo_cell.value = 'Lista de resultados agregados o indicadores'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        contsheet3 = 4
        arrayraId = []
        for indexIndicador, itemIndicador in enumerate(indicadorResultadoAgregadoTextList):
            indexIndicador =  itemIndicador.rraa_id + 3
            arrayraId.append(itemIndicador.rraa_id)
            
            sheet3.cell(row = contsheet3, column = 1).value = itemIndicador.rraa.codigo_rraa
            sheet3.cell(row = contsheet3, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet3.cell(row = contsheet3, column = 1).font = Font(size = "8", name='Barlow')

            sheet3.cell(row = contsheet3, column = 2).value = str(itemIndicador.rraa)
            sheet3.cell(row = contsheet3, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet3.cell(row = contsheet3, column = 2).font = Font(size = "8", name='Barlow')

            sheet3.cell(row = contsheet3, column = 3).value = str(itemIndicador.ind_res_agre)
            sheet3.cell(row = contsheet3, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet3.cell(row = contsheet3, column = 3).font = Font(size = "8", name='Barlow')
            contsheet3+=1
           
        for row in range(4, sheet3.max_row + 1):  ## definir tamaño de rows
            sheet3.row_dimensions[row].height = 90 


        ####################### Hoja 4 pregunta 18 #############################################


        def set_border(sheet4, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet4[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet4,'A1:AG'+str(rraas.count()+3))

		## size rows

        sheet4.row_dimensions[1].height = 55
        sheet4.row_dimensions[2].height = 40
        sheet4.row_dimensions[3].height = 50

        ## size column
        sheet4.column_dimensions['A'].width = 18
        sheet4.column_dimensions['B'].width = 18
        sheet4.column_dimensions['C'].width = 15
        sheet4.column_dimensions['D'].width = 15
        sheet4.column_dimensions['E'].width = 15
        sheet4.column_dimensions['F'].width = 15
        sheet4.column_dimensions['G'].width = 15
        sheet4.column_dimensions['H'].width = 15
        sheet4.column_dimensions['I'].width = 15
        sheet4.column_dimensions['J'].width = 15
        sheet4.column_dimensions['K'].width = 15
        sheet4.column_dimensions['L'].width = 15
        sheet4.column_dimensions['M'].width = 15
        sheet4.column_dimensions['N'].width = 15
        sheet4.column_dimensions['O'].width = 15
        sheet4.column_dimensions['P'].width = 15
        sheet4.column_dimensions['Q'].width = 15
        sheet4.column_dimensions['R'].width = 15
        sheet4.column_dimensions['S'].width = 15
        sheet4.column_dimensions['T'].width = 15
        sheet4.column_dimensions['U'].width = 15
        sheet4.column_dimensions['V'].width = 15
        sheet4.column_dimensions['W'].width = 15
        sheet4.column_dimensions['X'].width = 15
        sheet4.column_dimensions['Y'].width = 15
        sheet4.column_dimensions['Z'].width = 15
        sheet4.column_dimensions['AA'].width = 15
        sheet4.column_dimensions['AB'].width = 15
        sheet4.column_dimensions['AC'].width = 15
        sheet4.column_dimensions['AD'].width = 15
        sheet4.column_dimensions['AE'].width = 15
        sheet4.column_dimensions['AF'].width = 15
        sheet4.column_dimensions['AG'].width = 15
        sheet4.column_dimensions['AH'].width = 15
        

        title_cell = sheet4['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de Registros Administrativos'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')
    
        sheet4.merge_cells('A1:AG1')
        sheet4.merge_cells('A2:B2')
        sheet4.merge_cells('C2:AG2')

        ## heads

        codigo_cell = sheet4['A2']
        codigo_cell.value = 'REGISTRO ADMINISTRATIVO'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet4['C2']
        codigo_cell.value = 'Relacione las entidades que tienen acceso a los datos del RA:'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        
        codigo_cell = sheet4['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet4['B3']
        codigo_cell.value = 'Nombre del Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        
        contsheet4 = 4
        for rraa in rraas: 
            sheet4.cell(row = contsheet4, column = 1).value = rraa.codigo_rraa
            sheet4.cell(row = contsheet4, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet4.cell(row = contsheet4, column = 1).font = Font(size = "8", name='Barlow')

            sheet4.cell(row = contsheet4, column = 2).value = rraa.nombre_ra
            sheet4.cell(row = contsheet4, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet4.cell(row = contsheet4, column = 2).font = Font(size = "8", name='Barlow')

            contsheet4+= 1
    
        entidadraid = []
        for indexEntidAcc, itemEntidAcc in enumerate(entidadesAccesoRRAATextList): 
            indexEntidAcc = itemEntidAcc.rraa_id + 3
                    
            entidadraid.append(itemEntidAcc.rraa_id)
           
            count_id = entidadraid.count(itemEntidAcc.rraa_id)
            c=0 #inicializamos el contador  
            n=5*count_id
            for i in range(1,n+1):  
                if i%5 == 0:  
                    i = i - 2  ##posición de celda
                    c+=1

            for indexra, itemra in enumerate(rraas):
                if itemEntidAcc.rraa_id == itemra.pk:
                    posRow = indexra + 4
                     
                    sheet4.cell(row = 3, column = i ).value = "Nombre de Entidad"
                    sheet4.cell(row = 3, column = i ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = 3, column = i ).font = Font(bold=True) 

                    sheet4.cell(row = posRow, column = i ).value = str(itemEntidAcc.nomb_entidad_acc) 
                    sheet4.cell(row = posRow, column = i ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = posRow, column = i ).font = Font(size = "8", name='Barlow')

                    sheet4.cell(row = 3, column = i+4).value = "¿cuál?"
                    sheet4.cell(row = 3, column = i+4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = 3, column = i+4).font = Font(bold=True)

                    sheet4.cell(row = posRow, column = i+4).value = str(itemEntidAcc.otro_cual) 
                    sheet4.cell(row = posRow, column = i+4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = posRow, column = i+4).font = Font(size = "8", name='Barlow')


            
            listaFinalidad = list(itemEntidAcc.opcion_pr.all())  ##iterar para traer las fases seleccionadas

            for indexFinalida, itemFinalidad in enumerate(listaFinalidad):

                if str(itemFinalidad) == "Estadístico":
                    
                    countFin=0 #inicializamos el contador  
                    finSelected=5*count_id 
                    for posFin in range(1,n+1):  
                        if posFin%5 == 0:
                            posFin = posFin - 1 ##posicion celda
                            countFin+=1

                    sheet4.cell(row = 3, column = posFin ).value = "Estadístico"
                    sheet4.cell(row = 3, column = posFin ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = 3, column = posFin ).font = Font(bold=True)

                    sheet4.cell(row = posRow, column = posFin).value = str(itemFinalidad)
                    sheet4.cell(row = posRow, column = posFin).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = posRow, column = posFin).font = Font(size = "8", name='Barlow')
                                
                if str(itemFinalidad) == "No sabe":

                    countFin=0 #inicializamos el contador  
                    finSelected=5*count_id 
                    for posFin in range(1,n+1):  
                        if posFin%5 == 0:  
                            posFin = posFin ##posición celda 
                            countFin+=1

                    sheet4.cell(row = 3, column = posFin ).value = "No sabe"
                    sheet4.cell(row = 3, column = posFin ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = 3, column = posFin ).font = Font(bold=True)

                    sheet4.cell(row = posRow, column = posFin).value = str(itemFinalidad)
                    sheet4.cell(row = posRow, column = posFin).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = posRow, column = posFin).font = Font(size = "8", name='Barlow')

                if str(itemFinalidad) == "Otro, ¿cuál?":

                    countFin=0 #inicializamos el contador  
                    finSelected=5*count_id 
                    for posFin in range(1,n+1):  
                        if posFin%5 == 0:  
                            posFin = posFin + 1
                            countFin+=1

                    sheet4.cell(row = 3, column = posFin ).value = "Otro"
                    sheet4.cell(row = 3, column = posFin ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = 3, column = posFin ).font = Font(bold=True)

                    sheet4.cell(row = posRow, column = posFin).value = str(itemFinalidad)
                    sheet4.cell(row = posRow, column = posFin).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = posRow, column = posFin).font = Font(size = "8", name='Barlow')
                


        for row in range(4, sheet4.max_row + 1):
            sheet4.row_dimensions[row].height = 90


        #################################### Hoja 5 ###########################

        def set_border(sheet5, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet5[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet5,'A1:AZ'+str(rraas.count()+3))

		## size rows

        sheet5.row_dimensions[1].height = 55
        sheet5.row_dimensions[2].height = 40
        sheet5.row_dimensions[3].height = 50

        ## size column
        sheet5.column_dimensions['A'].width = 18
        sheet5.column_dimensions['B'].width = 18
        sheet5.column_dimensions['C'].width = 24
        sheet5.column_dimensions['D'].width = 24
        sheet5.column_dimensions['E'].width = 24
        sheet5.column_dimensions['F'].width = 24
        sheet5.column_dimensions['G'].width = 24
        sheet5.column_dimensions['H'].width = 24
        sheet5.column_dimensions['I'].width = 24
        sheet5.column_dimensions['J'].width = 24
        sheet5.column_dimensions['K'].width = 24
        sheet5.column_dimensions['L'].width = 24
        sheet5.column_dimensions['M'].width = 24
        sheet5.column_dimensions['N'].width = 24
        sheet5.column_dimensions['O'].width = 24
        sheet5.column_dimensions['P'].width = 24
        sheet5.column_dimensions['Q'].width = 24
        sheet5.column_dimensions['R'].width = 24
        sheet5.column_dimensions['S'].width = 24
        sheet5.column_dimensions['T'].width = 24
        sheet5.column_dimensions['U'].width = 24
        sheet5.column_dimensions['V'].width = 24
        sheet5.column_dimensions['W'].width = 24
        sheet5.column_dimensions['X'].width = 24
        sheet5.column_dimensions['Y'].width = 24
        sheet5.column_dimensions['Z'].width = 24
        sheet5.column_dimensions['AA'].width = 24
        sheet5.column_dimensions['AB'].width = 24
        sheet5.column_dimensions['AC'].width = 24
        sheet5.column_dimensions['AD'].width = 24
        sheet5.column_dimensions['AE'].width = 24
        sheet5.column_dimensions['AF'].width = 24
        sheet5.column_dimensions['AG'].width = 24
        sheet5.column_dimensions['AH'].width = 24
        
        sheet5.merge_cells('A1:AZ1')
        sheet5.merge_cells('A2:F2')
        sheet5.merge_cells('G2:AZ2')

        ## heads
        
        title_cell = sheet5['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de Registros Administrativos'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet5['A2']
        codigo_cell.value = 'IDENTIFICACIÓN'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet5['G2']
        codigo_cell.value = 'Fortalecimiento RRAA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['A3']
        resultados_cell.value = 'Área Temática'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['B3']
        resultados_cell.value = 'Tema'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['C3']
        resultados_cell.value = 'Código Entidad'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['D3']
        resultados_cell.value = 'Nombre Entidad'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['E3']
        resultados_cell.value = 'Código RRAA'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['F3']
        resultados_cell.value = 'Nombre del registro administrativo'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        
        contsheet5 = 4
        for rraa in rraas: 
            
            sheet5.cell(row = contsheet5, column = 1).value = str(rraa.area_temat)
            sheet5.cell(row = contsheet5, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 1).font = Font(size = "8", name='Barlow') 
            
            sheet5.cell(row = contsheet5, column = 2).value = str(rraa.tema)
            sheet5.cell(row = contsheet5, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 2).font = Font(size = "8", name='Barlow') 
            
            sheet5.cell(row = contsheet5, column = 3).value = rraa.entidad_pri.codigo
            sheet5.cell(row = contsheet5, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 3).font = Font(size = "8", name='Barlow')

            sheet5.cell(row = contsheet5, column = 4).value = str(rraa.entidad_pri)
            sheet5.cell(row = contsheet5, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 4).font = Font(size = "8", name='Barlow')

            sheet5.cell(row = contsheet5, column = 5).value = rraa.codigo_rraa
            sheet5.cell(row = contsheet5, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 5).font = Font(size = "8", name='Barlow')

            sheet5.cell(row = contsheet5, column = 6).value = rraa.nombre_ra
            sheet5.cell(row = contsheet5, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 6).font = Font(size = "8", name='Barlow')
            
            contsheet5+= 1

        idraFort = []
        for indexFort, itemFort in enumerate(fortalecimientoTextList): 
            indexFort = itemFort.post_ra_id + 3
            idraFort.append(itemFort.post_ra_id)

            ### 1 Diagnóstico del RA
            
            count_diagnostico = idraFort.count(itemFort.post_ra_id)
            
            contrField1=0 #inicializamos el contador  
            pos_diagno_rraa = 9*count_diagnostico
            for incField1 in range(1,pos_diagno_rraa+1):  
                if incField1%9 == 0:  
                    incField1 = incField1 - 2
                    contrField1+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField1).value = "Diagnóstico del RRAA"
                    sheet5.cell(row = 3, column = incField1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField1).font = Font(bold=True) 

                    if str(itemFort.diagnostico_ra) == "True":

                        sheet5.cell(row = posiRow, column = incField1).value = "Si"
                        sheet5.cell(row = posiRow, column = incField1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField1).font = Font(size = "8", name='Barlow')

                    elif str(itemFort.diagnostico_ra) == "False":

                        sheet5.cell(row = posiRow, column = incField1).value = "No"
                        sheet5.cell(row = posiRow, column = incField1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField1).font = Font(size = "8", name='Barlow')

  
            ### 2 Año de diagnóstico de rraa

            count_id_ra = idraFort.count(itemFort.post_ra_id)
            contadorFort=0 #inicializamos el contador  
            posFields = 9*count_id_ra
            
            for incFort in range(1,posFields+1):  
                if incFort%9 == 0:
                    incFort = incFort - 1
                    contadorFort+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incFort).value = "Año de diagnóstico de RRAA"
                    sheet5.cell(row = 3, column = incFort).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incFort).font = Font(bold=True) 

                    if itemFort.year_diagnostico == None:

                        sheet5.cell(row = posiRow, column = incFort).value = ""
                        sheet5.cell(row = posiRow, column = incFort).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incFort).font = Font(size = "8", name='Barlow')
                    else:
                        sheet5.cell(row = posiRow, column = incFort).value = itemFort.year_diagnostico.strftime('%Y')
                        sheet5.cell(row = posiRow, column = incFort).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incFort).font = Font(size = "8", name='Barlow')
            
            ### 3 Módulo o sección diagnosticado:

            contadorField3=0 #inicializamos el contador  
            posFields3 = 9*count_id_ra
            for incField3 in range(1,posFields3+1):  
                if incField3%9 == 0:  
                    incField3 = incField3
                    contadorField3+=1
                
            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField3).value = "Módulo o sección diagnosticado:"
                    sheet5.cell(row = 3, column = incField3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField3).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField3).value = str(itemFort.mod_sec_diagn)
                    sheet5.cell(row = posiRow, column = incField3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField3).font = Font(size = "8", name='Barlow')

            ### 4 Plan de fortalecimiento aprobado por la entidad:

            contadorField4=0 #inicializamos el contador  
            posFields4 = 9*count_id_ra
            for incField4 in range(1,posFields4+1):  
                if incField4%9 == 0:
                    incField4 = incField4 + 1
                    contadorField4+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField4).value = "Plan de fortalecimiento aprobado por la entidad:"
                    sheet5.cell(row = 3, column = incField4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField4).font = Font(bold=True)

                    if str(itemFort.plan_fort_aprob) == "True":
                        sheet5.cell(row = posiRow, column = incField4).value = "Si"
                        sheet5.cell(row = posiRow, column = incField4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField4).font = Font(size = "8", name='Barlow')

                    elif str(itemFort.plan_fort_aprob) == "False":
                        sheet5.cell(row = posiRow, column = incField4).value = "No"
                        sheet5.cell(row = posiRow, column = incField4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField4).font = Font(size = "8", name='Barlow')
            
        ### 5 Fecha de aprobación del Plan de Fortalecimiento

            contadorField5=0 #inicializamos el contador  
            posFields5 = 9*count_id_ra
            for incField5 in range(1,posFields5+1):  
                if incField5%9 == 0:  
                    incField5 = incField5 + 2
                    contadorField5+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField5).value = "Fecha de aprobación del Plan de Fortalecimiento"
                    sheet5.cell(row = 3, column = incField5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField5).font = Font(bold=True)

                    if itemFort.fecha_aprobacion == None:

                        sheet5.cell(row = posiRow, column = incField5).value = ""
                        sheet5.cell(row = posiRow, column = incField5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField5).font = Font(size = "8", name='Barlow')
                    else:
                        sheet5.cell(row = posiRow, column = incField5).value = itemFort.fecha_aprobacion
                        sheet5.cell(row = posiRow, column = incField5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField5).font = Font(size = "8", name='Barlow')
                    
        ### 6 Seguimiento a la implementación del Plan de fortalecimiento:

            contadorField6=0 #inicializamos el contador  
            posFields6 = 9*count_id_ra
            for incField6 in range(1,posFields6+1):  
                if incField6%9 == 0:  
                    incField6 = incField6 + 3
                    contadorField6+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField6).value = "Seguimiento a la implementación del Plan de fortalecimiento"
                    sheet5.cell(row = 3, column = incField6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField6).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField6).value = str(itemFort.seg_imple_plan)
                    sheet5.cell(row = posiRow, column = incField6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField6).font = Font(size = "8", name='Barlow')

        ### 7 Fecha de inicio de la implementación del Plan de Fortaleciemiento

            contadorField7=0 #inicializamos el contador  
            posFields7 = 9*count_id_ra
            for incField7 in range(1,posFields7+1):  
                if incField7%9 == 0:  
                    incField7 = incField7 + 4
                    contadorField7+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField7).value = "Fecha de inicio de la implementación del Plan de Fortaleciemiento"
                    sheet5.cell(row = 3, column = incField7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField7).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField7).value = str(itemFort.fecha_inicio_plan)
                    sheet5.cell(row = posiRow, column = incField7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField7).font = Font(size = "8", name='Barlow')

         ### 8 Fecha de último seguimiento a la implementación del Plan de Fortalecimiento   
            
            contadorField8=0 #inicializamos el contador  
            posFields8 = 9*count_id_ra
            for incField8 in range(1,posFields8+1):  
                if incField8%9 == 0:  
                    incField8 = incField8 + 5
                    contadorField8+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField8).value = "Fecha de último seguimiento a la implementación del Plan de Fortalecimiento "
                    sheet5.cell(row = 3, column = incField8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField8).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField8).value = str(itemFort.fecha_ultimo_seguimiento)
                    sheet5.cell(row = posiRow, column = incField8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField8).font = Font(size = "8", name='Barlow')

        ###  9 Fecha de finalización de la implementación del Plan de fortalecimiento
            
            contadorField9=0 #inicializamos el contador  
            posFields9 = 9*count_id_ra
            for incField9 in range(1,posFields9+1):  
                if incField9%9 == 0:  
                    incField9 = incField9 + 6
                    contadorField9+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField9).value = "Fecha de finalización de la implementación del Plan de fortalecimiento"
                    sheet5.cell(row = 3, column = incField9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField9).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField9).value = str(itemFort.fecha_finalizacion)
                    sheet5.cell(row = posiRow, column = incField9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField9).font = Font(size = "8", name='Barlow')

       
        for row in range(4, sheet5.max_row + 1):
            sheet5.row_dimensions[row].height = 90

        ############# END Hoja 5 #######################

        ## ocultar información si no esta autenticado
        if user.is_authenticated == True and user.profile.role.id != 2:
        
            ############# Hoja 6 Criticas #######################

            sheet6.merge_cells('A1:F1')
            sheet6.merge_cells('A2:F2')

            def set_border(sheet6, cell_range):
                border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                rows = sheet6[cell_range]
                for row in rows:
                    for cell in row:
                        cell.border = border

            set_border(sheet6,'A1:F1')
            
            #row dimensions
            sheet6.row_dimensions[1].height = 55
            sheet6.row_dimensions[2].height = 40
            sheet6.row_dimensions[3].height = 40
            sheet6.row_dimensions[4].height = 40
            sheet6.row_dimensions[5].height = 40
            sheet6.row_dimensions[6].height = 40
            

            # column width
            sheet6.column_dimensions['A'].width = 20
            sheet6.column_dimensions['B'].width = 20
            sheet6.column_dimensions['C'].width = 20
            sheet6.column_dimensions['D'].width = 20
            sheet6.column_dimensions['E'].width = 20
            sheet6.column_dimensions['F'].width = 20
        

            title_cell = sheet6['A1']
            title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
            title_cell.value = 'Directorio de los Registros Administrativos'
            title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
            title_cell.alignment =  Alignment(horizontal='center', vertical='center')

            codigo_cell = sheet6['A2']
            codigo_cell.value = 'Críticas'
            codigo_cell.font = Font(bold=True)
            codigo_cell.alignment = Alignment(horizontal='center')
            codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')


            resultados_cell = sheet6['A3']
            resultados_cell.value = 'Código'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet6['B3']
            resultados_cell.value = 'Nombre del Registro Administrativo'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
                
            resultados_cell = sheet6['C3']
            resultados_cell.value = 'Estado de la crítica'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet6['D3']
            resultados_cell.value = 'Observaciones de la crítica'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet6['E3']
            resultados_cell.value = 'Funcionario que realiza la crítica'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet6['F3']
            resultados_cell.value = 'Fecha en que se realiza la crítica'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            idraCritica = []
            contsheet6 = 4
            for indexCritica, itemCritica in enumerate(criticaTextList):
                indexCritica =  itemCritica.post_ra_id + 3
                idraCritica.append(itemCritica.post_ra_id)

                sheet6.cell(row = contsheet6, column = 1).value = itemCritica.post_ra.codigo_rraa
                sheet6.cell(row = contsheet6, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet6.cell(row = contsheet6, column = 1).font = Font(bold=True) 

                sheet6.cell(row = contsheet6, column = 2).value = str(itemCritica.post_ra)
                sheet6.cell(row = contsheet6, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet6.cell(row = contsheet6, column = 2).font = Font(bold=True) 

                sheet6.cell(row = contsheet6, column = 3).value = str(itemCritica.estado_critica_ra)
                sheet6.cell(row = contsheet6, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet6.cell(row = contsheet6, column = 3).font = Font(bold=True)

                sheet6.cell(row = contsheet6, column = 4).value = str(itemCritica.observa_critica)
                sheet6.cell(row = contsheet6, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet6.cell(row = contsheet6, column = 4).font = Font(bold=True)

                sheet6.cell(row = contsheet6, column = 5).value = str(itemCritica.user_critico)
                sheet6.cell(row = contsheet6, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet6.cell(row = contsheet6, column = 5).font = Font(bold=True)

                sheet6.cell(row = contsheet6, column = 6).value = str(itemCritica.fecha_critica)
                sheet6.cell(row = contsheet6, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet6.cell(row = contsheet6, column = 6).font = Font(bold=True)

                contsheet6+=1

            for row in range(4, sheet6.max_row + 1):  ## definir tamaño de rows
                sheet6.row_dimensions[row].height = 90

            ############# END Hoja 6 Criticas #######################

            ############# Hoja 7 Novedad #######################

            sheet7.merge_cells('A1:H1')
            sheet7.merge_cells('A2:H2')

            def set_border(sheet7, cell_range):
                border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                rows = sheet7[cell_range]
                for row in rows:
                    for cell in row:
                        cell.border = border

            set_border(sheet7,'A1:H1')
            

            #row dimensions
            sheet7.row_dimensions[1].height = 55
            sheet7.row_dimensions[2].height = 40
            sheet7.row_dimensions[3].height = 40
            sheet7.row_dimensions[4].height = 40
            sheet7.row_dimensions[5].height = 40
            sheet7.row_dimensions[6].height = 40
            sheet7.row_dimensions[7].height = 40
            sheet7.row_dimensions[8].height = 40

            # column width
            sheet7.column_dimensions['A'].width = 20
            sheet7.column_dimensions['B'].width = 20
            sheet7.column_dimensions['C'].width = 20
            sheet7.column_dimensions['D'].width = 20
            sheet7.column_dimensions['E'].width = 20
            sheet7.column_dimensions['F'].width = 20
            sheet7.column_dimensions['G'].width = 20
            sheet7.column_dimensions['H'].width = 20

            title_cell = sheet7['A1']
            title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
            title_cell.value = 'Directorio de los Registros Administrativos'
            title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
            title_cell.alignment =  Alignment(horizontal='center', vertical='center')

            codigo_cell = sheet7['A2']
            codigo_cell.value = 'Novedades'
            codigo_cell.font = Font(bold=True)
            codigo_cell.alignment = Alignment(horizontal='center')
            codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')


            resultados_cell = sheet7['A3']
            resultados_cell.value = 'Código'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet7['B3']
            resultados_cell.value = 'Nombre del  Registro Administrativo'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet7['C3']
            resultados_cell.value = 'Novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet7['D3']
            resultados_cell.value = 'Estado'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet7['E3']
            resultados_cell.value = 'Descripción de la novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet7['F3']
            resultados_cell.value = 'Observaciones de la novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet7['G3']
            resultados_cell.value = 'Funcionario que realiza la novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet7['H3']
            resultados_cell.value = 'Fecha en que se realiza la novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            idraNovedad = []
            contsheet7 = 4
            for indexNovedad, itemNovedad in enumerate(novedadTextList):
                indexNovedad =  itemNovedad.post_ra_id + 3
                idraNovedad.append(itemNovedad.post_ra_id)

                sheet7.cell(row = contsheet7, column = 1).value = itemNovedad.post_ra.codigo_rraa
                sheet7.cell(row = contsheet7, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 1).font = Font(bold=True) 

                sheet7.cell(row = contsheet7, column = 2).value = str(itemNovedad.post_ra)
                sheet7.cell(row = contsheet7, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 2).font = Font(bold=True) 

                sheet7.cell(row = contsheet7, column = 3).value = str(itemNovedad.novedad)
                sheet7.cell(row = contsheet7, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 3).font = Font(bold=True)

                sheet7.cell(row = contsheet7, column = 4).value = str(itemNovedad.est_actualiz)
                sheet7.cell(row = contsheet7, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 4).font = Font(bold=True)

                sheet7.cell(row = contsheet7, column = 5).value = str(itemNovedad.descrip_novedad)
                sheet7.cell(row = contsheet7, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 5).font = Font(bold=True)

                sheet7.cell(row = contsheet7, column = 6).value = str(itemNovedad.obser_novedad)
                sheet7.cell(row = contsheet7, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 6).font = Font(bold=True)

                sheet7.cell(row = contsheet7, column = 7).value = str(itemNovedad.name_nov)
                sheet7.cell(row = contsheet7, column = 7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 7).font = Font(bold=True)

                sheet7.cell(row = contsheet7, column = 8).value = str(itemNovedad.fecha_actualiz)
                sheet7.cell(row = contsheet7, column = 8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 8).font = Font(bold=True)
                contsheet7+=1

            for row in range(4, sheet7.max_row + 1):  ## definir tamaño de rows
                sheet7.row_dimensions[row].height = 90

        ############# END Hoja 7 Criticas #######################

        file_name = "reporte_RRAA.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(file_name)
        response['Content-Disposition'] = content
        wb.save(response)
        return response 

        ######################################## REPORTE RRAA COMPLETO ########################################################
class reportRRAAComp_xls(TemplateView): 
    def get(self, request, *args, **kwargs):
             
        rraas = RegistroAdministrativo.objects.all()
        user = request.user
        
        normaTextList = []
        documentoMetodTextList = []
        variableRecolectadaTextList = []
        conceptosEstandarizadosTextList = []
        clasificacionTextList = []
        recoleccionDatosTextList = []
        FrecuenciaRecoleccionDatoTextList = []
        herramientasUtilProcesaTextList = []
        seguridadInformTextList = []
        frecuenciaAlmacenamientobdTextList = []
        coberturaGeograficaTextList = []
        indicadorResultadoAgregadoTextList = []
        entidadesAccesoRRAATextList = []
        noAccesoDatosTextList = []
        fortalecimientoTextList = []
        criticaTextList = [] #critica
        novedadTextList = [] #novedad
        
        objArray = list(rraas.values('id'))
        for obj in objArray:
            for key, value in obj.items():
                #print("value RRAA", value)
                normaQuestionList = MB_NormaRRAA.objects.filter(rraa_id=value)
                unionListNorma = list(chain(normaQuestionList))
                normaTextList.extend(unionListNorma)

                documentoMetodQuestionList = MB_DocumentoMetodologRRAA.objects.filter(rraa_id=value)
                unionListDocumentoMetod = list(chain(documentoMetodQuestionList))
                documentoMetodTextList.extend(unionListDocumentoMetod)

                variabRecolQuestionList = MB_VariableRecolectada.objects.filter(rraa_id=value)  # formset 7
                unionListVariabRecol = list(chain(variabRecolQuestionList))
                variableRecolectadaTextList.extend(unionListVariabRecol)

                concepEstandQuestionList = MB_ConceptosEstandarizadosRRAA.objects.filter(rraa_id=value)
                unionListConcepEstand = list(chain(concepEstandQuestionList))
                conceptosEstandarizadosTextList.extend(unionListConcepEstand)

                clasificacionQuestionList = MB_ClasificacionesRRAA.objects.filter(rraa_id=value)
                unionListClasificacion = list(chain(clasificacionQuestionList))
                clasificacionTextList.extend(unionListClasificacion)

                recolecDatoQuestionList = MB_RecoleccionDatosRRAA.objects.filter(rraa_id=value)
                unionListRecolecDato = list(chain(recolecDatoQuestionList))
                recoleccionDatosTextList.extend(unionListRecolecDato)

                frecRecolDatoQuestionList = MB_FrecuenciaRecoleccionDato.objects.filter(rraa_id=value)
                unionListFrecRecolDato = list(chain(frecRecolDatoQuestionList))
                FrecuenciaRecoleccionDatoTextList.extend(unionListFrecRecolDato)

                herramUtilProcQuestionList = MB_HerramientasUtilProcesa.objects.filter(rraa_id=value)
                unionListHerramUtilProc = list(chain(herramUtilProcQuestionList))
                herramientasUtilProcesaTextList.extend(unionListHerramUtilProc)

                seguridadInformQuestionList = MB_SeguridadInform.objects.filter(rraa_id=value)
                unionListSeguridadInform = list(chain(seguridadInformQuestionList))
                seguridadInformTextList.extend(unionListSeguridadInform)

                frecAlmacenQuestionList = MB_FrecuenciaAlmacenamientobd.objects.filter(rraa_id=value)
                unionListFrecAlmacen = list(chain(frecAlmacenQuestionList))
                frecuenciaAlmacenamientobdTextList.extend(unionListFrecAlmacen)

                coberGeogQuestionList = MB_CoberturaGeograficaRRAA.objects.filter(rraa_id=value)
                unionListCoberGeog = list(chain(coberGeogQuestionList))
                coberturaGeograficaTextList.extend(unionListCoberGeog)

                indResAgreQuestionList = MB_IndicadorResultadoAgregado.objects.filter(rraa_id=value)  ##formset 16
                unionListIndResAgre = list(chain(indResAgreQuestionList))
                indicadorResultadoAgregadoTextList.extend(unionListIndResAgre)

                entidadesAcceQuestionList = MB_EntidadesAccesoRRAA.objects.filter(rraa_id=value)  ##formset 18
                unionListEntidadesAcce = list(chain(entidadesAcceQuestionList))
                entidadesAccesoRRAATextList.extend(unionListEntidadesAcce)

                noAccesoDatosQuestionList = MB_NoAccesoDatos.objects.filter(rraa_id=value)
                unionListNoAccesoDatos = list(chain(noAccesoDatosQuestionList))
                noAccesoDatosTextList.extend(unionListNoAccesoDatos)

                fortalecimientoQuestionList = FortalecimientoRRAA.objects.filter(post_ra_id=value)  ## otra hoja
                unionListFortalecimiento = list(chain(fortalecimientoQuestionList))
                fortalecimientoTextList.extend(unionListFortalecimiento)

                criticaQuestionList = CriticaRRAA.objects.filter(post_ra_id=value)  ## critica
                unionListCritica = list(chain(criticaQuestionList))
                criticaTextList.extend(unionListCritica)

                novedadQuestionList = NovedadActualizacionRRAA.objects.filter(post_ra_id=value)  ## NovedadActualizacionRRAA
                unionListNovedad = list(chain(novedadQuestionList))
                novedadTextList.extend(unionListNovedad)

        wb = Workbook()
        ws = wb.active
        ws.title = "Directorio RRAA"
        sheet2 = wb.create_sheet('Variables RRAA')
        sheet3 = wb.create_sheet('Resultados RRAA')
        sheet4 = wb.create_sheet('Entidades RRAA')
        sheet5 = wb.create_sheet('Fortalecimiento RRAA')
        sheet6 = wb.create_sheet('Critica RRAA')
        sheet7 = wb.create_sheet('Novedad RRAA')

        
        def set_border(ws, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = ws[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(ws,'A1:DG'+str(rraas.count()+3))

		## size rows

        ws.row_dimensions[1].height = 55
        ws.row_dimensions[2].height = 40
        ws.row_dimensions[3].height = 80

        ## size column
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 15
        ws.column_dimensions['O'].width = 15
        ws.column_dimensions['P'].width = 15
        ws.column_dimensions['Q'].width = 15
        ws.column_dimensions['R'].width = 22
        ws.column_dimensions['S'].width = 15
        ws.column_dimensions['T'].width = 15
        ws.column_dimensions['U'].width = 15
        ws.column_dimensions['V'].width = 15
        ws.column_dimensions['W'].width = 15
        ws.column_dimensions['X'].width = 15
        ws.column_dimensions['Y'].width = 20
        ws.column_dimensions['Z'].width = 15
        ws.column_dimensions['AA'].width = 15
        ws.column_dimensions['AB'].width = 15
        ws.column_dimensions['AC'].width = 15
        ws.column_dimensions['AD'].width = 15
        ws.column_dimensions['AE'].width = 15
        ws.column_dimensions['AF'].width = 15
        ws.column_dimensions['AG'].width = 15
        ws.column_dimensions['AH'].width = 15
        ws.column_dimensions['AI'].width = 15
        ws.column_dimensions['AJ'].width = 15
        ws.column_dimensions['AK'].width = 15
        ws.column_dimensions['AL'].width = 15
        ws.column_dimensions['AM'].width = 15
        ws.column_dimensions['AN'].width = 15
        ws.column_dimensions['AO'].width = 15
        ws.column_dimensions['AP'].width = 15
        ws.column_dimensions['AQ'].width = 15
        ws.column_dimensions['AR'].width = 15
        ws.column_dimensions['AS'].width = 15
        ws.column_dimensions['AT'].width = 15
        ws.column_dimensions['AU'].width = 15
        ws.column_dimensions['AV'].width = 15
        ws.column_dimensions['AW'].width = 15
        ws.column_dimensions['AX'].width = 15
        ws.column_dimensions['AY'].width = 15
        ws.column_dimensions['AZ'].width = 15
        ws.column_dimensions['BA'].width = 15
        ws.column_dimensions['BB'].width = 15
        ws.column_dimensions['BC'].width = 15
        ws.column_dimensions['BD'].width = 15
        ws.column_dimensions['BE'].width = 15
        ws.column_dimensions['BF'].width = 15
        ws.column_dimensions['BG'].width = 15
        ws.column_dimensions['BH'].width = 15
        ws.column_dimensions['BI'].width = 15
        ws.column_dimensions['BJ'].width = 15
        ws.column_dimensions['BK'].width = 15
        ws.column_dimensions['BL'].width = 15
        ws.column_dimensions['BM'].width = 15
        ws.column_dimensions['BN'].width = 15
        ws.column_dimensions['BO'].width = 15
        ws.column_dimensions['BP'].width = 15
        ws.column_dimensions['BQ'].width = 15
        ws.column_dimensions['BR'].width = 15
        ws.column_dimensions['BS'].width = 15
        ws.column_dimensions['BT'].width = 15
        ws.column_dimensions['BU'].width = 15
        ws.column_dimensions['BV'].width = 15
        ws.column_dimensions['BW'].width = 15
        ws.column_dimensions['BX'].width = 15
        ws.column_dimensions['BY'].width = 15
        ws.column_dimensions['BZ'].width = 15
        ws.column_dimensions['CA'].width = 15
        ws.column_dimensions['CB'].width = 15
        ws.column_dimensions['CC'].width = 15
        ws.column_dimensions['CD'].width = 15
        ws.column_dimensions['CE'].width = 15
        ws.column_dimensions['CF'].width = 15
        ws.column_dimensions['CG'].width = 15
        ws.column_dimensions['CH'].width = 15
        ws.column_dimensions['CI'].width = 15
        ws.column_dimensions['CJ'].width = 15
        ws.column_dimensions['CK'].width = 15
        ws.column_dimensions['CL'].width = 15
        ws.column_dimensions['CM'].width = 15
        ws.column_dimensions['CN'].width = 15
        ws.column_dimensions['CO'].width = 15
        ws.column_dimensions['CP'].width = 15
        ws.column_dimensions['CQ'].width = 15
        ws.column_dimensions['CR'].width = 15
        ws.column_dimensions['CS'].width = 15
        ws.column_dimensions['CT'].width = 15
        ws.column_dimensions['CU'].width = 15
        ws.column_dimensions['CV'].width = 15
        ws.column_dimensions['CW'].width = 15
        ws.column_dimensions['CX'].width = 15
        ws.column_dimensions['CY'].width = 15
        ws.column_dimensions['CZ'].width = 15
        ws.column_dimensions['DA'].width = 15
        ws.column_dimensions['DB'].width = 22
        ws.column_dimensions['DC'].width = 22
        ws.column_dimensions['DD'].width = 22
        ws.column_dimensions['DE'].width = 22
        ws.column_dimensions['DF'].width = 22
        ws.column_dimensions['DG'].width = 25


        title_cell = ws['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de Registros Administrativos'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')
    
        ws.merge_cells('A1:DG1')

        ws.merge_cells('A2:C2')
        ws.merge_cells('D2:E2')
        ws.merge_cells('F2:H2')
        ws.merge_cells('I2:Q2')
        ws.merge_cells('R2:V2')
        ws.merge_cells('W2:W3')
        ws.merge_cells('X2:Y2')
        ws.merge_cells('Z2:AD2')
        ws.merge_cells('AE2:AR2')

        ws.merge_cells('AS2:AV2')
        ws.merge_cells('AW2:BC2')
        ws.merge_cells('BD2:BL2')
        ws.merge_cells('BM2:BU2')
        ws.merge_cells('BV2:BZ2')
        ws.merge_cells('CA2:CH2')
        ws.merge_cells('CI2:CS2')
        ws.merge_cells('CT2:CT3')
        ws.merge_cells('CU2:CU3')
        ws.merge_cells('CV2:DA2')
        ws.merge_cells('DB2:DB3') 

        ws.merge_cells('DC2:DC3') 
        ws.merge_cells('DD2:DD3') 
        ws.merge_cells('DE2:DE3') 
        ws.merge_cells('DF2:DF3') 
        ws.merge_cells('DG2:DG3')

        ## heads

        codigo_cell = ws['A2']
        codigo_cell.value = 'REGISTRO ADMINISTRATIVO'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['D2']
        codigo_cell.value = 'ENTIDAD'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['F2']
        codigo_cell.value = 'ÁREA TEMÁTICA / TEMA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['I2']
        codigo_cell.value = ' A. IDENTIFICACIÓN'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = ws['R2']
        codigo_cell.value = 'Bajo cuál(es) de las siguientes normas, se soporta la creación del RA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['W2']
        codigo_cell.value = 'Describa la razón por la cuál se creó el RA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['X2']
        codigo_cell.value = 'Indique desde y hasta cuando se ha recolectado el RA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['Z2']
        codigo_cell.value = '¿El RA cuenta con alguno de los siguientes documentos metodológicos o funcionales?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AE2']
        codigo_cell.value = 'Indique si el RA utiliza conceptos estandarizados provenientes de'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AS2']
        codigo_cell.value = '¿El RA utiliza nomenclaturas y/o clasificaciones?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AW2']
        codigo_cell.value = '¿Cuál es el medio de obtención o recolección de los datos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BD2']
        codigo_cell.value = '¿Con qué frecuencia se recolectan los datos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['BM2']
        codigo_cell.value = 'Indique cuáles de las siguientes herramientas son utilizadas en el procesamiento de los datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BV2']
        codigo_cell.value = '¿Con cuáles herramientas cuenta para garantizar la seguridad de la información del RA?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CA2']
        codigo_cell.value = '¿La información recolectada es acopiada o almacenada en una base de datos?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CI2']
        codigo_cell.value = '¿Cuál es la cobertura geográfica del RA?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CT2']
        codigo_cell.value = '¿La entidad hace uso de los datos del RA para generar información estadística (Agregados, indicadores, Resultados estadísticos, etc)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CU2']
        codigo_cell.value = '¿Usuarios externos a la entidad, tienen acceso a los datos del RA?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CV2']
        codigo_cell.value = '¿Cuál es la razón principal por la cual no se permite el acceso a los datos del RA?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DB2']
        codigo_cell.value = 'OBSERVACIONES'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DC2']
        codigo_cell.value = 'Registro administrativo activo (SI / NO)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DD2']
        codigo_cell.value = 'Usuario DANE'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DE2']
        codigo_cell.value = 'Responde a requerimientos ODS (SI / NO)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DF2']
        codigo_cell.value = 'Indicador ODS a que da respuesta'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DG2']
        codigo_cell.value = 'Estado del proceso del Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)



        #--------------------------------------------------------------------------------------------
        
        
        codigo_cell = ws['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['B3']
        codigo_cell.value = 'Nombre del Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['C3']
        codigo_cell.value = 'Objetivo del Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['D3']
        codigo_cell.value = 'Nombre de la Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center',  wrap_text=True)
       
        codigo_cell = ws['E3']
        codigo_cell.value = 'Código de la Entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['F3']
        codigo_cell.value = 'Área Tematica'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = ws['G3']
        codigo_cell.value = 'Tema'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['H3']
        codigo_cell.value = 'Tema Compartido'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['I3']
        codigo_cell.value = 'Nombre de la Dependencia'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['J3']
        codigo_cell.value = 'Nombre del Director'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['K3']
        codigo_cell.value = 'Cargo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['L3']
        codigo_cell.value = 'Correo Electrónico del Director'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = ws['M3']
        codigo_cell.value = 'Teléfono del director'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['N3']
        codigo_cell.value = 'Nombre del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['O3']
        codigo_cell.value = 'Cargo del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['P3']
        codigo_cell.value = 'Correo Electrónico del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['Q3']
        codigo_cell.value = 'Teléfono del Temático responsable'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['R3']
        codigo_cell.value = 'a. Constitución Política'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['S3']
        codigo_cell.value = 'b. Ley'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        codigo_cell = ws['T3']
        codigo_cell.value = 'c. Decreto (nacional, departamental, municipal)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['U3']
        codigo_cell.value = 'd. Otra (Resolución, ordenanza, acuerdo)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center',  wrap_text=True)

        codigo_cell = ws['V3']
        codigo_cell.value = 'e. Ninguna'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['X3']
        codigo_cell.value = 'Fecha de inicio de recolección de información'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['Y3']
        codigo_cell.value = 'Fecha de última recolección de información'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['Z3']
        codigo_cell.value = 'a) Ficha técnica'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AA3']
        codigo_cell.value = 'b) Manual de diligenciamiento'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AB3']
        codigo_cell.value = 'c) Diccionario de datos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AC3']
        codigo_cell.value = 'd) Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AD3']
        codigo_cell.value = 'd) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AE3']
        codigo_cell.value = 'a). DANE'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AF3']
        codigo_cell.value = 'b). OCDE'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AG3']
        codigo_cell.value = 'c). ONU'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AH3']
        codigo_cell.value = 'd). EUROSTAT'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AI3']
        codigo_cell.value = 'e). Otro organismo Internacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AJ3']
        codigo_cell.value = 'e) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
  
        codigo_cell = ws['AK3']
        codigo_cell.value = 'f). Otra entidad de orden nacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AL3']
        codigo_cell.value = 'f) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AM3']
        codigo_cell.value = 'g). Leyes, decretos, etc.'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AN3']
        codigo_cell.value = 'g) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AO3']
        codigo_cell.value = 'h). Creación propia'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AP3']
        codigo_cell.value = 'i). Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AQ3']
        codigo_cell.value = 'i) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AR3']
        codigo_cell.value = 'j). No utiliza conceptos estandarizados'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AS3']
        codigo_cell.value = '¿utiliza nomenclaturas y/o clasificaciones?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AT3']
        codigo_cell.value = 'Si: Cuáles?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AU3']
        codigo_cell.value = 'Si: Otras'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


        codigo_cell = ws['AV3']
        codigo_cell.value = 'No: ¿Por qué?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


        codigo_cell = ws['AW3']
        codigo_cell.value = 'a) Formulario físico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['AX3']
        codigo_cell.value = 'b) Formulario electrónico'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


        codigo_cell = ws['AY3']
        codigo_cell.value = 'c) Dispositivo Móvil de Captura [DMC]'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


        codigo_cell = ws['AZ3']
        codigo_cell.value = 'd) Sistema de información'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BA3']
        codigo_cell.value = 'd) Cuáles?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BB3']
        codigo_cell.value = 'e) Imágenes satelitales'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BC3']
        codigo_cell.value = 'f) Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BD3']
        codigo_cell.value = 'a) Anual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BE3']
        codigo_cell.value = 'b) Semestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BF3']
        codigo_cell.value = 'c) Bimensual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BG3']
        codigo_cell.value = 'd) Trimestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BH3']
        codigo_cell.value = 'e) Mensual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BI3']
        codigo_cell.value = 'f) Diaria'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BJ3']
        codigo_cell.value = 'g) Por evento'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BK3']
        codigo_cell.value = 'h) Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BL3']
        codigo_cell.value = 'h) ¿cuál(es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BM3']
        codigo_cell.value = 'a) Excel'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BN3']
        codigo_cell.value = 'b) Access'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BO3']
        codigo_cell.value = 'c) R.'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BP3']
        codigo_cell.value = 'd) SAS.'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BQ3']
        codigo_cell.value = 'e) SPSS.'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BR3']
        codigo_cell.value = 'f) Oracle'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BS3']
        codigo_cell.value = 'g) Stata.'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BT3']
        codigo_cell.value = 'h) Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BU3']
        codigo_cell.value = 'h) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BV3']
        codigo_cell.value = 'a) Backups'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BW3']
        codigo_cell.value = 'b) Aislamiento del servidor'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BX3']
        codigo_cell.value = 'c) Utilización de perfiles'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BY3']
        codigo_cell.value = 'd) Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['BZ3']
        codigo_cell.value = 'd) ¿Cuál?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CA3']
        codigo_cell.value = '¿La información es almacenada en una base de datos? '
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CB3']
        codigo_cell.value = 'a) Anual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CC3']
        codigo_cell.value = 'b) Semestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CD3']
        codigo_cell.value = 'c) Trimestral'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CE3']
        codigo_cell.value = 'd) Mensual'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CF3']
        codigo_cell.value = 'e) Diaria'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CG3']
        codigo_cell.value = 'f) Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CH3']
        codigo_cell.value = 'f) ¿cuál(es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CI3']
        codigo_cell.value = 'a) Nacional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CJ3']
        codigo_cell.value = 'b) Regional'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CK3']
        codigo_cell.value = 'b) ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CL3']
        codigo_cell.value = 'c) Departamental'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CM3']
        codigo_cell.value = 'c) ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CN3']
        codigo_cell.value = 'd) Áreas metropolitanas'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CO3']
        codigo_cell.value = 'd) ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CP3']
        codigo_cell.value = 'e) Municipal'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CQ3']
        codigo_cell.value = 'e) ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CR3']
        codigo_cell.value = 'f) Otro (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CS3']
        codigo_cell.value = 'f) ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CV3']
        codigo_cell.value = 'a) Políticas de la entidad'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CW3']
        codigo_cell.value = 'b) Disposición normativa'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CX3']
        codigo_cell.value = 'c) Acuerdos con el informante'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CY3']
        codigo_cell.value = 'd) Carácter confidencial'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['CZ3']
        codigo_cell.value = 'e) Otra (s)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = ws['DA3']
        codigo_cell.value = 'e) ¿Cuál (es)?'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
 
        cont = 4
           
        for rraa in rraas:
            
            ws.cell(row = cont, column = 1).value = rraa.codigo_rraa
            ws.cell(row = cont, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 1).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 2).value = rraa.nombre_ra
            ws.cell(row = cont, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 2).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 3).value = rraa.objetivo_ra
            ws.cell(row = cont, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 3).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 4).value = str(rraa.entidad_pri)
            ws.cell(row = cont, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 4).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 5).value =  rraa.entidad_pri.codigo
            ws.cell(row = cont, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 5).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 6).value = str(rraa.area_temat)
            ws.cell(row = cont, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 6).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 7).value = str(rraa.tema)
            ws.cell(row = cont, column = 7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 7).font = Font(size = "8", name='Barlow')

            ## TEMA COMPARTIDO

            listaTemaCompartido = rraa.tema_compart.all()
            temaCompartido_array = []
            for indexTemaCompa, itemTemaCompa in enumerate(listaTemaCompartido):
                temaCompartido_array.append(str(itemTemaCompa))
                ws.cell(row = cont, column = 8).value = str(temaCompartido_array)
                ws.cell(row = cont, column = 8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 8).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 9).value = rraa.nom_dep
            ws.cell(row = cont, column = 9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
            ws.cell(row = cont, column = 9).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 10).value = rraa.nom_dir
            ws.cell(row = cont, column = 10).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 10).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 11).value = rraa.car_dir
            ws.cell(row = cont, column = 11).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 11).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 12).value = rraa.cor_dir
            ws.cell(row = cont, column = 12).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 12).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 13).value = rraa.telef_dir
            ws.cell(row = cont, column = 13).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 13).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 14).value = rraa.nom_resp
            ws.cell(row = cont, column = 14).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 14).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 15).value = rraa.carg_resp
            ws.cell(row = cont, column = 15).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 15).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 16).value = rraa.cor_resp
            ws.cell(row = cont, column = 16).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 16).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 17).value = rraa.telef_resp
            ws.cell(row = cont, column = 17).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 17).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 18).value = rraa.pq_secreo
            ws.cell(row = cont, column = 18).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 18).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 23).value = rraa.pq_secreo
            ws.cell(row = cont, column = 23).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 23).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 24).value = rraa.fecha_ini_rec
            ws.cell(row = cont, column = 24).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 24).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 25).value = rraa.fecha_ult_rec
            ws.cell(row = cont, column = 25).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 25).font = Font(size = "8", name='Barlow')

            if str(rraa.clas_s_n) == "True":
                ws.cell(row = cont, column = 45).value = "Si"
                ws.cell(row = cont, column = 45).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 45).font = Font(size = "8", name='Barlow')
            
            elif str(rraa.clas_s_n) == "False":
                ws.cell(row = cont, column = 45).value = "No"
                ws.cell(row = cont, column = 45).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 45).font = Font(size = "8", name='Barlow')

            ##Clasificaciones
            listaClasificaciones = rraa.nomb_cla.all()
            clasificaciones_array = []

            for indexClasificaciones, itemClasificaciones in enumerate(listaClasificaciones):
                clasificaciones_array.append(str(itemClasificaciones))
                ws.cell(row = cont, column = 46).value = str(clasificaciones_array)
                ws.cell(row = cont, column = 46).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 46).font = Font(size = "8", name='Barlow')

            if str(rraa.almacen_bd_s_n) == "True":
                ws.cell(row = cont, column = 79).value = "Si"
                ws.cell(row = cont, column = 79).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 79).font = Font(size = "8", name='Barlow')
            
            elif str(rraa.almacen_bd_s_n) == "False":
                ws.cell(row = cont, column = 79).value = "No"
                ws.cell(row = cont, column = 79).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 79).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 98).value = str(rraa.uso_de_datos)
            ws.cell(row = cont, column = 98).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 98).font = Font(size = "8", name='Barlow')

            if str(rraa.user_exte_acceso) == "True":
                ws.cell(row = cont, column = 99).value = "Si"
                ws.cell(row = cont, column = 99).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 99).font = Font(size = "8", name='Barlow')
            
            elif str(rraa.user_exte_acceso) == "False":
                ws.cell(row = cont, column = 99).value = "No"
                ws.cell(row = cont, column = 99).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 99).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 106).value = rraa.observacion
            ws.cell(row = cont, column = 106).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 106).font = Font(size = "8", name='Barlow')

            if str(rraa.ra_activo) == "True":
                ws.cell(row = cont, column = 107).value = "Si"
                ws.cell(row = cont, column = 107).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 107).font = Font(size = "8", name='Barlow')
            
            elif str(rraa.ra_activo) == "False":
                ws.cell(row = cont, column = 107).value = "No"
                ws.cell(row = cont, column = 107).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 107).font = Font(size = "8", name='Barlow')
            
            ws.cell(row = cont, column = 108).value = rraa.user_dane
            ws.cell(row = cont, column = 108).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 108).font = Font(size = "8", name='Barlow')

            if str(rraa.responde_ods) == "True":
                ws.cell(row = cont, column = 109).value = "Si"
                ws.cell(row = cont, column = 109).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 109).font = Font(size = "8", name='Barlow')
            
            elif str(rraa.responde_ods) == "False":
                ws.cell(row = cont, column = 109).value = "No"
                ws.cell(row = cont, column = 109).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row = cont, column = 109).font = Font(size = "8", name='Barlow')

            ws.cell(row = cont, column = 110).value = rraa.indicador_ods
            ws.cell(row = cont, column = 110).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 110).font = Font(size = "8", name='Barlow')

            
            ws.cell(row = cont, column = 111).value = str(rraa.sist_estado)
            ws.cell(row = cont, column = 111).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = cont, column = 111).font = Font(size = "8", name='Barlow')
            
            ##Si desea continuar despues de la columna 111 _____________
        
            ## opciones de preguntas
            listaNorma = rraa.norma_ra.all()
            listaDocumentoMetod = rraa.doc_met_ra.all()
            listaConcepEstand = rraa.con_est_ra.all()
            listaRecolec = rraa.recole_dato.all()
            listaFrecRecol = rraa.fre_rec_dato.all()
            listaHerramUtil = rraa.herr_u_pro.all()
            listaSeguridadIn = rraa.seg_inf.all()
            listaFrecAlm = rraa.frec_alm_bd.all()
            listaCoberGeo = rraa.cob_geograf.all()
            listaNoAccesoDat = rraa.no_hay_acceso.all()

            for index, item in enumerate(listaNorma):
                indice = index
                
                if  str(item) == 'Ninguna' and index == indice:
                    #print("item", item)
                    ws.cell(row = cont, column = 22).value = str(item)
                    ws.cell(row = cont, column = 22).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 22).font = Font(size = "8", name='Barlow')

            for indexDocMet, itemDocMet in enumerate(listaDocumentoMetod):
        
                if  str(itemDocMet) == 'Ficha técnica':
                   
                    ws.cell(row = cont, column = 26).value = str(itemDocMet)
                    ws.cell(row = cont, column = 26).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 26).font = Font(size = "8", name='Barlow')

                if  str(itemDocMet) == 'Manual de diligenciamiento':

                    ws.cell(row = cont, column = 27).value = str(itemDocMet)
                    ws.cell(row = cont, column = 27).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 27).font = Font(size = "8", name='Barlow')
                
                if str(itemDocMet) == 'Diccionario de datos':

                    ws.cell(row = cont, column = 28).value = str(itemDocMet)
                    ws.cell(row = cont, column = 28).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 28).font = Font(size = "8", name='Barlow')

                if str(itemDocMet) == 'Otro (s)':

                    ws.cell(row = cont, column = 29).value = str(itemDocMet)
                    ws.cell(row = cont, column = 29).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 29).font = Font(size = "8", name='Barlow')

            for itemConcepEstand, indexConcepEstand in enumerate(listaConcepEstand):
                
                if str(itemConcepEstand) == 'DANE':

                    ws.cell(row = cont, column = 31).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 31).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 31).font = Font(size = "8", name='Barlow')

                if str(itemConcepEstand) == 'OCDE':
                    ws.cell(row = cont, column = 32).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 32).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 32).font = Font(size = "8", name='Barlow')

                if str(itemConcepEstand) == 'ONU':
                    ws.cell(row = cont, column = 33).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 33).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 33).font = Font(size = "8", name='Barlow')

                if str(itemConcepEstand) == 'EUROSTAT':
                    ws.cell(row = cont, column = 34).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 34).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 34).font = Font(size = "8", name='Barlow')

                if str(itemConcepEstand) == 'Otro organismo Internacional':
                    ws.cell(row = cont, column = 35).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 35).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 35).font = Font(size = "8", name='Barlow')

                if str(itemConcepEstand) == 'Otra entidad de orden nacional':
                    ws.cell(row = cont, column = 37).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 37).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 37).font = Font(size = "8", name='Barlow')
                
                if str(itemConcepEstand) == 'Leyes, decretos, etc':
                    ws.cell(row = cont, column = 39).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 39).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 39).font = Font(size = "8", name='Barlow')   
                
                if str(itemConcepEstand) == 'Creación propia':
                    ws.cell(row = cont, column = 41).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 41).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 41).font = Font(size = "8", name='Barlow')
                
                if str(itemConcepEstand) == 'Otra (s)':
                    ws.cell(row = cont, column = 42).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 42).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 42).font = Font(size = "8", name='Barlow')
                
                if str(itemConcepEstand) == 'No utiliza conceptos estandarizados':
                    ws.cell(row = cont, column = 44).value = str(itemConcepEstand)
                    ws.cell(row = cont, column = 44).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 44).font = Font(size = "8", name='Barlow')

            
            for indexRecolec, itemRecolec in enumerate(listaRecolec):

                if str(itemRecolec) == 'Formulario físico':
                    ws.cell(row = cont, column = 49).value = str(itemRecolec)
                    ws.cell(row = cont, column = 49).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 49).font = Font(size = "8", name='Barlow')

                if str(itemRecolec) == 'Formulario electrónico':
                    ws.cell(row = cont, column = 50).value = str(itemRecolec)
                    ws.cell(row = cont, column = 50).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 50).font = Font(size = "8", name='Barlow')

                if str(itemRecolec) == 'Dispositivo Móvil de Captura [DMC]':
                    ws.cell(row = cont, column = 51).value = str(itemRecolec)
                    ws.cell(row = cont, column = 51).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 51).font = Font(size = "8", name='Barlow')

                if str(itemRecolec) == 'Sistema de información':
                    ws.cell(row = cont, column = 52).value = str(itemRecolec)
                    ws.cell(row = cont, column = 52).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 52).font = Font(size = "8", name='Barlow')

                if str(itemRecolec) == 'Imágenes satelitales':
                    ws.cell(row = cont, column = 54).value = str(itemRecolec)
                    ws.cell(row = cont, column = 54).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 54).font = Font(size = "8", name='Barlow')

            for indexFrecRecol, itemFrecRecol in enumerate(listaFrecRecol):

                if str(itemFrecRecol) == 'Anual':
                    ws.cell(row = cont, column = 56).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 56).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 56).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Semestral':
                    ws.cell(row = cont, column = 57).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 57).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 57).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Bimensual':
                    ws.cell(row = cont, column = 58).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 58).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 58).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Trimestral':
                    ws.cell(row = cont, column = 59).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 59).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 59).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Mensual':
                    ws.cell(row = cont, column = 60).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 60).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 60).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Diaria':

                    ws.cell(row = cont, column = 61).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 61).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 61).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Por evento':

                    ws.cell(row = cont, column = 62).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 62).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 62).font = Font(size = "8", name='Barlow')

                if str(itemFrecRecol) == 'Otra (s)':

                    ws.cell(row = cont, column = 63).value = str(itemFrecRecol)
                    ws.cell(row = cont, column = 63).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 63).font = Font(size = "8", name='Barlow')

            
            for indexHerramUtil, itemHerramUtil in enumerate(listaHerramUtil):

                if str(itemHerramUtil) == 'Excel':
                    ws.cell(row = cont, column = 65).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 65).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 65).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'Access':
                    ws.cell(row = cont, column = 66).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 66).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 66).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'R':
                    ws.cell(row = cont, column = 67).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 67).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 67).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'SAS':
                    ws.cell(row = cont, column = 68).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 68).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 68).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'SPSS':
                    ws.cell(row = cont, column = 69).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 69).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 69).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'Oracle':
                    ws.cell(row = cont, column = 70).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 70).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 70).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'Stata':
                    ws.cell(row = cont, column = 71).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 71).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 71).font = Font(size = "8", name='Barlow')

                if str(itemHerramUtil) == 'Otra (s)':
                    ws.cell(row = cont, column = 72).value = str(itemHerramUtil)
                    ws.cell(row = cont, column = 72).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 72).font = Font(size = "8", name='Barlow')

            for indexSeguridadIn, itemSeguridadIn in enumerate(listaSeguridadIn):

                if str(itemSeguridadIn) == 'Backups':

                    ws.cell(row = cont, column = 74).value = str(itemSeguridadIn)
                    ws.cell(row = cont, column = 74).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 74).font = Font(size = "8", name='Barlow')

                if str(itemSeguridadIn) == 'Aislamiento del servidor':

                    ws.cell(row = cont, column = 75).value = str(itemSeguridadIn)
                    ws.cell(row = cont, column = 75).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 75).font = Font(size = "8", name='Barlow')

                if str(itemSeguridadIn) == 'Utilización de perfiles':

                    ws.cell(row = cont, column = 76).value = str(itemSeguridadIn)
                    ws.cell(row = cont, column = 76).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 76).font = Font(size = "8", name='Barlow')

                if str(itemSeguridadIn) == 'Otra (s)':

                    ws.cell(row = cont, column = 77).value = str(itemSeguridadIn)
                    ws.cell(row = cont, column = 77).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 77).font = Font(size = "8", name='Barlow')

            for indexFrecAlm, itemFrecAlm in enumerate(listaFrecAlm):

                if str(itemFrecAlm) == 'Anual':
                    ws.cell(row = cont, column = 80).value = str(itemFrecAlm)
                    ws.cell(row = cont, column = 80).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 80).font = Font(size = "8", name='Barlow')

                if str(itemFrecAlm) == 'Semestral':
                    ws.cell(row = cont, column = 81).value = str(itemFrecAlm)
                    ws.cell(row = cont, column = 81).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 81).font = Font(size = "8", name='Barlow')

                if str(itemFrecAlm) == 'Trimestral':
                    ws.cell(row = cont, column = 82).value = str(itemFrecAlm)
                    ws.cell(row = cont, column = 82).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 82).font = Font(size = "8", name='Barlow')

                if str(itemFrecAlm) == 'Mensual':
                    ws.cell(row = cont, column = 83).value = str(itemFrecAlm)
                    ws.cell(row = cont, column = 83).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 83).font = Font(size = "8", name='Barlow')

                if str(itemFrecAlm) == 'Diaria':
                    ws.cell(row = cont, column = 84).value = str(itemFrecAlm)
                    ws.cell(row = cont, column = 84).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 84).font = Font(size = "8", name='Barlow')

                if str(itemFrecAlm) == 'Otra (s)':
                    ws.cell(row = cont, column = 85).value = str(itemFrecAlm)
                    ws.cell(row = cont, column = 85).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 85).font = Font(size = "8", name='Barlow')

            for indexCoberGeo, itemCoberGeo in enumerate(listaCoberGeo):

                if str(itemCoberGeo) == 'Nacional':
                    ws.cell(row = cont, column = 87).value = str(itemCoberGeo)
                    ws.cell(row = cont, column = 87).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 87).font = Font(size = "8", name='Barlow')

                if str(itemCoberGeo) == 'Regional':

                    ws.cell(row = cont, column = 88).value = str(itemCoberGeo)
                    ws.cell(row = cont, column = 88).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 88).font = Font(size = "8", name='Barlow')

                if str(itemCoberGeo) == 'Departamental':

                    ws.cell(row = cont, column = 90).value = str(itemCoberGeo)
                    ws.cell(row = cont, column = 90).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 90).font = Font(size = "8", name='Barlow')

                if str(itemCoberGeo) == 'Áreas metropolitanas':

                    ws.cell(row = cont, column = 92).value = str(itemCoberGeo)
                    ws.cell(row = cont, column = 92).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 92).font = Font(size = "8", name='Barlow')

                if str(itemCoberGeo) == 'Municipal':

                    ws.cell(row = cont, column = 94).value = str(itemCoberGeo)
                    ws.cell(row = cont, column = 94).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 94).font = Font(size = "8", name='Barlow')

                if str(itemCoberGeo) == 'Otro (s)':

                    ws.cell(row = cont, column = 96).value = str(itemCoberGeo)
                    ws.cell(row = cont, column = 96).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 96).font = Font(size = "8", name='Barlow')

            for indexNoAccesoDat, itemNoAccesoDat in enumerate(listaNoAccesoDat):

                if str(itemNoAccesoDat) == 'Políticas de la entidad':

                    ws.cell(row = cont, column = 100).value = str(itemNoAccesoDat)
                    ws.cell(row = cont, column = 100).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 100).font = Font(size = "8", name='Barlow')

                if str(itemNoAccesoDat) == 'Disposición normativa':

                    ws.cell(row = cont, column = 101).value = str(itemNoAccesoDat)
                    ws.cell(row = cont, column = 101).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 101).font = Font(size = "8", name='Barlow')

                if str(itemNoAccesoDat) == 'Acuerdos con el informante':

                    ws.cell(row = cont, column = 102).value = str(itemNoAccesoDat)
                    ws.cell(row = cont, column = 102).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 102).font = Font(size = "8", name='Barlow')

                if str(itemNoAccesoDat) == 'Carácter confidencial':

                    ws.cell(row = cont, column = 103).value = str(itemNoAccesoDat)
                    ws.cell(row = cont, column = 103).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 103).font = Font(size = "8", name='Barlow')

                if str(itemNoAccesoDat) == 'Otra (s)':

                    ws.cell(row = cont, column = 104).value = str(itemNoAccesoDat)
                    ws.cell(row = cont, column = 104).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row = cont, column = 104).font = Font(size = "8", name='Barlow')

          
            cont+=1

        ## campos de texto   

        #print("---------------------->", normaTextList)
        for indexNorma, itemNorma in enumerate(normaTextList):
           
            indexNorma =  indexNorma + 4
            ws.cell(row = indexNorma, column = 18).value = str(itemNorma.cp_ra)
            ws.cell(row = indexNorma, column = 18).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 18).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexNorma, column = 19).value = str(itemNorma.ley_ra)
            ws.cell(row = indexNorma, column = 19).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 19).font = Font(size = "8", name='Barlow')
                    
            ws.cell(row = indexNorma, column = 20).value = str(itemNorma.decreto_ra)
            ws.cell(row = indexNorma, column = 20).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 20).font = Font(size = "8", name='Barlow')
                    
            ws.cell(row = indexNorma, column = 21).value = str(itemNorma.otra_ra)
            ws.cell(row = indexNorma, column = 21).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNorma, column = 21).font = Font(size = "8", name='Barlow')


        for indexDocumentoMetod, itemDocumentoMetod in enumerate(documentoMetodTextList):
           
            indexDocumentoMetod =  indexDocumentoMetod + 4
            ws.cell(row = indexDocumentoMetod, column = 30).value = str(itemDocumentoMetod.otra_doc_cual)
            ws.cell(row = indexDocumentoMetod, column = 30).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexDocumentoMetod, column = 30).font = Font(size = "8", name='Barlow')


        for indexConEst, itemConEst in enumerate(conceptosEstandarizadosTextList):
           
            indexConEst =  indexConEst + 4

            ws.cell(row = indexConEst, column = 36).value = str(itemConEst.org_in_cual)
            ws.cell(row = indexConEst, column = 36).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConEst, column = 36).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexConEst, column = 38).value = str(itemConEst.ent_ordnac_cual)
            ws.cell(row = indexConEst, column = 38).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConEst, column = 38).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexConEst, column = 40).value = str(itemConEst.leye_dec_cual)
            ws.cell(row = indexConEst, column = 40).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConEst, column = 40).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexConEst, column = 43).value = str(itemConEst.otra_ce_cual)
            ws.cell(row = indexConEst, column = 43).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexConEst, column = 43).font = Font(size = "8", name='Barlow')
        

        for indexClasifi, itemClasifi in enumerate(clasificacionTextList):
           
            indexClasifi =  indexClasifi + 4

            ws.cell(row = indexClasifi, column = 47).value = str(itemClasifi.otra_cual_clas)
            ws.cell(row = indexClasifi, column = 47).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexClasifi, column = 47).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexClasifi, column = 48).value = str(itemClasifi.no_pq)
            ws.cell(row = indexClasifi, column = 48).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexClasifi, column = 48).font = Font(size = "8", name='Barlow')

        
        for indexRecoleccionDatos, itemRecoleccionDatos in enumerate(recoleccionDatosTextList):

            indexRecoleccionDatos = indexRecoleccionDatos + 4

            ws.cell(row = indexRecoleccionDatos, column = 53).value = str(itemRecoleccionDatos.sistema_inf_cual)
            ws.cell(row = indexRecoleccionDatos, column = 53).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexRecoleccionDatos, column = 53).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexRecoleccionDatos, column = 55).value = str(itemRecoleccionDatos.otro_c)
            ws.cell(row = indexRecoleccionDatos, column = 55).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexRecoleccionDatos, column = 55).font = Font(size = "8", name='Barlow')

        for indexFrecuenciaRecoleccionDato, itemFrecuenciaRecoleccionDato in enumerate(FrecuenciaRecoleccionDatoTextList):
            
            indexFrecuenciaRecoleccionDato = indexFrecuenciaRecoleccionDato + 4

            ws.cell(row = indexFrecuenciaRecoleccionDato, column = 64).value = str(itemFrecuenciaRecoleccionDato.otra_cual_fre)
            ws.cell(row = indexFrecuenciaRecoleccionDato, column = 64).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexFrecuenciaRecoleccionDato, column = 64).font = Font(size = "8", name='Barlow')

         
        for indexherramientasUtilPro, itemherramientasUtilPro in enumerate(herramientasUtilProcesaTextList):
            
            indexherramientasUtilPro = indexherramientasUtilPro + 4

            ws.cell(row = indexherramientasUtilPro, column = 73).value = str(itemherramientasUtilPro.otra_herram)
            ws.cell(row = indexherramientasUtilPro, column = 73).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexherramientasUtilPro, column = 73).font = Font(size = "8", name='Barlow')

        
        for indexSeguridadInfo, itemSeguridadInfo in enumerate(seguridadInformTextList):

            indexSeguridadInfo = indexSeguridadInfo + 4

            ws.cell(row = indexSeguridadInfo, column = 78).value = str(itemSeguridadInfo.otra_cual_s)
            ws.cell(row = indexSeguridadInfo, column = 78).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexSeguridadInfo, column = 78).font = Font(size = "8", name='Barlow')

        for indexfrecuenciaAlmacen, itemfrecuenciaAlmacen in enumerate(frecuenciaAlmacenamientobdTextList):

            indexfrecuenciaAlmacen = indexfrecuenciaAlmacen + 4

            ws.cell(row = indexfrecuenciaAlmacen, column = 86).value = str(itemfrecuenciaAlmacen.otra_alm_bd)
            ws.cell(row = indexfrecuenciaAlmacen, column = 86).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexfrecuenciaAlmacen, column = 86).font = Font(size = "8", name='Barlow')

        for indexCoberturaGeog, itemCoberturaGeog in enumerate(coberturaGeograficaTextList):
            
            indexCoberturaGeog = indexCoberturaGeog + 4

            ws.cell(row = indexCoberturaGeog, column = 89).value = str(itemCoberturaGeog.cual_regio)
            ws.cell(row = indexCoberturaGeog, column = 89).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 89).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 91).value = str(itemCoberturaGeog.cual_depa)
            ws.cell(row = indexCoberturaGeog, column = 91).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 91).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 93).value = str(itemCoberturaGeog.cual_are_metrop)
            ws.cell(row = indexCoberturaGeog, column = 93).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 93).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 95).value = str(itemCoberturaGeog.cual_munic)
            ws.cell(row = indexCoberturaGeog, column = 95).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 95).font = Font(size = "8", name='Barlow')

            ws.cell(row = indexCoberturaGeog, column = 97).value = str(itemCoberturaGeog.cual_otro)
            ws.cell(row = indexCoberturaGeog, column = 97).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexCoberturaGeog, column = 97).font = Font(size = "8", name='Barlow')

        for indexNoAccesoDatos, itemNoAccesoDatos in enumerate(noAccesoDatosTextList):

            indexNoAccesoDatos = indexNoAccesoDatos + 4

            ws.cell(row = indexNoAccesoDatos, column = 105).value = str(itemNoAccesoDatos.otra_no_acceso)
            ws.cell(row = indexNoAccesoDatos, column = 105).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row = indexNoAccesoDatos, column = 105).font = Font(size = "8", name='Barlow')

        for row in range(4, ws.max_row + 1):
            ws.row_dimensions[row].height = 90


       ####################### HOJA 2 ######################## 

        sheet2.merge_cells('A1:C1')
        sheet2.merge_cells('A2:C2')

        def set_border(sheet2, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet2[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet2,'A1:C1')
    
        #row dimensions
        sheet2.row_dimensions[1].height = 55
        sheet2.row_dimensions[2].height = 40

        # column width
        sheet2.column_dimensions['A'].width = 20
        sheet2.column_dimensions['B'].width = 20
        sheet2.column_dimensions['C'].width = 20
        
        title_cell = sheet2['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de los Registros Administrativos'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        title_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
	
        codigo_cell = sheet2['A2']
        codigo_cell.value = 'Lista de variables de los Registros Administrativos'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)


        codigo_cell = sheet2['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = sheet2['B3']
        codigo_cell.value = 'Nombre Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet2['C3']
        codigo_cell.value = 'Variables que maneja el Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        contsheet2 = 4

        arrayraId = []
        for indexVariab, itemVariab in enumerate(variableRecolectadaTextList):
            indexVariab =  itemVariab.rraa_id + 3
            arrayraId.append(itemVariab.rraa_id)
            for radm in rraas:
                if str(itemVariab.rraa) == str(radm.nombre_ra):
                    sheet2.cell(row = contsheet2, column = 1).value = radm.codigo_rraa
                    sheet2.cell(row = contsheet2, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = contsheet2, column = 1).font = Font(size = "8", name='Barlow')

                    sheet2.cell(row = contsheet2, column = 2).value = str(itemVariab.rraa)
                    sheet2.cell(row = contsheet2, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = contsheet2, column = 2).font = Font(size = "8", name='Barlow')

                    sheet2.cell(row = contsheet2, column = 3).value = str(itemVariab.variableRec)
                    sheet2.cell(row = contsheet2, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet2.cell(row = contsheet2, column = 3).font = Font(size = "8", name='Barlow')
                    contsheet2+=1
           
        for row in range(4, sheet2.max_row + 1):  ## definir tamaño de rows
            sheet2.row_dimensions[row].height = 90

        ##################### HOJA 3  #################################################

        sheet3.merge_cells('A1:C1')
        sheet3.merge_cells('A2:C2')

        def set_border(sheet3, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet3[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet3,'A1:C1')
    
        #row dimensions
        sheet3.row_dimensions[1].height = 55
        sheet3.row_dimensions[2].height = 40

        # column width
        sheet3.column_dimensions['A'].width = 20
        sheet3.column_dimensions['B'].width = 20
        sheet3.column_dimensions['C'].width = 20
        
        title_cell = sheet3['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de los Registros Administrativos'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
        title_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
	
        codigo_cell = sheet3['A2']
        codigo_cell.value = 'La entidad hace uso de los datos del RA para generar información estadística (Agregados, indicadores, Resultados estadísticos, etc)'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet3['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       
        codigo_cell = sheet3['B3']
        codigo_cell.value = 'Nombre Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet3['C3']
        codigo_cell.value = 'Lista de resultados agregados o indicadores'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        contsheet3 = 4
        arrayraId = []
        for indexIndicador, itemIndicador in enumerate(indicadorResultadoAgregadoTextList):
            indexIndicador =  itemIndicador.rraa_id + 3
            arrayraId.append(itemIndicador.rraa_id)
            
            sheet3.cell(row = contsheet3, column = 1).value = itemIndicador.rraa.codigo_rraa
            sheet3.cell(row = contsheet3, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet3.cell(row = contsheet3, column = 1).font = Font(size = "8", name='Barlow')

            sheet3.cell(row = contsheet3, column = 2).value = str(itemIndicador.rraa)
            sheet3.cell(row = contsheet3, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet3.cell(row = contsheet3, column = 2).font = Font(size = "8", name='Barlow')

            sheet3.cell(row = contsheet3, column = 3).value = str(itemIndicador.ind_res_agre)
            sheet3.cell(row = contsheet3, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet3.cell(row = contsheet3, column = 3).font = Font(size = "8", name='Barlow')
            contsheet3+=1
           
        for row in range(4, sheet3.max_row + 1):  ## definir tamaño de rows
            sheet3.row_dimensions[row].height = 90

        ####################### Hoja 4 pregunta 18 #############################################


        def set_border(sheet4, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet4[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet4,'A1:AG'+str(rraas.count()+3))

		## size rows

        sheet4.row_dimensions[1].height = 55
        sheet4.row_dimensions[2].height = 40
        sheet4.row_dimensions[3].height = 50

        ## size column
        sheet4.column_dimensions['A'].width = 18
        sheet4.column_dimensions['B'].width = 18
        sheet4.column_dimensions['C'].width = 15
        sheet4.column_dimensions['D'].width = 15
        sheet4.column_dimensions['E'].width = 15
        sheet4.column_dimensions['F'].width = 15
        sheet4.column_dimensions['G'].width = 15
        sheet4.column_dimensions['H'].width = 15
        sheet4.column_dimensions['I'].width = 15
        sheet4.column_dimensions['J'].width = 15
        sheet4.column_dimensions['K'].width = 15
        sheet4.column_dimensions['L'].width = 15
        sheet4.column_dimensions['M'].width = 15
        sheet4.column_dimensions['N'].width = 15
        sheet4.column_dimensions['O'].width = 15
        sheet4.column_dimensions['P'].width = 15
        sheet4.column_dimensions['Q'].width = 15
        sheet4.column_dimensions['R'].width = 15
        sheet4.column_dimensions['S'].width = 15
        sheet4.column_dimensions['T'].width = 15
        sheet4.column_dimensions['U'].width = 15
        sheet4.column_dimensions['V'].width = 15
        sheet4.column_dimensions['W'].width = 15
        sheet4.column_dimensions['X'].width = 15
        sheet4.column_dimensions['Y'].width = 15
        sheet4.column_dimensions['Z'].width = 15
        sheet4.column_dimensions['AA'].width = 15
        sheet4.column_dimensions['AB'].width = 15
        sheet4.column_dimensions['AC'].width = 15
        sheet4.column_dimensions['AD'].width = 15
        sheet4.column_dimensions['AE'].width = 15
        sheet4.column_dimensions['AF'].width = 15
        sheet4.column_dimensions['AG'].width = 15
        sheet4.column_dimensions['AH'].width = 15
        

        title_cell = sheet4['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de Registros Administrativos'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')
    
        sheet4.merge_cells('A1:AG1')
        sheet4.merge_cells('A2:B2')
        sheet4.merge_cells('C2:AG2')

        ## heads

        codigo_cell = sheet4['A2']
        codigo_cell.value = 'REGISTRO ADMINISTRATIVO'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet4['C2']
        codigo_cell.value = 'Relacione las entidades que tienen acceso a los datos del RA:'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        
        codigo_cell = sheet4['A3']
        codigo_cell.value = 'Código'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        codigo_cell = sheet4['B3']
        codigo_cell.value = 'Nombre del Registro Administrativo'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        
        contsheet4 = 4
        for rraa in rraas: 
            sheet4.cell(row = contsheet4, column = 1).value = rraa.codigo_rraa
            sheet4.cell(row = contsheet4, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet4.cell(row = contsheet4, column = 1).font = Font(size = "8", name='Barlow')

            sheet4.cell(row = contsheet4, column = 2).value = rraa.nombre_ra
            sheet4.cell(row = contsheet4, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet4.cell(row = contsheet4, column = 2).font = Font(size = "8", name='Barlow')

            contsheet4+= 1
    
        entidadraid = []
        for indexEntidAcc, itemEntidAcc in enumerate(entidadesAccesoRRAATextList): 
            indexEntidAcc = itemEntidAcc.rraa_id + 3
                    
            entidadraid.append(itemEntidAcc.rraa_id)
           
            count_id = entidadraid.count(itemEntidAcc.rraa_id)
            c=0 #inicializamos el contador  
            n=5*count_id
            for i in range(1,n+1):  
                if i%5 == 0:  
                    i = i - 2  ##posición de celda
                    c+=1

            for indexra, itemra in enumerate(rraas):
                if itemEntidAcc.rraa_id == itemra.pk:
                    posRow = indexra + 4
                     
                    sheet4.cell(row = 3, column = i ).value = "Nombre de Entidad"
                    sheet4.cell(row = 3, column = i ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = 3, column = i ).font = Font(bold=True) 

                    sheet4.cell(row = posRow, column = i ).value = str(itemEntidAcc.nomb_entidad_acc) 
                    sheet4.cell(row = posRow, column = i ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = posRow, column = i ).font = Font(size = "8", name='Barlow')

                    sheet4.cell(row = 3, column = i+4).value = "¿cuál?"
                    sheet4.cell(row = 3, column = i+4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = 3, column = i+4).font = Font(bold=True)

                    sheet4.cell(row = posRow, column = i+4).value = str(itemEntidAcc.otro_cual) 
                    sheet4.cell(row = posRow, column = i+4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = posRow, column = i+4).font = Font(size = "8", name='Barlow')


            
            listaFinalidad = list(itemEntidAcc.opcion_pr.all())  ##iterar para traer las fases seleccionadas

            for indexFinalida, itemFinalidad in enumerate(listaFinalidad):

                if str(itemFinalidad) == "Estadístico":
                    
                    countFin=0 #inicializamos el contador  
                    finSelected=5*count_id 
                    for posFin in range(1,n+1):  
                        if posFin%5 == 0:
                            posFin = posFin - 1 ##posicion celda
                            countFin+=1

                    sheet4.cell(row = 3, column = posFin ).value = "Estadístico"
                    sheet4.cell(row = 3, column = posFin ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = 3, column = posFin ).font = Font(bold=True)

                    sheet4.cell(row = posRow, column = posFin).value = str(itemFinalidad)
                    sheet4.cell(row = posRow, column = posFin).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = posRow, column = posFin).font = Font(size = "8", name='Barlow')
                                
                if str(itemFinalidad) == "No sabe":

                    countFin=0 #inicializamos el contador  
                    finSelected=5*count_id 
                    for posFin in range(1,n+1):  
                        if posFin%5 == 0:  
                            posFin = posFin ##posición celda 
                            countFin+=1

                    sheet4.cell(row = 3, column = posFin ).value = "No sabe"
                    sheet4.cell(row = 3, column = posFin ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = 3, column = posFin ).font = Font(bold=True)

                    sheet4.cell(row = posRow, column = posFin).value = str(itemFinalidad)
                    sheet4.cell(row = posRow, column = posFin).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = posRow, column = posFin).font = Font(size = "8", name='Barlow')

                if str(itemFinalidad) == "Otro, ¿cuál?":

                    countFin=0 #inicializamos el contador  
                    finSelected=5*count_id 
                    for posFin in range(1,n+1):  
                        if posFin%5 == 0:  
                            posFin = posFin + 1
                            countFin+=1

                    sheet4.cell(row = 3, column = posFin ).value = "Otro"
                    sheet4.cell(row = 3, column = posFin ).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = 3, column = posFin ).font = Font(bold=True)

                    sheet4.cell(row = posRow, column = posFin).value = str(itemFinalidad)
                    sheet4.cell(row = posRow, column = posFin).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet4.cell(row = posRow, column = posFin).font = Font(size = "8", name='Barlow')
                


        for row in range(4, sheet4.max_row + 1):
            sheet4.row_dimensions[row].height = 90


        #################################### Hoja 5 ###########################

        def set_border(sheet5, cell_range):
            border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            rows = sheet5[cell_range]
            for row in rows:
                for cell in row:
                    cell.border = border

        set_border(sheet5,'A1:AZ'+str(rraas.count()+3))

		## size rows

        sheet5.row_dimensions[1].height = 55
        sheet5.row_dimensions[2].height = 40
        sheet5.row_dimensions[3].height = 50

        ## size column
        sheet5.column_dimensions['A'].width = 18
        sheet5.column_dimensions['B'].width = 18
        sheet5.column_dimensions['C'].width = 24
        sheet5.column_dimensions['D'].width = 24
        sheet5.column_dimensions['E'].width = 24
        sheet5.column_dimensions['F'].width = 24
        sheet5.column_dimensions['G'].width = 24
        sheet5.column_dimensions['H'].width = 24
        sheet5.column_dimensions['I'].width = 24
        sheet5.column_dimensions['J'].width = 24
        sheet5.column_dimensions['K'].width = 24
        sheet5.column_dimensions['L'].width = 24
        sheet5.column_dimensions['M'].width = 24
        sheet5.column_dimensions['N'].width = 24
        sheet5.column_dimensions['O'].width = 24
        sheet5.column_dimensions['P'].width = 24
        sheet5.column_dimensions['Q'].width = 24
        sheet5.column_dimensions['R'].width = 24
        sheet5.column_dimensions['S'].width = 24
        sheet5.column_dimensions['T'].width = 24
        sheet5.column_dimensions['U'].width = 24
        sheet5.column_dimensions['V'].width = 24
        sheet5.column_dimensions['W'].width = 24
        sheet5.column_dimensions['X'].width = 24
        sheet5.column_dimensions['Y'].width = 24
        sheet5.column_dimensions['Z'].width = 24
        sheet5.column_dimensions['AA'].width = 24
        sheet5.column_dimensions['AB'].width = 24
        sheet5.column_dimensions['AC'].width = 24
        sheet5.column_dimensions['AD'].width = 24
        sheet5.column_dimensions['AE'].width = 24
        sheet5.column_dimensions['AF'].width = 24
        sheet5.column_dimensions['AG'].width = 24
        sheet5.column_dimensions['AH'].width = 24
        
        sheet5.merge_cells('A1:AZ1')
        sheet5.merge_cells('A2:F2')
        sheet5.merge_cells('G2:AZ2')

        ## heads
        
        title_cell = sheet5['A1']
        title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
        title_cell.value = 'Directorio de Registros Administrativos'
        title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
        title_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet5['A2']
        codigo_cell.value = 'IDENTIFICACIÓN'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')

        codigo_cell = sheet5['G2']
        codigo_cell.value = 'Fortalecimiento RRAA'
        codigo_cell.font = Font(bold=True)
        codigo_cell.alignment = Alignment(horizontal='center')
        codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['A3']
        resultados_cell.value = 'Área Temática'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['B3']
        resultados_cell.value = 'Tema'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['C3']
        resultados_cell.value = 'Código Entidad'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['D3']
        resultados_cell.value = 'Nombre Entidad'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['E3']
        resultados_cell.value = 'Código RRAA'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center')
        
        resultados_cell = sheet5['F3']
        resultados_cell.value = 'Nombre del registro administrativo'
        resultados_cell.font = Font(bold=True)
        resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

        
        contsheet5 = 4
        for rraa in rraas: 
            
            sheet5.cell(row = contsheet5, column = 1).value = str(rraa.area_temat)
            sheet5.cell(row = contsheet5, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 1).font = Font(size = "8", name='Barlow') 
            
            sheet5.cell(row = contsheet5, column = 2).value = str(rraa.tema)
            sheet5.cell(row = contsheet5, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 2).font = Font(size = "8", name='Barlow') 
            
            sheet5.cell(row = contsheet5, column = 3).value = rraa.entidad_pri.codigo
            sheet5.cell(row = contsheet5, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 3).font = Font(size = "8", name='Barlow')

            sheet5.cell(row = contsheet5, column = 4).value = str(rraa.entidad_pri)
            sheet5.cell(row = contsheet5, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 4).font = Font(size = "8", name='Barlow')

            sheet5.cell(row = contsheet5, column = 5).value = rraa.codigo_rraa
            sheet5.cell(row = contsheet5, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 5).font = Font(size = "8", name='Barlow')

            sheet5.cell(row = contsheet5, column = 6).value = rraa.nombre_ra
            sheet5.cell(row = contsheet5, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet5.cell(row = contsheet5, column = 6).font = Font(size = "8", name='Barlow')
            
            contsheet5+= 1

        idraFort = []
        for indexFort, itemFort in enumerate(fortalecimientoTextList): 
            indexFort = itemFort.post_ra_id + 3
            idraFort.append(itemFort.post_ra_id)

            ### 1 Diagnóstico del RA
            
            count_diagnostico = idraFort.count(itemFort.post_ra_id)
            
            contrField1=0 #inicializamos el contador  
            pos_diagno_rraa = 9*count_diagnostico
            for incField1 in range(1,pos_diagno_rraa+1):  
                if incField1%9 == 0:  
                    incField1 = incField1 - 2
                    contrField1+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField1).value = "Diagnóstico del RRAA"
                    sheet5.cell(row = 3, column = incField1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField1).font = Font(bold=True) 

                    if str(itemFort.diagnostico_ra) == "True":

                        sheet5.cell(row = posiRow, column = incField1).value = "Si"
                        sheet5.cell(row = posiRow, column = incField1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField1).font = Font(size = "8", name='Barlow')

                    elif str(itemFort.diagnostico_ra) == "False":

                        sheet5.cell(row = posiRow, column = incField1).value = "No"
                        sheet5.cell(row = posiRow, column = incField1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField1).font = Font(size = "8", name='Barlow')

  
            ### 2 Año de diagnóstico de rraa

            count_id_ra = idraFort.count(itemFort.post_ra_id)
            contadorFort=0 #inicializamos el contador  
            posFields = 9*count_id_ra
            
            for incFort in range(1,posFields+1):  
                if incFort%9 == 0:
                    incFort = incFort - 1
                    contadorFort+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incFort).value = "Año de diagnóstico de RRAA"
                    sheet5.cell(row = 3, column = incFort).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incFort).font = Font(bold=True) 

                    if itemFort.year_diagnostico == None:

                        sheet5.cell(row = posiRow, column = incFort).value = ""
                        sheet5.cell(row = posiRow, column = incFort).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incFort).font = Font(size = "8", name='Barlow')
                    else:
                        sheet5.cell(row = posiRow, column = incFort).value = itemFort.year_diagnostico.strftime('%Y')
                        sheet5.cell(row = posiRow, column = incFort).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incFort).font = Font(size = "8", name='Barlow')
            
            ### 3 Módulo o sección diagnosticado:

            contadorField3=0 #inicializamos el contador  
            posFields3 = 9*count_id_ra
            for incField3 in range(1,posFields3+1):  
                if incField3%9 == 0:  
                    incField3 = incField3
                    contadorField3+=1
                
            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField3).value = "Módulo o sección diagnosticado:"
                    sheet5.cell(row = 3, column = incField3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField3).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField3).value = str(itemFort.mod_sec_diagn)
                    sheet5.cell(row = posiRow, column = incField3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField3).font = Font(size = "8", name='Barlow')

            ### 4 Plan de fortalecimiento aprobado por la entidad:

            contadorField4=0 #inicializamos el contador  
            posFields4 = 9*count_id_ra
            for incField4 in range(1,posFields4+1):  
                if incField4%9 == 0:
                    incField4 = incField4 + 1
                    contadorField4+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField4).value = "Plan de fortalecimiento aprobado por la entidad:"
                    sheet5.cell(row = 3, column = incField4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField4).font = Font(bold=True)

                    if str(itemFort.plan_fort_aprob) == "True":
                        sheet5.cell(row = posiRow, column = incField4).value = "Si"
                        sheet5.cell(row = posiRow, column = incField4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField4).font = Font(size = "8", name='Barlow')

                    elif str(itemFort.plan_fort_aprob) == "False":
                        sheet5.cell(row = posiRow, column = incField4).value = "No"
                        sheet5.cell(row = posiRow, column = incField4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField4).font = Font(size = "8", name='Barlow')
            
        ### 5 Fecha de aprobación del Plan de Fortalecimiento

            contadorField5=0 #inicializamos el contador  
            posFields5 = 9*count_id_ra
            for incField5 in range(1,posFields5+1):  
                if incField5%9 == 0:  
                    incField5 = incField5 + 2
                    contadorField5+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField5).value = "Fecha de aprobación del Plan de Fortalecimiento"
                    sheet5.cell(row = 3, column = incField5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField5).font = Font(bold=True)

                    if itemFort.fecha_aprobacion == None:

                        sheet5.cell(row = posiRow, column = incField5).value = ""
                        sheet5.cell(row = posiRow, column = incField5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField5).font = Font(size = "8", name='Barlow')
                    else:
                        sheet5.cell(row = posiRow, column = incField5).value = itemFort.fecha_aprobacion
                        sheet5.cell(row = posiRow, column = incField5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        sheet5.cell(row = posiRow, column = incField5).font = Font(size = "8", name='Barlow')
                    
        ### 6 Seguimiento a la implementación del Plan de fortalecimiento:

            contadorField6=0 #inicializamos el contador  
            posFields6 = 9*count_id_ra
            for incField6 in range(1,posFields6+1):  
                if incField6%9 == 0:  
                    incField6 = incField6 + 3
                    contadorField6+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField6).value = "Seguimiento a la implementación del Plan de fortalecimiento"
                    sheet5.cell(row = 3, column = incField6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField6).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField6).value = str(itemFort.seg_imple_plan)
                    sheet5.cell(row = posiRow, column = incField6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField6).font = Font(size = "8", name='Barlow')

        ### 7 Fecha de inicio de la implementación del Plan de Fortaleciemiento

            contadorField7=0 #inicializamos el contador  
            posFields7 = 9*count_id_ra
            for incField7 in range(1,posFields7+1):  
                if incField7%9 == 0:  
                    incField7 = incField7 + 4
                    contadorField7+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField7).value = "Fecha de inicio de la implementación del Plan de Fortaleciemiento"
                    sheet5.cell(row = 3, column = incField7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField7).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField7).value = str(itemFort.fecha_inicio_plan)
                    sheet5.cell(row = posiRow, column = incField7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField7).font = Font(size = "8", name='Barlow')

         ### 8 Fecha de último seguimiento a la implementación del Plan de Fortalecimiento   
            
            contadorField8=0 #inicializamos el contador  
            posFields8 = 9*count_id_ra
            for incField8 in range(1,posFields8+1):  
                if incField8%9 == 0:  
                    incField8 = incField8 + 5
                    contadorField8+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField8).value = "Fecha de último seguimiento a la implementación del Plan de Fortalecimiento "
                    sheet5.cell(row = 3, column = incField8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField8).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField8).value = str(itemFort.fecha_ultimo_seguimiento)
                    sheet5.cell(row = posiRow, column = incField8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField8).font = Font(size = "8", name='Barlow')

        ###  9 Fecha de finalización de la implementación del Plan de fortalecimiento
            
            contadorField9=0 #inicializamos el contador  
            posFields9 = 9*count_id_ra
            for incField9 in range(1,posFields9+1):  
                if incField9%9 == 0:  
                    incField9 = incField9 + 6
                    contadorField9+=1

            for indexra, itemra in enumerate(rraas):
                if itemFort.post_ra_id == itemra.pk:
                    posiRow = indexra + 4

                    sheet5.cell(row = 3, column = incField9).value = "Fecha de finalización de la implementación del Plan de fortalecimiento"
                    sheet5.cell(row = 3, column = incField9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = 3, column = incField9).font = Font(bold=True)

                    sheet5.cell(row = posiRow, column = incField9).value = str(itemFort.fecha_finalizacion)
                    sheet5.cell(row = posiRow, column = incField9).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    sheet5.cell(row = posiRow, column = incField9).font = Font(size = "8", name='Barlow')

       
        for row in range(4, sheet5.max_row + 1):
            sheet5.row_dimensions[row].height = 90

        ############# no mostrar información si no esta autenticado #######################
        if user.is_authenticated == True and user.profile.role.id != 2:
        ############# Hoja 6 Criticas #######################

            sheet6.merge_cells('A1:F1')
            sheet6.merge_cells('A2:F2')

            def set_border(sheet6, cell_range):
                border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                rows = sheet6[cell_range]
                for row in rows:
                    for cell in row:
                        cell.border = border

            set_border(sheet6,'A1:F1')
            
            #row dimensions
            sheet6.row_dimensions[1].height = 55
            sheet6.row_dimensions[2].height = 40
            sheet6.row_dimensions[3].height = 40
            sheet6.row_dimensions[4].height = 40
            sheet6.row_dimensions[5].height = 40
            sheet6.row_dimensions[6].height = 40
            

            # column width
            sheet6.column_dimensions['A'].width = 20
            sheet6.column_dimensions['B'].width = 20
            sheet6.column_dimensions['C'].width = 20
            sheet6.column_dimensions['D'].width = 20
            sheet6.column_dimensions['E'].width = 20
            sheet6.column_dimensions['F'].width = 20
        

            title_cell = sheet6['A1']
            title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
            title_cell.value = 'Directorio de los Registros Administrativos'
            title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
            title_cell.alignment =  Alignment(horizontal='center', vertical='center')

            codigo_cell = sheet6['A2']
            codigo_cell.value = 'Críticas'
            codigo_cell.font = Font(bold=True)
            codigo_cell.alignment = Alignment(horizontal='center')
            codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')


            resultados_cell = sheet6['A3']
            resultados_cell.value = 'Código'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet6['B3']
            resultados_cell.value = 'Nombre del Registro Administrativo'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet6['C3']
            resultados_cell.value = 'Estado de la crítica'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
                
            resultados_cell = sheet6['D3']
            resultados_cell.value = 'Observaciones de la crítica'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet6['E3']
            resultados_cell.value = 'Funcionario que realiza la crítica'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet6['F3']
            resultados_cell.value = 'Fecha en que se realiza la crítica'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            idraCritica = []
            contsheet6 = 4
            for indexCritica, itemCritica in enumerate(criticaTextList):
                indexCritica =  itemCritica.post_ra_id + 3
                idraCritica.append(itemCritica.post_ra_id)

                sheet6.cell(row = contsheet6, column = 1).value = itemCritica.post_ra.codigo_rraa
                sheet6.cell(row = contsheet6, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet6.cell(row = contsheet6, column = 1).font = Font(bold=True) 

                sheet6.cell(row = contsheet6, column = 2).value = str(itemCritica.post_ra)
                sheet6.cell(row = contsheet6, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet6.cell(row = contsheet6, column = 2).font = Font(bold=True)

                sheet6.cell(row = contsheet6, column = 3).value = str(itemCritica.estado_critica_ra)
                sheet6.cell(row = contsheet6, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet6.cell(row = contsheet6, column = 3).font = Font(bold=True)

                sheet6.cell(row = contsheet6, column = 4).value = str(itemCritica.observa_critica)
                sheet6.cell(row = contsheet6, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet6.cell(row = contsheet6, column = 4).font = Font(bold=True)

                sheet6.cell(row = contsheet6, column = 5).value = str(itemCritica.user_critico)
                sheet6.cell(row = contsheet6, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet6.cell(row = contsheet6, column = 5).font = Font(bold=True)

                sheet6.cell(row = contsheet6, column = 6).value = str(itemCritica.fecha_critica)
                sheet6.cell(row = contsheet6, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet6.cell(row = contsheet6, column = 6).font = Font(bold=True)

                contsheet6+=1

            for row in range(4, sheet6.max_row + 1):  ## definir tamaño de rows
                sheet6.row_dimensions[row].height = 90

            ############# END Hoja 6 Criticas #######################

            ############# Hoja 7 novedades #######################

            sheet7.merge_cells('A1:H1')
            sheet7.merge_cells('A2:H2')

            def set_border(sheet7, cell_range):
                border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                rows = sheet7[cell_range]
                for row in rows:
                    for cell in row:
                        cell.border = border

            set_border(sheet7,'A1:H1')
            

            #row dimensions
            sheet7.row_dimensions[1].height = 55
            sheet7.row_dimensions[2].height = 40
            sheet7.row_dimensions[3].height = 40
            sheet7.row_dimensions[4].height = 40
            sheet7.row_dimensions[5].height = 40
            sheet7.row_dimensions[6].height = 40
            sheet7.row_dimensions[7].height = 40
            sheet7.row_dimensions[8].height = 40

            # column width
            sheet7.column_dimensions['A'].width = 20
            sheet7.column_dimensions['B'].width = 20
            sheet7.column_dimensions['C'].width = 20
            sheet7.column_dimensions['D'].width = 20
            sheet7.column_dimensions['E'].width = 20
            sheet7.column_dimensions['F'].width = 20
            sheet7.column_dimensions['G'].width = 20
            sheet7.column_dimensions['H'].width = 20

            title_cell = sheet7['A1']
            title_cell.fill = PatternFill(fgColor='005E66', fill_type = 'solid') 
            title_cell.value = 'Directorio de los Registros Administrativos'
            title_cell.font = Font(bold=True, size = "16", name='Barlow', color='FFFFFF') 
            title_cell.alignment =  Alignment(horizontal='center', vertical='center')

            codigo_cell = sheet7['A2']
            codigo_cell.value = 'Novedades'
            codigo_cell.font = Font(bold=True)
            codigo_cell.alignment = Alignment(horizontal='center')
            codigo_cell.alignment =  Alignment(horizontal='center', vertical='center')


            resultados_cell = sheet7['A3']
            resultados_cell.value = 'Código'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet7['B3']
            resultados_cell.value = 'Nombre del  Registro Administrativo'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet7['C3']
            resultados_cell.value = 'Novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet7['D3']
            resultados_cell.value = 'Estado'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet7['E3']
            resultados_cell.value = 'Descripción de la novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet7['F3']
            resultados_cell.value = 'Observaciones de la novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            resultados_cell = sheet7['G3']
            resultados_cell.value = 'Funcionario que realiza la novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            resultados_cell = sheet7['H3']
            resultados_cell.value = 'Fecha en que se realiza la novedad'
            resultados_cell.font = Font(bold=True)
            resultados_cell.alignment =  Alignment(horizontal='center', vertical='center', wrap_text=True)

            idraNovedad = []
            contsheet7 = 4
            for indexNovedad, itemNovedad in enumerate(novedadTextList):
                indexNovedad =  itemNovedad.post_ra_id + 3
                idraNovedad.append(itemNovedad.post_ra_id)

                sheet7.cell(row = contsheet7, column = 1).value = itemNovedad.post_ra.codigo_rraa
                sheet7.cell(row = contsheet7, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 1).font = Font(bold=True) 

                sheet7.cell(row = contsheet7, column = 2).value = str(itemNovedad.post_ra)
                sheet7.cell(row = contsheet7, column = 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 2).font = Font(bold=True) 

                sheet7.cell(row = contsheet7, column = 3).value = str(itemNovedad.novedad)
                sheet7.cell(row = contsheet7, column = 3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 3).font = Font(bold=True)

                sheet7.cell(row = contsheet7, column = 4).value = str(itemNovedad.est_actualiz)
                sheet7.cell(row = contsheet7, column = 4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 4).font = Font(bold=True)

                sheet7.cell(row = contsheet7, column = 5).value = str(itemNovedad.descrip_novedad)
                sheet7.cell(row = contsheet7, column = 5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 5).font = Font(bold=True)

                sheet7.cell(row = contsheet7, column = 6).value = str(itemNovedad.obser_novedad)
                sheet7.cell(row = contsheet7, column = 6).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 6).font = Font(bold=True)

                sheet7.cell(row = contsheet7, column = 7).value = str(itemNovedad.name_nov)
                sheet7.cell(row = contsheet7, column = 7).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 7).font = Font(bold=True)

                sheet7.cell(row = contsheet7, column = 8).value = str(itemNovedad.fecha_actualiz)
                sheet7.cell(row = contsheet7, column = 8).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet7.cell(row = contsheet7, column = 8).font = Font(bold=True)
                contsheet7+=1

            for row in range(4, sheet7.max_row + 1):  ## definir tamaño de rows
                sheet7.row_dimensions[row].height = 90

        ############# END Hoja 7 novedades #######################


        file_name = "reporte_RRAA.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(file_name)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


############ reporte complemenatrio para administrador ######################

@method_decorator(login_required, name='dispatch')
class reporteUltimaNovedad(View):
    def get(self, request):
        
        # Get some data to write to the spreadsheet.
        registros_ad = RegistroAdministrativo.objects.all()
        ##NovedadActualizacionRRAA.filter(post_ra=1).order_by('-id')[:1] 
        # Create an in-memory output file for the new workbook.
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet(name="Ultima novedad")
        
        # Get some data to write to the spreadsheet.
        row = 4
        col = 0
        ## altura de celda
        worksheet.set_row(0, 40)
        worksheet.set_row(1, 30)
        worksheet.set_row(2, 30)
        worksheet.set_row(3, 30)
        worksheet.set_row(4, 30)
        worksheet.set_row(5, 30)
        worksheet.set_row(6, 30)

        #worksheet.set_default_row(20)
        ## definir ancho
        worksheet.set_column('A:A', 30)
        worksheet.set_column('B:B', 30)
        worksheet.set_column('C:C', 30)
        worksheet.set_column('D:D', 30)
        worksheet.set_column('E:E', 30)
        worksheet.set_column('F:F', 30)
        worksheet.set_column('G:G', 30)
        
        ## agregar estilos
        Backgroundcolor = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 18,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        Backgroundcolor2 = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 14,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        Backgroundcolor3 = workbook.add_format({ 'bg_color': '#005E66','font_color': '#FFFFFF',
            'font_size': 12,'text_wrap':True, 'align': 'center', 'valign': 'vcenter','border': 1})

        cell_format = workbook.add_format()
        cell_format.set_font_color('white')

        textInfo_format = workbook.add_format({'text_wrap': True, 'align': 'center', 'valign': 'vcenter'})
        

        formatfecha = workbook.add_format({'num_format':'yyyy-mm-dd', 'align': 'center', 'valign': 'vcenter'})

        ##FILA 1
        worksheet.merge_range('A1:G1', "Inventario de Registros Administrativos", Backgroundcolor)
        worksheet.conditional_format('A1:G1', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'Inventario de Demandas de Información',
                                                'format': cell_format})

        ##FILA 2

        worksheet.merge_range('A2:G3', "ÚLTIMAS NOVEDADES", Backgroundcolor2)
        worksheet.conditional_format('A2:G3', { 'type': 'text',
                                                'criteria': 'begins with',
                                                'value': 'ÚLTIMAS NOVEDADES',
                                                'format': cell_format })

        worksheet.write(3, 0, "Código", Backgroundcolor3)
        worksheet.write(3, 1, "Nombre del Registro Administrativo", Backgroundcolor3)
        worksheet.write(3, 2, "Fase en el sistema", Backgroundcolor3)
        worksheet.write(3, 3, "Área temática", Backgroundcolor3)
        worksheet.write(3, 4, "Tema", Backgroundcolor3)
        worksheet.write(3, 5, "Estado de la novedad", Backgroundcolor3)
        worksheet.write(3, 6, "Fecha en que se realiza la novedad", Backgroundcolor3)
         
        for item in registros_ad:
            item.codigo_rraa

            for reg in NovedadActualizacionRRAA.objects.filter(post_ra__codigo_rraa=item.codigo_rraa).order_by('-id')[:1]:

                worksheet.write(row, col, reg.post_ra.codigo_rraa, textInfo_format)
                worksheet.write(row, col + 1, reg.post_ra.nombre_ra, textInfo_format)
                worksheet.write(row, col + 2, str(reg.post_ra.sist_estado), textInfo_format)
                worksheet.write(row, col + 3, str(reg.post_ra.area_temat), textInfo_format)
                worksheet.write(row, col + 4, str(reg.post_ra.tema), textInfo_format)
                worksheet.write(row, col + 5, str(reg.est_actualiz), textInfo_format)
                worksheet.write(row, col + 6, str(reg.fecha_actualiz), formatfecha)
                row += 1

        # Close the workbook before sending the data.
        workbook.close()

        # Rewind the buffer.
        output.seek(0)

        filename = 'registros_admin.xlsx'
        # Set up the Http response.
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        return response
            






    









